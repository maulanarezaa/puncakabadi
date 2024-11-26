from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages
from django.db.models import F,FloatField
from django.db.models import Sum,Value
from . import models
from datetime import datetime,date
from django.db import IntegrityError
from urllib.parse import quote
from django.core.exceptions import ObjectDoesNotExist
import pandas as pd
from . import logindecorators
from django.contrib.auth.decorators import login_required
from .viewsproduksi import calculate_KSBB
from django.http import HttpResponse
from django.db.models.functions import Coalesce
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
import math
from io import BytesIO
import time
from openpyxl import Workbook, load_workbook
from .viewssignal import gethargapurchasingperbulanperproduk
"""PURCHASING"""


# # READ NOTIF BARANG MASUK PURCHASIN +SPK G+ACC
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def acc_subkon(request,id) :
    '''
    Fitur ini digunakan untuk melakukan ACC pada produk subkon yang masuk pada perusahaan dari vendor subkon
    Algoritma :
    1. Mencari data Detail Surat Jalan penerimaan produk subkon dengan parameter IDDetailSJPenerimaanSubkon = id (id didapatkan dari pasing values html)
    2. Menampilkan form update konfirmasi produk subkon masuk
    3. User mengisi form 
    4. Program mendapat input harga barang, supplier, keterangan, tanggal invoice, No invoice, hargapotongan dan status ppn (True = dipotong ppn, False = Tidak dipotong ppn)
    5. Mengupdate data pada database
    '''
    accobj = models.DetailSuratJalanPenerimaanProdukSubkon.objects.get(IDDetailSJPenerimaanSubkon=id)
    print('tes')
    if request.method == "GET" :
        valueppn = request.GET.get("input_ppn",2)
        try:
            valueppn = int(valueppn)
            if valueppn < 0:
                valueppn = 2
                messages.error(request, "Nilai Persentase Minus!")
        except ValueError:
            valueppn = 2
            messages.error(request, "Nilai PPN tidak valid!")
        inputppn = valueppn/100
        harga_total = accobj.Jumlah * accobj.Harga
        harga_potongan = harga_total * inputppn
        harga_total_setelah_potongan = harga_total - harga_potongan
        return render(
            request,"Purchasing/update_detailsubkon.html",
            {
                "accobj":accobj,
                "harga_total":harga_total,
                "harga_potongan":harga_potongan,
                "harga_total_potongan":harga_total_setelah_potongan,
            }
        )
    else :
        harga_barang = request.POST["harga_barang"]
        supplier = request.POST["supplier"]
        keterangan = request.POST["keterangan"]
        tanggalinvoice = request.POST['tanggalinvoice']
        noinvoice = request.POST['noinvoice']
        hargapotongan = request.POST['harga_satuan_setelah_pemotongan']
        # print(asd)

        accobj.KeteranganACC = True
        accobj.Harga = harga_barang
        accobj.NoSuratJalan.Supplier = supplier
        accobj.Keterangan= keterangan
        if tanggalinvoice != '':
            accobj.NoSuratJalan.TanggalInvoice = tanggalinvoice
        if noinvoice != '':
            accobj.NoSuratJalan.NoInvoice = noinvoice
        accobj.hargapotongan = hargapotongan
        accobj.save()
        accobj.NoSuratJalan.save()
        models.transactionlog(
            user="Purchasing",
            waktu=datetime.now(),
            jenis="ACC",
            pesan=f"No Surat Jalan Penerimaan Barang Subkon {accobj.NoSuratJalan} sudah ACC",
        ).save()
        messages.success(request,'Jangan lupa cek barang subkon')
        return redirect("notif_purchasing")
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])

def notif_barang_purchasing(request):
    '''
    Fitur ini digunakan untuk menampilkan dashboard dari Purchasing
    Isi dashboard Purchasing mencakup
    1. Rekap Pengadaan Bahan Baku
    2. Notifikasi Bahan Baku Masuk 
    3. Notifkasi Produk Subkon Masuk 
    4. Notifikasi SPK Belum ACC
    5. Notifikasi Bahan Baku keluar Belum ACC

    Algoritma
    1. Rekap Pengadaan Bahan Baku
        A. Mendapatkan data SPKArtikel aktif saat ini
        B. Agregasi jumlah artikel pada SPK Tersebut
        C. Mencari data Penyusun tiap artikel menggunakan versi Default
        D. Menghitung total kebutuhan Bahan Baku tiap Artikel
        E. TOtal kebutuhan = TOtal Stok Sekarang - Total kebutuhan + Delay PO
        F. Delay PO = Total Bahan Baku PO - TOtal bahan baku yang telah masuk ke perusahaan menggunakan PO Tersebut
    2. Notifikasi Bahan Baku Masuk
        A. Mendapatkan data DetailSuratJalanPembelian dengan KeteranganACC = False (menandakan barang belum di ACC Purchasing)
        B. Mengubah format tanggal menjadi YYYY-MM-DD
    3. Notifkasi Produk Subkon Masuk
        A. Mendapatkan data produk subkon masuk dengan keteranganACC = False (Menandakan barang belum di ACC Purchasing)
        B. Mengubah format tanggal menjadi YYYY-MM-DD
    4. Notifikasi SPK Belum ACC
        A. Mengambil data SPK Belum ACC dengan kriteria KeteranganACC = False (Menandakan belum ACC Purchasing)
        B. Mengubah format tanggal menjadi YYYY-MM-DD
    5. Notifikasi Bahan Baku Keluar Belum ACC
        A. Mengambil data TransaksiGudang dengan kriteria KeteranganACC = True (Sudah di acc Gudang), KeteranganACCPurchasing = False (Belum di ACC Purchasing), dan Jumlah > 0 (menandakan barang keluar)
        B. Mengubah format tanggal menjadi YYYY-MM-DD

    
    '''
    waktusekarang = datetime.now()
    filtersubkonobj = models.DetailSuratJalanPenerimaanProdukSubkon.objects.filter(KeteranganACC = False).order_by("NoSuratJalan__Tanggal")
    for x in filtersubkonobj :
        x.NoSuratJalan.Tanggal = x.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")
        if x.NoSuratJalan.TanggalInvoice:
            x.NoSuratJalan.TanggalInvoice = x.NoSuratJalan.TanggalInvoice.strftime("%Y-%m-%d")
    # print("Data subkon bos!", filtersubkonobj)
    filter_dataobj = models.DetailSuratJalanPembelian.objects.filter(
        KeteranganACC=False
    ).order_by("NoSuratJalan__Tanggal")
    for x in filter_dataobj:
        x.NoSuratJalan.Tanggal = x.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")
    filter_spkobj = models.SPK.objects.filter(KeteranganACC=False).order_by("Tanggal")
    for x in filter_spkobj:
        x.Tanggal = x.Tanggal.strftime("%Y-%m-%d")
    transaksikeluarbelumacc = models.TransaksiGudang.objects.filter(KeteranganACC = True, KeteranganACCPurchasing = False,jumlah__gte = 0).order_by('-tanggal')
    for i in transaksikeluarbelumacc:
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")
    # print(filter_spkobj)
    '''
    Algoritma mencari kebutuhan barang
    1. Ambil data SPK terlebih dahulu yang belum lunas
    2. Jumlahkan artikel detail SPK
    3. Hitung kebutuhan Artikel berdasarkan bahan baku
    4. Ambil nilai Bahan Baku 
    5. Hitung kurangnya
    '''
    '''REKAP PENGADAAN BAHAN BAKU'''
    dataspk = (
        models.DetailSPK.objects.filter(NoSPK__StatusAktif=True)
        .values(("KodeArtikel__KodeArtikel"))
        .annotate(kuantitas=Sum("Jumlah"))
        .order_by()
    )
    # print(dataspk)
    querysetartikel = []
    for item in dataspk:
        artikelobj  = models.Artikel.objects.get(KodeArtikel=item['KodeArtikel__KodeArtikel'])
        jumlah = item['kuantitas']
        artikelobj.Jumlah = jumlah
        querysetartikel.append(artikelobj)

        # listtipeartikel.append(item['KodeArtikel__KodeArtikel'])
        # listjumlah.append(item['kuantitas'])
    
    # querysetartikel = models.Artikel.objects.filter(KodeArtikel__in =listtipeartikel)
    # Cari versi terakhir dari tiap artikel
    listkebutuhanproduk = {}
    for item in querysetartikel:
        # versiterakhir = models.Penyusun.objects.filter(KodeVersi__KodeArtikel = item).values_list('versi',flat=True).distinct().order_by("versi").last()
        # print(versiterakhir)
        # penyusunobj = models.Penyusun.objects.filter(KodeArtikel = item, versi = versiterakhir)
        penyusunobj = models.Penyusun.objects.filter(KodeArtikel = item, KodeVersi__isdefault = True)
        print(penyusunobj)
        print(item)
        # print(asd)
        # print(len(penyusunobj))
        # print(konversimaster)
        # print(len(konversimaster))
        for datapenyusun in penyusunobj:
            bahanbakuobj = models.Produk.objects.get(KodeProduk = datapenyusun.KodeProduk.KodeProduk)
            # print(bahanbakuobj)
            if bahanbakuobj not in listkebutuhanproduk :
                listkebutuhanproduk[bahanbakuobj] = math.ceil(datapenyusun.Allowance * item.Jumlah)
            else:
                listkebutuhanproduk[bahanbakuobj] += math.ceil(datapenyusun.Allowance * item.Jumlah)


            # if bahanbakuobj.jumlahbahanbaku == None:
                # pass
    # print(listkebutuhanproduk)

    # print(asd)
    # Belum tambah logika PO
    '''Algoritma PO
    1. Ambil data PO Aktif
    2. Cek kondisi apakah ada data surat jalan masuk dengan PO tersebut
    3. Kalau ada maka kurangi agregat jumlah PO dengan total agregat jumlah masuk denganPO tersebut
    4. Kalau tidak ada maka pengurang dari total jumlah PO
    5. hitung total selisih ulang
    '''
    rekappengadaanbarang = {}
    rekapdata = {}
    for produk,jumlah in listkebutuhanproduk.items():

        cachevalue = models.CacheValue.objects.filter(KodeProduk = produk,Tanggal__month = waktusekarang.month).first()
        totalsaldosekarang = cachevalue.Jumlah
        kebutuhan = jumlah

        # Cari data PO Aktif 
        detailpoaktifobj = models.DetailPO.objects.filter(KodeProduk = produk,KodePO__Status = False)
        jumlahpo = 0
        if detailpoaktifobj.exists():
            jumlahagregatpo = detailpoaktifobj.aggregate(total = Sum('Jumlah'))['total'] 
            jumlahpo = jumlahagregatpo

        
        datatransaksisjp = models.DetailSuratJalanPembelian.objects.filter(PO__in = detailpoaktifobj)
        totalsjpberdasarkanpo = datatransaksisjp.aggregate(total=Sum('Jumlah'))['total']
        print(totalsjpberdasarkanpo,produk)
        if totalsjpberdasarkanpo != None:
            jumlahpo -= totalsjpberdasarkanpo


        selisih = totalsaldosekarang-kebutuhan+jumlahpo
        if selisih<0 : 
            rekappengadaanbarang[produk] = (abs(selisih))
            rekapdata[produk] = {'stokgudang':totalsaldosekarang,"jumlahminimal":produk.Jumlahminimal,'kebutuhanproduksi':kebutuhan,'totaldelaypo':jumlahpo ,'totalpengadaan':math.ceil(abs(selisih))+produk.Jumlahminimal}
        else:
            continue
        # print(cachevalue)
    rekappengadaanbarang = dict(sorted(rekappengadaanbarang.items(), key=lambda item: item[0].KodeProduk))

    # print(asd)
    '''BARANG DIBAWAH STOK
    ALGORITMA
    1. Ambil semua data bahan baku
    2. Ambil cache value
    3. Bandingkan antar jumlah minimal dan nilai sekarang
    '''

    rekapbahandibawahstok = {}
    allproduk = models.Produk.objects.all()
    for produk in allproduk:
        cachevalue = models.CacheValue.objects.filter(KodeProduk = produk,Tanggal__month = waktusekarang.month).first()
        print(cachevalue)
        if cachevalue:
            if cachevalue.Jumlah < produk.Jumlahminimal:
                rekapbahandibawahstok[produk] = cachevalue.Jumlah
            else:
                continue
        else:
            print(produk.KodeProduk)
            # print(asd)
            data = gethargapurchasingperbulanperproduk(waktusekarang,produk)
            print(data)
            print(data[-1])
            print(data[-1]['Sisahariini'])
            jumlahstokterakhir= data[-1]['Sisahariini']
            if jumlahstokterakhir < produk.Jumlahminimal:
                rekapbahandibawahstok[produk]= jumlahstokterakhir
            # print(ads)
    print(rekapbahandibawahstok)
    rekapbahandibawahstok = dict(sorted(rekapbahandibawahstok.items(), key=lambda item: item[0].KodeProduk))

    # print(asd)

    # print(asd)

    # list_artikel = []
    # list_q_gudang = []
    # list_data_art = []
    # list_hasil_conv = []
    # list_q_akhir = []
    # try:
    #     allartikel = models.Artikel.objects.all()
    #     for item in allartikel:
    #         nama_artikel = item.KodeArtikel
    #         list_artikel.append(nama_artikel)
    # except models.Artikel.DoesNotExist:
    #     messages.error(request, "Data Artikel tidak ditemukan")

    
    # awaltahun = datetime(year = waktusekarang.year,month=1,day=1)
    # datasjb = (
    #     models.DetailSuratJalanPembelian.objects.values("KodeProduk").filter(NoSuratJalan__Tanggal__range =(awaltahun,waktusekarang))
    #     .annotate(kuantitas=Sum("Jumlah"))
    #     .order_by()
    # )

    # if len(datasjb) == 0:
    #     messages.error(request, "Tidak ada barang masuk ke gudang")

    # datagudang = (
    #     models.TransaksiGudang.objects.values("KodeProduk").filter(tanggal__range =(awaltahun,waktusekarang))
    #     .annotate(kuantitas=Sum("jumlah"))
    #     .order_by()
    # )

    # datasaldoawal = models.SaldoAwalBahanBaku.objects.values("IDBahanBaku").filter(Tanggal__range =(awaltahun,waktusekarang)).filter(IDLokasi__NamaLokasi="Gudang").annotate(kuantitas=Sum("Jumlah")).order_by()

    # datapemusnahan = models.PemusnahanBahanBaku.objects.values("KodeBahanBaku").filter(Tanggal__range = (awaltahun,waktusekarang)).filter(lokasi__NamaLokasi="Gudang").annotate(kuantitas=Sum("Jumlah")).order_by()
    

    # print("data sjb", datasjb)
    # print("data gudang", datagudang)
    # print("data saldoawal", datasaldoawal)
    # print("data pemusnahan", datapemusnahan)
    # for item in datasjb:
    #     kode_produk = item["KodeProduk"]
    #     print('ini kode produk', kode_produk)
    #     try:
    #         corresponding_gudang_item = datagudang.get(KodeProduk=kode_produk)
    #         print("corresponding gudang :",corresponding_gudang_item)


    #     except models.TransaksiGudang.DoesNotExist:
    #         corresponding_gudang_item = {"KodeProduk" : kode_produk, "kuantitas" : 0}

    #     try :
    #         corresponding_saldo_awal_item = datasaldoawal.get(IDBahanBaku=kode_produk)
    #         print("corresponding saldo_awal :",corresponding_saldo_awal_item)
    #     except models.SaldoAwalBahanBaku.DoesNotExist :
    #         corresponding_saldo_awal_item = {"IDBahanBaku" : kode_produk, "kuantitas" : 0}
        
    #     try :
    #         corresponding_pemusnahan_item = datapemusnahan.get(KodeBahanBaku = kode_produk)
    #         print("corresponding pemusnahan :",corresponding_pemusnahan_item)
    #     except models.PemusnahanBahanBaku.DoesNotExist :
    #         corresponding_pemusnahan_item = {"KodeBahanBaku" : kode_produk, "kuantitas" : 0}
    #     print("kuantitas sjb :", item["kuantitas"])
    #     jumlah_akhir = item["kuantitas"] - corresponding_gudang_item["kuantitas"]+corresponding_saldo_awal_item["kuantitas"]-corresponding_pemusnahan_item["kuantitas"]
    #     # if jumlah_akhir < 0:
    #     #     messages.info(request,"Kuantitas gudang menjadi minus")

    #     list_q_gudang.append({kode_produk: jumlah_akhir})


    # print("LIST KODE ART :", list_artikel)
    # print("LIST Q GUDANG :", list_q_gudang)
    # # print(asd)
   
    # dataspk = (
    #     models.DetailSPK.objects.filter(NoSPK__StatusAktif=True)
    #     .values(("KodeArtikel__KodeArtikel"))
    #     .annotate(kuantitas2=Sum("Jumlah"))
    #     .order_by()
    # )

    # for item in dataspk:
    #     art_code = item["KodeArtikel__KodeArtikel"]
    #     jumlah_art = item["kuantitas2"]
    #     list_data_art.append({"Kode_Artikel": art_code, "Jumlah_Artikel": jumlah_art})

    # for item in list_data_art:
    #     kodeArt = item["Kode_Artikel"]

    #     jumlah_art = item["Jumlah_Artikel"]

    #     getidartikel = models.Artikel.objects.get(KodeArtikel=kodeArt)
    #     art_code = getidartikel.id

    #     try:
    #         konversi_art = (
    #             models.KonversiMaster.objects.filter(KodePenyusun__KodeArtikel=art_code)
    #             .annotate(
    #                 kode_art=F("KodePenyusun__KodeArtikel__KodeArtikel"),
    #                 kode_produk=F("KodePenyusun__KodeProduk"),
    #                 nilai_konversi=F("Allowance"),
    #                 nama_bb=F("KodePenyusun__KodeProduk__NamaProduk"),
    #                 unit_satuan=F("KodePenyusun__KodeProduk__unit"),
    #             )
    #             .values(
    #                 "kode_art", "kode_produk", "Kuantitas", "nama_bb", "unit_satuan"
    #             )
    #             .distinct()
    #         )
    #         # print(konversi_art)
    #         # for i in konversi_art:
    #         #     print(i)
    #         # print(asd)

    #         for item2 in konversi_art:
    #             kode_artikel = art_code
    #             kode_produk = item2["kode_produk"]
    #             nilai_conv = item2["Kuantitas"]
    #             nama_bb = item2["nama_bb"]
    #             unit_satuan = item2["unit_satuan"]
    #             hasil_conv = round(jumlah_art * nilai_conv)

    #             list_hasil_conv.append(
    #                 {
    #                     "Kode Artikel": kode_artikel,
    #                     "Jumlah Artikel": jumlah_art,
    #                     "Kode Produk": kode_produk,
    #                     "Nama Produk": nama_bb,
    #                     "Hasil Konversi": hasil_conv,
    #                     "Unit Satuan": unit_satuan,
    #                 }
    #             )

    #     except models.KonversiMaster.DoesNotExist:
    #         pass
    
    # for item in list_hasil_conv:
    #     kode_produk = item["Kode Produk"]
    #     hasil_konversi = item["Hasil Konversi"]
    #     for item2 in list_q_gudang:
    #         if kode_produk in item2:
    #             gudang_jumlah = item2[kode_produk]

    #             hasil_akhir = gudang_jumlah - hasil_konversi
    #             list_q_akhir.append(
    #                 {
    #                     "Kode_Artikel": item["Kode Artikel"],
    #                     "Jumlah_Artikel": item["Jumlah Artikel"],
    #                     "Kode_Produk": kode_produk,
    #                     "Nama_Produk": item["Nama Produk"],
    #                     "Unit_Satuan": item["Unit Satuan"],
    #                     "Kebutuhan": hasil_konversi,
    #                     "Stok_Gudang": gudang_jumlah,
    #                     "Selisih": hasil_akhir,
    #                 }
    #             )

    # # list_pengadaan = []
    # # for item in list_q_akhir:
    # #     kode_produk = item['Kode_Produk']
    # #     nama_produk = item['Nama_Produk']
    # #     satuan = item['Unit_Satuan']
    # #     selisih = item['Selisih']
    # #     if kode_produk in pengadaan['Kode_Produk'] :
    # #         pengadaan[]
    # pengadaan = {}

    # for item in list_q_akhir:
    #     produk = item["Kode_Produk"]
    #     pengadaan[produk] = [0, 0]

    # for item in list_q_akhir:
    #     produk = item["Kode_Produk"]
    #     nama_produk = item["Nama_Produk"]
    #     selisih = item["Selisih"]
    #     if produk in pengadaan:
    #         pengadaan[produk][0] = nama_produk
    #         pengadaan[produk][1] += selisih
    #     else:
    #         pengadaan[produk][0] = nama_produk
    #         pengadaan[produk][1] = selisih

    # rekap_pengadaan = {}

    # for key, value in pengadaan.items():
    #     if value[1] < 0:
    #         new_value = abs(value[1])
    #         for i, item in enumerate(list_q_akhir):
    #             if item["Kode_Produk"] == key:
    #                 key = models.Produk.objects.get(pk = key).KodeProduk
    #                 index = i
    #                 rekap_pengadaan[key] = [
    #                     value[0],
    #                     new_value,
    #                     list_q_akhir[index]["Unit_Satuan"],
    #                 ]

    # list_produk = []
    # jumlah_min = models.Produk.objects.values("KodeProduk", "Jumlahminimal")
    # print(jumlah_min)
    # datasjb = (
    #     models.DetailSuratJalanPembelian.objects.values(
    #         "KodeProduk",
    #         "KodeProduk__NamaProduk",
    #         "KodeProduk__unit",
    #         "KodeProduk__keteranganGudang",
    #     )
    #     .annotate(kuantitas=Sum("Jumlah"))
    #     .order_by()
    # )

    # datagudang = (
    #     models.TransaksiGudang.objects.values("KodeProduk")
    #     .annotate(kuantitas=Sum("jumlah"))
    #     .order_by()
    # )

    # # for item2 in jumlah_min :
    # #         kode = item2['KodeProduk']
    # #         print(kode)
    # for item in datasjb:
    #     kode_produk = item["KodeProduk"]
    #     nama_produk = item["KodeProduk__NamaProduk"]
    #     satuan = item["KodeProduk__unit"]
    #     try:
    #         corresponding_gudang_item = datagudang.get(KodeProduk=kode_produk)
    #         item["kuantitas"] -= corresponding_gudang_item["kuantitas"]
    #         for item2 in jumlah_min:
    #             print(item2)
    #             kode = item2["KodeProduk"]
    #             jumlah_minimal = item2["Jumlahminimal"]
    #             print(jumlah_minimal)
    #             print(item)
    #             # print(ads)
    #             print(kode)
    #             print(kode_produk)
    #             # print(asd)
    #             if kode == models.Produk.objects.get(pk = kode_produk).KodeProduk:
    #                 if item["kuantitas"] < jumlah_minimal:
    #                     list_produk.append(
    #                         {
    #                             "Kode_Produk": kode_produk,
    #                             "Nama_Produk": nama_produk,
    #                             "Satuan": satuan,
    #                             "Jumlah_minimal": jumlah_minimal,
    #                             "Jumlah_aktual": item["kuantitas"],
    #                         }
    #                     )
    #                     print(list_artikel)
    #                     # print(asd)
    #                 else:
    #                     continue
    #             else:
    #                 continue

    #     except models.TransaksiGudang.DoesNotExist:
    #         pass
    
    # print("\nrekap pengadaan :",rekap_pengadaan,"\n")
    # print("Ini list produk: ", list_produk)
    return render(
        request,
        "Purchasing/notif_purchasing.html",
        {
            "filterobj": filter_dataobj,
            "filter_spkobj": filter_spkobj,
            "rekap_pengadaan": rekapdata,
            "listproduk": rekapbahandibawahstok,
            "filtersubkonobj" : filtersubkonobj,
            "transaksibelumacc" : transaksikeluarbelumacc
        },
    )

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def accbarangkeluar(request, id):
    '''
    Fitur ini digunakan untuk melakukan ACC barang keluar dari Gudang ke produksi
    Algoritma 
    1. Mengambil data dari Transaksi Gudang dengan kriteria IDDetailTransaksiGudang = id (id didapatkan dari passing values html)
    2. Mengubah atribut KeteranganACCPurchasing = True (di acc purchasing)
    3. Menyimpan dalam database
    '''
    datagudang = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    datagudang.KeteranganACCPurchasing = True
    datagudang.save()
    models.transactionlog(user="Purchasing",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"ACC Barang Keluar {datagudang.tanggal} {datagudang.KodeProduk} {datagudang.jumlah} {datagudang.Lokasi.NamaLokasi}").save()
    return redirect("notif_purchasing")
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def verifikasi_data(request, id):
    '''
    Fitur ini digunakan untuk melakukan konfirmasi terhadap barang masuk pada gudang pada tabel Surat Jalan pembelian. 
    Algoritma
    1. Mendapatkan data detailsuratjalanpembelian dengan kriteria IDDetailSJPembelian = id (id didapatkan dari passing values html)
    2. mencari data DetailPO dengan kriteria KodeProduk sama seperti Kode produk bahan baku masuk yang akan di ACC dan Status PO = False (False = Belum lunas / Masih aktif, True = Lunas /non aktif)
    3. menampilkan form update acc bahan baku masuk
    4. Program mendapat inputan Mata uang, harga satuan, supplier, PO , No Invoice, tanggal invoice
    5. mengupdate data 
    '''
    verifobj = models.DetailSuratJalanPembelian.objects.get(IDDetailSJPembelian=id)
    opsipo = models.DetailPO.objects.filter(KodeProduk = verifobj.KodeProduk,KodePO__Status = False)
    if request.method == "GET":
        harga_total = verifobj.Jumlah * verifobj.Harga
        return render(
            request,
            "Purchasing/update_verif_data.html",
            {
                "verifobj": verifobj,
                "harga_total": harga_total,
                "opsipo": opsipo
            },
        )
    else:
        print(request.POST)
        # print(asd)
        try:
            isppn = request.POST['isppn']
        except KeyError:
            isppn = False
            print('error')
        # print(asd)
        matauang = request.POST['mata_uang']
        harga_barang = request.POST["harga_barang"]
        supplier = request.POST["supplier"]
        po_barang = request.POST["po_barang"]
        tanggalinvoice = request.POST['tanggalinvoice']

        if tanggalinvoice == '':
            tanggalinvoice = None
        noinvoice = request.POST['noinvoice']
        if matauang == "dollar":
            verifobj.HargaDollar = request.POST['harga_dollar']
        verifobj.KeteranganACC = True
        verifobj.Harga = harga_barang
        verifobj.NoSuratJalan.supplier = supplier
        verifobj.NoSuratJalan.PO = po_barang
        verifobj.NoSuratJalan.NoInvoice = noinvoice
        verifobj.NoSuratJalan.TanggalInvoice=tanggalinvoice
        verifobj.PPN = isppn

        verifobj.save()
        verifobj.NoSuratJalan.save()
        # print("verif:",verifobj.NoSuratJalan)
        models.transactionlog(
            user="Purchasing",
            waktu=datetime.now(),
            jenis="ACC",
            pesan=f"No Surat Jalan {verifobj.NoSuratJalan} sudah ACC",
        ).save()

        # tes = models.transactionlog.objects.filter(user = "Purchasing")
        # print("Tes ae : ",tes)
        messages.success(request,'Data berhasil disimpan')
        return redirect("notif_purchasing")


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def acc_notif_spk(request, id):
    '''
    Fitur ini digunakan untuk melakukan ACC terhadap SPK yang telah diterbitkan oleh produksi
    Algoritma 
    1. Mendapatkan data SPK dengan primarykey = id (id didapatkan dari passing html value)
    2. Mengubah status keteranganACC menjadi True (sudah acc)
    3. Menyimpan perubahan dalam database
    '''

    accobj = models.SPK.objects.get(pk=id)
    accobj.KeteranganACC = True
    accobj.save()
    models.transactionlog(
        user="Purchasing",
        waktu=datetime.now(),
        jenis="ACC",
        pesan=f"No SPK {accobj.NoSPK} sudah ACC",
    ).save()
    messages.success(request, "Jangan lupa cek data SPK!")
    return redirect("notif_purchasing")


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def exportbarang_excel(request):
    '''
    Fitur ini digunakan untuk melakukan Export data Detail Surat Jalan Pembelian lengkap dengan PPN 
    Algoritma : 
    1. Mendapatkan besar PPN
    2. Mendapatkan rentang waktu filter data
    3. Filter data Detail surat jalan pembelian dengan rentang waktu yang telah dipilih
    4. Mengeksport kedalam excel
    '''
    valueppn = request.GET.get("input_ppn", 11)
    try:
        valueppn = int(valueppn)
        if valueppn < 0:
            valueppn = 11
            messages.error(request, "Nilai Persentase Minus!")
    except ValueError:
        valueppn = 11
        messages.error(request, "Nilai PPN tidak valid!")

    inputppn = valueppn / 100
    print("INPUT PPN NI BANGG", inputppn)

    # Ambil nilai input_awal dan input_terakhir dari query string
    input_awal = request.GET.get("input_awal")
    input_terakhir = request.GET.get("input_terakhir")
    print("Ini input awal ama akhir", input_awal, input_terakhir)
    # print(asd)

    # Validasi format tanggal
    date_awal = parse_date(input_awal) if input_awal else None
    date_terakhir = parse_date(input_terakhir) if input_terakhir else None

    if date_awal is None or date_terakhir is None:
        sjball = models.DetailSuratJalanPembelian.objects.all().order_by("NoSuratJalan__Tanggal")
        print("len = 0, sjb all =", sjball)
    else:
        sjball = models.DetailSuratJalanPembelian.objects.filter(
            NoSuratJalan__Tanggal__range=(date_awal, date_terakhir)
        ).order_by("NoSuratJalan__Tanggal")

    print('TES SJB BANG', sjball)

    # Buat Workbook baru
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Barang Masuk'

    # Definisikan header untuk worksheet
    headers = ['Tanggal','Suppliers','Kode Bahan Baku','Nama Bahan Baku', 'Kuantitas', 'Harga', 'Harga Total', f'Harga PPN {valueppn}%', 'Harga Total PPN', "Tanggal Invoice","No Invoice"]
    worksheet.append(headers)

    total_harga = 0
    total_ppn = 0
    total_harga_ppn = 0
    # Tambahkan data ke worksheet
    for item in sjball:

        harga_total = item.Jumlah * item.Harga
        if item.PPN == True:
            harga_ppn = harga_total * inputppn
            harga_total_ppn = harga_total + harga_ppn
        else:
            harga_ppn = 0
            harga_total_ppn = harga_total
        total_ppn += harga_ppn
        total_harga_ppn += harga_total_ppn
        total_harga += harga_total
        row = [
            item.NoSuratJalan.Tanggal.strftime("%Y-%m-%d"),
            item.NoSuratJalan.supplier,
            item.KodeProduk.KodeProduk,
            item.KodeProduk.NamaProduk,
            item.Jumlah,
            item.Harga,
            harga_total,
            harga_ppn,
            harga_total_ppn,
            item.NoSuratJalan.TanggalInvoice,
            item.NoSuratJalan.NoInvoice

        ]
        worksheet.append(row)
    total_row = [
        "",  # Kosongkan kolom Tanggal
        "TOTAL",  # Tulis "TOTAL" di kolom Suppliers
        "",  # Kosongkan kolom Kode Bahan Baku
        "",  # Kosongkan kolom Nama Bahan Baku
        "",  # Kosongkan kolom Kuantitas
        "",  # Kosongkan kolom Harga
        total_harga,  # Total Harga
        total_ppn,  # Total Harga PPN
        total_harga_ppn,  # Total Harga Total PPN
        "",  # Kosongkan kolom Tanggal Invoice
        ""  # Kosongkan kolom No Invoice
    ]
    worksheet.append(total_row)

    # Menyesuaikan lebar kolom
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width
    for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=11):  # mulai dari baris 2 dan kolom 5 (Kuantitas)
        for cell in row:
            if cell.column in [5, 6, 7, 8, 9]:  # kolom Kuantitas dan Harga
                cell.number_format = '#,##0.00'

    # Menambahkan warna pada header
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = header_fill

        # Menambahkan border pada semua sel
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
    # Atur response untuk mengunduh file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Barang Masuk {input_awal}-{input_terakhir}.xlsx'
    workbook.save(response)

    return response
    
   
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def barang_masuk(request):
    '''
    Fitur ini digunakan untuk melihat register bahan baku masuk ke gudang
    Algoritma :
    1. Insiasi besar PPN. Secara default diatur menjadi 11 %
    2. Inisiasi Tanggal mulai dan tanggal akhir. Secara default akan diatur menjadi Awal bulan hingga tanggal user mengakses
    3. Filter data Detail Surat Jalan Pembelian 
    4. Mengjitung besar PPN tiap detail bahan baku masuk
    5. Apabila status PPN pada detail Bahan Baku False, maka tidak diikutkan perhitungan potongan PPN
    '''
    # if request.method ==  'POST' :
    if len(request.POST) == 0 :
        valueppn = 11
    else :
        valueppn = request.POST["input_ppn"]
        if int(valueppn) < 0 :
            valueppn = 11
            messages.error(request,"Nilai Persentase Minus!") 
        else :
            pass

    inputppn = int(valueppn)/100

    print("inputppn",inputppn)
 

    if len(request.GET) == 0:
        tanggalsaatini = datetime.now().date()
        input_awal = date(tanggalsaatini.year,tanggalsaatini.month,1).strftime('%Y-%m-%d')
        input_terakhir = tanggalsaatini.strftime('%Y-%m-%d')
        list_harga_total1 = []
        list_ppn = []
        list_total_ppn = []
        sjball = models.DetailSuratJalanPembelian.objects.filter(NoSuratJalan__Tanggal__range=(input_awal,input_terakhir)).order_by(
            "-NoSuratJalan__Tanggal"
        )
        print(sjball)
        if len(sjball) > 0:
            for x in sjball:
                harga_total = x.Jumlah * x.Harga
                x.NoSuratJalan.Tanggal = x.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")
                if x.NoSuratJalan.TanggalInvoice is not None:
                    x.NoSuratJalan.TanggalInvoice = x.NoSuratJalan.TanggalInvoice.strftime("%Y-%m-%d")
                print(harga_total)
                list_harga_total1.append(harga_total)
            for item in list_harga_total1:
                harga_ppn = item*inputppn
                harga_total_ppn = item+harga_ppn
                list_ppn.append(harga_ppn)
                list_total_ppn.append(harga_total_ppn)
            i = 0
            for item in sjball:
                
                item.harga_total = list_harga_total1[i]
                if item.PPN == True:
                    item.harga_ppn = list_ppn[i]
                    item.harga_total_ppn = list_total_ppn[i]
                else:
                    item.harga_ppn = 0
                    item.harga_total_ppn = item.harga_total
                i += 1
            print("list hartot", list_harga_total1)
            i = 0
            for item in sjball:
                # item.hargatotal = item.Harga * item.Jumlah
                # item.hargasatuanppn = item.hargappn * item.Jumlah
                # item.hargatotalppn = item.hargasatuanppn + item.hargatotal
                # item.harga_total = list_harga_total1[i]
                if item.PPN == True:
                    print(item,i)
                    item.harga_ppn = list_ppn[i]
                    item.harga_total_ppn = list_total_ppn[i]
                else:
                    item.harga_ppn = 0
                    item.harga_total_ppn = item.harga_total
                i += 1
            print("list hartot", list_harga_total1)
            
        
            return render(
                request,
                "Purchasing/masuk_purchasing.html",
                {
                    "sjball": sjball,
                    "harga_total": harga_total,
                    "harga_ppn" : harga_ppn,
                    "harga_total_ppn" : harga_total_ppn,
                    "valueppn" : valueppn,
                    'input_awal':input_awal,
                    'input_terakhir':input_terakhir
                },
            )
        else:
            messages.error(request, "Data tidak ditemukan")
            return render(
                request,
                "Purchasing/masuk_purchasing.html",
            )
    else:
        # if request.method == "POST": 
            
        # else :
        input_awal = request.GET["awal"]
        input_terakhir = request.GET["akhir"]
        # valueppn = request.POST["input_ppn"]
        # inputppn = request.POST["input_ppn"]
        list_harga_total = []
        list_ppn_1 = []
        list_total_ppn_1 = []
        
        try :
            filtersjb = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__range=(input_awal, input_terakhir)
            ).order_by("NoSuratJalan__Tanggal")
        except models.DetailSuratJalanPembelian.DoesNotExist :
            messages.error(request, "Data tidak ditemukan")
        if len(filtersjb) > 0:
            # if len(inputppn) <= 0 :
            #     inputppn = 0.11
            # else :
            #     inputppn = inputppn/100
            # ppn = 0.11
            for x in filtersjb:
                harga_total = x.Jumlah * x.Harga
                x.NoSuratJalan.Tanggal = x.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")
                list_harga_total.append(harga_total)
                if x.NoSuratJalan.TanggalInvoice is not None:
                    x.NoSuratJalan.TanggalInvoice = x.NoSuratJalan.TanggalInvoice.strftime("%Y-%m-%d")
            for item in list_harga_total:
                harga_ppn_1 = item*inputppn
                harga_total_ppn_1 = item+harga_ppn_1
                list_ppn_1.append(harga_ppn_1)
                list_total_ppn_1.append(harga_total_ppn_1)

            i = 0
            for item in filtersjb:
                item.harga_total = list_harga_total[i]
                if item.PPN == True:
                    item.harga_ppn_1 = list_ppn_1[i]
                    item.harga_total_ppn_1 = list_total_ppn_1[i]
                else:
                    item.harga_ppn_1 = 0
                    item.harga_total_ppn_1 = item.harga_total
                i += 1
            return render(
                request,
                "Purchasing/masuk_purchasing.html",
                {
                    "data_hasil_filter": filtersjb,
                    "harga_total": harga_total,
                    "harga_ppn_1" :harga_ppn_1,
                    "harga_total_ppn_1" : harga_total_ppn_1,
                    "input_awal": input_awal,
                    "input_terakhir": input_terakhir,
                    "valueppn" : valueppn
                },
            )
        else:
            messages.error(request, "Data tidak ditemukan")
            return redirect("barang_masuk")


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def update_barang_masuk(request, id):
    '''
    Fitur ini digunakan untuk mengupdate detail surat jalan pembelian bahan baku 
    Algoritma
    1. Mendapatkan data Detail Surat Jalan Pembelian dengan kriteria IDDetailSJPembelian = id (dimana id didapatkan dari passing values HTML)
    2. menampilkan form update detail surat jalan pembelian
    3. Program mendapatkan input berupa Harga Satuan, Harga Dollar (opsional, apabila menginput menggunakan rupiah akan diset menjadi 0),Supplier, PO, Tanggal Invoice (Opsional), No Invoice (Opsional), PPN (Apabila True maka diikutkan perhitungan potongan PPN, Apabila False maka tidak diikutkan perhitungan)
    4. mengupdate data 
    '''
    updateobj = models.DetailSuratJalanPembelian.objects.get(IDDetailSJPembelian=id)
    detailpotersedia = models.DetailPO.objects.filter(KodeProduk = updateobj.KodeProduk,KodePO__Status = False)
    print(detailpotersedia)
    print(updateobj.PO)
    if updateobj.HargaDollar > 0:
        updateobj.hargakonversi = updateobj.Harga / updateobj.HargaDollar
    else:
        updateobj.hargakonversi = 16000

    if request.method == "GET":
        harga_total = updateobj.Jumlah * updateobj.Harga
        if updateobj.NoSuratJalan.TanggalInvoice :
            updateobj.NoSuratJalan.TanggalInvoice =updateobj.NoSuratJalan.TanggalInvoice.strftime("%Y-%m-%d")
        
        return render(
            request,
            "Purchasing/update_barang_masuk.html",
            {
                "updateobj": updateobj,
                "harga_total": harga_total,
                "opsipo":detailpotersedia
            },
        )
    else:
        print(request.POST)
        # print(asd)
        try: 
            isppn = bool(request.POST['isppn'])
        except KeyError:
            isppn = False
            print('error')
        # print(asd)
        harga_barang = request.POST["harga_barang"]
        supplier = request.POST["supplier"]
        po_barang = request.POST["po_barang"]
        matauang = request.POST['mata_uang']
        noinvoice = request.POST['noinvoice']
        tanggalinvoice = request.POST['tanggalinvoice']
        # hargappn = request.POST['totalppn']

        if po_barang == "":
            po_barang = None
        else:
            po_barang = models.DetailPO.objects.filter(KodePO__KodePO = po_barang,KodeProduk = updateobj.KodeProduk).first()
        
        if noinvoice != "":
            updateobj.NoSuratJalan.NoInvoice = noinvoice
        else :
            updateobj.NoSuratJalan.NoInvoice = None
        
        if tanggalinvoice != "":
            updateobj.NoSuratJalan.TanggalInvoice = tanggalinvoice
        else:
            updateobj.NoSuratJalan.TanggalInvoice = None
        if matauang == "dollar" :
            updateobj.HargaDollar = request.POST['harga_dollar'] 
        updateobj.Harga = harga_barang
        updateobj.NoSuratJalan.supplier = supplier

        updateobj.PPN = isppn
        updateobj.PO = po_barang
        updateobj.NoSuratJalan.save()
        # updateobj.hargappn = hargappn
        updateobj.save()
        print(harga_barang, updateobj.Jumlah)
        harga_total = float(updateobj.Jumlah) * float(harga_barang)
        models.transactionlog(
            user="Purchasing",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"No Surat Jalan{updateobj.NoSuratJalan} sudah di Update",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("barang_masuk")
        # return JsonResponse({'harga_total': harga_total})

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def update_barangsubkon_masuk(request, id):
    '''
    Fitur ini digunakan untuk mengupdate data produk subkon masuk ke perusahaan
    Algoritma
    1. Mendapatkan data detail surat jalan penerimaan produk subkon dengan kriteria IDDetailSJPenerimaanSubkon = id (Dimana id didapatkan dari passing values HTML)
    2. Menampilkan form update Barang subkon masuk
    3. mendapatkan input dari user terkait Status potongan (True = dipotong, False = Tidak dipotong),Harga satuan, No Invoice (opsional), Tanggal invocie, Harga potongan, supplier
    4. Mengupdate data
    '''
    updateobj = models.DetailSuratJalanPenerimaanProdukSubkon.objects.get(IDDetailSJPenerimaanSubkon=id)
    print(updateobj)
    if  updateobj.NoSuratJalan.TanggalInvoice:
        updateobj.NoSuratJalan.TanggalInvoice = updateobj.NoSuratJalan.TanggalInvoice.strftime('%Y-%m-%d')
    # if updateobj.HargaDollar > 0:
    #     updateobj.hargakonversi = updateobj.Harga / updateobj.HargaDollar
    # else:
    #     updateobj.hargakonversi = 16000
    
    if request.method == "GET":
        harga_total = updateobj.Jumlah * updateobj.Harga
        return render(
            request,
            "Purchasing/update_barangsubkon_masuk.html",
            {
                "updateobj": updateobj,
                "harga_total": harga_total,
            },
        )
    else:
        print(request.POST)
        try:
            ispotongan = bool(request.POST['potongan'])
        except KeyError:
            ispotongan = False
            print('error')
        # print(asd)
        harga_barang = request.POST["harga_barang"]
        matauang = request.POST['mata_uang']
        noinvoice = request.POST['noinvoice']
        tanggalinvoice = request.POST['tanggalinvoice']
        hargapotongan = request.POST['harga_setelah_potongan']
        supplier = request.POST['supplier']
        if noinvoice != "":
            updateobj.NoSuratJalan.NoInvoice = noinvoice
        else :
            updateobj.NoSuratJalan.NoInvoice = None
        
        if tanggalinvoice != "":
            updateobj.NoSuratJalan.TanggalInvoice = tanggalinvoice
        else:
            updateobj.NoSuratJalan.TanggalInvoice = None
        if supplier != '':
            updateobj.NoSuratJalan.Supplier = supplier
        else: 
            updateobj.NoSuratJalan.Supplier = None

        if matauang == "dollar" :
            updateobj.HargaDollar = request.POST['harga_dollar'] 
        updateobj.Harga = harga_barang
        updateobj.Potongan = ispotongan
        updateobj.hargapotongan = hargapotongan
        updateobj.save()
        updateobj.NoSuratJalan.save()
        models.transactionlog(
            user="Purchasing",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"No Surat Jalan{updateobj.NoSuratJalan} sudah di Update",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("rekaphargasubkon")
        # return JsonResponse({'harga_total': harga_total})

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def rekap_gudang(request):
    '''
    Fitur ini digunakan untuk menghitung rekapitulasi bahan baku gudang saat ini
    Algoritma 
    1. Mendapatkan data waktu sekarang
    2. Mendapatkan semua data produk
    3. Mengiterasi semua data produk
    4. Menghitung data stok gudang dengan mengambil data Detail Surat Jalan Pembelian, Saldo Awal Bahan Baku,Transaksi Gudang, dan Pemusnahan Bahan Baku
    5. Rumus Stok = Saldo Awal Bahan Baku Gudang (lokasi = gudang) +total agregat detailsuratjalanpembelian (bahan baku masuk) + Transaksi Gudang retur (jumlah dibawah 0) - Transaksi Gudang Keluar (jumlah lebih dari sama dengan 0) - Transaksi Pemusnahan Bahan Baku,
    '''
    sekarang = datetime.now().strftime('%Y-%m-%d')
    listproduk = []
    listnama = []
    satuan = []
    liststokakhir = []

    dataproduk = models.Produk.objects.all()
    datenow = datetime.now()
    tahun = datenow.year
    mulai = datetime(year=tahun, month=1, day=1)
    date = request.GET.get("date")

    for i in dataproduk:
        listproduk.append(i.KodeProduk)
        listnama.append(i.NamaProduk)
        satuan.append(i.unit)

        if date is not None:
            datagudang = models.TransaksiGudang.objects.filter(
                tanggal__range=(mulai, date), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("jumlah"), Value(0,output_field=FloatField())))
            datasjp = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__range=(mulai, date), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            saldoawal = models.SaldoAwalBahanBaku.objects.filter(
                Tanggal__range=(mulai, date), IDBahanBaku=i, IDLokasi="3"
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            pemusnahan = models.PemusnahanBahanBaku.objects.filter(
                Tanggal__range=(mulai, date), KodeBahanBaku=i, lokasi="3"
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
        else:
            datagudang = models.TransaksiGudang.objects.filter(
                tanggal__range=(mulai, datenow), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("jumlah"), Value(0,output_field=FloatField())))
            datasjp = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__range=(mulai, datenow), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            saldoawal = models.SaldoAwalBahanBaku.objects.filter(
                Tanggal__range=(mulai, datenow), IDBahanBaku=i, IDLokasi="3"
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            pemusnahan = models.PemusnahanBahanBaku.objects.filter(
    Tanggal__range=(mulai, datenow), KodeBahanBaku=i, lokasi="3"
).aggregate(
    kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField()))
)

        stokakhir = (
            datasjp["kuantitas"]
            - datagudang["kuantitas"]
            + saldoawal["kuantitas"]
            - pemusnahan["kuantitas"]
        )
        liststokakhir.append(stokakhir)

        # print(datagudang)
        # print(datasjp)
        # print(saldoawal)
        # print(pemusnahan)
        # print(stokakhir)

    combined_list = zip(listproduk, listnama, satuan, liststokakhir,dataproduk)

    # Membuat dictionary sesuai template yang diinginkan
    produk_dict = {
        kode_produk: {
            "NamaProduk": nama_produk,
            "Satuan": satuan,
            "StokAkhir": stok_akhir,
            "produkobj" : produkobj
        }
        for kode_produk, nama_produk, satuan, stok_akhir,produkobj in combined_list
    }
    if date != "":
        sekarang = date

    return render(
        request,
        "Purchasing/rekapgudang2.html",
        {
            "kodeproduk": listproduk,
            "date": date,
            "dict_semua": produk_dict,
            'waktu' : sekarang
        },
    )
    # # batas
    # datasjb = (
    #     models.DetailSuratJalanPembelian.objects.values(
    #         "KodeProduk",
    #         "KodeProduk__NamaProduk",
    #         "KodeProduk__unit",
    #         "KodeProduk__keteranganGudang",
    #     )
    #     .annotate(kuantitas=Sum("Jumlah"))
    #     .order_by()
    # )
    # print(datasjb)
    # datenow = datetime.now()
    # tahun = datenow.year
    # mulai = datetime(year=tahun, month=1, day=1)
    # date = request.GET.get("date")
    # if date is not None:
    #     datasjb = (
    #         models.DetailSuratJalanPembelian.objects.filter(
    #             NoSuratJalan__Tanggal__range=(mulai, date)
    #         )
    #         .values(
    #             "KodeProduk",
    #             "KodeProduk__NamaProduk",
    #             "KodeProduk__unit",
    #             "KodeProduk__keteranganGudang",
    #         )
    #         .annotate(kuantitas=Sum("Jumlah"))
    #         .order_by()
    #     )

    # if len(datasjb) == 0:
    #     messages.error(request, "Tidak ada barang masuk ke gudang")

    # datagudang = (
    #     models.TransaksiGudang.objects.values("KodeProduk")
    #     .annotate(kuantitas=Sum("jumlah"))
    #     .order_by()
    # )
    # print(datasjb)
    # for item in datasjb:
    #     kode_produk = item["KodeProduk"]
    #     try:
    #         corresponding_gudang_item = datagudang.get(KodeProduk=kode_produk)
    #         item["kuantitas"] -= corresponding_gudang_item["kuantitas"]

    #         if item["kuantitas"] + corresponding_gudang_item["kuantitas"] < 0:
    #             messages.info("Kuantitas gudang menjadi minus")

    #     except models.TransaksiGudang.DoesNotExist:
    #         pass
    # return render(
    #     request, "Purchasing/rekapgudang2.html", {"datasjb": datasjb, "date": date}
    # )

    # datasjb = models.DetailSuratJalanPembelian.objects.values('KodeProduk','KodeProduk__NamaProduk','KodeProduk__unit','KodeProduk__keterangan').annotate(kuantitas=Sum('Jumlah')).order_by()
    # if len(datasjb) == 0 :
    #     messages.error(request, "Tidak ada barang masuk ke gudang")

    # datagudang = models.TransaksiGudang.objects.values('KodeProduk').annotate(kuantitas=Sum('jumlah')).order_by()

    # for item in datasjb:
    #     kode_produk = item['KodeProduk']
    #     try:
    #         corresponding_gudang_item = datagudang.get(KodeProduk=kode_produk)
    #         item['kuantitas'] += corresponding_gudang_item['kuantitas']

    #         if item['kuantitas'] + corresponding_gudang_item['kuantitas'] < 0 :
    #             messages.info("Kuantitas gudang menjadi minus")

    #     except models.TransaksiGudang.DoesNotExist:
    #         pass

    # return render(request,'Purchasing/rekapgudang2.html',{
    #     'datasjb' : datasjb,
    # })


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def view_rekapbarang(request):
    '''
    Fitur ini digunakan untuk melihat rekapitulasi barang produksi yang ada pada WIP dan FG
    Algoritma : 
    1. Mendapatkan tanggal saat ini 
    2. Mendapatkan data semua produk yang ada pada database
    3. Menghitung dengan memanggil fungsi calculate_KSBB dengan mengirimkan parameter Produk Objects, Tanggal mulai (awal tahun), dan tanggal akhir (tanggal sekarang/yang diinput user)
    4. Menampilkan rekapitulasi Bahan baku pada Produksi
    '''
    tanggal_akhir = request.GET.get("periode")
    if tanggal_akhir == '':
        tanggal_akhir = datetime.now().date().strftime('%Y-%m-%d')

    sekarang = datetime.now()
    tahun = sekarang.year

    tanggal_mulai = datetime(year=tahun, month=1, day=1)

    dataproduk = models.Produk.objects.all()
    dataproduk = models.Produk.objects.filter(KodeProduk = 'A-004-04')
    try:

        lokasi = request.GET['lokasi']
    except: 
        lokasi = "WIP"

    if tanggal_akhir:
        waktustart = time.time()
        for produk in dataproduk:
            # cektransaksiproduksi = models.TransaksiGudang.objects.filter(KodeProduk = produk).exists()
            # cekpemusnahanproduksi = models.PemusnahanBahanBaku.objects.filter(KodeBahanBaku = produk).exists()
            # cekpenyusun = models.Penyusun.objects.filter(KodeProduk = produk).values_list('KodeVersi__Versi',flat=True)
            # cekversitransaksiproduksi = models.TransaksiProduksi.objects.filter(VersiArtikel__Versi__in = cekpenyusun).exists()
            # cekpemusnahanartikel  = models.PemusnahanArtikel.objects.filter(VersiArtikel__in = cekpenyusun).exists()
            # cekmutasikodestok = models.transaksimutasikodestok.objects.filter(KodeProdukAsal=produk).exists()
            # cekmutasikodestokkelua = models.transaksimutasikodestok.objects.filter(KodeProdukTujuan=produk).exists()
            # if cektransaksiproduksi or cekpemusnahanartikel or cekpemusnahanproduksi or cekversitransaksiproduksi or cekmutasikodestok or cekmutasikodestokkelua:

                listdata, saldoawal = calculate_KSBB(produk, tanggal_mulai, tanggal_akhir,lokasi)

                if listdata:
                    produk.kuantitas = listdata[-1]["Sisa"][-1]
                else:
                    produk.kuantitas = 0
        endtime = time.time()
        print('waktu iterasi : ', endtime - waktustart)
    else:
        for produk in dataproduk:
            listdata, saldoawal = calculate_KSBB(produk, tanggal_mulai, sekarang,lokasi)

            if listdata:
                produk.kuantitas = listdata[-1]["Sisa"][-1]
            else:
                produk.kuantitas = 0
    return render(
        request,
        "Purchasing/rekapproduksi2.html",
        {"data": dataproduk, "tanggal_akhir": tanggal_akhir,'lokasi':lokasi},
    )


"""REVISI DELETE ADA YANG ERROR, TRS ERROR HANDLING GABOLE CREATE DENGAN KODE YG SAMA(done)"""


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def read_produk(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data terkait Produk
    Algoritma:
    1. Mengambil semua data produk dari database
    2. Menampilkan pada program
    '''
    produkobj = models.Produk.objects.all().order_by('KodeProduk')
    return render(request, "Purchasing/read_produk.html", {"produkobj": produkobj})
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def read_deletedproduk(request):
    '''
    Fitur ini digunakan untuk menampilkan produk yang telah terhapus
    '''
    produkobj = models.Produk.objects.isdeleted()
    print(produkobj)
    return render(request, "Purchasing/read_deletedproduk.html", {"produkobj": produkobj})

def restore_deletedproduk(request,id):
    '''
    Digunakan untuk merestore deleted Produk
    Algoritma
    1. Ambil data Objek produk yang akan direstore
    2. setting Value menjadi False pada atribut isdeleted
    '''
    dataobj = get_object_or_404(models.Produk.objects.isdeleted(), pk=id)  
    dataobj.is_deleted = False
    dataobj.save()
    return redirect('read_deletedproduk')

def harddelete_produk(request,id):
    dataobj = models.Produk.objects.with_deleted().get(pk = id)
    dataobj.hard_delete()
    return redirect('read_deletedproduk')


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def create_produk(request):
    '''
    Fitur ini digunakan untuk membuat data Produk Baru
    Algoritma : 
    1. Menampilkan form input produk 
    2. Program menerima inputan berupa Kode Produk, Nama Produk, Satuan, dan Keterangan Produk
    3. Apabila kode produk sudah ada pada sistem maka akan menampilkan pesan error. 
    4. menyimpan dalam database
    '''
    if request.method == "GET":
        return render(request, "Purchasing/create_produk.html")
    else:
        kode_produk = request.POST["kode_produk"]
        nama_produk = request.POST["nama_produk"]
        unit_produk = request.POST["unit_produk"]
        keterangan_produk = request.POST["keterangan_produk"]
        jumlah_minimal = 0
        produkobj = models.Produk.objects.filter(KodeProduk=kode_produk)
        print(produkobj)
        if len(produkobj) > 0:
            messages.error(request, "Kode Produk sudah ada")
            return redirect("create_produk")
        else:
            new_produk = models.Produk(
                KodeProduk=kode_produk,
                NamaProduk=nama_produk,
                unit=unit_produk,
                keteranganPurchasing=keterangan_produk,
                TanggalPembuatan=datetime.now(),
                Jumlahminimal=jumlah_minimal,
            )
            try:
                new_produk.save()
            except:
                messages.error(request,f'Kode Stok berada dalam daftar Soft Deleted. Silahkan kunjungi url read_deletedproduk untuk merestore atau menghapus permanen')
                return redirect('create_produk')
            models.transactionlog(
                user="Purchasing",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Kode Produk {kode_produk} sudah di Create",
            ).save()
            messages.success(request, "Data berhasil disimpan")
            return redirect("read_produk")


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def update_produk(request, id):
    '''
    Fitur ini digunakan untuk melakukan update Bahan Baku
    Algoritma : 
    1. Mendapatkan dara Bahan Baku dengan KodeProduk = id (id didapatkan dari passing values HTML)
    2. menampilkan form input update bahan baku
    3. Menerima input berupa Kode Produk, Nama Produk, Satuan, Keterangan
    4. Apabila mengganti Kode Produk yang sudah ada pada sistem maka akan menampilkan pesan error
    5. mengupdate data 

    '''
    produkobj = models.Produk.objects.get(KodeProduk=id)
    if request.method == "GET":
        return render(
            request, "Purchasing/update_produk.html", {"produkobj": produkobj}
        )
    else:

        kode_produk = request.POST["kode_produk"]
        nama_produk = request.POST["nama_produk"]
        unit_produk = request.POST["unit_produk"]
        keterangan_produk = request.POST["keterangan_produk"]
        jumlah_minimal = request.POST["jumlah_minimal"]
        namaproduk = models.Produk.objects.filter(KodeProduk=kode_produk).exists()
        print(namaproduk)
        print(request.POST)
        print(id)
        # print(asd)

        if namaproduk and kode_produk != id:
            messages.error(request,f'Kode Produk {kode_produk} sudah ada pada sistem')
            return redirect('update_produk',id = id)
        produkbaru = models.Produk.objects.get(KodeProduk=id)
        produkbaru.KodeProduk = kode_produk
        produkbaru.NamaProduk = nama_produk
        produkbaru.unit = unit_produk
        produkbaru.keteranganPurchasing = keterangan_produk
        produkbaru.Jumlahminimal = jumlah_minimal
        # print(asd)
        produkbaru.save()
        models.transactionlog(
            user="Purchasing",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Kode Produk {kode_produk} sudah di Update",
        ).save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("read_produk")


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def delete_produk(request, id):
    '''
    Fitur ini digunakan untuk menghapus data bahan baku
    Algoritma
    1. Mendapatkan data bahan baku dengan PrimaryKey = id (dimana id didapatkan dari passing values)
    2. Menghapus data bahan baku
    '''
    print(id)
    produkobj = models.Produk.objects.get(pk=id)
    # print("delete:",produkobj.KodeProduk)
    models.transactionlog(
        user="Purchasing",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Kode Produk {produkobj.KodeProduk} sudah di Delete",
    ).save()
    produkobj.delete()
    # print(asd)
    messages.success(request, "Data Berhasil dihapus")
    return redirect("read_produk")

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def read_spk(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data SPK 
    Algoritma:
    1. mendapatkan semua data SPK pada sistem
    2. Melakukan iterasi pada data SPK
    3. Mencari Detail SPK tiap SPK 
    4. Menyimpan detailspk pada atribut detailspk pada data SPK
    5. Menampilkan pada program
    '''
    dataspk = models.SPK.objects.all().order_by('-Tanggal')
    for spk in dataspk:
        if spk.StatusDisplay == False:
            detailspk = models.DetailSPK.objects.filter(NoSPK=spk.id)
        else:
            detailspk = models.DetailSPKDisplay.objects.filter(NoSPK=spk.id)
        spk.detailspk = detailspk
        spk.Tanggal = spk.Tanggal.strftime("%Y-%m-%d")
    return render(request, "Purchasing/read_spk.html", {"dataspk": dataspk})


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def track_spk(request, id):
    '''
    Fitur ini digunakan untuk melakukan tracking 
    algoritma : 
    1. Mengambil data SPK dengan kriteria id (primarykey) = id(id didapatkan dari passing values html)
    2. Jika SPK Tersebut adalah spk display maka ambil data detail SPK Display dengan kriteria No SPK = data SPK
    3. Jika SPK Tersebut adalah spk artikel maka ambil data detail spk dengan kriteria NoSPK = data spk
    4. Mengambil data transaksi Gudang dengan SPK Tersebut
    5. Mengambil data Transaksi Produksi dengan SPK Tersebut
    6. Mengambil data Pengiriman dengan SPK Tersebut
    7. Menghitung rekap permintaan bahan baku dengan SPK Tersebut
    8. Menghitung rekap pengiriman produk dengan SPK Tersebut.
    9. Menampilkan pada user 

    '''
    dataartikel = models.Artikel.objects.all()
    datadisplay =models.Display.objects.all()
    dataspk = models.SPK.objects.get(id=id)
    if dataspk.StatusDisplay == False:
        datadetail = models.DetailSPK.objects.filter(NoSPK=dataspk.id)
    else:
        datadetail = models.DetailSPKDisplay.objects.filter(NoSPK = dataspk.id)
    if dataspk.StatusDisplay ==True:
        
        # Data SPK terkait yang telah di request ke Gudang
        transaksigudangobj = models.TransaksiGudang.objects.filter(
            DetailSPKDisplay__NoSPK=dataspk.id, jumlah__gte=0
        )

        # Data SPK Terkait yang telah jadi di FG
        transaksiproduksiobj = models.TransaksiProduksi.objects.filter(
            DetailSPKDisplay__NoSPK=dataspk.id, Jenis="Mutasi"
        )
        print(transaksiproduksiobj)

        # Data SPK Terkait yang telah dikirim
        sppbobj = models.DetailSPPB.objects.filter(DetailSPKDisplay__NoSPK=dataspk.id)
        rekapjumlahpermintaanperbahanbaku = transaksigudangobj.values('KodeProduk__KodeProduk',"KodeProduk__NamaProduk","KodeProduk__unit").annotate(total = Sum('jumlah'))
        rekapjumlahpengirimanperartikel = sppbobj.values("DetailSPKDisplay__KodeDisplay__KodeDisplay").annotate(total=Sum('Jumlah'))
    else:
        # Data SPK terkait yang telah di request ke Gudang
        transaksigudangobj = models.TransaksiGudang.objects.filter(
            DetailSPK__NoSPK=dataspk.id, jumlah__gte=0
        )

        # Data SPK Terkait yang telah jadi di FG
        transaksiproduksiobj = models.TransaksiProduksi.objects.filter(
            DetailSPK__NoSPK=dataspk.id, Jenis="Mutasi"
        )

        # Data SPK Terkait yang telah dikirim
        sppbobj = models.DetailSPPB.objects.filter(DetailSPK__NoSPK=dataspk.id)
        rekapjumlahpermintaanperbahanbaku = transaksigudangobj.values('KodeProduk__KodeProduk',"KodeProduk__NamaProduk","KodeProduk__unit").annotate(total = Sum('jumlah'))
        rekapjumlahpengirimanperartikel = sppbobj.values("DetailSPK__KodeArtikel__KodeArtikel").annotate(total=Sum('Jumlah'))

    if request.method == "GET":
        tanggal = datetime.strftime(dataspk.Tanggal, "%Y-%m-%d")


    print(transaksigudangobj)

    return render(
        request,
        "Purchasing/trackspk.html",
        {
            "data": dataartikel,
            "datadisplay": datadisplay,
            "dataspk": dataspk,
            "datadetail": datadetail,
            "tanggal": tanggal,
            "transaksigudang": transaksigudangobj,
            "transaksiproduksi": transaksiproduksiobj,
            "transaksikeluar": sppbobj,
            "datarekappermintaanbahanbaku": rekapjumlahpermintaanperbahanbaku,
            "datarekappengiriman": rekapjumlahpengirimanperartikel,
        },
    )


# SPPB
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def view_sppb(request):
    '''
    Digunakan untuk melakukan melihat data SPPB
    Algoritma : 
    1. Mengambil semua data SPPB 
    2. mengubah format tanggal menjadi yyyy-mm-dd
    '''
    datasppb = models.SPPB.objects.all().order_by("-Tanggal")
    for i in datasppb:
        i.detailsppb = models.DetailSPPB.objects.filter(NoSPPB = i)
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")
    return render(request, "Purchasing/view_sppb2.html", {"datasppb": datasppb})


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def views_penyusun(request):
    '''
    Fitur ini digunakan untuk melihat penyusun/konversi dari artikel
    Algoritma : 
    1.Mengambil data semua artikel yang ada pada sistem
    2. Menampilkan form input artikel pada sistem
    3. Program menerima input berupa kode stok yang akan dimasukkan 
    4. menampilkan penyusun artikel

    '''
    print(request.GET)
    data = request.GET
    dataartikel  = models.Artikel.objects.all()
    if len(request.GET) == 0:
        data = models.Artikel.objects.all()

        return render(request, "Purchasing/penyusun.html", {"dataartikel": data})
    else:
        kodeartikel = request.GET["kodeartikel"]

        try:
            get_id_kodeartikel = models.Artikel.objects.get(KodeArtikel=kodeartikel)
            data = models.Penyusun.objects.filter(KodeArtikel=get_id_kodeartikel.id)
            versifiltered = models.Versi.objects.filter(KodeArtikel = get_id_kodeartikel)
            dataversi = versifiltered.values_list("Versi", flat=True).distinct()
            print(dataversi)
            if dataversi.exists():
                try:
                    if request.GET["versi"] == "":
                        versiterpilih = versifiltered.filter(isdefault=True).first().Versi
                        print("ini versi terbaru", versiterpilih)
                    else:
                        versiterpilih = request.GET["versi"]
                except:
                    versiterpilih = dataversi.order_by("-Versi").first()
                    print("ini versi terbaru", versiterpilih)
                
                data = data.filter(KodeVersi__Versi=versiterpilih)
                datakonversi = []
                nilaifg = 0
                sekarang = date.today()
                awaltahun = date(sekarang.year, 1, 1)
                akhirtahun = date(sekarang.year,12,31)
                print(data)
                if data.exists():
                    for item in data:
                        hargaterakhir = 0
                        print(item, item.IDKodePenyusun)
                        
                        kuantitaskonversi = item.Kuantitas
                        kuantitasallowance = item.Allowance
                        cachevalue = models.CacheValue.objects.filter(KodeProduk = item.KodeProduk, Tanggal__month =sekarang.month).first()
                        if cachevalue:
                            hargasatuanawal = cachevalue.Harga
                        else:
                        # print(konversidataobj.Kuantitas)
                            masukobj = models.DetailSuratJalanPembelian.objects.filter(
                KodeProduk=item.KodeProduk, NoSuratJalan__Tanggal__range=(awaltahun,akhirtahun)
            )

                            tanggalmasuk = masukobj.values_list("NoSuratJalan__Tanggal", flat=True)

                            keluarobj = models.TransaksiGudang.objects.filter(
                                jumlah__gte=0, KodeProduk=item.KodeProduk, tanggal__range=(awaltahun,akhirtahun)
                            )
                            returobj = models.TransaksiGudang.objects.filter(
                                jumlah__lt=0, KodeProduk=item.KodeProduk, tanggal__range=(awaltahun,akhirtahun)
                            )
                            pemusnahanobj = models.PemusnahanBahanBaku.objects.filter(
                                lokasi__NamaLokasi = 'Gudang',KodeBahanBaku = item.KodeProduk, Tanggal__range = (awaltahun,akhirtahun)
                            )
                            # print(pemusnahanobj)
                            # print(asd)
                            tanggalkeluar = keluarobj.values_list("tanggal", flat=True)
                            tanggalretur = returobj.values_list("tanggal", flat=True)
                            tanggalpemusnahan = pemusnahanobj.values_list("Tanggal",flat=True)
                            print("ini kode bahan baku", keluarobj)
                            saldoawalobj = (
                                models.SaldoAwalBahanBaku.objects.filter(
                                    IDBahanBaku=item.KodeProduk,
                                    IDLokasi__IDLokasi=3,
                                    Tanggal__range=(awaltahun,akhirtahun)
                                )
                                .order_by("-Tanggal")
                                .first()
                            )
                            print(saldoawalobj)
                            print('ini item',item)
                            if (
                                not keluarobj.exists()
                                and not returobj.exists()
                                and not masukobj.exists()
                                and saldoawalobj is None
                            ):
                                messages.error(request, f"Tidak ditemukan data Transaksi Barang pada penyusun {item.KodeProduk.KodeProduk} - {item.KodeProduk.NamaProduk}")
                                hargasatuanawal = 0
                                hargaterakhir = 0
                                hargaperkotak = 0
                                # print(asd
                                kuantitaskonversi = item.Kuantitas
                                kuantitasallowance = item.Allowance
                                print(item)
                                # print(asd)
                                # return redirect("rekapharga")
                            # print(asdas)
                            else :
                                if saldoawalobj:
                                    print("ada data")
                                    saldoawal = saldoawalobj.Jumlah
                                    hargasatuanawal = saldoawalobj.Harga
                                    hargatotalawal = saldoawal * hargasatuanawal
                                    tahun = saldoawalobj.Tanggal.year

                                else:
                                    saldoawal = 0
                                    hargasatuanawal = 0
                                    hargatotalawal = saldoawal * hargasatuanawal
                                    tahun = awaltahun.year

                                saldoawalobj = {
                                    "saldoawal": saldoawal,
                                    "hargasatuanawal": hargasatuanawal,
                                    "hargatotalawal": hargatotalawal,
                                    'tahun' : tahun
                                }
                                
                                listdata = []
                                print(tanggalmasuk)
                                print(tanggalkeluar)
                                listtanggal = sorted(
                                    list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalretur).union(tanggalpemusnahan)))
                                )
                                print(listtanggal)
                                statusmasuk = False
                                for i in listtanggal:
                                    jumlahmasukperhari = 0
                                    hargamasuktotalperhari = 0
                                    hargamasuksatuanperhari = 0
                                    jumlahkeluarperhari = 0
                                    hargakeluartotalperhari = 0
                                    hargakeluarsatuanperhari = 0
                                    sjpobj = masukobj.filter(NoSuratJalan__Tanggal=i)
                                    if sjpobj.exists():
                                        for j in sjpobj:
                                            hargamasuktotalperhari += j.Harga * j.Jumlah
                                            jumlahmasukperhari += j.Jumlah
                                        hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
                                        print("data SJP ada")
                                        print(hargamasuksatuanperhari)
                                        print(jumlahmasukperhari)
                                        dumy = {
                                            "Tanggal": i.strftime("%Y-%m-%d"),
                                            "Jumlahstokawal": saldoawal,
                                            "Hargasatuanawal": round(hargasatuanawal, 2),
                                            "Hargatotalawal": round(hargatotalawal, 2),
                                            "Jumlahmasuk": jumlahmasukperhari,
                                            "Hargamasuksatuan": round(hargamasuksatuanperhari, 2),
                                            "Hargamasuktotal": round(hargamasuktotalperhari, 2),
                                            "Jumlahkeluar": jumlahkeluarperhari,
                                            "Hargakeluarsatuan": round(hargakeluarsatuanperhari, 2),
                                            "Hargakeluartotal": round(hargakeluartotalperhari, 2),
                                        }
                                        saldoawal += jumlahmasukperhari - jumlahkeluarperhari
                                        hargatotalawal += hargamasuktotalperhari - hargakeluartotalperhari
                                        hargasatuanawal = hargatotalawal / saldoawal

                                        print("Sisa Stok Hari Ini : ", saldoawal)
                                        print("harga awal Hari Ini :", hargasatuanawal)
                                        print("harga total Hari Ini :", hargatotalawal, "\n")
                                        dumy["Sisahariini"] = saldoawal
                                        dumy["Hargasatuansisa"] = round(hargasatuanawal, 2)
                                        dumy["Hargatotalsisa"] = round(hargatotalawal, 2)
                                        print(dumy)
                                        statusmasuk = True
                                        listdata.append(dumy)
                                        # print(asdasd)

                                    hargamasuktotalperhari = 0
                                    jumlahmasukperhari = 0
                                    hargamasuksatuanperhari = 0
                                    transaksigudangobj = keluarobj.filter(tanggal=i)
                                    transaksipemusnahan = pemusnahanobj.filter(Tanggal=i)


                                    if transaksigudangobj.exists():
                                        for j in transaksigudangobj:
                                            jumlahkeluarperhari += j.jumlah
                                            hargakeluartotalperhari += j.jumlah * hargasatuanawal

                                    if transaksipemusnahan.exists():
                                        for j in transaksipemusnahan:
                                            jumlahkeluarperhari += j.Jumlah
                                            hargakeluartotalperhari += j.Jumlah * hargasatuanawal

                                    if jumlahkeluarperhari > 0:
                                        hargakeluarsatuanperhari = hargakeluartotalperhari / jumlahkeluarperhari
                                    else:
                                        if statusmasuk:
                                            statusmasuk = False
                                            continue
                                        hargakeluartotalperhari = 0
                                        hargakeluarsatuanperhari = 0
                                        jumlahkeluarperhari = 0

                                    # transaksigudangobj = keluarobj.filter(tanggal=i)
                                    # print(transaksigudangobj)
                                    # if transaksigudangobj.exists():
                                    #     for j in transaksigudangobj:
                                    #         jumlahkeluarperhari += j.jumlah
                                    #         hargakeluartotalperhari += j.jumlah * hargasatuanawal
                                    #     hargakeluarsatuanperhari += (
                                    #         hargakeluartotalperhari / jumlahkeluarperhari
                                    #     )
                                    # else:
                                    #     if statusmasuk:
                                    #         statusmasuk = False
                                    #         continue
                                    #     hargakeluartotalperhari = 0
                                    #     hargakeluarsatuanperhari = 0
                                    #     jumlahkeluarperhari = 0

                                    # transaksipemusnahan = pemusnahanobj.filter(Tanggal=i)
                                    # print(transaksipemusnahan)
                                    # if transaksipemusnahan.exists():
                                    #     for j in transaksipemusnahan:
                                    #         jumlahkeluarperhari += j.Jumlah
                                    #         hargakeluartotalperhari += j.Jumlah * hargasatuanawal
                                    #     hargakeluarsatuanperhari += (
                                    #         hargakeluartotalperhari / jumlahkeluarperhari
                                    #     )
                                    # else:
                                    #     if statusmasuk:
                                    #         statusmasuk = False
                                    #         continue
                                    #     hargakeluartotalperhari = 0
                                    #     hargakeluarsatuanperhari = 0
                                    #     jumlahkeluarperhari = 0

                                    transaksireturobj = returobj.filter(tanggal=i)
                                    if transaksireturobj.exists():
                                        for j in transaksireturobj:
                                            jumlahmasukperhari += j.jumlah * -1
                                            hargamasuktotalperhari += j.jumlah * hargasatuanawal * -1
                                        hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
                                    else:
                                        hargamasuktotalperhari = 0
                                        hargamasuksatuanperhari = 0
                                        jumlahmasukperhari = 0


                                    print("Tanggal : ", i)
                                    print("Sisa Stok Hari Sebelumnya : ", saldoawal)
                                    print("harga awal Hari Sebelumnya :", hargasatuanawal)
                                    print("harga total Hari Sebelumnya :", hargatotalawal)
                                    print("Jumlah Masuk : ", jumlahmasukperhari)
                                    print("Harga Satuan Masuk : ", hargamasuksatuanperhari)
                                    print("Harga Total Masuk : ", hargamasuktotalperhari)
                                    print("Jumlah Keluar : ", jumlahkeluarperhari)
                                    print("Harga Keluar : ", hargakeluarsatuanperhari)
                                    print(
                                        "Harga Total Keluar : ", hargakeluarsatuanperhari * jumlahkeluarperhari
                                    )
                                    

                                    dumy = {
                                        "Tanggal": i.strftime("%Y-%m-%d"),
                                        "Jumlahstokawal": saldoawal,
                                        "Hargasatuanawal": round(hargasatuanawal, 2),
                                        "Hargatotalawal": round(hargatotalawal, 2),
                                        "Jumlahmasuk": jumlahmasukperhari,
                                        "Hargamasuksatuan": round(hargamasuksatuanperhari, 2),
                                        "Hargamasuktotal": round(hargamasuktotalperhari, 2),
                                        "Jumlahkeluar": jumlahkeluarperhari,
                                        "Hargakeluarsatuan": round(hargakeluarsatuanperhari, 2),
                                        "Hargakeluartotal": round(hargakeluartotalperhari, 2),
                                    }
                                    """
                                    Rumus dari Excel KSBB Purchasing
                                    Sisa = Sisa hari sebelumnya + Jumlah masuk hari ini - jumlah keluar hari ini 
                                    harga sisa satuan = total sisa / harga total sisa
                                    Harga keluar = harga satuan hari sebelumnya

                                    """
                                    dummysaldoawal = saldoawal
                                    dummyhargatotalawal = hargatotalawal
                                    dummyhargasatuanawal = hargasatuanawal

                                    saldoawal += jumlahmasukperhari - jumlahkeluarperhari
                                    hargatotalawal += hargamasuktotalperhari - hargakeluartotalperhari
                                    print(hargatotalawal, saldoawal)
                                    try:
                                        hargasatuanawal = hargatotalawal / saldoawal
                                    except:
                                        hargasatuanawal = 0

                                    print("Sisa Stok Hari Ini : ", saldoawal)
                                    print("harga awal Hari Ini :", hargasatuanawal)
                                    print("harga total Hari Ini :", hargatotalawal, "\n")
                                    dumy["Sisahariini"] = saldoawal
                                    dumy["Hargasatuansisa"] = round(hargasatuanawal, 2)
                                    dumy["Hargatotalsisa"] = round(hargatotalawal, 2)

                                listdata.append(dumy)

                        hargaterakhir += hargasatuanawal
                        hargaperkotak = hargaterakhir * kuantitasallowance
                        # print("\n", hargaterakhir, "\n")
                        nilaifg += hargaperkotak

                        datakonversi.append(
                            {
                                "HargaSatuan": round(hargaterakhir, 2),
                                "Penyusunobj": item,
                                "Konversi": round(kuantitaskonversi, 5),
                                "Allowance": round(kuantitasallowance, 5),
                                "Hargakotak": round(hargaperkotak, 2),
                            }
                        )
                    HargaFGArtikel= None
                    hargaartikel = models.HargaArtikel.objects.filter(KodeArtikel =get_id_kodeartikel,Tanggal__month = sekarang.month)
                    if hargaartikel.exists():
                        HargaFGArtikel = hargaartikel.first().Harga

                    return render(
                        request,
                        "Purchasing/penyusun.html",
                        {
                            "data": datakonversi,
                            "kodeartikel": get_id_kodeartikel,
                            "nilaifg": nilaifg,
                            "versiterpilih": versiterpilih,
                            "dataversi": dataversi,
                            'dataartikel' : dataartikel,
                            "hargafgartikel" : HargaFGArtikel,
                            'versiterpilihobj': models.Versi.objects.get(Versi = versiterpilih,KodeArtikel=get_id_kodeartikel)
                        },
                    )

            else:
                messages.error(request, "Kode Artikel Belum memiliki penyusun")
                return render(
                    request,
                    "Purchasing/penyusun.html",
                    {"kodeartikel": get_id_kodeartikel},
                )
        except models.Artikel.DoesNotExist:
            messages.error(request, "Kode Artikel Tidak ditemukan")
            return render(
                request,
                "Purchasing/penyusun.html",
                {"dataartikel": models.Artikel.objects.all()},
            )
    # batas
    # print(request.GET)
    # data = request.GET
    # if len(request.GET) == 0:
    #     data = models.Artikel.objects.all()
    #     return render(request, "Purchasing/penyusun.html", {"dataartikel": data})
    # else:
    #     kodeartikel = request.GET["kodeartikel"]
    #     try:
    #         get_id_kodeartikel = models.Artikel.objects.get(KodeArtikel=kodeartikel)
    #         data = models.Penyusun.objects.filter(KodeArtikel=get_id_kodeartikel.id)
    #         datakonversi = []
    #         nilaifg = 0
    #         if data.exists():
    #             for item in data:
    #                 konversidataobj = models.KonversiMaster.objects.get(
    #                     KodePenyusun=item.IDKodePenyusun
    #                 )
    #                 print(konversidataobj.Kuantitas)
    #                 masukobj = models.DetailSuratJalanPembelian.objects.filter(
    #                     KodeProduk=item.KodeProduk
    #                 )
    #                 print("ini detail sjp", masukobj)
    #                 tanggalmasuk = masukobj.values_list(
    #                     "NoSuratJalan__Tanggal", flat=True
    #                 )
    #                 keluarobj = models.TransaksiGudang.objects.filter(
    #                     jumlah__gte=0, KodeProduk=item.KodeProduk
    #                 )
    #                 tanggalkeluar = keluarobj.values_list("tanggal", flat=True)
    #                 print(item)
    #                 saldoawalobj = (
    #                     models.SaldoAwalBahanBaku.objects.filter(
    #                         IDBahanBaku=item.KodeProduk.KodeProduk
    #                     )
    #                     .order_by("-Tanggal")
    #                     .first()
    #                 )
    #                 if saldoawalobj:
    #                     print(saldoawalobj)
    #                     saldoawal = saldoawalobj.Jumlah
    #                     hargasatuanawal = saldoawalobj.Harga
    #                     hargatotalawal = saldoawal * hargasatuanawal
    #                 else:
    #                     saldoawal = 0
    #                     hargasatuanawal = 0
    #                     hargatotalawal = saldoawal * hargasatuanawal

    #                 hargaterakhir = 0
    #                 listdata = []
    #                 listtanggal = sorted(list(set(tanggalmasuk.union(tanggalkeluar))))
    #                 print("inii", listtanggal)
    #                 for i in listtanggal:
    #                     jumlahmasukperhari = 0
    #                     hargamasuktotalperhari = 0
    #                     hargamasuksatuanperhari = 0
    #                     jumlahkeluarperhari = 0
    #                     hargakeluartotalperhari = 0
    #                     hargakeluarsatuanperhari = 0
    #                     sjpobj = masukobj.filter(NoSuratJalan__Tanggal=i)
    #                     if sjpobj.exists():
    #                         for j in sjpobj:
    #                             hargamasuktotalperhari += j.Harga * j.Jumlah
    #                             jumlahmasukperhari += j.Jumlah
    #                         hargamasuksatuanperhari += (
    #                             hargamasuktotalperhari / jumlahmasukperhari
    #                         )
    #                     else:
    #                         hargamasuktotalperhari = 0
    #                         jumlahmasukperhari = 0
    #                         hargamasuksatuanperhari = 0

    #                     transaksigudangobj = keluarobj.filter(tanggal=i)
    #                     print(transaksigudangobj)
    #                     if transaksigudangobj.exists():
    #                         for j in transaksigudangobj:
    #                             jumlahkeluarperhari += j.jumlah
    #                             hargakeluartotalperhari += j.jumlah * hargasatuanawal
    #                         hargakeluarsatuanperhari += (
    #                             hargakeluartotalperhari / jumlahkeluarperhari
    #                         )
    #                     else:
    #                         hargakeluartotalperhari = 0
    #                         hargakeluarsatuanperhari = 0
    #                         jumlahkeluarperhari = 0

    #                     saldoawal += jumlahmasukperhari - jumlahkeluarperhari
    #                     hargatotalawal += (
    #                         hargamasuktotalperhari - hargakeluartotalperhari
    #                     )
    #                     hargasatuanawal = hargatotalawal / saldoawal

    #                     print("ini hargasatuan awal : ", hargasatuanawal)

    #                 hargaterakhir += hargasatuanawal
    #                 kuantitaskonversi = konversidataobj.Kuantitas
    #                 kuantitasallowance = kuantitaskonversi + kuantitaskonversi * 0.025
    #                 hargaperkotak = hargaterakhir * kuantitasallowance
    #                 print("\n", hargaterakhir, "\n")
    #                 nilaifg += hargaperkotak

    #                 datakonversi.append(
    #                     {
    #                         "HargaSatuan": round(hargaterakhir, 2),
    #                         "Penyusunobj": item,
    #                         "Konversi": round(kuantitaskonversi, 5),
    #                         "Allowance": round(kuantitasallowance, 5),
    #                         "Hargakotak": round(hargaperkotak, 2),
    #                     }
    #                 )

    #             print(data)
    #             print(datakonversi)
    #             return render(
    #                 request,
    #                 "Purchasing/penyusun.html",
    #                 {
    #                     "data": datakonversi,
    #                     "kodeartikel": get_id_kodeartikel,
    #                     "nilaifg": nilaifg,
    #                 },
    #             )
    #         else:
    #             messages.error(request, "Kode Artikel Belum memiliki penyusun")
    #             return render(
    #                 request,
    #                 "Purchasing/penyusun.html",
    #                 {"kodeartikel": get_id_kodeartikel},
    #             )
    #     except models.Artikel.DoesNotExist:
    #         messages.error(request, "Kode Artikel Tidak ditemukan")
    #         return render(request, "Purchasing/penyusun.html")


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def kebutuhan_barang(request):
    '''
    Fitur ini digunakan untuk menghitung besar kebutuhan barang yang dibutuhkan untuk memproduksi SPK
    Algoritma : 
    1. Mendapatkan SPK dengan kriteria (StatusAktif = True (spk aktif/belum lunas), StatusDisplay = False (bukan SPK Display))
    2. Menampilkan form input pada user 
    3. Program mendapatkan inputan berupa kode SPK yang dipilih
    4. Menghitung Penggunaan Barang
        A. Mendapatkan kumulasi produksi artikel sesuai SPK
        B. Mengiterasi data artikel 
        C. Mencari Versi default pada tiap Artikel 
        D. Mengalikan konversi dengan jumlah pada SPK untuk setiap Artikel 
        E. Mendapatkan jumlah Penggunaan Barang
    5. Menghitung stok gudang saat ini 
    6. Menghitung kebutuhan barang dengan rumus = stok gudang - penggunaan barang
    7. Apabila ada data minus (Penggunaan barang lebih besar dari stok gudang maka akan masuk ke Rekap pengadaan)
    8. Menghitung rekap pengadaan untuk semua bahan baku
    '''
    list_q_gudang = []
    list_hasil_conv = []
    list_q_akhir = []
    list_kode_art = []
    if len(request.GET) == 0:
        spkall = models.SPK.objects.filter(StatusAktif=True,StatusDisplay=False)
        return render(request, "Purchasing/kebutuhan_barang.html", {"spkall": spkall})
    else:
        inputno_spk = request.GET["inputno_spk"]
        try:
            getspk = models.SPK.objects.get(NoSPK=inputno_spk)
        except ObjectDoesNotExist:
            messages.error(request, "Nomor SPK tidak ditemukan atau sudah tidak aktif")
            return redirect("kebutuhan_barang")

        filterspk = models.DetailSPK.objects.filter(NoSPK=getspk.id).filter(
            NoSPK__StatusAktif=True
        )

        if len(filterspk) == 0:
            messages.error(request, "Nomor SPK tidak ditemukan atau sudah tidak aktif")
            return redirect("kebutuhan_barang")
        else:
            dataspk = (
                models.DetailSPK.objects.filter(NoSPK=getspk.id)
                .annotate(kuantitas2=Sum("Jumlah"))
                .order_by()
            )

            for item in dataspk:
                art_code = (
                    item.KodeArtikel
                )  # Ini bukan objek Artikel, tapi kode artikel itu sendiri
                # artikel1 = models.Artikel.objects.get(KodeArtikel=art_code)  # Ambil objek Artikel berdasarkan kode
                artikel = art_code.KodeArtikel
                jumlah_art = item.kuantitas2
                list_kode_art.append(
                    {
                        "Kode_Artikel": artikel,
                        "Jumlah_Artikel": jumlah_art,
                    }  # Gunakan objek Artikel
                )

            if request.method == "POST":
                list_nama_art = request.POST.getlist("artikel[]")
                list_jumlah_art = request.POST.getlist("quantity[]")

                print("list nama", list_nama_art)
                print("list jumlah", list_jumlah_art)
                for item1, item2 in zip(list_nama_art, list_jumlah_art):
                   
                    kode_artikel_ada = False
                    jumlah_artikel = int(item2)
                    for i in list_kode_art:
                        if i["Kode_Artikel"] == item1:
                            i["Jumlah_Artikel"] += jumlah_artikel
                            kode_artikel_ada = True
                            break
                    if not kode_artikel_ada:
                        try:
                            artikelobj = models.Artikel.objects.get(KodeArtikel=item1)
                        except:
                            messages.error(request,f"Data artikel {item1} tidak ditemukan dalam sistem")
                            continue
                        list_kode_art.append(
                            {"Kode_Artikel": item1, "Jumlah_Artikel": jumlah_artikel}
                        )

                print(list_kode_art)

            artall = models.Artikel.objects.all()
            waktusekarang = datetime.now()
            awaltahun = datetime(year = waktusekarang.year,month=1,day=1)
            dataproduk = models.CacheValue.objects.filter(Tanggal__year = waktusekarang.year, Tanggal__month = waktusekarang.month)
            print(dataproduk)
            print(len(dataproduk))
            if dataproduk.exists():
                for item in dataproduk:
                    list_q_gudang.append({item.KodeProduk.id:item.Jumlah})

                # print(asd)
            else:
                datasjb = (
                    models.DetailSuratJalanPembelian.objects.values("KodeProduk").filter(NoSuratJalan__Tanggal__range =(awaltahun,waktusekarang))
                    .annotate(kuantitas=Sum("Jumlah"))
                    .order_by()
                )

                if len(datasjb) == 0:
                    messages.error(request, "Tidak ada barang masuk ke gudang")

                datagudang = (
                    models.TransaksiGudang.objects.values("KodeProduk").filter(tanggal__range =(awaltahun,waktusekarang))
                    .annotate(kuantitas=Sum("jumlah"))
                    .order_by()
                )

                datasaldoawal = models.SaldoAwalBahanBaku.objects.values("IDBahanBaku").filter(Tanggal__range =(awaltahun,waktusekarang)).filter(IDLokasi__NamaLokasi="Gudang").annotate(kuantitas=Sum("Jumlah")).order_by()

                datapemusnahan = models.PemusnahanBahanBaku.objects.values("KodeBahanBaku").filter(Tanggal__range = (awaltahun,waktusekarang)).filter(lokasi__NamaLokasi="Gudang").annotate(kuantitas=Sum("Jumlah")).order_by()
                

                print("data sjb", datasjb)
                print("data gudang", datagudang)
                print("data saldoawal", datasaldoawal)
                print("data pemusnahan", datapemusnahan)
                # print(asd)
                for item in datasjb:
                    kode_produk = item["KodeProduk"]
                    print('ini kode produk', kode_produk)
                    try:
                        corresponding_gudang_item = datagudang.get(KodeProduk=kode_produk)
                        print("corresponding gudang :",corresponding_gudang_item)


                    except models.TransaksiGudang.DoesNotExist:
                        corresponding_gudang_item = {"KodeProduk" : kode_produk, "kuantitas" : 0}

                    try :
                        corresponding_saldo_awal_item = datasaldoawal.get(IDBahanBaku=kode_produk)
                        print("corresponding saldo_awal :",corresponding_saldo_awal_item)
                    except models.SaldoAwalBahanBaku.DoesNotExist :
                        corresponding_saldo_awal_item = {"IDBahanBaku" : kode_produk, "kuantitas" : 0}
                    
                    try :
                        corresponding_pemusnahan_item = datapemusnahan.get(KodeBahanBaku = kode_produk)
                        print("corresponding pemusnahan :",corresponding_pemusnahan_item)
                    except models.PemusnahanBahanBaku.DoesNotExist :
                        corresponding_pemusnahan_item = {"KodeBahanBaku" : kode_produk, "kuantitas" : 0}
                    print("kuantitas sjb :", item["kuantitas"])
                    jumlah_akhir = item["kuantitas"] - corresponding_gudang_item["kuantitas"]+corresponding_saldo_awal_item["kuantitas"]-corresponding_pemusnahan_item["kuantitas"]
                    # if jumlah_akhir < 0:
                    #     messages.info(request,"Kuantitas gudang menjadi minus")

                    list_q_gudang.append({models.Produk.objects.get(pk = kode_produk): jumlah_akhir})
            

            print("LIST KODE ART :", list_kode_art)
            print("LIST Q GUDANG :", list_q_gudang)
            # print(asd)
            for item in list_kode_art:
                kodeArt = item["Kode_Artikel"]

                jumlah_art = item["Jumlah_Artikel"]
                try:

                    getidartikel = models.Artikel.objects.get(KodeArtikel=kodeArt)
                except:
                    messages.error(
                        request, f"Kode artikel {kodeArt} tidak ditemukan dalam sistem"
                    )
                    continue
                art_code = getidartikel

                # versipenyusunterakhir = models.Penyusun.objects.filter(KodeArtikel =art_code).order_by('-versi').first()
                versipenyusunterakhir = models.Versi.objects.filter(KodeArtikel =art_code,isdefault = True).first()
                print("versi terakhir",versipenyusunterakhir, art_code)
                # listpenyusun = models.Penyusun.objects.filter(KodeArtikel = art_code, versi = versipenyusunterakhir.versi).values_list('IDKodePenyusun',flat=True)
                listpenyusun = models.Penyusun.objects.filter(KodeArtikel = art_code, KodeVersi = versipenyusunterakhir).values_list('IDKodePenyusun',flat=True)
                print(listpenyusun)
                # print(asd)

                # print(asd)
                # konversi_art = (
                #     models.Penyusun.objects.filter(
                #         KodeArtikel=art_code, KodePenyusun__in=listpenyusun
                #     )
                #     .annotate(
                #         kode_art=F("KodePenyusun__KodeArtikel__KodeArtikel"),
                #         kode_produk=F("KodePenyusun__KodeProduk"),
                #         nilai_konversi=F("Allowance"),
                #         nama_bb=F("KodePenyusun__KodeProduk__NamaProduk"),
                #     )
                #     .values("kode_art", "kode_produk", "Kuantitas", "nama_bb")
                #     .distinct()
                # )
                konversi_art = listpenyusun.annotate(kode_art=F("KodeArtikel__KodeArtikel"),
                        kode_produk=F("KodeProduk"),
                        nilai_konversi=F("Allowance"),
                        nama_bb=F("KodeProduk__NamaProduk"),).values("kode_art", "kode_produk", "nilai_konversi", "nama_bb")

                for item2 in konversi_art:
                    print(item2)

                    kode_artikel = art_code.KodeArtikel
                    kode_produk = item2["kode_produk"]
                    nilai_conv = item2["nilai_konversi"]
                    nama_bb = item2["nama_bb"]

                    hasil_conv = math.ceil(jumlah_art * nilai_conv)

                    list_hasil_conv.append(
                        {
                            "Kode Artikel": kode_artikel,
                            "Jumlah Artikel": jumlah_art,
                            "Kode Produk": kode_produk,
                            "Nama Produk": nama_bb,
                            "Hasil Konversi": hasil_conv,
                        }
                    )
                # pritn(asd)

            for item in list_hasil_conv:
                print(list_hasil_conv)
                print(item)
                # print(asd)
                kode_produk = item["Kode Produk"]
                hasil_konversi = item["Hasil Konversi"]
                datacachevalue = models.CacheValue.objects.filter(KodeProduk = kode_produk,Tanggal__year = waktusekarang.year, Tanggal__month = waktusekarang.month).first()
                print(datacachevalue)
                print(kode_produk)
                if datacachevalue.KodeProduk.KodeProduk == 'A-001-02':
                    print(datacachevalue.Jumlah)
                    # print(asd)
                gudang_jumlah = datacachevalue.Jumlah
                hasil_akhir = gudang_jumlah - hasil_konversi
                # print(asd)
                list_q_akhir.append(
                            {
                                "Kode_Artikel": item["Kode Artikel"],
                                "Jumlah_Artikel": item["Jumlah Artikel"],
                                "Kode_Produk": datacachevalue.KodeProduk.KodeProduk,
                                "Nama_Produk": item["Nama Produk"],
                                "Kebutuhan": hasil_konversi,
                                "Stok_Gudang": gudang_jumlah,
                                "Selisih": hasil_akhir,
                            }
                        )
              
                
            pengadaan = {}

            for item in list_q_akhir:
                produk = item["Kode_Produk"]
                pengadaan[produk] = [0, 0,0]
            print(list_q_akhir)
            print(pengadaan)
            # print(asd)
            for item in list_q_akhir:
                produk = item["Kode_Produk"]
                nama_produk = item["Nama_Produk"]
                selisih = item['Kebutuhan']
                stoksekarnag = item['Stok_Gudang']
                if produk in pengadaan:
                    pengadaan[produk][0] = nama_produk
                    pengadaan[produk][1] += selisih
                    pengadaan[produk][2] = stoksekarnag
                else:
                    pengadaan[produk][0] = nama_produk
                    pengadaan[produk][1] = selisih
                    pengadaan[produk][2] = stoksekarnag

            rekap_pengadaan = {}

            print(pengadaan)
            for key, value in pengadaan.items():
                value[1] -= value[2]
                if value[1] > 0:
                    new_value = abs(value[1])
                    rekap_pengadaan[key] = [value[0], new_value]
            print(pengadaan)
            # print(asd)
            return render(
                request,
                "Purchasing/kebutuhan_barang.html",
                {
                    "artall": artall,
                    "filterspk": filterspk,
                    "list_kode_art": list_kode_art,
                    "inputno_spk": inputno_spk,
                    "list_q_akhir": list_q_akhir,
                    "rekap_pengadaan": rekap_pengadaan,
                },
            )



@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def views_rekapharga(request):
    '''
    Fitur ini digunakan untuk menghitung KSBB purchasing
    Algoritma 
    1. Mengambil data semua produk yang ada 
    2. Menampilkan form input Kode Produk 
    3. Program mendapatkan input berupa kode produk 
    4. Data tanggal mulai = Awal tahun, Data tanggal akhir = Akhir tahun 
    5. Mengambil data DetailSuratJalanPembelian dengan kriteria KodeProduk = kode produk input user, NoSuratJalan__Tanggal__range = awal tahun, akhir tahun. 
    6. mengambil data Transaksi Gudang keluar dengan kriteria Jumlah >= 0 (mengindikasikan barang keluar), KodeProduk = kode produk input user, Tanggal dalam rentang awal tahun hingga akhir tahun
    7. mengambil data Transaksi Gudang retur dengan kriteria Jumlah < 0 (mengindikasikan barang keluar), KodeProduk = kode produk input user, Tanggal dalam rentang awal tahun hingga akhir tahun
    8. mengambil data pemusnahan bahan baku  dengan lokasi Gudang, kodeproduk = kode produk input user, Tanggal berada dalam rentanga awal tahun, akhirtahun
    9. Mengambil data saldo awal bahan baku dengan kriteria kode produk = kode produk input user, Lokasi = Gudang, Tanggal berada dalam rentang awal tahun dan akhir tahun
    10. Mengiterasi setiap tanggal dari detailsuratjalanpembelian, transaksigudang keluar, transaksigudang retur, dan pemusnahan bahan baku 
    11. Total Stok = Saldoawal bahan baku + Detail surat jalan pembelian + Transaksi Gudang Retur - pemusnahan bahan baku - transaksi gudang keluar
    12. Menghitung harga satuan = harga total / total stok (perhitungan menggunakan weighted average)
    13. Apabila ada pembagian dengan hasil Null / pembagi 0 maka akan diset menjadi 0
    '''

    kodeprodukobj = models.Produk.objects.all()
    tahun = datetime.now().year
    if len(request.GET) == 0:
        return render(
            request, "Purchasing/views_ksbb.html", {"kodeprodukobj": kodeprodukobj,'tahun':tahun}
        )
    else:
        kode_produk = request.GET["kode_produk"]
        tahun = request.GET["tahun"]
        tahun_period = tahun
        produkobj = models.Produk.objects.with_deleted().filter(KodeProduk = kode_produk).first()
        if produkobj == None:
            messages.error(request,'Kode Stok tidak ditemukan')
            return redirect('rekapharga')

        listdata,saldoawalobj, hargaterakhir = getksbbproduk(kode_produk,tahun_period)
        

        return render(
            request,
            "Purchasing/views_ksbb.html",
            {
                "data": listdata,
                "Hargaakhir": hargaterakhir,
                "Saldoawal": saldoawalobj,
                "kodeprodukobj": kodeprodukobj,
                "kode_produk": produkobj,
                "tahun": tahun_period,
            },
        )

def getksbbproduk(kodeproduk,periode):
    '''Input berupa String Kode Bahan Baku  dan Periode tahun '''
    '''
    Fitur ini digunakan untuk menghitung KSBB purchasing
    Algoritma 
    1. Mengambil data semua produk yang ada 
    2. Menampilkan form input Kode Produk 
    3. Program mendapatkan input berupa kode produk 
    4. Data tanggal mulai = Awal tahun, Data tanggal akhir = Akhir tahun 
    5. Mengambil data DetailSuratJalanPembelian dengan kriteria KodeProduk = kode produk input user, NoSuratJalan__Tanggal__range = awal tahun, akhir tahun. 
    6. mengambil data Transaksi Gudang keluar dengan kriteria Jumlah >= 0 (mengindikasikan barang keluar), KodeProduk = kode produk input user, Tanggal dalam rentang awal tahun hingga akhir tahun
    7. mengambil data Transaksi Gudang retur dengan kriteria Jumlah < 0 (mengindikasikan barang keluar), KodeProduk = kode produk input user, Tanggal dalam rentang awal tahun hingga akhir tahun
    8. mengambil data pemusnahan bahan baku  dengan lokasi Gudang, kodeproduk = kode produk input user, Tanggal berada dalam rentanga awal tahun, akhirtahun
    9. Mengambil data saldo awal bahan baku dengan kriteria kode produk = kode produk input user, Lokasi = Gudang, Tanggal berada dalam rentang awal tahun dan akhir tahun
    10. Mengiterasi setiap tanggal dari detailsuratjalanpembelian, transaksigudang keluar, transaksigudang retur, dan pemusnahan bahan baku 
    11. Total Stok = Saldoawal bahan baku + Detail surat jalan pembelian + Transaksi Gudang Retur - pemusnahan bahan baku - transaksi gudang keluar
    12. Menghitung harga satuan = harga total / total stok (perhitungan menggunakan weighted average)
    13. Apabila ada pembagian dengan hasil Null / pembagi 0 maka akan diset menjadi 0
    '''
    kode_produk = kodeproduk
    tahun = periode
    tahun_period = tahun

    if tahun == "":
        tahun = datetime.now().year

    tahun = datetime.strptime(tahun, format("%Y"))
    awaltahun = datetime(tahun.year, 1, 1)
    akhirtahun = datetime(tahun.year, 12, 31)

    produkobj = models.Produk.objects.with_deleted().filter(KodeProduk=kode_produk).first()

    masukobj = models.DetailSuratJalanPembelian.objects.filter(
        KodeProduk__KodeProduk=produkobj.KodeProduk, NoSuratJalan__Tanggal__range=(awaltahun,akhirtahun)
    )
    tanggalmasuk = masukobj.values_list("NoSuratJalan__Tanggal", flat=True)

    keluarobj = models.TransaksiGudang.objects.filter(
        jumlah__gte=0, KodeProduk__KodeProduk=produkobj.KodeProduk, tanggal__range=(awaltahun,akhirtahun),KeteranganACC=True
    )
    returobj = models.TransaksiGudang.objects.filter(
        jumlah__lt=0, KodeProduk__KodeProduk=produkobj.KodeProduk, tanggal__range=(awaltahun,akhirtahun)
    )
    pemusnahanobj = models.PemusnahanBahanBaku.objects.filter(
        lokasi__NamaLokasi = 'Gudang',KodeBahanBaku__KodeProduk = produkobj.KodeProduk, Tanggal__range = (awaltahun,akhirtahun)
    )
    # print(pemusnahanobj)
    # print(asd)
    tanggalkeluar = keluarobj.values_list("tanggal", flat=True)
    # tanggalretur = returobj.values_list("tanggal", flat=True)
    tanggalpemusnahan = pemusnahanobj.values_list("Tanggal",flat=True)
    print("ini kode bahan baku", keluarobj)
    saldoawalobj = (
        models.SaldoAwalBahanBaku.objects.filter(
            IDBahanBaku__KodeProduk=produkobj.KodeProduk,
            IDLokasi__IDLokasi=3,
            Tanggal__range=(awaltahun,akhirtahun)
        )
        .order_by("-Tanggal")
        .first()
    )
    print(saldoawalobj)
    if saldoawalobj:
        print("ada data")
        saldoawal = saldoawalobj.Jumlah
        hargasatuanawal = saldoawalobj.Harga
        hargatotalawal = saldoawal * hargasatuanawal
        tahun = saldoawalobj.Tanggal.year

    else:
        saldoawal = 0
        hargasatuanawal = 0
        hargatotalawal = saldoawal * hargasatuanawal
        tahun = awaltahun.year

    saldoawalobj = {
        "saldoawal": saldoawal,
        "hargasatuanawal": hargasatuanawal,
        "hargatotalawal": hargatotalawal,
        'tahun' : tahun
    }
    hargaterakhir = 0
    listdata = []
    print(tanggalmasuk)
    print(tanggalkeluar)
    listtanggal = sorted(
        list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalpemusnahan)))
    )
    # listtanggal = sorted(
    #     list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalretur).union(tanggalpemusnahan)))
    # )
    print(listtanggal)
    statusmasuk = False
    for i in listtanggal:
        jumlahmasukperhari = 0
        hargamasuktotalperhari = 0
        hargamasuksatuanperhari = 0
        jumlahkeluarperhari = 0
        hargakeluartotalperhari = 0
        hargakeluarsatuanperhari = 0
        sjpobj = masukobj.filter(NoSuratJalan__Tanggal=i)
        if sjpobj.exists():
            for j in sjpobj:
                hargamasuktotalperhari += j.Harga * j.Jumlah
                jumlahmasukperhari += j.Jumlah
            try: 
                hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
            except ZeroDivisionError:
                hargamasuksatuanperhari = 0
            print("data SJP ada")
            print(hargamasuksatuanperhari)
            print(jumlahmasukperhari)
            dumy = {
                "Tanggal": i.strftime("%Y-%m-%d"),
                "Jumlahstokawal": saldoawal,
                "Hargasatuanawal": hargasatuanawal, 
                "Hargatotalawal": hargatotalawal, 
                "Jumlahmasuk": jumlahmasukperhari,
                "Hargamasuksatuan": hargamasuksatuanperhari, 
                "Hargamasuktotal": hargamasuktotalperhari, 
                "Jumlahkeluar": jumlahkeluarperhari,
                "Hargakeluarsatuan": hargakeluarsatuanperhari, 
                "Hargakeluartotal": hargakeluartotalperhari, 
            }
            saldoawal += jumlahmasukperhari - jumlahkeluarperhari
            hargatotalawal += hargamasuktotalperhari - hargakeluartotalperhari
            try:
                hargasatuanawal = hargatotalawal / saldoawal
            except ZeroDivisionError:
                hargasatuanawal = 0

            print("Sisa Stok Hari Ini : ", saldoawal)
            print("harga awal Hari Ini :", hargasatuanawal)
            print("harga total Hari Ini :", hargatotalawal, "\n")
            dumy["Sisahariini"] = saldoawal
            dumy["Hargasatuansisa"] = hargasatuanawal
            dumy["Hargatotalsisa"] = hargatotalawal
            statusmasuk = True
            listdata.append(dumy)
            # print(asdasd)

        hargamasuktotalperhari = 0
        jumlahmasukperhari = 0
        hargamasuksatuanperhari = 0
        transaksigudangobj = keluarobj.filter(tanggal=i)
        transaksipemusnahan = pemusnahanobj.filter(Tanggal=i)


        if transaksigudangobj.exists():
            for j in transaksigudangobj:
                jumlahkeluarperhari += j.jumlah
                hargakeluartotalperhari += j.jumlah * hargasatuanawal

        if transaksipemusnahan.exists():
            for j in transaksipemusnahan:
                jumlahkeluarperhari += j.Jumlah
                hargakeluartotalperhari += j.Jumlah * hargasatuanawal

        if jumlahkeluarperhari > 0:
            hargakeluarsatuanperhari = hargakeluartotalperhari / jumlahkeluarperhari
        else:
            if statusmasuk:
                statusmasuk = False
                continue
            hargakeluartotalperhari = 0
            hargakeluarsatuanperhari = 0
            jumlahkeluarperhari = 0

        # transaksigudangobj = keluarobj.filter(tanggal=i)
        # print(transaksigudangobj)
        # if transaksigudangobj.exists():
        #     for j in transaksigudangobj:
        #         jumlahkeluarperhari += j.jumlah
        #         hargakeluartotalperhari += j.jumlah * hargasatuanawal
        #     hargakeluarsatuanperhari += (
        #         hargakeluartotalperhari / jumlahkeluarperhari
        #     )
        # else:
        #     if statusmasuk:
        #         statusmasuk = False
        #         continue
        #     hargakeluartotalperhari = 0
        #     hargakeluarsatuanperhari = 0
        #     jumlahkeluarperhari = 0

        # transaksipemusnahan = pemusnahanobj.filter(Tanggal=i)
        # print(transaksipemusnahan)
        # if transaksipemusnahan.exists():
        #     for j in transaksipemusnahan:
        #         jumlahkeluarperhari += j.Jumlah
        #         hargakeluartotalperhari += j.Jumlah * hargasatuanawal
        #     hargakeluarsatuanperhari += (
        #         hargakeluartotalperhari / jumlahkeluarperhari
        #     )
        # else:
        #     if statusmasuk:
        #         statusmasuk = False
        #         continue
        #     hargakeluartotalperhari = 0
        #     hargakeluarsatuanperhari = 0
        #     jumlahkeluarperhari = 0

        # transaksireturobj = returobj.filter(tanggal=i)
        # if transaksireturobj.exists():
        #     for j in transaksireturobj:
        #         jumlahmasukperhari += j.jumlah * -1
        #         hargamasuktotalperhari += j.jumlah * hargasatuanawal * -1
        #     try:
        #         hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
        #     except ZeroDivisionError:
        #         hargamasuksatuanperhari = 0
        # else:
        #     hargamasuktotalperhari = 0
        #     hargamasuksatuanperhari = 0
        #     jumlahmasukperhari = 0


        print("Tanggal : ", i)
        print("Sisa Stok Hari Sebelumnya : ", saldoawal)
        print("harga awal Hari Sebelumnya :", hargasatuanawal)
        print("harga total Hari Sebelumnya :", hargatotalawal)
        print("Jumlah Masuk : ", jumlahmasukperhari)
        print("Harga Satuan Masuk : ", hargamasuksatuanperhari)
        print("Harga Total Masuk : ", hargamasuktotalperhari)
        print("Jumlah Keluar : ", jumlahkeluarperhari)
        print("Harga Keluar : ", hargakeluarsatuanperhari)
        print(
            "Harga Total Keluar : ", hargakeluarsatuanperhari * jumlahkeluarperhari
        )
        

        dumy = {
            "Tanggal": i.strftime("%Y-%m-%d"),
            "Jumlahstokawal": saldoawal,
            "Hargasatuanawal": hargasatuanawal, 
            "Hargatotalawal": hargatotalawal, 
            "Jumlahmasuk": jumlahmasukperhari,
            "Hargamasuksatuan": hargamasuksatuanperhari, 
            "Hargamasuktotal": hargamasuktotalperhari, 
            "Jumlahkeluar": jumlahkeluarperhari,
            "Hargakeluarsatuan": hargakeluarsatuanperhari, 
            "Hargakeluartotal": hargakeluartotalperhari, 
        }
        """
        Rumus dari Excel KSBB Purchasing
        Sisa = Sisa hari sebelumnya + Jumlah masuk hari ini - jumlah keluar hari ini 
        harga sisa satuan = total sisa / harga total sisa
        Harga keluar = harga satuan hari sebelumnya

        """
        dummysaldoawal = saldoawal
        dummyhargatotalawal = hargatotalawal
        dummyhargasatuanawal = hargasatuanawal

        saldoawal += jumlahmasukperhari - jumlahkeluarperhari
        hargatotalawal += hargamasuktotalperhari - hargakeluartotalperhari
        print(hargatotalawal, saldoawal)
        try:
            hargasatuanawal = hargatotalawal / saldoawal
        except ZeroDivisionError:
            hargasatuanawal = 0

        print("Sisa Stok Hari Ini : ", saldoawal)
        print("harga awal Hari Ini :", hargasatuanawal)
        print("harga total Hari Ini :", hargatotalawal, "\n")
        dumy["Sisahariini"] = saldoawal
        dumy["Hargasatuansisa"] = hargasatuanawal
        dumy["Hargatotalsisa"] = hargatotalawal

        listdata.append(dumy)

    hargaterakhir += hargasatuanawal
    return listdata,saldoawalobj,hargaterakhir

def exportexcelksbb(request,kodeproduk,periode):
    '''
    fitur ini digunakan untuk melakukan eksport data KSBB kedalam excel
    Algoritma
    1. Mengambil data Produk
    2. Menghitung menggunakan fungsi getksbbproduk
    3. Mengeksport kedalam excel
    '''
    waktuawalproses = time.time()
    kodeprodukobj =models.Produk.objects.get(KodeProduk = kodeproduk)
    listdata,saldoawalobj,hargaterakhir = getksbbproduk(kodeproduk,periode)
    waktupersediaan = time.time()
    datamodelsksbb ={
 
        "Tanggal":[],
        "Kuantitas Masuk":[],
        "Harga Satuan Masuk":[],
        "Harga Total Masuk":[],
        "Kuantitas Keluar":[],
        "Harga Satuan keluar":[],
        "Harga Total keluar":[],
        "Kuantitas Sisa":[],
        "Harga Satuan Sisa":[],
        "Harga Total Sisa":[],
    }
    if saldoawalobj:
        datamodelsksbb["Tanggal"].append(date(saldoawalobj['tahun'],1,1).strftime('%Y-%m-%d'))
        datamodelsksbb["Kuantitas Masuk"].append('')
        datamodelsksbb['Harga Satuan Masuk'].append('')
        datamodelsksbb['Harga Total Masuk'].append('')
        datamodelsksbb["Kuantitas Keluar"].append('')
        datamodelsksbb['Harga Satuan keluar'].append('')
        datamodelsksbb['Harga Total keluar'].append('')
        datamodelsksbb["Kuantitas Sisa"].append(saldoawalobj['saldoawal'])
        datamodelsksbb['Harga Satuan Sisa'].append(saldoawalobj['hargasatuanawal'])
        datamodelsksbb['Harga Total Sisa'].append(saldoawalobj['hargatotalawal'])
    for item in listdata:
        datamodelsksbb["Tanggal"].append(item['Tanggal'])
        datamodelsksbb["Kuantitas Masuk"].append(item['Jumlahmasuk'])
        datamodelsksbb['Harga Satuan Masuk'].append(item['Hargamasuksatuan'])
        datamodelsksbb['Harga Total Masuk'].append(item['Hargamasuktotal'])
        datamodelsksbb["Kuantitas Keluar"].append(item['Jumlahkeluar'])
        datamodelsksbb['Harga Satuan keluar'].append(item['Hargakeluarsatuan'])
        datamodelsksbb['Harga Total keluar'].append(item['Hargakeluartotal'])
        datamodelsksbb["Kuantitas Sisa"].append(item['Sisahariini'])
        datamodelsksbb['Harga Satuan Sisa'].append(item['Hargasatuansisa'])
        datamodelsksbb['Harga Total Sisa'].append(item['Hargatotalsisa'])
    dfksbb = pd.DataFrame(datamodelsksbb)
    if dfksbb.empty:
        messages.error(request,f'Tidak ditemukan data transaksi')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return redirect('rekapharga')
    print(dfksbb)

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfksbb.to_excel(writer, index=False, startrow=5, sheet_name=f"{kodeprodukobj.KodeProduk}")
        writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {kodeprodukobj.KodeProduk}")
        writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {kodeprodukobj.NamaProduk}")
        writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {kodeprodukobj.unit}")
        maxrow = len(dfksbb)+1
        maxcol = len(dfksbb.columns)
        apply_number_format(writer.sheets[f"{kodeprodukobj.KodeProduk}"],6,maxrow+5,1,maxcol)
        apply_borders_thin(writer.sheets[f"{kodeprodukobj.KodeProduk}"],6,maxrow+5,maxcol)
        adjust_column_width(writer.sheets[f"{kodeprodukobj.KodeProduk}"],dfksbb,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBB {kodeprodukobj.KodeProduk}.xlsx"
    )
    
    print('Waktu generate laporan : ',time.time()-waktupersediaan)
    print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def exportkeseluruhanksbb (request,periode):
    '''
    Fitur ini digunakan untuk  mengeksport KSBB keseluruhan produk
    Algoritma
    1. Mengambil semua data produk
    2. Mengiterasi semua data produk 
    3. Menghitung KSBB dengan fungsi getksbbproduk
    4. Membuat datamodels representasi file excel
    5. Membuat dataframe dari datamodels
    6. Menambahkan dataframe kedalam listdataframe
    7. Mengiterasi semua data listdataframe untuk dieksport
    '''
    waktuawalproses = time.time()
    allproduk = models.Produk.objects.all().order_by('KodeProduk')
    print(allproduk)

    # allproduk = models.Produk.objects.filter(KodeProduk__in=['A-001-01','A-001-02','A-001-04'])
    listdataframe = []
    listkodestok = []
    for produk in allproduk:
        print('Produk', produk)
        listdata,saldoawalobj,hargaterakhir = getksbbproduk(produk.KodeProduk,periode)
        print('Done iterate')
        datamodelsksbb ={
 
        "Tanggal":[],
        "Kuantitas Masuk":[],
        "Harga Satuan Masuk":[],
        "Harga Total Masuk":[],
        "Kuantitas Keluar":[],
        "Harga Satuan keluar":[],
        "Harga Total keluar":[],
        "Kuantitas Sisa":[],
        "Harga Satuan Sisa":[],
        "Harga Total Sisa":[],
        }
        if saldoawalobj:
            datamodelsksbb["Tanggal"].append(date(saldoawalobj['tahun'],1,1).strftime('%Y-%m-%d'))
            datamodelsksbb["Kuantitas Masuk"].append('')
            datamodelsksbb['Harga Satuan Masuk'].append('')
            datamodelsksbb['Harga Total Masuk'].append('')
            datamodelsksbb["Kuantitas Keluar"].append('')
            datamodelsksbb['Harga Satuan keluar'].append('')
            datamodelsksbb['Harga Total keluar'].append('')
            datamodelsksbb["Kuantitas Sisa"].append(saldoawalobj['saldoawal'])
            datamodelsksbb['Harga Satuan Sisa'].append(saldoawalobj['hargasatuanawal'])
            datamodelsksbb['Harga Total Sisa'].append(saldoawalobj['hargatotalawal'])
        for item in listdata:
            datamodelsksbb["Tanggal"].append(item['Tanggal'])
            datamodelsksbb["Kuantitas Masuk"].append(item['Jumlahmasuk'])
            datamodelsksbb['Harga Satuan Masuk'].append(item['Hargamasuksatuan'])
            datamodelsksbb['Harga Total Masuk'].append(item['Hargamasuktotal'])
            datamodelsksbb["Kuantitas Keluar"].append(item['Jumlahkeluar'])
            datamodelsksbb['Harga Satuan keluar'].append(item['Hargakeluarsatuan'])
            datamodelsksbb['Harga Total keluar'].append(item['Hargakeluartotal'])
            datamodelsksbb["Kuantitas Sisa"].append(item['Sisahariini'])
            datamodelsksbb['Harga Satuan Sisa'].append(item['Hargasatuansisa'])
            datamodelsksbb['Harga Total Sisa'].append(item['Hargatotalsisa'])
        dfksbb = pd.DataFrame(datamodelsksbb)
        if not dfksbb.empty:
            listdataframe.append(dfksbb)
            listkodestok.append(produk)
        print(dfksbb)
    buffer = BytesIO()
    waktupersediaan = time.time()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for kodestok,data in zip(listkodestok,listdataframe):

        # Laporan Persediaan Section
            # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
            data.to_excel(writer, index=False, startrow=5, sheet_name=kodestok.KodeProduk)
            writer.sheets[kodestok.KodeProduk].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU {kodestok.KodeProduk}")
            writer.sheets[kodestok.KodeProduk].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {kodestok.NamaProduk}")
            writer.sheets[kodestok.KodeProduk].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {kodestok.unit}")
            maxrow = len(data)+1
            maxcol = len(data.columns)
            apply_number_format(writer.sheets[kodestok.KodeProduk],6,maxrow+5,1,maxcol)
            apply_borders_thin(writer.sheets[kodestok.KodeProduk],6,maxrow+5,maxcol)
            adjust_column_width(writer.sheets[kodestok.KodeProduk],data,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBB Purchasing - {periode}.xlsx"
    )
    
    print('Waktu generate laporan : ',time.time()-waktupersediaan)
    print('Waktu Proses : ', time.time()-waktuawalproses)
    return response


        
def apply_number_format(worksheet, start_row, end_row, start_col, end_col, number_format='#,##0.00'):
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Apply format only to numeric cells
                cell.number_format = number_format

def adjust_column_width(worksheet, df, start_row, start_col):
    for i, col in enumerate(df.columns):
        max_length = max(df[col].astype(str).map(len).max(), len(col))
        col_letter = get_column_letter(start_col + i)
        worksheet.column_dimensions[col_letter].width = max_length + 4

def apply_borders_thin(worksheet, start_row, end_row, max_col,min_col=1):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def exportbarangsubkon_excel(request):
    '''
    Fitur ini digunakan untuk eksport produk subkon (register subkon)
    Algoritma
    1. Mengambil data produk subkon
    2. Mengiterasi produk subkon 
    3. Apabila status potongan transaksi detail surat jalan penerimaan = True, maka harga akan dipotong, apabila False maka tidak dipotong
    4. Eksport ke Excel
    '''
    valueppn = request.GET.get("input_ppn", 2)
    try:
        valueppn = int(valueppn)
        if valueppn < 0:
            valueppn = 2
            messages.error(request, "Nilai Persentase Minus!")
    except ValueError:
        valueppn = 2
        messages.error(request, "Nilai PPN tidak valid!")

    inputppn = valueppn / 100
    print("INPUT PPN NI BANGG", inputppn)

    # Ambil nilai input_awal dan input_terakhir dari query string
    input_awal = request.GET.get("input_awal")
    input_terakhir = request.GET.get("input_terakhir")
    print("Ini input awal ama akhir", input_awal, input_terakhir)

    # Validasi format tanggal
    date_awal = parse_date(input_awal) if input_awal else None
    date_terakhir = parse_date(input_terakhir) if input_terakhir else None

    if date_awal is None or date_terakhir is None:
        sjball = models.DetailSuratJalanPenerimaanProdukSubkon.objects.all().order_by("NoSuratJalan__Tanggal")
        print("len = 0, sjb all =", sjball)
    else:
        sjball = models.DetailSuratJalanPenerimaanProdukSubkon.objects.filter(
            NoSuratJalan__Tanggal__range=(date_awal, date_terakhir)
        ).order_by("NoSuratJalan__Tanggal")

    print('TES SJB BANG', sjball)

    # Buat Workbook baru
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Barang Masuk'

    # Definisikan header untuk worksheet
    headers = ['Tanggal', 'Supplier', 'Kode Bahan Baku Subkon', 'Nama Bahan Baku Subkon', 'Satuan', 'Kuantitas', 'Harga', 'Harga Total', f'Harga Potongan ', 'Harga Total Setelah Potongan', 'Tanggal Invoice', 'No Invoice']
    worksheet.append(headers)

    # Tambahkan data ke worksheet
    total_potongan=0
    total_harga = 0
    totalharga_potongan = 0
    for item in sjball:
        harga_total = item.Jumlah * item.Harga
        item.hargatotalsebelumpotongan = item.Harga * item.Jumlah
        if item.Potongan:
            item.hargatotalsetelahpemotongan = item.hargapotongan * item.Jumlah
            # harga_potongan = harga_total * inputppn
            # harga_satuan_setelah_pemotognan = math.ceil(item.Harga - (item.Harga * inputppn))
        else:
            item.hargapotongan = 0
            item.hargatotalsetelahpemotongan = item.hargatotalsebelumpotongan
        #     harga_potongan = 0
        #     harga_satuan_setelah_pemotognan = 0
        # harga_total_setelah_potongan = harga_satuan_setelah_pemotognan * item.Jumlah
        total_harga +=harga_total
        # print('HARGA POTONGAN :',harga_potongan,total_potongan)
        total_potongan+=item.hargapotongan
        totalharga_potongan += item.hargatotalsetelahpemotongan
        row = [
            item.NoSuratJalan.Tanggal.strftime("%Y-%m-%d"),
            str(item.NoSuratJalan.Supplier),
            str(item.KodeProduk),
            str(item.KodeProduk.NamaProduk),
            str(item.KodeProduk.Unit),  # Assuming you have a field for 'Satuan'
            item.Jumlah,
            item.Harga,
            harga_total,
            item.hargapotongan,
            item.hargatotalsetelahpemotongan,
            item.NoSuratJalan.TanggalInvoice.strftime("%Y-%m-%d") if item.NoSuratJalan.TanggalInvoice else '',
            str(item.NoSuratJalan.NoInvoice) if item.NoSuratJalan.NoInvoice else ''
        ]
        worksheet.append(row)
    total_row = [
        "",  # Kosongkan kolom Tanggal
        "TOTAL",  # Tulis "TOTAL" di kolom Suppliers
        "",  # Kosongkan kolom Kode Bahan Baku
        "",  # Kosongkan kolom Nama Bahan Baku
        "",  # Kosongkan kolom Kuantitas
        "",
        
        "",  # Kosongkan kolom Harga    
        total_harga,  # Total Harga
        total_potongan,  # Total Harga PPN
        totalharga_potongan,  # Total Harga Total PPN
        "",  # Kosongkan kolom Tanggal Invoice
        ""  # Kosongkan kolom No Invoice
    ]
    worksheet.append(total_row)

    # Menyesuaikan lebar kolom
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    # Menambahkan warna pada header
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = header_fill

    # Menambahkan border pada semua sel
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
    for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=10):  # mulai dari baris 2 dan kolom 5 (Kuantitas)
        for cell in row:
            if cell.column in [6,7, 8, 9, 10]:  # kolom Kuantitas dan Harga
                cell.number_format = '#,##0.00'

    # Atur response untuk mengunduh file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Bahan Baku Subkon Masuk {input_awal} - {input_terakhir}.xlsx'
    workbook.save(response)

    return response


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def views_rekaphargasubkon(request):
    '''
    Fitur ini digunakan untuk melihat register produk subkon
    Algoritma 
    1. Mengambil data Detail Surat Jalan Penerimaan Produk Subkon dengan kriteria Tanggal pada surat jalan berada dalam rentang tanggal mulai dan tanggal akhir
    2. Mengiterasi Detail surat jalan penerimaan produk subkon
    3. Mengubah format tanggal menjadi YYYY-MM-DD
    '''
    # if request.method ==  'POST' :
    if len(request.POST) == 0 :
        valueppn = 2
    else :

        valueppn = request.POST["input_ppn"]
        if int(valueppn) < 0 :
            valueppn = 2
            messages.error(request,"Nilai Persentase Minus!") 
        else :
            pass

    inputppn = int(valueppn)/100

    print("inputppn",inputppn)
 

    if len(request.GET) == 0:
        sekarang = datetime.now().date()
        input_awal = date(sekarang.year,sekarang.month,1).strftime('%Y-%m-%d')
        input_terakhir = sekarang.strftime('%Y-%m-%d')
        
        harga_setelah_ppn = []
        list_total_ppn = []
        sjball = models.DetailSuratJalanPenerimaanProdukSubkon.objects.filter(NoSuratJalan__Tanggal__range=(input_awal,input_terakhir)).order_by(
            "NoSuratJalan__Tanggal"
        )
        print(sjball)
        # print(asd)
        if len(sjball) > 0:
       
            # for x in sjball:
            #     harga_total = x.Jumlah * x.Harga
            #     x.NoSuratJalan.Tanggal = x.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")
            #     if x.NoSuratJalan.TanggalInvoice is not None:
            #         x.NoSuratJalan.TanggalInvoice = x.NoSuratJalan.TanggalInvoice.strftime("%Y-%m-%d")
            #     print(harga_total)
            #     list_harga_total1.append(harga_total)
            #     harga_setelah_pemotongan = math.ceil(x.Harga - (x.Harga * inputppn))
            #     total_harga_setelah_pemotongan = harga_setelah_pemotongan * x.Jumlah
            #     harga_setelah_ppn.append(harga_setelah_pemotongan)
            #     list_total_ppn.append(total_harga_setelah_pemotongan)
            # i = 0
            # for item in sjball:
            #     item.harga_total = list_harga_total1[i]
            #     if item.Potongan:
            #         item.harga_ppn = harga_setelah_ppn[i]
            #         item.harga_total_ppn = list_total_ppn[i]
            #     else:
            #         item.harga_ppn = 0
            #         item.harga_total_ppn =0

            #     i += 1
            # print("list hartot", list_harga_total1)
            for item in sjball:
                item.NoSuratJalan.Tanggal = item.NoSuratJalan.Tanggal.strftime('%Y-%m-%d')
                item.hargatotalsebelumpotongan = item.Harga * item.Jumlah
                item.hargatotalsetelahpemotongan = item.hargapotongan * item.Jumlah
                if not item.Potongan:
                    item.hargatotalsetelahpemotongan = item.Harga * item.Jumlah
                    item.hargapotongan = 0
                    
            
        
            return render(
                request,
                "Purchasing/masuk_subkon.html",
                {
                    "sjball": sjball,
                    # "harga_total": harga_total,
                    "valueppn" : valueppn,
                                    "input_awal": input_awal,
                    "input_terakhir": input_terakhir,
                },
            )
        else:
            messages.error(request, "Data tidak ditemukan")
            return render(
                request,
                "Purchasing/masuk_subkon.html",
            )
    else:
        # if request.method == "POST": 
            
        # else :
        input_awal = request.GET["awal"]
        input_terakhir = request.GET["akhir"]
        # valueppn = request.POST["input_ppn"]
        # inputppn = request.POST["input_ppn"]
        list_harga_total = []
        list_ppn_1 = []
        list_total_ppn = []
        harga_setelah_ppn = []
        try :
            filtersjb = models.DetailSuratJalanPenerimaanProdukSubkon.objects.filter(
                NoSuratJalan__Tanggal__range=(input_awal, input_terakhir)
            ).order_by("NoSuratJalan__Tanggal")
        except models.DetailSuratJalanPenerimaanProdukSubkon.DoesNotExists :
            messages.error(request, "Data tidak ditemukan")
        if len(filtersjb) > 0:

            for item in filtersjb:
                item.NoSuratJalan.Tanggal = item.NoSuratJalan.Tanggal.strftime('%Y-%m-%d')
                item.hargatotalsebelumpotongan = item.Harga * item.Jumlah
                item.hargatotalsetelahpemotongan = item.hargapotongan * item.Jumlah
                if not item.Potongan:
                    item.hargatotalsetelahpemotongan = item.Harga * item.Jumlah
                    item.hargapotongan = 0
            return render(
                request,
                "Purchasing/masuk_subkon.html",
                {
                    "data_hasil_filter": filtersjb,
                    # "harga_total": harga_total,
                    "input_awal": input_awal,
                    "input_terakhir": input_terakhir,
                    "valueppn" : valueppn
                },
            )
        else:
            messages.error(request, "Data tidak ditemukan")
            return redirect("rekaphargasubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def accspk2(request, id):
    '''
    Fitur ini digunakan untuk melakukan ACC SPK pada halaman menu awal fitur SPK
    Algoritma
    1. Mengambil data SPK dengan kriteria id (Primary Key) =  seama seperti id(id didapatkan dari passing values HTML)
    2. Mengubah status keteranganACC menjadi True
    3. Menyimpan perubahan
    '''
    accobj = models.SPK.objects.get(pk=id)
    accobj.KeteranganACC = True
    accobj.save()
    models.transactionlog(
        user="Purchasing",
        waktu=datetime.now(),
        jenis="ACC",
        pesan=f"No SPK {accobj.NoSPK} sudah ACC",
    ).save()
    messages.success(request, "SPK berhasil diacc")
    return redirect("read_spk")

'''

REVISI 6/12/2024
1. Update Harga Saldo Awal 
'''
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing",'ppic'])
def read_saldoawal(request):
    '''
    Fitur ini digunakan untuk manajemen saldo awal bahan baku pada purchasing
    Algoritma
    1. Mendapatkan data Saldo Awal Bahan Baku pada sistem 
    2. Mengubah tanggal dengan format YYYY-MM-DD
    '''
    dataproduk = models.SaldoAwalBahanBaku.objects.all().order_by("-Tanggal")
    print(dataproduk)
    for i in dataproduk:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "Purchasing/read_saldoawalbahan.html", {"dataproduk": dataproduk}
    )
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def update_saldoawal(request,id):
    '''
    Fitur ini digunakan untuk mengupdate data saldoawal bahan baku
    Algoritma
    1. Mengambil data Saldoawal Bahan baku dalam sistem dengan kriteria IDSaldoAwalBahanBaku = id (id didapatkan dari passing values HTML)
    2. Mengubah format tanggal menjadi YYYY-MM-DD
    3. Menampilkan form update saldo awal
    4. Program menerima input Harga satuan
    5. Mengupdate harga 
    6. Menyimpan update
    '''
    databarang = models.Produk.objects.all()
    dataobj = models.SaldoAwalBahanBaku.objects.get(IDSaldoAwalBahanBaku=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "Purchasing/update_saldobahan.html",
            {"data": dataobj, "nama_lokasi": lokasiobj, "databarang": databarang},
        )

    else:

        harga = request.POST["harga"]
        
        hargalama = dataobj.Harga
        dataobj.Harga = harga
        models.transactionlog(
            user="Purchasing",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Update Harga Bahan Baku {dataobj.IDBahanBaku.KodeProduk}Jumlah lama {dataobj.Harga}, Jumlah baru {hargalama}",
        ).save()
        dataobj.save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("saldobahanbakupurchasing")


# @login_required
# @logindecorators.allowed_users(allowed_roles=['produksi'])

# @login_required
# @logindecorators.allowed_users(allowed_roles=['produksi'])



# @login_rexs.allowed_users(allowed_roles=['produksi'])

def bulk_createproduk(request):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        df = pd.read_excel(file, engine='openpyxl')
        df = df.fillna('-')
        for index,row in df.iterrows():
            bahanbakubaru = models.Produk(
                KodeProduk = row['Kode Stock'],
                NamaProduk = row['Nama Barang'],
                unit = row['Satuan'],
                TanggalPembuatan = datetime.now(),
                Jumlahminimal = 0,
                keteranganPurchasing = row['Keterangan']
            )
            print(bahanbakubaru.keteranganPurchasing)
            bahanbakubaru.save()
        return HttpResponse("Data successfully uploaded and processed!")
    
    return render(request, 'Purchasing/bulk_createproduk.html')


def bulk_createsjp(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        print(sheet_names)
        for item in sheet_names:
            data = models.Artikel(KodeArtikel=item, keterangan="-")
            print(data)
            data.save()
        return HttpResponse("Berhasil Upload")

    return render(request, "Purchasing/bulk_createproduk.html")


# Saldo Awal Artikel
@login_required
@logindecorators.allowed_users(allowed_roles=['purchasing','ppic'])
def view_saldoartikel(request):
    '''
    Fitur ini digunakan untuk melihat saldo awal artikel yang ada pada sistem
    Algoritma
    1. Mengambil data saldo awal artikel 
    2. Mengubah Tanggal dengan format YYYY-MM-DD
    '''
    dataartikel = models.SaldoAwalArtikel.objects.all().order_by("-Tanggal")
    for i in dataartikel:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "Purchasing/views_saldoartikel.html", {"dataartikel": dataartikel}
    )

# Saldo Bahan Subkon
@login_required
@logindecorators.allowed_users(allowed_roles=['purchasing','ppic'])
def view_saldobahansubkon(request):
    '''
    Fitur ini digunakan untuk melihat saldo awal bahan baku subkon yang ada pada sistem
    Algoritma
    1. Mengambil data saldo awal bahan baku subkon 
    2. Mengubah Tanggal dengan format YYYY-MM-DD
    '''
    datasubkon = models.SaldoAwalBahanBakuSubkon.objects.all().order_by("-Tanggal")
    for i in datasubkon:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "Purchasing/views_saldobahansubkon.html", {"datasubkon": datasubkon}
    )

# Saldo Awal Produk Subkon
@login_required
@logindecorators.allowed_users(allowed_roles=['purchasing','ppic'])
def view_saldosubkon(request):
    '''
    Fitur ini digunakan untuk melihat saldo awal produk subkon yang ada pada sistem
    Algoritma
    1. Mengambil data saldo awal produk subkon 
    2. Mengubah Tanggal dengan format YYYY-MM-DD
    '''
    datasubkon = models.SaldoAwalSubkon.objects.all().order_by("-Tanggal")
    for i in datasubkon:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "Purchasing/views_saldoproduksubkon.html", {"datasubkon": datasubkon}
    )


# Purchase Order
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing","ppic"])
def view_purchaseorder(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data Purchase Order pada sistem
    Algoritma 
    1. Mengambil data semua purchase order pada sistem
    2. Mengiterasi semua data PO
    3. mengambil data detail PO dengan kriteria KodePO = nomor PO
    '''
    datapo = models.PurchaseOrder.objects.all().order_by("Tanggal")
    # datasjb = models.DetailSuratJalanPembelian.objects.all().order_by(
    #     "NoSuratJalan__Tanggal"
    # )
    date = request.GET.get("mulai")
    dateakhir = request.GET.get("akhir")
    print(date, dateakhir)
    if date =="":
        date = datetime.min.date()
    if dateakhir == "":
        dateakhir = datetime.max.date()
    print(date,dateakhir)
    if date is not None and dateakhir is not None:
        datapo = models.PurchaseOrder.objects.filter(
            Tanggal__range=(date, dateakhir)
        ).order_by("Tanggal")
    for item in datapo:
        item.detailpo = models.DetailPO.objects.filter(KodePO = item.pk)
        item.Tanggal = item.Tanggal.strftime('%Y-%m-%d')
    # for i in date:
    #     i.KodePO.Tanggal = i.KodePO.Tanggal.strftime("%Y-%m-%d")

    if len(datapo) == 0:
        messages.info(request, "Tidak ada data PO")

    return render(
        request,
        "purchasing/purchaseorder.html",
        {"datasjb": datapo, "date": date, "mulai": date, "akhir": dateakhir},
    )


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def trackingpurchaseorder(request,id):
    '''
    Fitur ini digunakan untuk melakukan tracking pada Purchase Order yang ada pada sistem
    1. Mengambil data purchase order dengan kriteria id (primary key) = id (id didapatkan dari passing values HTML)
    2. Mengambil data detail PO dengan kriteria KodePO = data purchase order poin 1
    3. Mengambil data transaksi detail surat jalan pembelian yang menginputkan dengan kode po yang dipilih. 
    4. Mengiterasi data detail PO
    5. mencari selisih dengan cara selisih = Total jumlah PO - TOtal transaksi gudang masuk. Selisih PO artininya selisih kurangnya PO yang belum datang ke perusahaan
    

    '''
    datapo = models.PurchaseOrder.objects.get(id=id)
    datadetailpo =models.DetailPO.objects.filter(KodePO = datapo)
    transaksigudang = models.DetailSuratJalanPembelian.objects.filter(PO__KodePO= datapo)
    # Rekap dan kurang
    datarekap = transaksigudang.values('pk','KodeProduk__KodeProduk',"KodeProduk__NamaProduk","KodeProduk__unit").annotate(total = Sum('Jumlah'))
    print(datarekap)
    print(datadetailpo)
    # print(asd)
    datapo.Tanggal = datapo.Tanggal.strftime('%Y-%m-%d')
    for item in datadetailpo:
        totalpo = item.Jumlah
        totaltransaksigudangmasuk = 0
        transaksigudangmasuk = transaksigudang.filter(PO__pk = item.pk)
        if transaksigudangmasuk.exists():
            totaltransaksigudangmasuk = transaksigudangmasuk.aggregate(total = Sum('Jumlah'))['total']
        selisih = totalpo-totaltransaksigudangmasuk
        item.jumlahmasuk = totaltransaksigudangmasuk
        item.selisih = selisih

    return render(request,'Purchasing/trackingpo.html',{'datapo':datapo,'datadetailpo':datadetailpo,"transaksigudang":transaksigudang,})


@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def add_purchaseorder(request):
    '''
    Fitur ini digunakan untuk menambahkan data purchase order
    Algoritma : 
    1. mengambil semua data purchase order
    2. Mengambil semua data produk
    3. Menampilkan form input data purchase order
    4. program mendapatkan input berupa Nomor PO, Tanggal, list kode produk, list jumlah
    5. Membuat Purchase Order object dengan  kriteria Nomor Po = input nomor PO,Tanggal = input tanggal, dan Status = False (False = status aktif, True = status non aktif / lunas)
    6. Mengiterasi list kode produk dan list jumlah
    7. membuat object sejumlah kode produk dan jumlah dengan nomor po adalah Object Purchase Order
    8. menyimpan data
    '''
    if request.method == "GET":
        detailsj = models.PurchaseOrder.objects.all()
        getproduk = models.Produk.objects.all()

        return render(
            request,
            "Purchasing/addpurchaseorder.html",
            { "detailsj": detailsj, "getproduk": getproduk},
        )
    elif request.method == "POST":
        print(request.POST)
        # print(asd)
        kode = request.POST.getlist("kodeproduk")
        kodepo = request.POST["nomorpo"]
        tanggal = request.POST["tanggal"]
        supplier = request.POST['supplier']

        
    #     existing_entry = models.SuratJalanPembelian.objects.filter
    # (NoSuratJalan=kodepo).exists()
    #     print(existing_entry)
    #     # print(asd)
    #     if existing_entry:
    #         messages.warning(request,(f'No Surat Jalan {kodepo} sudah terdaftar pada sistem'))
    #         return redirect("addgudang")
        # Cari data kode po yang sama 
        purchaseorderobj = models.PurchaseOrder.objects.filter(KodePO = kodepo)
        if purchaseorderobj.exists():
            messages.error(request,f'Kode PO {kodepo} telah terdaftar pada sistem')
            return redirect('add_purchaseorder')
        nomorpoobj = models.PurchaseOrder(
            KodePO=kodepo, Tanggal=tanggal, Status = False,Supplier = supplier
        )
        listkodeproduk = request.POST.getlist("kodeproduk")
        error = 0
        for i in listkodeproduk:
            try:
                kodeproduk = models.Produk.objects.get(KodeProduk=i)
            except:
                error += 1
        if error == len(listkodeproduk):
            messages.error(request, "Data tidak ditemukan dalam sistem")
            return redirect("add_purchaseorder")
        nomorpoobj.save()
        nomorpoobj = models.PurchaseOrder.objects.get(
            KodePO=kodepo
        )
        for kodeproduk, jumlah in zip(
            request.POST.getlist("kodeproduk"), request.POST.getlist("jumlah")
        ):
            # print(kodeproduk)
            try:
                kodeprodukobj = models.Produk.objects.get(KodeProduk=kodeproduk)
            except:
                messages.error(
                    request, f"Data Bahan Baku {kodeproduk} tidak terdapat dalam sistem"
                )
                continue
            newprodukobj = models.DetailPO(
                KodeProduk=kodeprodukobj,
                Jumlah=jumlah,
                KodePO = nomorpoobj
            )
            models.transactionlog(
                user="Purchasing",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"No PO: {nomorpoobj.KodePO} Kode Barang : {newprodukobj.KodeProduk} Jumlah : {newprodukobj.Jumlah}",
            ).save()
            newprodukobj.save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("view_purchaseorder")

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def update_purchaseorder(request, id):
    '''
    Fitur ini digunakan untuk mengupdate data PO
    Algoritma : 
    1. Mengambil data PO dengan kriteria Primarykey = id (id didapatkan dari passing values HTML)
    2. Mengambil data detail PO
    3. mengubah format tanggal menjadi YYYY-MM-DD
    4. Menampilkan form update PO
    5. Program menerima input berupa Status, Tanggal, Kode PO 
    6. Program mengupdate data Purchase Order
    7. Program menerima input berupa list id existing detail po, list existing jumlah detail po, list existing kodeproduk existing
    8. Mengiterasi list id detailpo, list Jumlah, dan list kode produk
    9. Mengupdate data detail PO Existing
    10. Program menerima input berupa list kode produk baru, list jumlah baru
    11. Mengiterasi list produk baru dan list jumlah baru
    13. menyimpan detail purchase order dengan produk baru dan jumlah baru denga Kode PO = data po poin 1
    '''
    datapo = models.PurchaseOrder.objects.get(pk=id)
    datapo.detailpo = models.DetailPO.objects.filter(KodePO = datapo)
    getproduk = models.Produk.objects.all()
    datapo.Tanggal = datapo.Tanggal.strftime('%Y-%m-%d')
    
    
    if request.method == "GET":

        return render(
            request,
            "Purchasing/update_purchaseorder.html",
            {

                "datapo": datapo,
               
                "getproduk": getproduk,
            },
        )

    else:
        tanggal = request.POST["tanggal"]
        print(request.POST)

        # print(asd)
        kodepo = request.POST.get("kodepo")
        supplier = request.POST['supplier']
        # try:
        #     kode_produkobj = models.Produk.objects.get(KodeProduk=kode_produk)
        # except:
        #     messages.error(request, f"Data bahan baku {kode_produk} tidak ditemukan")
        #     return redirect("update_purchaseorder", id=id)
        status = request.POST.get('status')
        jumlah = request.POST.get("Jumlah")
        status = request.POST.get('status')
        cekexistingrecoed = models.PurchaseOrder.objects.filter(KodePO = kodepo).exclude(pk = id).exists()
        if cekexistingrecoed:
            messages.error(request, f"Kode PO {kodepo} Sudah ada dalam sistem")
            return redirect("update_purchaseorder", id=id)
        
        datapo.KodePO = kodepo
        datapo.Tanggal = tanggal
        datapo.Status = (status)
        datapo.Supplier = supplier
        datapo.save()

        # iterasi existing obj 
        idexisting = request.POST.getlist('idexisting')
        kodeprodukexisting = request.POST.getlist('kodeprodukexisting')
        jumlahexisting = request.POST.getlist('jumlahexisting')
        print(idexisting,kodeprodukexisting,jumlahexisting)

        kodeprodukbaru = request.POST.getlist('kodeproduk[]')
        listjumlahbaru = request.POST.getlist('jumlah[]')
        for idexist,kodeproduk,jumlah in zip(idexisting,kodeprodukexisting,jumlahexisting):
            try:
                produkobj = models.Produk.objects.get(KodeProduk = kodeproduk)
            except models.Produk.DoesNotExist:
                messages.error(request,f"Kode Bahan Baku {kodeproduk} tidak ditemukan")
                continue
            detailpoobj = models.DetailPO.objects.get(pk=idexist)
            detailpoobj.KodeProduk = produkobj
            detailpoobj.Jumlah = jumlah
            print(idexist,kodeproduk,jumlah)
            detailpoobj.save()
        
        for kodeproduk,jumlahbaru in zip(kodeprodukbaru,listjumlahbaru):
            try:
                produkobj = models.Produk.objects.get(KodeProduk = kodeproduk)
            except models.Produk.DoesNotExist:
                messages.error(request,f"Kode Bahan Baku {kodeproduk} tidak ditemukan")
                continue
            detailpoobj = models.DetailPO(
                KodePO = datapo,
                KodeProduk = produkobj,
                Jumlah = jumlahbaru
            )
            detailpoobj.save()
            

        # print(asd)
        # datasjp.save()
        # datasjp.NoSuratJalan.save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("view_purchaseorder")

@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def delete_purchaseorder(request, id):
    '''
    Fitur ini digunakan untuk menghapus data Purchase Order
    Algoritma
    1. Mengambil data Purchase Order dengan PrimaryKey = id (id didapatkan dari passing values HTML)
    2. menghapus data Purchase ORder
    '''
    datasbj = models.PurchaseOrder.objects.get(pk=id)
    models.transactionlog(
        user="Purchasing",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Purchase Order: {datasbj.KodePO} Tanggal : {datasbj.Tanggal} ",
    ).save()
    datasbj.delete()
    messages.success(request,'Data berhasil dihapus')
    return redirect("view_purchaseorder")
@login_required
@logindecorators.allowed_users(allowed_roles=["purchasing"])
def delete_detailpurchaseorder(request, id):
    '''
    Fitur ini digunakan untuk menghapus data Detail Purchase Order
    Algoritma 
    1. Mengambil detail PO dengan Primary Key = id (id didapatkan dari passing values HTML)
    2. menghapus data detail PO 
    '''
    datapo = models.DetailPO.objects.get(pk=id)
    idpo = datapo.KodePO.pk
    models.transactionlog(
        user="Purchasing",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Purchase Order: {datapo.KodePO} Tanggal : {datapo.KodePO.Tanggal} ",
    ).save()
    datapo.delete()
    messages.success(request,'Data berhasil dihapus')
    return redirect("update_purchaseorder",id=idpo)


