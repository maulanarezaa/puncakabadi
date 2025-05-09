from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import Http404, JsonResponse, HttpResponse
from django.urls import reverse
from . import models
from django.db.models import Sum, Value
from django.db.models.functions import Coalesce
from datetime import datetime
from datetime import timedelta,date
from django.db.models.functions import ExtractYear
from . import logindecorators
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.db.models import FloatField
import time
import re
from django.db.models import Q
from openpyxl import Workbook, load_workbook
from io import BytesIO



@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def view_gudang(request):
    '''
    Fitur ini digunakan untuk menampilkan dashboard Gudang.
    Dashboard Gudang berisi 3 bagian 
    1. Menampilkan Notifikasi Barang keluar belum ACC
    2. Notifikasi barang retur belum ACC
    3. Notifikasi SPK 30 hari kebelakang

    Algortima
    1. Menampilkan Notifikasi Barang keluar belum ACC
        A. Mengambil data dari tabel TransaksiGudang dengan kriteria Jumlah lebih dari 0 dan KeteranganACCPurchasing = False.
        B. Mengubah data Tanggal sesuai format (YYYY-MM-DD) yang awalnya adalah (Tanggal,Bulan,Tahun)
    2. Notifikasi Barang retur belum ACC
        A. Mengambil data barang retur dari TransaksiGudang dengan kriteria KeteranganACC = False dan jumlah kurang dari 0 
        B. Mengubah data Tanggal sesuai format (YYYY-MM-DD) yang awalnya adalah (Tanggal,Bulan,Tahun)
    3. Notifikasi SPK 30 hari kebelakang
        B. Mengubah data Tanggal sesuai format (YYYY-MM-DD) yang awalnya adalah (Tanggal,Bulan,Tahun)
    
    '''
    getretur = (
        models.TransaksiGudang.objects.filter(KeteranganACC=False)
        .filter(jumlah__lt=0)
        .order_by("tanggal")
    )

    getkeluar = (
        models.TransaksiGudang.objects.filter(KeteranganACCPurchasing=False)
        .filter(jumlah__gt=0)
        .order_by("tanggal")
    )
    print(getkeluar)
    akhir = datetime.now()

    mulai = akhir - timedelta(days=30)
    # print(mulai)
    allspk = models.SPK.objects.filter(
        Tanggal__range=(mulai, akhir), StatusAktif=True
    ).order_by("Tanggal")
    # print(allspk)
    # print(asd)
    for i in allspk:
        if i.StatusDisplay == 0:
            detailspk = models.DetailSPK.objects.filter(NoSPK=i.id)
            i.detailspk = detailspk
        else:
            detailspk = models.DetailSPKDisplay.objects.filter(NoSPK=i.id)
            i.detailspk = detailspk

    for a in getretur:
        a.jumlah = a.jumlah * -1
    for i in getretur:
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")
    for i in getkeluar:
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")
    for i in allspk:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    if len(getretur) == 0:
        messages.info(request, "Tidak ada barang retur yang belum ACC")
    if len(getkeluar) == 0:
        messages.info(request, "Tidak ada barang keluar yang belum ACC")
    if len(allspk) == 0:
        messages.warning(request, "Tidak ada SPK selama 30 hari terakhir")
    print(allspk)
    return render(
        request,
        "gudang/viewgudang.html",
        {
            "getkeluar": getkeluar,
            "getretur": getretur,
            "allspk": allspk,
        },
    )


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def masuk_gudang(request):
    '''
    Fitur ini digunakan untuk menampilkan semua data Transaksi Pembelian Barang pada Surat Jalan Pembelian
    Secara default akan menampilkan data kedatangan barang pada rentang bulan ini
    Algoritma
    1. Mengambil data DetailSuratJalanPembelian 
    2. Mengubah data Tanggal sesuai format (YYYY-MM-DD) yang awalnya adalah (Tanggal,Bulan,Tahun)
    '''
    datasjb = models.DetailSuratJalanPembelian.objects.all().order_by(
        "NoSuratJalan__Tanggal"
    )
    if len(request.GET) == 0 :
        sekarang = datetime.now().date()
        tanggalawal = date(sekarang.year,sekarang.month,1).strftime('%Y-%m-%d')
        tanggalakhir = sekarang.strftime('%Y-%m-%d')
    else:
        tanggalawal = request.GET.get("mulai")
        tanggalakhir = request.GET.get("akhir")
    if tanggalawal == '':
        tanggalawal = datetime.min
    if tanggalakhir == '':
        tanggalakhir = datetime.max
    if tanggalawal is not None and tanggalakhir is not None:
        datasjb = models.DetailSuratJalanPembelian.objects.filter(
            NoSuratJalan__Tanggal__range=(tanggalawal, tanggalakhir)
        ).order_by("NoSuratJalan__Tanggal")

    for i in datasjb:
        i.NoSuratJalan.Tanggal = i.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")

    if len(datasjb) == 0:
        messages.info(request, "Tidak ada barang masuk ke gudang")

    return render(
        request,
        "gudang/baranggudang.html",
        {"datasjb": datasjb, "date": tanggalawal, "mulai": tanggalawal, "akhir": tanggalakhir},
    )

def load_detailpo(request):
    kodeproduk = request.GET.get("kodeproduk")
    detailpoobj = models.DetailPO.objects.filter(KodeProduk__KodeProduk=kodeproduk,KodePO__Status = False)

    print(detailpoobj)
    print('tets')
    # if id_spk.StatusDisplay == False :
    #     detailspk = models.DetailSPK.objects.filter(NoSPK=id_spk.id,)
    # else :
    #     detailspk = models.DetailSPKDisplay.objects.filter(NoSPK = id_spk.id,)

    return render(request, "gudang/opsi_po.html", {"kodepo": detailpoobj})



@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def add_gudang(request):
    '''
    Fitur ini digunakan untuk menambahkan data Surat Jalan Pembelian dan Detail Surat Jalan Pembelian
    Algoritma
    1. Menampilkan form input SJP ( Nomor Surat Jalan, Tanggal, dan Supplier untuk membuat data SJP)
    2. Menampilkan form input Detail SJP  ( Kode Stok bahan baku, Jumlah kedatangan, Po, KeteranganACC diset menjadi False (untuk di ACC Purcahsing), Harga diset 0 (Gudang tidak boleh tau Harga))
    3. Menyimpan SJP pada database
    4. Mengambil data SJP Terakhir
    5. Apabila detail PO ada, maka akan mengambil data detail PO. Apabila tidak ada maka data detail PO dikosongkan
    5. Menyimpan Detail SJP pada database
    '''
    if request.method == "GET":
        detailsjp = models.DetailSuratJalanPembelian.objects.all()
        detailsj = models.SuratJalanPembelian.objects.all()
        getproduk = models.Produk.objects.all()
        

        return render(
            request,
            "gudang/addgudang.html",
            {"detailsjp": detailsjp, "detailsj": detailsj, "getproduk": getproduk},
        )
    if request.method == "POST":
        print(request.POST)
        # print(asd)
        kode = request.POST.getlist("kodeproduk")
        nosuratjalan = request.POST["nosuratjalan"]
        tanggal = request.POST["tanggal"]
        supplier = request.POST["supplier"]
        if supplier == "":
            supplier = "-"
    #     existing_entry = models.SuratJalanPembelian.objects.filter
    # (NoSuratJalan=nosuratjalan).exists()
    #     print(existing_entry)
    #     # print(asd)
    #     if existing_entry:
    #         messages.warning(request,(f'No Surat Jalan {nosuratjalan} sudah terdaftar pada sistem'))
    #         return redirect("addgudang")
        filternosuratjalan = models.SuratJalanPembelian.objects.filter(NoSuratJalan = nosuratjalan)
        if filternosuratjalan.exists():
            messages.error(request,f'Nomor Surat Jalan sudah terdaftar dalam sistem')
            return redirect('addgudang')

        nosuratjalanobj = models.SuratJalanPembelian(
            NoSuratJalan=nosuratjalan, Tanggal=tanggal, supplier=supplier
        )
        listkodeproduk = request.POST.getlist("kodeproduk")
        error = 0
        for i in listkodeproduk:
            try:
                kodeproduk = models.Produk.objects.get(KodeProduk=i)
            except:
                error += 1
        if error == len(listkodeproduk):
            messages.error(request, "Data tidak ditemukan dalam sistem")
            return redirect("addgudang")
        nosuratjalanobj.save()
        nosuratjalanobj = models.SuratJalanPembelian.objects.get(
            NoSuratJalan=nosuratjalan
        )
        issaved = False
        for kodeproduk, jumlah, kodepo in zip(
            request.POST.getlist("kodeproduk"), request.POST.getlist("jumlah"),request.POST.getlist('detailpo')
        ):
            # print(kodeproduk)
            try:
                kodeprodukobj = models.Produk.objects.get(KodeProduk=kodeproduk)
            except:
                messages.error(
                    request, f"Data Bahan Baku {kodeproduk} tidak terdapat dalam sistem"
                )
                continue
            if kodepo == "":
                detailpoobj = None
            else:
                try:
                    detailpoobj = models.DetailPO.objects.get(id = kodepo)
                except models.DetailPO.DoesNotExist:
                    messages.error(request,f'Detail PO tidak ditemukan {kodepo}')
                    detailpoobj = None
            newprodukobj = models.DetailSuratJalanPembelian(
                KodeProduk=kodeprodukobj,
                Jumlah=jumlah,
                KeteranganACC=0,
                Harga=0,
                NoSuratJalan=nosuratjalanobj,
                PO = detailpoobj
            )
            models.transactionlog(
                user="Gudang",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"No Surat Jalan : {newprodukobj.NoSuratJalan} Kode Barang : {newprodukobj.KodeProduk}",
            ).save()
            newprodukobj.save()
            issaved = True
        messages.success(request, "Data berhasil disimpan")
        if issaved == False:
            nosuratjalanobj.delete()
        return redirect("baranggudang")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def add_gudang2(request):
    if request.method == "GET":
        detailsjb = models.SuratJalanPembelian.objects.all()
        return render(request, "gudang/addgudang2.html", {"detailsj": detailsjb})
    if request.method == "POST":
        no_surat = request.POST["no_surat"]
        tanggal = request.POST["Tanggal"]
        supplier = request.POST["supplier"]
        datasjb = models.SuratJalanPembelian(
            NoSuratJalan=no_surat,
            Tanggal=tanggal,
            supplier=supplier,
        ).save()
        return redirect("baranggudang")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def accgudang(request, id):
    '''
    Fitur ini digunakan untuk melakukan ACC Gudang pada transaksi Notifikasi Barang Keluar dan Notifikasi Barang Retur 
    Algoritma
    1. Mengambil data TransaksiGudang dengan id yang dipilih
    2. Mengubah keteranganACC menjadi True yang menandakan sudah di ACC oleh Gudang
    3. Menyimpan perubahan pada database.
    '''
    datagudang = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    datagudang.KeteranganACC = True
    datagudang.save()
    return redirect("viewgudang")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def delete_gudang(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi DetailSuratJalanPembelian atau data kedatangan barang
    Algoritma
    1. Mendapatkan data DetailSuratJalanPembelian dengan parameter IDDetailSJPembelian = id. id didapatkan dari passing value halaman awal menu bahan baku masuk
    2. Menghapus data detail surat jalan pembelian
    '''
    datasbj = models.DetailSuratJalanPembelian.objects.get(IDDetailSJPembelian=id)
    nosuratjalan = datasbj.NoSuratJalan
    models.transactionlog(
        user="Gudang",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"No Surat Jalan : {datasbj.NoSuratJalan} Kode Barang : {datasbj.KodeProduk}",
    ).save()
    filterceknosuratjalan = models.DetailSuratJalanPembelian.objects.filter(NoSuratJalan = nosuratjalan).exclude(pk = datasbj.pk).count()
    print(filterceknosuratjalan)
    if filterceknosuratjalan == 0:
        datasuratjalanpembelian = models.SuratJalanPembelian.objects.get(pk = datasbj.NoSuratJalan.pk)
        datasuratjalanpembelian.delete()
    else:
        datasbj.delete()
    messages.success(request,'Data berhasil dihapus')
    return redirect("baranggudang")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def rekap_gudang(request):
    '''
    Fitur ini digunakan untuk mendapatkan rekapitulasi hasil akhir dari setiap bahan baku  per tanggal pada gudang
    Algoritma
    1. Ambil data semua bahan baku 
    2. Mengiterasi setiap bahan baku
        3. Mengambil data pemusnahan bahan baku dengan kriteria KodeBahanBaku = input kodeproduk, Tahun = input tahun, lokasi = Gudang (untuk melihat pemusnahan bahan baku pada gudang saja)
        4. Mengambil data TransaksiGudang dengan kriteria KodeProduk = input kode stok, Tahun = input tahun, Jumlah > 0 (mengindikasikan barang keluar)
        5. Mengambil data retur TransaksiGudang dengan kriteria KodeProduk = input kode stok, Tahun = input tahun, Jumlah < 0 (mengindikasikan barang retur)
        6. Mengambil data Saldo Awal Bahan Baku dengan kriteria KodeProduk = input kode stok, Tahun = input tahun, lokasi = Gudang 
        7. Mengambil data kedatangan barang pada DetailSuratJalanPembelian dengan kritreria KodeProduk = input kode stok, tahun = input tahun
        8. Membuat listtanggal dari semua data pemusnahan bahan baku, transaksi gudang, retur gudang, dan kedatangan barang
        9. Mengiterasi setiap tanggal dari  data pemusnahan bahan baku, transaksi gudang, retur gudang, dan kedatangan barang
        10. Untuk setiap hari, rumus kode stok adalah
            Sisa = Jumlah Barang Masuk + Jumlah Barang retur - Jumlah Transaksi Gudang - Jumlah pemusnahan bahan baku
        11. Pada tanggal pertama Sisa akan ditambahkan dengan total saldo awal pada Gudang
            Sisa hari 1 = Sisa + saldo awal gudang
        12. Data sisa pada hari terakhir akan dijadikan hasil akhir bahan baku tersebut
    13. Menampilkan rekapitulasi semua sisa bahan baku

    '''
    sekarang = datetime.now().strftime('%Y-%m-%d')
    listproduk = []
    listnama = []
    satuan = []
    liststokakhir = []

    dataproduk = models.Produk.objects.all()
    datenow = datetime.now()
    tahun = datenow.year
    mulai = datetime(year=tahun, month=1, day=1)
    date = request.GET.get("date")
    try:
        datestrp = datetime.strptime(date,'%Y-%m-%d')
    except:
        datestrp = datetime.now()
    awaltahunterpilih = datetime(year=datestrp.year,month=1,day=1)
    print(date)
    print(datenow)
    

    for i in dataproduk:
        listproduk.append(i.KodeProduk)
        listnama.append(i.NamaProduk)
        satuan.append(i.unit)
        datagudang = models.TransaksiGudang.objects.filter(
                tanggal__range=(mulai, date), KodeProduk=i
            )
        print(datagudang)
        # print(asd)

        if date is not None:
            datagudang = models.TransaksiGudang.objects.filter(
                tanggal__range=(awaltahunterpilih, date), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("jumlah"), Value(0,output_field=FloatField())))
            datasjp = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__range=(awaltahunterpilih, date), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            saldoawal = models.SaldoAwalBahanBaku.objects.filter(
                Tanggal__range=(awaltahunterpilih, date), IDBahanBaku=i, IDLokasi="3"
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            pemusnahan = models.PemusnahanBahanBaku.objects.filter(
                Tanggal__range=(awaltahunterpilih, date), KodeBahanBaku=i, lokasi="3"
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
        else:
            datagudang = models.TransaksiGudang.objects.filter(
                tanggal__range=(mulai, datenow), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("jumlah"), Value(0,output_field=FloatField())))
            datasjp = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__range=(mulai, datenow), KodeProduk=i
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            saldoawal = models.SaldoAwalBahanBaku.objects.filter(
                Tanggal__range=(mulai, datenow), IDBahanBaku=i, IDLokasi="3"
            ).aggregate(kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField())))
            pemusnahan = models.PemusnahanBahanBaku.objects.filter(
    Tanggal__range=(mulai, datenow), KodeBahanBaku=i, lokasi="3"
).aggregate(
    kuantitas=Coalesce(Sum("Jumlah"), Value(0,output_field=FloatField()))
)

        stokakhir = (
            datasjp["kuantitas"]
            - datagudang["kuantitas"]
            + saldoawal["kuantitas"]
            - pemusnahan["kuantitas"]
        )
        liststokakhir.append(stokakhir)

        # print(datagudang)
        # print(datasjp)
        # print(saldoawal)
        # print(pemusnahan)
        # print(stokakhir)

    combined_list = zip(listproduk, listnama, satuan, liststokakhir,dataproduk)

    # Membuat dictionary sesuai template yang diinginkan
    produk_dict = {
        kode_produk: {
            "NamaProduk": nama_produk,
            "Satuan": satuan,
            "StokAkhir": stok_akhir,
            "produkobj" : produkobj
        }
        for kode_produk, nama_produk, satuan, stok_akhir,produkobj in combined_list
    }
    if date != "":
        sekarang = date

    return render(
        request,
        "gudang/rekapgudang.html",
        {
            "kodeproduk": listproduk,
            "date": date,
            "dict_semua": produk_dict,
            'waktu' : sekarang
        },
    )


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def detail_barang(request):
    '''
    Fitur ini digunakan untuk menampilkan KSBB Gudang 
    Algoritma
    1. User menginputkan tahun dan kode stok bahan baku yang akan dicek
    2. Mengambil data pemusnahan bahan baku dengan kriteria KodeBahanBaku = input kodeproduk, Tahun = input tahun, lokasi = Gudang (untuk melihat pemusnahan bahan baku pada gudang saja)
    3. Mengambil data TransaksiGudang dengan kriteria KodeProduk = input kode stok, Tahun = input tahun, Jumlah > 0 (mengindikasikan barang keluar)
    4. Mengambil data retur TransaksiGudang dengan kriteria KodeProduk = input kode stok, Tahun = input tahun, Jumlah < 0 (mengindikasikan barang retur)
    5. Mengambil data Saldo Awal Bahan Baku dengan kriteria KodeProduk = input kode stok, Tahun = input tahun, lokasi = Gudang 
    6. Mengambil data kedatangan barang pada DetailSuratJalanPembelian dengan kritreria KodeProduk = input kode stok, tahun = input tahun
    7. Membuat listtanggal dari semua data pemusnahan bahan baku, transaksi gudang, retur gudang, dan kedatangan barang
    8. Mengiterasi setiap tanggal dari  data pemusnahan bahan baku, transaksi gudang, retur gudang, dan kedatangan barang
    9. Untuk setiap hari, rumus kode stok adalah
        Sisa = Jumlah Barang Masuk + Jumlah Barang retur - Jumlah Transaksi Gudang - Jumlah pemusnahan bahan baku
    10. Pada tanggal pertama Sisa akan ditambahkan dengan total saldo awal pada Gudang
        Sisa hari 1 = Sisa + saldo awal gudang
    '''
    datagudang = models.TransaksiGudang.objects.all()
    dataproduk = models.Produk.objects.all()
    if len(request.GET) == 0:
        return render(
            request,
            "gudang/detailbarang.html",
            {
                "datagudang": datagudang,
                "dataproduk": dataproduk,
            },
        )

    else:
        dict_semua = []
        list_masuk = []
        list_keluar = []
        list_sisa = []
        dictdata = []
        input_kode = request.GET.get("input_kode")
        input_tahun = request.GET.get("input_tahun")
        pemusnahan = models.PemusnahanBahanBaku.objects.filter(
            KodeBahanBaku__KodeProduk=input_kode,
            Tanggal__year=input_tahun,
            lokasi__NamaLokasi="Gudang",
        ).order_by("Tanggal")
        datagudang2 = models.TransaksiGudang.objects.filter(
            KodeProduk__KodeProduk=input_kode, tanggal__year=input_tahun, jumlah__gte=0,KeteranganACC = True
        ).order_by("tanggal")
        dataretur = models.TransaksiGudang.objects.filter(
            KodeProduk__KodeProduk=input_kode, tanggal__year=input_tahun, jumlah__lt=0
        ).order_by("tanggal")

        saldo_awal = (
            models.SaldoAwalBahanBaku.objects.filter(IDBahanBaku__KodeProduk=input_kode)
            .filter(Tanggal__year=input_tahun, IDLokasi__NamaLokasi="Gudang")
            .first()
        )

        if saldo_awal:
            datasaldoawal = saldo_awal.Jumlah
            sisa = saldo_awal.Jumlah
            tanggalsaldoawal = saldo_awal.Tanggal.year
        else:
            datasaldoawal = 0
            sisa = 0
            tanggalsaldoawal = datetime.now().year
        datasjp = (
            models.DetailSuratJalanPembelian.objects.filter(KodeProduk__KodeProduk=input_kode)
            .filter(NoSuratJalan__Tanggal__year=input_tahun)
            .order_by("NoSuratJalan__Tanggal")
        )
        tanggalgudang = list(datagudang2.values_list("tanggal", flat=True).distinct())
        tanggalgudang2 = list(
            datasjp.values_list("NoSuratJalan__Tanggal", flat=True).distinct()
        )
        tanggalgudang3 = list(dataretur.values_list("tanggal", flat=True).distinct())

        tanggalgudang4 = list(pemusnahan.values_list("Tanggal", flat=True).distinct())
        tanggaltotal = tanggalgudang + tanggalgudang2 + tanggalgudang3 + tanggalgudang4
        tanggaltotal = sorted(list(set(tanggaltotal)))

        if len(tanggaltotal) == 0:
            messages.error(
                request, "Tidak ada barang masuk ke gudang, keluar, dan retur"
            )
        for i in tanggaltotal:
            # Masuk
            # sjp
            # retur
            # Saldo awal
            sjpobj = datasjp.filter(NoSuratJalan__Tanggal=i).aggregate(
                total=Sum("Jumlah")
            )
            totalsjp = sjpobj["total"]
            if totalsjp is None:
                totalsjp = 0

            returobj = dataretur.filter(tanggal=i).aggregate(total=Sum("jumlah"))
            totalretur = returobj["total"]
            if totalretur is None:
                totalretur = 0
            else:
                totalretur = abs(returobj["total"])
            totalmasuk = totalsjp + totalretur

            datakeluar = datagudang2.filter(tanggal=i).aggregate(total=Sum("jumlah"))
            jumlahkeluar = datakeluar["total"]
            if jumlahkeluar is None:
                jumlahkeluar = 0
            pemusnahanobj = pemusnahan.filter(Tanggal=i).aggregate(total=Sum("Jumlah"))
            jumlahpemusnahan = pemusnahanobj["total"]
            if jumlahpemusnahan is None:
                jumlahpemusnahan = 0
            totalkeluar = jumlahkeluar + jumlahpemusnahan
            sisa += totalmasuk - totalkeluar
            if sisa < 0:
                messages.warning(request, f"Sisa stok menjadi negatif pada tanggal {i}")
            print(totalkeluar, totalmasuk)

            i = i.strftime("%Y-%m-%d")
            dummy = {"masuk": totalmasuk, "keluar": totalkeluar, "saldo": sisa,'Tanggal':i}
            dictdata.append(dummy)
        print(dictdata)
        produkobj = models.Produk.objects.get(KodeProduk = input_kode)
        return render(
            request,
            "gudang/detailbarang.html",
            {
                "datagudang2": datagudang2,
                "dataproduk": dataproduk,
                "list_keluar": list_keluar,
                "dict_semua": dict_semua,
                "kodeproduk": input_kode,
                "saldoawal": saldo_awal,
                "input_tahun": input_tahun,
                "datasjp": datasjp,
                "dictdata": dictdata,
                "datasaldoawal": datasaldoawal,
                'lokasi' : 'Gudang',
                'produkobj' : produkobj,
                'tanggalsaldoawal':str(input_tahun)
            },
        )

def detailksbb(request, id, tanggal,lokasi):
    '''
    Fitur ini digunakan untuk menampilkan detail transaksi dari KSBB
    pada fitur ini akan ada 4 bagian detail yang mencakup
    1. Barang Masuk : Diambil dari data DetailSuratJalanPembelian
    2. Barang keluar Gudang : Diambil dari data TransaksiGudang dengan kriteria Jumlah > 0
    3. Barang retur ke Gudang : Diambil dari data TransaksiGudang dengan kriteria Jumlah < 0
    4. Pemusnahan barang : Diambil dari data pemusnahan bahan baku gudang 
    '''
    tanggal = datetime.strptime(tanggal, "%Y-%m-%d")
    tanggal = tanggal.strftime("%Y-%m-%d")
    # SJP
    datamasuk = models.DetailSuratJalanPembelian.objects.filter(KodeProduk__KodeProduk = id, NoSuratJalan__Tanggal = tanggal)
    # Transaksi Gudang
    datagudang = models.TransaksiGudang.objects.filter(tanggal=tanggal, KodeProduk__KodeProduk=id,jumlah__gte=0)
    dataretur = models.TransaksiGudang.objects.filter(tanggal=tanggal, KodeProduk__KodeProduk=id,jumlah__lt=0)
    for item in dataretur:
        item.jumlah = item.jumlah *-1
    print(datagudang,dataretur)
    # Transaksi Pemusnahan Bahan Baku
    datapemusnahanbahanbaku  =models.PemusnahanBahanBaku.objects.filter(Tanggal = tanggal,KodeBahanBaku__KodeProduk = id,lokasi__NamaLokasi=lokasi)
    return render(
        request,
        "gudang/view_detailksbb.html",
        {
            "datagudang": datagudang,
            'datapemusnahanbahanbaku' : datapemusnahanbahanbaku,
            "dataretur" : dataretur,
            'datamasuk':datamasuk
        },
    )


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def barang_keluar(request):
    '''
    Fitur ini digunakan untuk menampilkan semua Transaksi Barang keluar pada TransaksiGudang
    Algoritma
    1. mengambil data TransaksiGudang dengan kriteria jumlah > 0 (mengindikasikan barang keluar)
    2. Mengubah tanggal dengan format (YYYY-MM-DD)
    '''
    datalokasi = models.Lokasi.objects.filter(NamaLokasi__in=("WIP", "FG","Lain-Lain"))
    data = models.TransaksiGudang.objects.filter(jumlah__gt=0).order_by("tanggal")
    
    
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
        data = data.filter(tanggal__range=(tanggalawal,tanggalakhir))
        for i in data:
            i.tanggal = i.tanggal.strftime("%Y-%m-%d")

        return render(
            request,
            "gudang/barangkeluar.html",
            {
                "datalokasi": datalokasi,
                "data": data,
                "date": tanggalawal,
                "date2": tanggalakhir,
            },
        )
    else:
        tanggalawal = request.GET.get("mulai")
        tanggalakhir = request.GET.get("akhir")
        lok = request.GET.get("lokasi")
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == "":
            tanggalakhir = datetime.max

        data = data.filter(
            tanggal__range=(tanggalawal, tanggalakhir), Lokasi__NamaLokasi=lok, jumlah__gt=0
        ).order_by("tanggal")

        for i in data:
            i.tanggal = i.tanggal.strftime("%Y-%m-%d")

        if len(data) == 0:
            messages.error(request, "Tidak ada barang keluar dari gudang")

        return render(
            request,
            "gudang/barangkeluar.html",
            {
                "datalokasi": datalokasi,
                "data": data,
                "date": tanggalawal,
                "date2": tanggalakhir,
                "lok": lok,
            },
        )


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def barang_retur(request):
    '''
    Fitur ini digunakan untuk menampilkan semua data barang retur ke gudang
    Algoritma
    1. Mengambil semua data transaksigudang dengan kriteria jumlah < 0 (mengindikasikan retur)
    2. Mengubah tanggal seusai format (YYYY-MM-DD)
    '''
    datalokasi = models.Lokasi.objects.all()
    data = models.TransaksiGudang.objects.filter(jumlah__lt=0).order_by("tanggal")
    for i in data:
        i.jumlah = i.jumlah * -1
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")
    if len(request.GET) == 0:
        return render(
            request,
            "gudang/barangretur.html",
            {
                "datalokasi": datalokasi,
                "data": data,
            },
        )
    else:
        date = request.GET.get("mulai")
        date2 = request.GET.get("akhir")
        lok = request.GET.get("lokasi")
        if date == '':
            date = datetime.min
        if date2 == "":
            date2 = datetime.max

        data = data.filter(
            tanggal__range=(date, date2), Lokasi__NamaLokasi=lok, jumlah__lt=0
        ).order_by("tanggal")


        for i in data:
            i.jumlah = i.jumlah * -1
            i.tanggal = i.tanggal.strftime("%Y-%m-%d")

        if len(data) == 0:
            messages.error(request, "Tidak ada barang keluar dari gudang")

        print(data)

        return render(
            request,
            "gudang/barangretur.html",
            {
                "datalokasi": datalokasi,
                "data": data,
                "date": date,
                "date2": date2,
                "lok": lok,
            },
        )


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def accgudang2(request, id):
    '''
    Fitur ini digunakan untuk melakukan ACC pada Transaksi Gudang barang retur melalui fitur halaman Barang Retur
    '''
    datagudang = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    datagudang.KeteranganACC = True
    datagudang.save()

    return redirect("barangretur")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def accgudang3(request, id):
    '''
        Fitur ini digunakan untuk melakukan ACC pada Transaksi Gudang barang keluar melalui fitur halaman Bahan Baku Keluar

    '''
    datagudang = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    datagudang.KeteranganACC = True
    datagudang.save()
    data = models.transactionlog(
        user="Gudang",
        waktu=datetime.now(),
        jenis="Update",
        pesan=f"Status Transaksi Data Gudang {datagudang.tanggal} - {datagudang.KodeProduk} - {datagudang.KodeProduk.NamaProduk}  ",
    )

    return redirect("barangkeluar")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def spk(request):
    '''
    Fitur ini digunakan untuk menampilkan data SPK 
    Algoritma
    1. Mendapatkan data spk dari tabel SPK
    2. Mengiterasi setiap item SPK
    3. Memfilter detail SPK tiap kode SPK untuk mendapatkan Artikel yang diproduksi pada SPK Terkait
    4. Menyesuaikan tanggal menjadi (YYYY-MM-DD)
    '''
    dataspk = models.SPK.objects.all().order_by("-Tanggal")
    for i in dataspk:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")
        if i.StatusDisplay == True:
            i.detailspk = models.DetailSPKDisplay.objects.filter(NoSPK = i)
        else:
            i.detailspk = models.DetailSPK.objects.filter(NoSPK = i)
    return render(request, "gudang/spkgudang.html", {"dataspk": dataspk})


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def tracking_spk(request, id):
    '''
    Fitur ini digunakan untuk melakukan tracking pada SPK
    pada bagian ini ada beberapa section antara lain
    1. Informasi SPK
    2. Rekap Kumulasi permintaan barang
    3. Tracking detail permitnaan barang ke SPK

    Algoritma
    1. Informasi SPK
        A. Mengambil data detail SPK dan SPK terkait dengan kode SPK id = id, dimana id didapatkan dari passing values pada halaman awal SPK
    2. Rekap Kumulasi permintaan barang
        A. Menghitung agregasi jumlah permintaan barang tiap bahan baku melalui tabel Transaksi Gudang dengan kriteria NoSPK = id dan jumlah > 0
    3. Tracking detail permitnaan barang ke SPK
        A. Mengambil data transaksi permintaan barang tiap bahan baku melalui tabel Transaksi Gudang dengan kriteria NoSPK = id dan jumlah > 0
    '''
    dataartikel = models.Artikel.objects.all()
    dataspk = models.SPK.objects.get(id=id)
    datadetail = models.DetailSPK.objects.filter(NoSPK=dataspk.id)

    # Data SPK terkait yang telah di request ke Gudang
    transaksigudangobj = models.TransaksiGudang.objects.filter(
        DetailSPK__NoSPK=dataspk.id, jumlah__gte=0
    )

    # Data SPK Terkait yang telah jadi di FG
    transaksiproduksiobj = models.TransaksiProduksi.objects.filter(
        DetailSPK__NoSPK=dataspk.id, Jenis="Mutasi"
    )

    # Data SPK Terkait yang telah dikirim
    sppbobj = models.DetailSPPB.objects.filter(DetailSPK__NoSPK=dataspk.id)

    if dataspk.StatusDisplay == True:

        # Data SPK terkait yang telah di request ke Gudang
        transaksigudangobj = models.TransaksiGudang.objects.filter(
            DetailSPKDisplay__NoSPK=dataspk.id, jumlah__gte=0
        )

        # Data SPK Terkait yang telah jadi di FG
        transaksiproduksiobj = models.TransaksiProduksi.objects.filter(
            DetailSPKDisplay__NoSPK=dataspk.id, Jenis="Mutasi"
        )
        print(transaksiproduksiobj)

        # Data SPK Terkait yang telah dikirim
        sppbobj = models.DetailSPPB.objects.filter(DetailSPKDisplay__NoSPK=dataspk.id)
        rekapjumlahpermintaanperbahanbaku = transaksigudangobj.values(
            "KodeProduk__KodeProduk", "KodeProduk__NamaProduk", "KodeProduk__unit"
        ).annotate(total=Sum("jumlah"))
    else:
        # Data SPK terkait yang telah di request ke Gudang
        transaksigudangobj = models.TransaksiGudang.objects.filter(
            DetailSPK__NoSPK=dataspk.id, jumlah__gte=0
        )

        # Data SPK Terkait yang telah jadi di FG
        transaksiproduksiobj = models.TransaksiProduksi.objects.filter(
            DetailSPK__NoSPK=dataspk.id, Jenis="Mutasi"
        )

        # Data SPK Terkait yang telah dikirim
        sppbobj = models.DetailSPPB.objects.filter(DetailSPK__NoSPK=dataspk.id)
        rekapjumlahpermintaanperbahanbaku = transaksigudangobj.values(
            "KodeProduk__KodeProduk", "KodeProduk__NamaProduk", "KodeProduk__unit"
        ).annotate(total=Sum("jumlah"))

    if request.method == "GET":
        tanggal = datetime.strftime(dataspk.Tanggal, "%Y-%m-%d")

        return render(
            request,
            "gudang/trackingspk.html",
            {
                "data": dataartikel,
                "dataspk": dataspk,
                "datadetail": datadetail,
                "tanggal": tanggal,
                "transaksigudang": transaksigudangobj,
                "transaksiproduksi": transaksiproduksiobj,
                "transaksikeluar": sppbobj,
                "datarekappermintaanbahanbaku": rekapjumlahpermintaanperbahanbaku,
            },
        )

'''Tidak dipakai'''
# @login_required
# @logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
# def cobaform(request):
#     databahanbaku = models.Produk.objects.all()
#     if request.method == "POST":
#         print(request.POST)
#         nomor_nota = request.POST.get("nomor_nota")
#         produk_list = request.POST.getlist("produk[]")
#         print(len(produk_list))
#         print(produk_list)

#     return render(request, "gudang/cobaform.html", {"data": databahanbaku})


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def addgudang3(request):
    '''
    Fitur ini digunakan untuk menambahkan data transaksi lain-lain pada gudang
    Transaksi Lain-Lain = Semua transaksi gudang yang ditujukan untuk keperluan selain non produksi (tidak masuk data KSBB Produksi WIP/FG)
    Algoritma
    1. Menampilkan form Transaksi Gudang 
    2. Mengambil data input berupa Kode Stok Bahan Baku, Tanggal, Keterangan, Jumlah, Lokasi =  Lain-Lain, Keterangan ACC = True
    3. Menyimpan data Transaksi Gudang 
    '''
    if request.method == "GET":
        detailspk = models.DetailSPK.objects.all()
        getproduk = models.Produk.objects.all()
        getlokasi = models.Lokasi.objects.all()
        return render(
            request,
            "gudang/addgudang3.html",
            {"detailspk": detailspk, "getproduk": getproduk, "getlokasi": getlokasi},
        )
    if request.method == "POST":
        kode = request.POST["kodeproduk"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]
        jumlah = request.POST["jumlah"]
        acc = True
        lokasi = request.POST["lokasi"]
        print(lokasi)
        try:
            kodeprodukobj = models.Produk.objects.get(KodeProduk=kode)
        except:
            messages.error(request, f"Data bahan baku {kode} tidak ditemukan")
            return redirect("addgudang3")
        savetrans = models.TransaksiGudang(
            KodeProduk=kodeprodukobj,
            keterangan=keterangan,
            jumlah=jumlah,
            tanggal=tanggal,
            KeteranganACC=acc,
            Lokasi=models.Lokasi.objects.get(NamaLokasi=lokasi),
            DetailSPK=None,
            KeteranganACCPurchasing = False

        )
        models.transactionlog(
            user="Gudang",
            waktu=datetime.now(),
            jenis="Create",
            pesan=f"Kode Barang : {savetrans.KodeProduk} Jumlah : {savetrans.jumlah} Lokasi : {lokasi}",
        ).save()
        savetrans.save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("barangkeluar")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def updatetransaksilainlain(request, id):
    '''
    Fitur ini digunakan untuk melakukan update transaksi lain lain
    Algoritma
    1. Ambil data transaksi lain lain dengan kriteria IDDetailTransaksiGudang = id (id didapatkan dari passing values HTML)
    2. Menampilkan form input update data transaksi 
    3. Program menerima input berupa Kode Produk, Lokasi, Tanggal, Jumlah, dan Keterangan
    4. Mengupdate data transaksi lain lain
    '''
    gudangobj = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    allproduk = models.Produk.objects.all()
    if request.method == "GET":
        tanggal = datetime.strftime(gudangobj.tanggal, "%Y-%m-%d")
        return render(
            request,
            "gudang/update_transaksilainlain.html",
            {
                "gudang": gudangobj,
                "tanggal": tanggal,
                "getproduk": allproduk,
            },
        )
    elif request.method == "POST":
        kode_produk = request.POST["kodeproduk"]
        try:
            getproduk = models.Produk.objects.get(KodeProduk=kode_produk)
        except:
            messages.error(request, f"Data bahan baku {kode_produk} tidak ditemukan")
            return redirect("updatetransaksilainlain", id=id)
        lokasi = request.POST["lokasi"]
        getlokasi = models.Lokasi.objects.get(NamaLokasi=lokasi)
        tanggal = request.POST["tanggal"]
        jumlah = request.POST["jumlah"]
        keterangan = request.POST["keterangan"]

        gudangobj.KodeProduk = getproduk
        gudangobj.Lokasi = getlokasi
        gudangobj.tanggal = tanggal
        gudangobj.jumlah = jumlah
        gudangobj.keterangan = keterangan
        gudangobj.KeteranganACCPurchasing = False

        gudangobj.save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("barangkeluar")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def deletetransaksilainlain(request, id):
    '''
    fitur ini digunakan untuk menghapus data transaksi lain lain
    Algoritma:
    1. Mengambil data transaksi lain lain dengan kriteria IDDetailTransaksiGudang = id (id didapatkan dari passing values HTML)
    2. Menghapus data transaksi lain lain 
    '''
    datagudang = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    datagudang.delete()
    messages.success(request, "Data Berhasil dihapus")

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Transaksi Barang Masuk. Kode Produk : {datagudang.KodeProduk} Jumlah : {datagudang.jumlah} Keterangan : {datagudang.keterangan}",
    ).save()

    return redirect("barangkeluar")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def read_produk(request):
    '''Fitur ini digunakan untuk menampilkan semua bahan baku yang memiliki kode stok pada database'''
    produkobj = models.Produk.objects.all()
    return render(request, "gudang/read_produk.html", {"produkobj": produkobj})


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def update_produk_gudang(request, id):
    '''
    Fitur ini digunakan untuk melakukan update bahan baku.
    Data yang dapat diubah adalah KeteranganGudang, dan jumlah minimal saja sedangkan sisanya dari Purchasing
    Algoritma
    1. Mengambil data produk gudang dengan kriteria KodeProduk = id (id didapatkan dari passing values HTML)
    2. Menampilkan form update bahan baku 
    3. Program mendapat input berupa data keteragan gudang dan Jumlah Minimal
    4. Mengupdate data produk
    '''
    produkobj = models.Produk.objects.get(KodeProduk=id)
    if request.method == "GET":
        return render(request, "gudang/update_produk.html", {"produkobj": produkobj})
    else:
        print(request.POST)
        keterangan_produk = request.POST["keterangan_produk"]
        jumlah_minimal = request.POST["jumlah_minimal"]
        produkobj.keteranganGudang = keterangan_produk
        produkobj.Jumlahminimal = jumlah_minimal
        produkobj.save()
        models.transactionlog(
            user="Gudang",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Jumlah Minimal : {produkobj.Jumlahminimal} Keterangan : {produkobj.keteranganGudang}",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("readprodukgudang")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def read_saldoawal(request):
    '''
    Fitur ini digunakan untuk melihat data saldo awal bahan baku pada gudang
    Algorimta
    1. Mengambil semua data saldo awal bahan baku pada gudang 
    2. Mengubah format tanggal menjadi (YYYY-MM-DD)
    '''
    dataproduk = models.SaldoAwalBahanBaku.objects.filter(
        IDLokasi__NamaLokasi="Gudang"
    ).order_by("-Tanggal")
    for i in dataproduk:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "gudang/read_saldoawalbahan.html", {"dataproduk": dataproduk}
    )


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def addsaldo(request):
    '''
    Fitur ini digunakan untuk menambahkan data Saldo awal bahan baku pada gudang.
    Hanya ada 1 data saldo awal pada bahan baku pada setiap tahunnya
    Algoritma
    1. Menampilkan form input data saldo awal bahan baku. (Kode Stok, Nama Lokasi disetting Gudang, Jumlah, dan Tanggal (hanya diambil tahun)
    2. Melakukan cek apakah data saldo awal untuk bahan baku terpilih dan tahun terpilih sudah ada ?. Apabila ada maka menampilkan tulisan error apabila tidak maka lanjut ke tahapn selanjutnya
    2. Menyimpan data saldoawal pada database
    '''
    databarang = models.Produk.objects.all()
    datalokasi = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "gudang/addsaldobahan.html",
            {"nama_lokasi": datalokasi, "databarang": databarang},
        )
    else:
        produk_list = request.POST.getlist("produk[]")
        lokasi_list = request.POST.getlist("nama_lokasi[]")
        jumlah_list = request.POST.getlist("jumlah[]")
        # harga = 0
        tanggal = request.POST["tanggal"]
        print(request.POST)
        # print(asd)

        # Ubah format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")
        # Periksa apakah entri sudah ada
        for kodeproduk, lokasi, jumlah in zip(produk_list, lokasi_list, jumlah_list):
            try:
                produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)
            except models.Produk.DoesNotExist:
                messages.error(request, f"Data bahan baku {kodeproduk} tidak ditemukan.")
                return redirect("addsaldobahan")

            try:
                lokasiobj = models.Lokasi.objects.get(NamaLokasi=lokasi)
            except models.Lokasi.DoesNotExist:
                messages.error(request, f"Lokasi {lokasi} tidak ditemukan.")
                return redirect("addsaldobahan")

            existing_entry = models.SaldoAwalBahanBaku.objects.filter(
                Tanggal__year=datetime.strptime(tanggal, "%Y-%m-%d").year,
                IDBahanBaku__KodeProduk=kodeproduk,
                IDLokasi__NamaLokasi=lokasi,
            ).exists()

            if existing_entry:
                messages.warning(
                    request, f"Sudah ada entry pada tahun {datetime.strptime(tanggal, '%Y-%m-%d').year} untuk produk {kodeproduk} di lokasi {lokasi}."
                )
                continue

            saldoawalobj = models.SaldoAwalBahanBaku(
                Tanggal=tanggal,
                Jumlah=jumlah,
                IDBahanBaku=produkobj,
                IDLokasi=lokasiobj,
                Harga=0,  # Sesuaikan jika ada input harga
            )

            saldoawalobj.save()

            models.transactionlog(
                user="Gudang",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Kode Barang: {kodeproduk}, Lokasi: {lokasi}.",
            ).save()

            messages.success(request, f"Data {produkobj.KodeProduk} berhasil ditambahkan.")
    return redirect("read_saldoawalbahan")



@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def delete_saldo(request, id):
    '''
    Fitur ini digunakan untuk menghapus data saldo awal bahan baku gudang
    Algoritma
    1. Mengambil data saldo awal bahan baku gudang dengan parameter IDSaldoAwalBahanBaku = id
    2. Menghapus data saldo awal bahan baku gudang
    '''
    dataobj = models.SaldoAwalBahanBaku.objects.get(IDSaldoAwalBahanBaku=id)
    models.transactionlog(
        user="Gudang",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Kode Barang : {dataobj.IDBahanBaku} Lokasi : {dataobj.IDLokasi}",
    ).save()
    dataobj.delete()
    messages.success(request,"Data berhasil dihapus")
    return redirect("read_saldoawalbahan")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def update_saldo(request, id):
    '''
    Fitur ini digunakan untuk melakukan update data saldo awal bahan baku gudang
    Algoritma
    1. Mengambil data saldo awal bahan baku dengan kriteria IDSaldoAwalBahanBaku = id. id didapatkan dari passing values pada halaman awal saldo awal bahan baku
    2. mengambil data Kode Produk, Lokasi (disetting gudang), Jumlah , dan tanggal
    3. Cek apakah sudah ada data kodeproduk terkait pada tahun yang diinput dengan mengecualikan id saldo awal bahan baku terpilh. Apabila ada maka akan menamplkan pesan error, apabila tidak maka lanjut
    4. Menyimpan perubahan pada database
    '''
    databarang = models.Produk.objects.all()
    dataobj = models.SaldoAwalBahanBaku.objects.get(IDSaldoAwalBahanBaku=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "gudang/update_saldobahan.html",
            {"data": dataobj, "nama_lokasi": lokasiobj, "databarang": databarang},
        )

    else:
        print(request.POST)
        kodeproduk = request.POST["produk"]
        lokasi = request.POST["nama_lokasi"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        try:
            produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except:
            messages.warning(
                request, f"Tidak ditemukan bahan bau {kodeproduk} dalam sistem"
            )
            return redirect("updatesaldobahan", id=id)
        lokasiobj = models.Lokasi.objects.get(NamaLokasi=lokasi)
        lokasi = str(lokasiobj.IDLokasi)
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")

        existing_entry = (
            models.SaldoAwalBahanBaku.objects.filter(
                Tanggal__year=tanggal_formatted.year,
                IDBahanBaku__KodeProduk=kodeproduk,
                IDLokasi=lokasi,
            )
            .exclude(IDSaldoAwalBahanBaku=id)
            .exists()
        )
        if existing_entry:
            # Jika sudah ada, beri tanggapan atau lakukan tindakan yang sesuai
            messages.warning(
                request, ("Sudah ada Entry pada tahun", tanggal_formatted.year)
            )
            return redirect("updatesaldobahan", id=id)

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.IDBahanBaku = produkobj
        dataobj.IDLokasi_id = lokasi
        models.transactionlog(
            user="Gudang",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Kode Barang Lama : {dataobj.IDBahanBaku} Jumlah Lama : {dataobj.Jumlah} Harga Lama : {dataobj.Harga} Kode Barang Baru : {kodeproduk} Jumlah Baru : {jumlah}",
        ).save()
        dataobj.save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("read_saldoawalbahan")


"""REVISI 5/18/2024"""


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def update_gudang(request, id):
    '''
    Fitur ini digunakan untuk melakukan update pada Transaksi DetailSuratJalanPembelian yang sudah ada
    Algoritma
    1. Mendapatkan data DetailSuratJalanPembelian dengan kode IDDetailSJPembelian = id. id didapatkan dari passing value url pada halaman awal menu bahan baku masuk
    2. Mengirimkan parameter berupa Kode Produk, Tanggal, Jumlah, DetailPO (berisi id detailPO)
    3. Menyimpan data update
    '''
    datasjp = models.DetailSuratJalanPembelian.objects.get(IDDetailSJPembelian=id)
    availablepo= models.DetailPO.objects.filter(KodeProduk=datasjp.KodeProduk,KodePO__Status = 0 )
    datasjp2 = models.DetailSuratJalanPembelian.objects.all()
    datasj = models.SuratJalanPembelian.objects.all()
    getproduk = models.Produk.objects.all()
    datasjp_getobj = models.SuratJalanPembelian.objects.get(
        NoSuratJalan=datasjp.NoSuratJalan.NoSuratJalan
    )
   
    if request.method == "GET":

        return render(
            request,
            "gudang/updategudang2.html",
            {
                "datasjp": datasjp_getobj,
                "detailsjp": datasjp,
                "datasj": datasj,
                "detailsj": datasjp2,
                "tanggal": datetime.strftime(
                    datasjp_getobj.Tanggal,
                    "%Y-%m-%d",
                ),
                "getproduk": getproduk,
                'availablepo':availablepo,
            },
        )

    else:
        tanggal = request.POST["tanggal"]
        supplier = request.POST['supplier']
        nosuratjalan = request.POST['nosuratjalan']
        print(request.POST)
        # print(asd)
        kode_produk = request.POST.get("kodeproduk")
        try:
            kode_produkobj = models.Produk.objects.get(KodeProduk=kode_produk)
        except:
            messages.error(request, f"Data bahan baku {kode_produk} tidak ditemukan")
            return redirect("updategudang", id=id)
        jumlah = request.POST["jumlah"]
        detailpo = request.POST['detailpo']
        if detailpo != '':
            try:
                detailpoobj = models.DetailPO.objects.get(pk = detailpo)
            except models.DetailPO.DoesNotExist:
                messages.error(request,f'Detail PO tidak ditemukan {detailpo}')
                detailpoobj = None
        else:
            detailpoobj = None

        datasjp.KodeProduk = kode_produkobj
        datasjp.Jumlah = jumlah
        datasjp.KeteranganACC = datasjp.KeteranganACC
        datasjp.Harga = datasjp.Harga
        datasjp.NoSuratJalan.NoSuratJalan = nosuratjalan
        datasjp.NoSuratJalan.Tanggal = tanggal
        datasjp.NoSuratJalan.supplier = supplier
        datasjp.KeteranganACC = False
        datasjp.PO = detailpoobj
        datasjp.NoSuratJalan.save()
        datasjp.save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("baranggudang")


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","purchasing","ppic"])
def load_produk(request):
    '''
    AJAX Support : untuk meloading bahan baku 
    '''
    print(request.GET)
    artikel = request.GET.get("artikel")
    produkobj = models.Produk.objects.get(KodeProduk=artikel)
    data = {
        "KodeProduk": artikel,
        "NamaProduk": produkobj.NamaProduk,
        "unit": produkobj.unit,
    }
    return JsonResponse(data, safe=False)


def bulk_createsjp(request):
    '''PAKAI FILE KSBB PRC A'''
    if request.method == "POST" and request.FILES["file"]:
        kodebahanerror = []
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        # sheet_names = ['A-004-154','A-004-155','A-005-158','B-012-10']
        # sheet_names = ['A-004-117']

        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=3)
            print(item)
            print(df)
            # print(asd)

            # i = 0
            for index, row in df.iterrows():
                # datadelete = models.DetailSuratJalanPembelian.objects.filter(
                #             KodeProduk=models.Produk.objects.get(KodeProduk=item),
                #             )
                # print(datadelete)
                # for data in datadelete:
                #     print(data.Jumlah, data.NoSuratJalan.Tanggal)
                #     data.delete()
                # print(asd)
                # if i < 2:
                #     i += 1
                #     continue
                # print(row["Tanggal"])
                # print(row)
                if pd.isna(row["Tanggal"]):
                    print(f"Index {index}: Tanggal adalah NaT")
                else:
                    try:
                        # print(index, row["Tanggal"])
                        # print(row["Masuk "])
                        if pd.isna(row["Masuk "]):
                            continue
                        data = models.SuratJalanPembelian(
                            NoSuratJalan=f"SJP/{row['Tanggal']}",
                            Tanggal=row["Tanggal"],
                            supplier="-",
                            
                        ).save()
                        detailsjp = models.DetailSuratJalanPembelian(
                            Jumlah=row["Masuk "],
                            KeteranganACC=1,
                            Harga=row["Unnamed: 3"],
                            KodeProduk=models.Produk.objects.get(KodeProduk=item),
                            NoSuratJalan=models.SuratJalanPembelian.objects.get(
                                Tanggal=row["Tanggal"]
                            ),
                        ).save()
                    except Exception as e:
                        kodebahanerror.append([item,e])

                        continue

        return render(request,'error/errorsjp.html',{'data':kodebahanerror})

    return render(request, "Purchasing/bulk_createproduk.html")


def bulk_createsaldoawal(request):
    '''PAKAI FILE KSBB PRC'''
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        sheet_names = ['A-004-154','A-004-155','A-005-158','B-012-10']


        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=4)
            print(item)
            print(df)
            # print(asd)


            for index, row in df.iterrows():
                    print("Saldo Akhir")
                    print(row)
                    if pd.isna(row["Harga.2"]):
                        print(f"Data Kosong, Lanjut")
                        break
                    else:
                        saldoawalwip = models.SaldoAwalBahanBaku.objects.update_or_create(
                            IDBahanBaku=models.Produk.objects.get(KodeProduk=item),
                            IDLokasi=models.Lokasi.objects.get(pk=3),
                            defaults={
                            'Harga':row["Harga.2"],
                            'Jumlah':row["Quantity.2"],
                            'Tanggal':"2024-01-01",
                            }
                        )
                        break

        return HttpResponse("Berhasil Upload")

    return render(request, "Purchasing/bulk_createproduk.html")


# def bulk_createtransaksigudang(request):
#     '''
#     UNTUK MENAMBAHKAN DATA TRANSAKSI GUDANG MELALUI KSBJ TIAP ARTIKEL 
#     '''
#     if request.method == "POST" and request.FILES["file"]:
#         file = request.FILES["file"]
#         nama_artikel = '9010 AC'
#         artikelobj = models.Artikel.objects.get(KodeArtikel = nama_artikel)
#         print(nama_artikel)
#         # print(asd)
#         excel_file = pd.ExcelFile(file)
        

#         # Mendapatkan daftar nama sheet
#         sheet_names = excel_file.sheet_names
#         produkerror = []

#         sheetname = 'WIP'
#         df = pd.read_excel(file, engine="openpyxl", sheet_name=sheetname, header=5)
#         print(sheetname)
#         print(df)
#         # print(asd)

#         i = 0
#         for index, row in df.iterrows():
#             try:
#                 print(row)
#                 print(asd)
#                 if pd.isna(row["Unnamed: 4"]):
#                     print(f"Index {index}: Tanggal adalah NaT")
#                 else:

#                     transaksiobj = models.TransaksiGudang(
#                             keterangan="-",
#                             jumlah=row['Unnamed: 4'],
#                             tanggal=row['Unnamed: 0'],
#                             KeteranganACC=True,
#                             KodeProduk=models.Produk.objects.get(KodeProduk=item),
#                             Lokasi=models.Lokasi.objects.get(IDLokasi=1),
#                         )
#             except Exception as e:
#                     produkerror.append([item,e])
#                     continue


#         return HttpResponse(f"Berhasil Upload, {produkerror}")

#     return render(request, "Purchasing/bulk_createproduk.html")

def bulk_createtransaksigudang(request):
    '''
    UNTUK MENAMBAHKAN DATA TRANSAKSI GUDANG MELALUI KSBB WIP DAN FG
    KS Bahan Baku A.xlsx (WIP)
    '''
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        # print(asd)
        excel_file = pd.ExcelFile(file)
        

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        # sheet_names = ['A-006-41']
        produkerror = []

        for item in sheet_names:
            # datadelete = models.TransaksiGudang.objects.filter( KodeProduk=models.Produk.objects.get(KodeProduk=item),
            #                     Lokasi=models.Lokasi.objects.get(IDLokasi=1))
            # print(datadelete)
            # for item in datadelete:
            #     print(item.tanggal,item.jumlah)
            #     item.delete()
            # print(asd)
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=6)
            print(item)
            print(df)
            # print(asd)

            i = 0
            tanggal = None
            listtanggal = []
            for index, row in df.iterrows():
            
                print(row)
                print(item)
                # print(asd)
                if not pd.isna(row["Tanggal"]):
                    listtanggal.append(row['Tanggal'])

                if pd.isna(row["Masuk"]):
                    print(f"Index {index}: Tanggal adalah NaT")
                else:
                    if pd.isna(row['Tanggal']):
                        try:
                            tanggal = listtanggal[-1]
                        except:
                            continue
                    else:
                        tanggal = row['Tanggal']
                    # try:
                    try:
                        tanggal = datetime.strftime(tanggal,"%Y-%m-%d")
                    except:
                        continue
                    transaksiobj = models.TransaksiGudang(
                            keterangan="-",
                            jumlah=row['Masuk'],
                            tanggal=tanggal,
                            KeteranganACC=True,
                            KodeProduk=models.Produk.objects.get(KodeProduk=item),
                            Lokasi=models.Lokasi.objects.get(NamaLokasi = 'FG'),
                            KeteranganACCPurchasing = True
                        )
                    if not pd.isna(row['Keterangan']):
                        keterangan = clean_string(row['Keterangan'])
                        tesartikel = models.Artikel.objects.filter(KodeArtikel=keterangan)
                        detailspk = models.DetailSPK.objects.filter(NoSPK__NoSPK='dummyspk',KodeArtikel__KodeArtikel__icontains=keterangan).first()
                        print(keterangan)
                        print(detailspk)
                        if detailspk == None and tesartikel.exists():
                            detailspk = models.DetailSPK(
                                NoSPK = models.SPK.objects.get(NoSPK = 'dummyspk'),
                                KodeArtikel = tesartikel.first(),
                                Jumlah = 0
                            )
                            print(tesartikel.first())
                            simpan = detailspk.save()
                        
                        transaksiobj.DetailSPK = detailspk
                        
                        # print(asd)
                    print(row['Tanggal'])
                    print(item)
                    print(tanggal)
                    transaksiobj.save()
                    tanggal = row['Tanggal']
                    # except Exception as e:
                    #     produkerror.append([item,e])
                # except Exception as e:
                #         produkerror.append([item,e])
                #         continue
                


        return HttpResponse(f"Berhasil Upload, {produkerror}")

    return render(request, "Purchasing/bulk_createproduk.html")

def clean_string(s):
    # Remove "Art" and any non-alphanumeric characters
    s = re.sub(r'Art', '', s)
    return re.sub(r'[^a-zA-Z0-9]', ' ', s).strip()

@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def view_pemusnahanbarang(request):
    '''
    Fitur ini digunakan untuk menampilkan semua data pemusnahan bahan baku pada sistem
    Algoritma
    1. Mengambil data Pemusnahan Bahan Baku dengan kriteria lokasi = Gudang
    2. Mengubah format tanggal menjadi YYYY-MM-DD
    3. Menampilkan data pemusnahan 
    '''
    dataproduksi = models.PemusnahanBahanBaku.objects.filter(
        lokasi__NamaLokasi="Gudang"
    ).order_by("-Tanggal")
    for i in dataproduksi:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "gudang/view_pemusnahanbarang.html", {"dataproduksi": dataproduksi}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def add_pemusnahanbarang(request):
    '''
    Fitur ini digunakan untuk menambahkan data pemusnahan bahan baku pada gudang
    Algoritma
    1. Mengambil semua data bahan baku pada sistem
    2. Menampilkan form input pemusnhan bahan baku 
    3. Program mendapatkan input berupa Kode Produk, Jumlah, Tanggal, dan Keterangan
    4. Menyimpan transaksi pemusnahan bahan baku pada gudang
    '''
    databarang = models.Produk.objects.all()
    datalokasi = models.Lokasi.objects.filter(NamaLokasi="Gudang")
    if request.method == "GET":
        return render(
            request,
            "gudang/add_pemusnahanbarang.html",
            {"nama_lokasi": datalokasi, "databarang": databarang},
        )
    else:

        kodeproduk = request.POST["produk"]
        lokasi = "Gudang"
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST['keterangan']
        lokasiobj = models.Lokasi.objects.get(NamaLokasi=lokasi)
        try:
            produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Bahan Baku tidak ditemukan")
            return redirect("add_pemusnahangudang")

        pemusnahanobj = models.PemusnahanBahanBaku(
            Tanggal=tanggal, Jumlah=jumlah, KodeBahanBaku=produkobj, lokasi=lokasiobj,Keterangan=keterangan
        )
        pemusnahanobj.save()

        models.transactionlog(
            user="Gudang",
            waktu=datetime.now(),
            jenis="Create",
            pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {produkobj.KodeProduk} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("read_pemusnahanbahangudang")

@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def update_pemusnahanbarang(request, id):
    '''
    Fitur ini digunakan untuk melakukan update transaksi pemusnahan bahan baku pada gudang
    Algoritma
    1. Mendapatkan semua data bahan baku
    2. Mendaatkan data transaksi pemusnahna bahan baku dengan kriteria IDPemusnahanBahanBaku = id (id didapatkan dari passing values HTML)
    3. Program mendapatkan input berupa kode produk, Lokasi, Jumlah, Tanggal, dan Keterangan
    '''
    databarang = models.Produk.objects.all()
    dataobj = models.PemusnahanBahanBaku.objects.get(IDPemusnahanBahanBaku=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.filter(NamaLokasi = "Gudang")
    if request.method == "GET":
        return render(
            request,
            "gudang/update_pemusnahanbarang.html",
            {"data": dataobj, "nama_lokasi": lokasiobj, "dataproduk": databarang},
        )

    else:
        kodeproduk = request.POST["produk"]
        lokasi = "Gudang"
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST['keterangan']
        try:
            produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Bahan Baku tidak ditemukan")
            return redirect("update_pemusnahanbaranggudang", id=id)
        lokasiobj = models.Lokasi.objects.get(NamaLokasi=lokasi)
        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.KodeBahanBaku = produkobj
        dataobj.lokasi = lokasiobj
        dataobj.Keterangan = keterangan

        dataobj.save()

        models.transactionlog(
            user="Gudang",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {produkobj.KodeProduk} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request, "Data berhasil diupdate")
        return redirect("read_pemusnahanbahangudang")

@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def delete_pemusnahanbarang(request, id):
    '''
    Fitur ini digunakan untuk menghapus data pemusnahan bahan baku gudang
    Algoritma
    1. Mendapatkan data pemusnahan bahan baku dengan kriteria IDPemusnahanBahanBaku = id (id didapatkan dari passing values HTML)
    2. Menghapus data pemusnahan bahan baku
    '''
    dataobj = models.PemusnahanBahanBaku.objects.get(IDPemusnahanBahanBaku=id)

    dataobj.delete()

    models.transactionlog(
        user="Gudang",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {dataobj.KodeBahanBaku.KodeProduk} Jumlah : {dataobj.Jumlah} Lokasi : {dataobj.lokasi.NamaLokasi}",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect('read_pemusnahanbahangudang')

# Purchase Order
@login_required
@logindecorators.allowed_users(allowed_roles=["gudang","ppic"])
def view_purchaseorder(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data Purchase Order pada sistem
    Algoritma 
    1. Mengambil data semua purchase order pada sistem
    2. Mengiterasi semua data PO
    3. mengambil data detail PO dengan kriteria KodePO = nomor PO
    '''
    datapo = models.PurchaseOrder.objects.all().order_by("Tanggal")
    # datasjb = models.DetailSuratJalanPembelian.objects.all().order_by(
    #     "NoSuratJalan__Tanggal"
    # )
    date = request.GET.get("mulai")
    dateakhir = request.GET.get("akhir")
    print(date, dateakhir)
    if date =="":
        date = datetime.min.date()
    if dateakhir == "":
        dateakhir = datetime.max.date()
    print(date,dateakhir)
    if date is not None and dateakhir is not None:
        datapo = models.PurchaseOrder.objects.filter(
            Tanggal__range=(date, dateakhir)
        ).order_by("Tanggal")
    for item in datapo:
        item.detailpo = models.DetailPO.objects.filter(KodePO = item.pk)
        item.Tanggal = item.Tanggal.strftime('%Y-%m-%d')
    # for i in date:
    #     i.KodePO.Tanggal = i.KodePO.Tanggal.strftime("%Y-%m-%d")

    if len(datapo) == 0:
        messages.info(request, "Tidak ada data PO")

    return render(
        request,
        "gudang/purchaseorder.html",
        {"datasjb": datapo, "date": date, "mulai": date, "akhir": dateakhir},
    )


@login_required
@logindecorators.allowed_users(allowed_roles=["gudang"])
def trackingpurchaseorder(request,id):
    '''
    Fitur ini digunakan untuk melakukan tracking pada Purchase Order yang ada pada sistem
    1. Mengambil data purchase order dengan kriteria id (primary key) = id (id didapatkan dari passing values HTML)
    2. Mengambil data detail PO dengan kriteria KodePO = data purchase order poin 1
    3. Mengambil data transaksi detail surat jalan pembelian yang menginputkan dengan kode po yang dipilih. 
    4. Mengiterasi data detail PO
    5. mencari selisih dengan cara selisih = Total jumlah PO - TOtal transaksi gudang masuk. Selisih PO artininya selisih kurangnya PO yang belum datang ke perusahaan
    

    '''
    datapo = models.PurchaseOrder.objects.get(id=id)
    datadetailpo =models.DetailPO.objects.filter(KodePO = datapo)
    transaksigudang = models.DetailSuratJalanPembelian.objects.filter(PO__KodePO= datapo)
    # Rekap dan kurang
    datarekap = transaksigudang.values('pk','KodeProduk__KodeProduk',"KodeProduk__NamaProduk","KodeProduk__unit").annotate(total = Sum('Jumlah'))
    print(datarekap)
    print(datadetailpo)
    # print(asd)
    datapo.Tanggal = datapo.Tanggal.strftime('%Y-%m-%d')
    for item in datadetailpo:
        totalpo = item.Jumlah
        totaltransaksigudangmasuk = 0
        transaksigudangmasuk = transaksigudang.filter(PO__pk = item.pk)
        if transaksigudangmasuk.exists():
            totaltransaksigudangmasuk = transaksigudangmasuk.aggregate(total = Sum('Jumlah'))['total']
        selisih = totalpo-totaltransaksigudangmasuk
        item.jumlahmasuk = totaltransaksigudangmasuk
        item.selisih = selisih

    return render(request,'gudang/trackingpo.html',{'datapo':datapo,'datadetailpo':datadetailpo,"transaksigudang":transaksigudang,})


def readcachevalue(request):
    cachevalue = models.CacheValue.objects.all()
    for item in cachevalue:
        item.Tanggal = item.Tanggal.strftime("%Y-%m-%d")
    return render(request, 'gudang/cachevalue.html',{'data' : cachevalue})

def updatecache(request):
    waktustart = time.time()
    allprodukobj = models.Produk.objects.all()
    # allprodukobj = models.Produk.objects.filter(KodeProduk = 'A-004-01')
    user_groups = request.user.groups.values_list('name', flat=True)
    print(user_groups)
    # print(asd)
    for produk in allprodukobj:
        produk.save()
    waktuakhir = time.time()
    if 'rnd' in user_groups:
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        return redirect('dashboardrnd')
    return HttpResponse(f'Waktu proses :{waktuakhir-waktustart}')

def exportbackup(request):
    datasjp = models.SuratJalanPembelian.objects.all()
    datadsjp = models.DetailSuratJalanPembelian.objects.all()
    datamodelssjp = {
        'NoSuratJalan' : [],
        'Tanggal' : [],
        'Supplier' : [],
        'NoInvoice' : [],
        'TanggalInvoice':[],
    }
    datamodelsdsjp ={
        'IDDetailSJPembelian' : [], 
    'NoSuratJalan' :[],
    'KodeProduk' :[],
    'Jumlah' :[],
    'KeteranganACC' :[],
    'Harga' :[],
    'HargaDollar' :[],
    'PPN' :[],
    'PO' :[],
    'hargappn' :[],
    }
    for sjp in datasjp:
        datamodelssjp['NoSuratJalan'].append(sjp.NoSuratJalan)
        datamodelssjp['Tanggal'].append(sjp.Tanggal)
        datamodelssjp['Supplier'].append(sjp.supplier)
        datamodelssjp['NoInvoice'].append(sjp.NoInvoice)
        datamodelssjp['TanggalInvoice'].append(sjp.TanggalInvoice)

    # Iterasi untuk DetailSuratJalanPembelian
    for dsjp in datadsjp:
        datamodelsdsjp['IDDetailSJPembelian'].append(dsjp.IDDetailSJPembelian)
        datamodelsdsjp['NoSuratJalan'].append(dsjp.NoSuratJalan.NoSuratJalan)  # Relasi FK ke NoSuratJalan
        datamodelsdsjp['KodeProduk'].append(dsjp.KodeProduk.KodeProduk)  # Relasi FK ke Produk
        datamodelsdsjp['Jumlah'].append(dsjp.Jumlah)
        datamodelsdsjp['KeteranganACC'].append(dsjp.KeteranganACC)
        datamodelsdsjp['Harga'].append(dsjp.Harga)
        datamodelsdsjp['HargaDollar'].append(dsjp.HargaDollar)
        datamodelsdsjp['PPN'].append(dsjp.PPN)
        datamodelsdsjp['PO'].append(dsjp.PO.KodePO if dsjp.PO else None)  # Jika ada PO, ambil KodePO
        datamodelsdsjp['hargappn'].append(dsjp.hargappn)


    dfsjp = pd.DataFrame(datamodelssjp)
    dfdsjp = pd.DataFrame(datamodelsdsjp)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dfsjp.to_excel(
                                    writer, index=False, sheet_name="SJP"
                    )
        dfdsjp.to_excel(
                                    writer, index=False, sheet_name="DSJP"
                    )
        

    buffer.seek(0)
    wb = load_workbook(buffer)
    # ws = wb["Laporan Persediaan"]
    # ws2 = wb["Barang Masuk"]
    # ws3 = wb["Barang Keluar"]

    # # Insert header "Januari" above the table
    # ws["A1"] = listbulan[waktuobj.month - 1]
    # ws2["A1"] = listbulan[waktuobj.month - 1]
    # ws3["A1"] = listbulan[waktuobj.month - 1]

    # Custom WS Barang Keluar
    # baristerakhirws3 = ws3.max_row
    # kolomterakhirws3 = ws3.max_column
    # hurufkolomws3 = get_column_letter(kolomterakhirws3)
    # print("Baris Terakhir : ", baristerakhirws3)
    # print("Kolom Terakhir : ", kolomterakhirws3)
    # print("Huruf Terakhir : ", hurufkolomws3)
    # ws3.cell(row=baristerakhirws3 + 1, column=kolomterakhirws3 - 1, value="Total Biaya")
    # ws3.cell(row=baristerakhirws3 + 1, column=kolomterakhirws3, value=totalbiayakeluar)
    # # print(asdas)
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=laporanpersediaan.xlsx"
    )
    return response

def bulkaddsjpdsjp (request):
    if request.method == "POST" and request.FILES["file"]:
        kodebahanerror = []
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        # sheet_names = ['A-004-154','A-004-155','A-005-158','B-012-10']
        sheet_names = ['DSJP']

        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item)
            print(item)
            print(df)
            # print(asd)
            # for index,row in df.iterrows():
            #     if pd.isna(row['NoInvoice']):
            #         invoice = None
            #     else:
            #         invoice = row['NoInvoice']
            #     if pd.isna(row['TanggalInvoice']):
            #         tanggalinvoice = None
            #     else:
            #         tanggalinvoice = row['TanggalInvoice']
            #     dataobj = models.SuratJalanPembelian(
            #         NoSuratJalan = row['NoSuratJalan'],
            #         Tanggal = row['Tanggal'],
            #         supplier = row['Supplier'],
            #         NoInvoice = invoice,
            #         TanggalInvoice = tanggalinvoice
            #     )
            #     dataobj.save()

            # i = 0
            for index, row in df.iterrows():
                print(row)
                # print(asd)

                po = None
                

                dataobj= models.DetailSuratJalanPembelian(
                     NoSuratJalan = models.SuratJalanPembelian.objects.get(NoSuratJalan = row['NoSuratJalan']),
    KodeProduk = models.Produk.objects.get(KodeProduk = row['KodeProduk']),
    Jumlah = row['Jumlah'],
    KeteranganACC = row['KeteranganACC'],
    Harga = row['Harga'],
    HargaDollar =row['HargaDollar'],
    PPN = row['PPN'],
    PO = po,
    hargappn = row['hargappn'],
                )
                dataobj.save()
                # datadelete = models.DetailSuratJalanPembelian.objects.filter(
                #             KodeProduk=models.Produk.objects.get(KodeProduk=item),
                #             )
                # print(datadelete)
                # for data in datadelete:
                #     print(data.Jumlah, data.NoSuratJalan.Tanggal)
                #     data.delete()
                # print(asd)
                # if i < 2:
                #     i += 1
                #     continue
                # print(row["Tanggal"])
                # print(row)
                

        return render(request,'error/errorsjp.html',{'data':kodebahanerror})

    return render(request, "Purchasing/bulk_createproduk.html")
