from django.shortcuts import render, redirect, HttpResponse
from django.contrib import messages
from . import models
from django.db.models import Sum
from datetime import datetime, timedelta,date
import pandas as pd
import re
from . import logindecorators
from django.contrib.auth.decorators import login_required
import math
from urllib.parse import quote
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl import Workbook, load_workbook
import time
from openpyxl.styles import PatternFill, Border, Side
from itertools import zip_longest


# Dashboard Produksi
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def dashboard(request):
    '''
    Fitur ini digunakan untuk menampilkan halaman dashboard dari produksi
    Algoritma:
    1. Mengambil semua data Artikel 
    2. Mencari data penyusun dengan last edit kurang dari 7 hari dari sekarang.
    '''
    harisekarang = (datetime.now()+timedelta(days=1)).date().strftime('%Y-%m-%d')
    harisebelumnya = (datetime.now() - timedelta(days=7)).date().strftime('%Y-%m-%d')
    datapenyusunfiltered = models.Penyusun.objects.filter(lastedited__range =(harisebelumnya,harisekarang) )
    artikels = models.Artikel.objects.all()
    for data in datapenyusunfiltered:
        data.lastedited = data.lastedited.strftime('%Y-%m-%d %H:%M:%S')
    alldata = []
    print(datapenyusunfiltered)
    return render(request, "produksi/dashboard.html",{"data": datapenyusunfiltered})


#Load dropdown
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])    
def load_detailspk(request):
    '''
    Fitur ini digunakan untuk supporting Ajax dalam memuat detail SPK berdasarkan Nomor SPK
    '''
    no_spk = request.GET.get("nomor_spk")
    id_spk = models.SPK.objects.get(NoSPK=no_spk)
    if id_spk.StatusDisplay == False :
        detailspk = models.DetailSPK.objects.filter(NoSPK=id_spk.id,)
    else :
        detailspk = models.DetailSPKDisplay.objects.filter(NoSPK = id_spk.id,)

    return render(request, "produksi/opsi_spk.html", {"detailspk": detailspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_htmx(request):
    no_spk = request.GET.get("nomor_spk")
    id_spk = models.SPK.objects.get(NoSPK=no_spk)
    detailspk = models.DetailSPK.objects.filter(NoSPK=id_spk.id)

    return render(request, "produksi/opsi_htmx.html", {"detailspk": detailspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_artikel(request):
    kode_artikel = request.GET.get("kode_artikel")
    artikelobj = models.Artikel.objects.get(KodeArtikel=kode_artikel)
    detailspk = models.DetailSPK.objects.filter(KodeArtikel=artikelobj, NoSPK__StatusAktif=1)
    print(kode_artikel)

    return render(request, "produksi/opsi_artikel.html", {"detailspk": detailspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_detailspkfromartikel(request):
    kode_artikel = request.GET.get("kode_artikel")
    artikelobj = models.Artikel.objects.get(KodeArtikel=kode_artikel)
    detailspk = models.DetailSPK.objects.filter(KodeArtikel=artikelobj).distinct()
    print(detailspk)

    return render(request, "produksi/opsi_detailspkfromartikel.html", {"detailspk": detailspk})
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_detailspkfromartikeltambahan(request):
    kode_artikel = request.GET.get("kode_artikel")
    artikelobj = models.Artikel.objects.get(KodeArtikel=kode_artikel)
    detailspk = models.DetailSPK.objects.filter(KodeArtikel=artikelobj).distinct()
    print(detailspk)

    return render(request, "produksi/opsi_detailspkfromartikeltambahan.html", {"detailspk": detailspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_penyusun(request):
    kodeartikel = request.GET.get("artikel")
    penyusun = models.Penyusun.objects.filter(KodeArtikel=kodeartikel)

    return render(request, "produksi/opsi_penyusun.html", {"penyusun": penyusun})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_display(request):
    print(request.GET)
    kode_display = request.GET.get("kode_artikel")
    displayobj = models.Display.objects.get(KodeDisplay=kode_display)
    detailspk = models.DetailSPKDisplay.objects.filter(KodeDisplay=displayobj.id,NoSPK__StatusDisplay = 1).distinct()
    print(detailspk)

    return render(request, "produksi/opsi_spkdisplay.html", {"detailspk": detailspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_spkmutasi(request):
    print(request.GET)
    kodeartikel = request.GET.get("kode_artikel")
    displayobj = models.Artikel.objects.get(KodeArtikel=kodeartikel)
    detailspk = models.DetailSPK.objects.filter(KodeArtikel=displayobj.id,NoSPK__StatusDisplay = 0, NoSPK__StatusAktif=1)
    print(detailspk)

    return render(request, "produksi/opsi_spkmutasi.html", {"detailspk": detailspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_spkmutasidisplay(request):
    print(request.GET)
    kodeartikel = request.GET.get("kode_artikel")
    displayobj = models.Display.objects.get(KodeDisplay=kodeartikel)
    detailspk = models.DetailSPKDisplay.objects.filter(KodeDisplay=displayobj.id,NoSPK__StatusDisplay = 1, NoSPK__StatusAktif=1)
    print(detailspk)

    return render(request, "produksi/opsi_spkmutasi.html", {"detailspk": detailspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_versi(request):
    kodeartikel = request.GET.get("artikel")
    print(request.GET)
    penyusun = models.Penyusun.objects.filter(KodeArtikel__id = kodeartikel).values_list('versi',flat=True).distinct()
    print(penyusun)

    return render(request, "produksi/opsi_versi.html", {"penyusun": penyusun,'kodeartikel':kodeartikel})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def load_penyusun2(request):
    print(request.GET)
    kodeartikel = request.GET.get("artikel")
    versi = request.GET.get("versi")
    tanggalversi = datetime.strptime(versi, '%b. %d, %Y')
    versi = tanggalversi.strftime('%Y-%m-%d')

    penyusun = models.Penyusun.objects.filter(KodeArtikel=kodeartikel,versi = versi)
    print(penyusun)

    return render(request, "produksi/opsi_penyusun.html", {"penyusun": penyusun})


# SPK
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_spk(request):
    '''
    Fitur ini digunakan untuk manajamen data SPK 
    Algoritma
    1. mengambil data SPK
    2. Mengiterasi data SPK
    3. Mengambik data detailSPK berdasarkan iterasi data spk
    '''
    dataspk = models.SPK.objects.all().order_by("-Tanggal")

    for j in dataspk:
        j.Tanggal = j.Tanggal.strftime('%Y-%m-%d')
        total_detail_spk = models.DetailSPK.objects.filter(NoSPK=j).values('KodeArtikel').annotate(total=Sum('Jumlah'))
        if j.StatusDisplay ==False:
            j.detailspk = models.DetailSPK.objects.filter(NoSPK=j)
        else:
            j.detailspk = models.DetailSPKDisplay.objects.filter(NoSPK=j)
        for detail_spk in total_detail_spk:
            kode_artikel = detail_spk['KodeArtikel']
            total_requested = detail_spk['total']
            
            total_processed = models.DetailSPPB.objects.filter(DetailSPK__NoSPK=j, DetailSPK__KodeArtikel=kode_artikel).aggregate(total=Sum('Jumlah'))['total'] or 0
            
            if total_processed < total_requested:
                is_lunas = False
                break
            

    return render(request, "produksi/view_spk.html", {"dataspk": dataspk})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_spk(request):
    '''
    Fitur ini digunakan untuk menambahkan data SPK
    Algoritma
    1. Mengambil data Artikel
    2. Mengambil data Display
    3. Menampilkan form input data SPK
    4. Program mendapatkan input Nomor SPK, Tanggal (terbit), Keterangan, Jenisspk
    5. Apabila Nomor SPK sudah ada pada sistem sebelumnya maka akan menampilkan pesan error
    6. membuat object data SPK dengan kriteria NoSPK = nomor spk input, Tanggal = tanggal input, Keterangan = keterangan input, KeteranganACC = False (belum di acc purchasig), StatusAktif = 1 (Aktif), JenisSPK = jenis spk input
    7. Apabila jenis spk adalah Artikel maka 
    8. Mengambil list artikel[]
    9. mengambil list quantity[]
    10. Mengiterasi artikel dan jumlah 
    11. membuat data detailSPk Artikel dengan kriteria NoSPK = nospk terakhir, KodeArtikel = kode artikel iterasi, Jumlah = jumlah iterasi
    12. menyimpan data detailspk
    '''
    dataartikel = models.Artikel.objects.all()
    datadisplay = models.Display.objects.all()
    if request.method == "GET":
        return render(request, "produksi/add_spk.html", {"data": dataartikel,'datadisplay':datadisplay})

    if request.method == "POST":
        print(request.POST)
        # print(asd)
        nomor_spk = request.POST["nomor_spk"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]
        try:
            jenisspk = request.POST['jenisspk']
        except:
            messages.error(request, "Masukkan Jenis SPK")
            return redirect("add_spk")

        dataspk = models.SPK.objects.filter(NoSPK=nomor_spk).exists()
        if dataspk:
            messages.error(request, "Nomor SPK sudah ada")
            return redirect("add_spk")
        else:
            
            data_spk = models.SPK(
                NoSPK=nomor_spk,
                Tanggal=tanggal,
                Keterangan=keterangan,
                KeteranganACC=False,
                StatusAktif = 1
            )
            if jenisspk == 'spkartikel' : 
                StatusDisplay = 0
            else : 
                StatusDisplay = 1
            
            data_spk.StatusDisplay = StatusDisplay
            

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Surat Perintah Kerja. No SPK : {nomor_spk} Keterangan : {keterangan} Status Display : {StatusDisplay}",
            ).save()
            
            artikel_list = request.POST.getlist("artikel[]")
            jumlah_list = request.POST.getlist("quantity[]")

            for produk, jumlah in zip(artikel_list, jumlah_list):
                # Pisahkan KodeArtikel dari jumlah dengan delimiter '/'
                if jenisspk == "spkartikel":
                    try:
                        kode_artikel = models.Artikel.objects.get(KodeArtikel=produk)
                        data_spk.save()
                        no_spk = models.SPK.objects.get(NoSPK=nomor_spk)


                        # Simpan data ke dalam model DetailSPK
                        datadetailspk = models.DetailSPK(
                            NoSPK=no_spk, KodeArtikel=kode_artikel, Jumlah=jumlah
                        )
                        models.transactionlog(
                            user="Produksi",
                            waktu=datetime.now(),
                            jenis="Create",
                            pesan=f"Detail Surat Perintah Kerja. No SPK : {no_spk.NoSPK} Kode Artikel : {kode_artikel.KodeArtikel} Jumlah : {jumlah}",
                        ).save()
                        datadetailspk.save()
                        messages.success(request, "Data berhasil disimpan")
                    except models.Artikel.DoesNotExist:
                        messages.error(request,f"Artikel {produk} tidak ditemukan dalam sistem")

                elif jenisspk == "spkdisplay":
                    try :
                        kode_display = models.Display.objects.get(KodeDisplay = produk)
                        data_spk.save()
                        no_spk = models.SPK.objects.get(NoSPK=nomor_spk)

                        # Simpan data dalam model detailSPK
                        datadetailspk = models.DetailSPKDisplay(
                            NoSPK = no_spk, KodeDisplay = kode_display, Jumlah = jumlah
                        )
                        models.transactionlog(
                            user="Produksi",
                            waktu=datetime.now(),
                            jenis="Create",
                            pesan=f"Detail Surat Perintah Kerja. No SPK : {no_spk.NoSPK} Kode Display : {kode_display.KodeDisplay} Jumlah : {jumlah}",
                        ).save()
                        datadetailspk.save()
                        messages.success(request, "Data berhasil disimpan")
                    except models.Display.DoesNotExist:
                        messages.error(request,f'Display {produk} tidak ditemukan dalam sistem')

            return redirect("view_spk")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def detail_spk(request, id):
    '''
    fitur ini digunakan untuk mengupdate data detail SPK
    Algoritma
    1. Mengambil data artikel 
    2. Mengambil data Display
    3. Mengambil data SPK dengan kriteria id (primarykey) = id (id didapatkan dari passing values HTML)
    4. memfilter data detailSPK dengan kode SPK = dataspk
    5. Menampilkan form update data SPK dan detail SPK
    6. Program menerima input Nomor SPK, Tnaggal, Keterangan, list Artikel, List Jumlah, Status Aktif
    7. Mengupdate data nomorspk dengan No SPK input
    8. Cek apabila sudah ada data dengan nomor SPK tersebut maka akan menampilkan pesan error
    9. mengupdate data Tnaggal, Nomor SPk, keterangan, Status Aktif
    10. Mengiterasi List Artikel/display, List Jumlah
    11. Mengupdate data detailSPK
    '''
    dataartikel = models.Artikel.objects.all().order_by('KodeArtikel')
    datadisplay =models.Display.objects.all().order_by('KodeDisplay')
    dataspk = models.SPK.objects.get(id=id)
    if dataspk.StatusDisplay == False:
        datadetail = models.DetailSPK.objects.filter(NoSPK=dataspk.id)
    else:
        datadetail = models.DetailSPKDisplay.objects.filter(NoSPK = dataspk.id)

    if request.method == "GET":
        tanggal = datetime.strftime(dataspk.Tanggal, "%Y-%m-%d")

        return render(
            request,
            "produksi/detail_spk.html",
            {
                "data": dataartikel,
                "datadisplay" : datadisplay,
                "dataspk": dataspk,
                "datadetail": datadetail,
                "tanggal": tanggal,
            },
        )

    elif request.method == "POST":
        nomor_spk = request.POST["nomor_spk"]
        tanggall = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]
        artikel_list = request.POST.getlist("artikel[]")
        jumlah_list = request.POST.getlist("quantity[]")
        statusaktif = request.POST['StatusAktif']

        spkbaru = models.SPK.objects.filter(NoSPK=nomor_spk)
        existspk = spkbaru.exists()
        print(dataspk)
        print(spkbaru)
        if existspk and dataspk != spkbaru.first():
            messages.error(request, "Nomor SPK sudah ada")
            return redirect("detail_spk", id=id)
        else:
            dataspk.NoSPK = nomor_spk
            dataspk.Tanggal = tanggall
            dataspk.Keterangan = keterangan
            dataspk.StatusAktif = statusaktif
            dataspk.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"Surat Perintah Kerja. No SPK : {nomor_spk} Keterangan : {keterangan} Status Aktif : {statusaktif}",
            ).save()
        
        if dataspk.StatusDisplay == False:
            for detail, artikel_id, jumlah in zip(datadetail, artikel_list, jumlah_list):
                kode_artikel = models.Artikel.objects.get(KodeArtikel=artikel_id)
                detail.KodeArtikel = kode_artikel
                detail.Jumlah = jumlah
                detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Detail Surat Perintah Kerja. No SPK : {dataspk.NoSPK} Kode Artikel : {kode_artikel.KodeArtikel} Jumlah : {jumlah}",
                ).save()
            
            no_spk = models.SPK.objects.get(NoSPK=nomor_spk)

            for artikel_id, jumlah in zip(artikel_list[len(datadetail) :], jumlah_list[len(datadetail) :]):
                kode_artikel = models.Artikel.objects.get(KodeArtikel=artikel_id)
                new_detail = models.DetailSPK.objects.create(
                    NoSPK=no_spk,  # Assuming NoSPK is the ForeignKey field to SPK in DetailSPK model
                    KodeArtikel=kode_artikel,
                    Jumlah=jumlah,
                )

                new_detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail Surat Perintah Kerja. Kode Artikel : {kode_artikel.KodeArtikel} Jumlah : {jumlah}",
                ).save()

        else:
            for detail, artikel_id, jumlah in zip(datadetail, artikel_list, jumlah_list):
                kode_artikel = models.Display.objects.get(KodeDisplay=artikel_id)
                detail.KodeDisplay = kode_artikel
                detail.Jumlah = jumlah
                detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Detail Surat Perintah Kerja. No SPK : {dataspk.NoSPK} Kode Artikel : {kode_artikel.KodeDisplay} Jumlah : {jumlah}",
                ).save()

            no_spk = models.SPK.objects.get(NoSPK=nomor_spk)
            
            for artikel_id, jumlah in zip(artikel_list[len(datadetail) :], jumlah_list[len(datadetail) :]):
                kode_artikel = models.Display.objects.get(KodeDisplay=artikel_id)
                new_detail = models.DetailSPKDisplay.objects.create(
                    NoSPK=no_spk,  # Assuming NoSPK is the ForeignKey field to SPK in DetailSPK model
                    KodeDisplay=kode_artikel,
                    Jumlah=jumlah,
                )

                new_detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail Surat Perintah Kerja. Kode Display : {kode_artikel.KodeDisplay} Jumlah : {jumlah}",
                ).save()

        messages.success(request,'Data berhasil di update')
        return redirect("detail_spk", id=id)

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_spk(request, id):
    '''
    Fitur ini digunakan untuk menghapus data SPK
    Algoritma
    1. Mendapatkan data SPK dengan id (primarykey) = id ( id berasal dari passing values HTML)
    2. Menghapus data SPK
    '''
    dataspk = models.SPK.objects.get(id=id)
    dataspk.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Surat Perintah Kerja. Nomor SPK : {dataspk.NoSPK} ",
    ).save()
    messages.success(request,'SPK Berhasil dihapus')
    return redirect("view_spk")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_detailspk(request, id):
    '''
    Fitur ini digunaakn untuk menhhapus detail SPK Artikel
    1. mengambil data detailSPK dengan kriteria IDDetailSPK = id (id didapatkan dari passing values HTML)
    2. Menghapus data detailSPK
    '''
    datadetailspk = models.DetailSPK.objects.get(IDDetailSPK=id)
    dataspk = models.SPK.objects.get(NoSPK=datadetailspk.NoSPK)
    datadetailspk.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Detail Surat Perintah Kerja. Nomor SPK : {dataspk.NoSPK} Artikel : {datadetailspk.KodeArtikel.KodeArtikel}",
    ).save()

    return redirect("detail_spk", id=dataspk.id)

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_detailspkdisplay(request, id):
    '''
    Fitur ini digunaakn untuk menhhapus detail SPK Display
    1. mengambil data detailSPKDisplay dengan kriteria IDDetailSPK = id (id didapatkan dari passing values HTML)
    2. Menghapus data detailSPKDisplay
    '''
    datadetailspk = models.DetailSPKDisplay.objects.get(IDDetailSPK=id)
    dataspk = models.SPK.objects.get(NoSPK=datadetailspk.NoSPK)
    datadetailspk.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Detail Surat Perintah Kerja. Nomor SPK : {dataspk.NoSPK} Kode Artikel : {datadetailspk.KodeDisplay.KodeDisplay}",
    ).save()

    return redirect("detail_spk", id=dataspk.id)

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def track_spk(request, id):
    '''
    Fitur ini digunakan untuk melakukan tracking pada SPK
    Algoritma
    1. Mengambil data Artikel
    2. Mengambil data Display
    3. Mengambil data SPK dengan kriteria id (primarykey) = id (id didapatkan dari passing values HTML)
    4. Mengambil data DetailSPK dengan kriteria NoSPK poin 3
    5. Apabila SPK adalah Display 
        A. Mengambil data transaksi Gudang  dengan NO SPK Display sana dengan NoSPK poin 3
        B. Mengambil data transaksi Produksi dengan NO SPK Display sama dengan NoSPK poin 3
        C. Mengambil data SPPB degan no SPK Display sama dengan NoSPK poin 3
        D. Menghitung agregasi total permintaan bahan baku
        E. Menghitung agregasi total pengiriman display
    6. Apabila SPK adalah Artikel
        A. Mengambil data transaksi Gudang  dengan NO SPK Artikel sama dengan NoSPK poin 3
        B. Mengambil data transaksi Produksi dengan NO SPK Artikel sama dengan NoSPK poin 3
        C. Mengambil data SPPB degan no SPK Artikel sama dengan NoSPK poin 3
        D. Menghitung agregasi total permintaan bahan baku
        E. Menghitung agregasi total pengiriman display
    '''
    dataartikel = models.Artikel.objects.all()
    datadisplay =models.Display.objects.all()
    dataspk = models.SPK.objects.get(id=id)
    if dataspk.StatusDisplay == False:
        datadetail = models.DetailSPK.objects.filter(NoSPK=dataspk.id)
    else:
        datadetail = models.DetailSPKDisplay.objects.filter(NoSPK = dataspk.id)
    if dataspk.StatusDisplay ==True:
        
        # Data SPK terkait yang telah di request ke Gudang
        transaksigudangobj = models.TransaksiGudang.objects.filter(
            DetailSPKDisplay__NoSPK=dataspk.id, jumlah__gte=0
        )

        # Data SPK Terkait yang telah jadi di FG
        transaksiproduksiobj = models.TransaksiProduksi.objects.filter(
            DetailSPKDisplay__NoSPK=dataspk.id, Jenis="Mutasi"
        )
        print(transaksiproduksiobj)

        # Data SPK Terkait yang telah dikirim
        sppbobj = models.DetailSPPB.objects.filter(DetailSPKDisplay__NoSPK=dataspk.id)
        rekapjumlahpermintaanperbahanbaku = transaksigudangobj.values('KodeProduk__KodeProduk',"KodeProduk__NamaProduk","KodeProduk__unit").annotate(total = Sum('jumlah'))
        rekapjumlahpengirimanperartikel = sppbobj.values("DetailSPKDisplay__KodeDisplay__KodeDisplay").annotate(total=Sum('Jumlah'))
    else:
        # Data SPK terkait yang telah di request ke Gudang
        transaksigudangobj = models.TransaksiGudang.objects.filter(
            DetailSPK__NoSPK=dataspk.id, jumlah__gte=0
        )

        # Data SPK Terkait yang telah jadi di FG
        transaksiproduksiobj = models.TransaksiProduksi.objects.filter(
            DetailSPK__NoSPK=dataspk.id, Jenis="Mutasi"
        )

        # Data SPK Terkait yang telah dikirim
        sppbobj = models.DetailSPPB.objects.filter(DetailSPK__NoSPK=dataspk.id)
        rekapjumlahpermintaanperbahanbaku = transaksigudangobj.values('KodeProduk__KodeProduk',"KodeProduk__NamaProduk","KodeProduk__unit").annotate(total = Sum('jumlah'))
        rekapjumlahpengirimanperartikel = sppbobj.values("DetailSPK__KodeArtikel__KodeArtikel").annotate(total=Sum('Jumlah'))

    if request.method == "GET":
        tanggal = datetime.strftime(dataspk.Tanggal, "%Y-%m-%d")


    print(transaksigudangobj)
    return render(
            request,
            "produksi/trackingspk.html",
            {
                "data": dataartikel,
                "datadisplay": datadisplay,
                "dataspk": dataspk,
                "datadetail": datadetail,
                "tanggal": tanggal,
                "transaksigudang": transaksigudangobj,
                "transaksiproduksi": transaksiproduksiobj,
                "transaksikeluar": sppbobj,
                'datarekappermintaanbahanbaku': rekapjumlahpermintaanperbahanbaku,
                'datarekappengiriman' : rekapjumlahpengirimanperartikel
            },
        )


# SPPB   
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_sppb(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data SPPB
    Algoritma:
    1. mengambil data SPPB 
    2. mengubah format tanggal menjadi YYYY-MM-DD
    '''
    datasppb = models.SPPB.objects.all().order_by("-Tanggal")
    for i in datasppb:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")
        i.detailsppb = models.DetailSPPB.objects.filter(NoSPPB = i)

    return render(request, "produksi/view_sppb.html", {"datasppb": datasppb})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_sppb(request):
    '''
    Fitur ini digunakan untuk menambahkan data SPPB
    1. Mengambil data bahan baku
    2. Mengambil data Artikel
    3. Mengambil data Display 
    4. Mengambil data Confirmationorder dengan kriteria StatusAktif = True (mengindikasikan CO masih aktif)
    5. Menampilkan form input SPPB
    6. Program mendapatkan input Nomor SPBB, Tanggal, Keterangan, list detailspkdisplay, list quantyty display, list purchase order
    7. Apabila no spbb sudah ada pada sistem maka akan menampilkan pesan error
    8. mengambil data list detailspk, list quantity, list purchase order, list versiartikel, list kodebahan, list quantity bahan, list purchaseorder
    9. mengiterasi list bahanlist, list jumlahbahan, list confirmation order
    10. membuat objek detailSPPB dengan kriteria noSPPB = no SPPB, DetailBahan= kodebahan, Jumlah = jumlah bahan iterasi
    11. Menyimpan data detailSPPB
    '''
    databahan = models.Produk.objects.all()
    dataartikel = models.Artikel.objects.all()
    datadisplay = models.Display.objects.all()
    purchaseorder = models.confirmationorder.objects.filter(StatusAktif = True)

    if request.method == "GET":
        return render(
            request,
            "produksi/add_sppb.html",
            {
                "bahan": databahan,
                "artikel": dataartikel,
                "display": datadisplay,
                "purchaseorder":purchaseorder
            },
        )

    if request.method == "POST":
        nomor_sppb = request.POST["nomor_sppb"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]
        
        displaylist = request.POST.getlist("detail_spkdisplay[]")
        jumlahdisplay = request.POST.getlist('quantitydisplay[]')
        confirmationorderdisplay = request.POST.getlist("purchaseorderdisplay")
        print(request.POST)
        # print(asd)
        datasppb = models.SPPB.objects.filter(NoSPPB=nomor_sppb).exists()
        if datasppb:
            messages.error(request, "Nomor SPPB sudah ada")
            return redirect("add_sppb")
        else:
            artikel_list = request.POST.getlist("detail_spk[]")
            jumlah_list = request.POST.getlist("quantity[]")
            confirmationorderartikel = request.POST.getlist("purchaseorderartikel")
            listversi = request.POST.getlist('versiartikel')

            bahan_list = request.POST.getlist("kode_bahan[]")
            jumlahbahan = request.POST.getlist("quantitybahan[]")
            confirmationorderbahan = request.POST.getlist("purchaseorderbahan")

            # Periksa apakah setidaknya satu item dari artikel atau display list memiliki data
            valid_bahan_list = any(bahan_list) and any(jumlahbahan)
            valid_artikel_list = any(artikel_list) and any(jumlah_list)
            valid_display_list = any(displaylist) and any(jumlahdisplay)

            if valid_artikel_list or valid_display_list or valid_bahan_list:

                data_sppb = models.SPPB(
                    NoSPPB=nomor_sppb, Tanggal=tanggal, Keterangan=keterangan
                ).save()
                messages.success(request, "Data berhasil disimpan")

                no_sppb = models.SPPB.objects.get(NoSPPB=nomor_sppb)

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Surat Perintah Pengiriman Barang. Nomor SPPB : {no_sppb.NoSPPB} Keterangan : {no_sppb.Keterangan}",
                ).save()

                for bahan, jumlah, confirmationorder in zip(bahan_list, jumlahbahan, confirmationorderbahan):
                    if bahan == '' or jumlah == '':
                        continue

                    kode_bahan = models.Produk.objects.get(KodeProduk=bahan)
                    jumlah_bahan = jumlah

                    # Simpan data ke dalam model DetailSPK
                    datadetailspk = models.DetailSPPB(
                        NoSPPB=no_sppb, DetailBahan=kode_bahan, Jumlah=jumlah_bahan
                    )

                    if not confirmationorder == "":
                        datadetailspk.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)

                    datadetailspk.save()

                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {no_sppb.NoSPPB} Kode Bahan Baku : {kode_bahan.NamaProduk} Jumlah : {jumlah} CO : { datadetailspk.IDCO if datadetailspk.IDCO is not None  else None}",
                    ).save()

                for artikel, jumlah, confirmationorder,versi in zip(artikel_list, jumlah_list,confirmationorderartikel,listversi):
                    if artikel == '' or jumlah == '':
                        continue

                    kode_artikel = models.DetailSPK.objects.get(IDDetailSPK=artikel)
                    jumlah_produk = jumlah

                    try:
                        versiobj = models.Versi.objects.get(pk = versi)
                    except:
                        versiobj = models.Versi.objects.filter(KodeArtikel = kode_artikel.KodeArtikel,isdefault = True)

                    # Simpan data ke dalam model DetailSPK
                    datadetailspk = models.DetailSPPB(
                        NoSPPB=no_sppb, DetailSPK=kode_artikel, Jumlah=jumlah_produk,VersiArtikel = versiobj
                    )

                    if not confirmationorder == "":
                        datadetailspk.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)

                    datadetailspk.save()

                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {no_sppb.NoSPPB} Kode Artikel : {kode_artikel.KodeArtikel} Jumlah : {jumlah} CO : { datadetailspk.IDCO if datadetailspk.IDCO is not None  else None}",
                    ).save()

                for display,jumlah,confirmationorder in zip(displaylist,jumlahdisplay,confirmationorderdisplay):

                    if display == "" or jumlah == "":
                        continue

                    detailspkdisplayobj = models.DetailSPKDisplay.objects.get(IDDetailSPK = display)
                    
                    datadetailspkdisplay = models.DetailSPPB(
                    NoSPPB=no_sppb, DetailSPKDisplay=detailspkdisplayobj, Jumlah=jumlah
                    )

                    if not confirmationorder == "":
                        datadetailspkdisplay.IDCO = models.confirmationorder.objects.get(pk = confirmationorder)

                    datadetailspkdisplay.save()

                    

                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {no_sppb.NoSPPB} Kode Display : {detailspkdisplayobj.KodeDisplay} Jumlah : {jumlah} CO : { datadetailspkdisplay.IDCO if datadetailspkdisplay.IDCO is not None  else None}",
                    ).save()

                return redirect("view_sppb")
            
            else:
                messages.error(request, "Masukkan Bahan, Artikel atau Display")
                return redirect("add_sppb")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def detail_sppb(request, id):
    '''
    Fitur ini digunakan untuk mengupdate data detailSPPB
    Algoritma:
    1. Mendapatkan data Produk 
    2. mendapatkan data Artikel
    3. Mendapatkdan data Display
    4. Mendapatkan data detail SPK
    5. Menampilkan data detailSPK
    6. Menampilkan data detailSPKDisplay
    7. Mendapatkan data DetailSPPB
    8. Mendapatkan data detailSPPBDisplay
    9. mendapatkan data Confirmationorder
    10. Menampilkan form update data SPPB
    11. Mendapatkan data No SPPB, Tanggal, Keterangan, list kode bahan baku awal,  list kuantitas bahn baku awal, list purchase order bahan baku awal, list Artikel awal, List Jumlah awal, Lust confirmation order awal,list versi awal, list display awal, list jumlah display awal, list purchase order display awal, list artikel baru, list jumlah artikel baru, list versi artikel baru, list purchaseorderartikel baru, list display baru, ist jumlah display baru, list purchase order display baru, list bahan baku baru, list jumlah bahan baku baru, list purchase order bahan baku baru.
    12. Cek apakah nosppb ada update atau tidak
    13. Apabila ada update maka akan cek apakah kode sppb sudah ada pada aplikasi atau belum
    14. Apabila sudah ada kode sppb sebelumnya maka akan menampilkan pesan eror
    15. Apabila tidak ada maka lanjut ke nomor 16
    16. mengupdate object SPPB dengan kriteria NoSPPB = nomor sppb, Tanggal = tanggal sppb, Keterangan = keterangan sppb
    17. Cek apakah ada data detail pengiriman bahan baku pada sppb sebelumnya
        A. Apabila ada maka akan mengiterasi tiap detail bahan baku sppb yang ada untuk di update
    18. Cek apakah ada data detail pengiriman artikel pada sppb sebelumnya
        A. Apabila ada maka akan mengiterasi tiap detail artikel dan mengupdate data detail sppb
    19. Cek apakah ada detail pengiriman display pada sppb sebelumnya
        A. Apabila ada maka akan mengiterasi tiap detail display dan mengupdate data detail sppb display
    20. Cek apakah ada pengiriman bahan baku baru ?
        A. Apabila ada maka mengiterasi list bahan baku baru 
        B. Menyimpan data bahan baku baru
    21. Cek apakah ada pengiriman artikel baru ? 
        A. Apabila ada maka mengiterasi list artikel baru 
        B. Menyimpan data artikel baru
    22. Cek apakah ada pengiriman display baru ?
        A. Apabila ada maka mengiterasi list display baru
        B. Menyimpan data display baru
    '''
    
    databahan = models.Produk.objects.all()
    dataartikel = models.Artikel.objects.all().order_by('KodeArtikel')
    datadisplay = models.Display.objects.all().order_by('KodeDisplay')
    datadetailspk = models.DetailSPK.objects.all()
    datadetailspkdisplay = models.DetailSPKDisplay.objects.all()
    datasppb = models.SPPB.objects.get(id=id)
    datadetailsppbbahan = models.DetailSPPB.objects.filter(NoSPPB=datasppb.id,DetailSPK = None,DetailSPKDisplay = None)
    datadetailsppbArtikel = models.DetailSPPB.objects.filter(NoSPPB=datasppb.id,DetailSPKDisplay = None,DetailBahan = None)
    datadetailsppbdisplay = models.DetailSPPB.objects.filter(NoSPPB=datasppb.id,DetailSPK = None,DetailBahan = None)
    purchaseorderdata = models.confirmationorder.objects.filter(StatusAktif =True)
    for item in datadetailsppbArtikel:
        dataspk = models.DetailSPK.objects.filter(KodeArtikel = item.DetailSPK.KodeArtikel).distinct()
        item.opsispk = dataspk
        print(item)
        print(dataspk)
    

    for item in datadetailsppbArtikel:
        item.opsiversi = models.Versi.objects.filter(KodeArtikel = item.DetailSPK.KodeArtikel)

    if request.method == "GET":
        tanggal = datetime.strftime(datasppb.Tanggal, "%Y-%m-%d")
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
        else:
            back_url = '/produksi/viewdetailsppb'

        return render(
            request,
            "produksi/detail_sppb.html",
            {
                "bahan": databahan,
                "dataartikel": dataartikel,
                "datadisplay": datadisplay,
                "data": datadetailspk,
                "dataspkdisplay": datadetailspkdisplay,
                "datasppb": datasppb,
                "datadetailbahan" : datadetailsppbbahan,
                "datadetail": datadetailsppbArtikel,
                "datadetaildisplay": datadetailsppbdisplay,
                "tanggal": tanggal,
                'purchaseorder':purchaseorderdata,
                'backurl':back_url
            },
        )

    elif request.method == "POST":
        print(request.POST)
        # print
        nomor_sppb = request.POST["nomor_sppb"]
        tanggall = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]

        bahan_list = request.POST.getlist("kode_bahanawal[]")
        jumlahbahan = request.POST.getlist("quantitybahan[]")
        confirmationorderbahan = request.POST.getlist("purchaseorderbahan")

        artikel_list = request.POST.getlist("detail_spkawal[]")
        jumlah_list = request.POST.getlist("quantity[]")
        confirmationorderartikel = request.POST.getlist("purchaseorderartikel")
        listversi = request.POST.getlist('versiasli')

        display_list = request.POST.getlist("detail_spkdisplayawal[]")
        jumlahdisplay_list = request.POST.getlist('quantitydisplay[]')
        confirmationorderdisplay = request.POST.getlist("purchaseorderdisplay")

        artikel_baru = request.POST.getlist('detail_spktambahan[]')
        jumlah_baru = request.POST.getlist('quantitybaru[]')
        purchaseorder_artikelbaru = request.POST.getlist('purchaseorderartikelbaru')
        listversibaru = request.POST.getlist('versiartikel')
        

        display_baru = request.POST.getlist('detail_spkdisplay[]')
        jumlahdisplay_baru = request.POST.getlist('quantitydisplaybaru[]')
        purchaseorder_displaybaru = request.POST.getlist('purchaseorderdisplaybaru')

        bahan_baru = request.POST.getlist("kode_bahan[]")
        jumlahbahan_baru = request.POST.getlist("quantitybahanbaru[]")
        purchaseorder_bahanbaru = request.POST.getlist("purchaseorderbahanbaru")

        datasppbbaru = models.SPPB.objects.filter(NoSPPB=nomor_sppb)
        existsppb = datasppbbaru.exists()
        # print(request.POST['detail_spk[]'])
        # print(asd)
        if existsppb and datasppb != datasppbbaru.first():
            messages.error(request, "Nomor SPPB sudah ada")
            return redirect("detail_sppb", id=id)
        else:
            datasppb.NoSPPB = nomor_sppb
            datasppb.Tanggal = tanggall
            datasppb.Keterangan = keterangan
            datasppb.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB} Keterangan : {datasppb.Keterangan}",
            ).save()

        if datadetailsppbbahan:
            # print(asd)
            for detail, bahan, jumlah, confirmationorder in zip(datadetailsppbbahan,bahan_list, jumlahbahan, confirmationorderbahan):

                kode_bahan = models.Produk.objects.get(KodeProduk=bahan)

                detail.DetailBahan = kode_bahan
                detail.Jumlah = jumlah

                if not confirmationorder == "":
                    detail.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)
                    
                detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB} Kode Bahan Baku: {kode_bahan.NamaProduk} Jumlah : {jumlah}",
                ).save()


        if datadetailsppbArtikel:
            print(datadetailsppbArtikel, artikel_list, jumlah_list, confirmationorderartikel,listversi)
            # print(asd)
            for detail, artikel_id, jumlah, confirmationorder,versi in zip(
                datadetailsppbArtikel, artikel_list, jumlah_list, confirmationorderartikel,listversi
            ):
                kode_artikel = models.DetailSPK.objects.get(IDDetailSPK=artikel_id)
                detail.DetailSPK = kode_artikel
                detail.Jumlah = jumlah
                try:
                    versiobj = models.Versi.objects.get(id = versi)
                except:
                    versiobj = models.Versi.objects.filter(kode_artikel = detail.DetailSPK.KodeArtikel, isdefault = True).first()
                detail.VersiArtikel  = versiobj

                if not confirmationorder == "":
                    detail.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)

                detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB} Kode Artikel : {kode_artikel.KodeArtikel} Jumlah : {jumlah}",
                ).save()
                
        if datadetailsppbdisplay:
            for detail, artikel_id, jumlah, confirmationorder in zip(
                datadetailsppbdisplay, display_list, jumlahdisplay_list, confirmationorderdisplay
            ):
                kode_display = models.DetailSPKDisplay.objects.get(IDDetailSPK=artikel_id)
                detail.DetailSPKDisplay = kode_display
                detail.Jumlah = jumlah

                if not confirmationorder == "":
                    detail.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)

                detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB} Kode Artikel : {kode_display.KodeDisplay} Jumlah : {jumlah}",
                ).save()

        no_sppb = models.SPPB.objects.get(NoSPPB=nomor_sppb)

        if bahan_baru:
            for bahan, jumlah, confirmationorder in zip(bahan_baru, jumlahbahan_baru, purchaseorder_bahanbaru):
                if bahan == '' or jumlah == '':
                    continue

                kode_bahan = models.Produk.objects.get(KodeProduk=bahan)
                jumlah_bahan = jumlah

                # Simpan data ke dalam model DetailSPK
                datadetailspk = models.DetailSPPB(
                    NoSPPB=no_sppb, DetailBahan=kode_bahan, Jumlah=jumlah_bahan
                )

                print(confirmationorder)

                if not confirmationorder == "":
                    datadetailspk.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)

                datadetailspk.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {no_sppb.NoSPPB} Kode Bahan Baku : {kode_bahan.NamaProduk} Jumlah : {jumlah} CO : { datadetailspk.IDCO if datadetailspk.IDCO is not None  else None}",
                ).save()

        if artikel_baru:
            for artikel_id, jumlah,confirmationorder,versi in zip(
                artikel_baru,jumlah_baru,purchaseorder_artikelbaru,listversibaru
            ):
                kode_artikel = models.DetailSPK.objects.get(IDDetailSPK=artikel_id)
                new_detail = models.DetailSPPB(
                    NoSPPB=no_sppb,  # Assuming NoSPK is the ForeignKey field to SPK in DetailSPK model
                    DetailSPK=kode_artikel,
                    Jumlah=jumlah,
                )
                try:
                    versiobj = models.Versi.objects.get(pk = versi)
                except:
                    versiobj = models.Versi.objects.filter(isdefault = True,KodeArtikel = kode_artikel.KodeArtikel)
                new_detail.VersiArtikel = versiobj
                if not confirmationorder == "":
                    new_detail.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)
                new_detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {no_sppb.NoSPPB} Kode Artikel : {kode_artikel.KodeArtikel} Jumlah : {jumlah}",
                ).save()

        if display_baru:
            for display_id, jumlah, confirmationorder in zip(display_baru,jumlahdisplay_baru,purchaseorder_displaybaru):
                kode_display = models.DetailSPKDisplay.objects.get(IDDetailSPK = display_id)
                new_detail = models.DetailSPPB(
                    NoSPPB=no_sppb,  # Assuming NoSPK is the ForeignKey field to SPK in DetailSPK model
                    DetailSPKDisplay=kode_display,
                    Jumlah=jumlah
                    )
                
                if not confirmationorder == "":
                    new_detail.IDCO = models.confirmationorder.objects.get(pk=confirmationorder)

                new_detail.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {no_sppb.NoSPPB} Kode Display : {kode_display.KodeDisplay} Jumlah : {jumlah}",
                ).save()

        messages.success(request,"Data berhasil disimpan")
        return redirect("detail_sppb", id=id)
    
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_sppb(request, id):
    '''
    Fitur ini digunakan untuk menghapus data SPPB
    Algoritma : 
    1. Mendapatkan data SPPB dengan id (primarykey) =  id (id didapatkan dari passing values id)
    2. menghapus data SPPB
    '''
    datasppb = models.SPPB.objects.get(id=id)
    datasppb.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB} ",
    ).save()
    messages.success(request,'Data SPPB Berhasil dihapus')
    return redirect("view_sppb")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_detailsppb(request, id):
    '''
    Fitur ini digunakan untuk menghapus detail SPPB
    Algoritma:
    1. mengambil data detailSPPB dengan kriteria IDDetailSPPB = id (id didapatkan dari passing values html)
    2. menghapus data detailSPPB
    '''
    datadetailsppb = models.DetailSPPB.objects.get(IDDetailSPPB=id)
    datasppb = models.SPPB.objects.get(NoSPPB=datadetailsppb.NoSPPB)
    datadetailsppb.delete()

    if datadetailsppb.DetailSPK :
        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Delete",
            pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB} Artikel : {datadetailsppb.DetailSPK.KodeArtikel.KodeArtikel} ",
        ).save()
    elif datadetailsppb.DetailSPKDisplay:
        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Delete",
            pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB}  Display : {datadetailsppb.DetailSPKDisplay.KodeDisplay.KodeDisplay} ",
        ).save()
    else:
        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Delete",
            pesan=f"Detail Surat Perintah Pengiriman Barang. Nomor SPPB : {datasppb.NoSPPB}  Display : {datadetailsppb.DetailBahan.KodeProduk} ",
        ).save()

    return redirect("detail_sppb", id=datasppb.id)


#Transaksi Produksi
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_mutasi(request):
    '''
    Fitur ini diguankan untuk melakukan manajemen data Transaksi Mutasi WIP-FG
    Algoritma:
    1. Menentukan data Tanggal awal default adalah awal bulan 
    2. Menentukan data tanggal akhir,default adalah tanggal user mengakses
    3. mengambil data transaksi produki dengan kriteria Tanggal berada pada rentang tanggal awal dan tanggal akhir
    4. mengubah format tanggal menjadi yyyy-mm-dd
    '''
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    dataproduksi = models.TransaksiProduksi.objects.filter(Jenis="Mutasi",Tanggal__range = (tanggalawal,tanggalakhir)).order_by(
        "-Tanggal", "KodeArtikel"
    )
    for i in dataproduksi:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(request, "produksi/view_mutasi.html", {"dataproduksi": dataproduksi,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_mutasidetailsppb(request):
    '''
    Fitur ini digunakan untuk melihat data detail mutasi FG - Keluar
    Algoritma :
    1. Menentukan tanggal awal, default adalah awal bulan 
    2. Menentukan tanggal akhir, default adalah tanggal user mengakses
    3. mengambil data detailSPBB dengan kriteria Tanggal berada dalam rentang tanggal awal dan tanggalakhir
    4. Mengubah format tanggal menjadi YYYY-MM-DD
    '''
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    dataproduksi = models.DetailSPPB.objects.filter(NoSPPB__Tanggal__range=(tanggalawal,tanggalakhir)).order_by(
        "-NoSPPB__Tanggal"
    )
    for i in dataproduksi:
        i.NoSPPB.Tanggal = i.NoSPPB.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_mutasisppb.html", {"dataproduksi": dataproduksi,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir}
    )

def loadversiartikel(request):
    print(request.GET)
    kodeartikel = request.GET.get('kodeartikel')
    opsiversi = models.Versi.objects.filter(KodeArtikel__KodeArtikel = kodeartikel).distinct()
    # id_spk = models.SPK.objects.get(NoSPK=no_spk)
    print(opsiversi)
    return render(request, "produksi/opsi_versitransaksiproduksi.html", {"opsiversi": opsiversi})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_mutasi(request):
    '''
    Fitur ini digunakan untuk menambahkan data mutasi WIP-FG pada produksi
    Algoritma : 
    1. Mengambil data ARtikel
    2. Mengambil data DIsplay
    3. Mengambil data SPK dengan kriteria StatusAKtif = True 
    4. Menampilkan input form transaksi mutasi 
    5. Program mendapatkan input List KodeArtikel, List Kode Display, tanggal, List Jumlah, list keterangan, list detailspk, list versiartikel
    6. Apabila mutasi artikel
        A. mengiterasi List kode artikel, list jumlah, ist keterangan, list detail spk, list versiartikel
        B. membuat object Transaksi produksi dengan KodeArtikel = kode artikel iterasi, lokasi = WIP, Tanggal = tanggal input, Jumlah = jumlah iterasi, Keterangan = keterangan iterasi, Jenis = 'Mutasi', Detailspk = detailspk iterasi, versiartikel = detailversi iterasi
        C. Menyimpan data mutasi artikel
    7. Apabila mutasi DIsplay
        A. Mengiterasi list kode display, list jumlah, list keterangan, list detailspk
        B. membuat object transaksi produksi dengan kode display = kode display iterasi, lokasi = WIP, Tanggal = tanggal input, Jumlah = jumlah iterasi, Keterangan = keterangan iterasi, Jenis ' Mutasi', Detail SPK DIsplau = detail spk iterasi
        C. menyimapn data mutasi display
    '''

    if request.method == "GET":
        data_artikel = models.Artikel.objects.all()
        data_display = models.Display.objects.all()
        data_lokasi = models.Lokasi.objects.all()
        data_spk = models.SPK.objects.filter(StatusAktif=True)

        return render(
            request,
            "produksi/add_mutasi.html",
            {
                "kode_artikel": data_artikel,
                "kode_display": data_display,
                "nama_lokasi": data_lokasi,
                "data_spk": data_spk,
            },
        )

    if request.method == "POST":
        print(request.POST)
        # print(asd)
        listkode_artikel = request.POST.getlist("kode_artikel[]")
        listkode_display = request.POST.getlist("kode_display[]")
        tanggal = request.POST["tanggal"]
        listjumlah = request.POST.getlist("jumlah[]")
        listketerangan = request.POST.getlist("keterangan[]")
        listdetail_spk = request.POST.getlist("detail_spk[]")


        if listkode_artikel:
            listversiartikel = request.POST.getlist('versiartikel')
            print(True)
            for kode, jumlah, keterangan, detail_spk, versi in zip(
                listkode_artikel, listjumlah, listketerangan, listdetail_spk,listversiartikel
            ):
                try:
                    artikelref = models.Artikel.objects.get(KodeArtikel=kode)
                except:
                    messages.error(request, "Kode Artikel tidak ditemukan")
                    continue
                
                try:
                    detailspkref = models.DetailSPK.objects.get(IDDetailSPK=detail_spk)
                except:
                    detailspkref = None
                
                try:
                    detailversi = models.Versi.objects.get(pk = versi)
                except:
                    detailversi = models.Versi.objects.filter(KodeArtikel = artikelref).order_by('Tanggal').last()
                    if detailversi == None:
                        messages.error(request,f'Artikel {artikelref.KodeArtikel} belum memiliki kode Versi')
                        continue
                    messages.error(request,f'Detail Versi {versi} tidak ditemukan pada artikel {kode}\nVersi diset pada versi Tanggal terakhir {detailversi.Versi} - {detailversi.Tanggal}')

                data_produksi = models.TransaksiProduksi(
                    KodeArtikel=artikelref,
                    Lokasi=models.Lokasi.objects.get(IDLokasi=1),
                    Tanggal=tanggal,
                    Jumlah=jumlah,
                    Keterangan=keterangan,
                    Jenis="Mutasi",
                    DetailSPK=detailspkref,
                    VersiArtikel = detailversi
                ).save()
                messages.success(request, "Data berhasil disimpan")

                if detailspkref:
                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Transaksi Mutasi. Kode Artikel : {artikelref.KodeArtikel} Jumlah : {jumlah} No SPK : {detailspkref.NoSPK} Keterangan : {keterangan}",
                    ).save()
                else:
                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Transaksi Mutasi. Kode Artikel : {artikelref.KodeArtikel} Jumlah : {jumlah} Keterangan : {keterangan}",
                    ).save()

        else:

            for kode, jumlah, keterangan, detail_spk in zip(
            listkode_display, listjumlah, listketerangan, listdetail_spk
            ):
                try:
                    displayref = models.Display.objects.get(KodeDisplay=kode)
                except:
                    messages.error(request, "Kode Artikel tidak ditemukan")
                    continue
                
                try:
                    detailspkref = models.DetailSPKDisplay.objects.get(IDDetailSPK=detail_spk)
                except:
                    detailspkref = None
                
                if detail_spk == "":
                    messages.error(request,"SPK Display Belum di set")
                    continue

                data_produksi = models.TransaksiProduksi(
                    KodeDisplay=displayref,
                    Lokasi=models.Lokasi.objects.get(IDLokasi=1),
                    Tanggal=tanggal,
                    Jumlah=jumlah,
                    Keterangan=keterangan,
                    Jenis="Mutasi",
                    DetailSPKDisplay=detailspkref,
                ).save()
                messages.success(request, "Data berhasil disimpan")

                if detailspkref:
                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Transaksi Mutasi. Kode Display : {displayref.KodeDisplay} Jumlah : {jumlah} No SPK : {detailspkref.NoSPK} Keterangan : {keterangan}",
                    ).save()
                else:
                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Transaksi Mutasi. Kode Display : {displayref.KodeDisplay} Jumlah : {jumlah} Keterangan : {keterangan}",
                    ).save()

        return redirect("view_mutasi")


# @login_required
# @logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
# def update_produksi(request, id):
    
#     produksiobj = models.TransaksiProduksi.objects.get(idTransaksiProduksi=id)
#     data_artikel = models.Artikel.objects.all()
#     data_spk = models.SPK.objects.all()

#     try:
#         data_detailspk = models.DetailSPK.objects.filter(
#             NoSPK=produksiobj.DetailSPK.NoSPK.id
#         )
#     except:
#         data_detailspk = None

#     if request.method == "GET":
#         tanggal = datetime.strftime(produksiobj.Tanggal, "%Y-%m-%d")
#         return render(
#             request,
#             "produksi/update_produksi.html",
#             {
#                 "produksi": produksiobj,
#                 "tanggal": tanggal,
#                 "kode_artikel": data_artikel,
#                 "data_spk": data_spk,
#                 "data_detailspk": data_detailspk,
#             },
#         )

#     elif request.method == "POST":
#         kode_artikel = request.POST["kode_artikel"]
#         getartikel = models.Artikel.objects.get(KodeArtikel=kode_artikel)
#         tanggal = request.POST["tanggal"]
#         jumlah = request.POST["jumlah"]
#         keterangan = request.POST["keterangan"]
#         try:
#             detail_spk = request.POST["detail_spk"]
#             detspkobj = models.DetailSPK.objects.get(IDDetailSPK=detail_spk)
#         except:
#             detspkobj = None

#         produksiobj.KodeArtikel = getartikel
#         produksiobj.Tanggal = tanggal
#         produksiobj.Jumlah = jumlah
#         produksiobj.Keterangan = keterangan
#         produksiobj.DetailSPK = detspkobj
#         produksiobj.save()
#         messages.success(request, "Data berhasil diupdate")

#         return redirect("view_produksi")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_mutasi(request, id):
    '''
    FItur ini digunakan untuk melakukan update data transaksi produksi
    Algoritma : 
    1. mendapatkan data transaksi produksi dengan kriteria idTransaksiProduksi = id (id didapatkan dari passing values HTML)
    2. Mendapatkan data Artikel
    3. Mendapatkan data SPK
    4. Menampilkan form update transaksi produksi 
    5. program mendapatkan inputan kode artikel, tanggal , jumlah, keterangan, Versi ARtikel
    6. mengupdate data transaksi produski
    '''
    produksiobj = models.TransaksiProduksi.objects.get(idTransaksiProduksi=id)
    data_artikel = models.Artikel.objects.all()
    data_display = models.Display.objects.all()
    selectedspkartikel = models.DetailSPK.objects.filter(KodeArtikel = produksiobj.KodeArtikel).distinct()
    produksiobj.spkartikel = selectedspkartikel
    selectedspkdisplay = models.DetailSPKDisplay.objects.filter(KodeDisplay = produksiobj.KodeDisplay).distinct()
    produksiobj.spkdisplay = selectedspkdisplay
    data_spk = models.SPK.objects.all()
    datadetail_spk = models.DetailSPK.objects.all()
    datadetail_spkdisplay = models.DetailSPKDisplay.objects.all()
    dataversi = models.Versi.objects.filter(KodeArtikel = produksiobj.KodeArtikel)

    if request.method == "GET":
        tanggal = datetime.strftime(produksiobj.Tanggal, "%Y-%m-%d")
        return render(
            request,
            "produksi/update_mutasi.html",
            {
                "produksi": produksiobj,
                "tanggal": tanggal,
                "kode_artikel": data_artikel,
                "kode_display": data_display,
                "data_spk": data_spk,
                "datadetail_spk" : datadetail_spk,
                "datadetail_spkdisplay" : datadetail_spkdisplay,
                'dataversi':dataversi
            },
        )

    elif request.method == "POST":
        print(request.POST)
        # print(asd)
        tanggal = request.POST["tanggal"]
        jumlah = request.POST["jumlah"]
        keterangan = request.POST["keterangan"]

        if produksiobj.KodeArtikel:
            versiartikel = request.POST['versiartikel']
            kode_artikel = request.POST["kode_artikel"]
            getartikel = models.Artikel.objects.get(KodeArtikel=kode_artikel)
            try:
                nomorspk = request.POST['nomor_spk']
                detspkobj = models.DetailSPK.objects.filter(KodeArtikel = getartikel,NoSPK__NoSPK = nomorspk ).first()
            except:
                detspkobj = None
            try:
                detailversi = models.Versi.objects.get(pk = versiartikel)
                produksiobj.VersiArtikel = detailversi
            except:
                messages.error(request,'Kode Versi Error Tidak ditemukan')
            produksiobj.KodeArtikel = getartikel
            produksiobj.DetailSPK = detspkobj
            

            if detspkobj:
                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Transaksi Mutasi. Kode Artikel : {getartikel.KodeArtikel} Jumlah : {jumlah} No SPK : {detspkobj.NoSPK} Keterangan : {keterangan}",
                ).save()
            else:
                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Transaksi Mutasi. Kode Artikel : {getartikel.KodeArtikel} Jumlah : {jumlah}  Keterangan : {keterangan}",
                ).save()

        else:
            kode_display = request.POST["kode_display"]
            getdisplay = models.Display.objects.get(KodeDisplay=kode_display)
            
            nomorspk = request.POST["nomor_spk"]
            detspkobj = models.DetailSPKDisplay.objects.filter(NoSPK__NoSPK=nomorspk,KodeDisplay = getdisplay).first()

            produksiobj.KodeDisplay = getdisplay
            produksiobj.DetailSPKDisplay = detspkobj

            if detspkobj:
                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Transaksi Mutasi. Kode Artikel : {getdisplay.KodeDisplay} Jumlah : {jumlah} No SPK : {detspkobj.NoSPK} Keterangan : {keterangan}",
                ).save()
            else:
                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Transaksi Mutasi. Kode Artikel : {getdisplay.KodeDisplay} Jumlah : {jumlah}  Keterangan : {keterangan}",
                ).save()
        
        produksiobj.Tanggal = tanggal
        produksiobj.Jumlah = jumlah
        produksiobj.Keterangan = keterangan
        produksiobj.save()

        messages.success(request, "Data berhasil diupdate")

        return redirect("view_mutasi")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_mutasi(request, id):
    '''
    Fitur ini digunakan untuk mengahapus data mutasi WIP FG
    Algoritma : 
    1. Mengambil data transaksi produksi dengan kriteria IDTransaksiProduksi = id (id didapatkan dari passing values HTML)
    2. Menghapus Data Transaksi 
    '''
    dataproduksi = models.TransaksiProduksi.objects.get(idTransaksiProduksi=id)
    dataproduksi.delete()
    messages.success(request, "Data Berhasil dihapus")

    if dataproduksi.DetailSPK:
        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Delete",
            pesan=f"Transaksi Mutasi. Kode Artikel : {dataproduksi.KodeArtikel} Jumlah : {dataproduksi.Jumlah} No SPK : {dataproduksi.DetailSPK.NoSPK} Keterangan : {dataproduksi.Keterangan}",
        ).save()
    else:
        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Delete",
            pesan=f"Transaksi Mutasi. Kode Artikel : {dataproduksi.KodeArtikel} Jumlah : {dataproduksi.Jumlah} Keterangan : {dataproduksi.Keterangan}",
        ).save()

    return redirect("view_mutasi")


# Transaksi Gudang
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_gudang(request):
    '''
    Fitur ini digunakan untuk menampilkan total transaksi mutasi kode stok dari gudang menuju produksi (WIP/FG)
    Algoritma :
    1. Mengambil data transaksi gudang dengan kriteria Jumlah > 0 (mengidikasi mutasi), dan rentang tanggal antara tangga awal dan tanggal akhir
    2. menampilkan data
    '''

    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    datagudang = models.TransaksiGudang.objects.filter(jumlah__gt=0,tanggal__range=(tanggalawal,tanggalakhir),Lokasi__NamaLokasi__in=('WIP','FG')).order_by(
        "-tanggal", "KodeProduk"
    )
    for i in datagudang:
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")

    return render(request, "produksi/view_gudang.html", {"datagudang": datagudang,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir})

@login_required  
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_gudangretur(request):
    '''
    Fitur ini digunakan untuk menampilkan total transaksi mutasi kode stok dari gudang menuju produksi (WIP/FG)
    Algoritma :
    1. Mengambil data transaksi gudang dengan kriteria Jumlah < 0 (mengidikasi retur), dan rentang tanggal antara tangga awal dan tanggal akhir
    2. menampilkan data
    '''
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    datagudang = models.TransaksiGudang.objects.filter(jumlah__lt=0,tanggal__range =(tanggalawal,tanggalakhir)).order_by(
        "-tanggal", "KodeProduk"
    )
    for data in datagudang:
        data.retur = -data.jumlah
    for i in datagudang:
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")

    return render(request, "produksi/view_gudangretur.html", {"datagudang": datagudang,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_gudang(request):
    '''
    Fitur ini digunakan untuk menambahkan data transaksi mutasi kode stok dari Gudang ke Produksi (WIP/FG)
    Algoritma
    1. Mengambil seluruh data Bahan Baku 
    2. Mengambil data Lokasi
    3. Mengambil data SPK dengan kriteria Status Aktif = True (mengidikasikan SPK Aktif saja)
    4. Program menampilkan form input transaksi Bahan Baku
    5. Program menerima input user berupa list kode stok, list lokasi, list tanggal, list jumlah, list keterangan, dan list detail SPK
    6. mengiterasi semua input dalam 1 set index
    7. Menyimpan data transaksi transaksi Bahan
    '''
    if request.method == "GET":
        data_produk = models.Produk.objects.all()
        data_lokasi = models.Lokasi.objects.all()
        data_spk = models.SPK.objects.filter(StatusAktif=True)

        listproduk = [produk.NamaProduk for produk in data_produk]

        return render(
            request,
            "produksi/add_gudang.html",
            {
                "kode_produk": data_produk,
                "nama_lokasi": data_lokasi[:2],
                "data_spk": data_spk,
                "listproduk": listproduk,
            },
        )

    if request.method == "POST":
        listkode = request.POST.getlist("kode_produk[]")
        listlokasi = request.POST.getlist("nama_lokasi[]")
        tanggal = request.POST["tanggal"]
        listjumlah = request.POST.getlist("jumlah[]")
        listketerangan = request.POST.getlist("keterangan[]")
        listdetail = request.POST.getlist("detail_spk[]")
        listtransaksiretur = request.POST.getlist('transaksiretur')
        print(request.POST)
        # print(asd)

        i = 1
        for produk, lokasi, jumlah, keterangan, detail,transaksiretur in zip(
            listkode, listlokasi, listjumlah, listketerangan, listdetail,listtransaksiretur
        ):
            nomorspk = request.POST[f"nomor_spk-{i}"]

            if nomorspk != '':
                spkobj = models.SPK.objects.get(NoSPK = nomorspk)
            try:
                produkref = models.Produk.objects.get(KodeProduk=produk)
            except:
                messages.error(request,f'Kode stok {produk} tidak ditemukan')
                continue
            lokasiref = models.Lokasi.objects.get(IDLokasi=lokasi)
            if transaksiretur == "True":
                transaksiretur = True
            else:
                transaksiretur = False

            data_gudang = models.TransaksiGudang(
                KodeProduk=produkref,
                Lokasi=lokasiref,
                tanggal=tanggal,
                jumlah=jumlah,
                keterangan=keterangan,
                KeteranganACC=False,
                KeteranganACCPurchasing = False,
                TransaksiRetur = transaksiretur
                
            )
            
            if detail != "":
                if spkobj.StatusDisplay == True:
                    try:
                        detailspkref = models.DetailSPKDisplay.objects.get(IDDetailSPK=detail)
                    except:
                        detailspkref = None
                    data_gudang.DetailSPKDisplay = detailspkref

                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Transaksi Barang Masuk. Kode Produk : {produkref.KodeProduk} Jumlah : {jumlah} No SPK Display: {detailspkref.NoSPK} Kode Display : {detailspkref.KodeDisplay} Keterangan : {keterangan}",
                    ).save()
                    
                else:
                    try:
                        detailspkref = models.DetailSPK.objects.get(IDDetailSPK=detail)
                    except:
                        detailspkref = None

                    data_gudang.DetailSPK = detailspkref

                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Create",
                        pesan=f"Transaksi Barang Masuk. Kode Produk : {produkref.KodeProduk} Jumlah : {jumlah} No SPK : {detailspkref.NoSPK} Kode Artikel : {detailspkref.KodeArtikel} Keterangan : {keterangan}",
                    ).save()
            else:
                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Transaksi Barang Masuk. Kode Produk : {produkref.KodeProduk} Jumlah : {jumlah} Keterangan : {keterangan}",
                ).save()

            i+=1
            data_gudang.save()
            messages.success(request, "Data berhasil ditambahkan")
                
        
        return redirect("view_gudang")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_gudangretur(request):
    '''
    Fitur ini digunakan untuk menambahkan data transaksi mutasi kode stok dari Gudang ke Produksi (WIP/FG)
    Algoritma
    1. Mengambil seluruh data Bahan Baku 
    2. Mengambil data Lokasi
    4. Program menampilkan form input transaksi Bahan Baku
    5. Program menerima input user berupa list kode stok, list lokasi, list tanggal, list jumlah dan list keterangan
    6. mengiterasi semua input dalam 1 set index
    7. Menyimpan data transaksi transaksi Bahan Baku Retur
    '''
    if request.method == "GET":
        data_produk = models.Produk.objects.all()
        data_lokasi = models.Lokasi.objects.all()
        data_spk = models.SPK.objects.all()

        listproduk = [produk.NamaProduk for produk in data_produk]

        return render(
            request,
            "produksi/add_gudangretur.html",
            {
                "kode_produk": data_produk,
                "nama_lokasi": data_lokasi[:2],
                "data_spk": data_spk,
                "listproduk": listproduk,
            },
        )

    if request.method == "POST":
        tanggal = request.POST["tanggal"]
        listproduk = request.POST.getlist("kode_produk[]")
        listlokasi = request.POST.getlist("nama_lokasi[]")
        listjumlah = request.POST.getlist("jumlah[]")
        listketerangan = request.POST.getlist("keterangan[]")

        for produk, lokasi, jumlah, keterangan in zip(
            listproduk, listlokasi, listjumlah, listketerangan
        ):
            try:
                produkref = models.Produk.objects.get(KodeProduk=produk)
            except:
                messages.error(request,f'Produk {produk} tidak ditemukan dalam database')
                continue
            lokasiref = models.Lokasi.objects.get(IDLokasi=lokasi)

            data_gudang = models.TransaksiGudang(
                KodeProduk=produkref,
                Lokasi=lokasiref,
                tanggal=tanggal,
                jumlah=-int(jumlah),
                keterangan=keterangan,
                KeteranganACC=False,
                KeteranganACCPurchasing = False
            ).save()
            messages.success(request, "Data berhasil ditambahkan")

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Transaksi Barang Retur. Kode Produk : {produkref.KodeProduk} Jumlah : {jumlah} Keterangan : {keterangan}",
            ).save()

        return redirect("view_gudangretur")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_gudang(request, id):
    '''
    Fitur ini digunakan untuk mengupdate data transaksi gudang
    Algoritma
    1. Mengambil data TransaksiGudang dengan kriteria IDDetailTransaksiGudang = id (id didapatkan dari passing values HTML)
    2. mengambil data kode stok
    3. Mengambil data Lokasi 
    5. Menampilkan input from update bahan baku
    6. Program mendapatkan input berupa Kode Stok, lokasi, tanggal, jumlah, dan keterangan
    7. Menyimpan perubahan 
    '''
    gudangobj = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    data_produk = models.Produk.objects.all()
    data_lokasi = models.Lokasi.objects.all()
    data_spk = models.SPK.objects.filter(StatusAktif=True)

    try:
        data_detailspk = models.DetailSPK.objects.filter(
            NoSPK=gudangobj.DetailSPK.NoSPK.id
        )
    except:
        data_detailspk = None

    try:
        data_detailspkdisplay = models.DetailSPKDisplay.objects.filter(
            NoSPK=gudangobj.DetailSPKDisplay.NoSPK.id
        )
    except:
        data_detailspkdisplay = None

    if request.method == "GET":
        tanggal = datetime.strftime(gudangobj.tanggal, "%Y-%m-%d")
        return render(
            request,
            "produksi/update_gudang.html",
            {
                "gudang": gudangobj,
                "tanggal": tanggal,
                "kode_produk": data_produk,
                "nama_lokasi": data_lokasi[:2],
                "data_spk": data_spk,
                "data_detailspk": data_detailspk,
                "data_detailspkdisplay": data_detailspkdisplay,
            },
        )

    elif request.method == "POST":
        kode_produk = request.POST["kode_produk"]
        lokasi = request.POST["nama_lokasi"]
        tanggal = request.POST["tanggal"]
        jumlah = request.POST["jumlah"]
        keterangan = request.POST["keterangan"]
        transaksiretur = request.POST['transaksiretur']

        getproduk = models.Produk.objects.get(KodeProduk=kode_produk)
        getlokasi = models.Lokasi.objects.get(IDLokasi=lokasi)
        if transaksiretur == "True":
            transaksiretur = True
        else:
            transaksiretur = False

        gudangobj.KodeProduk = getproduk
        gudangobj.Lokasi = getlokasi
        gudangobj.tanggal = tanggal
        gudangobj.jumlah = jumlah
        gudangobj.keterangan = keterangan
        gudangobj.KeteranganACCPurchasing = False
        gudangobj.KeteranganACC = False
        gudangobj.TransaksiRetur = transaksiretur
        print('tes uopdate keterangan')
        detail_spk = request.POST["detail_spk[]"]
        nomorspk = request.POST["nomor_spk"]

        try:
            detail_spk = request.POST["detail_spk[]"]
            nomorspk = request.POST["nomor_spk"]
            if detail_spk == "tot":
                gudangobj.DetailSPK = None
                gudangobj.DetailSPKDisplay = None

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Update",
                    pesan=f"Transaksi Barang Masuk. Kode Produk : {getproduk.KodeProduk} Jumlah : {jumlah} Keterangan : {keterangan}",
                ).save()

            try:
                spkobj = models.SPK.objects.get(NoSPK=nomorspk)
            except models.SPK.DoesNotExist:
                spkobj = None
                print("SPK object not found for NoSPK:", nomorspk)

            if spkobj.StatusDisplay==True:

                try:
                    detspkobj = models.DetailSPKDisplay.objects.get(IDDetailSPK=detail_spk)
                    print(detspkobj)
                    gudangobj.DetailSPKDisplay = detspkobj
                    gudangobj.DetailSPK = None

                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Update",
                        pesan=f"Transaksi Barang Masuk. Kode Produk : {getproduk.KodeProduk} Jumlah : {jumlah} No SPK Display: {detspkobj.NoSPK} Kode Display : {detspkobj.KodeDisplay} Keterangan : {keterangan}",
                    ).save()
                except models.DetailSPKDisplay.DoesNotExist:
                    gudangobj.DetailSPKDisplay = None
                    print("DetailSPKDisplay object not found for IDDetailSPKDisplay:", detail_spk)
            else:
                try:
                    detspkobj = models.DetailSPK.objects.get(IDDetailSPK=detail_spk)
                    print(detspkobj)
                    gudangobj.DetailSPK = detspkobj
                    gudangobj.DetailSPKDisplay = None

                    models.transactionlog(
                        user="Produksi",
                        waktu=datetime.now(),
                        jenis="Update",
                        pesan=f"Transaksi Barang Masuk. Kode Produk : {getproduk.KodeProduk} Jumlah : {jumlah} No SPK : {detspkobj.NoSPK} Kode Artikel : {detspkobj.KodeArtikel} Keterangan : {keterangan}",
                    ).save()

                except models.DetailSPK.DoesNotExist:
                    gudangobj.DetailSPK = None
                    print("DetailSPK object not found for IDDetailSPK:", detail_spk)

        except:
            detspkobj = None

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"Transaksi Barang Masuk. Kode Produk : {getproduk.KodeProduk} Jumlah : {jumlah} Keterangan : {keterangan}",
            ).save()

        gudangobj.save()
        messages.success(request, "Data berhasil diupdate")

        return redirect("view_gudang")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_gudangretur(request, id):
    '''
    Fitur ini digunakan untuk mengupdate data transaksi barang retur
    Algoritma
    1. Mengambil data TransaksiGudang dengan kriteria IDDetailTransaksiGudang = id (id didapatkan dari passing values HTML)
    2. mengambil data kode stok
    3. Mengambil data Lokasi 
    5. Menampilkan input from update barang retur
    6. Program mendapatkan input berupa Kode Stok, lokasi, tanggal, jumlah, dan keterangan
    7. Menyimpan perubahan 
    '''
    gudangobj = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    data_produk = models.Produk.objects.all()
    data_lokasi = models.Lokasi.objects.filter(NamaLokasi__in=("WIP","FG"))

    if request.method == "GET":
        tanggal = datetime.strftime(gudangobj.tanggal, "%Y-%m-%d")
        jumlah = -gudangobj.jumlah
        return render(
            request,
            "produksi/update_gudangretur.html",
            {
                "gudang": gudangobj,
                "tanggal": tanggal,
                "jumlah": jumlah,
                "kode_produk": data_produk,
                "nama_lokasi": data_lokasi,
            },
        )

    elif request.method == "POST":
        kode_produk = request.POST["kode_produk"]
        try:
            getproduk = models.Produk.objects.get(KodeProduk=kode_produk)
        except:
            messages.error(request,f"Data bahan baku {kode_produk} tidak ditemukan")
            return redirect("update_gudangretur",id=id)
        lokasi = request.POST["nama_lokasi"]
        getlokasi = models.Lokasi.objects.get(IDLokasi=lokasi)
        tanggal = request.POST["tanggal"]
        jumlah = request.POST["jumlah"]
        keterangan = request.POST["keterangan"]

        gudangobj.KodeProduk = getproduk
        gudangobj.Lokasi = getlokasi
        gudangobj.tanggal = tanggal
        gudangobj.jumlah = float(jumlah) * -1
        gudangobj.keterangan = keterangan

        gudangobj.save()
        messages.success(request, "Data berhasil diupdate")

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Transaksi Barang Retur. Kode Produk : {getproduk.KodeProduk} Jumlah : {jumlah} Keterangan : {keterangan}",
        ).save()

        return redirect("view_gudangretur")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_gudang(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi Gudang  pada produksi
    Algoritma 
    1. Mengambil data transaksi gudang dengan kriteria IDDetailTransaksiGudang = id (id didapatkan dari passing values HTML)
    2. Menghapus data Transaksi barang masuk produksi
    '''
    datagudang = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    datagudang.delete()
    messages.success(request, "Data Berhasil dihapus")

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Transaksi Barang Masuk. Kode Produk : {datagudang.KodeProduk} Jumlah : {datagudang.jumlah} Keterangan : {datagudang.keterangan}",
    ).save()

    return redirect("view_gudang")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_gudangretur(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi barang retur pada produksi
    Algoritma 
    1. Mengambil data transaksi gudang dengan kriteria IDDetailTransaksiGudang = id (id didapatkan dari passing values HTML)
    2. Menghapus data Transaksi barang retur
    '''
    datagudang = models.TransaksiGudang.objects.get(IDDetailTransaksiGudang=id)
    datagudang.delete()
    messages.success(request, "Data Berhasil dihapus")

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Transaksi Barang Retur. Kode Produk : {datagudang.KodeProduk} Jumlah : {-int(datagudang.jumlah)} Keterangan : {datagudang.keterangan}",
    ).save()

    return redirect("view_gudangretur")


# Rekapitulasi
def calculate_KSBB(produk,tanggal_mulai,tanggal_akhir,lokasi,kalkulator = False,startdate=None,enddate=None,requests=None,konversibaru = None):
    '''
    Perhitungan KSBB : 
    1. Jumlah Transaksi Gudang dalam rentang 1 tahun per produk 
    2. Jumlah Pemusnahan Bahan baku di produksi per bahan baku
    3. Jumlah konversi Bahan yang Bermutasi ke FG dalam bentuk kotak
    4. Jumlah Konversi Pemusnahan Kotak di area Produksi
    5. Jumlah Saldo awal Bahan baku pada WIP/FG
    6. Jumlah Transaksi mutasi kode stok masuk
    7. Jumlah Transaksi Mutasi Kode Stok keluar
    5. Total Stok = Saldo awal + Jumlah Transakasi Gudang (jumlah >= 0) +Jumlah transaksi mutasi kode stok masuk - konversi mutasi artikel WIP ke FG (berdasarkan versi) - Konversi pemusnahan artikel WIP -Pemusnahan bahan baku pada WIP atau FG
    '''
    # Menceri data transaksi gudang dengan kode 
    datagudang = models.TransaksiGudang.objects.filter(
        KodeProduk=produk, tanggal__range=(tanggal_mulai, tanggal_akhir),Lokasi__NamaLokasi=(lokasi),jumlah__gte=0,DetailSPKDisplay__isnull = True,KeteranganACC = True
    )
    dataretur = models.TransaksiGudang.objects.filter(
        KodeProduk=produk, tanggal__range=(tanggal_mulai, tanggal_akhir),Lokasi__NamaLokasi=(lokasi),jumlah__lt=0
    )
    # Kode Artikel yang di susun oleh bahan baku 
    penyusun_produk = (
        models.Penyusun.objects.filter(KodeProduk=produk,Lokasi__NamaLokasi = lokasi)
        .values_list("KodeArtikel", flat=True)
        .distinct()
    )
    penyusun_contain_produk =models.Penyusun.objects.filter(KodeProduk=produk,Lokasi__NamaLokasi = lokasi).values_list('KodeVersi__id').distinct()
    # print(penyusun_produk)
    # print(penyusun_contain_produk)
    # print(asd)
    # print('ini penyusun',penyusun_produk)
    # print(asd)

    '''BELUM DITES '''
    # Mencari data pemusnahan artikel yang disupport
    pemusnahanobj = models.PemusnahanArtikel.objects.filter(
        VersiArtikel__id__in=penyusun_contain_produk,
        Tanggal__range=(tanggal_mulai, tanggal_akhir),lokasi__NamaLokasi=lokasi
    )
    # Mencari data pemusnahan bahan baku 
    pemusnahanbahanbakuobj = models.PemusnahanBahanBaku.objects.filter(KodeBahanBaku = produk,Tanggal__range=(tanggal_mulai,tanggal_akhir),lokasi__NamaLokasi=(lokasi))

    dataproduksi = models.TransaksiProduksi.objects.filter(
        KodeArtikel__id__in=penyusun_produk,
        Jenis="Mutasi",
        Tanggal__range=(tanggal_mulai, tanggal_akhir),
    )
    dataproduksi = models.TransaksiProduksi.objects.filter(
        VersiArtikel__id__in=penyusun_contain_produk,
        Jenis="Mutasi",
        Tanggal__range=(tanggal_mulai, tanggal_akhir),
    )
    datasppb = models.DetailSPPB.objects.filter(
        NoSPPB__Tanggal__range = (tanggal_mulai,tanggal_akhir),DetailSPK__KodeArtikel__id__in=penyusun_produk
    ).exclude(DetailSPKDisplay__isnull = False)
    # print(datasppb)
    # print(dataproduksi)
    # print(asd)

    # Mencari data Bahan Display. Transaksi Produksi-->SPPBDisplay-->SPK-->FIlter TransaksiGudang by SPK
    datadisplay = datagudang.filter(DetailSPKDisplay__NoSPK__StatusDisplay =1).values_list('DetailSPKDisplay__NoSPK',flat=True).distinct()
    datadisplay2 = models.TransaksiProduksi.objects.filter(Jenis ="Mutasi",Tanggal__range = (tanggal_mulai,tanggal_akhir)).exclude(DetailSPKDisplay = None)
    datadisplaykeluar = datadisplay2.filter(DetailSPKDisplay__NoSPK__in = datadisplay)
    
    datadisplaykeluar=datadisplaykeluar.aggregate(total =Sum('Jumlah'))

    # Mencari data bahan transaksimutasi kodestok 
    datamutasikodestokkeluar = models.transaksimutasikodestok.objects.filter(Tanggal__range=(tanggal_mulai,tanggal_akhir),KodeProdukAsal = produk,Lokasi__NamaLokasi = lokasi)
    datamutasikodestokmasuk = models.transaksimutasikodestok.objects.filter(Tanggal__range=(tanggal_mulai,tanggal_akhir),KodeProdukTujuan = produk,Lokasi__NamaLokasi = lokasi)


    # print(asdas)
    listartikelmaster = []

    '''Masukan Logika : 
    Jumlah Barang yang di request pakai kode SPK diasumsikan langsung keluar walaupun belum ada Transaksi Keluar pada SPPB
    Pertimbangakan tidak menggunakan SPPB karena bisa mempengaruhi besar nilai penyesuaian apabila menunggu hasil SPPB karena penyesuaian melihat STOK KSBB REAL. Stok Display berkurang secara dinamis apabila menggunakan SPPB 
    contoh : Total request ke gudang 10 MDF pada SPK untuk produksi 10 Display. Ketika ada pengiriman 5 Unit maka pada KSBB total untuk display 10 Masuk, keluar 5 Maka masih sisa 5 Lembar. Apabila dikemudian hari ada penyesuaian maka yang dihitung adalah menggunakan penyesuaian dengan sisa 5 lembar MDF masih dalam produksi Hal ini akan mempengaruhi besar kecilnya penyesuaian dikemudian hari.
    '''

    ''' PENYESUAIAN SECTION '''
    '''
    Data Models
    Artikel :{
    tanggal1 : 0.033,
    tanggal2 : 0.0xx 
    }
    Output datamodelskonversimaster:
    {1: {datetime.date(2024, 1, 1): 0.01357, datetime.date(2024, 2, 1): 0.5}}
    '''
    datamodelskonversimaster = {}
    # for artikel in penyusun_produk:
    #     artikelmaster = models.Artikel.objects.get(id = artikel)
    #     # print(artikelmaster)
    #     konversi = models.Penyusun.objects.filter(KodeArtikel = artikel,KodeProduk = produk, Lokasi__NamaLokasi=lokasi).order_by("KodeVersi__Versi")
    #     # konversi = models.KonversiMaster.objects.filter(
    #     #     KodePenyusun__KodeArtikel=artikel, KodePenyusun__KodeProduk=produk, KodePenyusun__Lokasi__NamaLokasi = lokasi
    #     # ).order_by('KodePenyusun__versi')
    #     # print(konversi)

        
    #     tanggalversi = konversi.values_list('KodeVersi__Versi',flat=True).distinct()
    #     # print(tanggalversi)
    #     # print(asd)
    #     print(tanggalversi,artikelmaster)
    #     listkonversi = []
    #     if konversi.exists():
    #         dummy = {}
    #         for tanggal in tanggalversi:
    #             # print(tanggal,tanggalversi,konversi)
    #             datakonversi = konversi.filter(KodePenyusun__versi = tanggal)
    #             # print(datakonversi)
    #             kuantitas = datakonversi.aggregate(total = Sum('Allowance'))
    #             # print(kuantitas)
    #             listkonversi.append(kuantitas['total'])
    #             dummy[tanggal] = kuantitas['total']
    #             # print(asd)
    #         datamodelskonversimaster[artikel] = {'konversi':dummy,'penyesuaian':{}}
    #         # print(datamodelskonversimaster)
    #         # print(listkonversi,tanggalversi)
    #         # print(asd)
    #     artikelmaster.listkonversi = listkonversi
    #     artikelmaster.tanggalversi = tanggalversi

    #     # Data Penyesuaian 
    #     penyesuaianobj  = models.Penyesuaian.objects.filter( KodeProduk = produk ,KodeArtikel = artikel, TanggalMulai__range = (tanggal_mulai,tanggal_akhir)).order_by('TanggalMulai')
    #     # print(penyesuaianobj)
    #     # print(asd)
    #     if penyesuaianobj.exists:
    #         dummy ={}
    #         for penyesuaian in penyesuaianobj:
    #             dummy[penyesuaian.TanggalMulai] = penyesuaian.konversi
    #         datamodelskonversimaster[artikel]['penyesuaian'] = dummy
    #     penyesuaiandataperartikel = [i.konversi for i in penyesuaianobj]
    #     tanggalpenyesuaianperartikel = [i.TanggalMulai for i in penyesuaianobj]
    #     tanggalakhirpenyesuaian = [i.TanggalMinus for i in penyesuaianobj]

    #     artikelmaster.listpenyesuaian = penyesuaiandataperartikel
    #     artikelmaster.tanggalpenyesuaian =tanggalpenyesuaianperartikel
    #     artikelmaster.tanggalakhirpenyesuaian = tanggalakhirpenyesuaian
    #     listartikelmaster.append(artikelmaster)
        # print(penyesuaiandataperartikel, tanggalpenyesuaianperartikel,tanggalakhirpenyesuaian)
        # print(asdas)
        # print(asdas)
    # print(listartikelmaster)
    # print(asd)
    ''' TANGGAL SECTION '''
    tanggalmasuk = datagudang.values_list("tanggal", flat=True)
    tanggalretur = dataretur.values_list('tanggal',flat=True)
    tanggalkeluar = dataproduksi.values_list("Tanggal", flat=True)
    tanggalpemusnahan = pemusnahanobj.values_list("Tanggal", flat=True)
    tanggalpemusnahanbahanbaku = pemusnahanbahanbakuobj.values_list('Tanggal',flat=True)
    tanggalpengirimanbarang = datasppb.values_list('NoSPPB__Tanggal',flat=True)
    tanggalmutasikodestokkeluar = datamutasikodestokkeluar.values_list('Tanggal',flat=True)
    tanggalmutasikodestokmasuk = datamutasikodestokmasuk.values_list('Tanggal',flat=True)
    # Belum Mempertimbangkan SPK Display

    '''BELUM MEMPERTIMBANGKAN KELUAR DISPLAY'''
    if lokasi == 'WIP':
        listtanggal = sorted(
        list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalpemusnahan).union(tanggalpemusnahanbahanbaku).union(tanggalretur).union(tanggalmutasikodestokmasuk).union(tanggalmutasikodestokkeluar)))
    )
    else:
            listtanggal = sorted(
        list(set(tanggalmasuk.union(tanggalpemusnahan).union(tanggalpemusnahanbahanbaku).union(tanggalretur).union(tanggalpengirimanbarang).union(tanggalmutasikodestokmasuk).union(tanggalmutasikodestokkeluar)))
    )


    ''' SALDO AWAL SECTION '''
    try:
        saldoawal = models.SaldoAwalBahanBaku.objects.get(
            IDBahanBaku=produk,
            IDLokasi__NamaLokasi=lokasi,
            Tanggal__range=(tanggal_mulai, tanggal_akhir),
        )
        saldo = saldoawal.Jumlah
        saldoawal.Tanggal = saldoawal.Tanggal.strftime("%Y-%m-%d")

    except models.SaldoAwalBahanBaku.DoesNotExist:
        saldo = 0
        saldoawal = None

    sisa = saldo

    ''' PENGOLAHAN DATA '''
    listdata =[]
    # print(listtanggal)
    # print(asd)
    for i in listtanggal:
        # Data Models

        datamodelsartikel = []
        datamodelsperkotak = []
        datamodelskonversi = []
        datamodelskeluar = []
        datamodelssisa = []
        datadetailmasuk = []
        data = {
            'Tanggal': None,
            'Artikel': datamodelsartikel,
            'Perkotak' : datamodelsperkotak,
            "Konversi" : datamodelskonversi,
            'Masuk' : None,
            'detailmasuk' :datadetailmasuk,
            'Keluar' : datamodelskeluar,
            'Sisa' : datamodelssisa
            
        }
        data['Tanggal'] = i.strftime("%Y-%m-%d")
        # Data Masuk
        masuk = 0
        masukdisplay = 0
        datamasuk = datagudang.filter(tanggal=i)
        datakeluarretur = dataretur.filter(tanggal=i)
        datakeluar = dataproduksi.filter(Tanggal = i)
        datapemusnahan = pemusnahanobj.filter(Tanggal = i)
        datapemusnahanbahanbaku = pemusnahanbahanbakuobj.filter(Tanggal = i)
        datapengiriman = datasppb.filter(NoSPPB__Tanggal = i)
        datamutasikodestokkeluarfiltered = datamutasikodestokkeluar.filter(Tanggal=i)
        datamutasikodestokmasukfiltered = datamutasikodestokmasuk.filter(Tanggal=i)

        for k in datamasuk:
            masuk += k.jumlah
            # print(k.DetailSPKDisplay != None)
            # if k.DetailSPKDisplay != None and k.DetailSPKDisplay.NoSPK.StatusDisplay == True:
            #     masukdisplay +=k.jumlah
            
            #     print(k.DetailSPKDisplay)
            #     print(i)
            #     print(asd)
            # else:
            datadetailmasuk.append(k)

        for k in datamutasikodestokmasukfiltered:
            masuk += k.Jumlah
            # print(datamutasikodestokmasukfiltered)
            k.jumlah = k.Jumlah
            
            # print(k.DetailSPKDisplay != None)
            datadetailmasuk.append(k)

        sisa  += masuk
        
        # Data Keluar
        data['Masuk'] = masuk

        # artikelkeluar = datakeluar.values_list('KodeArtikel',flat=True).distinct()
        # artikelpemusnahan = datapemusnahan.values_list('KodeArtikel',flat=True).distinct()
        # artikelkirim = datapengiriman.values_list('DetailSPK__KodeArtikel',flat=True).distinct()
        # if lokasi == 'FG':
            # artikelkeluar = artikelkirim
            # print(artikelkeluar)
            # print(asd)

        if masukdisplay != 0:
            datamodelskeluar.append(masukdisplay)
            sisa -=masukdisplay
        '''
        NEW ARTIKEL KELUAR ALGORITHM
        1. Ambil data Transaksi Keluar perhari ini (mutasi )
        2. Cek menggunakan versi berapa pada transaksi produksi
        3. Hitung penggunaan Bahan Baku

        '''
        '''KONDISI CEK UNTUK PENYUSUN DENGAN VERSI PERLU DITINJAU ULANG'''
        if lokasi == 'WIP':
            for j in datakeluar:
                print(j)
                # print(datakeluar)
                penyesuaianobj = models.Penyesuaian.objects.filter( KodeProduk = produk ,KodeArtikel = j.KodeArtikel, TanggalMulai__range = (tanggal_mulai,tanggal_akhir)).order_by('TanggalMulai')
                penyusunartikelkeluar = models.Penyusun.objects.filter(KodeArtikel = j.KodeArtikel, KodeProduk = produk, KodeVersi = j.VersiArtikel).aggregate(total = Sum('Allowance'))['total']
                if penyusunartikelkeluar == None:
                    penyusunartikelkeluar = 0
                    continue
                totalartikelmutasi = j.Jumlah
                konversi = penyusunartikelkeluar
                penyesuaianobj = models.Penyesuaian.objects.filter(KodeProduk = produk,KodeArtikel = j.KodeArtikel, TanggalMulai__lte = i, TanggalMinus__gte = i).last()
                # print(penyesuaianobj)
                if penyesuaianobj:
                    # print(penyesuaianobj)
                    konversi = penyesuaianobj.konversi
                    # print('Masuk penyesuaian')
                    # print(penyesuaianobj)
                    # print(asd)
                if startdate != None and enddate !=None:
                    if startdate.date()<=i<=enddate.date():
                        if kalkulator and konversibaru :
                            konversi = konversibaru[j.KodeArtikel.KodeArtikel]['konversiakhir']
                            print(konversi)
                            print(konversibaru)
                # print(konversi)
                # print(produk,i,konversi,j.KodeArtikel)
                # print(asd)
                totalpenggunaanbahanbaku = totalartikelmutasi * konversi
                # print(produk,totalartikelmutasi,totalpenggunaanbahanbaku,j.KodeArtikel)
                datamodelskonversi.append(konversi)
                datamodelskeluar.append(totalpenggunaanbahanbaku)
                datamodelsartikel.append(j.KodeArtikel.KodeArtikel)
                datamodelsperkotak.append(j.Jumlah)
                sisa -= totalpenggunaanbahanbaku
                datamodelssisa.append(sisa)
                if sisa < 0  and requests != None:
                    messages.warning(requests,f'Sisa stok menjadi negatif pada tanggal {i}')
                print(totalpenggunaanbahanbaku)
                print(totalartikelmutasi)
                print(konversi)
                print(j)
                # if konversibaru:
                    # print(asd)

                    
        else:
            print(i)
            for j in datapengiriman:
                penyusunartikelkeluar = models.Penyusun.objects.filter(KodeArtikel = j.DetailSPK.KodeArtikel, KodeProduk = produk, KodeVersi = j.VersiArtikel,Lokasi__NamaLokasi = lokasi).aggregate(total = Sum('Allowance'))['total']
                print(penyusunartikelkeluar)
                # print(asd)
                if penyusunartikelkeluar == None:
                    penyusunartikelkeluar = 0
                    continue
                if penyusunartikelkeluar == None:
                    penyusunartikelkeluar = 0
                    continue
                totalartikelmutasi = j.Jumlah
                konversi = penyusunartikelkeluar
                penyesuaianobj = models.Penyesuaian.objects.filter(KodeProduk = produk,KodeArtikel = j.DetailSPK.KodeArtikel, TanggalMulai__lte = i, TanggalMinus__gte = i).last()
                # print(penyesuaianobj)
                if penyesuaianobj:
                    # print(penyesuaianobj)
                    konversi = penyesuaianobj.konversi
                    # print('Masuk penyesuaian')
                    # print(penyesuaianobj)
                    # print(asd)
                if startdate != None and enddate !=None:
                    if startdate.date()<=i<=enddate.date():
                        if kalkulator and konversibaru:
                            konversi = konversibaru[j.KodeArtikel.KodeArtikel]['konversiakhir']
                # print(konversi)
                # print(produk,i,konversi,j.KodeArtikel)
                # print(asd)
                totalpenggunaanbahanbaku = totalartikelmutasi * konversi
                # print(produk,totalartikelmutasi,totalpenggunaanbahanbaku,j.KodeArtikel)
                datamodelskonversi.append(konversi)
                datamodelskeluar.append(totalpenggunaanbahanbaku)
                datamodelsartikel.append(j.DetailSPK.KodeArtikel)
                datamodelsperkotak.append(j.Jumlah)
                sisa -= totalpenggunaanbahanbaku
                datamodelssisa.append(sisa)

        
        for j in datapemusnahan:
            penyesuaianobj = models.Penyesuaian.objects.filter( KodeProduk = produk ,KodeArtikel = j.KodeArtikel, TanggalMulai__range = (tanggal_mulai,tanggal_akhir)).order_by('TanggalMulai')
            penyusunartikelkeluar = models.Penyusun.objects.filter(KodeArtikel = j.KodeArtikel, KodeProduk = produk, KodeVersi = j.VersiArtikel).aggregate(total = Sum('Allowance'))['total']
            if penyusunartikelkeluar == None:
                penyusunartikelkeluar = 0
            totalartikelpemusnahan = j.Jumlah
            konversi = penyusunartikelkeluar
            penyesuaianobj = models.Penyesuaian.objects.filter(KodeProduk = produk,KodeArtikel = j.KodeArtikel, TanggalMulai__lte = i, TanggalMinus__gte = i).last()
            # print(penyesuaianobj)
            if penyesuaianobj:
                # print(penyesuaianobj)
                konversi = penyesuaianobj.konversi
            if startdate != None and enddate !=None:
                if startdate.date()<=i<=enddate.date():
                    if kalkulator and konversibaru:
                        konversi = konversibaru[j.KodeArtikel.KodeArtikel]['konversiakhir']
            totalpenggunaanbahanbaku = totalartikelpemusnahan * konversi
            datamodelskonversi.append(konversi)
            datamodelskeluar.append(totalpenggunaanbahanbaku)
            datamodelsartikel.append(j.KodeArtikel.KodeArtikel)
            datamodelsperkotak.append(j.Jumlah)
            sisa -= totalpenggunaanbahanbaku
            datamodelssisa.append(sisa)
            # print(asd)
        # for j in artikelkeluar:
        #     artikelkeluarobj = models.Artikel.objects.get(id = j)
        #     penyesuaianobj = models.Penyesuaian.objects.filter( KodeProduk = produk ,KodeArtikel = artikelkeluarobj, TanggalMulai__range = (tanggal_mulai,tanggal_akhir)).order_by('TanggalMulai')
        #     if lokasi == 'FG':
        #         total = datapengiriman.filter(DetailSPK__KodeArtikel__id = j).aggregate(total = Sum('Jumlah'))
        #     else:
        #         total = datakeluar.filter(KodeArtikel__id = j).aggregate(total = Sum('Jumlah'))
        #     # print(listartikelmaster)
        #     print(artikelkeluarobj,i,total)
        #     print(j)
        #     # print(asd)
        #     indexartikel = listartikelmaster.index(artikelkeluarobj)
        #     filtered_data = [d for d in listartikelmaster[indexartikel].tanggalversi if d <= i]
        #     # print(indexartikel, listartikelmaster,listartikelmaster[indexartikel].tanggalversi)
        #     # print('ini i ',i, filtered_data,'\n\n\n\n\n\n')
        #     filtered_data.sort(reverse=True)
        #     # print(filtered_data)
        #     # print(asd)

        #     if not filtered_data:
        #         filtered_data = 0


        #     if filtered_data != 0:
        #         tanggalversiterdekat = max(filtered_data)
        #         indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalversi).index(tanggalversiterdekat)
        #         konversiterdekat = listartikelmaster[indexartikel].listkonversi[indextanggalterdekat]
        #         # print(i,filtered_data)
        #         # print(asd)
        #     else:
        #         tanggalversiterdekat = i
        #         konversiterdekat = 0

        #     if listartikelmaster[indexartikel].tanggalpenyesuaian  and filtered_data != 0:
        #         # print('masuk1')
        #         # print(asd)
                
        #         cektanggal  = penyesuaianobj.filter(TanggalMulai__lte = i, TanggalMinus__gte = i).last()
        #         # print(cektanggal)
        #         # print(penyesuaianobj,i,j)
        #         # print(asd)
        #         if cektanggal:
        #             # print('Ada data')
        #             filtered_data.sort(reverse=True)
        #             tanggalversiterdekat = cektanggal.TanggalMulai
        #             # print(tanggalversiterdekat)
        #             indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalpenyesuaian).index(tanggalversiterdekat)
        #             konversiterdekat = listartikelmaster[indexartikel].listpenyesuaian[indextanggalterdekat]
        #             # print(i,cektanggal,konversiterdekat)
        #             # print(asd)
        #         else :
        #             filtered_data.sort(reverse=True)
        #             tanggalversiterdekat = max(filtered_data)
        #             # print(tanggalversiterdekat)
        #             indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalversi).index(tanggalversiterdekat)
        #             konversiterdekat = listartikelmaster[indexartikel].listkonversi[indextanggalterdekat]
        #     # elif listartikelmaster[indexartikel].tanggalpenyesuaian:
        #     #     cektanggal  = penyesuaianobj.filter(TanggalMulai__lte = i, TanggalMinus__gte = i).last()
        #     #     print(cektanggal)
        #     #     print(penyesuaianobj,i)
        #     #     print(asd)

        #     konversiterdekat= (konversiterdekat)
        #     datamodelskonversi.append(konversiterdekat)
        #     # print(konversiterdekat)
        #     # print(total)
        #     # print(listartikelmaster)
        #     # print(asd)

        #     datamodelskeluar.append(konversiterdekat*total['total'])
        #     datamodelsartikel.append(artikelkeluarobj)
        #     datamodelsperkotak.append(total['total'])
        #     sisa -= konversiterdekat*total['total']
        #     sisa = (sisa)
        #     datamodelssisa.append(sisa)

        # for j in artikelpemusnahan:
        #     print(datapemusnahan)
            
        #     artikelkeluarobj = models.Artikel.objects.get(id = j)
        #     total = datapemusnahan.filter(KodeArtikel__id = j).aggregate(total=Sum('Jumlah'))
        #     indexartikel = listartikelmaster.index(artikelkeluarobj)
        #     filtered_data = [d for d in listartikelmaster[indexartikel].tanggalversi if d <= i]
            
        #     filtered_data.sort(reverse=True)
        #     # print(filtered_data,i)
        #     # print(asd)
            
        #     if not filtered_data:
        #         filtered_data = 0

        #     if filtered_data !=  0:
        #         tanggalversiterdekat = max(filtered_data)
        #         indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalversi).index(tanggalversiterdekat)
        #         konversiterdekat = listartikelmaster[indexartikel].listkonversi[indextanggalterdekat]
        #     else:
        #         tanggalversiterdekat = i
        #         konversiterdekat = 0
                
        #     if listartikelmaster[indexartikel].tanggalpenyesuaian and filtered_data != 0 :
        #         # filtered_data = [d for d in listartikelmaster[indexartikel].tanggalpenyesuaian if d <= i]
        #         cektanggal  = penyesuaianobj.filter(TanggalMulai__lte = i, TanggalMinus__gte = i).order_by('TanggalMulai').first()
        #         # print(i,cektanggal)
        #         # print(asd)
        #         if cektanggal:
        #             # print('Ada data')
        #             tanggalversiterdekat = cektanggal.TanggalMulai
        #             # print(tanggalversiterdekat)
        #             indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalpenyesuaian).index(tanggalversiterdekat)
        #             konversiterdekat = listartikelmaster[indexartikel].listpenyesuaian[indextanggalterdekat]
        #             # print(i,cektanggal,konversiterdekat)
        #             # print(asd)

        #         else :
        #             filtered_data.sort(reverse=True)
        #             tanggalversiterdekat = max(filtered_data)
        #             # print(tanggalversiterdekat)
        #             indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalversi).index(tanggalversiterdekat)
        #             konversiterdekat = listartikelmaster[indexartikel].listkonversi[indextanggalterdekat]

               
        #     datamodelskonversi.append(konversiterdekat)
        #     datamodelskeluar.append(konversiterdekat*total['total'])
        #     datamodelsartikel.append(artikelkeluarobj)
        #     datamodelsperkotak.append(total['total'])
        #     sisa -= konversiterdekat*total['total']
        #     sisa = (sisa)
        #     datamodelssisa.append(sisa)

        # Pemusnahan Bahan Baku
        if datapemusnahanbahanbaku.exists():
            # Mengagregat Jumlah Bahan Baku rusak
            totalbahanbakurusak = datapemusnahanbahanbaku.aggregate(total=Sum('Jumlah'))

            sisa -= totalbahanbakurusak['total']
            datamodelssisa.append(sisa)
            datamodelskeluar.append(totalbahanbakurusak['total'])
        # Mutasi Kode Stok Keluar
        if datamutasikodestokkeluarfiltered.exists():
            # Mengagregat Jumlah Bahan Baku rusak
            totalkodestokkeluar = datamutasikodestokkeluarfiltered.aggregate(total=Sum('Jumlah'))

            sisa -= totalkodestokkeluar['total']
            datamodelssisa.append(sisa)
            datamodelskeluar.append(totalkodestokkeluar['total'])
        # Barang retur
        if datakeluarretur.exists():
            # Mengagregat Jumlah Bahan Baku rusak
            totalbahanbakuretur = datakeluarretur.aggregate(total=Sum('jumlah'))

            sisa -= totalbahanbakuretur['total']*-1
            datamodelssisa.append(sisa)
            datamodelskeluar.append(totalbahanbakuretur['total']*-1)
            # print(asdasd)

        if not datamodelssisa :
            sisa = (sisa)
            datamodelssisa.append(sisa)

        data['Sisa'] = datamodelssisa
        listdata.append(data)
    # print(listdata)
    # print(asd)
        # Indikasi salah di bagian konversi penyusun yang mencari nilai atasnya. Kalau gaada bisa di set 0 

    return listdata,saldoawal

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_ksbb3(request):
    '''
    Fitur ini digunakan untuk menampilkan halaman KSBB bahan Baku Produksi
    Algoritma
    1. Mendapatkan data Produk 
    2. Menampilkan form input pada user
    3. Program mendapatkan input berupa Kodeproduk dan periode
    4. Memanggil fungsi CalculateKSBB
    5. Menampilkan data KSBB produksi
    '''
    kodeproduk = models.Produk.objects.all()
    sekarang = datetime.now().year
    
    if len(request.GET) == 0:
        return render(request, "produksi/view_ksbb.html", {"kodeprodukobj": kodeproduk, "sekarang": sekarang})
    else:
        try:
            produk = models.Produk.objects.get(KodeProduk=request.GET["kodebarang"])
            nama = produk.NamaProduk
            satuan = produk.unit
        except models.Produk.DoesNotExist:
            messages.error(request, "Data Produk tidak ditemukan")
            return redirect("view_ksbb")
        
        if "periode" in request.GET and request.GET["periode"]:
            tahun = int(request.GET["periode"])
        else:
            tahun = sekarang
        
        
        lokasi = request.GET['lokasi']
        print(lokasi)

        tanggal_mulai = datetime(year=tahun, month=1, day=1)
        tanggal_akhir = datetime(year=tahun, month=12, day=31)

        listdata, saldoawal = calculate_KSBB(produk, tanggal_mulai, tanggal_akhir,lokasi,requests=request)


        # print(asd)
        if saldoawal != None:
            saldoawal.Tanggal = datetime.strptime(saldoawal.Tanggal,"%Y-%m-%d")

        return render(request, "produksi/view_ksbb.html", {
            'data': listdata,
            'saldo': saldoawal,
            'kodebarang': request.GET["kodebarang"],
            "nama": nama,
            "satuan": satuan,
            'kodeprodukobj': kodeproduk,
            'tahun': tahun,
            'lokasi' : lokasi
        })


@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def detailksbb(request, id, tanggal,lokasi):
    '''
    Fitur ini digunakan untuk melihat detail tanggal KSBB
    Algoritma
    1. Mengambil data transaksi Gudang dengan kriteria Tanggal = tanggal (tanggal didapatkan dari passing values HTML), KodeProduk = id (id didapatkan dari passing values HTML), dan Jumlah > 0 (mengindikasikan permintaan barang)
    2. Mengambil data transaksi Gudang dengan kriteria Tanggal = tanggal (tanggal didapatkan dari passing values HTML), KodeProduk = id (id didapatkan dari passing values HTML), dan Jumlah < 0 (mengindikasikan retur barang)
    3. Mengiterasi data list barang retur 
    4. Mengalikan jumlah retur dengan -1 (untuk mendapatkan hasil Positif)
    5. Mendapatkan data list versi dari tabel penysuun dengan kriteria Kode Produk = id (id didapatkan dari passing values)
    6. Mendapatkan data list artikel dari tabel penyusun dengan kriteria Kode Produk = id (id didapatkan dari passing values)
    7. Mencari data Transaksi Produksi dengan kriteria VersiArtikel pada listversi, Jenis="Mutasi", dan Tanggal = tanggal(tanggal didapatkan dari passing values HTML)
    8. Mencari data pemusnahan bahan baku dengan kriteria Tanggal = tanggal (tanggal didapatkan dari passing values HTML), KodeProduk = id (id didapatkan dari passing values HTML), NamaLokasi = Lokasi (Lokkasi didapatkan dari passing values HTMl)
    9. mencari data pemusnahan artikel dengan kriteria Kode artikel berada dalam list kode artikel, Tanggal = tanggal (tanggal didapatkan dari passing values HTML)
    10. Mencari data transaksi mutasi kode stok keluar pada tabel transaksi mutasi kode stok dengan kriteria Tanggal = tanggal (tanggal didapatkan dari passing values HTML), KodeProdukAsal = id (id didapatkan dari passing values HTML), NamaLokasi = lokasi (lokasi didapatkan dari passing values)
    10. Mencari data transaksi mutasi kode stok Masuk pada tabel transaksi mutasi kode stok dengan kriteria Tanggal = tanggal (tanggal didapatkan dari passing values HTML), KodeProdukTujuan = id (id didapatkan dari passing values HTML), NamaLokasi = lokasi (lokasi didapatkan dari passing values)
    '''
    tanggal = datetime.strptime(tanggal, "%Y-%m-%d")
    tanggal = tanggal.strftime("%Y-%m-%d")
    print('ini lokasi',lokasi)
    print(id)
    

    # Transaksi Gudang
    datagudang = models.TransaksiGudang.objects.filter(tanggal=tanggal, KodeProduk__KodeProduk=id,Lokasi__NamaLokasi=(lokasi),jumlah__gte=0)
    dataretur = models.TransaksiGudang.objects.filter(tanggal=tanggal, KodeProduk__KodeProduk=id,Lokasi__NamaLokasi=(lokasi),jumlah__lt=0)
    penyusun_produk = (
        models.Penyusun.objects.filter(KodeProduk__KodeProduk=id,Lokasi__NamaLokasi = lokasi)
        .values_list("KodeArtikel", flat=True)
        .distinct()
    )
    datasppb = models.DetailSPPB.objects.filter(
        NoSPPB__Tanggal=tanggal,DetailSPK__KodeArtikel__id__in=penyusun_produk).exclude(DetailSPKDisplay__isnull = False)
    print(datasppb)
    # print(asd)
    for item in dataretur:
        item.jumlah = item.jumlah * -1
    listartikel = (
        models.Penyusun.objects.filter(KodeProduk__KodeProduk=id)
        .values_list("KodeArtikel__KodeArtikel", flat=True)
        .distinct()
    )
    listversi = models.Penyusun.objects.filter(KodeProduk__KodeProduk=id).values_list("KodeVersi", flat=True).distinct()
    print(listversi)
    # print(asd)
    # Transaksi Produksi
    dataproduksi = models.TransaksiProduksi.objects.filter(
         VersiArtikel__in=listversi, Jenis="Mutasi", Tanggal=tanggal
    )
    print(dataproduksi)
    # print(asd)

    # Transaksi Pemusnahan
    datapemusnahan = models.PemusnahanArtikel.objects.filter(
        KodeArtikel__KodeArtikel__in=listartikel, Tanggal=tanggal,
    )
    # Transaksi Pemusnahan Bahan Baku
    datapemusnahanbahanbaku  =models.PemusnahanBahanBaku.objects.filter(Tanggal = tanggal,KodeBahanBaku__KodeProduk = id,lokasi__NamaLokasi=lokasi)
    # Transaksi Mutasi Kode Stok
    datamutasikodestokkeluar = models.transaksimutasikodestok.objects.filter(Tanggal=tanggal,KodeProdukAsal__KodeProduk = id,Lokasi__NamaLokasi = lokasi)
    datamutasikodestokmasuk = models.transaksimutasikodestok.objects.filter(Tanggal=tanggal,KodeProdukTujuan__KodeProduk = id,Lokasi__NamaLokasi = lokasi)

    print(datapemusnahanbahanbaku)
    print(datagudang)
    return render(
        request,
        "produksi/view_detailksbb.html",
        {
            "datagudang": datagudang,
            "dataproduksi": dataproduksi,
            "datapemusnahan": datapemusnahan,
            'datapemusnahanbahanbaku' : datapemusnahanbahanbaku,
            "dataretur" : dataretur,
            'datamutasikodestokkeluar':datamutasikodestokkeluar,
            'datamutasikodestokmasuk': datamutasikodestokmasuk,
            'lokasi' : lokasi,
            'datasppb' : datasppb

        },
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_ksbj2(request):
    '''
    Fitur ini digunakan untuk menampilkan KSBJ Produksi 
    Algoritma : 
    1. Mendapatkan data artikel
    2. Menampilkan input artikel dan periode KSBJ
    3. Memanggil fungsi calculate_Ksbj
    '''
    dataartikel = models.Artikel.objects.all()
    if len(request.GET) == 0:
        return render(request,'produksi/view_ksbj.html', {"dataartikel": dataartikel})
    else :
        kodeartikel =  request.GET['kodeartikel']
        try:
            Artikelobj = models.Artikel.objects.get(KodeArtikel = kodeartikel)
        except:
            messages.error(request,'Data Artikel tidak ditemukan')
            return redirect('view_ksbj')
        
        if request.GET['tahun']:
            tahun = int(request.GET['tahun'])
        else:
            sekarang = datetime.now()
            tahun = sekarang.year
        
        tanggal_mulai = datetime(year=tahun, month=1, day=1)
        tanggal_akhir = datetime(year=tahun, month=12, day=31)

        lokasi = request.GET['lokasi']
        lokasiobj = models.Lokasi.objects.get(NamaLokasi = lokasi)
        try:
            listdata,saldoawal = calculate_ksbj(Artikelobj,lokasi,tanggal_mulai.year,requests=request)
        except Exception as e:
            print(e)
            messages.error(request,'Bahan Baku Utama belum diset')
            return redirect('view_ksbj')



        print(listdata)

        return render(
            request,
            "produksi/view_ksbj.html",
            {
                "kodeartikel": kodeartikel,
                "dataartikel": dataartikel,
                "lokasi": lokasi,
                "listdata": listdata,
                "saldoawal": saldoawal,
                "tahun": tahun,
                'artikelobj':Artikelobj
            },
        )
        
def calculate_ksbj(artikel,lokasi,tahun,requests=None):
    '''
    Fitur ini digunakan untuk menghitung besar KSBJ tiap artikel,tahun dan lokasi
    Parameter fungsi ini ada 3 yaitu Artikel (Artikejkobj, Lokasi dalam string, dan tahun
    Algortima
    1. Mencari bahan baku penyusun utama .
        A. Apabila data data bahan baku penyusun utama maka akan muncul tulisan error
        B. Apabila ada maka lanjut ke nomor 2
    2. Mengambil data Transaksi Produksi dengan kriteria KodeArtikel = artiktel (artikel didapatkan dari fungsi lainnya) dan Jenis = 'Mutasi'
    3. Mengambil data Transaksi Gudang dengan kriteria DetailSPK__KodeArtikel = artikel, KodeProduk = bahan baku utama
    4. Mendapatkan data List tanggal dengan berisi list tanggal masuk transaksi gudang
    5. Apabila Lokasi WIP :
        A. Mencari data pemusnahan artikel dengan kriteria Kode Artikel = artikel, Tanggal berada dalam rentang tanggal mulai dan tanggal akhir (didapatkan dari fungsi lain) dan Lokasi = 'WIP'
        B. Membuat data tanggal pemusnahan artikel
        C. Mencari data Saldo awal artikel dengan kriteria KodeArtikel = artikel, Lokasi = WIP, dan Tanggal berada dalam rentang tanggal mulai dan tanggal akhir
        D. Melakukan filter data pada Transaksi Produksi dengan tambahan filter lokasi = 'WIP'
        E. Membuat tanggal list yang terdiri dari tanggal Transaksi Produksi, Tanggal Transaksi Gudang, dan tanggal pemusnahan
        F. melooping semua tanggal 
        G. Rumus Sisa Stok Artikel = Saldo awal + Jumlah Transaksi Gudang * Konversi Penyusun - Jumlah Mutasi WIPFG - Jumlah Pemusnahan
        H. Menambahkan pada list data
    7. Apabila lokasi FG
        A. Mencari data tanggal Transaksi Produksi dengan lokasi WIP
        B. Mencari Saldo Awal Artikel dengan kriteria Kode Artikel = artikel, Lokasi = FG, Tanggal berada dalam rentang tanggal mulai dan tanggal akhir
        C. Mencari list tanggal mutasi dari tabel transaksi produksi dengan kriteria Tanggal berada dalam rentang tanggal mulai dan tanggal akhir, lokasi = FG dan Jneis = "Mutas"
        D. Mencari data Pengiriman dari tabel detailSPPB dengan kriteria KodeArtikel = artikel,Tanggal berada dalam rentang tanggal mulai dan akhir,
        E. Mencari list tanggal pengiriman detailSPPB
        F. Mencarai data Pemusnahan Artikel dengan kriteria kodeArtikel = artikel, Lokasi = "FG", Tanggal berada dalam rentang tanggal mulai dan akhir
        G. Mencari list tanggal pemusnahan
        F. Menggabungkan list tanggal yang terdiri dari tanggal mutasi, tanggal sppb, dan tanggal pemusnahan
        G. Melooping semua tanggal
        H. Rumus Sisa stok artikel = Saldo awal + Jumlah Mutasi WIPFG - Pemusnahan Artikel FG - Pengiriman SPPB
        I. Menambahkan pada list data

    
    '''
    tanggal_mulai = datetime(year=tahun, month=1, day=1)
    tanggal_akhir = datetime(year=tahun, month=12, day=31)
    getbahanbakuutama = models.Penyusun.objects.filter(KodeArtikel=artikel.id, Status=1, KodeVersi__isdefault = True).order_by('-KodeVersi__Tanggal').first()
    print(getbahanbakuutama)
    # print(asd)
    if not getbahanbakuutama :
        getbahanbakuutama = models.Penyusun.objects.filter(KodeArtikel=artikel.id, Status=1).order_by('-KodeVersi__Tanggal').first()
        if not getbahanbakuutama :

            raise ValueError('Belum di set')
    
    data = models.TransaksiProduksi.objects.filter(KodeArtikel=artikel.id,Jenis = "Mutasi")
    datamasuk = models.TransaksiGudang.objects.filter(DetailSPK__KodeArtikel = artikel.id,KodeProduk = getbahanbakuutama.KodeProduk,tanggal__range=(tanggal_mulai, tanggal_akhir))
    listtanggalmasuk = datamasuk.values_list('tanggal',flat=True).distinct()
    # print(datamasuk)
    listdata = []
    if lokasi == "WIP":
        datapemusnahan = models.PemusnahanArtikel.objects.filter(KodeArtikel = artikel,Tanggal__range=(tanggal_mulai,tanggal_akhir),lokasi__NamaLokasi = "WIP")
        listtanggalpemusnahan = datapemusnahan.values_list('Tanggal',flat=True).distinct()
        data = data.filter(Lokasi__pk=1)
        try:
            saldoawalobj = models.SaldoAwalArtikel.objects.get(IDArtikel__KodeArtikel=artikel, IDLokasi__NamaLokasi=lokasi,Tanggal__range =(tanggal_mulai,tanggal_akhir))
            saldo = saldoawalobj.Jumlah
            saldoawalobj.Tanggal = saldoawalobj.Tanggal.year

        except models.SaldoAwalArtikel.DoesNotExist :
            saldo = 0
            saldoawal = None
            saldoawalobj ={'Tanggal' : 'Belum ada Data','saldo' : saldo}

        tanggallist = (data.filter(Tanggal__range=(tanggal_mulai, tanggal_akhir)).values_list("Tanggal", flat=True).distinct())
        saldoawal = saldo
        tanggallist = sorted(list(set((tanggallist.union(listtanggalmasuk).union(listtanggalpemusnahan)))))

        for i in tanggallist:
            datamodels = {
                'Tanggal': None,
                "SPK" : None,
                "Kodeproduk" : None,
                "Masuklembar":None,
                "Masukkonversi" : None,
                'Hasil' : None,
                'Keluar' : None,
                "Sisa" : None
            }

            filtertanggal = data.filter(Tanggal=i)
            filtertanggaltransaksigudang = datamasuk.filter(tanggal=i)
            filtertanggalpemusnahan = datapemusnahan.filter(Tanggal = i)

            jumlahmutasi =  filtertanggal.filter(Jenis ="Mutasi").aggregate(total = Sum('Jumlah'))['total']
            jumlahmasuk = filtertanggaltransaksigudang.aggregate(total = Sum('jumlah'))['total']
            jumlahpemusnahan = filtertanggalpemusnahan.aggregate(total=Sum('Jumlah'))['total']

            if jumlahmutasi is None:
                jumlahmutasi = 0
            if jumlahmasuk is None :
                jumlahmasuk = 0
            if jumlahpemusnahan is None:
                jumlahpemusnahan = 0

            # Cari data penyusun sesuai tanggal 
            penyusunfiltertanggal = models.Penyusun.objects.filter(KodeArtikel = artikel.id,KodeProduk = getbahanbakuutama.KodeProduk,KodeVersi__isdefault = True).first()

            print(getbahanbakuutama)

            if not penyusunfiltertanggal:
                penyusunfiltertanggal = models.Penyusun.objects.filter(KodeArtikel = artikel.id, KodeProduk = getbahanbakuutama.KodeProduk, KodeVersi__Tanggal__lte = i).order_by('KodeVersi__Tanggal').first()

            # konversimasterobj = models.KonversiMaster.objects.get(KodePenyusun=penyusunfiltertanggal.IDKodePenyusun)
            # konversi =konversimasterobj.Allowance
            konversi = penyusunfiltertanggal.Allowance
            # penyesuaiaanfilter
            # except:
            #     konversi = round(0)
            #     messages.error(request,'Data allowance belum di set')
            print(getbahanbakuutama)
            # print(asd)d
            penyesuaiaanfilter = models.PenyesuaianArtikel.objects.filter(KodeArtikel = artikel.id,TanggalMulai__lte = i, TanggalMinus__gte = i).last()
            # if penyesuaiaanfilter:
            #     konversi = penyesuaiaanfilter.konversi
            #     print(konversi)
                # print(asd)
            # print(penyesuaiaanfilter)
        
            # print(konversi,i,jumlahmasuk)
            # print(penyusunfiltertanggal)
            # print(asd)
            masukpcs = math.floor(jumlahmasuk/konversi)
            if penyesuaiaanfilter:
                print(jumlahmasuk,konversi)
                konversi= penyesuaiaanfilter.konversi
                print(konversi,penyesuaiaanfilter,i)
                print(masukpcs)
                masukpcs = math.floor(masukpcs*konversi)
                print(konversi)
                print(masukpcs)
                # print(asd)
            # if masukpcs != 0 :
                # print(masukpcs)

                # print(jumlahmasuk, konversi)
                # print(asd)
            
                
            saldoawal = saldoawal - jumlahmutasi + masukpcs-jumlahpemusnahan
            if saldoawal < 0 and requests != None:
                messages.warning(requests,f'Sisa stok menjadi negatif pada {i}')

            datamodels['Tanggal'] = i.strftime("%Y-%m-%d")
            datamodels['Masuklembar'] = jumlahmasuk
            datamodels['Masukkonversi'] = masukpcs
            datamodels['Sisa'] = saldoawal
            datamodels['Hasil'] = jumlahmutasi
            datamodels['SPK'] = filtertanggal.filter(Jenis = 'Mutasi')
            datamodels["Kodeproduk"] = penyusunfiltertanggal
            datamodels['Keluar'] = jumlahpemusnahan

            # Cari data penyesuaian

            
            listdata.append(datamodels)
    else:
        data = data.filter(Lokasi=1)
        try:
            saldoawalobj = models.SaldoAwalArtikel.objects.get(IDArtikel__KodeArtikel=artikel, IDLokasi__NamaLokasi=lokasi,Tanggal__range =(tanggal_mulai,tanggal_akhir))
            saldo = saldoawalobj.Jumlah
            saldoawalobj.Tanggal = saldoawalobj.Tanggal.year
        except models.SaldoAwalArtikel.DoesNotExist :
            saldo = 0
            saldoawalobj ={
                'Tanggal' : 'Belum ada Data',
                'saldo' : saldo
            }


        tanggalmutasi = data.filter(Jenis = 'Mutasi',Tanggal__range=(tanggal_mulai,tanggal_akhir)).values_list('Tanggal',flat=True).distinct()
        sppb = models.DetailSPPB.objects.filter(DetailSPK__KodeArtikel = artikel, NoSPPB__Tanggal__range = (tanggal_mulai,tanggal_akhir))
        tanggalsppb = sppb.values_list('NoSPPB__Tanggal',flat=True).distinct()
        datapemusnahan = models.PemusnahanArtikel.objects.filter(KodeArtikel = artikel,lokasi__NamaLokasi = "FG", Tanggal__range=(tanggal_mulai,tanggal_akhir))
        tanggalpemusnahan = datapemusnahan.values_list('Tanggal',flat=True).distinct()
        tanggallist = sorted(list(set(tanggalmutasi.union(tanggalsppb).union(tanggalpemusnahan))))

        saldoawal = saldo

        for i in tanggallist:
            datamodels = {
                "Tanggal" : None,
                "Penyerahanwip": None,
                "DetailSPPB" : None,
                "Sisa" : None,
                "Keluar" : None,
                'Jumlahkirim':None
            }

            penyerahanwip = models.TransaksiProduksi.objects.filter(Tanggal = i, KodeArtikel = artikel, Jenis = "Mutasi", Lokasi__NamaLokasi = "WIP" )
            detailsppbjobj = sppb.filter(NoSPPB__Tanggal = i)
            filteredpemusnahan = datapemusnahan.filter(Tanggal = i)

            totalpenyerahanwip = penyerahanwip.aggregate(total=Sum('Jumlah'))['total']
            totalkeluar = detailsppbjobj.aggregate(total=Sum('Jumlah'))['total']
            totalpemusnahan = filteredpemusnahan.aggregate(total=Sum('Jumlah'))['total']
            totaldikirim = detailsppbjobj.aggregate(total = Sum('Jumlah'))['total']
            if not totalpenyerahanwip:
                totalpenyerahanwip = 0
            if not totalkeluar :
                totalkeluar = 0
            if not totalpemusnahan:
                totalpemusnahan = 0
            if not totaldikirim :
                totaldikirim = 0

            saldoawal += totalpenyerahanwip - totalkeluar - totalpemusnahan



            datamodels ['Tanggal'] = i.strftime('%Y-%m-%d')
            datamodels ['Penyerahanwip'] = totalpenyerahanwip
            datamodels['DetailSPPB'] = detailsppbjobj
            datamodels['Sisa'] = saldoawal
            datamodels['Keluar'] = totalpemusnahan
            datamodels['Jumlahkirim'] = totaldikirim

            listdata.append(datamodels)

        # print(listdata)
    return listdata,saldoawalobj
# def hitung_ksbj(kodeartikel, lokasi, tahun):
#     try:
#         artikel = models.Artikel.objects.get(KodeArtikel = kodeartikel)
#     except:
#         return redirect('view_ksbj')
    
#     if request.GET['tahun']:
#         tahun = int(request.GET['tahun'])
#     else:
#         sekarang = datetime.now()
#         tahun - sekarang.year

#     tanggal_mulai = datetime(year=tahun, month=1, day=1)
#     tanggal_akhir = datetime(year=tahun, month=12, day=31)

#     print(tanggal_mulai)
#     print(tanggal_akhir)

#     lokasi = request.GET['lokasi']
#     lokasiobj = models.Lokasi.objects.get(NamaLokasi = lokasi)

#     getbahanbakuutama = models.Penyusun.objects.filter(KodeArtikel=artikel.id, Status=1, KodeVersi__isdefault = True).order_by('-KodeVersi__Tanggal').first()
#     print(getbahanbakuutama)
#     if not getbahanbakuutama :
#         messages.error(request, "Bahan Baku utama belum di set")
#         return redirect("view_ksbj")
    
#     data = models.TransaksiProduksi.objects.filter(KodeArtikel=artikel.id,Jenis = "Mutasi")
#     datamasuk = models.TransaksiGudang.objects.filter(DetailSPK__KodeArtikel = artikel.id,KodeProduk = getbahanbakuutama.KodeProduk,tanggal__range=(tanggal_mulai, tanggal_akhir))
#     listtanggalmasuk = datamasuk.values_list('tanggal',flat=True).distinct()
#     print(datamasuk)
#     listdata = []
#     if lokasi == "WIP":
#         datapemusnahan = models.PemusnahanArtikel.objects.filter(KodeArtikel = artikel,Tanggal__range=(tanggal_mulai,tanggal_akhir),lokasi__NamaLokasi = "WIP")
#         listtanggalpemusnahan = datapemusnahan.values_list('Tanggal',flat=True).distinct()
#         data = data.filter(Lokasi=lokasiobj.IDLokasi)
#         try:
#             saldoawalobj = models.SaldoAwalArtikel.objects.get(IDArtikel__KodeArtikel=kodeartikel, IDLokasi=lokasiobj.IDLokasi,Tanggal__range =(tanggal_mulai,tanggal_akhir))
#             saldo = saldoawalobj.Jumlah
#             saldoawalobj.Tanggal = saldoawalobj.Tanggal.year

#         except models.SaldoAwalArtikel.DoesNotExist :
#             saldo = 0
#             saldoawal = None
#             saldoawalobj ={'Tanggal' : 'Belum ada Data','saldo' : saldo}

#         tanggallist = (data.filter(Tanggal__range=(tanggal_mulai, tanggal_akhir)).values_list("Tanggal", flat=True).distinct())
#         saldoawal = saldo
#         tanggallist = sorted(list(set((tanggallist.union(listtanggalmasuk).union(listtanggalpemusnahan)))))

#         for i in tanggallist:
#             datamodels = {
#                 'Tanggal': None,
#                 "SPK" : None,
#                 "Kodeproduk" : None,
#                 "Masuklembar":None,
#                 "Masukkonversi" : None,
#                 'Hasil' : None,
#                 'Keluar' : None,
#                 "Sisa" : None
#             }

#             filtertanggal = data.filter(Tanggal=i)
#             filtertanggaltransaksigudang = datamasuk.filter(tanggal=i)
#             filtertanggalpemusnahan = datapemusnahan.filter(Tanggal = i)

#             jumlahmutasi =  filtertanggal.filter(Jenis ="Mutasi").aggregate(total = Sum('Jumlah'))['total']
#             jumlahmasuk = filtertanggaltransaksigudang.aggregate(total = Sum('jumlah'))['total']
#             jumlahpemusnahan = filtertanggalpemusnahan.aggregate(total=Sum('Jumlah'))['total']

#             if jumlahmutasi is None:
#                 jumlahmutasi = 0
#             if jumlahmasuk is None :
#                 jumlahmasuk = 0
#             if jumlahpemusnahan is None:
#                 jumlahpemusnahan = 0

#             # Cari data penyusun sesuai tanggal 
#             penyusunfiltertanggal = models.Penyusun.objects.filter(KodeArtikel = artikel.id,KodeProduk = getbahanbakuutama.KodeProduk,KodeVersi__isdefault = True).first()

#             print(getbahanbakuutama)

#             if not penyusunfiltertanggal:
#                 penyusunfiltertanggal = models.Penyusun.objects.filter(KodeArtikel = artikel.id, KodeProduk = getbahanbakuutama.KodeProduk, KodeVersi__Tanggal__lte = i).order_by('KodeVersi__Tanggal').first()

#             # konversimasterobj = models.KonversiMaster.objects.get(KodePenyusun=penyusunfiltertanggal.IDKodePenyusun)
#             # konversi =konversimasterobj.Allowance
#             konversi = penyusunfiltertanggal.Allowance
#             # except:
#             #     konversi = round(0)
#             #     messages.error(request,'Data allowance belum di set')
#             print(getbahanbakuutama)
#             # print(asd)d
#             penyesuaiaanfilter = models.PenyesuaianArtikel.objects.filter(KodeArtikel = artikel.id,TanggalMulai__lte = i, TanggalMinus__gte = i).first()
#             print(penyesuaiaanfilter)
        
#             print(konversi,i,jumlahmasuk)
#             print(penyusunfiltertanggal)
#             # print(asd)
#             masukpcs = math.ceil(jumlahmasuk/konversi)
#             if penyesuaiaanfilter:
#                 konversi= penyesuaiaanfilter.konversi
#                 print(konversi,penyesuaiaanfilter,i)
#                 masukpcs = math.ceil(masukpcs*konversi)
#                 # print(asd)
#             if masukpcs != 0 :
#                 print(masukpcs)

#                 print(jumlahmasuk, konversi)
#                 # print(asd)
            
                
#             saldoawal = saldoawal - jumlahmutasi + masukpcs-jumlahpemusnahan

#             datamodels['Tanggal'] = i.strftime("%Y-%m-%d")
#             datamodels['Masuklembar'] = jumlahmasuk
#             datamodels['Masukkonversi'] = masukpcs
#             datamodels['Sisa'] = saldoawal
#             datamodels['Hasil'] = jumlahmutasi
#             datamodels['SPK'] = filtertanggal.filter(Jenis = 'Mutasi')
#             datamodels["Kodeproduk"] = penyusunfiltertanggal
#             datamodels['Keluar'] = jumlahpemusnahan

#     return listdata, saldoawal


@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_rekapbarang(request):
    '''
    Fitur ini digunakan untuk melihat rekapitulasi semua bahan baku pada area produksi
    Algoritma 
    1. Mengambil data produk
    2. mengiterasi data produk
        A. Memanggil fungsi calculate_KSBB
    3. Menampilkan hasil rekapitulasi bahan baku  
    '''
    tanggal_akhir = request.GET.get("periode")
    try:
        tanggalakhirstrp = datetime.strptime(tanggal_akhir,'%Y-%M-%d')
    except:
        tanggalakhirstrp = datetime.now()
    tanggalmulaiselected = datetime(year=tanggalakhirstrp.year,month=1,day=1)
    try:

        lokasi = request.GET['lokasi']
    except: 
        lokasi = "WIP"
    
    sekarang = datetime.now()
    tahun = sekarang.year

    tanggal_mulai = datetime(year=tahun, month=1, day=1)

    dataproduk = models.Produk.objects.all()
    # dataproduk = models.Produk.objects.filter(KodeProduk = "A-001-01")

    if tanggal_akhir:
        for produk in dataproduk:
            listdata, saldoawal = calculate_KSBB(produk, tanggalmulaiselected, tanggal_akhir,lokasi)

            if listdata:
                produk.kuantitas = listdata[-1]["Sisa"][-1]
            else:
                produk.kuantitas = 0
    else:
        for produk in dataproduk:
            listdata, saldoawal = calculate_KSBB(produk, tanggal_mulai, sekarang,lokasi)

            if listdata:
                produk.kuantitas = listdata[-1]["Sisa"][-1]
            else:
                produk.kuantitas = 0
    # print(dataproduk)
    return render(request, "produksi/rekap_barang.html", {'data':dataproduk , 'tanggal_akhir':tanggal_akhir,'lokasi':lokasi})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_rekaprusak(request):
    '''
    FItur ini digunakan untuk melihat rekapitulasi pemusnahan bahan baku dan artikel pada Produksi
    Algoritma:
    1. Mendapatkan data waktu akhir rekapitulasi
    2. Medapatkan data input lokasi 
    3. Mengambil data pemusnahan bahan baku dengan kriteria Lokasi = lokasi (lokasi didapatkan dari input user), tanggal berada dalam rentang tanggal mulai dan akhir. Tanggal mulai diseting awal tahun 
    4. Mengambil data pemusnahan Artikel bahan baku dengan kriteria Lokasi = lokasi (lokasi didapatkan dari input user), tanggal berada dalam rentang tanggal mulai dan akhir. Tanggal mulai diseting awal tahun 
    5. Mengakumulasi data pemusnahan bahan baku
    6. Mengakumulasi data pemusnahan artikel
    7. Menampilkan rekap ke user
     
    '''

    tanggal_akhir = request.GET.get("periode")
    
    sekarang = datetime.now()
    tahun = sekarang.year

    lokasi = request.GET.get("lokasi")
    if lokasi:
        lokasiobj = models.Lokasi.objects.get(NamaLokasi=lokasi)
    else:
        lokasiobj = models.Lokasi.objects.get(IDLokasi=1)
        
    tanggal_mulai = datetime(year=tahun, month=1, day=1)

    if tanggal_akhir:
        tanggal_akhirstrp = datetime.strptime(tanggal_akhir,'%Y-%M-%d')
        tanggal_mulai = datetime(year=tanggal_akhirstrp.year,month=1,day=1)

        databarang = models.PemusnahanBahanBaku.objects.filter(lokasi=lokasiobj,Tanggal__range=(tanggal_mulai, tanggal_akhir)).values('KodeBahanBaku__KodeProduk','KodeBahanBaku__NamaProduk','KodeBahanBaku__unit','KodeBahanBaku__keteranganProduksi').annotate(kuantitas=Sum('Jumlah'))
        dataartikel = models.PemusnahanArtikel.objects.filter(lokasi=lokasiobj,Tanggal__range=(tanggal_mulai, tanggal_akhir)).values('KodeArtikel__KodeArtikel','KodeArtikel__keterangan').annotate(kuantitas=Sum('Jumlah'))

    else:
        databarang = models.PemusnahanBahanBaku.objects.filter(lokasi=lokasiobj,Tanggal__range=(tanggal_mulai, sekarang)).values('KodeBahanBaku__KodeProduk','KodeBahanBaku__NamaProduk','KodeBahanBaku__unit','KodeBahanBaku__keteranganProduksi').annotate(kuantitas=Sum('Jumlah'))
        dataartikel = models.PemusnahanArtikel.objects.filter(lokasi=lokasiobj,Tanggal__range=(tanggal_mulai, sekarang)).values('KodeArtikel__KodeArtikel','KodeArtikel__keterangan').annotate(kuantitas=Sum('Jumlah'))

    return render(request, "produksi/rekap_rusak.html", {"databarang": databarang, "dataartikel": dataartikel, "lokasi":lokasi,'tanggal_akhir':tanggal_akhir})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_rekapproduksi(request):
    '''
    Fitur ini digunakan untuk membuat rekapitulasi artikel pada setiap bulan pada WIP dan FG dalam 1 tahun
    Algoritma
    1. Mengambil data tanggal mulai dan tanggal akhir
    2. Mengambil data artikel
    3. mengiterasi data artikel
        A. Mendapatkan data bahan baku penyusun utama pada tabel penyusun dengan kriteria Kode ARtikel = artikel., dan Status = 1
        B. Mengambil data mutasi dari tabel Transaksi Produksi dengan kriteria Kode Artikel = artikel, Jenis = Mutasi, Tanggal berada dalam rentang tanggal mulai dan tanggal akhir
        C. mengambil data masuk dari tabel transaksi gydang dengan kriteria Kode artikel = artikel.id, Tanggal berada adalam rentang tanggal mulai dan akhir
        D. Memanggil fungsi calculate KSBJ
    4. Membuat dataframe 
    5. Menampilkan data rekap

    '''
    if len(request.GET) == 0:
        return render(request, "produksi/rekap_produksi.html")
    else:
        if request.GET["periode"]:
            tahun = int(request.GET["periode"])
        else:
            sekarang = datetime.now()
            tahun = sekarang.year

        tanggal_mulai = datetime(year=tahun, month=1, day=1)
        tanggal_akhir = datetime(year=tahun, month=12, day=31)

        datarekap = []

        artikelobj = models.Artikel.objects.all()
        # artikelobj = models.Artikel.objects.filter(KodeArtikel = '9010 AC')
        for artikel in artikelobj:

            print(artikel)
            
            dataartikel = []
            getbahanbakuutama = models.Penyusun.objects.filter(KodeArtikel=artikel.id, Status=1)

            if not getbahanbakuutama :
                messages.error(request, ("Bahan Baku",artikel.KodeArtikel,"belum di set"))
                continue

            data = models.TransaksiProduksi.objects.filter(KodeArtikel=artikel,Jenis = "Mutasi",Tanggal__range=(tanggal_mulai,tanggal_akhir))
            print(data)
            datamasuk = models.TransaksiGudang.objects.filter(DetailSPK__KodeArtikel = artikel.id,tanggal__range=(tanggal_mulai, tanggal_akhir),KodeProduk = getbahanbakuutama.first().KodeProduk)
            print(datamasuk)
            # print(datapemusnahan)
            # print(asd)
            listtanggalmasuk = datamasuk.values_list('tanggal',flat=True).distinct()

            if not data :
                messages.error(request, ("Tidak ditemukan data Transaki Produksi untuk Artikel",artikel.KodeArtikel))
                continue
            else:
                dataartikel.append(artikel.KodeArtikel)
                
            lokasiobj = models.Lokasi.objects.all()

            for lokasi in lokasiobj[:2]:

                listdata = []

                if lokasi.NamaLokasi == "WIP":
                    listdata,saldoawal = calculate_ksbj(artikel,'WIP',tahun)

                else:
                    listdata,saldoawal = calculate_ksbj(artikel,'FG',tahun)

                if not listdata:
                    pass
                else:
                    df = pd.DataFrame(listdata)

                    # Convert 'Tanggal' column to datetime
                    df['Tanggal'] = pd.to_datetime(df['Tanggal'])

                    # Resampling to get the last day of each month
                    df_resampled = df.resample('M', on='Tanggal').last().ffill().reset_index()

                    # Creating a new DataFrame with all months from 1 to 12
                    all_months = pd.date_range(start=tanggal_mulai, end=tanggal_akhir, freq='M')
                    df_all_months = pd.DataFrame({'Tanggal': all_months})

                    # Merging the resampled data with all months
                    result_df = pd.merge(df_all_months, df_resampled, on='Tanggal', how='left').ffill()

                    # Getting the data for all months
                    result_data = result_df.to_dict('records')
                
                    dataartikel.append(result_data)
            
            if len(dataartikel) == 1:
                pass
            else:
                item3 = [{'Tanggal': d1['Tanggal'], 'Sisa': d1['Sisa'] + d2['Sisa']} for d1, d2 in zip(dataartikel[1], dataartikel[2])]
                
                dataartikel.append(item3)

                datarekap.append(dataartikel)
        print(datarekap)

        return render(request, "produksi/rekap_produksi.html", {'artikel':artikelobj, 'data':datarekap, 'tahun':tahun })


# Pemusnahan Artikel
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_pemusnahan(request):
    '''
    Fitur ini digunakan unutk melakukan manajemen data pemusnahan artikel pada sistem
    Algoritma
    1. Mendefinisikan tanggal awal dan tanggal akhir secara default tanggal awal adalah awal bulan ketika user mengakses, akhir bulan adalah hari user mengakases
    2. mengambil data Pemusnahan Artikel dengan kriteria rentang tanggal berada dalam rentang tanggal awal dan tanggal akhir
    3. Menampilkan data pemusnahan artikel
    '''
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    dataproduksi = models.PemusnahanArtikel.objects.filter(Tanggal__range=(tanggalawal,tanggalakhir)).order_by("-Tanggal")
    for i in dataproduksi:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_pemusnahan.html", {"dataproduksi": dataproduksi,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_pemusnahan(request):
    '''
    Fitur ini digunakan untuk menambahkan data pemusnahan artikel pada sistem
    Algoritma
    1. Mengambil data aritkel
    2. Mengambil data lokasi 
    3. Menampilkan form input data pemusnahan artikel
    4. program mendapatkan inputan Versi artikel,kodeartikel, lokasi, jumlah, tanggal, dan keterangan
    5. Cek apakah kode artikel valid ? apabila tidak akan menampilkan pesan error
    6. Menyimpan data transaksi pemusnahan artikel 
    '''
    dataartikel = models.Artikel.objects.all()
    datalokasi = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_pemusnahan.html",
            {"nama_lokasi": datalokasi[:2], "dataartikel": dataartikel},
        )
    else:
        kodeversi = request.POST['versiartikel']
        kodeartikel = request.POST["artikel"]
        lokasi = request.POST["nama_lokasi"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]

        try:
            artikelobj = models.Artikel.objects.get(KodeArtikel=kodeartikel)
        except:
            messages.error(request, "Kode Artikel tidak ditemukan")
            return redirect("add_pemusnahan")
        try:
            detailversi = models.Versi.objects.get(pk = kodeversi)
        except:
            detailversi = models.Versi.objects.filter(KodeArtikel = artikelobj).last()
            messages.error(request,f'Kode Versi tidak ditemukan, versi disetting pada versi terbaru {detailversi.Versi} Tanggal {detailversi.Tanggal}')
        lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)
        pemusnahanobj = models.PemusnahanArtikel(
            Tanggal=tanggal, Jumlah=jumlah, KodeArtikel=artikelobj, lokasi=lokasiobj, Keterangan=keterangan, VersiArtikel = detailversi
        )
        pemusnahanobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Create",
            pesan=f"Pemusnahan Artikel. Kode Artikel : {artikelobj.KodeArtikel} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request,'Data Berhasil ditambahkan')
        return redirect("view_pemusnahan")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_pemusnahan(request, id):
    '''
    Fitur ini digunakan untuk melakukan update pemusnahan artikel 
    Algoritma
    1. Mengambil data semua artikel
    2. Mengambil data Pemusnahan Artikel
    3. Mengambil data Versi dengan kriteria Kode Artikel = artikel pada transaksi pemusnahanArtikel
    4. Menampilkan form update pemusnahan 
    5. Program menerima input berupa Kode Artikel, Lokasi, Jumlah, Tanggal, Keterangan, dan Versi Artikel
    6. Mengupdate data transaksi pemusnahan artikel
    '''
    dataartikel = models.Artikel.objects.all()
    dataobj = models.PemusnahanArtikel.objects.get(IDPemusnahanArtikel=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    dataversi = models.Versi.objects.filter(KodeArtikel = dataobj.KodeArtikel)
    if request.method == "GET":
        return render(
            request,
            "produksi/update_pemusnahan.html",
            {"data": dataobj, "nama_lokasi": lokasiobj[:2], "dataartikel": dataartikel,'dataversi':dataversi},
        )

    else:
        
        kodeartikel = request.POST["artikel"]
        lokasi = request.POST["nama_lokasi"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]
        kodeversi = request.POST['versiartikel']
        
        try:
            artikelobj = models.Artikel.objects.get(KodeArtikel=kodeartikel)
        except:
            messages.error(request, "Kode Artikel tidak ditemukan")
            return redirect("update_pemusnahan" ,id=id)
        try:
            detailversi = models.Versi.objects.get(pk = kodeversi)
        except:
            detailversi = models.Versi.objects.filter(KodeArtikel = artikelobj).last()
            messages.error(request,f'Data versi tidak ditemukan, menggunakan versi terbaru {detailversi.Versi} - {detailversi.Tanggal}')
        lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.KodeArtikel = artikelobj
        dataobj.lokasi = lokasiobj
        dataobj.Keterangan = keterangan
        dataobj.VersiArtikel = detailversi

        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Pemusnahan Artikel. Kode Artikel : {artikelobj.KodeArtikel} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_pemusnahan")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_pemusnahan(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi pemusnahan artikel pada Produksi
    Algoritma:
    1. Mengambil data transaksi pemusnahna Artikel dengan kriteria IDPemusnahanARtikel = id (id didapatkan dari passing values HTML)
    2. Menghapus data 
    '''
    dataobj = models.PemusnahanArtikel.objects.get(IDPemusnahanArtikel=id)
    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Pemusnahan Artikel. Kode Artikel : {dataobj.KodeArtikel.KodeArtikel} Jumlah : {dataobj.Jumlah} Lokasi : {dataobj.lokasi.NamaLokasi}",
    ).save()
    messages.success(request,'Data berhasil dihapus')

    return redirect(view_pemusnahan)


# Pemusnahan Barang
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_pemusnahanbarang(request):
    '''
    Fitur ini digunakan unutk melakukan manajemen data pemusnahan Bahan Baku pada Produksi
    Algoritma
    1. Mendefinisikan tanggal awal dan tanggal akhir secara default tanggal awal adalah awal bulan ketika user mengakses, akhir bulan adalah hari user mengakases
    2. mengambil data Pemusnahan Bahan Baku dengan kriteria rentang tanggal berada dalam rentang tanggal awal dan tanggal akhir, Lokasi pada WIP dan FG
    3. Menampilkan data pemusnahan artikel
    '''
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    dataproduksi = models.PemusnahanBahanBaku.objects.filter(lokasi__NamaLokasi__in=("WIP","FG"),Tanggal__range=(tanggalawal,tanggalakhir)).order_by("-Tanggal")
    for i in dataproduksi:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_pemusnahanbarang.html", {"dataproduksi": dataproduksi,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_pemusnahanbarang(request):
    '''
    Fitur ini digunakan untuk menambahkan data pemusnahan Bahan Baku pada Produksi
    Algoritma
    1. Mengambil data Kode Bahan Baku
    2. Mengambil data lokasi 
    3. Menampilkan form input data pemusnahan artikel
    4. program mendapatkan inputan  kode bahan baku, lokasi, jumlah, tanggal, dan keterangan
    5. Cek apakah kode artikel valid ? apabila tidak akan menampilkan pesan error
    6. Menyimpan data transaksi pemusnahan artikel 
    '''
    databarang = models.Produk.objects.all()
    datalokasi = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_pemusnahanbarang.html",
            {"nama_lokasi": datalokasi[:2], "databarang": databarang},
        )
    else:
        print(request.POST)
        # Get object
        kodeproduk = request.POST["produk"]
        lokasi = request.POST["nama_lokasi"]
        jumlah = float(request.POST["jumlah"])
        tanggal = request.POST["tanggal"]
        lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)
        keterangan = request.POST["keterangan"]

        try:
            produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Bahan Baku tidak ditemukan")
            return redirect("add_pemusnahanbarang")
        
        pemusnahanobj = models.PemusnahanBahanBaku(
            Tanggal=tanggal, Jumlah=jumlah, KodeBahanBaku=produkobj, lokasi=lokasiobj, Keterangan=keterangan
        )
        pemusnahanobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Create",
            pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {produkobj.KodeProduk} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request, "Data berhasil disimpan")

        return redirect("view_pemusnahanbarang")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_pemusnahanbarang(request, id):
    '''
    Fitur ini digunakan untuk melakukan update pemusnahan Bahan Baku Produksi 
    Algoritma
    1. Mengambil data semua Produk
    2. Mengambil data Versi dengan kriteria IDPemusnahanBahanBaku = id (id didapatkan dari passing values HTML)
    3. Menampilkan form update pemusnahan 
    4. Program menerima input berupa Kode Bahan Baku, Lokasi, Jumlah, Tanggal, Keterangan
    5. Mengupdate data transaksi pemusnahan Bahan Baku
    '''
    databarang = models.Produk.objects.all()
    dataobj = models.PemusnahanBahanBaku.objects.get(IDPemusnahanBahanBaku=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    if request.method == "GET":

        return render(
            request,
            "produksi/update_pemusnahanbarang.html",
            {"data": dataobj, "nama_lokasi": lokasiobj[:2], 'dataproduk':databarang},
        )

    else:
        kodeproduk = request.POST["produk"]
        lokasi = request.POST["nama_lokasi"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]

        try:
            produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Bahan Baku tidak ditemukan")
            return redirect("update_pemusnahanbarang",id=id)
        lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.KodeBahanBaku = produkobj
        dataobj.lokasi = lokasiobj
        dataobj.Keterangan = keterangan
        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {produkobj.KodeProduk} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request,'Data berhasil diupdate')
        return redirect("view_pemusnahanbarang")
        

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_pemusnahanbarang(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi pemusnahan Bahan Baku pada Produksi
    Algoritma:
    1. Mengambil data transaksi pemusnahna BahanBaku dengan kriteria IDPemusnahanBahanBaku = id (id didapatkan dari passing values HTML)
    2. Menghapus data 
    '''
    dataobj = models.PemusnahanBahanBaku.objects.get(IDPemusnahanBahanBaku=id)

    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {dataobj.KodeBahanBaku.KodeProduk} Jumlah : {dataobj.Jumlah} Lokasi : {dataobj.lokasi.NamaLokasi}",
    ).save()

    return redirect(view_pemusnahanbarang)

# Pemusnahan Artikel SUBKON
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_pemusnahanproduksubkon(request):
    '''
    Fitur ini digunakan unutk melakukan manajemen data pemusnahan produk subkon pada Produksi
    Algoritma
    1. Mendefinisikan tanggal awal dan tanggal akhir secara default tanggal awal adalah awal bulan ketika user mengakses, akhir bulan adalah hari user mengakases
    2. mengambil data Pemusnahan produk subkon dengan kriteria rentang tanggal berada dalam rentang tanggal awal dan tanggal akhir
    3. Menampilkan data pemusnahan produk subkon
    '''
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    dataproduksi = models.PemusnahanProdukSubkon.objects.filter(Tanggal__range = (tanggalawal,tanggalakhir)).order_by("-Tanggal")
    for i in dataproduksi:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_pemusnahanproduksubkon.html", {"dataproduksi": dataproduksi,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_pemusnahanproduksubkon(request):
    '''
    Fitur ini digunakan untuk menambahkan data pemusnahan produk subkon pada Produksi
    Algoritma
    1. Mengambil data Kode produk subkon
    2. Menampilkan form input data pemusnahan Produk Subkon
    3. program mendapatkan inputan  kode bahan baku, jumlah, tanggal, dan keterangan
    4. Cek apakah kode produk subkon valid ? apabila tidak akan menampilkan pesan error
    5. Menyimpan data transaksi pemusnahan produk subkon 
    '''
    dataartikel = models.ProdukSubkon.objects.all()
    datalokasi = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_pemusnahanproduksubkon.html",
            {"nama_lokasi": datalokasi[:2], "dataartikel": dataartikel},
        )
    else:

        print(request.POST)
        kodeartikel = request.POST["kodebarangHidden"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]
        # print(asd)

        try:
            artikelobj = models.ProdukSubkon.objects.get(pk=kodeartikel)
        except:
            messages.error(request, "Kode Artikel tidak ditemukan")
            return redirect("add_pemusnahanproduksubkon")

        lokasiobj = models.Lokasi.objects.get(NamaLokasi="WIP")
        pemusnahanobj = models.PemusnahanProdukSubkon(
            Tanggal=tanggal, Jumlah=jumlah, KodeProdukSubkon=artikelobj, lokasi=lokasiobj, Keterangan=keterangan
        )
        pemusnahanobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Create",
            pesan=f"Pemusnahan Produk Subkon. Kode Produk Subkon : {artikelobj.NamaProduk}-{artikelobj.KodeArtikel} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request,'Data Berhasil ditambahkan')
        return redirect("view_pemusnahanproduksubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_pemusnahanproduksubkon(request, id):
    '''
    Fitur ini digunakan untuk melakukan update pemusnahan produk subkon  
    Algoritma
    1. Mengambil data semua Produksubkon
    2. Mengambil data pemusnahan produk subkon dengan kriteria IDPemusnahanArtikel = id (id didapatkan dari passing values HTML)
    3. Menampilkan form update pemusnahan 
    4. Program menerima input berupa Kode produk subkon, Jumlah, Tanggal, Keterangan
    5. Mengupdate data transaksi pemusnahan produk subkon
    '''
    dataartikel = models.ProdukSubkon.objects.all()
    dataobj = models.PemusnahanProdukSubkon.objects.get(IDPemusnahanArtikel=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/update_pemusnahanproduksubkon.html",
            {"data": dataobj, "nama_lokasi": lokasiobj[:2], "dataartikel": dataartikel},
        )

    else:
        print(request.POST)
        # print(asd)
        kodeartikel = request.POST["kodebarangHidden"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]
        
        try:
            artikelobj = models.ProdukSubkon.objects.get(pk=kodeartikel)
        except:
            messages.error(request, "Kode Artikel tidak ditemukan")
            return redirect("update_pemusnahan" ,id=id)
        lokasiobj = models.Lokasi.objects.get(NamaLokasi="WIP")

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.KodeProdukSubkon = artikelobj
        dataobj.lokasi = lokasiobj
        dataobj.Keterangan = keterangan

        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Pemusnahan Artikel. Kode Produk Subkon : {artikelobj.NamaProduk} {artikelobj.KodeArtikel} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_pemusnahanproduksubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_pemusnahanproduksubkon(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi pemusnahan Produk Subkon pada Produksi
    Algoritma:
    1. Mengambil data transaksi pemusnahna Produk Subkon dengan kriteria IDPemusnahanArtikel = id (id didapatkan dari passing values HTML)
    2. Menghapus data 
    '''
    dataobj = models.PemusnahanProdukSubkon.objects.get(IDPemusnahanArtikel=id)
    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Pemusnahan Produk Subkon. Kode Produk : {dataobj.KodeProdukSubkon.NamaProduk} {dataobj.KodeProdukSubkon.KodeArtikel} Jumlah : {dataobj.Jumlah} Lokasi : {dataobj.lokasi.NamaLokasi}",
    ).save()
    messages.success(request,'Data berhasil dihapus')

    return redirect('view_pemusnahanproduksubkon')


# Pemusnahan Barang
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_pemusnahanbarangsubkon(request):
    '''
    Fitur ini digunakan unutk melakukan manajemen data pemusnahan Bahan Baku subkon pada Produksi
    Algoritma
    1. Mendefinisikan tanggal awal dan tanggal akhir secara default tanggal awal adalah awal bulan ketika user mengakses, akhir bulan adalah hari user mengakases
    2. mengambil data Pemusnahan Bahan Baku subkon dengan kriteria rentang tanggal berada dalam rentang tanggal awal dan tanggal akhir
    3. Menampilkan data pemusnahan Bahan Baku subkon
    '''
    if len(request.GET) == 0:
        tanggalakhir = datetime.now().date()
        tanggalawal = date(tanggalakhir.year,tanggalakhir.month,1).strftime('%Y-%m-%d')
        tanggalakhir = tanggalakhir.strftime('%Y-%m-%d')
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    dataproduksi = models.PemusnahanBahanBakuSubkon.objects.filter(lokasi__NamaLokasi__in=("WIP","FG"), Tanggal__range = (tanggalawal,tanggalakhir)).order_by("-Tanggal")
    for i in dataproduksi:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_pemusnahanbarangsubkon.html", {"dataproduksi": dataproduksi,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_pemusnahanbarangsubkon(request):
    '''
    Fitur ini digunakan untuk menambahkan data pemusnahan Bahan Baku subkon pada Produksi
    Algoritma
    1. Mengambil data Kode Bahan Baku subkon
    2. Menampilkan form input data pemusnahan Bahan Baku Subkon
    3. program mendapatkan inputan  kode bahan baku, jumlah, tanggal, dan keterangan
    4. Cek apakah kode Bahan Baku subkon valid ? apabila tidak akan menampilkan pesan error
    5. Menyimpan data transaksi pemusnahan Bahan Baku subkon 
    '''
    databarang = models.BahanBakuSubkon.objects.all()
    datalokasi = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_pemusnahanbarangsubkon.html",
            {"nama_lokasi": datalokasi[:2], "databarang": databarang},
        )
    else:
        print(request.POST)
        # print(asd)
        # Get object
        kodeproduk = request.POST["produk"]
        jumlah = float(request.POST["jumlah"])
        tanggal = request.POST["tanggal"]
        lokasiobj = models.Lokasi.objects.get(NamaLokasi="WIP")
        keterangan = request.POST["keterangan"]

        try:
            produkobj = models.BahanBakuSubkon.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Bahan Baku tidak ditemukan")
            return redirect("add_pemusnahanbarang")
        
        pemusnahanobj = models.PemusnahanBahanBakuSubkon(
            Tanggal=tanggal, Jumlah=jumlah, KodeBahanBaku=produkobj, lokasi=lokasiobj, Keterangan=keterangan
        )
        pemusnahanobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Create",
            pesan=f"Pemusnahan Bahan Baku Subkon. Kode Bahan Baku : {produkobj.KodeProduk} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request, "Data berhasil disimpan")

        return redirect("view_pemusnahanbarangsubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_pemusnahanbarangsubkon(request, id):
    '''
    Fitur ini digunakan untuk melakukan update pemusnahan Bahan Baku subkon  
    Algoritma
    1. Mengambil data semua Produksubkon
    2. Mengambil data pemusnahan Bahan Baku subkon dengan kriteria IDPemusnahanPemusnahanBahanBaku = id (id didapatkan dari passing values HTML)
    3. Menampilkan form update pemusnahan 
    4. Program menerima input berupa Kode Bahan Baku subkon, Jumlah, Tanggal, Keterangan
    5. Mengupdate data transaksi pemusnahan Bahan Baku subkon
    '''
    databarang = models.BahanBakuSubkon.objects.all()
    dataobj = models.PemusnahanBahanBakuSubkon.objects.get(IDPemusnahanBahanBaku=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    if request.method == "GET":

        return render(
            request,
            "produksi/update_pemusnahanbarangsubkon.html",
            {"data": dataobj, "nama_lokasi": lokasiobj[:2], 'dataproduk':databarang},
        )

    else:
        print(request.POST)
        kodeproduk = request.POST["produk"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST["keterangan"]

        try:
            produkobj = models.BahanBakuSubkon.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Bahan Baku tidak ditemukan")
            return redirect("update_pemusnahanbarangsubkon",id=id)
        lokasiobj = models.Lokasi.objects.get(NamaLokasi='WIP')

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.KodeBahanBaku = produkobj
        dataobj.lokasi = lokasiobj
        dataobj.Keterangan = keterangan
        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Pemusnahan Bahan Baku Subkon. Kode Bahan Baku : {produkobj.KodeProduk} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request,'Data berhasil diupdate')
        return redirect("view_pemusnahanbarangsubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_pemusnahanbarangsubkon(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi pemusnahan Bahan Baku Subkon pada Produksi
    Algoritma:
    1. Mengambil data transaksi pemusnahna Produk Subkon dengan kriteria IDPemusnahanBahanBaku = id (id didapatkan dari passing values HTML)
    2. Menghapus data 
    '''
    dataobj = models.PemusnahanBahanBakuSubkon.objects.get(IDPemusnahanBahanBaku=id)

    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {dataobj.KodeBahanBaku.KodeProduk} Jumlah : {dataobj.Jumlah} Lokasi : {dataobj.lokasi.NamaLokasi}",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect("view_pemusnahanbarangsubkon")


# Penyesuaian
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def penyesuaianartikel(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data penyesuaian Artikel pada produksi
    Algoritma
    1. Mengambil data Penyesuaian Artikel 
    2. mengubah format tanggal minus dan tanggal mulai menjadi YYYY-mm-dd
    '''
    datapenyesuaian = models.PenyesuaianArtikel.objects.all()
    for item in datapenyesuaian:
        item.TanggalMinus = item.TanggalMinus.strftime('%Y-%m-%d')
        item.TanggalMulai = item.TanggalMulai.strftime('%Y-%m-%d')
    return render(
        request, "produksi/view_penyesuaianartikel.html", {"datapenyesuaian": datapenyesuaian}
    )
    
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def addpenyesuaianartikel(request):
    '''
    Fitur ini digunakan untuk menambahkan data penyesuaian artikel 
    Algoritma
    1. Mengambil data semua artikel 
    2. menampilkan form input penyesuaian artikel
    3. program menerima input berupa Tanggal mulai, Tanggal minus, list artikel, list kuantitas
    4. mengiterasi list artikel dan kuantitas
        A. Menyimpan data penyesuaian artikel
    '''
    dataartikel = models.Artikel.objects.all()
    if request.method == "GET":
        return render(
            request, "produksi/add_penyesuaianartikel.html", {"Artikel": dataartikel}
        )
    else:
        print(request.POST)
        # print(asd)
        # Add Penyesuaian
        tanggalmulai = request.POST["tanggalmulai"]
        tanggalminus = request.POST['tanggalminus']
        lokasi = request.POST['lokasi']
        lokasi = models.Lokasi.objects.get(NamaLokasi = lokasi)
        
        listidartikel = request.POST.getlist('artikel_display')
        listkuantitas = request.POST.getlist("kuantitas")

        for artikel,kuantitas in zip(listidartikel,listkuantitas):
            # print(artikel,bahanbaku,kuantitas)
            kodeartikel = models.Artikel.objects.get(KodeArtikel = artikel)
            penyesuaianobj = models.PenyesuaianArtikel(
                TanggalMulai=tanggalmulai,
                TanggalMinus = tanggalminus,
                lokasi = lokasi,
                KodeArtikel = kodeartikel,
                konversi = kuantitas

            )
            penyesuaianobj.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Penyesuaian. Kode Artikel : {kodeartikel} Tanggal Mulai : {tanggalmulai} Tanggal Minus : {tanggalminus} Konversi : {kuantitas} ",
            ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_penyesuaianartikel")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_penyesuaianartikel(request, id):
    '''
    Fitur ini digunakan untuk melakukan update penyesuaian artikel 
    Algoritma
    1. Mendapatkan data Artikel
    2. mendapatakan data transaksi penyesuaianArtikel dengan kriteria pk (primarykey) = id (id didapatkan dari passing values HTML)
    3.  menampilkan form update transaksi penyesuaian artikel
    4. program menerima input berupa tanggal mulai, tanggal minus, id penyeusian, kuantitas, dan artikel
    5. Mengupdate data penyesuaian artikel
    '''
    dataartikel = models.Artikel.objects.all()

    datapenyesuaianobj = models.PenyesuaianArtikel.objects.get(pk = id)
    datapenyesuaianobj.TanggalMulai = datapenyesuaianobj.TanggalMulai.strftime('%Y-%m-%d')
    datapenyesuaianobj.TanggalMinus = datapenyesuaianobj.TanggalMinus.strftime('%Y-%m-%d')

    if request.method == "GET":
        return render(
            request,
            "produksi/update_penyesuaianartikel.html",
            {"dataobj": datapenyesuaianobj, "Artikel": dataartikel},
        )
    else:
        print(request.POST)
        tanggalmulai = request.POST["tanggalmulai"]
        tanggalminus = request.POST['tanggalminus']
        idpenyesuaian = request.POST['idpenyesuaian']
        kuantitas = request.POST['kuantitas']
        kodeartikel = request.POST['artikel']
        lokasi = request.POST['lokasi']

        try:
            artikelobj=models.Artikel.objects.get(id = kodeartikel)
        except models.Artikel.DoesNotExist:
            messages.error(request,f"Data Artikel {kodeartikel} tidak ditemukan dalam database")
            return redirect('update_penyesuaianartikel',id = id)

       
        penyesuaianobj = models.PenyesuaianArtikel.objects.get(
            IDPenyesuaian=idpenyesuaian
        )
        penyesuaianobj.TanggalMinus = tanggalminus
        penyesuaianobj.TanggalMulai = tanggalmulai
        penyesuaianobj.KodeArtikel = artikelobj
        penyesuaianobj.lokasi = models.Lokasi.objects.get(NamaLokasi = lokasi)

        penyesuaianobj.konversi = kuantitas
        penyesuaianobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Penyesuaian. Kode Artikel : {datapenyesuaianobj.KodeArtikel} Tanggal Mulai : {datapenyesuaianobj.TanggalMulai} Tanggal Minus : {datapenyesuaianobj.TanggalMinus} Konversi : {kuantitas} ",
        ).save()
        messages.success(request,"Data berhasil disimpan")
        return redirect("view_penyesuaianartikel")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_penyesuaianartikel(request, id):
    '''
    Fitur ini diguanakan untuk menghapus data penyesuian Artikel
    Algoritma
    1. Mengambil data Penyesuaian Artikel dengan kriteria IDPenyesuaian = id (id didapatkan dari passing values HTML)
    2. Menghapus data penyesuaian
    '''
    datapenyesuaianobj = models.PenyesuaianArtikel.objects.get(IDPenyesuaian=id)
    datapenyesuaianobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Penyesuaian. Kode Artikel : {datapenyesuaianobj.KodeArtikel} Tanggal Mulai : {datapenyesuaianobj.TanggalMulai} Tanggal Minus : {datapenyesuaianobj.TanggalMinus} Konversi : {datapenyesuaianobj.konversi} ",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect("view_penyesuaianartikel")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def penyesuaian(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data penyesuaian Bahan Baku pada produksi
    Algoritma
    1. Mengambil data Penyesuaian Bahan Baku 
    2. mengubah format tanggal minus dan tanggal mulai menjadi YYYY-mm-dd
    '''
    datapenyesuaian = models.Penyesuaian.objects.all()
    for item in datapenyesuaian:
        item.TanggalMulai = item.TanggalMulai.strftime('%Y-%m-%d')
        item.TanggalMinus = item.TanggalMinus.strftime('%Y-%m-%d')
    return render(
        request, "produksi/view_penyesuaian.html", {"datapenyesuaian": datapenyesuaian}
    )
    
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def addpenyesuaian(request):
    '''
    Fitur ini digunakan untuk menambahkan data penyesuaian Kode Bahan Baku 
    Algoritma
    1. Mengambil data semua Kode Bahan Baku 
    2. menampilkan form input penyesuaian Kode Bahan Baku
    3. program menerima input berupa Tanggal mulai, Tanggal minus, list Kode Bahan Baku, list kuantitas
    4. mengiterasi list Kode Bahan Baku dan kuantitas
        A. Menyimpan data penyesuaian Kode Bahan Baku
    '''
    dataartikel = models.Artikel.objects.all()
    kodebahanbaku = models.Produk.objects.all()
    if request.method == "GET":
        return render(
            request, "produksi/add_penyesuaian.html", {"Artikel": dataartikel,'kodebahanbaku':kodebahanbaku}
        )
    else:
        print(request.POST)
        # print(asd)
        # Add Penyesuaian
        tanggalmulai = request.POST["tanggalmulai"]
        tanggalminus = request.POST['tanggalminus']
        lokasi = request.POST['lokasi']
        lokasi = models.Lokasi.objects.get(NamaLokasi = lokasi)
        
        listidartikel = request.POST.getlist('artikel_display')
        listkuantitas = request.POST.getlist("kuantitas")
        listbahanbaku = request.POST.getlist('kodebahanbaku')

        for artikel,bahanbaku,kuantitas in zip(listidartikel,listbahanbaku,listkuantitas):
            print(artikel,bahanbaku,kuantitas)
            kodeartikel = models.Artikel.objects.get(KodeArtikel = artikel)
            bahanbaku = models.Produk.objects.get(KodeProduk = bahanbaku)
            penyesuaianobj = models.Penyesuaian(
                TanggalMulai=tanggalmulai,
                TanggalMinus = tanggalminus,
                lokasi = lokasi,

                KodeProduk = bahanbaku,
                KodeArtikel = kodeartikel,
                konversi = kuantitas

            )
            penyesuaianobj.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Penyesuaian. Kode Artikel : {kodeartikel} Kode bahan baku : {bahanbaku} Tanggal Mulai : {tanggalmulai} Tanggal Minus : {tanggalminus} Konversi : {kuantitas} ",
            ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_penyesuaian")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_penyesuaian(request, id):
    '''
    Fitur ini digunakan untuk melakukan update penyesuaian Kode Bahan Baku 
    Algoritma
    1. Mendapatkan data Kode Bahan Baku
    2. mendapatakan data transaksi Penyesuaian dengan kriteria pk (primarykey) = id (id didapatkan dari passing values HTML)
    3.  menampilkan form update transaksi penyesuaian Kode Bahan Baku
    4. program menerima input berupa tanggal mulai, tanggal minus, id penyeusian, kuantitas, dan Kode Bahan Baku
    5. Mengupdate data penyesuaian Kode Bahan Baku
    '''
    
    dataartikel = models.Artikel.objects.all()
    dataproduk = models.Produk.objects.all()

    datapenyesuaianobj = models.Penyesuaian.objects.get(pk = id)
    datapenyesuaianobj.TanggalMulai = datapenyesuaianobj.TanggalMulai.strftime('%Y-%m-%d')
    datapenyesuaianobj.TanggalMinus = datapenyesuaianobj.TanggalMinus.strftime('%Y-%m-%d')

    if request.method == "GET":
        return render(
            request,
            "produksi/update_penyesuaian.html",
            {"dataobj": datapenyesuaianobj, "Artikel": dataartikel,'kodebahanbaku':dataproduk},
        )
    else:
        print(request.POST)
        tanggalmulai = request.POST["tanggalmulai"]
        tanggalminus = request.POST['tanggalminus']
        idpenyesuaian = request.POST['idpenyesuaian']
        kuantitas = request.POST['kuantitas']
        kodeartikel = request.POST['artikel']
        kodebahanbaku = request.POST['kodebahanbaku']
        lokasi = request.POST['lokasi']

        try:
            artikelobj=models.Artikel.objects.get(id = kodeartikel)
        except models.Artikel.DoesNotExist:
            messages.error(request,f"Data Artikel {kodeartikel} tidak ditemukan dalam database")
            return redirect('update_penyesuaian',id = id)

        try:
            produkobj=models.Produk.objects.get(KodeProduk = kodebahanbaku)
        except models.Produk.DoesNotExist:
            messages.error(request,f"Data Produk {kodebahanbaku} tidak ditemukan dalam database")
            return redirect('update_penyesuaian',id = id)

        penyesuaianobj = models.Penyesuaian.objects.get(
            IDPenyesuaian=idpenyesuaian
        )
        penyesuaianobj.TanggalMinus = tanggalminus
        penyesuaianobj.TanggalMulai = tanggalmulai
        penyesuaianobj.KodeArtikel = artikelobj
        penyesuaianobj.KodeProduk = produkobj
        penyesuaianobj.lokasi = models.Lokasi.objects.get(NamaLokasi = lokasi)

        penyesuaianobj.konversi = kuantitas
        penyesuaianobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Penyesuaian. Kode Artikel : {datapenyesuaianobj.KodeArtikel} Kode Bahan Baku : {datapenyesuaianobj.KodeProduk} Tanggal Mulai : {datapenyesuaianobj.TanggalMulai} Tanggal Minus : {datapenyesuaianobj.TanggalMinus} Konversi : {kuantitas} ",
        ).save()
        messages.success(request,"Data berhasil disimpan")
        return redirect("view_penyesuaian")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_penyesuaian(request, id):
    '''
    Fitur ini diguanakan untuk menghapus data penyesuian Bahan Baku
    Algoritma
    1. Mengambil data Penyesuaian Bahan Baku dengan kriteria IDPenyesuaian = id (id didapatkan dari passing values HTML)
    2. Menghapus data penyesuaian
    '''
    datapenyesuaianobj = models.Penyesuaian.objects.get(IDPenyesuaian=id)
    datapenyesuaianobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Penyesuaian. Kode Artikel : {datapenyesuaianobj.KodeArtikel} Kode Bahan Baku : {datapenyesuaianobj.KodeProduk} Tanggal Mulai : {datapenyesuaianobj.TanggalMulai} Tanggal Minus : {datapenyesuaianobj.TanggalMinus} Konversi : {datapenyesuaianobj.konversi} ",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect("view_penyesuaian")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def kalkulatorpenyesuaian2(request):
    '''
    Fitur ini digunakan untuk melakukan kalkulasi perhitungan penyesuaian
    Algoritma
    1. Mengambil data semua produk 
    2. Menampilkan form input kode produk
    3. Program menerima inputan Kode produk
    4. Memanggil fungsi Calculate_KSBB untuk menghitung KSBB dari produk yang dipilih
    4. Menampilkan data KSBB dari produk dan input Tanggal mulai, tagnggal akhir, Jumlah Aktual
    5. Apabila program mendapatkan input Tanggal Mulai, Tanggal Akhir dan jumlah aktual maka akan melanjutkan proses kalkulasi
    6. Mengiterasi data hasil Calculate_KSBB
        A. Apabila Tanggal iterasi berada pada rentang Tanggal mulai dan tanggal akhir maka melanjutkan program
        B. Menghitung Sumproduct dengan cara Jumlah kotak * Konversi untuk setiap data pada rentang tanggal mulai dan akhir
        C. Menghitung kumulasi Jumlah data masuk  pada rentang  tanggal mulai dan akhir
        D. Menghitung kumulasi jumlah keluar pada rentang tanggal mulai dan akhir
        E. Menambahkan data Jumlah x Konversi dan Jumlah bahan baku tiap artikel pada dictionary jumlahxkonversi
    7. Mendapatkan saldo data terakhir dengan cara mengambil saldo data terakhir dari rentang tanggal mulai dan akhir
    8. Mengambil data pemusnahan bahan baku dengan kriteria Kode Bahan Baku = bahan baku terpilih, Tanggal diantara tanggal mulai dan akhir
    9. Apabila ada data pemusnahan baku pada maka jumlah data aktual dijumlahkan dengan agregasi total pemusnahan bahan baku 
    10. Menghitung Keluar penyesuaian dengan rumus = Jumlah Keluar - (Jumlah Aktual - saldo data)
    11. Mengiterasi data jumlahxkonversi
        A. menghitung jumlah penyesuaian = keluar penyesuaian/jumlah barang keluar perartikel
        B. menghitung productpersumproduct = jumlahxkonversi / total sumproduct
        C. Menghitung konveri baru = Jumlah penyesuaian * Productpersumproduct
    12. Mengupdate perhitungan pada hasil calculate_KSBB

    '''
    kodeproduk = models.Produk.objects.all()
    if len(request.GET) == 0:
        return render(
            request,
            "produksi/kalkulator_penyesuaian.html",
            {"kodeprodukobj": kodeproduk},
        )
    else:
        print(request.GET)
        # print(asd)
        """
        1. Cari 
        """
        try:
            produk = models.Produk.objects.get(KodeProduk=request.GET["kodebarang"])
            nama = produk.NamaProduk
            satuan = produk.unit
        except:
            messages.error(request, "Data Produk tidak ditemukan")
            return redirect("view_ksbb")
        
        if request.GET["periode"]:
            tahun = int(request.GET["periode"])
        else:
            sekarang = datetime.now()
            tahun = sekarang.year
        try:
            tahun = int(request.GET['tahundata'])
        except:
            pass
        tanggal_mulai = datetime(year=tahun, month=1, day=1)
        tanggal_akhir = datetime(year=tahun, month=12, day=31)
        lokasi = request.GET['lokasi']
        listdata,saldoawal = calculate_KSBB(produk,tanggal_mulai,tanggal_akhir,lokasi)
        print(listdata) 
        if len(listdata) == 0:
            messages.warning(request,f'Tidak ditemukan transaksi ')
            return render(
            request,
            "produksi/kalkulator_penyesuaian.html",{"kodeprodukobj": kodeproduk})
        # print(asd)    

        datasisaminus = 0
        datajumlah = 0
        dataajumlahartikel = {}
        datakonversiartikel = {}
        # print(data)

        firstindex = True

        listartikel = models.Penyusun.objects.filter(KodeProduk = produk).values_list("KodeArtikel__KodeArtikel",flat=True).distinct()
        penyusun_contain_produk =models.Penyusun.objects.filter(KodeProduk=produk,Lokasi__NamaLokasi = lokasi).values_list('KodeVersi__id').distinct()

        # Mendapatkan data minus pertama
        tanggalminus = None
        datakuantitasperhitungan = {
            'saldodata': 0,
            'saldofisik':None,
            'datakeluar': 0,
        }
        print(request.GET)
        # print(asd)
        tanggalawalperhitungan = request.GET['tanggalawal']
        tanggalakhirperhitungan = request.GET['tanggalstokopname']
        jumlahaktual = (request.GET['jumlah'])
        # for item in listdata:
        #     sisa = item['Sisa']
        #     for j in sisa:
        #         if j <0 :
        #             sisa_minus_pertama = sisa.index(j)
        #             tanggalminus = item['Tanggal']
        #             datasisaminus = j
        #             lanjut = False
        #             break
        #     if not lanjut:
        #         break
        listdataperhitungan = {}
        print(tanggalminus)
        if tanggalawalperhitungan == "" or tanggalakhirperhitungan == "" or jumlahaktual == "":
            print('Masukkkk')
            return render(
            request,
            "produksi/newkalkulator_penyesuaian.html",
            {
                "kodebarang": request.GET["kodebarang"],
                "nama": nama,
                "satuan": satuan,
                "data": listdata,
                'dataperhitungan' : listdataperhitungan,
                "saldo": saldoawal,
                "tahun": tahun,
                "jumlahartikel": dataajumlahartikel,
                "konversiawal": datakonversiartikel,
                "datakuantitas" : datakuantitasperhitungan,
                'kodeproduk' : kodeproduk,
                'lokasi' : lokasi

            },
        )
            '''
            1. Kumpulkan data perartikel (v)
            2. Cek index dan sisa minus  (v)
            3. Datajumlah artikel tetap 
            4. datakonversiartikel tetap (v)
            '''
        else:
            jumlahaktual =float(jumlahaktual)
            # dataproduksi = models.TransaksiProduksi.objects.filter(Tanggal__range=(tanggal_mulai,tanggalminus),Jenis = 'Mutasi',KodeArtikel__KodeArtikel__in=listartikel)
            dataproduksi = models.TransaksiProduksi.objects.filter(Tanggal__range=(tanggal_mulai,tanggalminus),Jenis = 'Mutasi',VersiArtikel__id__in=listartikel)
            totaljumlahkotakperartikel = dataproduksi.values('KodeArtikel__KodeArtikel').annotate(total=Sum('Jumlah'))
            '''Coba V2'''
            sumproduct = 0
            jumlahkeluar = 0
            jumlahxkonversidictionary = {}

        #    Cek apakah ada data penyesuaian pada tanggal tersebut 
            startdate = datetime.strptime(tanggalawalperhitungan,"%Y-%m-%d")
            enddate = datetime.strptime(tanggalakhirperhitungan,"%Y-%m-%d")
            '''
            CASE Penyesuaian 
            1. Apabila input Tanggal Awal Perhitungan kurang dari tanggalmulai dari penyesuaian 
            2. Apabila input Tanggal Awal perhitungan sama seperti tanggal awal penyesuaian database
            3. Apabila input Tanggal Awal Perhitungan berada pada antara Tanggal mulai dan Tanggal minus dari penyesuaian 

            4. Apabila input tanggal akhir perhitungan kurang dari tanggal 
            '''
            # penyesuaianobj = models.Penyesuaian.objects.filter(KodeProduk=produk,TanggalMulai__range=(startdate,enddate))
            # if penyesuaianobj.exists():
            #     datapenyesuaianawal = penyesuaianobj.order_by('TanggalMulai').first()
            #     messages.warning(request,f"Sudah ada record penyesuai dengan tanggal mulai {datapenyesuaianawal.TanggalMulai}\nTanggal mulai disetting pada {datapenyesuaianawal.TanggalMulai + timedelta(1)}")
            # penyesuaianobj = models.Penyesuaian.objects.filter(KodeProduk=produk,TanggalMinus__range=(startdate,enddate))
            # if penyesuaianobj.exists():
            #     datapenyesuaianawal = penyesuaianobj.order_by('TanggalMulai').first()
            #     messages.warning(request,f"Sudah ada record penyesuai dengan tanggal mulai {datapenyesuaianawal.TanggalMulai}\nTanggal mulai disetting pada {datapenyesuaianawal.TanggalMulai + timedelta(1)}")
            # print(listdata[0])
            listdata,saldoawal = calculate_KSBB(produk,tanggal_mulai,tanggal_akhir,lokasi,True,startdate,enddate)
            print(listdata)
            if len(listdata) == 0:
                messages.warning(request,f'Tidak ditemukan transaksi ')
                return render(
                request,
                "produksi/kalkulator_penyesuaian.html",{"kodeprodukobj": kodeproduk},
)
            # print(asd)
            sisa = [0]

            for item in listdata:
                tanggal = item['Tanggal']
                datetimetanggal = datetime.strptime(tanggal,"%Y-%m-%d")
                if datetimetanggal > enddate:

                    print('tes',datetimetanggal)
                    break
                elif datetimetanggal < startdate:
                    continue
                print(datetimetanggal,startdate)
                if item['Artikel']:
                    print('masuk')
                    for artikel,jumlah,konversi in zip(item['Artikel'],item['Perkotak'],item['Konversi']):
                        sumproduct += jumlah * konversi
                        sisa = item['Sisa']
                        if artikel in jumlahxkonversidictionary:
                            jumlahxkonversidictionary[artikel]['jumlahxkonversi'] += jumlah * konversi
                            jumlahxkonversidictionary[artikel]['jumlah'] += jumlah 
                        else:
                            jumlahxkonversidictionary[artikel] = {'jumlahxkonversi': jumlah*konversi,'jumlah':jumlah}
                
                jumlahkeluar += sum(item['Keluar'])
            print(jumlahxkonversidictionary)
            # print(sisa,item)
            # print(asd)
            saldodata = sisa[-1]
            # print(asd)
            datakuantitasperhitungan["datakeluar"] = jumlahkeluar
            # print(jumlahkeluar)
            datakuantitasperhitungan["saldodata"] = saldodata
            datakuantitasperhitungan['Tanggalminus'] = enddate.strftime("%Y-%m-%d")
            # print(datakuantitasperhitungan)
            # print(adasd)
            
            # Cari jumlah pemusnahan
            jumlahpemusnahanobj = models.PemusnahanBahanBaku.objects.filter(KodeBahanBaku = produk,Tanggal__gt =startdate,Tanggal__lte = enddate).aggregate(total = Sum('Jumlah'))['total']
            if jumlahpemusnahanobj :
                print(jumlahpemusnahanobj)
                jumlahaktual += jumlahpemusnahanobj
            else:
                jumlahpemusnahanobj = 0
            print(jumlahpemusnahanobj)
            print(startdate,enddate)
            # print(asd)
            datakuantitasperhitungan["saldofisik"] = jumlahaktual
            saldoaktual = (jumlahaktual)
            print(datakuantitasperhitungan)
            keluarpenyesuaian = jumlahkeluar -  (saldoaktual - saldodata)
            print(keluarpenyesuaian)
            # print(asd)
            datakonversiakhir = {}

            for key,value in jumlahxkonversidictionary.items():
                try:
                    jumlahpenyesuaian =  keluarpenyesuaian/value['jumlah']
                    productpersumproduct = value['jumlahxkonversi'] / sumproduct
                    konversiakhir = jumlahpenyesuaian*productpersumproduct
                    print(f'Artikel : {key} Jumlah Penyesuaian : {jumlahpenyesuaian} Sumproductperproduct : {productpersumproduct}')
                except ZeroDivisionError:
                    jumlahxkonversidictionary[key].delete()
                datakonversiakhir[key] = {'jumlah': value['jumlah'],'konversiakhir':(konversiakhir)}
            
            print(f'Keluar Penyesuaian : {keluarpenyesuaian} Jumlah Keluar : {jumlahkeluar} Saldo Aktual : {saldoaktual} Saldo Data : {saldodata}')
            print(datakonversiakhir)
            # print(asd)
            listdata,saldoawal = calculate_KSBB(produk,tanggal_mulai,tanggal_akhir,lokasi,True,startdate,enddate,konversibaru=datakonversiakhir)
            print(listdata)
            # print(asd)

            saldoakhirperhari = None
            statusberubah = False
            print(listdata[0])
            # for i,item in enumerate(listdata):
            #     tanggal = item['Tanggal']
            #     datetimetanggal = datetime.strptime(tanggal,"%Y-%m-%d")
            #     if datetimetanggal < startdate:
            #         continue
            #     print(datetimetanggal,startdate)
            #     if item['Artikel']:
            #         print(item)
            #         for x,(artikel,jumlah,konversi) in enumerate(zip(item['Artikel'],item['Perkotak'],item['Konversi'])):
            #             print(artikel,jumlah,konversi)
            #             try:
            #                 konversi = datakonversiakhir[artikel]['konversiakhir']
            #             except KeyError:
            #                 konversi = konversi
            #             listdata[i]['Konversi'][x]= konversi
            #             if saldoakhirperhari == None:
            #                 saldoakhirperhari =  listdata[i]['Sisa'][x] + listdata[i]['Keluar'][x]
            #             listdata[i]['Keluar'][x] = konversi * jumlah
            #             saldoakhirperhari -= listdata[i]['Keluar'][x]
                    
            #         saldoakhirperhari += listdata[i]['Masuk']
            #         listdata[i]['Sisa'][x] = saldoakhirperhari
            #         if len(listdata[i]['Keluar']) != x+1:
            #             j = 1
            #             for pemusnahan in listdata[i]['Keluar'][x+1:]:
            #                 saldoakhirperhari -= pemusnahan
            #             print(item)
            #             print(saldoakhirperhari)
            #             print(x)
            #             listdata[i]['Sisa'][x+j] = saldoakhirperhari

            #         # print(asd)

            #     else:
            #         print(item)
            #         print(saldoakhirperhari)
            #         if saldoakhirperhari == None:
            #             saldoakhirperhari = 0
            #         jumlahkeluar = 0
            #         if listdata[i]['Keluar']:
            #             jumlahkeluar = sum(listdata[i]['Keluar'])
            #         print(jumlahkeluar)
            #         listdata[i]['Sisa'][0] = saldoakhirperhari +  listdata[i]['Masuk']-jumlahkeluar
            #         saldoakhirperhari =listdata[i]['Sisa'][0] 

            #         # print(asd)
                        

                        
            


            # print(sumproduct)
            # print(saldodata)
            # print(jumlahkeluar)
            # print(dataproduksi)
            # print('\n',listdata)
            # print('\njumlahxkonversi : ',jumlahxkonversidictionary)
            # print('\ntotaljumlahkotakperartikel :',totaljumlahkotakperartikel)
            # print(asdas)
            print(tanggalawalperhitungan)
            print(tanggalakhirperhitungan)
            
            return render(
                request,
                "produksi/newkalkulator_penyesuaian.html",
                {
                    "kodebarang": request.GET["kodebarang"],
                    "nama": nama,
                    "satuan": satuan,
                    "data": listdata,
                    'dataperhitungan' : listdataperhitungan,
                    "saldo": saldoawal,
                    "tahun": tahun,
                    "dataaktual" : float(jumlahaktual)-jumlahpemusnahanobj,
                    "jumlahartikel": dataajumlahartikel,
                    "konversiawal": datakonversiartikel,
                    "datakuantitas" : datakuantitasperhitungan,
                    "konversiakhirfix" : datakonversiakhir,
                    'tanggalstokopname' : tanggalakhirperhitungan,
                    'tanggalawal' : tanggalawalperhitungan,
                    'kodeproduk' : kodeproduk,
                                    'lokasi' : lokasi,

                },
            )

# @login_required
# @logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
# def kalkulatorpenyesuaianbahanbaku(request):
#     kodeproduk = models.Produk.objects.all()
#     if len(request.GET) == 0:
#         return render(
#             request,
#             "produksi/kalkulator_penyesuaian.html",
#             {"kodeprodukobj": kodeproduk},
#         )
#     else:
#         """
#         1. Cari 
#         """
#         try:
#             produk = models.Produk.objects.get(KodeProduk=request.GET["kodebarang"])
#             nama = produk.NamaProduk
#             satuan = produk.unit
#         except:
#             messages.error(request, "Data Produk tidak ditemukan")
#             return redirect("view_ksbb")
        
#         if request.GET["periode"]:
#             tahun = int(request.GET["periode"])
#         else:
#             sekarang = datetime.now()
#             tahun = sekarang.year

#         tanggal_mulai = datetime(year=tahun, month=1, day=1)
#         tanggal_akhir = datetime(year=tahun, month=12, day=31)
#         lokasi = request.GET['lokasi']
#         listdata,saldoawal = calculate_KSBB(produk,tanggal_mulai,tanggal_akhir,lokasi)
#         print(listdata) 
#         # print(asd)    

#         datasisaminus = 0
#         datajumlah = 0
#         dataajumlahartikel = {}
#         datakonversiartikel = {}
#         # print(data)

#         firstindex = True

#         listartikel = models.Penyusun.objects.filter(KodeProduk = produk).values_list("KodeArtikel__KodeArtikel",flat=True).distinct()
#         # Mendapatkan data minus pertama
#         sisa_minus_pertama = None
#         tanggalminus = None
#         lanjut = True
#         datakuantitasperhitungan = {
#             'saldodata': 0,
#             'saldofisik':None,
#             'datakeluar': 0,
#         }
#         print(request.GET)
#         # print(asd)
#         tanggalstokopname = request.GET['tanggalstokopname']
#         tanggalawalperhitungan = request.GET['tanggalawal']
#         for item in listdata:
#             sisa = item['Sisa']
#             for j in sisa:
#                 if j <0 :
#                     sisa_minus_pertama = sisa.index(j)
#                     tanggalminus = item['Tanggal']
#                     datasisaminus = j
#                     lanjut = False
#                     break
#             if not lanjut:
#                 break
#         listdataperhitungan = {}
#         print(tanggalminus)
#         if not tanggalminus and tanggalstokopname == "":
#             return render(
#             request,
#             "produksi/newkalkulator_penyesuaian.html",
#             {
#                 "kodebarang": request.GET["kodebarang"],
#                 "nama": nama,
#                 "satuan": satuan,
#                 "data": listdata,
#                 'dataperhitungan' : listdataperhitungan,
#                 "saldo": saldoawal,
#                 "tahun": tahun,
#                 "jumlahartikel": dataajumlahartikel,
#                 "konversiawal": datakonversiartikel,
#                 "datakuantitas" : datakuantitasperhitungan,
#                 'kodeproduk' : kodeproduk,
#                 'lokasi' : lokasi

#             },
#         )
#             '''
#             1. Kumpulkan data perartikel (v)
#             2. Cek index dan sisa minus  (v)
#             3. Datajumlah artikel tetap 
#             4. datakonversiartikel tetap (v)
#             '''
    
            
    
      
#         dataproduksi = models.TransaksiProduksi.objects.filter(Tanggal__range=(tanggal_mulai,tanggalminus),Jenis = 'Mutasi',KodeArtikel__KodeArtikel__in=listartikel)

#         totaljumlahkotakperartikel = dataproduksi.values('KodeArtikel__KodeArtikel').annotate(total=Sum('Jumlah'))

        
#         '''Coba V2'''
#         sumproduct = 0
#         jumlahkeluar = 0
#         jumlahxkonversidictionary = {}
#         datapenyesuaian = models.Penyesuaian.objects.filter(KodeProduk = produk).values_list('TanggalMinus',flat=True).distinct().order_by('-TanggalMinus')
#         print(datapenyesuaian)
#         # print(asd)
#         print(tanggalstokopname)
#         print(tanggalminus)

#         if tanggalstokopname != "":
#             enddate = tanggalstokopname
#         else:
#             enddate = tanggalminus
#         enddate = datetime.strptime(enddate,"%Y-%m-%d")

#         if tanggalawalperhitungan != "":
#             tanggalawalperhitungan = datetime.strptime(tanggalawalperhitungan,"%Y-%m-%d")
#             print('masuk')
#             startdate = tanggalawalperhitungan
#             if datapenyesuaian.exists():
#                 datapenyesuaianawal = datetime.strptime(str(datapenyesuaian[0]),"%Y-%m-%d")
#                 print(tanggalawalperhitungan, datapenyesuaian.first())
#                 if tanggalawalperhitungan < datapenyesuaianawal:
#                     print('tidak valid menggunakan data penyesuaian saat ini') 
#                     startdate = datapenyesuaianawal
#         else:
#             if datapenyesuaian.exists():
#                 datapenyesuaian = datapenyesuaian.filter(TanggalMinus__lte = enddate)
#                 print(datapenyesuaian)
#                 if datapenyesuaian:
#                     startdate =datetime.strptime(str(datapenyesuaian[0]),"%Y-%m-%d")
#                 else:
#                     startdate = tanggal_mulai
#             else:
#                 startdate = tanggal_mulai
#         print(startdate,enddate)
#         # print(listdata)
#         print('\n\n\n\n\n')
#         for item in listdata:

#             # print(asd)
            

#             tanggal = item['Tanggal']

#             datetimetanggal = datetime.strptime(tanggal,"%Y-%m-%d")

#             if datetimetanggal > enddate:
#                 break
#             elif datetimetanggal <= startdate:
#                 continue
#             print(datetimetanggal,startdate)
#             if item['Artikel']:
#                 for artikel,jumlah,konversi in zip(item['Artikel'],item['Perkotak'],item['Konversi']):
#                     sumproduct += jumlah * konversi
#                     sisa = item['Sisa']
#                     if artikel in jumlahxkonversidictionary:
#                         jumlahxkonversidictionary[artikel]['jumlahxkonversi'] += jumlah * konversi
#                         jumlahxkonversidictionary[artikel]['jumlah'] += jumlah 
#                     else:
#                         jumlahxkonversidictionary[artikel] = {'jumlahxkonversi': jumlah*konversi,'jumlah':jumlah}
            
#             jumlahkeluar += sum(item['Keluar'])
#             print(item)
        
#         # print(asdas)
#         print(jumlahxkonversidictionary)
#         # print(asd)
        
#         saldodata = sisa[-1]
#         print(sisa)
#         # print(asd)
#         datakuantitasperhitungan["datakeluar"] = jumlahkeluar
#         print(jumlahkeluar)
#         datakuantitasperhitungan["saldodata"] = saldodata
#         datakuantitasperhitungan['Tanggalminus'] = enddate.strftime("%Y-%m-%d")
#         # print(datasisaminus)
#         # print(adasd)
#         try:
#             dataaktual = int(request.GET["jumlah"])
#         except Exception:
#             return render(
#             request,
#             "produksi/newkalkulator_penyesuaian.html",
#             {
#                 "kodebarang": request.GET["kodebarang"],
#                 "nama": nama,
#                 "satuan": satuan,
#                 "data": listdata,
#                 "saldo": saldoawal,
#                 "tahun": tahun,
#                 "datakuantitas" : datakuantitasperhitungan,
#                 'tanggalstokopname' : tanggalstokopname,
#                 'kodeproduk' : kodeproduk,
#                                 'lokasi' : lokasi


#             },
#         )
#         # Cari jumlah pemusnahan
#         jumlahpemusnahanobj = models.PemusnahanBahanBaku.objects.filter(KodeBahanBaku = produk,Tanggal__gt =startdate,Tanggal__lte = enddate).aggregate(total = Sum('Jumlah'))['total']
#         if jumlahpemusnahanobj :
#             dataaktual += jumlahpemusnahanobj
#         else:
#             jumlahpemusnahanobj = 0
#         print(jumlahpemusnahanobj)
#         print(startdate,enddate)
#         # print(asd)
#         datakuantitasperhitungan["saldofisik"] = dataaktual
#         saldoaktual = dataaktual
#         print(saldoaktual)
#         keluarpenyesuaian = jumlahkeluar -  (saldoaktual - saldodata)
#         datakonversiakhir = {}
#         for key,value in jumlahxkonversidictionary.items():
#             try:
#                 jumlahpenyesuaian =  keluarpenyesuaian/value['jumlah']
#                 productpersumproduct = value['jumlahxkonversi'] / sumproduct
#                 konversiakhir = jumlahpenyesuaian*productpersumproduct
#                 print(f'Artikel : {key} Jumlah Penyesuaian : {jumlahpenyesuaian} Sumproductperproduct : {productpersumproduct}')
#             except ZeroDivisionError:
#                 jumlahxkonversidictionary[key].delete()
#             datakonversiakhir[key] = {'jumlah': value['jumlah'],'konversiakhir':(konversiakhir)}
        
#         print(f'Keluar Penyesuaian : {keluarpenyesuaian} Jumlah Keluar : {jumlahkeluar} Saldo Aktual : {saldoaktual} Saldo Data : {saldodata}')
#         print(datakonversiakhir)
        


#         # print(sumproduct)
#         # print(saldodata)
#         # print(jumlahkeluar)
#         # print(dataproduksi)
#         # print('\n',listdata)
#         # print('\njumlahxkonversi : ',jumlahxkonversidictionary)
#         # print('\ntotaljumlahkotakperartikel :',totaljumlahkotakperartikel)
#         # print(asdas)

        
#         return render(
#             request,
#             "produksi/newkalkulator_penyesuaian.html",
#             {
#                 "kodebarang": request.GET["kodebarang"],
#                 "nama": nama,
#                 "satuan": satuan,
#                 "data": listdata,
#                 'dataperhitungan' : listdataperhitungan,
#                 "saldo": saldoawal,
#                 "tahun": tahun,
#                 "dataaktual" : dataaktual-jumlahpemusnahanobj,
#                 "jumlahartikel": dataajumlahartikel,
#                 "konversiawal": datakonversiartikel,
#                 "datakuantitas" : datakuantitasperhitungan,
#                 "konversiakhirfix" : datakonversiakhir,
#                 'tanggalstokopname' : tanggalstokopname,
#                 'kodeproduk' : kodeproduk,
#                                 'lokasi' : lokasi

#             },
#         )


@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def kalkulatorpenyesuaianartikel(request):
    '''
    Fitur ini digunakan untuk melakukan kalkulasi penyeusian Artikel
    Algoritma
    1. Mendapatkan data Artikel
    2. Menampilkan form input kode artikel 
    3. Program mendapatkan input kode artikel, Tanggal Awal, Tanggal Akhir, dan Jumlah
    4. Memanggil fungsi Calculate_KSBJ
    5. Mengiterasi data hasil Calculate_KSBJ
        A. Apabila tanggal data diantara tanggal awal dan akhir maka melanjutkan program
        B. Menghitung jumlah masuk pada rentang tanggal awal dan akhir
    6. Menghitung selisih dengan cara menghitung Sisa data terakhir pada rentang tanggal awal dan akhir - jumlahaktual
    7. Menghitung jumlah selisih masuk dengan rumus = Jumlah Masuk - Selisih
    8. Menghitung penyesuaian Baru dengan rumus = selisih masuk / jumlah masuk
    9. Mengiterasi data hasil calculate_KSBJ
        A. Mengupdate konversi masuk pada rentang tnaggal awal dan akhir
    '''
    kodeproduk = models.Artikel.objects.all()
    if len(request.GET) == 0:
        return render(
            request,
            "produksi/kalkulator_penyesuaianartikel.html",
            {"kodeprodukobj": kodeproduk},
        )
    else:
        print(request.GET)
        
        """
        1. Cari 
        """
        try:
            Artikelobj = models.Artikel.objects.get(KodeArtikel=request.GET["kodebarang"])
            
        except:
            messages.error(request, "Data Produk tidak ditemukan")
            return redirect("kalkulatorpenyesuaianartikel")
        
        if request.GET["periode"]:
            tahun = int(request.GET["periode"])
        else:
            sekarang = datetime.now()
            tahun = sekarang.year
        try:
            tahun = int(request.GET['periodetahun'])
        except:
            pass
            

        tanggal_mulai = datetime(year=tahun, month=1, day=1)
        tanggal_akhir = datetime(year=tahun, month=12, day=31)
        lokasi = "WIP"
        listdata,saldoawal = calculate_ksbj(Artikelobj,lokasi,tanggal_mulai.year)
        print(listdata,saldoawal) 
        print(request.GET)
        # print(asd)
        tanggalawal = request.GET['tanggalawal']
        tanggalakhir = request.GET['tanggalakhir']
        jumlahaktual = request.GET['jumlah']
        versidefault = models.Versi.objects.filter(KodeArtikel = Artikelobj,isdefault = True).first()
        if tanggalawal == "" or tanggalakhir == "" or jumlahaktual== "":
            return render(
            request,
            "produksi/newkalkulator_penyesuaianartikel.html",
            {
                "kodebarang": request.GET["kodebarang"],
                "data": listdata,
                "saldo": saldoawal,
                "tahun": tahun,
                'lokasi' : lokasi,
                'versidefault':versidefault,
                'kodeprodukobj' : kodeproduk

            },
        
            )
        else:
            '''
            ALGORITMA
            1. Ambil semua data Bahan Baku hasil konversi masuk pertanggal
            2. Ambil selisih antara hasil terakhir pada tanggalakhir dan hasil stok opname
            3. Jumlah hasil konversi dari semua bahan - selisih
            4. Didapat nilai penyesuaian
            5. Nilai penyesuaian dikalikan dengan nilai hasil.
            
            '''
            jumlahmasuk = 0
            i=0
            tanggalawaldatetime =  datetime.strptime(tanggalawal,"%Y-%m-%d")
            tanggalakhirdatetime =  datetime.strptime(tanggalakhir,"%Y-%m-%d")
            # print(tanggalawaldatetime)
            # print(tanggalakhirdatetime)
            for item in listdata:
                tanggaldata = datetime.strptime(item['Tanggal'],"%Y-%m-%d")
                if tanggaldata < tanggalawaldatetime:
                    continue
                elif tanggaldata > tanggalakhirdatetime:
                    break
                else :
                    if i == 0:
                        jumlahawal = item['Masukkonversi']
                        i+=1
                    # print(item)
                    jumlahmasuk += item['Masukkonversi']
                    sisa = item['Sisa']
            selisih = sisa - float(jumlahaktual)
            selisihmasuk = jumlahmasuk - selisih
            print(selisihmasuk,jumlahmasuk)
            # print(asd)
            try:
                penyesuaianbaru = selisihmasuk/jumlahmasuk
            except ZeroDivisionError:
                penyesuaianbaru = 0
            selisihpotong = jumlahmasuk
            
            # penyesuaianbaru = 0.979471329730907

            i = 0
            hasilsisa = jumlahawal  # Inisialisasi `hasilsisa` dengan `jumlahawal`

            for item in listdata:
                tanggaldata = datetime.strptime(item['Tanggal'], "%Y-%m-%d")
                
                if tanggalawaldatetime <= tanggaldata <= tanggalakhirdatetime:
                    # print(item)
                    
                    # Update Masukkonversi dengan pembulatan
                    item['Masukkonversi'] =(item['Masukkonversi'] * penyesuaianbaru)
                    pemusnahan = item['Keluar']
                    
                    if i == 0:
                        # Hitung selisih dan update hasilsisa untuk item pertama
                        selisihhasilawal = item['Masukkonversi'] - jumlahawal
                        hasilsisa = item['Sisa'] + selisihhasilawal - pemusnahan
                        item['Sisa'] = hasilsisa
                    else:
                        # Update Sisa berdasarkan hasilsisa dan Masukkonversi - Hasil
                        hasilsisa = hasilsisa + item['Masukkonversi'] - item['Hasil'] - pemusnahan
                        item['Sisa'] = hasilsisa
                    
                    # print('Masukkonversi:', item['Masukkonversi'])
                    # print('Sisa setelah update:', item['Sisa'])
                    # Update iterasi
                    item['Masukkonversi'] =round(item['Masukkonversi'])
                    item['Sisa'] =round(item['Sisa'])
                    i += 1 
                
                elif tanggaldata < tanggalawaldatetime:
                    continue
                elif tanggaldata > tanggalakhirdatetime:
                    break
            # print(listdata)
            #  {'Tanggal': '2024-07-26', 'Masuklembar': 28.0, 'Masukkonversi': 2090, 'Hasil': 2100, 'Sisa': 71711}
            # print(selisih,selisihmasuk,jumlahmasuk,penyesuaianbaru,sisa,jumlahaktual)
            return render(
            request,
            "produksi/newkalkulator_penyesuaianartikel.html",
            {
                "kodebarang": request.GET["kodebarang"],
                "data": listdata,
                "saldo": saldoawal,
                "tahun": tahun,
                'lokasi' : lokasi,
                'tanggalawal' : tanggalawal,
                'tanggalakhir' : tanggalakhir,
                'dataaktual' : jumlahaktual,
                'penyesuaian' : penyesuaianbaru,
                'kodeprodukobj':kodeproduk,
                'versidefault' : versidefault


            },)





# Saldo Awal Bahan Baku
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_saldobahan(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data saldo awal bahan baku produksi
    Algoritma
    1. Mendapatkan data saldoawal bahan baku dengan kriteria Nama Lokasi = WIP/FG
    2. Menampilkan data saldo awal bahan baku
    '''
    dataproduk = models.SaldoAwalBahanBaku.objects.filter(IDLokasi__NamaLokasi__in=("WIP","FG")).order_by("-Tanggal")
    for i in dataproduk:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_saldobahan.html", {"dataproduk": dataproduk}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_saldobahan(request):
    '''
    Fitur ini digunakan untuk menambahkan data saldo awal bahan baku pada sistem 
    Algoritma
    1. Mendapatkan data semua bahan baku
    2. Menampilkan form input data saldo awal bahan baku 
    3. Program mendapatkan input berupa Kode bahan baku, Lokasi, Jumlah, tanggal
    4. Apabila tahun pada tanggal yang diinput sudah memiliki kode stok tersebut maka akan menampilkan pesan error
    5. menyimpan data saldo awal bahan baku
    '''
    databarang = models.Produk.objects.all()
    datalokasi = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_saldobahan.html",
            {"nama_lokasi": datalokasi[:2], "databarang": databarang},
        )
    else:
        print(request.POST)
        # print(asd)
        tanggal = request.POST.get("tanggal")
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")

        produk_list = request.POST.getlist("produk[]")
        lokasi_list = request.POST.getlist("nama_lokasi[]")
        jumlah_list = request.POST.getlist("jumlah[]")

        for kodeproduk, lokasi, jumlah in zip(produk_list, lokasi_list, jumlah_list):
            lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)
            try:
                # Periksa apakah entri sudah ada
                existing_entry = models.SaldoAwalBahanBaku.objects.filter(
                    Tanggal__year=tanggal_formatted.year,
                    IDBahanBaku__KodeProduk=kodeproduk,
                    IDLokasi=lokasi
                ).exists()
                if existing_entry:
                    messages.warning(
                        request,
                        f"Data untuk Bahan Baku {kodeproduk} di lokasi {lokasiobj.NamaLokasi} pada tahun {tanggal_formatted.year} sudah ada."
                    )
                    continue

                # Validasi produk
                produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)

                # Validasi lokasi

                # Simpan data
                saldo_bahan_obj = models.SaldoAwalBahanBaku(
                    Tanggal=tanggal,
                    Jumlah=jumlah,
                    IDBahanBaku=produkobj,
                    IDLokasi=lokasiobj,
                    Harga=0
                )
                saldo_bahan_obj.save()
                messages.success(request, f"Data {produkobj.KodeProduk} berhasil disimpan")

                # Log transaksi
                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Saldo Awal Bahan Baku. Kode Produk: {produkobj.KodeProduk}, Lokasi: {lokasiobj.NamaLokasi}, Jumlah: {jumlah}, Harga: 0"
                ).save()

            except models.Produk.DoesNotExist:
                messages.error(
                    request, f"Bahan Baku {kodeproduk} tidak ditemukan dalam sistem."
                )
            except models.Lokasi.DoesNotExist:
                messages.error(
                    request, f"Lokasi {lokasi} tidak ditemukan dalam sistem."
                )

        # messages.success(request, "Data berhasil diproses.")
        return redirect("view_saldobahanproduksi")

    # return redirect("add_saldobahan")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_saldobahan(request, id):
    '''
    Fitur ini digunakan untuk melakukan update saldo awal bahan baku pada sistem
    Algoritma
    1. Mendapatkan data semua produk
    2. Mendapatkan data saldo awal bahan baku dengan kriteria IDSaldoAwalBahanBaku = id (id didapatkan dari passing values HTML)
    3. menampilkan form update saldo awal bahan baku
    4. program menerima input berupa kode bahan baku, Lokasi, Jumlah, dan Tanggal
    5. Mengupdate data saldo awal bahan baku
    '''
    databarang = models.Produk.objects.all()
    dataobj = models.SaldoAwalBahanBaku.objects.get(IDSaldoAwalBahanBaku=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/update_saldobahan.html",
            {"data": dataobj, "nama_lokasi": lokasiobj[:2],"databarang": databarang},
        )

    else:
        kodeproduk = request.POST["produk"]
        lokasi = request.POST["nama_lokasi"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]

        # Ubah format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")
        # Periksa apakah entri sudah ada
        existing_entry = models.SaldoAwalBahanBaku.objects.filter(
            Tanggal__year=tanggal_formatted.year,
            IDBahanBaku__KodeProduk=kodeproduk,
            IDLokasi=lokasi
        ).exclude(IDSaldoAwalBahanBaku=id).exists()

        if existing_entry:
            # Jika sudah ada, beri tanggapan atau lakukan tindakan yang sesuai
            messages.warning(request,('Sudah ada Entry pada tahun',tanggal_formatted.year))
            return redirect("update_saldobahan",id=id)
        try :
            produkobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except :
            messages.error(request,f'Bahan Baku {kodeproduk} tidak ditemukan dalam sistem')
            return redirect("update_saldobahan", id=id)
        lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.IDBahanBaku = produkobj
        dataobj.IDLokasi = lokasiobj
        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Saldo Bahan Baku. Kode Bahan Baku : {produkobj.KodeProduk} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi} ",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_saldobahanproduksi")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_saldobahan(request, id):
    '''
    Fitur ini digunakan untuk menghapus data saldo awal bahan baku pada sistem
    Algoritma 
    1. Mengambil data saldo awal bahan baku dengan kriteria IDSaldoAwalBahanBaku = id (id didapatkan dari passing values HTML)
    2. Menghapus data saldo awal bahan baku
    '''
    dataobj = models.SaldoAwalBahanBaku.objects.get(IDSaldoAwalBahanBaku=id)

    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Saldo Bahan Baku. Kode Bahan Baku : {dataobj.IDBahanBaku.KodeProduk} Jumlah : {dataobj.Jumlah} Lokasi : {dataobj.IDLokasi.NamaLokasi} Harga : {dataobj.Harga}",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect(view_saldobahan)


# Saldo Awal Artikel
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_saldoartikel(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data saldo awal Artikel produksi
    Algoritma
    1. Mendapatkan data saldoawal Artikel 
    2. Menampilkan data saldo awal Artikel
    '''
    dataartikel = models.SaldoAwalArtikel.objects.all().order_by("-Tanggal")
    for i in dataartikel:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_saldoartikel.html", {"dataartikel": dataartikel}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_saldoartikel(request):
    '''
    Fitur ini digunakan untuk menambahkan data saldo awal Artikel pada sistem 
    Algoritma
    1. Mendapatkan data semua Artikel
    2. Menampilkan form input data saldo awal Artikel 
    3. Program mendapatkan input berupa Kode Artikel, Lokasi, Jumlah, tanggal
    4. Apabila tahun pada tanggal yang diinput sudah memiliki kode stok tersebut maka akan menampilkan pesan error
    5. menyimpan data saldo awal Artikel
    '''
    dataartikel = models.Artikel.objects.all()
    datalokasi = models.Lokasi.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_saldoartikel.html",
            {"nama_lokasi": datalokasi[:2], "dataartikel": dataartikel},
        )
    else:
        tanggal = request.POST["tanggal"]
        artikel_list = request.POST.getlist("artikel[]")
        lokasi_list = request.POST.getlist("nama_lokasi[]")
        jumlah_list = request.POST.getlist("jumlah[]")

        # Format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")

        for artikel, lokasi, jumlah in zip(artikel_list, lokasi_list, jumlah_list):
            # Periksa apakah entri sudah ada
            existing_entry = models.SaldoAwalArtikel.objects.filter(
                Tanggal__year=tanggal_formatted.year,
                IDArtikel__KodeArtikel=artikel,
                IDLokasi=lokasi,
            ).exists()

            if existing_entry:
                # Jika sudah ada, beri pesan peringatan
                messages.warning(request, f"Sudah ada entri untuk artikel {artikel} di lokasi {lokasi} pada tahun {tanggal_formatted.year}.")
                continue  # Lewati penyimpanan untuk artikel ini

            try:
                artikelobj = models.Artikel.objects.get(KodeArtikel=artikel)
            except models.Artikel.DoesNotExist:
                messages.error(request, f"Tidak ditemukan data artikel dengan kode {artikel}.")
                continue

            try:
                lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)
            except models.Lokasi.DoesNotExist:
                messages.error(request, f"Tidak ditemukan lokasi dengan ID {lokasi}.")
                continue

            # Simpan data ke dalam database
            saldo_awal_obj = models.SaldoAwalArtikel(
                Tanggal=tanggal,
                Jumlah=jumlah,
                IDArtikel=artikelobj,
                IDLokasi=lokasiobj,
            )
            saldo_awal_obj.save()
            messages.success(request, f"Data {artikelobj.KodeArtikel} - {tanggal_formatted.year} berhasil disimpan.")

            # Catat ke dalam log transaksi
            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Saldo Artikel. Kode Artikel: {artikelobj.KodeArtikel}, Jumlah: {jumlah}, Lokasi: {lokasiobj.NamaLokasi}.",
            ).save()

        # Berikan pesan sukses jika ada data yang berhasil disimpan
        return redirect("view_saldoartikel")

    # Render halaman form jika bukan POST
    # return render(request, "produksi/add_saldoartikel.html")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_saldoartikel(request, id):
    '''
    Fitur ini digunakan untuk melakukan update saldo awal Artikel pada sistem
    Algoritma
    1. Mendapatkan data semua produk
    2. Mendapatkan data saldo awal Artikel dengan kriteria IDSaldoAwalBahanBaku = id (id didapatkan dari passing values HTML)
    3. menampilkan form update saldo awal Artikel
    4. program menerima input berupa kode Artikel, Lokasi, Jumlah, dan Tanggal
    5. Mengupdate data saldo awal Artikel
    '''
    dataartikel = models.Artikel.objects.all()
    dataobj = models.SaldoAwalArtikel.objects.get(IDSaldoAwalBahanBaku=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    lokasiobj = models.Lokasi.objects.all()
    if request.method == "GET":

        return render(
            request,
            "produksi/update_saldoartikel.html",
            {"data": dataobj, "nama_lokasi": lokasiobj[:2], "dataartikel": dataartikel},
        )

    else:
        artikel = request.POST["artikel"]
        lokasi = request.POST["nama_lokasi"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]

        # Ubah format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")
        # Periksa apakah entri sudah ada
        existing_entry = models.SaldoAwalArtikel.objects.filter(
            Tanggal__year=tanggal_formatted.year,
            IDArtikel__KodeArtikel=artikel,
            IDLokasi=lokasi
        ).exclude(IDSaldoAwalBahanBaku=id).exists()

        if existing_entry:
            # Jika sudah ada, beri tanggapan atau lakukan tindakan yang sesuai
            messages.warning(request,('Sudah ada Entry pada tahun',tanggal_formatted.year))
            return redirect("update_saldoartikel",id=id)
        try:
            artikelobj = models.Artikel.objects.get(KodeArtikel=artikel)
            
        except:
            messages.error(request,f'Data artikel {artikel} tidak ditemukan dalam sistem')
            return redirect("update_saldoartikel",id=id)
        lokasiobj = models.Lokasi.objects.get(IDLokasi=lokasi)

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.IDArtikel = artikelobj
        dataobj.IDLokasi= lokasiobj

        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Saldo Artikel. Kode Bahan Baku : {artikelobj.KodeArtikel} Jumlah : {jumlah} Lokasi : {lokasiobj.NamaLokasi}",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_saldoartikel")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_saldoartikel(request, id):
    '''
    Fitur ini digunakan untuk menghapus data saldo awal artikel pada sistem
    Algoritma 
    1. Mengambil data saldo awal artikel dengan kriteria IDSaldoAwalArtikel = id (id didapatkan dari passing values HTML)
    2. Menghapus data saldo awal artikel
    '''
    dataobj = models.SaldoAwalArtikel.objects.get(IDSaldoAwalBahanBaku=id)

    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Saldo Artikel. Kode Bahan Baku : {dataobj.IDArtikel.KodeArtikel} Jumlah : {dataobj.Jumlah} Lokasi : {dataobj.IDLokasi.NamaLokasi}",
    ).save()
    messages.success(request,'Data berhasil dihapus')

    return redirect(view_saldoartikel)


# Saldo Awal Produk Subkon
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_saldosubkon(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data saldo awal Produk Subkon produksi
    Algoritma
    1. Mendapatkan data saldoawal Produk Subkon 
    2. Menampilkan data saldo awal Produk Subkon
    '''
    datasubkon = models.SaldoAwalSubkon.objects.all().order_by("-Tanggal")
    for i in datasubkon:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_saldosubkon.html", {"datasubkon": datasubkon}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_saldosubkon(request):
    '''
    Fitur ini digunakan untuk menambahkan data saldo awal Produk Subkon pada sistem 
    Algoritma
    1. Mendapatkan data semua Produk Subkon
    2. Menampilkan form input data saldo awal Produk Subkon 
    3. Program mendapatkan input berupa Kode Produk Subkon, Lokasi, Jumlah, tanggal
    4. Apabila tahun pada tanggal yang diinput sudah memiliki kode stok tersebut maka akan menampilkan pesan error
    5. menyimpan data saldo awal Produk Subkon
    '''
    datasubkon = models.ProdukSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_saldosubkon.html",
            { "datasubkon": datasubkon},
        )
    else:
        print(request.POST)
        # print(asd)
        kodebarang_hidden_list = request.POST.getlist("kodebarangHidden[]")
        jumlah_list = request.POST.getlist("jumlah[]")
        tanggal = request.POST.get("tanggal")

        # Ubah format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")
        log_entries = []  # Untuk menyimpan log transaksi yang akan dibuat

        for kodeproduk, jumlah in zip(kodebarang_hidden_list, jumlah_list):
            try:
                # Periksa apakah entri sudah ada
                existing_entry = models.SaldoAwalSubkon.objects.filter(
                    Tanggal__year=tanggal_formatted.year,
                    IDProdukSubkon__pk=kodeproduk,
                ).exists()

                produkobj = models.ProdukSubkon.objects.get(IDProdukSubkon=kodeproduk)
                if existing_entry:
                    messages.warning(
                        request,
                        f"Sudah ada entry pada tahun {tanggal_formatted.year} untuk produk dengan ID {produkobj.NamaProduk} artikel {produkobj.KodeArtikel}.",
                    )
                    continue

                # Ambil objek produk

                # Simpan saldo awal
                pemusnahanobj = models.SaldoAwalSubkon(
                    Tanggal=tanggal_formatted,
                    Jumlah=jumlah,
                    IDProdukSubkon=produkobj,
                )
                pemusnahanobj.save()
                messages.success(request,f"Data {produkobj.NamaProduk} artikel {produkobj.KodeArtikel} berhasil disimpan dalam sistem")

                # Tambahkan ke log transaksi
                log_entries.append(
                    f"Saldo Produk Subkon. Nama Produk: {produkobj.NamaProduk}, "
                    f"Kode Artikel: {produkobj.KodeArtikel}, Jumlah: {jumlah}"
                )

            except models.ProdukSubkon.DoesNotExist:
                messages.error(
                    request,
                    f"Produk dengan ID {kodeproduk} tidak ditemukan dalam sistem.",
                )
                continue

        # Simpan semua log transaksi
        for log_entry in log_entries:
            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=log_entry,
            ).save()

        
        return redirect("view_saldosubkonproduksi")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_saldosubkon(request, id):
    '''
    Fitur ini digunakan untuk melakukan update saldo awal Produk Subkon pada sistem
    Algoritma
    1. Mendapatkan data semua produk
    2. Mendapatkan data saldo awal Produk Subkon dengan kriteria IDSaldoAwalProdukSubkon = id (id didapatkan dari passing values HTML)
    3. menampilkan form update saldo awal Produk Subkon
    4. program menerima input berupa kode Produk Subkon, Jumlah, dan Tanggal
    5. Mengupdate data saldo awal Produk Subkon
    '''
    dataobj = models.SaldoAwalSubkon.objects.get(IDSaldoAwalProdukSubkon=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    datasubkon = models.ProdukSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/update_saldosubkon.html",
            {"data": dataobj,"datasubkon": datasubkon },
        )

    else:
        kodeproduk = request.POST["kodebarangHidden"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]

        # Ubah format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")
        # Periksa apakah entri sudah ada
        existing_entry = models.SaldoAwalSubkon.objects.filter(
            Tanggal__year=tanggal_formatted.year,
            IDProdukSubkon__NamaProduk=kodeproduk,
        ).exclude(IDSaldoAwalProdukSubkon=id).exists()
          
        if existing_entry:
            # Jika sudah ada, beri tanggapan atau lakukan tindakan yang sesuai
            messages.warning(request,('Sudah ada Entry pada tahun',tanggal_formatted.year))
            return redirect("view_saldosubkon")
        
        produkobj = models.ProdukSubkon.objects.get(IDProdukSubkon=kodeproduk)

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.IDProdukSubkon = produkobj
        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Saldo Produk Subkon. Nama Produk : {produkobj.NamaProduk} Kode Artikel : {produkobj.KodeArtikel} Jumlah : {jumlah}",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_saldosubkonproduksi")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_saldosubkon(request, id):
    '''
    Fitur ini digunakan untuk menghapus data saldo awal Produk Subkon pada sistem
    Algoritma 
    1. Mengambil data saldo awal Produk Subkon dengan kriteria IDSaldoAwalProdukSubkon = id (id didapatkan dari passing values HTML)
    2. Menghapus data saldo awal Produk Subkon
    '''
    dataobj = models.SaldoAwalSubkon.objects.get(IDSaldoAwalProdukSubkon=id)

    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Saldo Produk Subkon. Nama Produk : {dataobj.IDProdukSubkon.NamaProduk}  Kode Artikel : {dataobj.IDProdukSubkon.KodeArtikel} Jumlah : {dataobj.Jumlah}",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect('view_saldosubkonproduksi')


# Saldo Bahan Subkon
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_saldobahansubkon(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data saldo awal Bahan Baku Subkon produksi
    Algoritma
    1. Mendapatkan data saldoawal Bahan Baku Subkon 
    2. Menampilkan data saldo awal Bahan Baku Subkon
    '''
    datasubkon = models.SaldoAwalBahanBakuSubkon.objects.all().order_by("-Tanggal")
    for i in datasubkon:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_saldobahansubkon.html", {"datasubkon": datasubkon}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_saldobahansubkon(request):
    '''
    Fitur ini digunakan untuk menambahkan data saldo awal Bahan Baku Subkon pada sistem 
    Algoritma
    1. Mendapatkan data semua Bahan Baku Subkon
    2. Menampilkan form input data saldo awal Bahan Baku Subkon 
    3. Program mendapatkan input berupa Kode Bahan Baku Subkon, Lokasi, Jumlah, tanggal
    4. Apabila tahun pada tanggal yang diinput sudah memiliki kode stok tersebut maka akan menampilkan pesan error
    5. menyimpan data saldo awal Bahan Baku Subkon
    '''
    datasubkon = models.BahanBakuSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/add_saldobahansubkon.html",
            { "datasubkon": datasubkon},
        )
    else:
        kodeproduk_list = request.POST.getlist("produk[]")
        jumlah_list = request.POST.getlist("jumlah[]")
        tanggal = request.POST["tanggal"]

        # Ubah format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")
        errors = []

        for kodeproduk, jumlah in zip(kodeproduk_list, jumlah_list):
            # Validasi apakah entri sudah ada
            existing_entry = models.SaldoAwalBahanBakuSubkon.objects.filter(
                Tanggal__year=tanggal_formatted.year,
                IDBahanBakuSubkon__KodeProduk=kodeproduk,
            ).exists()

            if existing_entry:
                errors.append(f"Data untuk kode produk {kodeproduk} pada tahun {tanggal_formatted.year} sudah ada.")
                continue

            # Validasi apakah kode produk ditemukan
            try:
                produkobj = models.BahanBakuSubkon.objects.get(KodeProduk=kodeproduk)
            except models.BahanBakuSubkon.DoesNotExist:
                errors.append(f"Kode produk {kodeproduk} tidak ditemukan.")
                continue

            # Simpan entri baru
            pemusnahanobj = models.SaldoAwalBahanBakuSubkon(
                Tanggal=tanggal_formatted,
                Jumlah=jumlah,
                IDBahanBakuSubkon=produkobj,
            )
            pemusnahanobj.save()
            messages.success(request,f'Data {produkobj.KodeProduk} - {tanggal_formatted.year} berhasil disimpan')

            # Catat transaksi
            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Saldo Bahan Baku Subkon. Kode Bahan Baku: {produkobj.KodeProduk}, Jumlah: {jumlah}",
            ).save()

        # Berikan pesan sesuai hasil
        if errors:
            for error in errors:
                messages.error(request, error)

        return redirect("view_saldobahansubkonproduksi")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_saldobahansubkon(request, id):
    '''
    Fitur ini digunakan untuk melakukan update saldo awal Bahan Baku Subkon pada sistem
    Algoritma
    1. Mendapatkan data semua produk
    2. Mendapatkan data saldo awal Bahan Baku Subkon dengan kriteria IDSaldoAwalBahan BakuSubkon = id (id didapatkan dari passing values HTML)
    3. menampilkan form update saldo awal Bahan Baku Subkon
    4. program menerima input berupa kode Bahan Baku Subkon, Jumlah, dan Tanggal
    5. Mengupdate data saldo awal Bahan Baku Subkon
    '''
    dataobj = models.SaldoAwalBahanBakuSubkon.objects.get(IDSaldoAwalBahanBakuSubkon=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    datasubkon = models.BahanBakuSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/update_saldobahansubkon.html",
            {"data": dataobj,"datasubkon": datasubkon },
        )

    else:
        kodeproduk = request.POST["produk"]
        jumlah = request.POST["jumlah"]
        tanggal = request.POST["tanggal"]

        # Ubah format tanggal menjadi YYYY-MM-DD
        tanggal_formatted = datetime.strptime(tanggal, "%Y-%m-%d")
        # Periksa apakah entri sudah ada
        existing_entry = models.SaldoAwalBahanBakuSubkon.objects.filter(
            Tanggal__year=tanggal_formatted.year,
            IDBahanBakuSubkon__KodeProduk=kodeproduk,
        ).exclude(IDSaldoAwalBahanBakuSubkon=id).exists()

        if existing_entry:
            # Jika sudah ada, beri tanggapan atau lakukan tindakan yang sesuai
            messages.warning(request,('Sudah ada Entry pada tahun',tanggal_formatted.year))
            return redirect("update_saldobahansubkonproduksi", id=id)
        try:
            produkobj = models.BahanBakuSubkon.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request,f"Kode Bahan Baku Subkon {kodeproduk} tidak ditemukan dalam sistem")
            return redirect("update_saldobahansubkonproduksi", id = id)

        dataobj.Tanggal = tanggal
        dataobj.Jumlah = jumlah
        dataobj.IDBahanBakuSubkon = produkobj
        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Saldo Bahan Baku Subkon. Kode Bahan Baku: {produkobj.KodeProduk} Jumlah : {jumlah}",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("view_saldobahansubkonproduksi")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_saldobahansubkon(request, id):
    '''
    Fitur ini digunakan untuk menghapus data saldo awal BahanBaku Subkon pada sistem
    Algoritma 
    1. Mengambil data saldo awal BahanBaku Subkon dengan kriteria IDSaldoAwalBahanBakuSubkon = id (id didapatkan dari passing values HTML)
    2. Menghapus data saldo awal BahanBaku Subkon
    '''
    dataobj = models.SaldoAwalBahanBakuSubkon.objects.get(IDSaldoAwalBahanBakuSubkon=id)

    dataobj.delete()

    models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Delete",
            pesan=f"Saldo Bahan Baku Subkon. Kode Bahan Baku: {dataobj.IDBahanBakuSubkon.KodeProduk} Jumlah : {dataobj.Jumlah}",
        ).save()
    messages.success(request,'Data berhasil dihapus')

    return redirect(view_saldobahansubkon)


# Keterangan Bahan Baku
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def read_bahanbaku(request):
    '''
    Fitur ini digunakan untuk melihat semua data bahan baku pada sistem
    Algoritma
    1. Mendapatkan data semua data Produk 
    2. Menampilkan pada sistem
    '''
    produkobj = models.Produk.objects.all()
    return render(request, "produksi/read_produk.html", {"produkobj": produkobj})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_produk_produksi(request, id):
    '''
    FItur ini digunakan untuk mengupdate data produk pada sistem
    Algoritma
    1. Mendapatkan data produk dengan kriteria KodeProduk = id (id didapatkan dari passing values HTML)
    2. Menampilkan form update bahan baku 
    3. Program mendapatkan input berupa Keterangan Produk 
    4. Mengupdate data bahan baku
    '''
    produkobj = models.Produk.objects.get(KodeProduk=id)
    if request.method == "GET":
        return render(request, "produksi/update_produk.html", {"produkobj": produkobj})
    else:
        keterangan_produk = request.POST["keterangan_produk"]
        produkobj.keteranganProduksi = keterangan_produk
        produkobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Bahan Baku. Kode Bahan Baku: {produkobj.KodeProduk} Nama Bahan Baku : {produkobj.NamaProduk}  Keterangan : {produkobj.keteranganProduksi}",
        ).save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("read_produk_produksi")


'''SUBKON SECTION'''
# Bahan Baku SUBKON
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def read_bahansubkon(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data bahan baku subkon
    Algoritma
    1. Mendapatkan semua data bahan baku subkon
    2. Menampilkan data bahan baku subkon
    '''
    produkobj = models.BahanBakuSubkon.objects.all()
    return render(request, "produksi/read_bahansubkon.html", {"produkobj": produkobj})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def create_bahansubkon(request):
    '''
    Fitur ini digunakan untuk menambahkan data bahan baku subkon pada sistem
    Algoritma
    1. Menampilkan form input bahan baku subkon
    2. Program mendapatkan input berupa Kode stok , Nama , dan unit dari bahan baku subkon
    3. Menyimpan dalam sistem
    '''
    if request.method == "GET":
        return render(
            request, "produksi/create_bahansubkon.html")
    else:
        kode_produk = request.POST["kode_produk"]
        nama_produk = request.POST["nama_produk"]
        unit_produk = request.POST["unit_produk"]

        databahan = models.BahanBakuSubkon.objects.filter(KodeProduk=kode_produk).exists()
        
        if databahan:
            messages.error(request, "Kode Produk sudah ada")
            return redirect("create_bahansubkon")
        else:
            new_produk = models.BahanBakuSubkon(
                KodeProduk = kode_produk,
                NamaProduk = nama_produk,
                unit = unit_produk,
            )
            new_produk.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Bahan Baku Subkon. Kode Bahan Baku: {kode_produk} Nama Bahan Baku : {nama_produk}  Unit : {unit_produk}",
            ).save()
            messages.success(request,"Data berhasil disimpan")
            return redirect("read_bahansubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_bahansubkon(request, id):
    '''
    Fitur ini digunakan untuk mengupdate data bahan baku subkon 
    Algoritma
    1. Mendapatkan data bahan baku subkon dengan kriteria pk (primary key) = id (id didapatkan dari passing values HTML)
    2. Menampilkan form update bahan baku subkon 
    3. Program mendapatkan input berupa kode stok, nama, dan unit bahan baku subkon
    '''
    produkobj = models.BahanBakuSubkon.objects.get(pk=id)

    if request.method == "GET":
        return render(
            request,
            "produksi/update_bahansubkon.html",
            {"produkobj": produkobj},
        )
    else:
        kode_produk = request.POST["kode_produk"]
        nama_produk = request.POST["nama_produk"]
        unit_produk = request.POST["unit_produk"]

        databahan = models.BahanBakuSubkon.objects.filter(KodeProduk=kode_produk).exclude(id=id).exists()
        
        if databahan:
            messages.error(request, "Kode Bahan Baku Subkon sudah ada")
            return redirect("update_bahansubkon",id=id)
        else:
            produkobj.KodeProduk = kode_produk
            produkobj.NamaProduk = nama_produk
            produkobj.unit = unit_produk

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"Bahan Baku Subkon. Kode Bahan Baku: {kode_produk} Nama Bahan Baku : {nama_produk}  Unit : {unit_produk}",
            ).save()

        produkobj.save()
        messages.success(request,"Data berhasil diupdate")
        return redirect("read_bahansubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_bahansubkon(request, id):
    '''
    Fitur ini digunakan untuk menghapus data bahan baku subkon 
    Algoritma :
    1. Mendapatkan data bahan baku subkon dengan kriteria id = id (id didapatkan dari passing values HTML)
    2. Menghapus data 
    '''
    produkobj = models.BahanBakuSubkon.objects.get(id=id)
    produkobj.delete()
    messages.success(request, "Data Berhasil dihapus")

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Bahan Baku Subkon. Kode Bahan Baku: {produkobj.KodeProduk} Nama Bahan Baku : {produkobj.NamaProduk}  Unit : {produkobj.unit}",
    ).save()
    
    return redirect("read_bahansubkon")


# Produk SUBKON
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def read_produksubkon(request):
    '''
    Fitur ini digunakan untuk melihat semua data Produk Subkon pada sistem
    Algoritma
    1. Mendapatkan data semua data Produk 
    2. Menampilkan pada sistem
    '''
    produkobj = models.ProdukSubkon.objects.all()
    return render(request, "produksi/read_produksubkon.html", {"produkobj": produkobj})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def create_produksubkon(request):
    '''
    Fitur ini digunakan untuk menambahkan data produk subkon pada sistem
    Algoritma
    1. Menampilkan form input produk subkon
    2. Program mendapatkan input berupa Kode stok , kode artikel peruntukan, Nama , dan unit dari bahan baku subkon
    3. Menyimpan dalam sistem
    '''
    kodeartikel = models.Artikel.objects.all()
    if request.method == "GET":
        return render(
            request, "produksi/create_produksubkon.html", {"kodeartikel": kodeartikel}
        )
    else:
        kode_produk = request.POST["kode_produk"]
        nama_produk = request.POST["nama_produk"]
        unit_produk = request.POST["unit_produk"]
        keterangan_produk = request.POST["keterangan_produk"]

        try:
            artikelobj = models.Artikel.objects.get(KodeArtikel=kode_produk)
        except models.Artikel.DoesNotExist:
            messages.error(request, "Kode Artikel Peruntukan tidak ditemukan")
            return redirect("create_produksubkon")
        
        listkodeproduk = (
            models.ProdukSubkon.objects.filter(KodeArtikel=artikelobj.id)
            .values_list("NamaProduk", flat=True)
            .distinct()
        )

        if nama_produk in listkodeproduk:
            messages.error(
                request, "Nama Produk untuk Artikel terkait sudah ada pada Database"
            )
            return redirect("create_produksubkon")
        else:
            new_produk = models.ProdukSubkon(
                NamaProduk=nama_produk,
                Unit=unit_produk,
                KodeArtikel=artikelobj,
                keterangan=keterangan_produk,
            )
            new_produk.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Produk Subkon. Nama Produk: {nama_produk} Artikel Peruntukan : {artikelobj.KodeArtikel}  Unit : {unit_produk} Keterangan : {keterangan_produk}",
            ).save()
            messages.success(request,'Data berhasil disimpan')
            return redirect("read_produksubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_produksubkon(request, id):
    '''
    Fitur ini digunakan untuk mengupdate data produk subkon 
    Algoritma
    1. Mendapatkan data produk subkon dengan kriteria pk (primary key) = id (id didapatkan dari passing values HTML)
    2. Menampilkan form update produk subkon 
    3. Program mendapatkan input berupa kode stok, nama, dan unit produk subkon
    '''
    produkobj = models.ProdukSubkon.objects.get(pk=id)
    dataartikel = models.Artikel.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/update_produksubkon.html",
            {"produkobj": produkobj, "dataartikel": dataartikel},
        )
    else:
        kode_produk = request.POST["kode_produk"]
        nama_produk = request.POST["nama_produk"]
        unit_produk = request.POST["unit_produk"]
        keterangan_produk = request.POST["keterangan_produk"]

        try:
            artikelobj = models.Artikel.objects.get(KodeArtikel=kode_produk)
        except models.Artikel.DoesNotExist:
            messages.error(request, "Kode Artikel Peruntukan tidak ditemukan")
            return redirect("update_produksubkon",id = id)
        
        listkodeproduk = (
            models.ProdukSubkon.objects.filter(KodeArtikel=artikelobj.id)
            .values_list("NamaProduk", flat=True).exclude(IDProdukSubkon=id)
            .distinct()
        )

        if nama_produk in listkodeproduk:
            messages.error(
                request, "Nama Produk untuk Artikel terkait sudah ada pada Database"
            )
            return redirect("update_produksubkon",id= id)
        else:
            produkobj.KodeArtikel= artikelobj
            produkobj.NamaProduk = nama_produk
            produkobj.Unit = unit_produk
            produkobj.keterangan = keterangan_produk
            produkobj.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"Produk Subkon. Nama Produk: {nama_produk} Artikel Peruntukan : {artikelobj.KodeArtikel}  Unit : {unit_produk} Keterangan : {keterangan_produk}",
            ).save()

        return redirect("read_produksubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_produksubkon(request, id):
    '''
    Fitur ini digunakan untuk menghapus data produk subkon 
    Algoritma :
    1. Mendapatkan data produk subkon dengan kriteria IDProdukSubkon = id (id didapatkan dari passing values HTML)
    2. Menghapus data 
    '''
    produkobj = models.ProdukSubkon.objects.get(IDProdukSubkon=id)
    produkobj.delete()
    messages.success(request, "Data Berhasil dihapus")

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Produk Subkon. Nama Produk: {produkobj.NamaProduk} Artikel Peruntukan : {produkobj.KodeArtikel.KodeArtikel}  Unit : {produkobj.Unit} Keterangan : {produkobj.keterangan}",
    ).save()

    return redirect("read_produksubkon")


# Surat Jalan Kirim Subkon
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_subkonbahankeluar(request):
    '''
    Fitur ini digunakan untuk pengiriman bahan baku subkon menuju vendor
    Algoritma 
    1. mendapatkan data detail surat jalan pengiriman bahan baku subkon 
    2. menampilkan pada sistem
    '''
    datasubkon = models.DetailSuratJalanPengirimanBahanBakuSubkon.objects.all().order_by("NoSuratJalan__Tanggal")
    for i in datasubkon:
        i.NoSuratJalan.Tanggal = i.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")

    return render(request, "produksi/view_subkonbahankeluar.html", {"datasubkon": datasubkon})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_subkonbahankeluar(request):
    '''
    Fitur ini digunakan untuk menambahkan data surat jalan pengiriman bahan baku subkon ke vendor
    Algoritma.
    1. Mendapatkan data detail surat jalan pengiriman bahan baku subkon 
    2. Mendapatkan data surat jalan pengiriman bahan baku subkon
    3. Mendapatkan data bahan baku subkon 
    4. menampilkan form input surat jalan pengiriman bahan baku subkon
    5. program mendapatkan input berupa no surat jalan, tanggal,list kode bahan baku subkon, list jumlah, dan list keterangan
    6. cek apakah no surat jalan subkon sudah ada pada sistem atau belum. Apabila ada maka akan menampilkan pesan error
    7. Menyimpan data surat jalan subkon pada sistem
    8. mengiterasi list kode bahan baku subko, list jumlah dan list keterangan
        A. membuat data detail transaksi bahan baku subkon pada surat jalan bahan baku subkon terakhir
    '''
    if request.method == "GET":
        subkonkirim = models.DetailSuratJalanPengirimanBahanBakuSubkon.objects.all()
        detailsk = models.SuratJalanPengirimanBahanBakuSubkon.objects.all()
        getproduk = models.BahanBakuSubkon.objects.all()

        return render(
            request,
            "produksi/add_subkonbahankeluar.html",
            {"subkonkirim": subkonkirim, "detailsk": detailsk, "getproduk": getproduk},
        )
    if request.method == "POST":
        nosuratjalan = request.POST["nosuratjalan"]
        tanggal = request.POST["tanggal"]

        datasj = models.SuratJalanPengirimanBahanBakuSubkon.objects.filter(NoSuratJalan=nosuratjalan).exists()
        if datasj:
            messages.error(request, "No Surat Jalan sudah ada")
            return redirect("add_subkonbahankeluar")
        else:
            subkonkirimobj = models.SuratJalanPengirimanBahanBakuSubkon(NoSuratJalan=nosuratjalan, Tanggal=tanggal)
            subkonkirimobj.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"SJ Kirim Bahan Subkon. No Surat Jalan: {nosuratjalan}" ,
            ).save()

            subkonkirimobj = models.SuratJalanPengirimanBahanBakuSubkon.objects.get(NoSuratJalan=nosuratjalan)

            listkode = request.POST.getlist("kodeproduk")
            listjumlah = request.POST.getlist("jumlah")
            listket = request.POST.getlist("keterangan")
            detail_added = False
            for kodeproduk, jumlah, keterangan in zip(listkode, listjumlah, listket):
                try:
                    bahanobj = models.BahanBakuSubkon.objects.get(KodeProduk=kodeproduk)
                except models.BahanBakuSubkon.DoesNotExist :
                    messages.error(request,f"Data Bahan Baku {kodeproduk} tidak ditemukan dalam sistem")    
                    continue
                newprodukobj = models.DetailSuratJalanPengirimanBahanBakuSubkon(
                    KodeBahanBaku = bahanobj,
                    Jumlah=jumlah,
                    Keterangan=keterangan,
                    NoSuratJalan=subkonkirimobj,
                )
                newprodukobj.save()
                detail_added = True
        
                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail SJ Kirim Bahan Subkon. Kode Bahan Baku: {bahanobj.KodeProduk} Nama Bahan Baku : {bahanobj.NamaProduk}  Jumlah : {jumlah} Keterangan : {keterangan}",
                ).save()
            if not detail_added:
                # If no detail is added, delete the main Surat Jalan object
                subkonkirimobj.delete()
                messages.error(request,"Transaksi Gagal, cek kembali kode bahan baku subkom")
                return redirect("add_subkonbahankeluar")
            messages.success(request,"Data berhasil disimpan")
            return redirect("view_subkonbahankeluar")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_subkonbahankeluar(request, id):
    '''
    Fitur ini digunakan untuk melakukan update surat jalan pengiriman bahan baku subkon 
    Algoritma
    1. MEndapatkan data Detail Surat Jalan pengiriman bahan baku subkon dengan kriteria IDDEtailSJPengriimanSubkon = id (id diapatkan dari passing values HTML)
    2. Mendapatkan data objec surat jalan pengiriman bahan baku subkon 
    3. mendapatkan semua data bahan baku subkon 
    4. menampilkan form input transaksi pengiriman bahan baku subkon 
    5. Program mendapatkan input berupa No Surat Jalan, Tanggal, Kode Produk, Jumlah dan Keterangan
    6. Mengupdate data suratjalan dan detail surat jalan pengiriman bahan baku subkon 
    '''
    datasjp = models.DetailSuratJalanPengirimanBahanBakuSubkon.objects.get(IDDetailSJPengirimanSubkon=id)

    datasjp_getobj = models.SuratJalanPengirimanBahanBakuSubkon.objects.get(
        NoSuratJalan = datasjp.NoSuratJalan.NoSuratJalan
    )
    getproduk = models.BahanBakuSubkon.objects.all()

    if request.method == "GET":
        return render(
            request,
            "produksi/update_subkonbahankeluar.html",
            {
                "datasjp": datasjp_getobj,
                "detailsjp" :datasjp,
                "tanggal": datetime.strftime(datasjp_getobj.Tanggal, "%Y-%m-%d"),
                "getproduk": getproduk
            },
        )
    else:
        nosuratjalan = request.POST["nosuratjalan"]
        tanggal = request.POST["tanggal"]
        kode_produk = request.POST["kodeproduk"]
        try:
            kode_produkobj = models.BahanBakuSubkon.objects.get(KodeProduk=kode_produk)
        except:
            messages.error(request,f"Data Bahan Baku {kode_produk} tidak ditemukan dalam sistem")    
            return redirect("update_subkonbahankeluar",id)
        jumlah = request.POST["jumlah"]
        keterangan = request.POST["keterangan"]

        datasj = models.SuratJalanPengirimanBahanBakuSubkon.objects.filter(NoSuratJalan=nosuratjalan).exclude(NoSuratJalan=datasjp.NoSuratJalan.NoSuratJalan).exists()
        print(datasj)
        if datasj:
            messages.error(request, "No Surat Jalan sudah ada")
            return redirect("update_subkonbahankeluar",id)
        else:

            datasjp.NoSuratJalan.NoSuratJalan = nosuratjalan
            datasjp.NoSuratJalan.Tanggal = tanggal


            datasjp.NoSuratJalan.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"SJ Kirim Bahan Subkon. No Surat Jalan: {nosuratjalan}" ,
            ).save()

            datasjp.KodeBahanBaku = kode_produkobj
            datasjp.Jumlah = jumlah
            datasjp.Keterangan = keterangan
            
            datasjp.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"Detail SJ Kirim Bahan Subkon. Kode Bahan Baku: {kode_produkobj.KodeProduk} Nama Bahan Baku : {kode_produkobj.NamaProduk}  Jumlah : {jumlah} Keterangan : {keterangan}",
            ).save()

            listkode = request.POST.getlist('kodeproduk[]')
            listjumlah = request.POST.getlist('jumlah[]')
            listket = request.POST.getlist('keterangan[]')

            for kodeproduk, jumlah, keterangan in zip(listkode, listjumlah, listket):
                bahanobj = models.BahanBakuSubkon.objects.get(KodeProduk=kodeproduk)
                newprodukobj = models.DetailSuratJalanPengirimanBahanBakuSubkon(
                    KodeBahanBaku = bahanobj,
                    Jumlah=jumlah,
                    Keterangan=keterangan,
                    NoSuratJalan=datasjp_getobj,
                )
                newprodukobj.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail SJ Kirim Bahan Subkon. Kode Bahan Baku: {bahanobj.KodeProduk} Nama Bahan Baku : {bahanobj.NamaProduk}  Jumlah : {jumlah} Keterangan : {keterangan}",
                ).save()
            messages.success(request,"Data berhasil disimpan")
            return redirect("view_subkonbahankeluar")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_subkonbahankeluar(request, id):
    '''
    Fitur ini digunakan untuk menghapus data detail surat jalan pengiriman bahan baku subkon 
    ALgoritma
    1. Mengambil data detail surat jalan pengiriman bahan baku subkon dengan kondeisi IDDetailSJPengirimanSubkon = id (id didapatkan dari passing values HTML)
    2. cek apakah memiliki detail surat jalan pengiriman lebih dari 1 ?
    3. APabila tidak ada maka akan menghapus data surat jalan pengiriman bahan baku subkon (terhubung kedalam detail suratjalan )
    4. Apabila ada data lebih dari 1 maka akan menghapus detail surat jalan pengiriman bahan baku subkon
    '''
    dataskk = models.DetailSuratJalanPengirimanBahanBakuSubkon.objects.get(IDDetailSJPengirimanSubkon=id)
    kodesuratjalan = dataskk.NoSuratJalan
    # dataskk.delete()
    ceksuratjalan = models.DetailSuratJalanPengirimanBahanBakuSubkon.objects.filter(NoSuratJalan = kodesuratjalan)
    ceksuratjalan = ceksuratjalan.exclude(pk = dataskk.pk)
    print(ceksuratjalan)
    # print(asd)
    if not ceksuratjalan.exists() :
        suratjalanobj = models.SuratJalanPengirimanBahanBakuSubkon.objects.get(NoSuratJalan = kodesuratjalan.NoSuratJalan)
        suratjalanobj.delete()
    else:
        dataskk.delete()
        


    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Detail SJ Kirim Bahan Subkon. Kode Bahan Baku: {dataskk.KodeBahanBaku.KodeProduk} Nama Bahan Baku : {dataskk.KodeBahanBaku.NamaProduk}  Jumlah : {dataskk.Jumlah} Keterangan : {dataskk.Keterangan}",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect("view_subkonbahankeluar")


# Surat Jalan Terima Subkon
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_subkonprodukmasuk(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data produk subkon yang masuk kedalam sistem
    Algoritma:
    1. Mendapatkan data detail surat jalan penerimaan produk subkon 
    2. Menampilkan data pada sistem
    '''
    datasubkon = models.DetailSuratJalanPenerimaanProdukSubkon.objects.all().order_by("NoSuratJalan__Tanggal")
    for i in datasubkon:
        i.NoSuratJalan.Tanggal = i.NoSuratJalan.Tanggal.strftime("%Y-%m-%d")

    return render(request, "produksi/view_subkonprodukmasuk.html", {"datasubkon": datasubkon})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_subkonprodukmasuk(request):
    '''
    Fitur ini digunakan untuk menambahkan data produk subkon yang masuk kedalam perusahaan
    Algoritma:
    1. Mendapatkan data produk subkon 
    2. Menampilkan form input data surat jalan penerimaan produk subkon
    3. Program mendapatkan inputan berupa No Surat Jalan, Tanggal, Supplier, list produk subkon, list jumlah, list keterangan
    4. Cek apakah sudah ada data No Srurat Jalan yang diinput pada sistem sebelumnya ? Apabila ada maka akan menampilkan pesan error
    5. melakukan iterasi pada list produk subkon untuk melihat apakah ada error produk subkon yang tidak terdaftar dalam sistem 
    6. membuat object surat jalan penerimaan produk subkon dengan kriteria No Surat Jalan = input no suratjalan, Tanggal = input tanggal, dan Supplier = input supplier
    7. Melakukan iterasi pada List Produk Subkon, List Jumlah, dan List Keterangan produk subkon
        A. Membuat Object Detail Surat Jalan Penerimaan Produk Subkon dengan kriteria Kode Produk = produksubkon iterasi, Jumlah = jumlah iterasi, dan keterangan = keterangan iterasi
        B. Menyimpan detail surat jalan penerimaan produk subkon
    '''
    if request.method == "GET":
        subkonkirim = models.DetailSuratJalanPenerimaanProdukSubkon.objects.all()
        detailsk = models.SuratJalanPenerimaanProdukSubkon.objects.all()
        getproduk = models.ProdukSubkon.objects.all().order_by('KodeArtikel__KodeArtikel','NamaProduk')

        return render(
            request,
            "produksi/add_subkonprodukmasuk.html",
            {"subkonkirim": subkonkirim, "detailsk": detailsk, "getproduk": getproduk},
        )
    
    if request.method == "POST":
        print(request.POST)
        # print(asd)
        nosuratjalan = request.POST["nosuratjalan"]
        tanggal = request.POST["tanggal"]
        supplier = request.POST['supplier']

        datasj = models.SuratJalanPenerimaanProdukSubkon.objects.filter(NoSuratJalan=nosuratjalan).exists()
        if datasj:
            messages.error(request, "No Surat Jalan sudah ada")
            return redirect("add_subkonprodukmasuk")
        else:
            listkode = request.POST.getlist("produk")
            listjumlah = request.POST.getlist("jumlah")
            listket = request.POST.getlist("keterangan")
            error = 0
            for item in listkode:
                try:
                    produksubkonobj = models.ProdukSubkon.objects.get(
                        IDProdukSubkon=item
                    )
                except :
                    error +=1
            if error == len(listkode):
                messages.error(request,"Kode Produk Subkon tidak terdapat dalam sistem")
                return redirect('add_subkonprodukmasuk')
            subkonkirimobj = models.SuratJalanPenerimaanProdukSubkon(NoSuratJalan=nosuratjalan, Tanggal=tanggal,Supplier = supplier)
            subkonkirimobj.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"SJ Terima Produk Subkon. No Surat Jalan: {nosuratjalan}" ,
            ).save()

            subkonkirimobj = models.SuratJalanPenerimaanProdukSubkon.objects.get(NoSuratJalan=nosuratjalan)



            for kodeproduk, jumlah, keterangan in zip(listkode, listjumlah, listket):

                try:
                    produksubkonobj = models.ProdukSubkon.objects.get(
                        IDProdukSubkon=kodeproduk
                    )
                except :
                    messages.error(request, f"Kode Produk Subkon {kodeproduk}tidak ditemukan")
                    continue
                
                newprodukobj = models.DetailSuratJalanPenerimaanProdukSubkon(
                    KodeProduk = produksubkonobj,
                    Jumlah=jumlah,
                    Keterangan=keterangan,
                    NoSuratJalan=subkonkirimobj,
                )
                newprodukobj.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail SJ Terima Produk Subkon. Nama Produk : {produksubkonobj.NamaProduk} Artikel Untuk : {produksubkonobj.KodeArtikel}  Jumlah : {jumlah} Keterangan : {keterangan}",
                ).save()
            messages.success(request,'Data Berhasil Disimpan')
            return redirect("view_subkonprodukmasuk")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_subkonprodukmasuk(request, id):
    '''
    Fitur ini digunakan untuk melakukan update data surat jalan penerimaan produk subkon
    Algoritma 
    1. Mengambil detail surat jalan penerimaan produk subkon dengan kriteria IDDetailSJPenerimaanSubkon = id (id didapatkan dari passing values HTML)
    2. Mendapatkan data Surat jalan penerimaan produk subkon
    3. Menampilkan data form update detail surat jalan produk subkon
    4. Program menerima input berupa update berupa No Surat Jalan, Tanggal, Produk, Supplier, Jumlah dan keterangan
    5. Cek apakah kode produk subkon ada pada sistem ? Jika tidak maka akan menampilkan pesan error
    6. Mengupdate data
    '''
    datasjp = models.DetailSuratJalanPenerimaanProdukSubkon.objects.get(IDDetailSJPenerimaanSubkon=id)

    datasjp_getobj = models.SuratJalanPenerimaanProdukSubkon.objects.get(
        NoSuratJalan = datasjp.NoSuratJalan.NoSuratJalan
    )
    getproduk = models.ProdukSubkon.objects.all().order_by('KodeArtikel__KodeArtikel','NamaProduk')

    if request.method == "GET":
        return render(
            request,
            "produksi/update_subkonprodukmasuk.html",
            {
                "datasjp": datasjp_getobj,
                "detailsjp" :datasjp,
                "tanggal": datetime.strftime(datasjp_getobj.Tanggal, "%Y-%m-%d"),
                "getproduk": getproduk
            },
        )
    else:
        nosuratjalan = request.POST["nosuratjalan"]
        tanggal = request.POST["tanggal"]
        kode_produk = request.POST["produk"]
        
        supplier = request.POST['supplier']
        print(request.POST)
        # print(asd)
        try:
            produksubkonobj = models.ProdukSubkon.objects.get(IDProdukSubkon=kode_produk)

        except models.ProdukSubkon.DoesNotExist:
            messages.error(request, "Kode Produk Subkon tidak ditemukan")
            return redirect("transaksi_subkon_terima")
    
        jumlah = request.POST["jumlah"]
        keterangan = request.POST["keterangan"]

        datasj = models.SuratJalanPenerimaanProdukSubkon.objects.filter(NoSuratJalan=nosuratjalan).exclude(NoSuratJalan=datasjp.NoSuratJalan.NoSuratJalan).exists()
        if datasj:
            messages.error(request, "No Surat Jalan sudah ada")
            return redirect("update_subkonprodukmasuk",id)
        else:
            datasjp.NoSuratJalan.NoSuratJalan = nosuratjalan
            datasjp.NoSuratJalan.Tanggal = tanggal
            datasjp.NoSuratJalan.Supplier = supplier
            datasjp.NoSuratJalan.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"SJ Terima Produk Subkon. No Surat Jalan: {nosuratjalan}" ,
            ).save()

            datasjp.KodeProduk = produksubkonobj
            datasjp.Jumlah = jumlah
            datasjp.Keterangan = keterangan
            
            datasjp.save()

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Update",
                pesan=f"Detail SJ Terima Produk Subkon. Nama Produk : {produksubkonobj.NamaProduk} Artikel Untuk : {produksubkonobj.KodeArtikel}  Jumlah : {jumlah} Keterangan : {keterangan}",
            ).save()

            listkode = request.POST.getlist("kodebarangHidden")
            listjumlah = request.POST.getlist("jumlah[]")
            listket = request.POST.getlist("keterangan[]")

            for kodeproduk, jumlah, keterangan in zip(listkode, listjumlah, listket):

                try:
                    produksubkon = models.ProdukSubkon.objects.get(
                        IDProdukSubkon=kodeproduk
                    )

                except models.ProdukSubkon.DoesNotExist:
                    messages.error(request, "Kode Produk Subkon tidak ditemukan")
                    return redirect("transaksi_subkon_terima")
                
                newprodukobj = models.DetailSuratJalanPenerimaanProdukSubkon(
                    KodeProduk = produksubkon,
                    Jumlah=jumlah,
                    Keterangan=keterangan,
                    NoSuratJalan=datasjp_getobj,
                )
                newprodukobj.save()

                models.transactionlog(
                    user="Produksi",
                    waktu=datetime.now(),
                    jenis="Create",
                    pesan=f"Detail SJ Terima Produk Subkon. Nama Produk : {produksubkon.NamaProduk} Artikel Untuk : {produksubkon.KodeArtikel}  Jumlah : {jumlah} Keterangan : {keterangan}",
                ).save()

            messages.success(request,'Data berhasil diupdate')
            return redirect("view_subkonprodukmasuk")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_subkonprodukmasuk(request, id):
    '''
    Fitur ini digunakan untuk menghapus detail surat jalan penerimaan produk subkon
    Algoritma : 
    1. Ambil data detail surat jalan penerimaan produk subkon dengan kriteria IDDetailaSJPenerimaanSubkon = id (id didapatkan dari passing values HTML)
    2. Cek apakah ada DetailSuratJalanPenerimaanProdukSubkon lagi selain object terpilih poin 1 pada Surat Jalan Penerimaan Produk Subkon
    3. Apabila Ada Maka hapus detailsuratjalanpenerimaanproduksubkon
    4. Apabila tidak ada maka hapus Surat Jalan Penerimaan Produk Subkon dengan No Surat Jalan yang sama pada object detail surat jalan penerimaan produk subkon poin 1

    '''
    dataskk = models.DetailSuratJalanPenerimaanProdukSubkon.objects.get(IDDetailSJPenerimaanSubkon=id)
    kodesuratjalan = dataskk.NoSuratJalan

    ceksuratjalan = models.DetailSuratJalanPenerimaanProdukSubkon.objects.filter(NoSuratJalan = kodesuratjalan)
    ceksuratjalan = ceksuratjalan.exclude(pk = dataskk.pk)

    if not ceksuratjalan.exists() :
        suratjalanobj = models.SuratJalanPenerimaanProdukSubkon.objects.get(NoSuratJalan = kodesuratjalan.NoSuratJalan)
        suratjalanobj.delete()
    else:
        dataskk.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Detail SJ Terima Produk Subkon. Nama Produk : {dataskk.KodeProduk.NamaProduk} Artikel Untuk : {dataskk.KodeProduk.KodeArtikel}  Jumlah : {dataskk.Jumlah} Keterangan : {dataskk.Keterangan}",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect("view_subkonprodukmasuk")


# Transaksi subkon bahan baku masuk
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def transaksi_subkonbahan_masuk(request):
    '''
    Fitur ini digunakan untuk manajemen data Transaksi Bahan Baku Subkon Masuk dari Produksi (dari Kode Stok Bahan Baku menajdi Kode stok bahan baku subkon)
    Algoritma
    1. Mengamil data transakasi bahan baku subkon 
    2. Menampilkan data pada sistem
    '''

    produkobj = models.TransaksiBahanBakuSubkon.objects.all().order_by("-Tanggal")
    for i in produkobj:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")
    return render(
        request, "produksi/read_transaksisubkonbahan_masuk.html", {"produkobj": produkobj}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def create_transaksi_subkonbahan_masuk(request):
    '''
    Fitur ini diguanakan untuk menambahkan data transaksi bahan baku masuk subkon pada sistem
    Algoritma
    1. Mengambil semua data bahan baku subkon 
    2. Menampilkan form input data transaksi subkon bahan baku masuk
    3. Program mendapatkan input berupa List Nama Kode, List Jumlah, List Keterangan 
    4. Mengiterasi list Nama Kode Bahan Baku, List Jumlah, List Keterangan
        A. Membuat Object Transaki bahan baku subkon dengan kriteria Tanggal = tanggal iterasi, Jumlah = jumlah iterasi, Kode Bahan Baku = kode bahan baku subkon iterasi, keterangan = keterangan iterasi
        B. Menyimpan data pada sistem
    '''
    produksubkon = models.BahanBakuSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/create_transaksisubkonbahan_masuk.html",
            {"produksubkon": produksubkon},
        )
    else:
        tanggal = request.POST["tanggal"]
        list_nama_kode = request.POST.getlist("nama_produk[]")
        listjumlah = request.POST.getlist("jumlah[]")
        listketerangan = request.POST.getlist("keterangan[]")

        for nama_kode, jumlah, keterangan in zip(list_nama_kode,listjumlah, listketerangan):
            try:
                produksubkonobj = models.BahanBakuSubkon.objects.get(
                    KodeProduk=nama_kode
                    )
            except models.BahanBakuSubkon.DoesNotExist:
                messages.error(request,f'Kode Produk Subkon {nama_kode} tidak ditemukan')
                continue

            new_produk = models.TransaksiBahanBakuSubkon(
                Tanggal=tanggal,
                Jumlah=jumlah,
                KodeBahanBaku=produksubkonobj,
                Keterangan=keterangan
            )
            new_produk.save()
            messages.success(request, "Data berhasil disimpan")

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Transaksi Bahan Baku Subkon. Kode Bahan Baku: {produksubkonobj.KodeProduk} Nama Bahan Baku : {produksubkonobj.NamaProduk}  Jumlah : {jumlah} Keterangan : {keterangan}",
            ).save()
        return redirect("transaksi_subkonbahan_masuk")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_transaksi_subkonbahan_masuk(request, id):
    '''
    Fitur ini digunakan untuk melakukan update transaksi subkon bahan baku masuk pada sistem 
    Algoritma
    1. Mendapatkan data transaksi bahan baku subkon dengan kriteraia pk (primary key) = id (id didapatkan dari passing values HTML)
    2. menampilkan form update transaksi bahan baku subkon 
    3. Program mendapatkan input berupa Jumlah, kode bahan baku, keterangan, dan tanggal 
    4. Mengupdate data 
    '''
    produkobj = models.TransaksiBahanBakuSubkon.objects.get(pk=id)
    produkobj.Tanggal = produkobj.Tanggal.strftime("%Y-%m-%d")
    produksubkon = models.BahanBakuSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/update_transaksisubkonbahan_masuk.html",
            {"produkobj": produkobj, "produksubkon": produksubkon},
        )
    else:
        jumlah = request.POST["jumlah"]
        nama_kode = request.POST["nama_produk"]
        keterangan = request.POST["keterangan"]
        tanggal = request.POST["tanggal"]
        try:
            produksubkonobj = models.BahanBakuSubkon.objects.get(KodeProduk=nama_kode)
        except models.BahanBakuSubkon.DoesNotExist:
            messages.error(request,"Kode produk subkon tidak ditemukan")
            return redirect('update_transaksi_subkonbahan_masuk',id=id)
        produkobj.KodeBahanBaku = produksubkonobj
        produkobj.Jumlah = jumlah
        produkobj.Tanggal = tanggal
        produkobj.Keterangan = keterangan

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Transaksi Bahan Baku Subkon. Kode Bahan Baku: {produksubkonobj.KodeProduk} Nama Bahan Baku : {produksubkonobj.NamaProduk}  Jumlah : {jumlah} Keterangan : {keterangan}",
        ).save()

        produkobj.save()
        messages.success(request,'Data berhasil disimpan')
        return redirect("transaksi_subkonbahan_masuk")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_transaksi_subkonbahan_masuk(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi subkon masuk
    Algoritma
    1. Mendapatkan data transaksi bahan baku subkon dengan kreiteria IDTransaksiBahanBakuSubkon = id (id didapatkan dari passing valeus HTML)
    2. Menghapus data transaksi bahan baku subkon 
    '''
    produkobj = models.TransaksiBahanBakuSubkon.objects.get(IDTransaksiBahanBakuSubkon=id)
    produkobj.delete()
    messages.success(request, "Data Berhasil dihapus")

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Transaksi Bahan Baku Subkon. Kode Bahan Baku: {produkobj.KodeBahanBaku.KodeProduk} Nama Bahan Baku : {produkobj.KodeBahanBaku.NamaProduk}  Jumlah : {produkobj.Jumlah} Keterangan : {produkobj.Keterangan}",
    ).save()
    
    return redirect("transaksi_subkonbahan_masuk")

# Tramsalsi Produksi Produk Subkon

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def views_produksiproduksubkon(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data transaksi produksi produk subkon pada perusahaan
    ALgoritma
    1. Mendapatkan data transaksi produksi produk subkon 
    2. Menampilkan pada user
    '''
    datasubkon = models.TransaksiProduksiProdukSubkon.objects.all().order_by("-Tanggal")
    for i in datasubkon:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(request, "produksi/view_produksiproduksubkon.html", {"datasubkon": datasubkon})

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_produksiproduksubkon(request):
    '''
    FItur ini digunakan unutk menambahkan data produksi produk subkon pada sistem
    Algoritma
    1. Mendapatkan data produk subkon 
    2. MEnampilkan form input data transaksi produksi produk subkon
    3. Program mendapatkan input Tanggal, list kode barang, list jumlah, dan list keterangan
    4. Mengiterasi List Kode Barang, List Jumlah, List Keterangan
        A. Menyimpan data transaksi produksi produk subkon
    '''
    if request.method == "GET":
        subkonkirim = models.DetailSuratJalanPenerimaanProdukSubkon.objects.all()
        detailsk = models.SuratJalanPenerimaanProdukSubkon.objects.all()
        getproduk = models.ProdukSubkon.objects.all()

        return render(
            request,
            "produksi/add_produksiproduksubkon.html",
            {"subkonkirim": subkonkirim, "detailsk": detailsk, "getproduk": getproduk},
        )
    
    if request.method == "POST":
        print(request.POST)
        # print(asd)
        tanggal = request.POST['tanggal']
        listkode = request.POST.getlist("kodebarangHidden")
        listjumlah = request.POST.getlist("jumlah")
        listket = request.POST.getlist("keterangan")
        error = 0
        for item in listkode:
            try:
                produksubkonobj = models.ProdukSubkon.objects.get(
                    IDProdukSubkon=item
                )
            except :
                error +=1
        if error == len(listkode):
            messages.error(request,"Kode Produk Subkon tidak terdapat dalam sistem")
            return redirect('add_produksiproduksubkon')

        for kode,jumlah,keterangan in zip(listkode,listjumlah,listket):
            try:
                produkobj = models.ProdukSubkon.objects.get(pk = kode)
            except:
                messages.error(request,f'Produk')
                continue

            dataobj = models.TransaksiProduksiProdukSubkon(
                Tanggal = tanggal,
                KodeProduk = produkobj,
                Jumlah = jumlah,
                Keterangan = keterangan
            )
            dataobj.save()
            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Produksi Produk Subkon. Kode Produk : {produkobj}, Jumlah : {jumlah}, Keterangan : {keterangan}" ,
            ).save()

        messages.success(request,'Data Berhasil Disimpan')
        return redirect("view_produksiproduksubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_produksiproduksubkon(request, id):
    '''
    Fitur ini digunakan untuk melakukan update transaksi produksi produk subkon pada sistem
    Algoritma
    1. Mendapatkan data transaksi produksi produk subkon dengan kriteria pk (primary key) = id (id didapatkan dari passing values HTML)
    2. Menampilkan form update pada user
    3. Program menerima inputan berupa Tanggal, Kode produk, Jumlah, dan Keterangan 
    4. Mengupdate data transaksi produksi produk subkon
    '''
    produkobj = models.TransaksiProduksiProdukSubkon.objects.get(pk = id)
    produkobj.Tanggal = produkobj.Tanggal.strftime('%Y-%m-%d')
    getproduk = models.ProdukSubkon.objects.all()


    if request.method == "GET":
        return render(
            request,
            "produksi/update_produksiproduksubkon.html",
            {
                "getproduk": getproduk,
                'produk' : produkobj
            },
        )
    else:
        print(request.POST)
        tanggal = request.POST["tanggal"]
        kode_produk = request.POST["kodebarangHiddens"]
        jumlah = request.POST['jumlah']
        keterangan = request.POST['keterangan']
        try:
            produksubkonobj = models.ProdukSubkon.objects.get(IDProdukSubkon=kode_produk)

        except models.ProdukSubkon.DoesNotExist:
            messages.error(request, "Kode Produk Subkon tidak ditemukan")
            return redirect("update_produksiproduksubkon", id = id)


        

        produkobj.KodeProduk = produksubkonobj
        produkobj.Jumlah = jumlah
        produkobj.Keterangan = keterangan
        
        produkobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Detail SJ Terima Produk Subkon. Nama Produk : {produksubkonobj.NamaProduk} Artikel Untuk : {produksubkonobj.KodeArtikel}  Jumlah : {jumlah} Keterangan : {keterangan}",
        ).save()
        messages.success(request,'Data berhasil disimpan')

        return redirect("view_produksiproduksubkon")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_produksiproduksubkon(request, id):
    '''
    Fitur ini digunakan untuk menghapus transaksi produksi produk subkon 
    Algoritma
    1. Mendapatkan data transaksi produksi produk subkon 
    2. Menghapus data
    '''
    dataskk = models.TransaksiProduksiProdukSubkon.objects.get(pk=id)
    dataskk.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Transaksi Produksi Produk Subkon. Nama Produk : {dataskk.KodeProduk.NamaProduk} Artikel Untuk : {dataskk.KodeProduk.KodeArtikel}  Jumlah : {dataskk.Jumlah} Keterangan : {dataskk.Keterangan}",
    ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect("view_produksiproduksubkon")


# Transaksi subkon produk keluar
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def transaksi_subkon_terima(request):
    '''
    Fitur ini digunakan untuk mengeluarkan produk subkon untuk keperluan Produksi pada WIP
    Algoritma
    1. Mengambil data transaksi subkon 
    2. Menampilkan data 
    '''
    produkobj = models.TransaksiSubkon.objects.all().order_by("-Tanggal")
    for i in produkobj:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")
    return render(
        request, "produksi/read_transaksisubkon_terima.html", {"produkobj": produkobj}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def create_transaksi_subkon_terima(request):
    '''
    Fitur ini digunakan untuk menambahkan transaksi produk subkon keluar
    Algoritma
    1. Mengambil semua data produk subkon 
    2. Menampilkan form input produk subkon 
    3. Program mendapatkan input berupa Tanggal, list kode produk subkon, list Jumlah, dan list keterangan
    4. Mengiterasi List Kode Produk Subkon, List Jumlah, dan List Keterangan
        A. Membuat object Transaksi Subkon dengan kriteria Tanggal = tanggal iterasi, Jumlah = jumlah iterasi, KodeProduk = produksubkon iterasi, Keterangan = keterangan iterasi
        B. Menyimpan object Transaksi Subkon 
    '''
    produksubkon = models.ProdukSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/create_transaksisubkon_terima.html",
            {"produksubkon": produksubkon},
        )
    else:
        tanggal = request.POST["tanggal"]
        list_nama_kode = request.POST.getlist("kodebarangHidden")
        listjumlah = request.POST.getlist("jumlah[]")
        listketerangan = request.POST.getlist('keterangan')

        print(list_nama_kode)

        for nama_kode, jumlah,keterangan in zip(list_nama_kode,listjumlah,listketerangan):

            try:
                produksubkonobj = models.ProdukSubkon.objects.get(IDProdukSubkon=nama_kode)

            except models.ProdukSubkon.DoesNotExist:
                messages.error(request, "Kode Produk Subkon tidak ditemukan")
                return redirect("transaksi_subkon_terima")
            
            new_produk = models.TransaksiSubkon(
                Tanggal=tanggal,
                Jumlah=jumlah,
                KodeProduk=produksubkonobj,
                Keterangan = keterangan
            )
            new_produk.save()
            messages.success(request, "Data berhasil disimpan")

            models.transactionlog(
                user="Produksi",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Transaksi Produk Subkon. Nama Produk: {produksubkonobj.NamaProduk} Artikel Peruntukan : {produksubkonobj.KodeArtikel}  Jumlah : {jumlah}",
            ).save()

        return redirect("transaksi_subkon_terima")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_transaksi_subkon_terima(request, id):
    '''
    FItur ini digunakan untuk melakukan update data transaksi subkon terima
    Algoritma
    1. Mendapatkan data TransaksiSubkon dengan kriteria pk (primary key) = id (id didapatkan dari passing values HTML)
    2. menampilkan form update transaksi subkon 
    3. Program menerima jumlah, Kode Produk, Tanggal, dan Keterangan 
    4.  Mengupdate data 
    '''
    produkobj = models.TransaksiSubkon.objects.get(pk=id)
    produkobj.Tanggal = produkobj.Tanggal.strftime("%Y-%m-%d")
    produksubkon = models.ProdukSubkon.objects.all()
    if request.method == "GET":
        return render(
            request,
            "produksi/update_transaksisubkon_terima.html",
            {"produkobj": produkobj, "produksubkon": produksubkon},
        )
    else:
        jumlah = request.POST["jumlah"]
        nama_kode = request.POST["kodebarangHidden"]
        tanggal = request.POST["tanggal"]
        keterangan = request.POST['keterangan']

        try:
            produksubkonobj = models.ProdukSubkon.objects.get(
                IDProdukSubkon=nama_kode
            )
            print(produksubkonobj)
        except models.ProdukSubkon.DoesNotExist:
            messages.error(request, "Kode Produk Subkon tidak ditemukan")
            return redirect("update_transaksi_subkon_terima",id=id)
        
        produkobj.KodeProduk = produksubkonobj
        produkobj.Jumlah = jumlah
        produkobj.Tanggal = tanggal
        produkobj.Keterangan = keterangan
        
        produkobj.save()
        messages.success(request, "Data berhasil disimpan")

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Transaksi Produk Subkon. Nama Produk: {produksubkonobj.NamaProduk} Artikel Peruntukan : {produksubkonobj.KodeArtikel}  Jumlah : {jumlah}",
        ).save()
        
        return redirect("transaksi_subkon_terima")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_transaksi_subkon_terima(request, id):
    '''
    Fitur ini digunakan untuk menghapus data transaksi subkon untuk produksi
    Algorima:
    1. Mengambil data transaksi produk subkon dengan kriteria IDTransaksiProdukSubkon = id (id didapatkan dari passing values HTML)
    2. Menghapus data transaksi produk subkon
    '''
    produkobj = models.TransaksiSubkon.objects.get(IDTransaksiProdukSubkon=id)
    produkobj.delete()
    messages.success(request, "Data Berhasil dihapus")

    models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Delete",
            pesan=f"Transaksi Produk Subkon. Nama Produk: {produkobj.KodeProduk.NamaProduk} Artikel Peruntukan : {produkobj.KodeProduk.KodeArtikel}  Jumlah : {produkobj.Jumlah}",
        ).save()
    messages.success(request,'Data berhasil dihapus')
    return redirect("transaksi_subkon_terima")

def calculateksbjsubkon(produk,tanggal_mulai,tanggal_akhir):
    '''
    Fitur ini digunakan untuk melakukan kalkulasi KSBJ Subkon
    Algoritma
    1. Mengamnbil data TransaksiSubkon dengan kriteria KodeProduk = produk (produk didapatkan dari fungsi lain berisi object kode produk subkon), tanggal berada dalam rentang Tanggal_mulai dan tanggal_akhir (tanggal Mulai dan Tanggal akhir didapatkan dari fungsi lain)
    2. Menambil data transaksi produksi produk subkon dengan kriteria KodeProduk = produk(produk didapatkan dari fungsi lain berisi object kode produk subkon), Tanggal berada dalam rentang tanggal mulai dan tanggal akhir
    3. Mengambil data Detail trasnaski surat jalan penerimaan produk subkon dengan kriteria KodeProduk = produk(produk didapatkan dari fungsi lain berisi object kode produk sibkon), tanggal berada dalam rentang tanggal mulai dan tanggal akhir
    4. mengambil data pemusnahan produk subkon dengan kriteria Kode Produk Subkon = produk, Tanggal berada dalam rentang tanggal mulai dan akhir
    5. Membuat tanggal masuk dari data Transaksi Subkon
    6. Membuat tanggal produksi dari data transaksi produksi produk subkon
    7. Membuat tanggal pemusnahan dari data pemusnahan produk subkon
    8. membuat tanggal keluar dari data detail transaksi surat jalan penerimaaan produk subkon
    9. Menggabungkan semua list tanggal 
    10. Mencari data Saldo awal Subkon dengan kriteria IDProdukSubkon = produk (produk didapat dari parameter fungsi ) dan tanggal berada dalam rentang tanggal mulai dan akhir, Apabila tidak ada maka set saldo awal menjadi 0
    11. Mengiterasi list tanggal 
        A. Membuat models data transaksi KSBJ Subkon 
        B. menghitung jumlah masuk dengan rumus = Jumlah masuk Detail Surat Jalan Penerimaan Produk Subkon + Data Produksi Subkon 
        C. Menghitung jumlah Keluar dengan rumus = Jumlah Pemusnahan Produk subkon + jumlah transaksi produk subkon 
        D. Menghitung stok setiap iterasi dengan rumus. Sisa = Jumlah masuk - jumlah keluar
    12. Mengembalikan data listdata dan saldoawal. List data berisi models data transaksi KSBJ Subkon, saldoawal berisi saldoawal object. Semisal tidak ada saldo awal maka akan None
    '''
    # Menceri data transaksi gudang dengan kode 
    dataterima = models.TransaksiSubkon.objects.filter(
        KodeProduk=produk.IDProdukSubkon, Tanggal__range=(tanggal_mulai, tanggal_akhir)
    )
    dataproduksisubkon = models.TransaksiProduksiProdukSubkon.objects.filter(KodeProduk = produk,Tanggal__range = (tanggal_mulai,tanggal_akhir))

    dataproduksi = models.DetailSuratJalanPenerimaanProdukSubkon.objects.filter(
        KodeProduk = produk.IDProdukSubkon,
        NoSuratJalan__Tanggal__range=(tanggal_mulai, tanggal_akhir),
    )
    datapemusnahan = models.PemusnahanProdukSubkon.objects.filter(KodeProdukSubkon = produk,Tanggal__range=(tanggal_mulai,tanggal_akhir))

    # ''' TANGGAL SECTION '''
    tanggalmasuk = dataterima.values_list("Tanggal", flat=True)
    tanggalkeluar = dataproduksi.values_list("NoSuratJalan__Tanggal", flat=True)
    tanggalpemusnahan = datapemusnahan.values_list('Tanggal',flat=True)
    tanggalprodukksi = dataproduksisubkon.values_list("Tanggal",flat=True)
    # tanggalpemusnahan = pemusnahanobj.values_list("Tanggal", flat=True)

    listtanggal = sorted(list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalpemusnahan).union(tanggalprodukksi))))

    ''' SALDO AWAL SECTION '''
    try:
        saldoawal = models.SaldoAwalSubkon.objects.get(
            IDProdukSubkon=produk.IDProdukSubkon,
            Tanggal__range=(tanggal_mulai, tanggal_akhir),
        )
        saldo = saldoawal.Jumlah
        saldoawal.Tanggal = saldoawal.Tanggal.year

    except models.SaldoAwalSubkon.DoesNotExist:
        saldo = 0
        saldoawal = None

    sisa = saldo

    # ''' PENGOLAHAN DATA '''
    listdata = [ ]
    for i in listtanggal:
        # Data Models
        data = {
            'Tanggal': None,
            'Masuk' : None,
            'Keluar' : None,
            'Sisa' : None,
            'Pemusnahan' : None
            
        }
        data['Tanggal'] = i.strftime("%Y-%m-%d")
        # Data Masuk
        masuk = 0
        datamasuk = dataproduksi.filter(NoSuratJalan__Tanggal = i)
        for m in datamasuk:
            masuk += m.Jumlah
        dataproduksimasuk = dataproduksisubkon.filter(Tanggal = i)
        for m in dataproduksimasuk:
            masuk += m.Jumlah
        sisa += masuk
        data['Masuk'] = masuk
        
        # Data Keluar
        datakeluar = dataterima.filter(Tanggal=i)
        keluar = 0
        pemusnahan = 0
        for k in datakeluar:
            keluar += k.Jumlah
        datapemusnahanproduk = datapemusnahan.filter(Tanggal = i)
        for k in datapemusnahanproduk:
            pemusnahan +=k.Jumlah
        sisa -= keluar + pemusnahan
        data['Keluar'] = keluar
        data['Pemusnahan'] = pemusnahan

        data['Sisa'] = sisa

        listdata.append(data)
    return listdata,saldoawal
# KSBB KSBJ Subkon
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_ksbjsubkon(request):
    '''
    FItur ini digunakan untuk melihat KSBJ Subkon yang ada pada sistem
    Algoritma
    1. Mengambil data produk subkon
    2. Menampilkan form inpiut kode produk subkon 
    3. Program menerima input berupa kode produk subkon 
    4. Memanggil fungsi Calculateksbjsubkon
    5. menampilkan data transaksi KSBJ Subkon
    '''
    kodeproduk = models.ProdukSubkon.objects.all()
    print(kodeproduk)
    if len(request.GET) == 0:
        return render(request, "produksi/view_ksbjsubkon.html", {"kodeprodukobj": kodeproduk})
    else:
        """
        1. Cari 
        """
        nama_kode = request.GET["kodebarangHidden"]
        print(nama_kode)
        try:
            produk = models.ProdukSubkon.objects.get(
               pk = nama_kode
            )
            nama = produk.NamaProduk
            satuan = produk.Unit
        except:
            messages.error(request, "Kode Produk Subkon tidak ditemukan")
            return redirect("ksbjsubkon")
        
        idartikel = produk.KodeArtikel
        artikel = models.Artikel.objects.get(KodeArtikel=idartikel)
        
        if request.GET["periode"]:
            tahun = int(request.GET["periode"])
        else:
            sekarang = datetime.now()
            tahun = sekarang.year

        tanggal_mulai = datetime(year=tahun, month=1, day=1)
        tanggal_akhir = datetime(year=tahun, month=12, day=31)
        listdata,saldoawal = calculateksbjsubkon(produk,tanggal_mulai,tanggal_akhir)
        
        

        return render(request, "produksi/view_ksbjsubkon.html",{"data":listdata,'saldo':saldoawal,"nama": nama,"satuan": satuan,"artikel":artikel,"kodeprodukobj": kodeproduk, 'produk':produk,'tahun':tahun})

def calculateksbbsubkon(produk,tanggal_mulai,tanggal_akhir):
    '''
    Fungsi ini digunakan untuk menghitung transaksi KSBB Subkon 
    Parameter : 
    1. Produk = digunakan untuk menampung Object bahan baku subkon dari fungsi lain 
    2. Tanggal_mulai = digunakan untuk menampung tanggal mulai dari fungsi lain 
    3. Tanggal_akhir = digunakan untuk menampung tanggal akhir dari fungsi lain 
    Algoritma
    1. Mengambil data transaksi bahan baku subkon dengan kriteria Kode Produk = produk (didapatkan dari parameter),Tanggal berada dalam rentang tanggal_mulai dan tanggal_akhir
    2. Mengambil data detail surat jalan pengiriman bahan baku subkon dengan kriteria Kode produk = produk (didapatan dari parameter), tanggal berada dalam rentang tanggal mulai dan tanggal akhir
    3. Mengambil data pemusnahan bahan baku subkon dengan kriteria Kode Bahan Baku = produk (parameter) dan tanggal berada dalam rentang tanggal mulai dan tanggal akhir 
    4. Membuat list tanggal dari transaksi bahan baku subkon 
    5. Membuat list tanggal dari transaksi detail surat jalan pengiriman bahan baku subkon 
    6. Membuat list tanggal dari transaksi pemusnahan bahan baku subkon 
    7. Menggabungkan list tanggal semua trasnaksi 
    8. Mencari saldo awal bahan baku subkon dengan kriteria KodeProduk = produk (parameter), Tanggal berada dalam rentang tanggal mulai dan akhir. Jika tidak ada maka set menjadi 0
    9. Mengiterasi data tanggal 
        A. Membuat models data transaksi KSBB Subkon
        B. menghitung jumlah masuk dengan rumus = Jumlah Transaksi Bahan Baku Subkon
        C. Menghitung jumlah Keluar dengan rumus = Jumlah Pemusnahan Produk subkon + jumlah detail surat jalan pengiriman bahan baku subkon 
        D. Menghitung stok setiap iterasi dengan rumus. Sisa = Jumlah masuk - jumlah keluar
    10. fungsi mengembalikan listdata dan saldoawal. List data berisi models data transaksi KSBB Subkon tiap tanggal, dan Saldo Awal berisi saldo awal object dari Bahan Baku subkon selama taun yang diinput

    '''
    # Menceri data transaksi gudang dengan kode 
    databahan = models.TransaksiBahanBakuSubkon.objects.filter(
        KodeBahanBaku__KodeProduk=produk.KodeProduk, Tanggal__range=(tanggal_mulai, tanggal_akhir)
    )

    datakirim = models.DetailSuratJalanPengirimanBahanBakuSubkon.objects.filter(
        KodeBahanBaku__KodeProduk=produk.KodeProduk, NoSuratJalan__Tanggal__range=(tanggal_mulai, tanggal_akhir),
    )
    datapemusnahan = models.PemusnahanBahanBakuSubkon.objects.filter(Tanggal__range=(tanggal_mulai,tanggal_akhir),KodeBahanBaku = produk)
    
    ''' TANGGAL SECTION '''
    tanggalmasuk = databahan.values_list("Tanggal", flat=True)
    tanggalkeluar = datakirim.values_list("NoSuratJalan__Tanggal", flat=True)
    tanggalpemusnahan = datapemusnahan.values_list('Tanggal',flat=True)

    listtanggal = sorted(
        list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalpemusnahan)))
    )
    # print(datapemusnahan)
    # print(asd)

    ''' SALDO AWAL SECTION '''
    try:
        saldoawal = models.SaldoAwalBahanBakuSubkon.objects.get(
            IDBahanBakuSubkon__KodeProduk=produk.KodeProduk,
            Tanggal__range=(tanggal_mulai, tanggal_akhir),
        )
        saldo = saldoawal.Jumlah
        saldoawal.Tanggal = saldoawal.Tanggal.year

    except models.SaldoAwalBahanBakuSubkon.DoesNotExist:
        saldo = 0
        saldoawal = None

    sisa = saldo

    ''' PENGOLAHAN DATA '''
    listdata =[]
    for i in listtanggal:
        data = {
            'Tanggal': None,
            'Masuk' : None,
            'Keluar' : None,
            'Sisa' : None,
            'Pemusnahan':None
        }

        data['Tanggal'] = i.strftime("%Y-%m-%d")
        # Data Masuk
        masuk = 0
        datamasuk = databahan.filter(Tanggal=i)
        for m in datamasuk:
            masuk += m.Jumlah
        sisa  += masuk
        data['Masuk'] = masuk
        
        # Data Keluar
        keluar = 0
        pemusnahan = 0
        datakeluar = datakirim.filter(NoSuratJalan__Tanggal = i)
        for k in datakeluar:
            keluar += k.Jumlah
        datapemusnahanbahanbaku = datapemusnahan.filter(Tanggal = i)
        # if datapemusnahanbahanbaku.exists():
        #     print(datapemusnahanbahanbaku)
        #     print(asd)
        for k in datapemusnahanbahanbaku:
            pemusnahan += k.Jumlah
        sisa -= keluar + pemusnahan
        data['Keluar'] = keluar
        data['Pemusnahan'] = pemusnahan



        data['Sisa'] = sisa
        listdata.append(data)
    return listdata,saldoawal

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_ksbbsubkon(request):
    '''
    Fitur ini digunakan untuk melihat KSBB Subkon yang ada pada perusahaan
    Algoritma
    1. Mengambil semua data bahan baku subkon
    2. menampilkan input form kode stok bahan baku subkon
    3. Program menerima input berupa kode stok bahan baku 
    4. Memanggil fungsi calculateksbbsubkon
    5. Menampilkan KSBB Subkon
    '''
    kodeproduk = models.BahanBakuSubkon.objects.all()
    if len(request.GET) == 0:
        return render(request, "produksi/view_ksbbsubkon.html", {"kodeprodukobj": kodeproduk})
    else:
        """
        1. Cari 
        """
        try:
            print(request)
            produk = models.BahanBakuSubkon.objects.get(KodeProduk=request.GET["kodebarang"])
            nama = produk.NamaProduk
            satuan = produk.unit
        except:
            messages.error(request, "Data Bahan tidak ditemukan")
            return redirect("ksbbsubkon")
        
        if request.GET["periode"]:
            tahun = int(request.GET["periode"])
        else:
            sekarang = datetime.now()
            tahun = sekarang.year

        tanggal_mulai = datetime(year=tahun, month=1, day=1)
        tanggal_akhir = datetime(year=tahun, month=12, day=31)
        listdata,saldoawal = calculateksbbsubkon(produk,tanggal_mulai,tanggal_akhir)
        

        

        return render(request, "produksi/view_ksbbsubkon.html",{'produk':produk,'data':listdata,'saldo':saldoawal,'kodebarang':request.GET["kodebarang"],"nama": nama,"satuan": satuan,"kodeprodukobj": kodeproduk,'tahun':tahun})

def ksbbcat (request):
    kodecat = 'A-004'
    sekarang = datetime.now().year
    tanggal_mulai = datetime(year=sekarang, month=1, day=1)
    tanggal_akhir = datetime(year=sekarang, month=12, day=31)
    kodeproduk = models.Produk.objects.filter(KodeProduk__startswith=kodecat)
    if len(request.GET) == 0:
        print(kodeproduk)
        return render(request, "produksi/view_ksbbcat.html", {"kodeprodukobj": kodeproduk, "sekarang": sekarang})
    else:
        getkodeproduk = request.GET['kodebarang']
        datakodeprodukobj = models.Produk.objects.get(KodeProduk = getkodeproduk)
        lokasi = 'WIP'
        data,saldoawal = calculateksbbcat(datakodeprodukobj,tanggal_mulai,tanggal_akhir,lokasi)
        print(data)
        # print(asd)
        return render(request, "produksi/view_ksbbcat.html", {"kodeprodukobj": kodeproduk, "sekarang": sekarang,'data':data,'saldo':saldoawal,'kodeproduk':getkodeproduk})

def calculateksbbcat(produk,tanggal_mulai,tanggal_akhir,lokasi):


    datagudang = models.TransaksiGudang.objects.filter(
        KodeProduk=produk, tanggal__range=(tanggal_mulai, tanggal_akhir),Lokasi__NamaLokasi=(lokasi),jumlah__gte=0
    )
    dataretur = models.TransaksiGudang.objects.filter(
        KodeProduk=produk, tanggal__range=(tanggal_mulai, tanggal_akhir),Lokasi__NamaLokasi=(lokasi),jumlah__lt=0
    )
    # Kode Artikel yang di susun oleh bahan baku 
    penyusun_produk = (
        models.Penyusun.objects.filter(KodeProduk=produk,Lokasi__NamaLokasi = lokasi)
        .values_list("KodeArtikel", flat=True)
        .distinct()
    )
    pemusnahanobj = models.PemusnahanArtikel.objects.filter(
        KodeArtikel__id__in=penyusun_produk,
        Tanggal__range=(tanggal_mulai, tanggal_akhir),lokasi__NamaLokasi=lokasi
    )
    # Mencari data pemusnahan bahan baku 
    pemusnahanbahanbakuobj = models.PemusnahanBahanBaku.objects.filter(KodeBahanBaku = produk,Tanggal__range=(tanggal_mulai,tanggal_akhir),lokasi__NamaLokasi=(lokasi))

    dataproduksi = models.TransaksiProduksi.objects.filter(
        KodeArtikel__id__in=penyusun_produk,
        Jenis="Mutasi",
        Tanggal__range=(tanggal_mulai, tanggal_akhir),
    )
    print(dataproduksi)
    
    # print(asd)
    datasppb = models.DetailSPPB.objects.filter(
        NoSPPB__Tanggal__range = (tanggal_mulai,tanggal_akhir),DetailSPK__KodeArtikel__id__in=penyusun_produk
    ).exclude(DetailSPKDisplay__isnull = False)

    datatransaksicat = models.TransaksiCat.objects.filter(
        KodeProduk = produk,Tanggal__range=(tanggal_mulai,tanggal_akhir)
    )
    print(datatransaksicat)
    # print(asd)
    
    listartikelmaster = []
    datamodelskonversimaster = {}
    for artikel in penyusun_produk:
        artikelmaster = models.Artikel.objects.get(id = artikel)
        # print(artikelmaster)
        konversi = models.KonversiMaster.objects.filter(
            KodePenyusun__KodeArtikel=artikel, KodePenyusun__KodeProduk=produk, KodePenyusun__Lokasi__NamaLokasi = lokasi
        ).order_by('KodePenyusun__versi')
        print(konversi)

        
        tanggalversi = konversi.values_list('KodePenyusun__versi',flat=True).distinct()
        # print(tanggalversi)
        # print(asd)
        listkonversi = []
        if konversi.exists():
            dummy = {}
            for tanggal in tanggalversi:
                print(tanggal,tanggalversi,konversi)
                datakonversi = konversi.filter(KodePenyusun__versi = tanggal)
                print(datakonversi)
                kuantitas = datakonversi.aggregate(total = Sum('Allowance'))
                print(kuantitas)
                listkonversi.append(kuantitas['total'])
                dummy[tanggal] = kuantitas['total']
                # print(asd)
            datamodelskonversimaster[artikel] = {'konversi':dummy,'penyesuaian':{}}
            print(datamodelskonversimaster)
            print(listkonversi,tanggalversi)
            # print(asd)
        artikelmaster.listkonversi = listkonversi
        artikelmaster.tanggalversi = tanggalversi

        # Data Penyesuaian 
        penyesuaianobj  = models.Penyesuaian.objects.filter( KodeProduk = produk ,KodeArtikel = artikel, TanggalMulai__range = (tanggal_mulai,tanggal_akhir)).order_by('TanggalMulai')
        print(penyesuaianobj)
        # print(asd)
        if penyesuaianobj.exists:
            dummy ={}
            for penyesuaian in penyesuaianobj:
                dummy[penyesuaian.TanggalMulai] = penyesuaian.konversi
            datamodelskonversimaster[artikel]['penyesuaian'] = dummy
        penyesuaiandataperartikel = [i.konversi for i in penyesuaianobj]
        tanggalpenyesuaianperartikel = [i.TanggalMulai for i in penyesuaianobj]
        tanggalakhirpenyesuaian = [i.TanggalMinus for i in penyesuaianobj]

        artikelmaster.listpenyesuaian = penyesuaiandataperartikel
        artikelmaster.tanggalpenyesuaian =tanggalpenyesuaianperartikel
        artikelmaster.tanggalakhirpenyesuaian = tanggalakhirpenyesuaian
        listartikelmaster.append(artikelmaster)
        print(penyesuaiandataperartikel, tanggalpenyesuaianperartikel,tanggalakhirpenyesuaian)
        # print(asdas)
        # print(asdas)
    print(listartikelmaster)

    tanggalmasuk = datagudang.values_list("tanggal", flat=True)
    tanggalretur = dataretur.values_list('tanggal',flat=True)
    tanggalkeluar = dataproduksi.values_list("Tanggal", flat=True)
    tanggalpemusnahan = pemusnahanobj.values_list("Tanggal", flat=True)
    tanggalpemusnahanbahanbaku = pemusnahanbahanbakuobj.values_list('Tanggal',flat=True)
    tanggalpengirimanbarang = datasppb.values_list('NoSPPB__Tanggal',flat=True)
    tanggaltransaksicat = datatransaksicat.values_list("Tanggal",flat=True)

    listtanggal = sorted(
        list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalpemusnahan).union(tanggalpemusnahanbahanbaku).union(tanggalretur).union(tanggaltransaksicat)))
    )
    ''' SALDO AWAL SECTION '''
    sisapengambilanawal = 0 
    try:
        saldoawal = models.SaldoAwalBahanBaku.objects.get(
            IDBahanBaku=produk,
            IDLokasi__NamaLokasi=lokasi,
            Tanggal__range=(tanggal_mulai, tanggal_akhir),
        )
        saldo = saldoawal.Jumlah

        sisapengambilanawal = saldoawal.SisaPengambilan

    except models.SaldoAwalBahanBaku.DoesNotExist:
        saldo = 0
        saldoawal = None

    sisa = saldo

    ''' PENGOLAHAN DATA '''
    listdata =[]
    print(listtanggal)
    print(saldoawal)

    for i in listtanggal:
        datamodelsartikel = []
        datamodelsperkotak = []
        datamodelskonversi = []
        datamodelskeluar = []
        datamodelssisa = []
        datadetailmasuk = []
        datasisapengambilan = []
        data = {
            'Tanggal': None,
            'Artikel': datamodelsartikel,
            "Konversi" : datamodelskonversi,
            'Masuk' : None,
            'SisaPengambilan':None,
            'Perkotak' : datamodelsperkotak,
            'Persatuan' : datamodelskeluar,
            'Sisa' : datamodelssisa
            
        }
        data['Tanggal'] = i.strftime("%Y-%m-%d")
        masuk = 0
        datamasuk = datagudang.filter(tanggal=i)
        datakeluarretur = dataretur.filter(tanggal=i)
        datakeluar = dataproduksi.filter(Tanggal = i)
        datapemusnahan = pemusnahanobj.filter(Tanggal = i)
        datapemusnahanbahanbaku = pemusnahanbahanbakuobj.filter(Tanggal = i)
        datapengiriman = datasppb.filter(NoSPPB__Tanggal = i)
        datatransaksicatfiltered = datatransaksicat.filter(Tanggal = i)
        print(datamasuk)
        # print(asd)
        '''Masuk dari Gudang'''
        for k in datamasuk:
            masuk += k.jumlah
            # print(k.DetailSPKDisplay != None)
            if k.DetailSPKDisplay != None and k.DetailSPKDisplay.NoSPK.StatusDisplay == True:
                masukdisplay +=k.jumlah
                print(masukdisplay)
            else:
                datadetailmasuk.append(k)

        sisa  += masuk
        print(sisa)
        
        # Data Keluar
        data['Masuk'] = masuk
        artikelkeluar = datakeluar.values_list('KodeArtikel',flat=True).distinct()
        artikelpemusnahan = datapemusnahan.values_list('KodeArtikel',flat=True).distinct()
        artikelkirim = datapengiriman.values_list('DetailSPK__KodeArtikel',flat=True).distinct()
        '''Transaksi Penimbangan Cat'''
        for j in datatransaksicatfiltered:
            total = datatransaksicatfiltered.aggregate(total=Sum('SisaPengambilan'))
            print(total)
            data['SisaPengambilan'] = total['total']
            persatuankeluar = data['Masuk'] + sisapengambilanawal- total['total']
            datamodelskeluar.append( persatuankeluar)
            sisapengambilanawal = total['total']
            print(persatuankeluar)
            print(sisa)
            sisa -= persatuankeluar
            print(sisa)
            # print(asd)
            
        '''Keluar Mutasi'''
        for j in artikelkeluar:
            print('\n',j)
            artikelkeluarobj = models.Artikel.objects.get(id = j)
            penyesuaianobj = models.Penyesuaian.objects.filter( KodeProduk = produk ,KodeArtikel = artikelkeluarobj, TanggalMulai__range = (tanggal_mulai,tanggal_akhir)).order_by('TanggalMulai')
            total = datakeluar.filter(KodeArtikel__id = j).aggregate(total = Sum('Jumlah'))
            print(total)
            indexartikel = listartikelmaster.index(artikelkeluarobj)
            filtered_data = [d for d in listartikelmaster[indexartikel].tanggalversi if d <= i]
            print(indexartikel, listartikelmaster,listartikelmaster[indexartikel].tanggalversi)
            print('ini i ',i, filtered_data,'\n\n\n\n\n\n')
            filtered_data.sort(reverse=True)
            # print(asd)
            if not filtered_data:
                filtered_data = 0


            if filtered_data != 0:
                tanggalversiterdekat = max(filtered_data)
                indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalversi).index(tanggalversiterdekat)
                konversiterdekat = listartikelmaster[indexartikel].listkonversi[indextanggalterdekat]
                print(i,filtered_data)
                # print(asd)
            else:
                tanggalversiterdekat = i
                konversiterdekat = 0

            if listartikelmaster[indexartikel].tanggalpenyesuaian  and filtered_data != 0:
                print('masuk1')
                # print(asd)
                
                cektanggal  = penyesuaianobj.filter(TanggalMulai__lte = i, TanggalMinus__gte = i).last()
                print(cektanggal)
                print(penyesuaianobj,i,j)
                # print(asd)
                if cektanggal:
                    print('Ada data')
                    filtered_data.sort(reverse=True)
                    tanggalversiterdekat = cektanggal.TanggalMulai
                    print(tanggalversiterdekat)
                    indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalpenyesuaian).index(tanggalversiterdekat)
                    konversiterdekat = listartikelmaster[indexartikel].listpenyesuaian[indextanggalterdekat]
                    print(i,cektanggal,konversiterdekat)
                    # print(asd)
                else :
                    filtered_data.sort(reverse=True)
                    tanggalversiterdekat = max(filtered_data)
                    print(tanggalversiterdekat)
                    indextanggalterdekat = list(listartikelmaster[indexartikel].tanggalversi).index(tanggalversiterdekat)
                    konversiterdekat = listartikelmaster[indexartikel].listkonversi[indextanggalterdekat]
            # elif listartikelmaster[indexartikel].tanggalpenyesuaian:
            #     cektanggal  = penyesuaianobj.filter(TanggalMulai__lte = i, TanggalMinus__gte = i).last()
            #     print(cektanggal)
            #     print(penyesuaianobj,i)
            #     print(asd)

            konversiterdekat= (konversiterdekat)
            datamodelskonversi.append(konversiterdekat)
            print(konversiterdekat)
            print(total)
            print(listartikelmaster)
            # print(asd)

            datamodelskeluar.append(konversiterdekat*total['total'])
            datamodelsartikel.append(artikelkeluarobj)
            datamodelsperkotak.append(total['total'])
            sisa -= konversiterdekat*total['total']
            sisa = (sisa)
            datamodelssisa.append(sisa)
        
        if datakeluarretur.exists():
            # Mengagregat Jumlah Bahan Baku rusak
            totalbahanbakuretur = datakeluarretur.aggregate(total=Sum('jumlah'))

            sisa -= totalbahanbakuretur['total']*-1
            datamodelssisa.append(sisa)
            datamodelskeluar.append(totalbahanbakuretur['total']*-1)
        

        if datapemusnahanbahanbaku.exists():
            # Mengagregat Jumlah Bahan Baku rusak
            totalbahanbakurusak = datapemusnahanbahanbaku.aggregate(total=Sum('Jumlah'))

            sisa -= totalbahanbakurusak['total']
            datamodelssisa.append(sisa)
            datamodelskeluar.append(totalbahanbakurusak['total'])
        
        if not datamodelssisa :
            sisa = (sisa)
            datamodelssisa.append(sisa)
        data['Sisa'] = sisa
        listdata.append(data)
        
    return listdata,saldoawal

def detailksbbcat(request, id, tanggal):
    tanggal = datetime.strptime(tanggal, "%Y-%m-%d")
    tanggal = tanggal.strftime("%Y-%m-%d")
    lokasi = 'WIP'

    # Transaksi Gudang
    datagudang = models.TransaksiGudang.objects.filter(tanggal=tanggal, KodeProduk__KodeProduk=id,Lokasi__NamaLokasi=(lokasi),jumlah__gte=0)
    dataretur = models.TransaksiGudang.objects.filter(tanggal=tanggal, KodeProduk__KodeProduk=id,Lokasi__NamaLokasi=(lokasi),jumlah__lt=0)
    for item in dataretur:
        item.jumlah = item.jumlah * -1
    listartikel = (
        models.Penyusun.objects.filter(KodeProduk__KodeProduk=id)
        .values_list("KodeArtikel__KodeArtikel", flat=True)
        .distinct()
    )

    # Transaksi Produksi
    dataproduksi = models.TransaksiProduksi.objects.filter(
        KodeArtikel__KodeArtikel__in=listartikel, Jenis="Mutasi", Tanggal=tanggal
    )
    print(dataproduksi)

    # Transaksi Pemusnahan
    datapemusnahan = models.PemusnahanArtikel.objects.filter(
        KodeArtikel__KodeArtikel__in=listartikel, Tanggal=tanggal,
    )
    # Transaksi Pemusnahan Bahan Baku
    datapemusnahanbahanbaku  =models.PemusnahanBahanBaku.objects.filter(Tanggal = tanggal,KodeBahanBaku__KodeProduk = id,lokasi__NamaLokasi=lokasi)
    print(datapemusnahanbahanbaku)
    print(datagudang)
    # Transaksi Cat
    datatransaksicat = models.TransaksiCat.objects.filter(Tanggal = tanggal,KodeProduk__KodeProduk = id)
    print(datatransaksicat)
    return render(
        request,
        "produksi/view_detailksbbcat.html",
        {
            "datagudang": datagudang,
            "dataproduksi": dataproduksi,
            "datapemusnahan": datapemusnahan,
            'datapemusnahanbahanbaku' : datapemusnahanbahanbaku,
            "dataretur" : dataretur,
            "datatransaksicat" : datatransaksicat
        },
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def view_transaksicat(request):
    dataproduksi = models.TransaksiCat.objects.all().order_by("-Tanggal")
    for i in dataproduksi:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(
        request, "produksi/view_transaksicat.html", {"dataproduksi": dataproduksi}
    )

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_transaksicat(request):
    kodeprodukcat = "A-004"
    dataprodukcat = models.Produk.objects.filter(KodeProduk__startswith=kodeprodukcat)
    if request.method == "GET":
        return render(
            request,
            "produksi/add_transaksicat.html",
            { "dataproduk": dataprodukcat},
        )
    else:

        kodeproduk = request.POST["kodeproduk"]
        sisapengambilan = request.POST["sisapengambilan"]
        tanggal = request.POST["tanggal"]


        try:
            kodeprodukobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Artikel tidak ditemukan")
            return redirect("add_transaksicat")

        transaksicatbj = models.TransaksiCat(
            KodeProduk = kodeprodukobj,
            Tanggal = tanggal,
            SisaPengambilan = sisapengambilan
        ).save()
        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Create",
            pesan=f"Transaksi Cat. Kode Bahan Baku : {kodeprodukobj.KodeProduk} SisaPengambilan : {sisapengambilan} ",
        ).save()
        messages.success(request,'Data Berhasil Disimpan')
        
        return redirect("view_transaksicat")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_transaksicat(request, id):
    kodeprodukcat = "A-004"
    dataprodukcat = models.Produk.objects.filter(KodeProduk__startswith=kodeprodukcat)
    dataobj = models.TransaksiCat.objects.get(id=id)
    dataobj.Tanggal = dataobj.Tanggal.strftime("%Y-%m-%d")
    if request.method == "GET":
        return render(
            request,
            "produksi/update_transaksicat.html",
            {"data": dataobj,'kodeproduk':dataprodukcat},
        )

    else:
        kodeproduk = request.POST["kodeproduk"]
        sisapengambilan = request.POST["sisapengambilan"]
        tanggal = request.POST["tanggal"]
        
        try:
            kodeprodukobj = models.Produk.objects.get(KodeProduk=kodeproduk)
        except:
            messages.error(request, "Kode Artikel tidak ditemukan")
            return redirect("update_pemusnahan" ,id=id)

        dataobj.Tanggal = tanggal
        dataobj.SisaPengambilan = sisapengambilan
        dataobj.KodeProduk = kodeprodukobj

        dataobj.save()

        models.transactionlog(
            user="Produksi",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Pemusnahan Artikel. Kode Bahan Baku : {kodeprodukobj.KodeProduk} Sisa Pengambilan : {sisapengambilan} ",
        ).save()
        messages.success(request,"Data Berhasil disimpan")
        return redirect("view_transaksicat")

def trackingartikelspksppb(request):
    '''
    Fitur ini digunakan untuk melakukan tracking artikel dan display berdasarkan SPK
    Algoritma
    1. Mengambil data artikel 
    2. Mengambil data display
    3. Menampilkan form input kode artikel atau kode display pada user
    4. Program menerima input kode artikel atau kode display
    5. Mengambil data object kode artikel / display
    6. Apabila data artikel ada 
        A. Mengambil data SPK dengan rkiteria Kode Artikel = object artikel
        B. Mengurutkan data SPK berdasarkan tanggal terlama ke terbaru
        C. Mengiterasi data SPK 
            I. Mengambil data detailSPPB dengan kriteria DetailSPK = spkiterasi
            II. Menghitung jumlah Masuk
            III. Menghitung sisa produksi SPK 
        D. Menambahkan data sisa produksi SPK ke dalam models data
    7. Apabila data display ada
        A. Mengambil data SPK dengan kriteria Kode Display = object display
        B. Mengurutkan data SPK berdasarkan tanggal terlama ke terbaru
        C. Mengiterasi data SPK
            I. Mengambil data detailSPPB dengan kriteria DetailSPK = spkiterasi
            II. Menghitung jumlah Masuk
            III. Menghitung sisa produksi SPK 
        D. Menambahkan data sisa produksi SPK ke dalam models data
    8. Menampilkan data
    '''
    dataartikel = models.Artikel.objects.all()
    datadisplay = models.Display.objects.all()

    if len(request.GET) == 0:
        return render(request,'produksi/trackingspksppb.html',{'dataartikel':dataartikel,'datadisplay':datadisplay})
        '''
        ALGORITMA
        1. Ambil data artikel
        2. Filter SPK Berdasarkan artikel tersebut
        3. Ambil data detail SPPB berdasarkan SPK Tersebut Diurutkan dari tanggal sppb
        4. Hitung sisa

        datamodels
        9010 = {SPK1:{detail}}
        '''
    else:
        kodeartikel = request.GET['kodeartikel']
        kodedisplay = request.GET['kodedisplay']
        
        cekartikel = models.Artikel.objects.filter(KodeArtikel = kodeartikel)
        cekdisplay = models.Display.objects.filter(KodeDisplay = kodedisplay)
        print(request.GET)
        print(cekdisplay)
        
        if cekartikel.exists():
            try:
                artikelobj = models.Artikel.objects.get(KodeArtikel = kodeartikel)
            except models.Artikel.DoesNotExist:
                messages.error(request,f'Artikel {kodeartikel} tidak ditemukan dalam sistem')
                return redirect('trackingartikelspksppb')
            filteredspk = models.DetailSPK.objects.filter(KodeArtikel = artikelobj).order_by("NoSPK__Tanggal")
            datamodels = {}
            for item in filteredspk:
                detailsppbobj = models.DetailSPPB.objects.filter(DetailSPK = item).order_by("NoSPPB__Tanggal")
                item.detailsppb = detailsppbobj
                sisaspk = item.Jumlah
                jumlahkirim = 0
                recordsisa = []
                item.NoSPK.Tanggal = item.NoSPK.Tanggal.strftime('%Y-%m-%d')
                for detailsppb in detailsppbobj:
                    jumlahkirim +=detailsppb.Jumlah
                    sisaspk -= detailsppb.Jumlah
                    dummy = {"detailsppb":detailsppb,"sisa":sisaspk}
                    recordsisa.append(dummy)
                    detailsppb.NoSPPB.Tanggal = detailsppb.NoSPPB.Tanggal.strftime('%Y-%m-%d')
                # datamodels.append({'SPK':item,"data":recordsisa})

                datamodels[item.NoSPK.NoSPK] = {'spk':item,"data":recordsisa,'total':{'jumlahkirim':jumlahkirim,'sisa':sisaspk}}
            return render(request,'produksi/trackingspksppb.html',{'dataartikel':dataartikel,'listdata':datamodels,'kodeartikel':kodeartikel,'kodedisplay':''})
        elif cekdisplay.exists():
            try:
                displayobj = models.Display.objects.get(KodeDisplay = kodedisplay)
            except models.Artikel.DoesNotExist:
                messages.error(request,f'Display {kodedisplay} tidak ditemukan dalam sistem')
                return redirect('trackingartikelspksppb')
            filteredspk = models.DetailSPKDisplay.objects.filter(KodeDisplay = displayobj).order_by("NoSPK__Tanggal")
            datamodels = {}
            for item in filteredspk:
                detailsppbobj = models.DetailSPPB.objects.filter(DetailSPKDisplay = item).order_by("NoSPPB__Tanggal")
                item.detailsppb = detailsppbobj
                sisaspk = item.Jumlah
                jumlahkirim = 0
                recordsisa = []
                item.NoSPK.Tanggal = item.NoSPK.Tanggal.strftime('%Y-%m-%d')
                for detailsppb in detailsppbobj:
                    jumlahkirim +=detailsppb.Jumlah
                    sisaspk -= detailsppb.Jumlah
                    dummy = {"detailsppb":detailsppb,"sisa":sisaspk}
                    recordsisa.append(dummy)
                    detailsppb.NoSPPB.Tanggal = detailsppb.NoSPPB.Tanggal.strftime('%Y-%m-%d')
                # datamodels.append({'SPK':item,"data":recordsisa})

                datamodels[item.NoSPK.NoSPK] = {'spk':item,"data":recordsisa,'total':{'jumlahkirim':jumlahkirim,'sisa':sisaspk}}
            return render(request,'produksi/trackingspksppb.html',{'dataartikel':dataartikel,'listdata':datamodels,'kodedisplay':kodedisplay,'kodeartikel':''})
        else:
            messages.error(request,f'Kode {kodeartikel} tidak ditemukan pada database Artikel dan Display')
            return redirect('trackingartikelspksppb')
        # print(datamodels)
        # print(datamodels[2])
        # for item in datamodels['01/SPK/I-2024']:
        #     print(item)
        # print(datamodels['01/SPK/I-2024'])
        # print(datamodels["01/SPK/I-2024"]['spk'].Jumlah)
        # for data in datamodels["01/SPK/I-2024"]['data']:
        #     print(data)
                


    

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_pemusnahancat(request, id):
    dataobj = models.TransaksiCat.objects.get(id=id)
    dataobj.delete()

    models.transactionlog(
        user="Produksi",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Pemusnahan Bahan Baku. Kode Bahan Baku : {dataobj.KodeProduk.KodeProduk} Sisa Pengambilan : {dataobj.SisaPengambilan}",
    ).save()
    messages.success(request,'Data Berhasil dihapus')
    return redirect("view_transaksicat")


@login_required
@logindecorators.allowed_users(allowed_roles=["produksi",'ppic'])
def views_penyusun(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data konversi pada tiap artikel
    Pada bagian ini user dapat melakukan Melihat Konversi Artikel, Melihat Versi pada Artikel, dan manajemen konversi data artikel
    Algoritma
    1. User mengisi form input kode artikel yang akan dilihat
    2. Secara default, program akan menampilkan data konversi untuk versi default dari kode artikel tersebut
    3. User dapat melakukan manajemen data konversi pada halaman tersebut.
    '''
    print(request.GET)
    data = request.GET
    if len(request.GET) == 0:
        data = models.Artikel.objects.all()

        return render(request, "produksi/penyusun.html", {"dataartikel": data})
    else:
        kodeartikel = request.GET["kodeartikel"]

        try:
            dataartikel = models.Artikel.objects.all()
            get_id_kodeartikel = models.Artikel.objects.get(KodeArtikel=kodeartikel)
            data = models.Penyusun.objects.filter(KodeArtikel=get_id_kodeartikel.id)
            versifiltered = models.Versi.objects.filter(KodeArtikel = get_id_kodeartikel)
            dataversi = versifiltered.values_list("Versi", flat=True).distinct()
            print(dataversi)
            if dataversi.exists():
                try:
                    if request.GET["versi"] == "":
                        versiterpilih = dataversi.order_by("-Versi").first()
                        print("ini versi terbaru", versiterpilih)
                    else:
                        versiterpilih = request.GET["versi"]
                except:
                    versiterpilih = dataversi.order_by("-Versi").first()
                    print("ini versi terbaru", versiterpilih)
                
                data = data.filter(KodeVersi__Versi=versiterpilih)
                datakonversi = []
                nilaifg = 0
                sekarang = date.today()
                awaltahun = date(sekarang.year, 1, 1)
                akhirtahun = date(sekarang.year,12,31)
                print(data)
                if data.exists():
                    for item in data:
                        hargaterakhir = 0
                        print(item, item.IDKodePenyusun)
                        
                        kuantitaskonversi = item.Kuantitas
                        kuantitasallowance = item.Allowance
                        cachevalue = models.CacheValue.objects.filter(KodeProduk = item.KodeProduk, Tanggal__month =sekarang.month).first()
                        if cachevalue:
                            hargasatuanawal = cachevalue.Harga
                        else:
                        # print(konversidataobj.Kuantitas)
                            masukobj = models.DetailSuratJalanPembelian.objects.filter(
                KodeProduk=item.KodeProduk, NoSuratJalan__Tanggal__range=(awaltahun,akhirtahun)
            )

                            tanggalmasuk = masukobj.values_list("NoSuratJalan__Tanggal", flat=True)

                            keluarobj = models.TransaksiGudang.objects.filter(
                                jumlah__gte=0, KodeProduk=item.KodeProduk, tanggal__range=(awaltahun,akhirtahun)
                            )
                            returobj = models.TransaksiGudang.objects.filter(
                                jumlah__lt=0, KodeProduk=item.KodeProduk, tanggal__range=(awaltahun,akhirtahun)
                            )
                            pemusnahanobj = models.PemusnahanBahanBaku.objects.filter(
                                lokasi__NamaLokasi = 'Gudang',KodeBahanBaku = item.KodeProduk, Tanggal__range = (awaltahun,akhirtahun)
                            )
                            # print(pemusnahanobj)
                            # print(asd)
                            tanggalkeluar = keluarobj.values_list("tanggal", flat=True)
                            tanggalretur = returobj.values_list("tanggal", flat=True)
                            tanggalpemusnahan = pemusnahanobj.values_list("Tanggal",flat=True)
                            print("ini kode bahan baku", keluarobj)
                            saldoawalobj = (
                                models.SaldoAwalBahanBaku.objects.filter(
                                    IDBahanBaku=item.KodeProduk,
                                    IDLokasi__IDLokasi=3,
                                    Tanggal__range=(awaltahun,akhirtahun)
                                )
                                .order_by("-Tanggal")
                                .first()
                            )
                            print(saldoawalobj)
                            print('ini item',item)
                            if (
                                not keluarobj.exists()
                                and not returobj.exists()
                                and not masukobj.exists()
                                and saldoawalobj is None
                            ):
                                messages.error(request, f"Tidak ditemukan data Transaksi Barang pada penyusun {item.KodeProduk.KodeProduk} - {item.KodeProduk.NamaProduk}")
                                hargasatuanawal = 0
                                hargaterakhir = 0
                                hargaperkotak = 0
                                # print(asd
                                kuantitaskonversi = item.Kuantitas
                                kuantitasallowance = item.Allowance
                                print(item)
                                # print(asd)
                                # return redirect("rekapharga")
                            # print(asdas)
                            else :
                                if saldoawalobj:
                                    print("ada data")
                                    saldoawal = saldoawalobj.Jumlah
                                    hargasatuanawal = saldoawalobj.Harga
                                    hargatotalawal = saldoawal * hargasatuanawal
                                    tahun = saldoawalobj.Tanggal.year

                                else:
                                    saldoawal = 0
                                    hargasatuanawal = 0
                                    hargatotalawal = saldoawal * hargasatuanawal
                                    tahun = awaltahun.year

                                saldoawalobj = {
                                    "saldoawal": saldoawal,
                                    "hargasatuanawal": hargasatuanawal,
                                    "hargatotalawal": hargatotalawal,
                                    'tahun' : tahun
                                }
                                
                                listdata = []
                                print(tanggalmasuk)
                                print(tanggalkeluar)
                                listtanggal = sorted(
                                    list(set(tanggalmasuk.union(tanggalkeluar).union(tanggalretur).union(tanggalpemusnahan)))
                                )
                                print(listtanggal)
                                statusmasuk = False
                                for i in listtanggal:
                                    jumlahmasukperhari = 0
                                    hargamasuktotalperhari = 0
                                    hargamasuksatuanperhari = 0
                                    jumlahkeluarperhari = 0
                                    hargakeluartotalperhari = 0
                                    hargakeluarsatuanperhari = 0
                                    sjpobj = masukobj.filter(NoSuratJalan__Tanggal=i)
                                    if sjpobj.exists():
                                        for j in sjpobj:
                                            hargamasuktotalperhari += j.Harga * j.Jumlah
                                            jumlahmasukperhari += j.Jumlah
                                        hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
                                        print("data SJP ada")
                                        print(hargamasuksatuanperhari)
                                        print(jumlahmasukperhari)
                                        dumy = {
                                            "Tanggal": i.strftime("%Y-%m-%d"),
                                            "Jumlahstokawal": saldoawal,
                                            "Hargasatuanawal": round(hargasatuanawal, 2),
                                            "Hargatotalawal": round(hargatotalawal, 2),
                                            "Jumlahmasuk": jumlahmasukperhari,
                                            "Hargamasuksatuan": round(hargamasuksatuanperhari, 2),
                                            "Hargamasuktotal": round(hargamasuktotalperhari, 2),
                                            "Jumlahkeluar": jumlahkeluarperhari,
                                            "Hargakeluarsatuan": round(hargakeluarsatuanperhari, 2),
                                            "Hargakeluartotal": round(hargakeluartotalperhari, 2),
                                        }
                                        saldoawal += jumlahmasukperhari - jumlahkeluarperhari
                                        hargatotalawal += hargamasuktotalperhari - hargakeluartotalperhari
                                        hargasatuanawal = hargatotalawal / saldoawal

                                        print("Sisa Stok Hari Ini : ", saldoawal)
                                        print("harga awal Hari Ini :", hargasatuanawal)
                                        print("harga total Hari Ini :", hargatotalawal, "\n")
                                        dumy["Sisahariini"] = saldoawal
                                        dumy["Hargasatuansisa"] = round(hargasatuanawal, 2)
                                        dumy["Hargatotalsisa"] = round(hargatotalawal, 2)
                                        print(dumy)
                                        statusmasuk = True
                                        listdata.append(dumy)
                                        # print(asdasd)

                                    hargamasuktotalperhari = 0
                                    jumlahmasukperhari = 0
                                    hargamasuksatuanperhari = 0
                                    transaksigudangobj = keluarobj.filter(tanggal=i)
                                    transaksipemusnahan = pemusnahanobj.filter(Tanggal=i)


                                    if transaksigudangobj.exists():
                                        for j in transaksigudangobj:
                                            jumlahkeluarperhari += j.jumlah
                                            hargakeluartotalperhari += j.jumlah * hargasatuanawal

                                    if transaksipemusnahan.exists():
                                        for j in transaksipemusnahan:
                                            jumlahkeluarperhari += j.Jumlah
                                            hargakeluartotalperhari += j.Jumlah * hargasatuanawal

                                    if jumlahkeluarperhari > 0:
                                        hargakeluarsatuanperhari = hargakeluartotalperhari / jumlahkeluarperhari
                                    else:
                                        if statusmasuk:
                                            statusmasuk = False
                                            continue
                                        hargakeluartotalperhari = 0
                                        hargakeluarsatuanperhari = 0
                                        jumlahkeluarperhari = 0

                                    # transaksigudangobj = keluarobj.filter(tanggal=i)
                                    # print(transaksigudangobj)
                                    # if transaksigudangobj.exists():
                                    #     for j in transaksigudangobj:
                                    #         jumlahkeluarperhari += j.jumlah
                                    #         hargakeluartotalperhari += j.jumlah * hargasatuanawal
                                    #     hargakeluarsatuanperhari += (
                                    #         hargakeluartotalperhari / jumlahkeluarperhari
                                    #     )
                                    # else:
                                    #     if statusmasuk:
                                    #         statusmasuk = False
                                    #         continue
                                    #     hargakeluartotalperhari = 0
                                    #     hargakeluarsatuanperhari = 0
                                    #     jumlahkeluarperhari = 0

                                    # transaksipemusnahan = pemusnahanobj.filter(Tanggal=i)
                                    # print(transaksipemusnahan)
                                    # if transaksipemusnahan.exists():
                                    #     for j in transaksipemusnahan:
                                    #         jumlahkeluarperhari += j.Jumlah
                                    #         hargakeluartotalperhari += j.Jumlah * hargasatuanawal
                                    #     hargakeluarsatuanperhari += (
                                    #         hargakeluartotalperhari / jumlahkeluarperhari
                                    #     )
                                    # else:
                                    #     if statusmasuk:
                                    #         statusmasuk = False
                                    #         continue
                                    #     hargakeluartotalperhari = 0
                                    #     hargakeluarsatuanperhari = 0
                                    #     jumlahkeluarperhari = 0

                                    transaksireturobj = returobj.filter(tanggal=i)
                                    if transaksireturobj.exists():
                                        for j in transaksireturobj:
                                            jumlahmasukperhari += j.jumlah * -1
                                            hargamasuktotalperhari += j.jumlah * hargasatuanawal * -1
                                        hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
                                    else:
                                        hargamasuktotalperhari = 0
                                        hargamasuksatuanperhari = 0
                                        jumlahmasukperhari = 0


                                    print("Tanggal : ", i)
                                    print("Sisa Stok Hari Sebelumnya : ", saldoawal)
                                    print("harga awal Hari Sebelumnya :", hargasatuanawal)
                                    print("harga total Hari Sebelumnya :", hargatotalawal)
                                    print("Jumlah Masuk : ", jumlahmasukperhari)
                                    print("Harga Satuan Masuk : ", hargamasuksatuanperhari)
                                    print("Harga Total Masuk : ", hargamasuktotalperhari)
                                    print("Jumlah Keluar : ", jumlahkeluarperhari)
                                    print("Harga Keluar : ", hargakeluarsatuanperhari)
                                    print(
                                        "Harga Total Keluar : ", hargakeluarsatuanperhari * jumlahkeluarperhari
                                    )
                                    

                                    dumy = {
                                        "Tanggal": i.strftime("%Y-%m-%d"),
                                        "Jumlahstokawal": saldoawal,
                                        "Hargasatuanawal": round(hargasatuanawal, 2),
                                        "Hargatotalawal": round(hargatotalawal, 2),
                                        "Jumlahmasuk": jumlahmasukperhari,
                                        "Hargamasuksatuan": round(hargamasuksatuanperhari, 2),
                                        "Hargamasuktotal": round(hargamasuktotalperhari, 2),
                                        "Jumlahkeluar": jumlahkeluarperhari,
                                        "Hargakeluarsatuan": round(hargakeluarsatuanperhari, 2),
                                        "Hargakeluartotal": round(hargakeluartotalperhari, 2),
                                    }
                                    """
                                    Rumus dari Excel KSBB Purchasing
                                    Sisa = Sisa hari sebelumnya + Jumlah masuk hari ini - jumlah keluar hari ini 
                                    harga sisa satuan = total sisa / harga total sisa
                                    Harga keluar = harga satuan hari sebelumnya

                                    """
                                    dummysaldoawal = saldoawal
                                    dummyhargatotalawal = hargatotalawal
                                    dummyhargasatuanawal = hargasatuanawal

                                    saldoawal += jumlahmasukperhari - jumlahkeluarperhari
                                    hargatotalawal += hargamasuktotalperhari - hargakeluartotalperhari
                                    print(hargatotalawal, saldoawal)
                                    try:
                                        hargasatuanawal = hargatotalawal / saldoawal
                                    except:
                                        hargasatuanawal = 0

                                    print("Sisa Stok Hari Ini : ", saldoawal)
                                    print("harga awal Hari Ini :", hargasatuanawal)
                                    print("harga total Hari Ini :", hargatotalawal, "\n")
                                    dumy["Sisahariini"] = saldoawal
                                    dumy["Hargasatuansisa"] = round(hargasatuanawal, 2)
                                    dumy["Hargatotalsisa"] = round(hargatotalawal, 2)

                                listdata.append(dumy)

                        hargaterakhir += hargasatuanawal
                        hargaperkotak = hargaterakhir * kuantitasallowance
                        # print("\n", hargaterakhir, "\n")
                        nilaifg += hargaperkotak

                        datakonversi.append(
                            {
                                "HargaSatuan": round(hargaterakhir, 2),
                                "Penyusunobj": item,
                                "Konversi": round(kuantitaskonversi, 5),
                                "Allowance": round(kuantitasallowance, 5),
                                "Hargakotak": round(hargaperkotak, 2),
                            }
                        )
                    HargaFGArtikel= None
                    hargaartikel = models.HargaArtikel.objects.filter(KodeArtikel =get_id_kodeartikel,Tanggal__month = sekarang.month)
                    if hargaartikel.exists():
                        HargaFGArtikel = hargaartikel.first().Harga

                    dataartikel = models.Artikel.objects.all()
                    return render(
                        request,
                        "produksi/penyusun.html",
                        {
                             "data": datakonversi,
                            "kodeartikel": get_id_kodeartikel,
                            "nilaifg": nilaifg,
                            "versiterpilih": versiterpilih,
                            "dataversi": dataversi,
                            'dataartikel' : dataartikel,
                            'versiterpilihobj': models.Versi.objects.get(Versi = versiterpilih,KodeArtikel=get_id_kodeartikel)
                        },
                    )
                else:
                    messages.error(request, "Versi tidak ditemukan")
                    return redirect(
                        f"/produksi/konversi?kodeartikel={quote(kodeartikel)}&versi="
                    )
            else:
                messages.error(request, "Kode Artikel Belum memiliki penyusun")
                return render(
                    request,
                    "produksi/penyusun.html",
                    {"kodeartikel": get_id_kodeartikel,'dataartikel':dataartikel},
                )
        except models.Artikel.DoesNotExist:
            messages.error(request, "Kode Artikel Tidak ditemukan")
            return render(
                request,
                "produksi/penyusun.html",
                {"dataartikel": models.Artikel.objects.all()},
            )

def view_transaksimutasistok(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data transaksi mutasi kode stok
    Algoritma
    1. Mendapatkan data transaksu mutasi kode stok
    2. Menampilkan data transaksi mutasi kode stok
    '''
    if len (request.GET)==0:
        sekarang = datetime.now().date()
        tanggalawal = date(sekarang.year,1,1)
        tanggalakhir = date(sekarang.year,12,31)
    else :
        tanggalawal = request.GET['mulai']
        tanggalakhir = request.GET['akhir']
        if tanggalawal == '':
            tanggalawal = datetime.min
        if tanggalakhir == '':
            tanggalakhir = datetime.max
    dataobj = models.transaksimutasikodestok.objects.filter(Tanggal__range=(tanggalawal,tanggalakhir)).order_by(
        "-Tanggal"
    )
    for i in dataobj:
        i.Tanggal = i.Tanggal.strftime("%Y-%m-%d")

    return render(request, "produksi/view_transaksimutasikodestok.html", {"dataobj": dataobj,'tanggalawal':tanggalawal,'tanggalakhir':tanggalakhir})
@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def add_mutasikodestok(request):
    '''
    Fitur ini digunakan untuk menambahkan transaski mutasi kode stok 
    Algoritma
    1. Mendapatkan data kode stok bahan baku pada sistem
    2. Menampilkan form input trasnasksi mutasi kode stok 
    3. Program menerima inputan berupa tanggal, list KodeAsal, list KodeTujuan, list jumlah, list nama lokasi, list keterangan
    4. Mengiterasi List Kode Asal, List Kode Tujuan, List Jumlah, dan list keterangan
        A. Cek apakah kode produk asal ada dalam database ? apabila tidak ada maka akan menampilkan pesan error dan melanjutkan iterasi
        B. Cek apakah kode produk tujuan ada dalam database ? apabila tidak ada maka akan menampilkan pesan error dan melanjutkan iterasi
        C. Cek apakah kode produk asal dan kode produk tujuan sama ? apabila sama maka akan menampilkan pesan error dan melanjutkan iterasi
        D. Membuat object Transaksi mutasi Kode Stok dengan kriteria Kode Produk Asal = produk asal iterasi, Kode Produk Tujuan = produk tujuan iterasi,Jumlah = jumlah iterasi, Keterangan = keterangan iterasi, Tanggal = tanggal iterasi, dan lokasi = lokasi iterasi
        E. Menyimpan data object trasnaksi mtuasi kode stok 
    '''
    if request.method == "GET":
        data_produk = models.Produk.objects.all()
        lokasi = models.Lokasi.objects.filter(NamaLokasi__in=('WIP','FG'))

        listproduk = [produk.NamaProduk for produk in data_produk]

        return render(
            request,
            "produksi/add_mutasikodestok.html",
            {
                "kode_produk": data_produk,
                "listproduk": listproduk,
                'nama_lokasi' : lokasi
            },
        )

    if request.method == "POST":
        print(request.POST)
        # print(asd)
        tanggal = request.POST['tanggal']
        kodeasal = request.POST.getlist('kodeasal')
        kodetujuan = request.POST.getlist('kodetujuan')
        listjumlah = request.POST.getlist('jumlah')
        listlokasi = request.POST.getlist('nama_lokasi')
        listketerangan = request.POST.getlist('keterangan')
        # listkode = request.POST.getlist("kode_produk[]")
        # listlokasi = request.POST.getlist("nama_lokasi[]")
        # tanggal = request.POST["tanggal"]
        # listjumlah = request.POST.getlist("jumlah[]")
        # listketerangan = request.POST.getlist("keterangan[]")
        # listdetail = request.POST.getlist("detail_spk[]")

        i = 0
        for produkasal, produktujuan, jumlah,lokasi, keterangan  in zip(
            kodeasal, kodetujuan, listjumlah, listlokasi, listketerangan
        ):
            try:
                produkasal = models.Produk.objects.get(KodeProduk = produkasal)
            except models.Produk.DoesNotExist:
                messages.error(request,f'Kode Stok Asal {produkasal} tidak ditemukan')
                i+=1
                continue
            try:
                produktujuan = models.Produk.objects.get(KodeProduk = produktujuan)
            except models.Produk.DoesNotExist:
                messages.error(request,f'Kode Stok Tujuan {produktujuan} tidak ditemukan')
                i+=1
                continue
            if produkasal == produktujuan:
                messages.error(request,'Kode Stok Asal dan Tujuan Sama')
                i+=1
                continue
            
            newdataobj = models.transaksimutasikodestok(
                KodeProdukAsal = produkasal,
                KodeProdukTujuan = produktujuan,
                Jumlah = jumlah,
                Keterangan = keterangan,
                Tanggal = tanggal,
                Lokasi = models.Lokasi.objects.get(pk = lokasi)
                )
            


            
            newdataobj.save()
            messages.success(request, "Data berhasil ditambahkan")
            print('success')
                
        if i == len(kodeasal):
            return redirect("addmutasikodestok")
        
        return redirect("mutasikodestok")

@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def delete_mutasikodestok(request, id):
    '''
    Fitur ini digunakan untuk menghapus data mutasi kodestok
    Algoritma
    1. Mengambil transaksi mutasi kode stok dengan kriteria pk (primary key) = id (id didapatkan dari passing values HTML)
    2. Menghapus data 
    '''
    dataproduksi = models.transaksimutasikodestok.objects.get(pk=id)
    dataproduksi.delete()
    messages.success(request, "Data Berhasil dihapus")

    
    return redirect("mutasikodestok")


@login_required
@logindecorators.allowed_users(allowed_roles=['produksi','ppic'])
def update_mutasikodestok(request, id):
    '''
    Fitur ini digunakan untuk update transaksi mutasi kode stok
    Algoritma
    1. mengambil data transaksi mutasi kode stok dengan kriteria id = id(id didapatkan dari passing values HTML)
    2. menampilkan data pada user
    3. Program mendapatkan input Kode Stok asal, kode stok tujuan, tanggal, jumlah, keterangan, dan lokasi
    4. Mengupdate data
    '''
    produksiobj = models.transaksimutasikodestok.objects.get(id=id)
    dataproduk = models.Produk.objects.all()
    lokasi = models.Lokasi.objects.filter(NamaLokasi__in=('WIP','FG'))


    if request.method == "GET":
        tanggal = datetime.strftime(produksiobj.Tanggal, "%Y-%m-%d")
        return render(
            request,
            "produksi/update_mutasikodestock.html",
            {
                "datamutasi": produksiobj,
                "tanggal": tanggal,
               'allproduk':dataproduk,
               'datalokasi':lokasi
            },
        )

    elif request.method == "POST":
        print(request.POST)
        # print(asd)
        kodestokasal = request.POST['kodestokasal']
        kodestoktujuan = request.POST['kodestoktujuan'] 
        tanggal = request.POST["tanggal"]
        jumlah = request.POST["jumlah"]
        keterangan = request.POST["keterangan"]
        lokasi = request.POST['lokasi']
        try:
            kodeprodukasalobj = models.Produk.objects.get(KodeProduk=kodestokasal)
        except models.Produk.DoesNotExist:
            messages.error(request,f'Kode Stok Asal {kodestokasal} tidak ditemukan ')
            return redirect('updatemutasikodestok',id=id)
        try:
            kodeproduktujuanobj = models.Produk.objects.get(KodeProduk=kodestoktujuan)
        except models.Produk.DoesNotExist:
            messages.error(request,f'Kode Stok Tujuan {kodestoktujuan} tidak ditemukan ')
            return redirect('updatemutasikodestok',id=id)

        produksiobj.KodeProdukAsal = kodeprodukasalobj
        produksiobj.KodeProdukTujuan = kodeproduktujuanobj
        produksiobj.Jumlah = float(jumlah)
        produksiobj.Keterangan = keterangan
        produksiobj.Lokasi = models.Lokasi.objects.get(pk = lokasi)
        produksiobj.Tanggal = tanggal
        produksiobj.save()
        
        messages.success(request,f'Data berhasil disimpan ')

        return redirect("mutasikodestok")

def eksportksbbproduksi(request,id,lokasi,tahun):
    # waktuawalproses = time.time()
    print(lokasi)
    kodeprodukobj =models.Produk.objects.get(KodeProduk = id)
    tanggalmulai = date(int(tahun),1,1)
    tanggalakhir = date(int(tahun),12,31)
    dfksbb = pd.DataFrame()
    dfksbbfg = pd.DataFrame()

    listdata,saldoawal = calculate_KSBB(kodeprodukobj,tanggalmulai,tanggalakhir,'WIP')
    # print(listdata[0].keys())
    # print(listdata)
    # print(asd)
    if len(listdata) != 0:
        sisa_terakhir = 0
        # print(asd)
        output_data = []
        if saldoawal and  saldoawal.Jumlah != None:
            jumlahsaldoawal = saldoawal.Jumlah
        else:
            jumlahsaldoawal = 0
        output_data.append({
                    'Tanggal': tanggalmulai.strftime('%Y-%m-%d'),
                    'Masuk': None,
                    'Artikel': None,
                    'Perkotak': None,
                    'Konversi':None,
                    'Keluar': None,
                    'Saldo': jumlahsaldoawal,
                })
        
        sisa_terakhir = jumlahsaldoawal

    # Memetakan setiap artikel dan perkotak per tanggal, dan menghitung Saldo
        for record in listdata:
            tanggal = record['Tanggal']
            artikel_list = record.get('Artikel', [])
            perkotak_list = record.get('Perkotak', [])
            konversi_list = record.get('Konversi', [])
            keluar_list = record.get('Keluar', [])
            masuk = record['Masuk']
            sisa_list = record.get('Saldo', [])

            # Jika ini adalah hari pertama, ambil Saldo awal dari data
            print(konversi_list)
            # print(asd)
            if sisa_terakhir is None:
                sisa_terakhir = sisa_list[0] if sisa_list else 0

            # Jika artikel, perkotak, dan konversi kosong, tetap memasukkan minimal satu entri
            if not artikel_list and not perkotak_list and not konversi_list:
                sisa_awal = sisa_terakhir
                sisa_akhir = sisa_awal + masuk - sum(keluar_list)  # Hitung sisa total
                sisa_terakhir = sisa_akhir
                if keluar_list:
                    jumlahkeluar = sum(keluar_list)
                else:
                    jumlahkeluar = None

                output_data.append({
                    'Tanggal': tanggal,
                    'Artikel': None,
                    'Masuk': masuk,
                    'Perkotak': None,
                    'Konversi': None,
                    'Keluar': jumlahkeluar,
                    'Saldo': sisa_akhir,
                })
            else:
                # Menggunakan zip_longest untuk memasukkan setiap artikel, perkotak, konversi, dan keluar
                statusmasuk = False
                for artikel, perkotak, konversi, keluar, sisa in zip_longest(artikel_list, perkotak_list, konversi_list, keluar_list, sisa_list, fillvalue=None):
                    # Round konversi to 5 decimal places if it exists
                    konversi = round(konversi, 5) if konversi is not None else None
                    print(konversi)
                    # Hitung Saldo: menggunakan sisa sebelumnya jika ada
                    if statusmasuk == True:
                        masuk = " "
                    else:
                        sisa_awal = sisa_terakhir
                        sisa_akhir = sisa_awal + masuk - (keluar if keluar is not None else 0)  # Hitung sisa berdasarkan baris
                        sisa_terakhir = sisa_akhir

                    output_data.append({
                        'Tanggal': tanggal,
                        'Artikel': artikel,
                        'Masuk': masuk,
                        'Perkotak': perkotak,
                        'Konversi': konversi,
                        'Keluar': keluar,
                        'Saldo': sisa_akhir,
                    })
                    statusmasuk = True
        dfksbb = pd.DataFrame(output_data)
    print(dfksbb)
    # print(asd)
    listdata,saldoawal = calculate_KSBB(kodeprodukobj,tanggalmulai,tanggalakhir,'FG')
    print(listdata == True)
    # print(asd)
    if len(listdata) != 0:
        print(listdata)
        # print(asd)
        sisa_terakhir = 0
        output_data = []
        if saldoawal and  saldoawal.Jumlah != None:
            jumlahsaldoawal = saldoawal.Jumlah
        else:
            jumlahsaldoawal = 0
        output_data.append({
                    'Tanggal': '2024',
                    'Artikel': None,
                    'Masuk': None,
                    'Perkotak': None,
                    'Konversi':None,
                    'Keluar': None,
                    'Saldo': jumlahsaldoawal,
                })
        
        sisa_terakhir = jumlahsaldoawal

    # Memetakan setiap artikel dan perkotak per tanggal, dan menghitung Saldo
        for record in listdata:
            tanggal = record['Tanggal']
            artikel_list = record.get('Artikel', [])
            perkotak_list = record.get('Perkotak', [])
            konversi_list = record.get('Konversi', [])
            keluar_list = record.get('Keluar', [])
            masuk = record['Masuk']
            sisa_list = record.get('Saldo', [])

            # Jika ini adalah hari pertama, ambil Sisa awal dari data
            print(konversi_list)
            print(keluar_list)
            # print(asd)
            if sisa_terakhir is None:
                sisa_terakhir = sisa_list[0] if sisa_list else 0

            # Jika artikel, perkotak, dan konversi kosong, tetap memasukkan minimal satu entri
            if not artikel_list and not perkotak_list and not konversi_list:
                sisa_awal = sisa_terakhir
                sisa_akhir = sisa_awal + masuk - sum(keluar_list)  # Hitung sisa total
                sisa_terakhir = sisa_akhir
                if keluar_list:
                    jumlahkeluar = sum(keluar_list)
                else:
                    jumlahkeluar = None


                output_data.append({
                    'Tanggal': tanggal,
                    'Artikel': None,
                    'Masuk': masuk,
                    'Perkotak': None,
                    'Konversi': None,
                    'Keluar': jumlahkeluar,
                    'Saldo': sisa_akhir,
                })
            else:
                statusmasuk = True
                # Menggunakan zip_longest untuk memasukkan setiap artikel, perkotak, konversi, dan keluar
                for artikel, perkotak, konversi, keluar, sisa in zip_longest(artikel_list, perkotak_list, konversi_list, keluar_list, sisa_list, fillvalue=None):
                    # Round konversi to 5 decimal places if it exists
                    konversi = round(konversi, 5) if konversi is not None else None
                    print(keluar,' keluar')
                    # Hitung Saldo: menggunakan sisa sebelumnya jika ada
                    if statusmasuk == False:
                        masuk = ' '
                    else:
                        sisa_awal = sisa_terakhir
                        sisa_akhir = sisa_awal + masuk - (keluar if keluar is not None else 0)  # Hitung sisa berdasarkan baris
                        sisa_terakhir = sisa_akhir
                    output_data.append({
                        'Tanggal': tanggal,
                        'Artikel': artikel,
                        'Masuk': masuk,
                        'Perkotak': perkotak,
                        'Konversi': konversi,
                        'Keluar': keluar,
                        'Saldo': sisa_akhir,
                    })
                    statusmasuk = False
                    print(record)
                
        dfksbbfg = pd.DataFrame(output_data)

    print(dfksbb)
    print(dfksbbfg)
    buffer = BytesIO()

    if dfksbbfg.empty and dfksbb.empty:
        messages.error(request,f'Tidak dapat mengeksport ke Excel karena tidak ada data Transaksi')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return redirect('view_ksbb')

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        if not dfksbb.empty:
            dfksbb.to_excel(writer, index=False, startrow=5, sheet_name=f"{kodeprodukobj.KodeProduk}")
            writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {kodeprodukobj.KodeProduk}")
            writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {kodeprodukobj.NamaProduk}")
            writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {kodeprodukobj.unit}")
            maxrow = len(dfksbb)+1
            maxcol = len(dfksbb.columns)
            apply_number_format(writer.sheets[f"{kodeprodukobj.KodeProduk}"],6,maxrow+5,1,maxcol)
            apply_borders_thin(writer.sheets[f"{kodeprodukobj.KodeProduk}"],6,maxrow+5,maxcol)
            adjust_column_width(writer.sheets[f"{kodeprodukobj.KodeProduk}"],dfksbb,1,1)
        if not dfksbbfg.empty:
            dfksbbfg.to_excel(writer, index=False, startrow=5, sheet_name=f"{kodeprodukobj.KodeProduk} FG")
            writer.sheets[f"{kodeprodukobj.KodeProduk} FG"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {kodeprodukobj.KodeProduk} FG")
            writer.sheets[f"{kodeprodukobj.KodeProduk} FG"].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {kodeprodukobj.NamaProduk}")
            writer.sheets[f"{kodeprodukobj.KodeProduk} FG"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {kodeprodukobj.unit}")
            maxrow = len(dfksbbfg)+1
            maxcol = len(dfksbbfg.columns)
            apply_number_format(writer.sheets[f"{kodeprodukobj.KodeProduk} FG"],6,maxrow+5,1,maxcol)
            apply_borders_thin(writer.sheets[f"{kodeprodukobj.KodeProduk} FG"],6,maxrow+5,maxcol)
            adjust_column_width(writer.sheets[f"{kodeprodukobj.KodeProduk} FG"],dfksbbfg,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBB {kodeprodukobj.KodeProduk}.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response
def eksportksbbproduksikeseluruhan(request,id,lokasi,tahun):
    waktuawalproses = time.time()
    print(str)
    kodeprodukobj =models.Produk.objects.get(KodeProduk = id)
    tanggalmulai = date(int(tahun),1,1)
    tanggalakhir = date(int(tahun),12,31)
    dfksbb = pd.DataFrame()
    dfksbbfg = pd.DataFrame()
    produkall = models.Produk.objects.all()
    # produkall = models.Produk.objects.filter(KodeProduk__in = ['A-001-01','A-001-02','A-001-03','A-001-04','C-005-06'])
    listdf = []
    listproduk = []
    for produk in produkall:


        listdata,saldoawal = calculate_KSBB(produk,tanggalmulai,tanggalakhir,'WIP')
        # print(listdata[0].keys())
        # print(listdata)
        # print(asd)
        if len(listdata) != 0:
            sisa_terakhir = 0
            # print(asd)
            output_data = []
            if saldoawal and  saldoawal.Jumlah != None:
                jumlahsaldoawal = saldoawal.Jumlah
            else:
                jumlahsaldoawal = 0
            output_data.append({
                        'Tanggal': tanggalmulai.strftime('%Y-%m-%d'),
                        'Artikel': None,
                        'Perkotak': None,
                        'Konversi':None,
                        'Masuk': None,
                        'Keluar': None,
                        'Sisa': jumlahsaldoawal,
                    })
            
            sisa_terakhir = jumlahsaldoawal

        # Memetakan setiap artikel dan perkotak per tanggal, dan menghitung Sisa
            for record in listdata:
                tanggal = record['Tanggal']
                artikel_list = record.get('Artikel', [])
                perkotak_list = record.get('Perkotak', [])
                konversi_list = record.get('Konversi', [])
                keluar_list = record.get('Keluar', [])
                masuk = record['Masuk']
                sisa_list = record.get('Sisa', [])

                # Jika ini adalah hari pertama, ambil Sisa awal dari data
                # print(konversi_list)
                # print(asd)
                if sisa_terakhir is None:
                    sisa_terakhir = sisa_list[0] if sisa_list else 0

                # Jika artikel, perkotak, dan konversi kosong, tetap memasukkan minimal satu entri
                if not artikel_list and not perkotak_list and not konversi_list:
                    sisa_awal = sisa_terakhir
                    sisa_akhir = sisa_awal + masuk - sum(keluar_list)  # Hitung sisa total
                    sisa_terakhir = sisa_akhir

                    output_data.append({
                        'Tanggal': tanggal,
                        'Artikel': None,
                        'Perkotak': None,
                        'Konversi': None,
                        'Masuk': masuk,
                        'Keluar': None,
                        'Sisa': sisa_akhir,
                    })
                else:
                    # Menggunakan zip_longest untuk memasukkan setiap artikel, perkotak, konversi, dan keluar
                    for artikel, perkotak, konversi, keluar, sisa in zip_longest(artikel_list, perkotak_list, konversi_list, keluar_list, sisa_list, fillvalue=None):
                        # Round konversi to 5 decimal places if it exists
                        konversi = round(konversi, 5) if konversi is not None else None
                        print(konversi)
                        # Hitung Sisa: menggunakan sisa sebelumnya jika ada
                        sisa_awal = sisa_terakhir
                        sisa_akhir = sisa_awal + masuk - (keluar if keluar is not None else 0)  # Hitung sisa berdasarkan baris
                        sisa_terakhir = sisa_akhir

                        output_data.append({
                            'Tanggal': tanggal,
                            'Artikel': artikel,
                            'Perkotak': perkotak,
                            'Konversi': konversi,
                            'Masuk': masuk,
                            'Keluar': keluar,
                            'Sisa': sisa_akhir,
                        })
            dfksbb = pd.DataFrame(output_data)
            listdf.append(dfksbb)
            listproduk.append([produk,'WIP'])
        # print(dfksbb)
        # print(asd)
        listdata,saldoawal = calculate_KSBB(produk,tanggalmulai,tanggalakhir,'FG')
        # print(listdata == True)
        # print(asd)
        if len(listdata) != 0:
            # print(listdata)
            # print(asd)
            sisa_terakhir = 0
            output_data = []
            if saldoawal and  saldoawal.Jumlah != None:
                jumlahsaldoawal = saldoawal.Jumlah
            else:
                jumlahsaldoawal = 0
            output_data.append({
                        'Tanggal': tanggalmulai.strftime('%Y-%m-%d'),
                        'Artikel': None,
                        'Perkotak': None,
                        'Konversi':None,
                        'Masuk': None,
                        'Keluar': None,
                        'Sisa': jumlahsaldoawal,
                    })
            
            sisa_terakhir = jumlahsaldoawal

        # Memetakan setiap artikel dan perkotak per tanggal, dan menghitung Sisa
            for record in listdata:
                tanggal = record['Tanggal']
                artikel_list = record.get('Artikel', [])
                perkotak_list = record.get('Perkotak', [])
                konversi_list = record.get('Konversi', [])
                keluar_list = record.get('Keluar', [])
                masuk = record['Masuk']
                sisa_list = record.get('Sisa', [])

                # Jika ini adalah hari pertama, ambil Sisa awal dari data
                # print(konversi_list)
                # print(asd)
                if sisa_terakhir is None:
                    sisa_terakhir = sisa_list[0] if sisa_list else 0

                # Jika artikel, perkotak, dan konversi kosong, tetap memasukkan minimal satu entri
                if not artikel_list and not perkotak_list and not konversi_list:
                    sisa_awal = sisa_terakhir
                    sisa_akhir = sisa_awal + masuk - sum(keluar_list)  # Hitung sisa total
                    sisa_terakhir = sisa_akhir

                    output_data.append({
                        'Tanggal': tanggal,
                        'Artikel': None,
                        'Masuk': masuk,
                        'Perkotak': None,
                        'Konversi': None,
                        'Keluar': None,
                        'Sisa': sisa_akhir,
                    })
                else:
                    # Menggunakan zip_longest untuk memasukkan setiap artikel, perkotak, konversi, dan keluar
                    for artikel, perkotak, konversi, keluar, sisa in zip_longest(artikel_list, perkotak_list, konversi_list, keluar_list, sisa_list, fillvalue=None):
                        # Round konversi to 5 decimal places if it exists
                        konversi = round(konversi, 5) if konversi is not None else None
                        # print(konversi)
                        # Hitung Sisa: menggunakan sisa sebelumnya jika ada
                        sisa_awal = sisa_terakhir
                        sisa_akhir = sisa_awal + masuk - (keluar if keluar is not None else 0)  # Hitung sisa berdasarkan baris
                        sisa_terakhir = sisa_akhir

                        output_data.append({
                            'Tanggal': tanggal,
                            'Artikel': artikel,
                            'Masuk': masuk,
                            'Perkotak': perkotak,
                            'Konversi': konversi,
                            'Keluar': keluar,
                            'Sisa': sisa_akhir,
                        })
            dfksbbfg = pd.DataFrame(output_data)
            listdf.append(dfksbbfg)
            listproduk.append([produk,'FG'])

        # print(dfksbb)
        # print(dfksbbfg)
    # print(listdf)
    # print(listproduk)
    waktupersediaan = time.time()
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        for item,data in zip(listdf,listproduk):
            item.to_excel(writer, index=False, startrow=5, sheet_name=f"{data[0].KodeProduk}-{data[1]}")
            writer.sheets[f"{data[0].KodeProduk}-{data[1]}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {data[0].KodeProduk}")
            writer.sheets[f"{data[0].KodeProduk}-{data[1]}"].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {data[0].NamaProduk}")
            writer.sheets[f"{data[0].KodeProduk}-{data[1]}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {data[0].unit}")
            writer.sheets[f"{data[0].KodeProduk}-{data[1]}"].cell(row=4, column = 1,value =f"LOKASI : {data[1]}")
            maxrow = len(item)+1
            maxcol = len(item.columns)
            apply_number_format(writer.sheets[f"{data[0].KodeProduk}-{data[1]}"],6,maxrow+5,1,maxcol)
            apply_borders_thin(writer.sheets[f"{data[0].KodeProduk}-{data[1]}"],6,maxrow+5,maxcol)
            adjust_column_width(writer.sheets[f"{data[0].KodeProduk}-{data[1]}"],item,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBB Produksi {tahun}.xlsx"
    )
    
    print('Waktu generate laporan : ',time.time()-waktupersediaan)
    print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def eksportksbjsubkonperartikel(request,id,tahun):
    print(str)
    kodeprodukobj =models.ProdukSubkon.objects.get(pk = id)
    kodeproduksubkonperartikel = models.ProdukSubkon.objects.filter(KodeArtikel = kodeprodukobj.KodeArtikel)
    tanggalmulai = date(int(tahun),1,1)
    tanggalakhir = date(int(tahun),12,31)
    dfksbb = pd.DataFrame()
    dfksbbfg = pd.DataFrame()
    listdf = []
    listproduksubkon =[]
    print(kodeproduksubkonperartikel)
    # print(asd)
    for data in kodeproduksubkonperartikel:
        
        listdata,saldoawal = calculateksbjsubkon(data,tanggalmulai,tanggalakhir)
        print(listdata)
        datamodels = {
            'Tanggal':[],
            'Masuk' : [],
            'Keluar' : [],
            'Sisa' : []
        }
        if saldoawal:
            datamodels["Tanggal"].append(tanggalmulai)
            datamodels['Masuk'].append('')
            datamodels['Keluar'].append('')
            datamodels['Sisa'].append(saldoawal.Jumlah)

        for item in listdata:
            datamodels["Tanggal"].append(item['Tanggal'])
            datamodels['Masuk'].append(item['Masuk'])
            datamodels['Keluar'].append(item['Keluar'])
            datamodels['Sisa'].append(item['Sisa'])
        dfksbjsubkon = pd.DataFrame(datamodels)
        if not dfksbjsubkon.empty:
            listdf.append(dfksbjsubkon)
            listproduksubkon.append(data)
        print(data)
        print(dfksbjsubkon)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        for data,produk in zip(listdf,listproduksubkon):
            sheettitle = clean_string(produk.NamaProduk)

            data.to_excel(writer, index=False, startrow=5, sheet_name=f"{sheettitle}")
            writer.sheets[f"{sheettitle}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {produk.NamaProduk}")
            writer.sheets[f"{sheettitle}"].cell(row=2, column = 1,value =f"ARTIKEL PERUNTUKAN: {produk.KodeArtikel}")
            writer.sheets[f"{sheettitle}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {produk.Unit}")
            maxrow = len(data)+1
            maxcol = len(data.columns)
            apply_number_format(writer.sheets[f"{sheettitle}"],6,maxrow+5,1,maxcol)
            apply_borders_thin(writer.sheets[f"{sheettitle}"],6,maxrow+5,maxcol)
            adjust_column_width(writer.sheets[f"{sheettitle}"],data,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBJ Subkon {kodeprodukobj.KodeArtikel}.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response
    
def eksportksbjsubkon(request,id,tahun):
    print(str)
    kodeprodukobj =models.ProdukSubkon.objects.get(pk = id)
    tanggalmulai = date(int(tahun),1,1)
    tanggalakhir = date(int(tahun),12,31)
    dfksbb = pd.DataFrame()
    dfksbbfg = pd.DataFrame()
    print
    listdata,saldoawal = calculateksbjsubkon(kodeprodukobj,tanggalmulai,tanggalakhir)
    print(listdata)
    datamodels = {
        'Tanggal':[],
        'Masuk' : [],
        'Keluar' : [],
        'Pemusnahan' :[],
        'Sisa' : [],
    }
    if saldoawal:
        datamodels["Tanggal"].append(tanggalmulai)
        datamodels['Masuk'].append('')
        datamodels['Keluar'].append('')
        datamodels['Pemusnahan'].append('')
        datamodels['Sisa'].append(saldoawal.Jumlah)

    for item in listdata:
        print(item)
        datamodels["Tanggal"].append(item['Tanggal'])
        datamodels['Masuk'].append(item['Masuk'])
        datamodels['Keluar'].append(item['Keluar'])
        datamodels['Pemusnahan'].append(item['Pemusnahan'])
        datamodels['Sisa'].append(item['Sisa'])
    dfksbjsubkon = pd.DataFrame(datamodels)
    buffer = BytesIO()

    if dfksbjsubkon.empty:
        messages.error(request,f'Tidak dapat mengeksport, tidak ada data transaksi')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return('ksbjsubkon')
    sheettitle = clean_string(kodeprodukobj.NamaProduk)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfksbjsubkon.to_excel(writer, index=False, startrow=5, sheet_name=f"{sheettitle}")
        writer.sheets[f"{sheettitle}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {kodeprodukobj.NamaProduk}")
        writer.sheets[f"{sheettitle}"].cell(row=2, column = 1,value =f"ARTIKEL PERUNTUKAN: {kodeprodukobj.KodeArtikel}")
        writer.sheets[f"{sheettitle}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {kodeprodukobj.Unit}")
        maxrow = len(dfksbjsubkon)+1
        maxcol = len(dfksbjsubkon.columns)
        apply_number_format(writer.sheets[f"{sheettitle}"],6,maxrow+5,1,maxcol)
        apply_borders_thin(writer.sheets[f"{sheettitle}"],6,maxrow+5,maxcol)
        adjust_column_width(writer.sheets[f"{sheettitle}"],dfksbjsubkon,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBJ Subkon {kodeprodukobj.NamaProduk} - {kodeprodukobj.KodeArtikel}.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response
    
def eksportpemusnahanbahanbaku (request,tanggalawal,tanggalakhir):
    print(tanggalawal)
    print(tanggalakhir)
    pemusnahanobj = models.PemusnahanBahanBaku.objects.filter(Tanggal__range = (tanggalawal,tanggalakhir),lokasi__NamaLokasi__in=('WIP','FG')).order_by('Tanggal')
    datamodels ={
        'Tanggal' : [],
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan': [],
        'Lokasi' : [],
        'Jumlah' : [],
        'Keterangan' : [],
    }
    for data in pemusnahanobj:
        datamodels['Tanggal'].append(data.Tanggal.strftime('%Y-%m-%d'))
        datamodels['Kode Bahan Baku'].append(data.KodeBahanBaku.KodeProduk)
        datamodels['Nama Bahan Baku'].append(data.KodeBahanBaku.NamaProduk)
        datamodels['Satuan'].append(data.KodeBahanBaku.unit)
        datamodels['Lokasi'].append(data.lokasi)
        datamodels['Jumlah'].append(data.Jumlah)
        datamodels['Keterangan'].append(data.Keterangan)
    dfpemsunahan = pd.DataFrame(datamodels)
    buffer = BytesIO()

    if dfpemsunahan.empty:
        messages.error(request,f'Tidak dapat mengeksport, tidak ada data transaksi')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return('view_pemusnahanbarang')
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfpemsunahan.to_excel(writer, index=False, startrow=2, sheet_name=f"Pemusnahan Bahan Baku Produksi")
        writer.sheets[f"Pemusnahan Bahan Baku Produksi"].cell(row=1, column = 1,value =f"PEMUSNAHAN BAHAN BAKU PRODUKSI")
        writer.sheets[f"Pemusnahan Bahan Baku Produksi"].cell(row=2, column = 1,value =f"PERIODE : {tanggalawal} - {tanggalakhir}")


        maxrow = len(dfpemsunahan)+1
        maxcol = len(dfpemsunahan.columns)
        apply_number_format(writer.sheets[f"Pemusnahan Bahan Baku Produksi"],3,maxrow+2,1,maxcol)
        apply_borders_thin(writer.sheets[f"Pemusnahan Bahan Baku Produksi"],3,maxrow+2,maxcol)
        adjust_column_width(writer.sheets[f"Pemusnahan Bahan Baku Produksi"],dfpemsunahan,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=Pemusnahan Bahan Baku Produksi.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response
def eksportpemusnahanbahanbakusubkon (request,tanggalawal,tanggalakhir):
    pemusnahanobj = models.PemusnahanBahanBakuSubkon.objects.filter(Tanggal__range = (tanggalawal,tanggalakhir),lokasi__NamaLokasi__in=('WIP','FG')).order_by('Tanggal')
    datamodels ={
        'Tanggal' : [],
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan': [],
        'Lokasi' : [],
        'Jumlah' : [],
        'Keterangan' : [],
    }
    for data in pemusnahanobj:
        datamodels['Tanggal'].append(data.Tanggal.strftime('%Y-%m-%d'))
        datamodels['Kode Bahan Baku'].append(data.KodeBahanBaku.KodeProduk)
        datamodels['Nama Bahan Baku'].append(data.KodeBahanBaku.NamaProduk)
        datamodels['Satuan'].append(data.KodeBahanBaku.unit)
        datamodels['Lokasi'].append(data.lokasi)
        datamodels['Jumlah'].append(data.Jumlah)
        datamodels['Keterangan'].append(data.Keterangan)
    dfpemsunahan = pd.DataFrame(datamodels)
    buffer = BytesIO()

    if dfpemsunahan.empty:
        messages.error(request,f'Tidak dapat mengeksport, tidak ada data transaksi')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return('view_pemusnahanbarangsubkon')
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfpemsunahan.to_excel(writer, index=False, startrow=2, sheet_name=f"Pemusnahan Bahan Baku Subkon")
        writer.sheets[f"Pemusnahan Bahan Baku Subkon"].cell(row=1, column = 1,value =f"PEMUSNAHAN BAHAN BAKU SUBKON")
        writer.sheets[f"Pemusnahan Bahan Baku Subkon"].cell(row=2, column = 1,value =f"PERIODE : {tanggalawal} - {tanggalakhir}")


        maxrow = len(dfpemsunahan)+1
        maxcol = len(dfpemsunahan.columns)
        apply_number_format(writer.sheets[f"Pemusnahan Bahan Baku Subkon"],3,maxrow+2,1,maxcol)
        apply_borders_thin(writer.sheets[f"Pemusnahan Bahan Baku Subkon"],3,maxrow+2,maxcol)
        adjust_column_width(writer.sheets[f"Pemusnahan Bahan Baku Subkon"],dfpemsunahan,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=Pemusnahan Bahan Baku Subkon.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def eksportpemusnahanproduksubkon (request,tanggalawal,tanggalakhir):
    pemusnahanobj = models.PemusnahanProdukSubkon.objects.filter(Tanggal__range = (tanggalawal,tanggalakhir)).order_by('Tanggal')
    datamodels ={
        'Tanggal' : [],
        'Artikel Peruntukan' : [],
        'Nama Bahan Baku' : [],
        'Satuan': [],
        'Jumlah' : [],
        'Keterangan' : [],
    }
    for data in pemusnahanobj:
        datamodels['Tanggal'].append(data.Tanggal.strftime('%Y-%m-%d'))
        datamodels['Artikel Peruntukan'].append(data.KodeProdukSubkon.KodeArtikel.KodeArtikel)
        datamodels['Nama Bahan Baku'].append(data.KodeProdukSubkon.NamaProduk)
        datamodels['Satuan'].append(data.KodeProdukSubkon.Unit)
        datamodels['Jumlah'].append(data.Jumlah)
        datamodels['Keterangan'].append(data.Keterangan)
    dfpemsunahan = pd.DataFrame(datamodels)
    buffer = BytesIO()

    if dfpemsunahan.empty:
        messages.error(request,f'Tidak dapat mengeksport, tidak ada data transaksi')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return('view_pemusnahanproduksubkon')
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfpemsunahan.to_excel(writer, index=False, startrow=2, sheet_name=f"Pemusnahan Produk Subkon")
        writer.sheets[f"Pemusnahan Produk Subkon"].cell(row=1, column = 1,value =f"PEMUSNAHAN PRODUK SUBKON")
        writer.sheets[f"Pemusnahan Produk Subkon"].cell(row=2, column = 1,value =f"PERIODE : {tanggalawal} - {tanggalakhir}")


        maxrow = len(dfpemsunahan)+1
        maxcol = len(dfpemsunahan.columns)
        apply_number_format(writer.sheets[f"Pemusnahan Produk Subkon"],3,maxrow+2,1,maxcol)
        apply_borders_thin(writer.sheets[f"Pemusnahan Produk Subkon"],3,maxrow+2,maxcol)
        adjust_column_width(writer.sheets[f"Pemusnahan Produk Subkon"],dfpemsunahan,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=Pemusnahan Produk Subkon.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def eksportpemusnahanartikel (request,tanggalawal,tanggalakhir):
    pemusnahanobj = models.PemusnahanArtikel.objects.filter(Tanggal__range = (tanggalawal,tanggalakhir),lokasi__NamaLokasi__in=('WIP','FG')).order_by('Tanggal')
    datamodels ={
        'Tanggal' : [],
        'Kode Artikel' : [],
        'Versi' : [],
        'Lokasi' : [],
        'Jumlah' : [],
        'Keterangan' : [],
    }
    for data in pemusnahanobj:
        datamodels['Tanggal'].append(data.Tanggal.strftime('%Y-%m-%d'))
        datamodels['Kode Artikel'].append(data.KodeArtikel.KodeArtikel)
        datamodels['Versi'].append(data.VersiArtikel.Versi)
        datamodels['Lokasi'].append(data.lokasi)
        datamodels['Jumlah'].append(data.Jumlah)
        datamodels['Keterangan'].append(data.Keterangan)
    dfpemsunahan = pd.DataFrame(datamodels)
    buffer = BytesIO()

    if dfpemsunahan.empty:
        messages.error(request,f'Tidak dapat mengeksport, tidak ada data transaksi')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return('ksbjsubkon')
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfpemsunahan.to_excel(writer, index=False, startrow=2, sheet_name=f"Pemusnahan Artikel Produksi")
        writer.sheets[f"Pemusnahan Artikel Produksi"].cell(row=1, column = 1,value =f"PEMUSNAHAN ARTIKEL PRODUKSI")
        writer.sheets[f"Pemusnahan Artikel Produksi"].cell(row=2, column = 1,value =f"PERIODE : {tanggalawal} - {tanggalakhir}")


        maxrow = len(dfpemsunahan)+1
        maxcol = len(dfpemsunahan.columns)
        apply_number_format(writer.sheets[f"Pemusnahan Artikel Produksi"],3,maxrow+2,1,maxcol)
        apply_borders_thin(writer.sheets[f"Pemusnahan Artikel Produksi"],3,maxrow+2,maxcol)
        adjust_column_width(writer.sheets[f"Pemusnahan Artikel Produksi"],dfpemsunahan,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=Pemusnahan Artikel Produksi.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def apply_number_format(worksheet, start_row, end_row, start_col, end_col, number_format='#,##0.00'):
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Apply format only to numeric cells
                cell.number_format = number_format

def adjust_column_width(worksheet, df, start_row, start_col):
    for i, col in enumerate(df.columns):
        max_length = max(df[col].astype(str).map(len).max(), len(col))
        col_letter = get_column_letter(start_col + i)
        worksheet.column_dimensions[col_letter].width = max_length + 4

def apply_borders_thin(worksheet, start_row, end_row, max_col,min_col=1):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def eksportksbjproduksi(request,id,lokasi,tahun):
    # waktuawalproses = time.time()
    tahun = int(tahun)
    kodeartikelobj =models.Artikel.objects.get(pk = id)
    tanggalmulai = date(int(tahun),1,1)
    tanggalakhir = date(int(tahun),12,31)
    lokasi = "WIP"
    listdata,saldoawal = calculate_ksbj(kodeartikelobj,lokasi,tahun)
    print(listdata)
    print(kodeartikelobj.KodeArtikel)
    # print(asd)
    # waktupersediaan = time.time()
    kodeartikel = clean_string(kodeartikelobj.KodeArtikel)
    datamodelsksbj ={
 
        "Tanggal":[],
        "SPK (Keluar)":[],
        "Kode Stok":[],
        "Masuk (Lembar)":[],
        "Masuk (Konversi)":[],
        "Keluar (FG)":[],
        "Pemusnahan":[],
        "Saldo":[],
    }
    print(saldoawal)
    print(listdata)
    if saldoawal:
        # print(saldoawal)
        try:
            if saldoawal.Tanggal:
                datamodelsksbj["Tanggal"].append(tanggalmulai.strftime('%Y-%m-%d'))
                datamodelsksbj['SPK (Keluar)'].append('')
                datamodelsksbj['Kode Stok'].append('')
                datamodelsksbj['Masuk (Lembar)'].append('')
                datamodelsksbj["Masuk (Konversi)"].append('')
                datamodelsksbj['Keluar (FG)'].append('')
                datamodelsksbj['Pemusnahan'].append('')
                datamodelsksbj["Saldo"].append(saldoawal.Jumlah)
        except:
            pass
    for item in listdata:

        datamodelsksbj["Tanggal"].append(item['Tanggal'])
        dummy = []
        for data in item['SPK']:
            if data.DetailSPK == None:
                dummy.append('')
            else:
                dummy.append(data.DetailSPK.NoSPK.NoSPK)
        spk_string = ', '.join(dummy)  # Gabung elemen-elemen list menjadi string
        datamodelsksbj['SPK (Keluar)'].append(spk_string)
        datamodelsksbj['Kode Stok'].append(item['Kodeproduk'].KodeProduk)
        datamodelsksbj['Masuk (Lembar)'].append(item['Masuklembar'])
        datamodelsksbj["Masuk (Konversi)"].append(item['Masukkonversi'])
        datamodelsksbj['Keluar (FG)'].append(item['Hasil'])
        datamodelsksbj['Pemusnahan'].append(item['Keluar'])
        datamodelsksbj["Saldo"].append(item['Sisa'])
    dfksbj = pd.DataFrame(datamodelsksbj)
    print(dfksbj)

    lokasi = "FG"
    listdata,saldoawal = calculate_ksbj(kodeartikelobj,lokasi,tahun)
    # print(listdata)
    # print(listdata[0].keys())
    # print(asd)
    datamodelsksbjfg = {
        'Tanggal': [],
        'Masuk (WIP)': [],
        "SPK (Keluar)":[],
        'SPPB' : [],
        'Keluar (Kirim)' : [],
        'Pemusnahan' : [],
        'Saldo': [],
    }
    sisaakhir = 0
    if saldoawal:
        print(saldoawal)
        try:
            if saldoawal.Tanggal:
                datamodelsksbjfg["Tanggal"].append(tanggalmulai.strftime('%Y-%m-%d'))
                datamodelsksbjfg["SPK (Keluar)"].append('')
                datamodelsksbjfg["Masuk (WIP)"].append('')
                datamodelsksbjfg["Keluar (Kirim)"].append('')
                datamodelsksbjfg["SPPB"].append('')
                datamodelsksbjfg["Pemusnahan"].append('')
                datamodelsksbjfg["Saldo"].append(saldoawal.Jumlah)
        except:
            pass
    for item in listdata:
        index = 0
        if len(item['DetailSPPB']) > 1:
            for data in item['DetailSPPB']:
                # print(item)
                # print(sisaakhir)
                # print(data.Jumlah)
                # print(asd)
                datamodelsksbjfg["Tanggal"].append(item['Tanggal'])
                if index == 0:
                    jumlahkirim = data.Jumlah
                    jumlahkeluar = item['Keluar']
                    jumlahpenyerahan = item['Penyerahanwip']
                    datamodelsksbjfg["Masuk (WIP)"].append(jumlahpenyerahan)
                    datamodelsksbjfg["Pemusnahan"].append(jumlahkeluar)
                    datamodelsksbjfg["Keluar (Kirim)"].append(jumlahkirim)
                    sisaakhir += jumlahpenyerahan - jumlahkeluar - jumlahkirim
                else:
                    jumlahkirim = data.Jumlah
                    jumlahkeluar = 0
                    jumlahpenyerahan = 0
                    datamodelsksbjfg["Masuk (WIP)"].append(0)
                    datamodelsksbjfg["Pemusnahan"].append(0)
                    datamodelsksbjfg["Keluar (Kirim)"].append(data.Jumlah)
                    sisaakhir -= jumlahkirim
                # sisaakhir += jumlahpenyerahan - jumlahkeluar - jumlahkeluar
                datamodelsksbjfg["Saldo"].append(sisaakhir)
                datamodelsksbjfg["SPPB"].append(data.NoSPPB.NoSPPB)
                datamodelsksbjfg["SPK (Keluar)"].append(data.DetailSPK)
                # datamodelsksbjfg["Keluar (Kirim)"].append(data['Tanggal'])
                index+=1
        else:
            datamodelsksbjfg["Tanggal"].append(item['Tanggal'])
            datamodelsksbjfg["Masuk (WIP)"].append(item['Penyerahanwip'])
            datamodelsksbjfg["Pemusnahan"].append(item['Keluar'])
            datamodelsksbjfg["Keluar (Kirim)"].append(item['Jumlahkirim'])
            datamodelsksbjfg["Saldo"].append(item['Sisa'])
            if len(item['DetailSPPB']) == 0:

                datamodelsksbjfg["SPPB"].append('-')
                datamodelsksbjfg["SPK (Keluar)"].append('-')
            else:
                datamodelsksbjfg["SPPB"].append(item['DetailSPPB'][0].NoSPPB)
                datamodelsksbjfg["SPK (Keluar)"].append(item['DetailSPPB'][0].DetailSPK)
        sisaakhir = item['Sisa']
    print(datamodelsksbjfg)
    print(len(datamodelsksbjfg['SPK (Keluar)']))
    print(len(datamodelsksbjfg['SPPB']))
    print(len(datamodelsksbjfg['Tanggal']))
    print(len(datamodelsksbjfg['Keluar (Kirim)']))
    print(len(datamodelsksbjfg['Masuk (WIP)']))
    print(len(datamodelsksbjfg['Pemusnahan']))
    dfksbjfg = pd.DataFrame(datamodelsksbjfg)
    # print(dfksbjfg)
    if dfksbj.empty and dfksbjfg.empty:
        messages.error(request,'Tidak dapat menemukan data transaksi untuk di eksport')
        if 'HTTP_REFERER' in request.META:
            back_url = request.META['HTTP_REFERER']
            return redirect(back_url)
        else:
            return redirect('view_ksbj')
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfksbj.to_excel(writer, index=False, startrow=5, sheet_name=f"WIP")
        writer.sheets[f"WIP"].cell(row=1, column = 1,value =f"KARTU STOK Artikel : {kodeartikel}")
        maxrow = len(dfksbj)+1
        maxcol = len(dfksbj.columns)
        apply_number_format(writer.sheets[f"WIP"],6,maxrow+5,1,maxcol)
        apply_borders_thin(writer.sheets[f"WIP"],6,maxrow+5,maxcol)
        adjust_column_width(writer.sheets[f"WIP"],dfksbj,1,1)

        dfksbjfg.to_excel(writer, index=False, startrow=5, sheet_name=f"FG")
        writer.sheets[f"FG"].cell(row=1, column = 1,value =f"KARTU STOK Artikel : {kodeartikel}")
        maxrow = len(dfksbjfg)+1
        maxcol = len(dfksbjfg.columns)
        apply_number_format(writer.sheets[f"FG"],6,maxrow+5,1,maxcol)
        apply_borders_thin(writer.sheets[f"FG"],6,maxrow+5,maxcol)
        adjust_column_width(writer.sheets[f"FG"],dfksbjfg,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBJ {kodeartikel}.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response
    # waktuawalproses = time.time()
    print(str)
    kodeprodukobj =models.Produk.objects.get(KodeProduk = id)
    tanggalmulai = date(int(tahun),1,1)
    tanggalakhir = date(int(tahun),12,31)

    listdata,saldoawal = calculate_KSBB(kodeprodukobj,tanggalmulai,tanggalakhir,lokasi)
    print(listdata)
    print(asd)
    # waktupersediaan = time.time()
    datamodelsksbb ={
 
        "Tanggal":[],
        "Kuantitas Masuk":[],
        "Harga Satuan Masuk":[],
        "Harga Total Masuk":[],
        "Kuantitas Keluar":[],
        "Harga Satuan keluar":[],
        "Harga Total keluar":[],
        "Kuantitas Sisa":[],
        "Harga Satuan Sisa":[],
        "Harga Total Sisa":[],
    }
    for item in listdata:
        # <tr>
        #                         <td>{{i.Tanggal}}</td>
        #                         <td>{{i.Jumlahmasuk|separator_ribuan}}</td>
        #                         <td>{{i.Hargamasuksatuan|custom_thousands_separator}}</td>
        #                         <td>{{i.Hargamasuktotal|custom_thousands_separator}}</td>
        #                         <td>{{i.Jumlahkeluar|separator_ribuan}}</td>
        #                         <td>{{i.Hargakeluarsatuan|custom_thousands_separator}}</td>
                                # <td>{{i.Hargakeluartotal|custom_thousands_separator}}</td>
        #                         <td>{{i.Sisahariini|separator_ribuan}}</td>
        #                         <td>{{i.Hargasatuansisa|custom_thousands_separator}}</td>
        #                         <td>{{i.Hargatotalsisa|custom_thousands_separator}}</td>

        #                     </tr>
        datamodelsksbb["Tanggal"].append(item['Tanggal'])
        datamodelsksbb["Kuantitas Masuk"].append(item['Jumlahmasuk'])
        datamodelsksbb['Harga Satuan Masuk'].append(item['Hargamasuksatuan'])
        datamodelsksbb['Harga Total Masuk'].append(item['Hargamasuktotal'])
        datamodelsksbb["Kuantitas Keluar"].append(item['Jumlahkeluar'])
        datamodelsksbb['Harga Satuan keluar'].append(item['Hargakeluarsatuan'])
        datamodelsksbb['Harga Total keluar'].append(item['Hargakeluartotal'])
        datamodelsksbb["Kuantitas Sisa"].append(item['Sisahariini'])
        datamodelsksbb['Harga Satuan Sisa'].append(item['Hargasatuansisa'])
        datamodelsksbb['Harga Total Sisa'].append(item['Hargatotalsisa'])
    dfksbb = pd.DataFrame(datamodelsksbb)
    print(dfksbb)

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfksbb.to_excel(writer, index=False, startrow=5, sheet_name=f"{kodeprodukobj.KodeProduk}")
        writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {kodeprodukobj.KodeProduk}")
        writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {kodeprodukobj.NamaProduk}")
        writer.sheets[f"{kodeprodukobj.KodeProduk}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {kodeprodukobj.unit}")
        maxrow = len(dfksbb)+1
        maxcol = len(dfksbb.columns)
        apply_number_format(writer.sheets[f"{kodeprodukobj.KodeProduk}"],6,maxrow+5,1,maxcol)
        apply_borders_thin(writer.sheets[f"{kodeprodukobj.KodeProduk}"],6,maxrow+5,maxcol)
        adjust_column_width(writer.sheets[f"{kodeprodukobj.KodeProduk}"],dfksbb,1,1)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBB {kodeprodukobj.KodeProduk}.xlsx"
    )
    
    print('Waktu generate laporan : ',time.time()-waktupersediaan)
    print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def rekapakumulasiksbbsubkon(request,id):
    produk = models.BahanBakuSubkon.objects.get(pk = id)
    if len(request.GET) == 0:
        return render(request,'produksi/view_rekapksbbsubkon.html',{'produk':produk})
    else:
        '''
        Logika rekapitulasi 
        '''
        tanggalmulai = request.GET['tanggalawal']
        tanggalakhir = request.GET['tanggalakhir']
        produkobj = models.BahanBakuSubkon.objects.get(pk = id)
        pemusnahanobj = models.PemusnahanBahanBakuSubkon.objects.filter(KodeBahanBaku = produkobj,Tanggal__range = (tanggalmulai,tanggalakhir))
        jumlahpemusnahan = pemusnahanobj.aggregate(total = Sum('Jumlah'))['total']
        

        listdata,saldoawal = calculateksbbsubkon(produkobj,tanggalmulai,tanggalakhir)
        masuk = 0
        keluar = 0
        print(listdata)
        for item in listdata:
            masuk += item['Masuk']
            keluar += (item['Keluar'])
        
        return render(request,'produksi/view_rekapksbbsubkon.html',{'produk':produk,'jumlahpemusnahan':jumlahpemusnahan,'masuk':masuk,'keluar':keluar,'tanggalawal':tanggalmulai,'tanggalakhir':tanggalakhir})

def rekapakumulasiksbb(request,id,lokasi):
    if len(request.GET) == 0:
        return render(request,'produksi/view_rekapksbb.html')
    else:
        '''
        Logika rekapitulasi 
        '''
        tanggalmulai = request.GET['tanggalawal']
        tanggalakhir = request.GET['tanggalakhir']
        produkobj = models.Produk.objects.get(KodeProduk = id)
        pemusnahanbahanbakuobj = models.PemusnahanBahanBaku.objects.filter(KodeBahanBaku = produkobj,lokasi__NamaLokasi="WIP",Tanggal__range=(tanggalmulai,tanggalakhir))
        pemusnahanbahanbakuobjfg = models.PemusnahanBahanBaku.objects.filter(KodeBahanBaku = produkobj,lokasi__NamaLokasi="FG",Tanggal__range=(tanggalmulai,tanggalakhir))
        jumlahpemusnahwip = pemusnahanbahanbakuobj.aggregate(total = Sum('Jumlah'))['total']
        jumlahpemusnahanfg = pemusnahanbahanbakuobjfg.aggregate(total=Sum('Jumlah'))['total']

        if jumlahpemusnahanfg == None:
            jumlahpemusnahanfg = 0
        if jumlahpemusnahwip == None:
            jumlahpemusnahwip = 0
        listdata,saldoawal = calculate_KSBB(produkobj,tanggalmulai,tanggalakhir,lokasi)
        masuk = 0
        keluar = 0
        for item in listdata:
            masuk += item['Masuk']
            keluar += sum(item['Keluar'])
            print(item)
        print(pemusnahanbahanbakuobjfg)
        return render(request,'produksi/view_rekapksbb.html',{'jumlahpemusnahanwip':jumlahpemusnahwip,'jumlahpemusnahanfg':jumlahpemusnahanfg,'masuk':masuk,'keluar':keluar,'tanggalawal':tanggalmulai,'tanggalakhir':tanggalakhir,'produk':produkobj})

def rekapitulasiksbj (request,id,lokasi):
    if len(request.GET)==0:
        return render(request,'produksi/view_rekapksbj.html')
    else:
        tanggalmulai = request.GET['tanggalawal']
        tanggalakhir = request.GET['tanggalakhir']
        Artikelobj = models.Artikel.objects.get(pk = id)
        pemusnahanobjwip = models.PemusnahanArtikel.objects.filter(KodeArtikel = Artikelobj,Tanggal__range=(tanggalmulai,tanggalakhir),lokasi__NamaLokasi = 'WIP')
        pemusnahanobjfg = models.PemusnahanArtikel.objects.filter(KodeArtikel = Artikelobj,Tanggal__range=(tanggalmulai,tanggalakhir),lokasi__NamaLokasi = 'FG')
        totalpemusnahanwip = pemusnahanobjwip.aggregate(total = Sum('Jumlah'))['total']
        totalpemusnahanfg = pemusnahanobjfg.aggregate(total = Sum('Jumlah'))['total']
        if totalpemusnahanwip == None:
            totalpemusnahanwip = 0
        if totalpemusnahanfg == None:
            totalpemusnahanfg = 0
        tanggalmulaidatetime = datetime.strptime(tanggalmulai,'%Y-%m-%d')
        tanggalakhirdatetime = datetime.strptime(tanggalakhir,'%Y-%m-%d')
        listdata,saldoawal = calculate_ksbj(Artikelobj,lokasi,tanggalmulaidatetime.year)
        print(listdata)
        masuk = 0
        keluar = 0
        # for item in listdata:
        #     masuk += item['Masukkonversi']
        #     keluar += item['Keluar'] + item['Hasil']
            
        
        print(masuk)
        print(keluar)
        for item in listdata:
            print(item.keys())  
            tanggaldata = datetime.strptime(item['Tanggal'], "%Y-%m-%d")
            
            if tanggalmulaidatetime <= tanggaldata <= tanggalakhirdatetime:
                if lokasi == 'WIP':
                    masuk += item['Masukkonversi']
                    keluar += item['Keluar'] + item['Hasil']
                else:
                    masuk += item['Penyerahanwip']
                    keluar += item['Jumlahkirim'] 
                print(tanggaldata)
            
            elif tanggaldata < tanggalmulaidatetime:
                continue
            elif tanggaldata > tanggalakhirdatetime:
                break
        if lokasi == 'WIP':
            keluar -= totalpemusnahanwip
        else:
            keluar -= totalpemusnahanfg
        print(keluar)
        print(pemusnahanobjwip)

        return render(request,'produksi/view_rekapksbj.html',{'totalpemusnahanfg':totalpemusnahanfg,'totalpemusnahanwip':totalpemusnahanwip,'masuk':masuk,'keluar':keluar,'tanggalawal':tanggalmulai,'tanggalakhir':tanggalakhir,'kodeartikel':Artikelobj})

def rekapitulasiksbjsubkon (request,id):
    produk = models.ProdukSubkon.objects.get(pk = id)
    if len(request.GET)==0:
        return render(request,'produksi/view_rekapksbjsubkon.html',{'produk':produk})
    else:
        tanggalmulai = request.GET['tanggalawal']
        tanggalakhir = request.GET['tanggalakhir']
        Artikelobj = models.ProdukSubkon.objects.get(pk = id)
        tanggalmulaidatetime = datetime.strptime(tanggalmulai,'%Y-%m-%d')
        tanggalakhirdatetime = datetime.strptime(tanggalakhir,'%Y-%m-%d')
        pemusnahanobj = models.PemusnahanProdukSubkon.objects.filter(KodeProdukSubkon = Artikelobj,Tanggal__range = (tanggalmulai,tanggalakhir))
        jumlahpemusnahan = pemusnahanobj.aggregate(total = Sum('Jumlah'))['total']
        if jumlahpemusnahan == None:
            jumlahpemusnahan = 0
        listdata,saldoawal = calculateksbjsubkon(Artikelobj,tanggalmulai,tanggalakhir)
        print(listdata)
        masuk = 0
        keluar = 0
        for item in listdata:
            masuk += item['Masuk']
            keluar += item['Keluar']
            
        
        print(masuk)
        print(keluar)
        return render(request,'produksi/view_rekapksbjsubkon.html',{'produk':produk,'jumlahpemusnahan':jumlahpemusnahan,'masuk':masuk,'keluar':keluar,'tanggalawal':tanggalmulai,'tanggalakhir':tanggalakhir})
def eksportksbbsubkonkeseluruhan(request,id,tahun):
    '''
    Logika rekapitulasi 
    '''
    tahun= int(tahun)
    tanggalmulai = date(tahun,1,1)
    tanggalakhir = date(tahun,12,31)
    dfksbbsubkon = pd.DataFrame()

    produkobj = models.BahanBakuSubkon.objects.all()
    listdf = []
    listproduk = []
    for item in produkobj:
        saldoawalsubkon = models.SaldoAwalBahanBakuSubkon.objects.filter(IDBahanBakuSubkon= item, )
        listdata,saldoawal = calculateksbbsubkon(item,tanggalmulai,tanggalakhir)
        print(listdata)
        datamodels = {
            'Tanggal': [],
            'Masuk' : [],
            'Keluar' : [],
            'Sisa' : [],
        }
        print(saldoawal)
        if saldoawal:
            print(saldoawal.Tanggal)
            datamodels['Tanggal'].append(tanggalmulai)
            datamodels['Masuk'].append('')
            datamodels['Keluar'].append('')
            datamodels['Sisa'].append(saldoawal.Jumlah)
            # print(asd)
        for data in listdata:
            datamodels['Tanggal'].append(data['Tanggal'])
            datamodels['Masuk'].append(data['Masuk'])
            datamodels['Keluar'].append(data['Keluar'])
            datamodels['Sisa'].append(data['Sisa'])
        
        dfksbbsubkon = pd.DataFrame(datamodels)
        if dfksbbsubkon.empty:
            continue
        listdf.append(dfksbbsubkon)
        item.KodeProduk = clean_string(item.KodeProduk)
        listproduk.append(item)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        for item,produk in zip(listdf,listproduk):
            item.to_excel(writer, index=False, startrow=5, sheet_name=f"{produk.KodeProduk}")
            writer.sheets[f"{produk.KodeProduk}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {produk.KodeProduk}")
            writer.sheets[f"{produk.KodeProduk}"].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {produk.NamaProduk}")
            writer.sheets[f"{produk.KodeProduk}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {produk.unit}")
            maxrow = len(item)+1
            maxcol = len(item.columns)
            apply_number_format(writer.sheets[f"{produk.KodeProduk}"],6,maxrow+5,1,maxcol)
            apply_borders_thin(writer.sheets[f"{produk.KodeProduk}"],6,maxrow+5,maxcol)
            adjust_column_width(writer.sheets[f"{produk.KodeProduk}"],item,1,1)
    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBB Subkon {tahun}.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def eksportksbbsubkon(request,id,tahun):
    '''
    Logika rekapitulasi 
    '''
    tahun= int(tahun)
    tanggalmulai = date(tahun,1,1)
    tanggalakhir = date(tahun,12,31)
    dfksbbsubkon = pd.DataFrame()

    produkobj = models.BahanBakuSubkon.objects.get(pk = id)
    pemusnahanobj = models.PemusnahanBahanBakuSubkon.objects.filter(KodeBahanBaku = produkobj,Tanggal__range = (tanggalmulai,tanggalakhir))
    listdata,saldoawal = calculateksbbsubkon(produkobj,tanggalmulai,tanggalakhir)
    print(listdata)
    datamodels = {
        'Tanggal': [],
        'Masuk' : [],
        'Keluar' : [],
        'Sisa' : [],
    }
    for item in listdata:
        datamodels['Tanggal'].append(item['Tanggal'])
        datamodels['Masuk'].append(item['Masuk'])
        datamodels['Keluar'].append(item['Keluar'])
        datamodels['Sisa'].append(item['Sisa'])
    
    dfksbbsubkon = pd.DataFrame(datamodels)
    buffer = BytesIO()
    produkobj.KodeProduk = clean_string(produkobj.KodeProduk)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfksbbsubkon.to_excel(writer, index=False, startrow=5, sheet_name=f"{produkobj.KodeProduk}")
        writer.sheets[f"{produkobj.KodeProduk}"].cell(row=1, column = 1,value =f"KARTU STOK BAHAN BAKU : {produkobj.KodeProduk}")
        writer.sheets[f"{produkobj.KodeProduk}"].cell(row=2, column = 1,value =f"NAMA BAHAN BAKU : {produkobj.NamaProduk}")
        writer.sheets[f"{produkobj.KodeProduk}"].cell(row=3, column = 1,value =f"SATUAN BAHAN BAKU : {produkobj.unit}")
        maxrow = len(dfksbbsubkon)+1
        maxcol = len(dfksbbsubkon.columns)
        apply_number_format(writer.sheets[f"{produkobj.KodeProduk}"],6,maxrow+5,1,maxcol)
        apply_borders_thin(writer.sheets[f"{produkobj.KodeProduk}"],6,maxrow+5,maxcol)
        adjust_column_width(writer.sheets[f"{produkobj.KodeProduk}"],dfksbbsubkon,1,1)
    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=KSBB Subkon {produkobj.KodeProduk}.xlsx"
    )
    
    # print('Waktu generate laporan : ',time.time()-waktupersediaan)
    # print('Waktu Proses : ', time.time()-waktuawalproses)
    return response
def bulkcreate_transaksiproduksi(request):
    '''
    Input dari KSBB Produksi GOLONGAN A
    '''
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        # print(asd)
        excel_file = pd.ExcelFile(file)
        wb = openpyxl.load_workbook(file, data_only=True)
        

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        produkerror = []

        for item in sheet_names:
            sheet = wb[item]
    
   
            if sheet.sheet_state != 'visible':
                continue
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=6)
            print(item)
            print(df)
            # print(asd)

            i = 0
            tanggal = None
            listtanggal = []
            for index, row in df.iterrows():
                if i == 0:
                    i+=1
                    continue
            
                print(row)
                print(item)
                # print(asd)
                if not pd.isna(row["Tanggal"]):
                    listtanggal.append(row['Tanggal'])
                try:
                    
                    if  pd.isna(row['Unnamed: 3']):
                        print(f"Index {index}: Unnamed: 3 adalah number")
                    else:
                        try:

                            if pd.isna(row['Tanggal']):
                                tanggal = listtanggal[-1]
                            else:
                                tanggal = row['Tanggal']
                            if not pd.isna(row['Keterangan']):
                                keterangan = clean_string(row['Keterangan'])
                            print(tanggal)
                            existing_transaction = models.TransaksiProduksi.objects.filter(
                                KodeArtikel__KodeArtikel=keterangan,
                                Tanggal=tanggal
                            ).exists()
                            if not existing_transaction:
                                transaksiobj = models.TransaksiProduksi(
                                    KodeArtikel = models.Artikel.objects.get(KodeArtikel = keterangan ),
                                    Lokasi = models.Lokasi.objects.get(NamaLokasi='WIP'),
                                    Jumlah = row['Unnamed: 3'],
                                    Jenis = 'Mutasi',
                                    Tanggal = tanggal,
                                    Keterangan = '-'
                                )
                                transaksiobj.save()
                    
                        
                        except Exception as e:
                                produkerror.append([item,(e,keterangan)])
                                continue
                
                except KeyError :
                    break


        return render(request,'error/errorsjp.html',{'data':produkerror})

    return render(request, "produksi/bulk_createtransaksiproduksi.html")

def clean_string(s):
    # Remove "Art" and any non-alphanumeric characters
    s = re.sub(r'Art', '', s)
    return re.sub(r'[^a-zA-Z0-9]', ' ', s).strip()

def bulk_createsaldoawalproduksi(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        listerror = []

        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=6)
            print(item)
            print(df)
            # print(asd)


            for index, row in df.iterrows():
                    print("Saldo Akhir")
                    print(row)
                    # print(asd)
                    if pd.isna(row["Sisa"]):
                        print(f"Data Kosong, Lanjut")
                        break
                    else:
                        try:
                            saldoawalwip = models.SaldoAwalBahanBaku(
                                Harga=0,
                                Jumlah=row['Sisa'],
                                Tanggal="2024-01-01",
                                IDBahanBaku=models.Produk.objects.get(KodeProduk=item),
                                IDLokasi=models.Lokasi.objects.get(pk=2),
                            ).save()
                        except Exception as e:
                            listerror.append([item,e])
                        break

        return render(request,'error/errorsjp.html',{'data':listerror})

    return render(request, "produksi/bulk_createproduk.html")

def bulkcreate_saldoawalartikel(request):
    '''INPUT KSBJ TIAP ARTIKEL, JANGAN LUPA GANTI KODE ARTIKEL'''
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        kodeartikel = '5111 BW'
        artikeobj = models.Artikel.objects.get(KodeArtikel = kodeartikel)

        # Mendapatkan daftar nama sheet
        sheet_names = ['WIP','FG']
        listerror = []
        '''INPUT SALDO AWAL ARTIKEL'''
        item = 'WIP'
        df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=5)
        print(item)
        print(df)
        # print(asd)

        i = 0
        for index, row in df.iterrows():
                if i == 0 :
                    i+=1
                    continue
                print("Saldo Akhir")
                print(row)
                print(row['Saldo Opname'])
                # print(asd)
                try:
                    saldoawalwip = models.SaldoAwalArtikel(
                        IDArtikel = artikeobj,
                        Jumlah=row['Saldo Opname'],
                        Tanggal="2024-01-01",
                        IDLokasi=models.Lokasi.objects.get(pk=1),
                    ).save()
                except Exception as e:
                    listerror.append([item,e])
                break
        
                
        # print(asd)
        item = 'FG'
        df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=5)
        print(item)
        print(df)
        # print(asd)

        i = 0
        for index, row in df.iterrows():
                if i == 0 :
                    i+=1
                    continue
                print("Saldo Akhir")
                print(row)
                print(row['Saldo Opname'])
                # print(asd)
                try:
                    saldoawalwip = models.SaldoAwalArtikel(
                        IDArtikel = artikeobj,
                        Jumlah=row['Saldo Opname'],
                        Tanggal="2024-01-01",
                        IDLokasi=models.Lokasi.objects.get(pk=2),
                    ).save()
                except Exception as e:
                    listerror.append([item,e])
                break
        
        i = 0
        nosppb = None
        for index, row in df.iterrows():
                if i == 0 :
                    i+=1
                    continue
                print("Saldo Akhir")
                print(row)
                print(row['Kirim Barang'])
                if not pd.isna(row['Kirim Barang']):
                    if not pd.isna(row['Unnamed: 4']):
                        nosppb = row['Unnamed: 4']
                    try:
                        try:
                            nomorsppb = models.SPPB.objects.get(NoSPPB= nosppb)
                        except models.SPPB.DoesNotExist:
                            sppbobj = models.SPPB(
                                NoSPPB = nosppb,
                                Tanggal = datetime.now(),
                                Keterangan = 'Cek kembali tanggal'
                            ).save()
                            # nomorsppb = models.SPPB.objects.get(NoSPPB = nosppb)

                        saldoawalwip = models.DetailSPPB(
                            NoSPPB = models.SPPB.objects.get(NoSPPB= nosppb),
                            DetailSPK = models.DetailSPK.objects.get(NoSPK__NoSPK = row['Kirim Barang'],KodeArtikel = artikeobj),
                            Jumlah=row['Unnamed: 5'],
        
                        ).save()
                    except Exception as e:
                        listerror.append([item,(e,nosppb,row['Kirim Barang'])])
                

       
        return render(request,'error/errorsjp.html',{'data':listerror})

    return render(request, "produksi/bulk_createproduk.html")

def bulk_createspk(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = ['SPK','SPK 2024']
        listerror = []

        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item, header=4)
            print(item)
            print(df)
            # print(asd)

            tanggal = None
            NoSPK = None
            for index, row in df.iterrows():
                    if not pd.isna(row['Tanggal']):
                        tanggal = row['Tanggal']
                        NoSPK = row['No. SPK']
                        spkobj=models.SPK(
                            NoSPK = NoSPK,
                            Tanggal = tanggal,
                            KeteranganACC = True,
                        )
                        spkobj.save()
                        
                    print(tanggal)
                    print(NoSPK)
                    print(row)
                    if not pd.isna(row['Artikel']):
                        try:
                            print(row['Artikel'])
                            detailspkobj = models.DetailSPK(
                                NoSPK = models.SPK.objects.get(NoSPK = NoSPK),
                                KodeArtikel = models.Artikel.objects.get(KodeArtikel = clean_string(row['Artikel'])),
                                Jumlah = row['Jumlah']
                            ).save()
                        except Exception as e:
                            listerror.append([NoSPK,(e,row['Artikel'],item)])


                        # print(asd)
                    # if pd.isna(row["Sisa"]):
                    #     print(f"Data Kosong, Lanjut")
                    #     break
                    # else:
                    #     try:
                    #         saldoawalwip = models.SaldoAwalBahanBaku(
                    #             Harga=0,
                    #             Jumlah=row['Sisa'],
                    #             Tanggal="2024-01-01",
                    #             IDBahanBaku=models.Produk.objects.get(KodeProduk=item),
                    #             IDLokasi=models.Lokasi.objects.get(pk=2),
                    #         ).save()
                    #     except Exception as e:
                    #         listerror.append([item,e])
                    #     break

        return render(request,'error/errorsjp.html',{'data':listerror})

    return render(request, "produksi/bulk_createproduk.html")

def bulk_createsppb(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = ['SPPB2024']
        listerror = []

        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item)
            print(item)
            print(df)
            # print(asd)


            for index, row in df.iterrows():
                    print("Saldo Akhir")
                    print(row)
                    # print(asd)
                    try:
                        if not pd.isna(row['Tanggal']):
                            tanggal = row['Tanggal']
                            NoSPK = row['No. SPPB']
                            spkobj=models.SPPB(
                                NoSPPB = NoSPK,
                                Tanggal = tanggal,
                                Keterangan = "-",
                            )
                            spkobj.save()
                    except Exception as e :
                        listerror.append([row,e])

        return render(request,'error/errorsjp.html',{'data':listerror})

    return render(request, "produksi/bulk_createproduk.html")

def bulk_createpenyesuaian(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)

        # Mendapatkan daftar nama sheet
        sheet_names = excel_file.sheet_names
        # sheet_names = ['A-001-02']
        listerror = []

        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item)
            produkobj = models.Produk.objects.get(KodeProduk = item)

            print(item)
            print(df)
            TanggalAwalPenyesuaian = '2024-1-1'
            TanggalAkhirPenyesuan = "2024-7-30"
            # print(asd)

            i = 0
            for index, row in df.iterrows():
                if i == 0 :
                    i+=1
                    continue
                try:
                    jumlah = (row['Unnamed: 10'])
                    if type(jumlah) != float and type(jumlah) != int:
                        listerror.append([item,'Tipe data bukan number'])
                        break
                    if jumlah == 1 :
                        listerror.append([item,'Jumlah Hanya 1'])
                        break
                except Exception as e :
                    listerror.append([item,e])
                    break
                
                    
                print(jumlah)
                print("Saldo Akhir")
                print(row)
                penyesuaianartikel = models.Penyusun.objects.filter(KodeProduk__KodeProduk = item).values_list('KodeArtikel__KodeArtikel',flat=True).distinct()
                print(penyesuaianartikel)
                for artikel in penyesuaianartikel:
                    print(artikel)
                    konversiobj =models.KonversiMaster.objects.filter(KodePenyusun__KodeArtikel__KodeArtikel = artikel, KodePenyusun__KodeProduk__KodeProduk = item)
                    if len(konversiobj) != 1 :
                        kuantitas = konversiobj.aggregate(konversi = Sum('Kuantitas'))['konversi']
                        # print(kuantitas)
                        # print(asd)
                    else:
                        kuantitas = konversiobj.first().Kuantitas
                    # kuantitas = konversiobj.Kuantitas
                    print(kuantitas,jumlah)
                    # print(asd)
                    konversi = kuantitas * jumlah
                
                    try:
                        spkobj=models.Penyesuaian(
                            KodeArtikel = models.Artikel.objects.get(KodeArtikel = artikel),
                            KodeProduk = produkobj,
                            TanggalMulai = TanggalAwalPenyesuaian,
                            TanggalMinus = TanggalAkhirPenyesuan,
                            lokasi = models.Lokasi.objects.get(NamaLokasi = 'WIP'),
                            konversi =  konversi
                        )
                        spkobj.save()
                    except Exception as e:
                        listerror.append([item,(e,artikel,item)])
                    
                break

        return render(request,'error/errorsjp.html',{'data':listerror})

    return render(request, "produksi/bulk_createproduk.html")

def updatetransaksiproduksiversi(request):
    # datatransaksiproduksiall = models.TransaksiProduksi.objects.filter(Jenis = 'Mutasi')
    # print(datatransaksiproduksiall.count())
    # for item in datatransaksiproduksiall:
    #     versiobj = models.Versi.objects.filter(KodeArtikel = item.KodeArtikel).first()
    #     item.VersiArtikel = versiobj
    #     print(item)
    #     item.save()
    datadetailsppb = models.DetailSPPB.objects.filter(DetailSPK__isnull = False)
    for item in datadetailsppb:
        versiobj = models.Versi.objects.filter(KodeArtikel = item.DetailSPK.KodeArtikel).first()
        if versiobj == None:
            print(item.DetailSPK.KodeArtikel)
        item.VersiArtikel = versiobj
        item.save()

def createhargafg (request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        excel_file = pd.ExcelFile(file)
        print(excel_file)
        # print(asd)

        # Mendapatkan daftar nama sheet
        sheet_names = ['SPPB2024']
        listerror = []

        for item in sheet_names:
            df = pd.read_excel(file, engine="openpyxl", sheet_name=item)
            print(item)
            print(df)
            # print(asd)


            for index, row in df.iterrows():
                    print("Saldo Akhir")
                    print(row)
                    # print(asd)
                    try:
                        if not pd.isna(row['Tanggal']):
                            tanggal = row['Tanggal']
                            NoSPK = row['No. SPPB']
                            spkobj=models.SPPB(
                                NoSPPB = NoSPK,
                                Tanggal = tanggal,
                                Keterangan = "-",
                            )
                            spkobj.save()
                    except Exception as e :
                        listerror.append([row,e])

        return render(request,'error/errorsjp.html',{'data':listerror})

    return render(request, "produksi/bulk_createproduk.html")

@login_required
@logindecorators.allowed_users(allowed_roles=["produksi",'ppic'])
def views_artikel(request):
    '''
    FItur ini digunakan untuk manajemen data Artikel pada sistem
    Algoritma
    1. Mendapatkan semua data Artikel pada tabel Artikel
    2. Mengiterasi semua data Artikel pada tabel
    3. Mencari kode bahan baku penyusun utama tiap artikel 
    4. Apabila ada data penysuun utama maka akan menambahkan kedapat detailartikelobj untuk objek penyusunnya
    5. Apabila ada lebih dari 1 bahan baku utama (multiple version) maka akan diambil yang terakhir
    6. Apabila tidak ada bahan baku penyusun utama maka akan menambhakan keterangan "Belum diset"
    '''
    datakirim = []
    data = models.Artikel.objects.all()
    for item in data:
        detailartikelobj = models.Penyusun.objects.filter(KodeArtikel=item.id).filter(
            Status=1
        )
        if detailartikelobj.exists():
            datakirim.append([item, detailartikelobj.last()])
        else:
            datakirim.append([item, "Belum diset"])
    return render(request, "produksi/views_artikel.html", {"data": datakirim})


@login_required
@logindecorators.allowed_users(allowed_roles=["produksi",'ppic'])
def views_display(request):
    '''
    Fitur ini digunakan untuk melakukan manajemen data Display pada sistem
    '''
    data = models.Display.objects.all()
    return render(request, "produksi/views_display.html", {"data": data})