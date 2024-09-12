from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import Http404, HttpResponse
from django.urls import reverse
from . import models
from django.db.models import Sum, Max
from io import BytesIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import calendar
from . import logindecorators
from django.contrib.auth.decorators import login_required
import time
from django.core.cache import cache
from django.contrib.admin.models import LogEntry


# Create your views here.


def gethargabahanbaku(
    listtanggal, hargamasukobj, hargakeluarobj, hargareturobj,hargapemusnahanobj, hargaawal, jumlahawal
):
    listdata = []
    totalharga = hargaawal * jumlahawal
    for j in listtanggal:
        jumlahkeluarperhari = 0
        hargakeluartotalperhari = 0
        hargakeluarsatuanperhari = 0
        hargamasuktotalperhari = 0
        hargamasuksatuanperhari = 0
        jumlahmasukperhari = 0

        sjpobj = hargamasukobj.filter(NoSuratJalan__Tanggal=j)
        returobj = hargareturobj.filter(tanggal=j)
        if sjpobj.exists() or returobj.exists():
            
            if sjpobj.exists():
                for k in sjpobj:
                    hargamasuktotalperhari += k.Harga * k.Jumlah
                    jumlahmasukperhari += k.Jumlah
            if returobj.exists():
                for k in returobj:
                    jumlahmasukperhari += k.jumlah *-1
                    hargamasuktotalperhari += k.jumlah * hargaawal * -1
            try:
                hargamasuksatuanperhari += hargamasuktotalperhari / jumlahmasukperhari
            except ZeroDivisionError:
                hargamasuksatuanperhari += 0
            dumy = {
                "Tanggal": j,
                "Jumlahstokawal": jumlahawal,
                "Hargasatuanawal": round(hargaawal, 2),
                "Hargatotalawal": round(totalharga, 2),
                "Jumlahmasuk": jumlahmasukperhari,
                "Hargamasuksatuan": round(hargamasuksatuanperhari, 2),
                "Hargamasuktotal": round(hargamasuktotalperhari, 2),
                "Jumlahkeluar": jumlahkeluarperhari,
                "Hargakeluarsatuan": round(hargakeluarsatuanperhari, 2),
                "Hargakeluartotal": round(hargakeluartotalperhari, 2),
            }
            jumlahawal += jumlahmasukperhari - jumlahkeluarperhari
            totalharga += hargamasuktotalperhari - hargakeluartotalperhari
            try:
                hargaawal = totalharga / jumlahawal
            except ZeroDivisionError :
                hargaawal =  0

            # print("Sisa Stok Hari Ini : ", jumlahawal)
            # print("harga awal Hari Ini :", hargaawal)
            # print("harga total Hari Ini :", totalharga, "\n")
            dumy["Sisahariini"] = jumlahawal
            dumy["Hargasatuansisa"] = round(hargaawal, 2)
            dumy["Hargatotalsisa"] = round(totalharga, 2)
            listdata.append(dumy)
            # print(dumy)

        hargamasuktotalperhari = 0
        hargamasuksatuanperhari = 0
        jumlahmasukperhari = 0

        # print(j)
        transaksigudangobj = hargakeluarobj.filter(tanggal=j)
        if transaksigudangobj.exists():
            for k in transaksigudangobj:
                jumlahkeluarperhari += k.jumlah
                hargakeluartotalperhari += k.jumlah * hargaawal
            hargakeluarsatuanperhari += hargakeluartotalperhari / jumlahkeluarperhari
        else:
            hargakeluartotalperhari += 0
            hargakeluarsatuanperhari += 0
            jumlahkeluarperhari += 0
        transaksipemusnahanobj = hargapemusnahanobj.filter(Tanggal=j)
        if transaksipemusnahanobj.exists():
            for k in transaksipemusnahanobj:
                jumlahkeluarperhari += k.Jumlah
                hargakeluartotalperhari += k.Jumlah * hargaawal
            try:
                hargakeluarsatuanperhari += hargakeluartotalperhari / jumlahkeluarperhari
            except : 
                hargakeluarsatuanperhari = 0
        else:
            hargakeluartotalperhari += 0
            hargakeluarsatuanperhari += 0
            jumlahkeluarperhari += 0
        # print("Tanggal : ", j)
        # print("Sisa Stok Hari Sebelumnya : ", jumlahawal)
        # print("harga awal Hari Sebelumnya :", hargaawal)
        # print("harga total Hari Sebelumnya :", totalharga)
        # print("Jumlah Masuk : ", jumlahmasukperhari)
        # print("Harga Satuan Masuk : ", hargamasuksatuanperhari)
        # print("Harga Total Masuk : ", hargamasuktotalperhari)
        # print("Jumlah Keluar : ", jumlahkeluarperhari)
        # print("Harga Keluar : ", hargakeluarsatuanperhari)
        # print(
        #     "Harga Total Keluar : ",
        #     hargakeluarsatuanperhari * jumlahkeluarperhari,
        # )
        jumlahawal += jumlahmasukperhari - jumlahkeluarperhari
        totalharga += hargamasuktotalperhari - hargakeluartotalperhari
        try:
            hargaawal = totalharga / jumlahawal
        except ZeroDivisionError:
            hargaawal = 0

    return hargaawal, jumlahawal, totalharga


@login_required
@logindecorators.allowed_users(allowed_roles=["ppic"])
def dashboard(request):
    data = models.confirmationorder.objects.filter(StatusAktif=True)
    print(data)
    for i in data:
        detailcopo = models.detailconfirmationorder.objects.filter(
            confirmationorder=i.id
        )
        i.detailcopo = detailcopo
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")
    return render(request, "ppic/dashboard.html", {"data": data})


def gethargafgterakhirberdasarkanmutasi(KodeArtikel, Tanggaltes, HargaPurchasing):
    manualinputharga = False
    hargaartikelfg = models.HargaArtikel.objects.filter(KodeArtikel = KodeArtikel, Tanggal__month = Tanggaltes.month)
    if hargaartikelfg.exists():
        totalbiayafg = hargaartikelfg.first().Harga
        manualinputharga = True
    else:
        cekmutasiwipfg = cekmutasiwipfgterakhir(KodeArtikel, Tanggaltes)
        cekmutasiwip = cekmutasiwipterakhir(KodeArtikel, Tanggaltes)
        # cekmutasifg = cekmutasifgterakhir(KodeArtikel, Tanggaltes)
        # print("\n\n ", Tanggaltes)
        # print("Mutasi WIP - FG Terakhir", cekmutasiwipfg)
        # print("Mutasi WIP ", cekmutasiwip)
        # print("Mutasi FG ", cekmutasifg)
        hargakomponen_fg_fgterakhir = gethargaartikelfgperbulan(
            KodeArtikel, cekmutasiwipfg, HargaPurchasing
        )
        # print("\n\n", hargakomponen_fg_fgterakhir)

        hargakomponen_wip_fgterakhir = gethargaartikelwipperbulan(
            KodeArtikel, cekmutasiwipfg, HargaPurchasing
        )

        if cekmutasiwipfg.month == Tanggaltes.month:
            # print(f"Ada Mutasi WIP ke FG pada {Tanggaltes}, Harga FG Total di update")
            komponen_fg_terakhir = gethargaartikelfgperbulan(
                KodeArtikel, cekmutasiwipfg, HargaPurchasing
            )
            if komponen_fg_terakhir:
                hargakomponen_fg_fgterakhir = komponen_fg_terakhir[KodeArtikel]["hargafg"]
            else:
                hargakomponen_fg_fgterakhir = 0

            komponen_wip_terakhir = gethargaartikelwipperbulan(
                KodeArtikel, cekmutasiwipfg, HargaPurchasing
            )

            if hargakomponen_wip_fgterakhir:
                hargakomponen_wip_fgterakhir = komponen_wip_terakhir[KodeArtikel]["hargawip"]
            else :
                hargakomponen_wip_fgterakhir = 0
            totalbiayafg = hargakomponen_wip_fgterakhir + hargakomponen_fg_fgterakhir

        else:
            # print(
            #     f"Tidak ada Mutasi FG pada {Tanggaltes}, Harga FG menggunakan harga FG terakhir",
            #     cekmutasiwipfg,
            # )
            komponen_wip_terakhir = gethargaartikelwipperbulan(
                KodeArtikel, cekmutasiwipfg, HargaPurchasing
            )
            hargakomponen_wip_fgterakhir = komponen_wip_terakhir[KodeArtikel]["hargawip"]
            komponen_fg_terakhir = gethargaartikelfgperbulan(
                KodeArtikel, cekmutasiwipfg, HargaPurchasing
            )
            hargakomponen_fg_fgterakhir = komponen_fg_terakhir[KodeArtikel]["hargafg"]
            totalbiayafg = hargakomponen_wip_fgterakhir + hargakomponen_fg_fgterakhir
            if cekmutasiwip.month == Tanggaltes.month:
                # print("Ada mutasi WIP bulan : ", cekmutasiwip.month)
                komponen_wip_terakhir = gethargaartikelwipperbulan(
                    KodeArtikel, cekmutasiwip, HargaPurchasing
                )
                hargakomponen_wip_fgterakhir = komponen_wip_terakhir[KodeArtikel][
                    "hargawip"
                ]

            else:
                # print("Tidak Ada mutasi WIP bulan : ", cekmutasiwip.month)
                komponen_wip_terakhir = gethargaartikelwipperbulan(
                    KodeArtikel, cekmutasiwip, HargaPurchasing
                )
                hargakomponen_wip_fgterakhir = komponen_wip_terakhir[KodeArtikel][
                    "hargawip"
                ]
        # print(hargakomponen_wip_fgterakhir,
        #     hargakomponen_fg_fgterakhir,
        #     )
    # print(asd)
    return (
        totalbiayafg,
        manualinputharga
        # hargakomponen_wip_fgterakhir,
        # hargakomponen_wip_fgterakhir,
        # komponen_wip_terakhir,
        # komponen_fg_terakhir,
    )


def laporanbarangjadi(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanstokfg.html")
    else:
        # Rumus = Saldo awal periode sampai tanggal akhir - Keluar awal periode sampai tanggal akhir
        tanggal_akhir = request.GET["tanggalakhir"]
        tanggalakhir_obj = datetime.strptime(tanggal_akhir, "%Y-%m-%d")
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(tanggalakhir_obj.year, month)[1]
            last_days.append(date(tanggalakhir_obj.year, month, last_day))
        index = int(tanggalakhir_obj.month)
        awaltahun = date(tanggalakhir_obj.year, 1, 1)
        hargaakhirbulanperproduk = gethargapurchasingperbulan(
            last_days, index, awaltahun
        )
        data = models.Artikel.objects.all()
        # Transaksi Masuk ke WIP
        transaksimasukfg = models.TransaksiProduksi.objects.filter(
            KodeArtikel__in=data, Jenis="Mutasi", Tanggal__lte=tanggalakhir_obj
        )
        totaltransaksimasuk = transaksimasukfg.values(
            "KodeArtikel__KodeArtikel"
        ).annotate(total=Sum("Jumlah"))
        transaksikeluarfg = models.DetailSPPB.objects.filter(
            DetailSPK__KodeArtikel__in=data, NoSPPB__Tanggal__lte=tanggalakhir_obj
        )
        totaltransaksikeluar = transaksikeluarfg.values(
            "DetailSPK__KodeArtikel__KodeArtikel"
        ).annotate(total=Sum("Jumlah"))
        print(transaksimasukfg)
        grandtotal = 0
        data2_dict = {
            item["DetailSPK__KodeArtikel__KodeArtikel"]: item["total"]
            for item in totaltransaksikeluar
        }

        # Mengurangi nilai total dari data2 pada data1 dan menghasilkan hasil dalam satu langkah
        result = [
            {
                "KodeArtikel": item["KodeArtikel__KodeArtikel"],
                "total": item["total"]
                - data2_dict.get(item["KodeArtikel__KodeArtikel"], 0),
            }
            for item in totaltransaksimasuk
        ]
        for data in result:
            dataArtikel = models.Artikel.objects.get(KodeArtikel=data["KodeArtikel"])
            totalbiayafgterbaru = gethargafgterakhirberdasarkanmutasi(
                dataArtikel, tanggalakhir_obj, hargaakhirbulanperproduk
            )[0]
            data["biayafg"] = totalbiayafgterbaru
            data["hargatotal"] = totalbiayafgterbaru * data["total"]
            grandtotal += data["hargatotal"]

        return render(
            request,
            "ppic/views_laporanstokfg.html",
            {
                "data": result,
                "tanggalakhir": tanggal_akhir,
                "grandtotal": grandtotal,
            },
        )


def laporanbarangmasuk(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanbarangmasuk.html")
    else:
        tanggalawal = request.GET["tanggalawal"]
        tanggalakhir = request.GET["tanggalakhir"]

        dataspk = models.SuratJalanPembelian.objects.filter(
            Tanggal__range=(tanggalawal, tanggalakhir)
        ).order_by("Tanggal")
        print(dataspk)
        listdetailsjp = []
        grandtotal = 0
        for i in dataspk:
            detailsjpembelianobj = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan=i.NoSuratJalan
            )
            for j in detailsjpembelianobj:
                j.supplier = i.supplier
                j.totalharga = j.Jumlah * j.Harga
                grandtotal += j.totalharga
                j.NoSuratJalan.Tanggal = date.strftime(
                    j.NoSuratJalan.Tanggal, "%Y-%m-%d"
                )
                listdetailsjp.append(j)

        print(listdetailsjp)

        return render(
            request,
            "ppic/views_laporanbarangmasuk.html",
            {
                "data": listdetailsjp,
                "tanggalawal": tanggalawal,
                "tanggalakhir": tanggalakhir,
                "grandtotal": grandtotal,
            },
        )


# def gethargafg(penyusunobj):
#     # Mengambil data Konversi master
#     konversiobj = models.KonversiMaster.objects.get(
#         KodePenyusun=penyusunobj.IDKodePenyusun
#     )
#     # Mengambil kuantitas dan ditambahkan 2.5%
#     konversialowance = konversiobj.Allowance
#     # Mengambil data Detail surrat jalan pembelian untuk kode produk sesuai dengan penyusun obj (1 penyusun 1 produk 1 artkel)
#     detailsjpembelian = models.DetailSuratJalanPembelian.objects.filter(
#         KodeProduk=penyusunobj.KodeProduk
#     )
#     hargatotalkodeproduk = 0
#     jumlahtotalkodeproduk = 0
#     for m in detailsjpembelian:
#         hargatotalkodeproduk += m.Harga * m.Jumlah
#         jumlahtotalkodeproduk += m.Jumlah
#     # print("ini jumlah harga total ", hargatotalkodeproduk)
#     try:
#         rataratahargakodeproduk = hargatotalkodeproduk / jumlahtotalkodeproduk
#     except ZeroDivisionError:
#         rataratahargakodeproduk = 0
#     # print("selesai")
#     # print(rataratahargakodeproduk)
#     nilaifgperkodeproduk = rataratahargakodeproduk * konversialowance
#     # print("Harga Konversi : ", nilaifgperkodeproduk)
#     return nilaifgperkodeproduk


def laporanbarangkeluar(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanbarangkeluar.html")
    else:
        tanggalawal = request.GET["tanggalawal"]
        tanggalakhir = request.GET["tanggalakhir"]
        tanggalawal_obj = datetime.strptime(tanggalawal, "%Y-%m-%d")
        tanggalakhir_obj = datetime.strptime(tanggalakhir, "%Y-%m-%d")
        tahun_awal = tanggalawal_obj.year
        tahun_akhir = tanggalakhir_obj.year
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(tanggalakhir_obj.year, month)[1]
            last_days.append(date(tanggalakhir_obj.year, month, last_day))
        index = int(tanggalakhir_obj.month)
        awaltahun = date(tanggalakhir_obj.year, 1, 1)
        hargaakhirbulanperproduk = gethargapurchasingperbulan(
            last_days, index, awaltahun
        )

        data = models.DetailSPPB.objects.filter(
            NoSPPB__Tanggal__range=(tanggalawal, tanggalakhir), DetailSPKDisplay=None
        ).order_by("NoSPPB__Tanggal")

        print("ini data", data)
        grandtotal = 0
        datakirim = []
        if not data.exists():
            messages.warning(
                request, "Data SPPB tidak ditemukan pada rentang tanggal tersebut"
            )
            return redirect("laporanbarangkeluar")
        for i in data:
            (
                totalbiayafg,
                hargakomponenwip,
                hargakomponenfg,
                datakomponenwip,
                datakomponenfg,
            ) = gethargafgterakhirberdasarkanmutasi(
                i.DetailSPK.KodeArtikel, i.NoSPPB.Tanggal, hargaakhirbulanperproduk
            )

            i.HargaFG = totalbiayafg
            i.TotalBiayaKeluar = i.HargaFG * i.Jumlah
            i.NoSPPB.Tanggal = date.strftime(i.NoSPPB.Tanggal, ("%Y-%m-%d"))
            grandtotal += totalbiayafg
            print(
                f"\n\n\n\n Harga FG {i.DetailSPK.KodeArtikel} - {i.NoSPPB.Tanggal} ",
                totalbiayafg,
            )
            """KONDISI
            1. MUTASI WIP MUTASI FG --> Harga total FG di update
            2. Mutasi WIP, tidak mutasi FG --> Harga total tidak diupdate
            3. tidka Mutasi WIP, MUtasi FG --> Harga total tidak diupdate
            4. Tidak keduanya --> harga tidak di update
            """
            # print(datakomponenwip)
            # print(datakomponenwip[dataartikel]['hargawip'])

        # print(listdata)

        return render(
            request,
            "ppic/views_laporanbarangkeluar.html",
            {
                "tanggalawal": tanggalawal,
                "tanggalakhir": tanggalakhir,
                "data": data,
                "grandtotal": grandtotal,
            },
        )



"""
Revisi 4/21/2024
1. Co PO
2. Transaction Log
3. Perhitungan Laporan
"""


def viewconfirmationorder(request):
    data = models.confirmationorder.objects.all()

    print(data)
    for i in data:
        detailcopo = models.detailconfirmationorder.objects.filter(
            confirmationorder=i.id
        )

        i.detailcopo = detailcopo
        i.tanggal = i.tanggal.strftime("%Y-%m-%d")
    if len(request.GET) == 0:
        return render(request, "ppic/views_confirmationorder.html", {"data": data})


def tambahconfirmationorder(request):
    dataartikel = models.Artikel.objects.all()
    datadisplay = models.Display.objects.all()
    if request.method == "GET":
        return render(
            request,
            "ppic/add_co.html",
            {"dataartikel": dataartikel, "datadisplay": datadisplay},
        )
    else:
        print(request.POST)
        tanggaladd = request.POST["tanggal"]
        nomorco = request.POST["nomorco"]
        kepada = request.POST["kepada"]
        perihal = request.POST["perihal"]
        artikel = request.POST.getlist("artikel[]")
        kuantitas = request.POST.getlist("kuantitas[]")
        harga = request.POST.getlist("harga[]")
        deskripsi = request.POST.getlist("deskripsi[]")
        displaylist = request.POST.getlist("display[]")
        deskripsidisplay = request.POST.getlist("deskripsidisplay[]")
        kuantitasdisplay = request.POST.getlist("kuantitasdisplay[]")
        hargadisplay = request.POST.getlist("hargadisplay[]")

        valid_artikel_list = any(artikel)
        valid_display_list = any(displaylist)
        if valid_artikel_list or valid_display_list:

            # Cek CO sudah ada atau belum
            ceknopo = models.confirmationorder.objects.filter(NoCO=nomorco).exists()
            if ceknopo:
                messages.error(
                    request, "Nomor Confirmation Order telah ada pada sistem"
                )
                return redirect("addco")

            confirmationorderobj = models.confirmationorder(
                NoCO=nomorco, kepada=kepada, perihal=perihal, tanggal=tanggaladd
            )
            confirmationorderobj.save()
            print(confirmationorderobj.id)
            for (
                artikelterpilih,
                kuantitasterpilih,
                hargaterpilih,
                deskripsiterpilih,
            ) in zip(artikel, kuantitas, harga, deskripsi):
                if artikelterpilih == "":
                    continue
                # print(artikel, kuantitas, harga, deskripsi)
                try:
                    artikelobj = models.Artikel.objects.get(KodeArtikel=artikelterpilih)

                except:
                    messages.error(
                        request, f"Data Artikel {artikel} tidak ditemukan dalam sistem"
                    )
                    continue
                print(artikelobj)
                detailconfirmationobj = models.detailconfirmationorder(
                    confirmationorder=confirmationorderobj,
                    Artikel=artikelobj,
                    Harga=hargaterpilih,
                    kuantitas=kuantitasterpilih,
                    deskripsi=deskripsiterpilih,
                )
                print(dir(detailconfirmationobj))
                detailconfirmationobj.save()
            for display, kuantitas, harga, deskripsi in zip(
                displaylist, kuantitasdisplay, hargadisplay, deskripsidisplay
            ):
                if display == "":
                    continue
                try:
                    displayobj = models.Display.objects.get(KodeDisplay=display)
                except:
                    messages.error(request, f"Data display {display} tidak ditemukan")
                    continue
                detailconfirmationobj = models.detailconfirmationorder(
                    confirmationorder=confirmationorderobj,
                    Display=displayobj,
                    Harga=harga,
                    kuantitas=kuantitas,
                    deskripsi=deskripsi,
                ).save()
                messages.success(request, f"Data berhasil disimpan")
            return redirect("confirmationorder")
        else:
            messages.error(request, "Masukkan Artikel atau Display")
            return redirect("addco")


def detailco(request, id):
    data = models.confirmationorder.objects.get(id=id)
    detailcopo = models.detailconfirmationorder.objects.filter(
        confirmationorder=data.id, Display=None
    )
    data.detailcopo = detailcopo
    detailcopodisplay = models.detailconfirmationorder.objects.filter(
        confirmationorder=data.id, Artikel=None
    )
    data.tanggal = data.tanggal.strftime("%Y-%m-%d")
    data.detailcopodisplay = detailcopodisplay
    detailsppb = models.DetailSPPB.objects.filter(IDCO=data)
    detailsppbartikel = detailsppb.filter(DetailSPK__isnull = False)
    detailsppbdisplay = detailsppb.filter(DetailSPKDisplay__isnull = False)
    detailsppbbahan = detailsppb.filter(DetailBahan__isnull = False)
    data.detailsppb = detailsppb
    print(detailsppbdisplay)
    datajumlah = detailsppbartikel.values("DetailSPK__KodeArtikel__KodeArtikel").annotate(
        total=Sum("Jumlah")

    )
    datajumlahdisplay = detailsppbdisplay.values("DetailSPKDisplay__KodeDisplay__KodeDisplay").annotate(
        total=Sum("Jumlah")

    )
    datajumlahbahan = detailsppbbahan.values("DetailBahan__KodeProduk").annotate(
        total=Sum("Jumlah")

    )

    

    print(datajumlah)
    print(datajumlahdisplay)
    print(datajumlahbahan)

    return render(request, "ppic/detailco.html", {"dataco": data, "jumlah": datajumlah,'jumlahdisplay':datajumlahdisplay,'jumlahbahan':datajumlahbahan})


def updateco(request, id):
    data = models.confirmationorder.objects.get(id=id)
    detailcopo = models.detailconfirmationorder.objects.filter(
        confirmationorder=data.id, Display=None
    )
    data.detailcopo = detailcopo
    detailcopodisplay = models.detailconfirmationorder.objects.filter(
        confirmationorder=data.id, Artikel=None
    )
    data.detailcopodisplay = detailcopodisplay
    data.tanggal = data.tanggal.strftime("%Y-%m-%d")
    print(len(data.detailcopo))
    dataartikel = models.Artikel.objects.all()
    datadisplay = models.Display.objects.all()
    if request.method == "GET":
        return render(
            request,
            "ppic/updateco.html",
            {"dataco": data, "dataartikel": dataartikel, "datadisplay": datadisplay},
        )
    else:
        print(request.POST)
        tanggaladd = request.POST["tanggal"]
        nomorco = request.POST["nomorco"]
        kepada = request.POST["kepada"]
        perihal = request.POST["perihal"]
        status = request.POST["status"]
        listartikel = request.POST.getlist("artikel[]")
        listkuantitas = request.POST.getlist("kuantitas[]")
        listharga = request.POST.getlist("harga[]")
        listdeskripsi = request.POST.getlist("deskripsi[]")
        listid = request.POST.getlist("id[]")
        listiddisplay = request.POST.getlist("iddisplay[]")
        display = request.POST.getlist("display[]")
        kuantitasdisplay = request.POST.getlist("kuantitasdisplay[]")
        hargadisplay = request.POST.getlist("hargadisplay[]")
        deskripsidisplay = request.POST.getlist("deskripsidisplay[]")

        # print(listid, artikel, kuantitas, harga, deskripsi)
        # print(asd)
        data.tanggal = tanggaladd
        data.nomorco = nomorco
        data.kepada = kepada
        data.perihal = perihal
        data.StatusAktif = status

        data.save()

        for artikelid, artikel, kuantitas, harga, deskripsi in zip(
            listid, listartikel, listkuantitas, listharga, listdeskripsi
        ):
            # print(artikel, kuantitas, harga, deskripsi)
            if artikelid == "":
                detailconfirmationobj = models.detailconfirmationorder(
                    confirmationorder=data,
                    Artikel=models.Artikel.objects.get(KodeArtikel=artikel),
                    Harga=harga,
                    kuantitas=kuantitas,
                    deskripsi=deskripsi,
                )
            else:
                detailconfirmationobj = models.detailconfirmationorder.objects.get(
                    id=artikelid
                )
                detailconfirmationobj.confirmationorder = data
                detailconfirmationobj.Artikel = models.Artikel.objects.get(
                    KodeArtikel=artikel
                )
                detailconfirmationobj.Harga = harga
                detailconfirmationobj.kuantitas = kuantitas
                detailconfirmationobj.deskripsi = deskripsi

            detailconfirmationobj.save()
        for (
            iddisplay,
            displayterpilih,
            displayqty,
            displayharga,
            displaydeskripsi,
        ) in zip(
            listiddisplay, display, kuantitasdisplay, hargadisplay, deskripsidisplay
        ):
            print("<Masuk")
            if iddisplay == "":
                detailconfirmationobj = models.detailconfirmationorder(
                    confirmationorder=data,
                    Display=models.Display.objects.get(KodeDisplay=displayterpilih),
                    Harga=displayharga,
                    kuantitas=displayqty,
                    deskripsi=displaydeskripsi,
                )
            else:
                detailconfirmationobj = models.detailconfirmationorder.objects.get(
                    id=iddisplay
                )
                detailconfirmationobj.confirmationorder = data

                detailconfirmationobj.Diplay = models.Display.objects.get(
                    KodeDisplay=displayterpilih
                )
                detailconfirmationobj.Harga = displayharga
                detailconfirmationobj.kuantitas = displayqty
                detailconfirmationobj.deskripsi = displaydeskripsi
            detailconfirmationobj.save()

        return redirect("confirmationorder")


def deletedetailco(request, id):
    data = models.detailconfirmationorder.objects.get(id=id)
    idco = data.confirmationorder.id
    data.delete()
    messages.success(request,'Data berhasil dihapus')
    return redirect("updateco", id=idco)


def deleteco(request, id):
    data = models.confirmationorder.objects.get(id=id)
    data.delete()
    return redirect("confirmationorder")

def detaillaporanbarangkeluar(request):
    if len(request.GET) > 0:
        bulan = request.GET["bulan"]
        waktu = request.GET["waktu"]
        waktuobj = datetime.strptime(waktu, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = listbulan.index(request.GET["bulan"]) + 1
        # print(index)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))
        """ SPPB Section """
        (

            datadetailsppb,
            totalbiayakeluar,
            datapenyusun,
            datalistbarangkeluar,
            transaksilainlain,
            transaksigold,
            detailbiaya,
            datatransaksikeluar
        ) = getbarangkeluar(last_days, index, awaltahun)
        datatransaksikeluar = datatransaksikeluar[index-1]
        rekapartikel,rekaptransaksidisplay,rekapdatabahankeluar,rekapgolongand,rekaptransaksilainlain,rekappemusnahanartikel,rekappemusnahanbahanbaku = create_rekaptransaksikeluar(datatransaksikeluar)
        # print(rekapartikel)
        # print(datatransaksikeluar['Pemusnahanartikel'])
        print(rekappemusnahanbahanbaku)
        print(rekappemusnahanartikel)
        return render(
            request,
            "ppic/detaillaporanbarangkeluar.html",
            {
                "totalbiayakeluar": datatransaksikeluar['totalkeluar'],
                "bulan": bulan,
                "penyusun": datapenyusun,
                "sppb": datatransaksikeluar["SPPBArtikel"]["SPPBArtikel"],
                "nilaibarangkeluar": datatransaksikeluar['SPPBArtikel']['totalbiayasppb'],
                'display' : datatransaksikeluar['SPPBDisplay']['SPPBDisplay'],
                'nilaidisplaykeluar' : datatransaksikeluar['SPPBDisplay']['totalbiayasppb'],
                "lainlain": datatransaksikeluar['Transaksilainlain']['datatransaksi'],
                "nilailainlain": datatransaksikeluar['Transaksilainlain']['totalbiaya'],
                "transaksigold": datatransaksikeluar['Transaksigolongand']['datatransaksi'],
                "nilaigold": datatransaksikeluar['Transaksigolongand']['totalbiaya'],
                "transaksibahanbaku":datatransaksikeluar["Transaksibahanbaku"]['datatransaksi'],
                "nilaitransaksibahanbaku":datatransaksikeluar["Transaksibahanbaku"]['totalbiaya'],
                'pemusnahanbahanbaku' : datatransaksikeluar['Pemusnahanbahanbaku']['datatransaksi'],
                'nilaipemusnahanbahanbaku' : datatransaksikeluar['Pemusnahanbahanbaku']['totalbiaya'],
                'pemusnahanartikel' : datatransaksikeluar['Pemusnahanartikel']['transaksipemusnahan'],
                'nilaipemusnahanartikel':datatransaksikeluar['Pemusnahanartikel']['totalbiayapemusnahan'],
                "rekapartikel" : rekapartikel,
                'rekapgolongand':rekapgolongand,
                'rekaptransaksidisplay' : rekaptransaksidisplay,
                'rekapbahankeluar' : rekapdatabahankeluar,
                'rekaptransaksilainlain' : rekaptransaksilainlain,
                'rekappemusnahanartikel' : rekappemusnahanartikel,
                'rekappemusnahanbahanbaku' : rekappemusnahanbahanbaku
                
                
            },
        )

def create_rekaptransaksikeluar (datatransaksikeluar):
    '''Digunakan untuk membuat rekapitulasi barang keluar dengan input adalah datatransaksikeluar dari fungsi getbarangkeluar'''
    rekapsppb = datatransaksikeluar["SPPBArtikel"]["SPPBArtikel"]
    # Rekap Artikel Keluar
    rekapartikel = {}
    for i in rekapsppb:
        print(i.inputmanual)
        if i.DetailSPK.KodeArtikel in rekapartikel:
            rekapartikel[i.DetailSPK.KodeArtikel]['Jumlah'] += i.Jumlah
        else:
            rekapartikel[i.DetailSPK.KodeArtikel] = {'Jumlah': i.Jumlah, "HargaFG":i.hargafg}
        rekapartikel[i.DetailSPK.KodeArtikel]['hargatotal'] = rekapartikel[i.DetailSPK.KodeArtikel]['Jumlah'] * rekapartikel[i.DetailSPK.KodeArtikel]['HargaFG']
        rekapartikel[i.DetailSPK.KodeArtikel]['inputmanual'] = i.inputmanual

    '''Digunakan untuk membuat rekapitulasi pemusnahan artikel'''
    datapemusnahanartikel = datatransaksikeluar["Pemusnahanartikel"]["transaksipemusnahan"]
    # Rekap Artikel Keluar
    rekappemusnahanartikel = {}
    for i in datapemusnahanartikel:
        # print(i.inputmanual)
        if i.KodeArtikel in rekappemusnahanartikel:
            rekappemusnahanartikel[i.KodeArtikel]['Jumlah'] += i.Jumlah
        else:
            rekappemusnahanartikel[i.KodeArtikel] = {'Jumlah': i.Jumlah, "HargaFG":i.hargafg}
        rekappemusnahanartikel[i.KodeArtikel]['hargatotal'] = rekappemusnahanartikel[i.KodeArtikel]['Jumlah'] * rekappemusnahanartikel[i.KodeArtikel]['HargaFG']
        rekappemusnahanartikel[i.KodeArtikel]['inputmanual'] = i.inputmanual
    
    print(rekappemusnahanartikel)
    # print(asd)

    
    # Rekap Display Keluar
    '''
    A : {NoSPK : {1: {Jumlah : x,Harga:y},}}
    '''
    datatransaksidisplay =datatransaksikeluar['SPPBDisplay']['SPPBDisplay']
    rekaptransaksidisplay = {}
    for i in datatransaksidisplay:
        # print(dir(i))
        if i.DetailSPKDisplay.KodeDisplay in rekaptransaksidisplay:
            print('Masuk',rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK'])
            
            if i.DetailSPKDisplay.NoSPK.NoSPK in rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK']:
                rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK'][i.DetailSPKDisplay.NoSPK.NoSPK]['Jumlah'] += i.Jumlah
                rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK'][i.DetailSPKDisplay.NoSPK.NoSPK]['Hargatotal'] += i.totalbiaya
                # TOTAL BIAYA MASIH ERROR DARI RUMUS PERHITUNGAN BARANG KELUAR DI FG 
                # print('masuksini')
            else:
                rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK'][i.DetailSPKDisplay.NoSPK.NoSPK ]= {'Jumlah' : i.Jumlah, 'Hargatotal':i.totalbiaya}

        else:
            rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay] = {"NoSPK" : {i.DetailSPKDisplay.NoSPK.NoSPK: {'Jumlah' : i.Jumlah, 'Hargatotal':i.totalbiaya}}}
        hargatotal = rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK'][i.DetailSPKDisplay.NoSPK.NoSPK ]['Hargatotal']
        if hargatotal == 0 :
            hargasatuan = 0
        else:
            hargasatuan = hargatotal /rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK'][i.DetailSPKDisplay.NoSPK.NoSPK]['Jumlah']
        rekaptransaksidisplay[i.DetailSPKDisplay.KodeDisplay]['NoSPK'][i.DetailSPKDisplay.NoSPK.NoSPK ]['hargasatuan'] = hargasatuan

    print(rekaptransaksidisplay)
    # Rekap Bahan Baku Keluar
    databahanbakukeluar = datatransaksikeluar["Transaksibahanbaku"]['datatransaksi']
    rekapdatabahankeluar = {}
    for i in databahanbakukeluar:
        if i.DetailBahan in rekapdatabahankeluar:
            rekapdatabahankeluar[i.DetailBahan]['Jumlah'] += i.Jumlah
        else:
            rekapdatabahankeluar[i.DetailBahan] = {'Jumlah': i.Jumlah, "HargaFG":i.harga}
        rekapdatabahankeluar[i.DetailBahan]['hargatotal'] = rekapdatabahankeluar[i.DetailBahan]['Jumlah'] * rekapdatabahankeluar[i.DetailBahan]['HargaFG']
    
    # Rekap pemusnahan bahan baku
    datapemusnahanbahanbaku = datatransaksikeluar["Pemusnahanbahanbaku"]['datatransaksi']
    rekapdatapemusnahanbahanbaku = {}
    for i in datapemusnahanbahanbaku:
        if i.KodeBahanBaku in rekapdatapemusnahanbahanbaku:
            rekapdatapemusnahanbahanbaku[i.KodeBahanBaku]['Jumlah'] += i.Jumlah
        else:
            rekapdatapemusnahanbahanbaku[i.KodeBahanBaku] = {'Jumlah': i.Jumlah, "HargaFG":i.harga}
        rekapdatapemusnahanbahanbaku[i.KodeBahanBaku]['hargatotal'] = rekapdatapemusnahanbahanbaku[i.KodeBahanBaku]['Jumlah'] * rekapdatapemusnahanbahanbaku[i.KodeBahanBaku]['HargaFG']

    print(rekapdatapemusnahanbahanbaku)
    # Rekap Golongan D Keluar
    rekapgold = datatransaksikeluar['Transaksigolongand']['datatransaksi']
    rekapgolongand = {}
    for i in rekapgold:
        if i.KodeProduk in rekapgolongand:
            rekapgolongand[i.KodeProduk]['Jumlah'] += i.jumlah
        else:
            rekapgolongand[i.KodeProduk] = {"Jumlah" : i.jumlah,"Harga":i.harga}
        rekapgolongand[i.KodeProduk]['hargatotal'] = rekapgolongand[i.KodeProduk]['Jumlah'] * rekapgolongand[i.KodeProduk]['Harga']

    # print(rekapgolongand)
    # Rekap transaksi lain lain
    datatransaksilainlain = datatransaksikeluar['Transaksilainlain']['datatransaksi']
    print(datatransaksilainlain)
    rekaptransaksilainlain={}
    for i in datatransaksilainlain:
        if i.KodeProduk in rekaptransaksilainlain:
            rekaptransaksilainlain[i.KodeProduk]['Jumlah'] += i.jumlah
        else:
            rekaptransaksilainlain[i.KodeProduk] = {'Jumlah' : i.jumlah, "Harga" : i.harga}
        rekaptransaksilainlain[i.KodeProduk]['hargatotal'] = rekaptransaksilainlain[i.KodeProduk]['Jumlah'] * rekaptransaksilainlain[i.KodeProduk]['Harga']
    return rekapartikel,rekaptransaksidisplay,rekapdatabahankeluar,rekapgolongand,rekaptransaksilainlain,rekappemusnahanartikel,rekapdatapemusnahanbahanbaku


def gethargapurchasingperbulanperproduk(tanggal,kodeproduk):
    cachevalueobj = models.CacheValue.objects.filter(
        KodeProduk__KodeProduk = kodeproduk, Tanggal__month = tanggal.month
    ).first()
    print(cachevalueobj)
    print(kodeproduk,tanggal)
    print(cachevalueobj.Jumlah, cachevalueobj.Harga)
    print(tanggal)
    if cachevalueobj != None:
        totalsaldo = cachevalueobj.Harga * cachevalueobj.Jumlah
        datahargaperbulan = (cachevalueobj.Harga,cachevalueobj.Jumlah,totalsaldo)
    else:

        bahanbaku = models.Produk.objects.get(KodeProduk = kodeproduk)
        awaltahun = date(tanggal.year,1,1)
        akhirtahun = date(tanggal.year,12,31)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(tanggal.year, month)[1]
            last_days.append(date(tanggal.year, month, last_day))
        indexbulan = tanggal.month
        filtertanggalakhir = last_days[indexbulan-1]



        # Saldoawal Section
        saldoawalobj = models.SaldoAwalBahanBaku.objects.filter(
                IDBahanBaku=bahanbaku,
                Tanggal__range=(awaltahun, akhirtahun),
                IDLokasi__NamaLokasi="Gudang",
            )
        if saldoawalobj.exists():
            saldoawalobj = saldoawalobj.first()
            totalbiayaawal = saldoawalobj.Harga * saldoawalobj.Jumlah
            hargaawal = saldoawalobj.Harga
            jumlahawal = saldoawalobj.Jumlah
        else:
            totalbiayaawal = 0
            hargaawal = 0
            jumlahawal = 0

        hargamasukobj = models.DetailSuratJalanPembelian.objects.filter(
                KodeProduk=bahanbaku, NoSuratJalan__Tanggal__range=(awaltahun,filtertanggalakhir)
            )
        hargakeluarobj = models.TransaksiGudang.objects.filter(
                KodeProduk=bahanbaku, jumlah__gte=0,tanggal__range=(awaltahun,filtertanggalakhir))
        hargareturobj = models.TransaksiGudang.objects.filter(
                KodeProduk=bahanbaku, jumlah__lt = 0,tanggal__range=(awaltahun,filtertanggalakhir))
        pemusnahanobj = models.PemusnahanBahanBaku.objects.filter(
                lokasi__NamaLokasi = 'Gudang',KodeBahanBaku = bahanbaku, Tanggal__range = (awaltahun,filtertanggalakhir)
            )
        tanggalhargamasukobj = (
                    hargamasukobj
                    .values_list("NoSuratJalan__Tanggal", flat=True)
                    .distinct()
                )
        tanggalhargakeluarobj = (
                    hargakeluarobj
                    .values_list("tanggal", flat=True)
                    .distinct()
                )
        tanggalretur = hargareturobj.values_list('tanggal',flat=True).distinct()
        tanggalpemusnahan = pemusnahanobj.values_list('Tanggal',flat=True).distinct()
        listtanggal = sorted(list(set(tanggalhargamasukobj.union(tanggalhargakeluarobj).union(tanggalretur).union(tanggalpemusnahan))))

        datahargaperbulan = gethargabahanbaku(listtanggal,hargamasukobj,hargakeluarobj,hargareturobj,pemusnahanobj,hargaawal,jumlahawal)

    print(datahargaperbulan)
    # print(tes)
    return datahargaperbulan[0]




def getbarangkeluar(last_days, stopindex, awaltahun,hargapurchasing=None):
    hargapurchasingperbulan = hargapurchasing
    if hargapurchasingperbulan == None:
        hargapurchasingperbulan = gethargapurchasingperbulan(
            last_days, stopindex, awaltahun
        )

    datapenyusun = {}
    listdatadetailsppb = []
    biayakeluar = {}
    detailbiaya = {}
    listdatakeluar = []
    
    for index, hari in enumerate(last_days[:stopindex]):
        datamodelssppb = {
            'SPPBArtikel' : None,
            'totalbiayasppb' : None,
            'detailpenyusun' : None,

    }
        datamodelspemusnahanartikel = {
            'transaksipemusnahan':None,
            'totalbiayapemusnahan':None

    }
        datamodelspemusnahandisplay = {
            'transaksipemusnahan':None,
            'totalbiayapemusnahan':None

    }

        datamodeldisplay ={
            'SPPBDisplay': None,
            'totalbiayasppb' : None,
            'detailpenyusun' : None
        }
        datamodeltransaksigolongand = {
            'datatransaksi': None,
            'totalbiaya' : None
        }
        datamodeltransaksilainlain = {
            'datatransaksi' : None,
            'totalbiaya' : None
        }
        datamodeltransaksibahanbaku = {
            'datatransaksi' : None,
            'totalbiaya' : None
        }
        datamodelpemusnahanbahanbaku = {
            'datatransaksi' : None,
            'totalbiaya' : None
        }

        totalbiayakeluar = 0

        if index == 0:
            datadetailsppb = models.DetailSPPB.objects.filter(
                NoSPPB__Tanggal__lte=hari,
                NoSPPB__Tanggal__gte=awaltahun,
                DetailSPK__isnull=False,
            )
            # Transaksi Golongan D
            datatransaksigold = models.TransaksiGudang.objects.filter(
                KodeProduk__KodeProduk__istartswith="D",
                tanggal__range=(awaltahun, hari),
                jumlah__gte=0,
            )
            datatransaksilainlain = models.TransaksiGudang.objects.filter(
                tanggal__range=(awaltahun, hari),
                jumlah__gte=0,
                Lokasi__NamaLokasi="Lain-Lain",
            )
            datatransaksisppbdisplay = models.DetailSPPB.objects.filter(
            DetailSPKDisplay__isnull = False,NoSPPB__Tanggal__range = (awaltahun,hari)
            )
            datatransaksibahanbaku = models.DetailSPPB.objects.filter(
            DetailBahan__isnull = False,NoSPPB__Tanggal__range = (awaltahun,hari)
            )
            datapemusnahanbahanbaku = models.PemusnahanBahanBaku.objects.filter(Tanggal__range = (awaltahun,hari))
            datapemusnahanartikel = models.PemusnahanArtikel.objects.filter(Tanggal__range = (awaltahun,hari))
            
        else:
            datadetailsppb = models.DetailSPPB.objects.filter(
                NoSPPB__Tanggal__lte=hari,
                NoSPPB__Tanggal__gt=last_days[index - 1],
                DetailSPK__isnull=False,
            )
            # Transaksi Golongan D
            datatransaksigold = models.TransaksiGudang.objects.filter(
                KodeProduk__KodeProduk__istartswith="D",
                tanggal__range=(last_days[index - 1], hari),
                jumlah__gte=0,
            )
            datatransaksilainlain = models.TransaksiGudang.objects.filter(
                tanggal__range=(last_days[index - 1], hari),
                jumlah__gte=0,
                Lokasi__NamaLokasi="Lain-Lain",
            )
            datatransaksisppbdisplay = models.DetailSPPB.objects.filter(
            DetailSPKDisplay__isnull = False,NoSPPB__Tanggal__range = (last_days[index-1],hari)
            )
            datatransaksibahanbaku = models.DetailSPPB.objects.filter(
            DetailBahan__isnull = False,NoSPPB__Tanggal__range = (last_days[index-1],hari)
            )
            datapemusnahanbahanbaku = models.PemusnahanBahanBaku.objects.filter(Tanggal__range = (last_days[index-1],hari))
            datapemusnahanartikel = models.PemusnahanArtikel.objects.filter(Tanggal__range = (last_days[index-1],hari))


        if datadetailsppb.exists():
            for detailsppb in datadetailsppb:
                ge = gethargafgterakhirberdasarkanmutasi(
                    detailsppb.DetailSPK.KodeArtikel, hari, hargapurchasingperbulan
                )

                harga = detailsppb.Jumlah * ge[0]
                detailsppb.hargafg = ge[0]
                detailsppb.inputmanual = ge[1]
                detailsppb.totalharga = harga

                totalbiayakeluar += detailsppb.totalharga
                # print(totalbiayakeluar, detailsppb)
                # print(asd)
                # datapenyusun[detailsppb.DetailSPK.KodeArtikel] = {
                #     "WIP": ge[3][detailsppb.DetailSPK.KodeArtikel]["penyusun"],
                #     "FG": ge[4][detailsppb.DetailSPK.KodeArtikel]["penyusun"],
                #     "hargafg": ge[0],
                # }
                # detailsppb.datapenyusun=datapenyusun
                
        listdatadetailsppb.append(datadetailsppb)
        datamodelssppb["SPPBArtikel"] = datadetailsppb
        # datamodelssppb["detailpenyusun"] = datapenyusun
        datamodelssppb['totalbiayasppb'] = totalbiayakeluar
        #  Bahan Baku golongan D
        nilaigold = 0
        for bahand in datatransaksigold:
            harga = hargapurchasingperbulan[bahand.KodeProduk]["data"][index][
                "hargasatuan"
            ]
            bahand.harga = harga
            bahand.hargatotal = harga * bahand.jumlah
            # print(harga)
            nilaigold += bahand.hargatotal

        datamodeltransaksigolongand['datatransaksi'] = datatransaksigold
        datamodeltransaksigolongand["totalbiaya"] = nilaigold

        # Transaksi Lain lain
        nilailainlain = 0
        for tlainlain in datatransaksilainlain:
            harga = hargapurchasingperbulan[tlainlain.KodeProduk]["data"][index][
                "hargasatuan"
            ]
            tlainlain.harga = harga
            tlainlain.hargatotal = harga * tlainlain.jumlah
            nilailainlain += tlainlain.hargatotal
        
        datamodeltransaksilainlain["datatransaksi"] = (datatransaksilainlain)
        datamodeltransaksilainlain["totalbiaya"] = (nilailainlain)
        totalbiayadisplay = 0

        # Transaksi Pengiriman Bahan Baku
        nilaitransaksibahan = 0
        for transaksibahanbaku in datatransaksibahanbaku:
            harga = hargapurchasingperbulan[transaksibahanbaku.DetailBahan]["data"][index][
                "hargasatuan"
            ]
            transaksibahanbaku.harga = harga
            transaksibahanbaku.hargatotal = harga * transaksibahanbaku.Jumlah
            # print(harga)
            nilaitransaksibahan += transaksibahanbaku.hargatotal
        
        datamodeltransaksibahanbaku["datatransaksi"] = (datatransaksibahanbaku)
        datamodeltransaksibahanbaku["totalbiaya"] = (nilaitransaksibahan)

        nilaipemusnahanbahanbaku = 0
        for item in datapemusnahanbahanbaku:
            
            harga = hargapurchasingperbulan[item.KodeBahanBaku]["data"][index][
                "hargasatuan"
            ]
            item.harga = harga
            item.hargatotal = harga * item.Jumlah
            # print(harga)
            nilaipemusnahanbahanbaku += item.hargatotal

        datamodelpemusnahanbahanbaku['datatransaksi'] = datapemusnahanbahanbaku
        datamodelpemusnahanbahanbaku["totalbiaya"] = nilaipemusnahanbahanbaku

        nilaipemusnahanartikel = 0
        for item in datapemusnahanartikel:
            ge = gethargafgterakhirberdasarkanmutasi(
                   item.KodeArtikel, hari, hargapurchasingperbulan
                )
            

            harga = item.Jumlah * ge[0]
            item.hargafg = ge[0]
            item.inputmanual = ge[1]
            item.totalharga = harga

            nilaipemusnahanartikel += item.totalharga
        
        datamodelspemusnahanartikel['transaksipemusnahan'] = datapemusnahanartikel
        # datamodelssppb["detailpenyusun"] = datapenyusun
        datamodelspemusnahanartikel['totalbiayapemusnahan'] = nilaipemusnahanartikel

        for display in datatransaksisppbdisplay:

            
            '''
            1. Cari Jumlah berdasarkan SPK dulu 
            2.  Melihat jumlah bahan yang sudah diminta berdasarkan spk
            3. Menghitung berdasarkan pengeluaran 
            # '''

            dataproduksispk = display.DetailSPKDisplay.Jumlah
            totalpermintaanmenggunakanspk = models.TransaksiGudang.objects.filter(DetailSPKDisplay__NoSPK=display.DetailSPKDisplay.NoSPK).values('KodeProduk__KodeProduk').annotate(total = Sum('jumlah'))
            totalpengirimanproduk = (display.Jumlah / dataproduksispk) 

            penggunaanproduk = {}
            totalbiayakomponendisplay = 0
            for dataproduk in totalpermintaanmenggunakanspk:
                datapermintaan = models.TransaksiGudang.objects.filter(DetailSPKDisplay__NoSPK = display.DetailSPKDisplay.NoSPK,KodeProduk__KodeProduk = dataproduk['KodeProduk__KodeProduk'])
                totalpermintaan = datapermintaan.aggregate(total = Sum('jumlah'))['total']
                tanggalpermintaanawal = datapermintaan.order_by('tanggal').first()
                hargaawal = gethargapurchasingperbulanperproduk(tanggalpermintaanawal.tanggal,dataproduk['KodeProduk__KodeProduk'])
                print(hargaawal)
                print(tanggalpermintaanawal)
                # print(asd)
                # print(dataproduk)
                totalpermintaanperproduk = dataproduk['total']
                totalpenggunaanperdisplay = totalpengirimanproduk * totalpermintaanperproduk
                # print(totalpenggunaanperdisplay)
                # print(display.Jumlah, dataproduksispk)
                penggunaanproduk[dataproduk['KodeProduk__KodeProduk']] = {'jumlahpenggunaan':totalpenggunaanperdisplay,"biayaawal":hargaawal,"totalbiaya":hargaawal*totalpenggunaanperdisplay,'totalpermintaan':totalpermintaan}
                totalbiayakomponendisplay += hargaawal * totalpenggunaanperdisplay
            # print(penggunaanproduk)
            totalbiayadisplay += totalbiayakomponendisplay
            display.penggunaanbahan = penggunaanproduk
            display.totalbiaya = totalbiayakomponendisplay
            
        
        datamodeldisplay['SPPBDisplay']= (datatransaksisppbdisplay)
        datamodeldisplay['totalbiayasppb']= (totalbiayadisplay)


        biayakeluar[index] = totalbiayakeluar + nilaigold + nilailainlain + totalbiayadisplay + nilaitransaksibahan + nilaipemusnahanartikel + nilaipemusnahanbahanbaku
        detailbiaya[index] = {
            "artikelkeluar": totalbiayakeluar,
            "nilaigold": nilaigold,
            "nilailainlain": nilailainlain,
            'displaykeluar' : totalbiayadisplay
        }

        # print("\n\n\n\n", biayakeluar, detailbiaya)
        # pritn(asd)
        
        listdatakeluar.append({'SPPBArtikel':datamodelssppb,'SPPBDisplay': datamodeldisplay,'Transaksigolongand':datamodeltransaksigolongand,"Transaksibahanbaku":datamodeltransaksibahanbaku,"Transaksilainlain":datamodeltransaksilainlain,'datapenyusunartikel':datapenyusun,"totalkeluar":totalbiayakeluar + nilaigold + nilailainlain + totalbiayadisplay + nilaitransaksibahan + nilaipemusnahanbahanbaku + nilaipemusnahanartikel,'Pemusnahanbahanbaku':datamodelpemusnahanbahanbaku,'Pemusnahanartikel':datamodelspemusnahanartikel,})

        
    return (
        datadetailsppb,
        biayakeluar,
        datapenyusun,
        listdatadetailsppb,
        datatransaksilainlain,
        datatransaksigold,
        detailbiaya,
        listdatakeluar
    )


def gethargapurchasingperbulan(last_days, stopindex, awaltahun):
    bahanbaku = models.Produk.objects.all()
    akhirtahun = date(awaltahun.year,12,31)
    # bahanbaku = models.Produk.objects.filter(KodeProduk="A-101")

    hargaakhirbulanperproduk = {}
    for i in bahanbaku:
        saldoawalobj = models.SaldoAwalBahanBaku.objects.filter(
            IDBahanBaku=i,
            Tanggal__range=(awaltahun, akhirtahun),
            IDLokasi__NamaLokasi="Gudang",
        )
        # print(saldoawalobj)
        # print(asd)
        if saldoawalobj.exists():
            saldoawalobj = saldoawalobj.first()
            totalbiayaawal = saldoawalobj.Harga * saldoawalobj.Jumlah
            hargaawal = saldoawalobj.Harga
            jumlahawal = saldoawalobj.Jumlah
        else:
            totalbiayaawal = 0
            hargaawal = 0
            jumlahawal = 0
        hargamasukobj = models.DetailSuratJalanPembelian.objects.filter(
            KodeProduk=i, NoSuratJalan__Tanggal__range=(awaltahun,akhirtahun)
        )
        hargakeluarobj = models.TransaksiGudang.objects.filter(
            KodeProduk=i, tanggal__range=(awaltahun,akhirtahun),jumlah__gte= 0
        )
        hargareturobj =  models.TransaksiGudang.objects.filter(
            KodeProduk=i, tanggal__range=(awaltahun,akhirtahun),jumlah__lt = 0
        )
        pemusnahanobj = models.PemusnahanBahanBaku.objects.filter(
            lokasi__NamaLokasi = 'Gudang',KodeBahanBaku = i, Tanggal__range = (awaltahun,akhirtahun)
        )
        hargaakhirbulan = {}
        totalhargabahanbakugudangperbulan = 0
        for j, k in enumerate(last_days[:stopindex]):
            # print(j,k)
            tesvaluecache = models.CacheValue.objects.filter(KodeProduk = i,Tanggal = k)
            # print(tesvaluecache)
            # print(i)
            if tesvaluecache.exists():
                dataobj = tesvaluecache.first()
                hargaakhirbulan[j] = {
                "hargasatuan": dataobj.Harga,
                "jumlah": dataobj.Jumlah,
                "hargatotal": dataobj.Harga*dataobj.Jumlah,
            }
                totalhargabahanbakugudangperbulan += dataobj.Harga*dataobj.Jumlah
                continue

            else:
                # print(asd)
                tanggalhargamasukobj = (
                    hargamasukobj.filter(NoSuratJalan__Tanggal__lte=k)
                    .values_list("NoSuratJalan__Tanggal", flat=True)
                    .distinct()
                )
                tanggalhargakeluarobj = (
                    hargakeluarobj.filter(tanggal__lte=k)
                    .values_list("tanggal", flat=True)
                    .distinct()
                )
                listtanggal = sorted(
                    list(set(tanggalhargamasukobj.union(tanggalhargakeluarobj)))
                )

                suratjalanpembelianakhirbulanobj = hargamasukobj.filter(
                    NoSuratJalan__Tanggal__lte=k
                )
                transaksigudangakhirbulanobj = hargakeluarobj.filter(tanggal__lte=k)
                tanggalsuratjalanpembelianakhirbulanobj = (
                    suratjalanpembelianakhirbulanobj.values_list(
                        "NoSuratJalan__Tanggal", flat=True
                    ).distinct()
                )
                tanggaltransaksigudangakhirbulanobj = (
                    transaksigudangakhirbulanobj.values_list(
                        "tanggal", flat=True
                    ).distinct()
                )
                transaksireturakhirbulanobj = hargareturobj.filter(tanggal__lte=k)
                tanggalreturobj = transaksireturakhirbulanobj.values_list("tanggal", flat=True).distinct()
                transaksipemusnahanobj = pemusnahanobj.filter(Tanggal__lte=k)
                tanggalpemusnahanobj = transaksipemusnahanobj.values_list("Tanggal",flat=True).distinct()
                tanggalkeluarmasukperbulan = sorted(
                    list(
                        set(
                            tanggalsuratjalanpembelianakhirbulanobj.union(
                                tanggaltransaksigudangakhirbulanobj
                            ).union(tanggalpemusnahanobj).union(tanggalreturobj)
                        )
                    )
                )
                datahargaperbulan = gethargabahanbaku(
                    tanggalkeluarmasukperbulan,
                    suratjalanpembelianakhirbulanobj,
                    transaksigudangakhirbulanobj,
                    transaksireturakhirbulanobj,
                    transaksipemusnahanobj,
                    hargaawal,
                    jumlahawal,
                )
                # print(j, k, datahargaperbulan)
                hargaakhirbulan[j] = {
                    "hargasatuan": datahargaperbulan[0],
                    "jumlah": datahargaperbulan[1],
                    "hargatotal": datahargaperbulan[2],
                }
                totalhargabahanbakugudangperbulan += datahargaperbulan[2]
                dataobj = models.CacheValue.objects.update_or_create(
                    KodeProduk = i,
                    Tanggal = k,
                    defaults={
                    'Jumlah': datahargaperbulan[1],
                    'Harga': datahargaperbulan[0]
                    }
                )
        hargaakhirbulanperproduk[i] = {
            "data": hargaakhirbulan,
            "total": totalhargabahanbakugudangperbulan,
        }
    # print(hargaakhirbulanperproduk)
    # print(asdasd)
    return hargaakhirbulanperproduk


def gethargaartikelfgperbulan(artikel, tanggal, hargaakhirbulanperproduk):
    last_days = []
    for month in range(1, 13):
        last_day = calendar.monthrange(tanggal.year, month)[1]
        last_days.append(date(tanggal.year, month, last_day))

    dataartikel = models.Artikel.objects.filter(KodeArtikel=artikel)
    datahargawipartikel = {}
    print(tanggal)
    for artikel in dataartikel:
        dataperbulan = {}
        """
        Models perbulan
        {}
        """
        modelsperbulan = {}
        # print("\n\n\n\n", last_days[:tanggal.month])
        for index, hari in enumerate(last_days[: tanggal.month]):
            # Mengambil data terakhir versi penyusun tiap bulannya
            versiterakhirperbulan = (
                models.Penyusun.objects.filter(
                    KodeArtikel=artikel, KodeVersi__isdefault=True, Lokasi__NamaLokasi="FG"
                )
                .values_list("KodeVersi__Tanggal", flat=True)
                .distinct()
                .order_by("KodeVersi__Tanggal")
                .last()
            )
            if versiterakhirperbulan is None:
                versiterakhirperbulan = (
                    models.Penyusun.objects.filter(
                        KodeArtikel=artikel, KodeVersi__isdefault = True, Lokasi__NamaLokasi="FG"
                    )
                    .values_list("KodeVersi__Tanggal", flat=True)
                    .distinct()
                    .order_by("-KodeVersi__Tanggal")
                    .last()
                )
            penyusunversiterpilih = models.Penyusun.objects.filter(
                KodeArtikel=artikel,
                KodeVersi__isdefault=True,
                Lokasi__NamaLokasi="FG",
            )
            datapenyusun = {}
            hargawip = 0
            for penyusun in penyusunversiterpilih:
                dummy = {}
                hargapenyusun = hargaakhirbulanperproduk[penyusun.KodeProduk]["data"][
                    index
                ]["hargasatuan"]
                kuantitas =penyusun.Allowance

                hargabahanbakuwip = hargapenyusun * kuantitas
                hargawip += hargabahanbakuwip
                dummy["totalharga"] = hargabahanbakuwip
                dummy["kuantitas"] = kuantitas
                dummy["harga"] = hargapenyusun
                if penyusun.KodeProduk in datapenyusun:
                    datapenyusun[penyusun.KodeProduk]["kuantitas"] += dummy["kuantitas"]
                    datapenyusun[penyusun.KodeProduk]["totalharga"] += (
                        dummy["kuantitas"] * dummy["harga"]
                    )
                else:
                    datapenyusun[penyusun.KodeProduk] = dummy

            dummy2 = {}
            dummy2["penyusun"] = datapenyusun
            dummy2["hargafg"] = hargawip

        datahargawipartikel[artikel] = dummy2

    return datahargawipartikel


def gethargaartikelwipperbulan(artikel, tanggal, hargaakhirbulanperproduk):
    last_days = []
    for month in range(1, 13):
        last_day = calendar.monthrange(tanggal.year, month)[1]
        last_days.append(date(tanggal.year, month, last_day))

    dataartikel = models.Artikel.objects.filter(KodeArtikel=artikel)
    datahargawipartikel = {}
    for artikel in dataartikel:
        dataperbulan = {}
        """
        Models perbulan
        {}
        """
        modelsperbulan = {}
        for index, hari in enumerate(last_days[: tanggal.month]):
            # Mengambil data terakhir versi penyusun tiap bulannya
            versiterakhirperbulan = (
                models.Penyusun.objects.filter(
                    KodeArtikel=artikel, KodeVersi__isdefault = True, Lokasi__NamaLokasi="WIP"
                )
                .values_list("KodeVersi__Tanggal", flat=True)
                .distinct()
                .order_by("KodeVersi__Tanggal")
                .last()
            )
            if versiterakhirperbulan is None:
                versiterakhirperbulan = (
                    models.Penyusun.objects.filter(
                        KodeArtikel=artikel, KodeVersi__isdefault = True, Lokasi__NamaLokasi="WIP"
                    )
                    .values_list("KodeVersi__Tanggal", flat=True)
                    .distinct()
                    .order_by("-KodeVersi__Tanggal")
                    .last()
                )

            penyusunversiterpilih = models.Penyusun.objects.filter(
                KodeArtikel=artikel,
                KodeVersi__isdefault = True,
                Lokasi__NamaLokasi="WIP",
            )
            # print(versiterakhirperbulan)
            # print(penyusunversiterpilih)

            datapenyusun = {}
            hargawip = 0
            for penyusun in penyusunversiterpilih:
                dummy = {}
                hargapenyusun = hargaakhirbulanperproduk[penyusun.KodeProduk]["data"][
                    index
                ]["hargasatuan"]
                kuantitas =penyusun.Allowance
                hargabahanbakuwip = hargapenyusun * kuantitas
                hargawip += hargabahanbakuwip
                dummy["totalharga"] = hargabahanbakuwip
                dummy["kuantitas"] = kuantitas
                dummy["harga"] = hargapenyusun
                if penyusun.KodeProduk in datapenyusun:
                    # print(datapenyusun[penyusun.KodeProduk]["kuantitas"])
                    datapenyusun[penyusun.KodeProduk]["kuantitas"] += dummy["kuantitas"]
                    datapenyusun[penyusun.KodeProduk]["totalharga"] += (
                        dummy["kuantitas"] * dummy["harga"]
                    )
                    # print(datapenyusun[penyusun.KodeProduk]["kuantitas"])

                else:
                    datapenyusun[penyusun.KodeProduk] = dummy

            dummy2 = {}
            dummy2["penyusun"] = datapenyusun
            dummy2["hargawip"] = hargawip

        datahargawipartikel[artikel] = dummy2
    return datahargawipartikel


# def gethargafgperbulan(last_days, stopindex, awaltahun):
#     hargaakhirbulanperproduk = gethargapurchasingperbulan(
#         last_days, stopindex, awaltahun
#     )
#     # print(hargaakhirbulanperproduk)
#     # print(asd)
#     dataartikel = models.Artikel.objects.all()
#     datahargafgartikel = {}
#     for artikel in dataartikel:
#         dataperbulan = {}
#         """
#         Models perbulan
#         {}
#         """
#         modelsperbulan = {}
#         for index, hari in enumerate(last_days[:stopindex]):
#             # Mengambil data terakhir versi penyusun tiap bulannya
#             versiterakhirperbulan = (
#                 models.Penyusun.objects.filter(KodeArtikel=artikel, versi__lte=hari)
#                 .values_list("versi", flat=True)
#                 .distinct()
#                 .order_by("versi")
#                 .last()
#             )
#             # SEMENTARA PAKAI .LAST()
#             penyusunversiterpilih = models.Penyusun.objects.filter(
#                 KodeArtikel=artikel, versi=versiterakhirperbulan
#             )
#             datapenyusun = {}
#             hargafg = 0
#             for penyusun in penyusunversiterpilih:
#                 dummy = {}
#                 hargapenyusun = hargaakhirbulanperproduk[penyusun.KodeProduk]["data"][
#                     index
#                 ]["hargasatuan"]
#                 kuantitas = models.KonversiMaster.objects.get(
#                     KodePenyusun=penyusun
#                 ).Allowance

#                 hargabahanbakufg = hargapenyusun * kuantitas
#                 hargafg += hargabahanbakufg
#                 dummy["totalharga"] = hargabahanbakufg
#                 dummy["kuantitas"] = kuantitas
#                 dummy["harga"] = hargapenyusun
#                 datapenyusun[penyusun.KodeProduk] = dummy

#             dummy2 = {}
#             dummy2["penyusun"] = datapenyusun
#             dummy2["hargafg"] = hargafg

#             dataperbulan[index] = dummy2
#         datahargafgartikel[artikel] = dataperbulan

#     """Section Display"""
#     datadisplay = models.Display.objects.all()
#     datahargafgdisplay = {}

#     # for display in datadisplay:
#     #     dataperbulan = {}
#     #     modelsperbulan = {}
#     #     for index,hari in enumerate(last_days[:stopindex]):
#     #         ambilpenyu

#     # print(datahargafgartikel)
#     # print(asdasd)
#     return datahargafgartikel


def getbarangmasuk(last_days, stopindex, awaltahun):
    rekapbarangmasukperbulan = {}
    for index, hari in enumerate(last_days[:stopindex]):
        if index == 0:
            databahanbakumasuk = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__lte=hari, NoSuratJalan__Tanggal__gte=awaltahun
            )
        else:
            databahanbakumasuk = models.DetailSuratJalanPembelian.objects.filter(
                NoSuratJalan__Tanggal__lte=hari,
                NoSuratJalan__Tanggal__gt=last_days[index - 1],
            )
        nilaibahanbakumasuk = 0
        for item in databahanbakumasuk:
            biaya = item.Harga * item.Jumlah
            nilaibahanbakumasuk += biaya
            item.totalbiaya = biaya

        rekapbarangmasukperbulan[index] = {
            "data": databahanbakumasuk,
            "total": nilaibahanbakumasuk,
        }
    # print("tes")
    # print(rekapbarangmasukperbulan)
    # print(asdasd)
    return rekapbarangmasukperbulan


def detaillaporanbarangmasuk(request):
    if len(request.GET) > 0:
        print(request.GET)
        bulan = request.GET["bulan"]
        waktu = request.GET["waktu"]
        waktuobj = datetime.strptime(waktu, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = listbulan.index(request.GET["bulan"]) + 1
        # print(index)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))
        # print(last_days)
        # print(last_days[:index+1])
        bahanbakumasuk = getbarangmasuk(last_days, index, awaltahun)
        print(bahanbakumasuk)

        return render(
            request,
            "ppic/detaillaporanbarangmasuk.html",
            {
                "sjp": bahanbakumasuk[index - 1]["data"],
                "totalbiayamasuk": bahanbakumasuk[index - 1]["total"],
                "bulan": bulan,
            },
        )


def getstokgudang(awaltahun, last_days, stopindex, spessifikproduk=None):
    bahanbaku = models.Produk.objects.all()
    bahangudangperbulan = {}
    rekaphargabahanbakugudangperbulan = 0
    for index, hari in enumerate(last_days[:stopindex]):
        dummy = {}
        for produk in bahanbaku:

            barangmasukobj = models.DetailSuratJalanPembelian.objects.filter(
                KodeProduk=produk,
                NoSuratJalan__Tanggal__gte=awaltahun,
                NoSuratJalan__Tanggal__lte=hari,
            )
            barangkeluarobj = models.TransaksiGudang.objects.filter(
                KodeProduk=produk, tanggal__gte=awaltahun, tanggal__lte=hari
            )

            tanggalbarangmasukobj = barangmasukobj.values_list(
                "NoSuratJalan__Tanggal", flat=True
            ).distinct()
            tanggalbarangkeluar = barangkeluarobj.values_list(
                "tanggal", flat=True
            ).distinct()
            listtanggal = sorted(
                list(set(tanggalbarangmasukobj.union(tanggalbarangkeluar)))
            )
            print(index, hari, produk, listtanggal)
            jumlahawal = 0
            totalbiayaawal = 0
            for tanggal in listtanggal:
                jumlahmasukperhari = 0
                hargamasuktotalperhari = 0
                hargamasuksatuanperhari = 0
                jumlahkeluarperhari = 0
                hargakeluartotalperhari = 0
                hargakeluarsatuanperhari = 0
                jumlahmasukperhari = 0

                sjpobj = barangmasukobj.filter(NoSuratJalan__Tanggal=tanggal)
                if sjpobj.exists():
                    for k in sjpobj:
                        hargamasuktotalperhari += k.Harga * k.Jumlah
                        jumlahmasukperhari += k.Jumlah
                        hargamasuksatuanperhari += (
                            hargamasuktotalperhari / jumlahmasukperhari
                        )
                else:
                    hargamasuktotalperhari = 0
                    jumlahmasukperhari = 0
                    hargamasuksatuanperhari = 0

                transaksigudangobj = barangkeluarobj.filter(tanggal=tanggal)
                if transaksigudangobj.exists():
                    for k in transaksigudangobj:
                        jumlahkeluarperhari += k.jumlah
                        hargakeluartotalperhari += k.jumlah * hargasatuanawal
                    hargakeluarsatuanperhari += (
                        hargakeluartotalperhari / jumlahkeluarperhari
                    )
                else:
                    hargakeluartotalperhari = 0
                    hargakeluarsatuanperhari = 0
                    jumlahkeluarperhari = 0

                print(produk)
                print("Tanggal : ", tanggal)
                print("Sisa Stok Hari Sebelumnya : ", jumlahawal)
                print("harga awal Hari Sebelumnya :", hargasatuanawal)
                print("harga total Hari Sebelumnya :", totalbiayaawal)
                print("Jumlah Masuk : ", jumlahmasukperhari)
                print("Harga Satuan Masuk : ", hargamasuksatuanperhari)
                print("Harga Total Masuk : ", hargamasuktotalperhari)
                print("Jumlah Keluar : ", jumlahkeluarperhari)
                print("Harga Keluar : ", hargakeluarsatuanperhari)
                print(
                    "Harga Total Keluar : ",
                    hargakeluarsatuanperhari * jumlahkeluarperhari,
                )
                jumlahawal += jumlahmasukperhari - jumlahkeluarperhari
                totalbiayaawal += hargamasuktotalperhari - hargakeluartotalperhari
                try:
                    hargasatuanawal = totalbiayaawal / jumlahawal
                except ZeroDivisionError:
                    hargasatuanawal = 0

                print("Sisa Stok Hari Ini : ", jumlahawal)
                print("harga awal Hari Ini :", hargasatuanawal)
                print("harga total Hari Ini :", totalbiayaawal, "\n")
            dummy[produk] = {
                "hargasatuan": hargasatuanawal,
                "jumlah": jumlahawal,
                "totalbiaya": totalbiayaawal,
            }
            rekaphargabahanbakugudangperbulan += totalbiayaawal
        bahangudangperbulan[index] = {
            "data": dummy,
            "total": rekaphargabahanbakugudangperbulan,
        }
    return bahangudangperbulan


def detaillaporanbaranstokgudang(request):
    if len(request.GET) > 0:
        bulan = request.GET["bulan"]
        waktu = request.GET["waktu"]
        waktuobj = datetime.strptime(waktu, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)

        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = listbulan.index(request.GET["bulan"])
        # print(index)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))
        # print(last_days)
        # print(last_days[:index+1])
        bahanbakuperbulan = getstokgudang(awaltahun, last_days, index + 1)
        print(bahanbakuperbulan)
        return render(
            request,
            "ppic/detaillaporanstokgudang.html",
            {"stokgudang": bahanbakuperbulan, "bulan": bulan},
        )

def getpenyusunartikelpertanggal(tanggal,kodeartikel):
    versiterakhirperbulan = (
                models.Penyusun.objects.filter(
                    KodeArtikel__KodeArtikel=kodeartikel, KodeVersi__isdefault=True, Lokasi__NamaLokasi="FG"
                )
                .values_list("KodeVersi__Tanggal", flat=True)
                .distinct()
                .order_by("KodeVersi__Tanggal")
                .last()
            )
    if versiterakhirperbulan is None:
        versiterakhirperbulan = (
                    models.Penyusun.objects.filter(
                        KodeArtikel__KodeArtikel=kodeartikel, KodeVersi__isdefault = True, Lokasi__NamaLokasi="FG"
                    )
                    .values_list("KodeVersi__Tanggal", flat=True)
                    .distinct()
                    .order_by("-KodeVersi__Tanggal")
                    .last()
                )
    penyusunversiterpilih = models.Penyusun.objects.filter(
                KodeArtikel__KodeArtikel=kodeartikel,
                KodeVersi__isdefault = True,
                Lokasi__NamaLokasi="FG",
            )
    # datakonversimaster = models.KonversiMaster.objects.filter(KodePenyusun__in = penyusunversiterpilih)

    datapenyusun = penyusunversiterpilih.values('KodeProduk__KodeProduk').annotate(total = Sum('Allowance'))
    # print(datapenyusun)

    return datapenyusun
def getstokfg(request,lastdays, stopindex,awaltahun,hargapurchasing=None):
    waktuawal = time.time()
    if hargapurchasing == None:
        hargapurchasing = gethargapurchasingperbulan(lastdays,stopindex,awaltahun)
    waktuakhir = time.time()

    # print(asd)
    akhirtahun = date(awaltahun.year,12,31)
    data = []


    '''
    LOGIKA 
    1. Hitung Jumlah mutasi perartikel/display dari WIP ke FG
    2. Hitung Jumlah mutasi perartikel/display dari FG keluar (SPPB)
    4. Ambil Data produk yang diminta ke gudang dari awal tahun - sekarang
    3. Hitung kumulasi jumlah permintaan gudang ke FG
    '''

    for index, hari in enumerate(lastdays[:stopindex]):
        awal = time.time()
        if index == 0 :
            # Mengambil Saldo Awal Artikel pada FG
            datasaldoawalartikel = models.SaldoAwalArtikel.objects.filter(Tanggal__range=(awaltahun,akhirtahun),IDLokasi__NamaLokasi="FG")
            jumlahdatasaldoawal = datasaldoawalartikel.values("IDArtikel__KodeArtikel").annotate(total = Sum('Jumlah'))
            # Mengambil Saldo Awal Bahan Baku pada FG
            datasaldoawalbahanbaku = models.SaldoAwalBahanBaku.objects.filter(Tanggal__range=(awaltahun,akhirtahun),IDLokasi__NamaLokasi = "FG")
            jumlahdatasaldoawalbahanbaku = datasaldoawalbahanbaku.values("IDBahanBaku__KodeProduk").annotate(total = Sum('Jumlah'))

            #  Mengambil total data transaksi mutasi WIP ke FG
            datamutasiartikel = models.TransaksiProduksi.objects.filter(Lokasi__NamaLokasi="WIP",Tanggal__range =(awaltahun,hari),Jenis = "Mutasi", KodeArtikel__isnull = False )
            # Mengambil data transaksi display
            datamutasidisplay = models.TransaksiProduksi.objects.filter(Lokasi__NamaLokasi="WIP",Tanggal__range =(awaltahun,hari),Jenis = "Mutasi", KodeDisplay__isnull = False ).order_by('Tanggal')
            # Mengambil total data transaksi pengeluaran barang dari FG Artikel
            datapengeluaranartikel = models.DetailSPPB.objects.filter(NoSPPB__Tanggal__range = (awaltahun,hari),DetailSPK__isnull=False)
            # Mengambil total data transaksi pengeluaran barang dari FG Display
            datapengeluarandisplay = models.DetailSPPB.objects.filter(NoSPPB__Tanggal__range = (awaltahun,hari),DetailSPKDisplay__isnull=False)
            # Mengambil total permintaan transaksi bahan baku 
            datapermintaanbahanbaku = models.TransaksiGudang.objects.filter(Lokasi__NamaLokasi = "FG",tanggal__range=(awaltahun,hari))
        else:
            #  Mengambil total data transaksi mutasi WIP ke FG
            datamutasiartikel = models.TransaksiProduksi.objects.filter(Lokasi__NamaLokasi="WIP",Tanggal__range =(awaltahun,hari),Jenis = "Mutasi", KodeArtikel__isnull = False )
            # Mengambil data transaksi display
            datamutasidisplay = models.TransaksiProduksi.objects.filter(Lokasi__NamaLokasi="WIP",Tanggal__range =(awaltahun,hari),Jenis = "Mutasi", KodeDisplay__isnull = False ).order_by('Tanggal')
            # Mengambil total data transaksi pengeluaran barang dari FG Artikel
            datapengeluaranartikel = models.DetailSPPB.objects.filter(NoSPPB__Tanggal__range = (awaltahun,hari),DetailSPK__isnull=False)
            # Mengambil total data transaksi pengeluaran barang dari FG Display
            datapengeluarandisplay = models.DetailSPPB.objects.filter(NoSPPB__Tanggal__range = (awaltahun,hari),DetailSPKDisplay__isnull=False)
            # Mengambil total permintaan transaksi bahan baku 
            datapermintaanbahanbaku = models.TransaksiGudang.objects.filter(Lokasi__NamaLokasi = "FG",tanggal__range=(awaltahun,hari))
        waktuambildata = time.time()
        
        jumlahmutasiartikel = datamutasiartikel.values('KodeArtikel__KodeArtikel').annotate(total = Sum('Jumlah'))
        # print(jumlahmutasiartikel)

        jumlahmutasidisplay = datamutasidisplay.values('KodeDisplay__KodeDisplay').annotate(total = Sum('Jumlah'))
        # print(jumlahmutasidisplay)

        jumlahpengirimanartikel = datapengeluaranartikel.values('DetailSPK__KodeArtikel__KodeArtikel').annotate(total = Sum('Jumlah'))
        
        jumlahpengirimandisplay = datapengeluarandisplay.values('DetailSPKDisplay__KodeDisplay__KodeDisplay').annotate(total = Sum('Jumlah'))
        # print(jumlahpengirimandisplay)
        # print(asd)

        totalpermintaanbahanbaku = datapermintaanbahanbaku.values("KodeProduk__KodeProduk").annotate(total = Sum('jumlah'))
        datajenisproduk = datapermintaanbahanbaku.values_list("KodeProduk__KodeProduk",flat=True).distinct()
        # print(totalpermintaanbahanbaku)
        waktuvalues = time.time()

        # Mengurangi Antara jumlah Mutasi dan SPPB
        
        mutasi_dict = {item['KodeArtikel__KodeArtikel']: item['total'] for item in jumlahmutasiartikel}
        pengiriman_dict = {item['DetailSPK__KodeArtikel__KodeArtikel']: item['total'] for item in jumlahpengirimanartikel}
        saldoawal_dict = {item['IDArtikel__KodeArtikel']: item['total'] for item in jumlahdatasaldoawal}
        permintaanbahanbaku_dict = {item['KodeProduk__KodeProduk']: item['total'] for item in totalpermintaanbahanbaku}
        saldoawalbahanbaku_dict = {item['IDBahanBaku__KodeProduk']: item['total'] for item in jumlahdatasaldoawalbahanbaku}
        mutasidisplay_dict = {item['KodeDisplay__KodeDisplay']: item['total'] for item in jumlahmutasidisplay}
        pengirimandisplay_dict = {item['DetailSPKDisplay__KodeDisplay__KodeDisplay']: item['total'] for item in jumlahpengirimandisplay}
        all_kode_artikel = set(mutasi_dict.keys()).union(set(pengiriman_dict.keys()))
        # print(permintaanbahanbaku_dict)
        # print(pengiriman_dict)
        # print(asd)
        all_kode_artikel = models.Artikel.objects.all().values_list("KodeArtikel",flat=True)
        # all_kode_artikel = models.Artikel.objects.filter(KodeArtikel = "penyusun display AC Medium").values_list("KodeArtikel",flat=True)
        all_kode_display = models.Display.objects.all()
        result = []
        resultdengansaldoawal = []
        totalpenggunaanbahanbaku = {}
        totalsaldoartikel = 0
        timegeneratedata = time.time()
        # print('waktugeneratedata : ',timegeneratedata-awal)
        waktuartikel = time.time()
        nilaiwaktukonversi = 0
        nilaiwaktuharga = 0
        for kode_artikel in all_kode_artikel:
            ceksaldoawalartikel = 0
            has_saldoawalartikel = models.SaldoAwalArtikel.objects.filter(IDArtikel__KodeArtikel = kode_artikel)
            if has_saldoawalartikel:
                ceksaldoawalartikel = has_saldoawalartikel.first().Jumlah
            has_transaksiproduksi = models.TransaksiProduksi.objects.filter(KodeArtikel__KodeArtikel = kode_artikel, Tanggal__year = hari.year)
            has_detailsppb = models.DetailSPPB.objects.filter(DetailSPK__KodeArtikel__KodeArtikel = kode_artikel, NoSPPB__Tanggal__year =hari.year)
            total_mutasi = mutasi_dict.get(kode_artikel, 0)
            total_pengiriman = pengiriman_dict.get(kode_artikel, 0)
            total_saldoawal = saldoawal_dict.get(kode_artikel, 0)
            startkonversi = time.time()
            konversibahanbaku = getpenyusunartikelpertanggal(hari,kode_artikel)
            akhirkonversi = time.time()
            nilaiwaktukonversi += akhirkonversi - startkonversi
            if not has_saldoawalartikel.exists() and  not has_transaksiproduksi.exists() and  not has_detailsppb.exists():
                print(total_mutasi,total_pengiriman,total_saldoawal)
                print('tidak ada data artikel ', kode_artikel)
                resultdengansaldoawal.append({'KodeArtikel': kode_artikel, 'total':total,"penyusunfg":konversibahanbaku,'hargafg':0,'totalsaldo':0})

                continue
            # print('waktu cek penyusun artikel : ',kode_artikel,akhirkonversi-startkonversi)
            
            # print(konversibahanbaku)
            # print(total_pengiriman)
            # print(asd)
            
            if (total_mutasi != 0) and konversibahanbaku.exists():
                for i in konversibahanbaku:
                    # print(i)
                    # print(asd)
                    total_artikeljadi = abs(total_mutasi)
                    penggunaanbahanbakufg = i['total'] * total_artikeljadi
                    # print(penggunaanbahanbakufg)
                    if i['KodeProduk__KodeProduk'] in totalpenggunaanbahanbaku:
                        totalpenggunaanbahanbaku[i['KodeProduk__KodeProduk']] += penggunaanbahanbakufg
                    else:
                        totalpenggunaanbahanbaku[i['KodeProduk__KodeProduk']] = penggunaanbahanbakufg
                # print(total_artikeljadi,kode_artikel,konversibahanbaku)
                # print(totalpenggunaanbahanbaku)
                # print(asd)
            waktuhargaawal = time.time()
            hargaterakhir = gethargafgterakhirberdasarkanmutasi(models.Artikel.objects.get(KodeArtikel = kode_artikel),hari,hargapurchasing)
            waktuhargaakhir = time.time()
            nilaiwaktuharga += waktuhargaakhir-waktuhargaawal
            total = total_saldoawal + total_mutasi - total_pengiriman
            totalsaldo = hargaterakhir[0] * total
            totalsaldoartikel += totalsaldo
            resultdengansaldoawal.append({'KodeArtikel': kode_artikel, 'total':total,"penyusunfg":konversibahanbaku,'hargafg':hargaterakhir[0],'totalsaldo':totalsaldo})
        waktuakhirartikel = time.time()
        print("waktu ambil data database : ", waktuambildata - waktuawal )
        print('Waktu Values : ',waktuvalues-waktuambildata )
        print("Waktu DIcitionary : ",timegeneratedata - waktuvalues)
        print('waktu proses artikel : ', waktuakhirartikel-waktuartikel)
        print('Waktu Konversi : ',nilaiwaktukonversi)
        print('Waktu Harga : ',nilaiwaktuharga)
        print('watku proses - artikel section ',waktuakhirartikel - waktuawal)
        # print(resultdengansaldoawal)
        # print(asd)
        timedataartikel = time.time()
        # print('waktu artikel : ',timedataartikel-timegeneratedata)
        # print(asd)

        resultdisplay = []
        totalsaldodisplay = 0
        waktudisplay = time.time()
        for kode_display in all_kode_display:
            total_mutasi = mutasidisplay_dict.get(kode_display.KodeDisplay, 0)
            total_pengiriman = pengirimandisplay_dict.get(kode_display.KodeDisplay, 0)
            total = total_mutasi - total_pengiriman
            print(kode_display,total_mutasi,total_pengiriman,total)
            print(datamutasidisplay)
            # print(asd)
            nilaijumlahkirim = total_pengiriman
           
            if total_mutasi > 0 or total_pengiriman > 0:
                valid = False
                for num,datadetailmutasidisplay in enumerate(datamutasidisplay):
                    sisapermintaan = datadetailmutasidisplay.Jumlah - nilaijumlahkirim
                    if sisapermintaan < 0 :
                        nilaijumlahkirim = abs(sisapermintaan)
                        continue
                    else:
                        datamutasidisplay = datamutasidisplay[num:]
                        sisapermintaan = sisapermintaan
                        valid = True
                        break
                # print(datamutasidisplay)
                # print(asd)
                if valid : 
                    stokxhargaperspk = 0
                    jumlahspk = 0
                    weightedaverage = 0

                    for num,datadetailmutasidisplay in enumerate(datamutasidisplay):
                        hargafgdisplay = 0
                        if num == 0 :
                            jumlahstokdisplay = sisapermintaan
                        #    print(sisapermintaan)
                        #    print(asd)
                        else:
                            jumlahstokdisplay = datadetailmutasidisplay.Jumlah
                            # print(jumlahstokdisplay)
                            # print(asd)
                        spkpermintaanproduk = datadetailmutasidisplay.DetailSPKDisplay
                        permintaanproduk = models.TransaksiGudang.objects.filter(DetailSPKDisplay = spkpermintaanproduk).values("KodeProduk__KodeProduk").annotate(total=Sum('jumlah'))
                        # print(permintaanproduk,spkpermintaanproduk,spkpermintaanproduk.Jumlah)
                        for k in permintaanproduk:
                            hargaterakhir = gethargapurchasingperbulanperproduk(hari,k['KodeProduk__KodeProduk'])
                            # print(hargaterakhir,k,jumlahstokdisplay)
                            jumlahkonversispk = k['total']/spkpermintaanproduk.Jumlah
                            jumlahhargakonversifgperbahanbaku = hargaterakhir * jumlahkonversispk * jumlahstokdisplay
                            hargafgdisplay += jumlahhargakonversifgperbahanbaku
                        stokxhargaperspk += hargafgdisplay
                        jumlahspk += jumlahstokdisplay
                        print('ini jumlah',jumlahstokdisplay)

                        print(hargafgdisplay)
                    
                    try:
                        weightedaverage = stokxhargaperspk / jumlahspk
                    except ZeroDivisionError:
                        weightedaverage = 0
                    print(stokxhargaperspk,jumlahspk)
                    print(weightedaverage)
                    # print(asd)
                    
                else:
                    stokxhargaperspk = 0
                    jumlahspk = sisapermintaan
                    weightedaverage = 0
                    if request != None:
                        messages.error(request,f'Terjadi Kesalahan perhitungan pada {kode_display} Cek kembali mutasi dan Saldo awal display')
            else:
                stokxhargaperspk = 0
                jumlahspk = 0
                weightedaverage = 0
            totalsaldodisplay += stokxhargaperspk
            resultdisplay.append({'KodeDisplay': kode_display, 'total':jumlahspk,'hargafg':weightedaverage,'totalsaldo':stokxhargaperspk})

        waktuakhirdisplay = time.time()
        print('waktu proses display : ', waktuakhirdisplay-waktudisplay)

        sisabahanbaku = []
        totalsaldobahanbaku = 0
        bahanbakuall = models.Produk.objects.all()
        # bahanbakuall = models.Produk.objects.filter(KodeProduk = "tesbahanbaku")
        waktubahan = time.time()
        for item in bahanbakuall:
            total_permintaanbarang = permintaanbahanbaku_dict.get(item.KodeProduk, 0)
            total_saldoawalbahanbaku = saldoawalbahanbaku_dict.get(item.KodeProduk,0)
            totalbahanbaku = total_permintaanbarang + total_saldoawalbahanbaku
            if item.KodeProduk in totalpenggunaanbahanbaku.keys():
                # print(f"item {item} ada di penggunaanbahanbaku")
                jumlahpenggunaan = totalpenggunaanbahanbaku[item.KodeProduk]
                totalbahanbaku -= jumlahpenggunaan
            permintaanterakhir = models.TransaksiGudang.objects.filter(Lokasi__NamaLokasi = "FG",tanggal__range=(awaltahun,hari),KodeProduk = item).order_by('-tanggal').first()
            hargaterakhir = 0
            if permintaanterakhir:
                hargaterakhir = gethargapurchasingperbulanperproduk(permintaanterakhir.tanggal,item)
            totalsaldo = totalbahanbaku * hargaterakhir
            totalsaldobahanbaku += totalsaldo
            sisabahanbaku.append({'KodeProduk':item,'total':totalbahanbaku,'hargasatuan':hargaterakhir,"hargatotal":totalsaldo})

        waktuakhirbahan = time.time()
        print('waktu akhir bahan : ',waktuakhirbahan - waktubahan)
        total_saldofg = totalsaldoartikel + totalsaldobahanbaku + totalsaldodisplay
        
        data.append({'Artikel':resultdengansaldoawal,"BahanBaku":sisabahanbaku,"Display":resultdisplay,"totalsaldo":total_saldofg,'totalperbagian':{'artikel':totalsaldoartikel,'display':totalsaldodisplay,'bahanbaku':totalsaldobahanbaku}})
        
        akhir1loop = time.time()
    print('waktu akhir 1 loop : ',akhir1loop-waktuawal)
    # print(asd)
        
    return data
    
def detaillaporanstokfg(request):
    if len(request.GET) > 0:
        startawal = time.time()
        bulan = request.GET["bulan"]
        waktu = request.GET["waktu"]
        waktuobj = datetime.strptime(waktu, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = listbulan.index(request.GET["bulan"]) + 1
        # print(index)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))
        # print(last_days)
        # print(last_days[:index+1])
        hargaakhirbulanperproduk = gethargapurchasingperbulan(
        last_days, index, awaltahun
    )
        tes = getstokfg(request,last_days,index,awaltahun,hargaakhirbulanperproduk)
        # datastokgudang, bahanbakusisafg = getstokartikelfg(last_days, index, awaltahun)
        waktuproses = time.time()
        print(waktuproses-startawal)

        return render(
            request,
            "ppic/detaillaporanstokfg.html",
            {
                "artikelfg": tes[-1]['Artikel'],
                "sisabahanfg": tes[-1]['BahanBaku'],
                'displayfg' : tes[-1]['Display'],
                'totalsaldofg' : tes[-1]['totalsaldo'],
                'saldoperbagian' : tes[-1]['totalperbagian'],
                "bulan": bulan,
            },
        )



def cekmutasifgterakhir(artikel, tanggaltes):
    mutasifg = (
        models.DetailSPPB.objects.filter(
            DetailSPK__KodeArtikel=artikel, NoSPPB__Tanggal__lte=tanggaltes
        )
        .values_list("NoSPPB__Tanggal", flat=True)
        .distinct()
        .order_by("-NoSPPB__Tanggal")
        .first()
    )
    mutasimasukfg = (
        models.TransaksiProduksi.objects.filter(
            KodeArtikel=artikel, Tanggal__lte=tanggaltes, Jenis="Mutasi"
        )
        .values_list("Tanggal", flat=True)
        .distinct()
        .order_by("-Tanggal")
        .first()
    )

    if mutasifg or mutasimasukfg:
        if mutasifg is not None and mutasimasukfg is not None:
            return max(mutasifg, mutasimasukfg)
        if mutasifg:
            return mutasifg
        else:
            return mutasimasukfg

    else:
        return tanggaltes


def cekmutasiwipfgterakhir(artikel, tanggaltes):
    tanggalawalbulan = date(tanggaltes.year, tanggaltes.month, 1)
    mutasiwip = (
        models.TransaksiProduksi.objects.filter(
            KodeArtikel=artikel, Jenis="Mutasi", Tanggal__lte=tanggaltes
        )
        .values_list("Tanggal", flat=True)
        .distinct()
        .order_by("-Tanggal")
        .first()
    )
    if mutasiwip:
        return mutasiwip
    else:
        return tanggaltes


def cekmutasiwipterakhir(artikel, tanggaltes):
    tanggalawalbulan = date(tanggaltes.year, tanggaltes.month, 1)
    mutasiwip = (
        models.TransaksiProduksi.objects.filter(
            KodeArtikel=artikel, Jenis="Mutasi", Tanggal__lte=tanggaltes
        )
        .values_list("Tanggal", flat=True)
        .distinct()
        .order_by("-Tanggal")
        .first()
    )
    getbahanbakuutama = models.Penyusun.objects.filter(KodeArtikel=artikel.id, Status=1, KodeVersi__isdefault = True).order_by('-KodeVersi__Tanggal').first()


    if getbahanbakuutama:
        transaksigudang = (
            models.TransaksiGudang.objects.filter(
                DetailSPK__KodeArtikel=artikel,
                tanggal__lte=tanggaltes,
                KodeProduk=getbahanbakuutama.KodeProduk,
            )
            .values_list("tanggal", flat=True)
            .distinct()
            .order_by("-tanggal")
            .first()
        )
        # if transaksigudang:
        #     print(artikel)
        #     print(getbahanbakuutama)
        #     print(transaksigudang)
            # print(asd)
    else:
        # print("Versi belum di set")
        transaksigudang = None
    # print(
    #     "\n\n\n\mutasi akhir wip : ",
    #     mutasiwip,
    #     "Mutasi Transaksi Gudnag : ",
    #     transaksigudang,
    # )
    if mutasiwip or transaksigudang:
        if mutasiwip is None and transaksigudang is not None:
            return transaksigudang
        elif mutasiwip is not None and transaksigudang is None:
            return mutasiwip
        else:
            return max(mutasiwip, transaksigudang)
    else:
        return tanggaltes


def perhitunganmutasiwipterakhir(artikel, tanggaltes):
    mutasiwip = (
        models.TransaksiProduksi.objects.filter(
            KodeArtikel=artikel, Jenis="Mutasi", Tanggal__lte=tanggaltes
        )
        .values_list("Tanggal", flat=True)
        .distinct()
        .order_by("-Tanggal")
        .first()
    )
    if mutasiwip:
        index = int(mutasiwip.month)
        awaltahun = datetime(mutasiwip.year, 1, 1)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(mutasiwip.year, month)[1]
            last_days.append(date(mutasiwip.year, month, last_day))

        hargamutasiakhirbulan = gethargaartikelwipperbulan(
            last_days, index, awaltahun, artikel
        )

        hargamutasiakhirbulan[artikel][last_days[index]] = hargamutasiakhirbulan[
            artikel
        ][last_days[index - 1]]

        return hargamutasiakhirbulan
    else:
        print("Todal ada data mutasi WIP terekam pada database")
        print(tanggaltes)
        index = int(tanggaltes.month)
        awaltahun = datetime(tanggaltes.year, 1, 1)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(tanggaltes.year, month)[1]
            last_days.append(date(tanggaltes.year, month, last_day))

        hargafgperbulan = gethargapurchasingperbulan(last_days, index, awaltahun)
        return hargafgperbulan


def perhitunganmutasifgterakhir(artikel, tanggaltes):
    print("cek FG", artikel, tanggaltes)
    mutasifg = (
        models.DetailSPPB.objects.filter(
            DetailSPK__KodeArtikel=artikel, NoSPPB__Tanggal__lte=tanggaltes
        )
        .values_list("NoSPPB__Tanggal", flat=True)
        .distinct()
        .order_by("-NoSPPB__Tanggal")
        .first()
    )
    mutasifgtanggal = (
        models.DetailSPPB.objects.filter(
            DetailSPK__KodeArtikel=artikel, NoSPPB__Tanggal__lte=tanggaltes
        )
        .values_list("NoSPPB__Tanggal", flat=True)
        .distinct()
        .order_by("-NoSPPB__Tanggal")
    )
    if mutasifg:
        index = int(mutasifg.month)
        awaltahun = datetime(mutasifg.year, 1, 1)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(mutasifg.year, month)[1]
            last_days.append(date(mutasifg.year, month, last_day))
        # print(last_days)
        hargafgperbulan = gethargaartikelfgperbulan(
            last_days, index, awaltahun, artikel
        )
        # print("\n\n Ini Harga FG Bulan : ", mutasifg, hargafgperbulan[artikel])
        # print(artikel)
        return hargafgperbulan
    else:
        print("Todal ada data mutasi terekam pada database")
        print(tanggaltes)
        index = int(tanggaltes.month)
        awaltahun = datetime(tanggaltes.year, 1, 1)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(tanggaltes.year, month)[1]
            last_days.append(date(tanggaltes.year, month, last_day))

        hargafgperbulan = gethargapurchasingperbulan(last_days, index, awaltahun)
        return hargafgperbulan


def getstokartikelfg(last_days, stopindex, awaltahun):
    datastokfgperbulan = {}
    starttime = time.time()
    hargaakhirbulanperproduk = gethargapurchasingperbulan(
        last_days, stopindex, awaltahun
    )
    endtime = time.time()
    print("Waktu selesai : ", endtime - starttime)
    # print(hargaakhirbulanperproduk)

    datastokbahanbakufg = {}
    for index, hari in enumerate(last_days[:stopindex]):
        if index == 0:
            datatransaksiproduksi = models.TransaksiProduksi.objects.filter(
                Tanggal__lte=hari, Tanggal__gte=awaltahun, Jenis="Mutasi"
            )
            databarangkeluar = models.DetailSPPB.objects.filter(
                NoSPPB__Tanggal__lte=hari, NoSPPB__Tanggal__gte=awaltahun
            )

        else:
            datatransaksiproduksi = models.TransaksiProduksi.objects.filter(
                Tanggal__lte=hari, Tanggal__gt=last_days[index - 1], Jenis="Mutasi"
            )
            databarangkeluar = models.DetailSPPB.objects.filter(
                NoSPPB__Tanggal__lte=hari, NoSPPB__Tanggal__gt=last_days[index - 1]
            )

        jumlahkumulatifbiayaperbulan = 0
        dumystok = {}
        dummy = {}

        datamodelsisabarangfg = {}
        bahanbakurequestkefg = models.Produk.objects.all()
        listbahanbaku = []
        for bahanbaku in bahanbakurequestkefg:
            if index == 0:

                stokdifg = 0
                datatransaksigudangfg = models.TransaksiGudang.objects.filter(
                    tanggal__gte=awaltahun,
                    tanggal__lte=hari,
                    Lokasi__NamaLokasi="FG",
                    KodeProduk=bahanbaku,
                )
            else:
                stokdifg = datastokbahanbakufg[index - 1][bahanbaku]["stok"]

                datatransaksigudangfg = models.TransaksiGudang.objects.filter(
                    tanggal__gt=last_days[index - 1],
                    tanggal__lte=hari,
                    Lokasi__NamaLokasi="FG",
                    KodeProduk=bahanbaku,
                )
            if datatransaksigudangfg.exists():
                stokdifg += datatransaksigudangfg.aggregate(total=Sum("jumlah"))[
                    "total"
                ]
                # print(index, stokdifg, bahanbaku)
                # print(datatransaksigudangfg)

            hargasatuan = hargaakhirbulanperproduk[bahanbaku]["data"][index][
                "hargasatuan"
            ]
            datamodelsisabarangfg[bahanbaku] = {
                "stok": stokdifg,
                "Hargasatuan": hargasatuan,
                "Hargatotal": 0,
            }
        # print('sebelum',datastokbahanbakufg)
        datastokbahanbakufg[index] = datamodelsisabarangfg
        # print('sesudah',datastokbahanbakufg)

        # semuaartikel = models.Artikel.objects.all()
        semuaartikel = models.Artikel.objects.filter(KodeArtikel="artikelmultiple")
        print(semuaartikel)
        # print(asd)
        for dataartikel in semuaartikel:
            if index == 0:
                stokawal = 0
            else:
                stokawal = datastokfgperbulan[index - 1]["data"][dataartikel]["jumlah"]
            # cek mutasi produk
            totalbiayafg, datakomponenwip, datakomponenfg, komponenwip, komponenfg = (
                gethargafgterakhirberdasarkanmutasi(
                    dataartikel, hari, hargaakhirbulanperproduk
                )
            )

            """KONDISI
            1. MUTASI WIP MUTASI FG --> Harga total FG di update
            2. Mutasi WIP, tidak mutasi FG --> Harga total tidak diupdate
            3. tidka Mutasi WIP, MUtasi FG --> Harga total tidak diupdate
            4. Tidak keduanya --> harga tidak di update
            """

            jumlahartikelmutasiperbulan = datatransaksiproduksi.filter(
                KodeArtikel=dataartikel
            )
            if jumlahartikelmutasiperbulan.exists():
                totaltransaksimutasi = jumlahartikelmutasiperbulan.aggregate(
                    total=Sum("Jumlah")
                )["total"]
                stokawal += totaltransaksimutasi

            jumlahkeluarartikelmutasiperbulan = databarangkeluar.filter(
                DetailSPK__KodeArtikel=dataartikel
            )
            if jumlahkeluarartikelmutasiperbulan.exists():
                totaltransaksikeuar = jumlahkeluarartikelmutasiperbulan.aggregate(
                    total=Sum("Jumlah")
                )["total"]
                stokawal -= totaltransaksikeuar

            """Data Request ke FG"""
            totalbiaya = totalbiayafg * stokawal

            dummy[dataartikel] = {
                "jumlah": stokawal,
                "hargawip": datakomponenfg,
                "hargafg": totalbiayafg,
                "biaya": totalbiaya,
            }
            jumlahkumulatifbiayaperbulan += totalbiaya
        datastokfgperbulan[index] = {
            "data": dummy,
            "total": jumlahkumulatifbiayaperbulan,
        }
        """Menghitung total persediaan bahan baku fg"""
        total = 0
        for item in datamodelsisabarangfg.values():
            item["Hargatotal"] = item["stok"] * item["Hargasatuan"]
            total += item["Hargatotal"]
        datamodelsisabarangfg["total"] = total
    """
    Data Models
    2: {'data': {<Artikel: 9010/AC>: {'jumlah': 1365, 'hargafg': 40993.59298941417, 'biaya': 55956254.43055034}, <Artikel: 5111/EXP>: {'jumlah': 222, 'hargafg': 400.0974675766822, 'biaya': 88821.63780202346}}, 'total': 56045076.068352364}}  
    """
    # print(datastokfgperbulan)
    return datastokfgperbulan, datastokbahanbakufg


def getsaldoawalgudang(request):
    if len(request.GET) > 0:
        bulan = request.GET["bulan"]
        waktu = request.GET["waktu"]
        waktuobj = datetime.strptime(waktu, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = listbulan.index(request.GET["bulan"]) + 1
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))
        data = saldogudang(last_days, index, awaltahun)
        # print(data)
        # print(adas)
        return render(
            request,
            "ppic/detailstockgudang.html",
            {
                "data": data,
            },
        )


def saldogudang(last_days, stopindex, awaltahun,hargapurchasing=None):
    hargaakhirbulanperproduk = hargapurchasing
    if hargaakhirbulanperproduk == None:
        hargaakhirbulanperproduk = gethargapurchasingperbulan(
        last_days, stopindex, awaltahun
        )
    # print(hargaakhirbulanperproduk)
    # print(asd)
    saldoawal = {}
    saldoakhirgudang = {}
    totalbiayasaldoawal = 0
    bahanbaku = models.Produk.objects.all()
    # bahanbaku = models.Produk.objects.filter(KodeProduk="A-001-06")
    for produk in bahanbaku:
        saldoawalobj = models.SaldoAwalBahanBaku.objects.filter(
            IDBahanBaku=produk, Tanggal__gte=awaltahun, IDLokasi__NamaLokasi="Gudang"
        )
        if saldoawalobj.exists():
            saldoawalobj = saldoawalobj.first()
            totalbiayaawal = saldoawalobj.Harga * saldoawalobj.Jumlah
            hargasatuanawal = saldoawalobj.Harga
            jumlahawal = saldoawalobj.Jumlah

        else:
            totalbiayaawal = 0
            hargasatuanawal = 0
            jumlahawal = 0

        totalbiayasaldoawal += totalbiayaawal
        produk.hargasatuanawal = hargasatuanawal
        produk.totalbiayaawal = totalbiayaawal
        produk.jumlahawal = jumlahawal

    
    # print(hargaakhirbulanperproduk)
    # print(asd)
    dataproduk = models.Produk.objects.all()
    for index, hari in enumerate(last_days[:stopindex]):
        totalbiayasaldoakhirperbulan = 0
        for produk in dataproduk:
            hargaproduk = hargaakhirbulanperproduk[produk]["data"][index]["hargatotal"]
            totalbiayasaldoakhirperbulan += hargaproduk
        saldoakhirgudang[index] = totalbiayasaldoakhirperbulan
        if index == 0:
            saldoawal[index] = totalbiayasaldoawal
        else:
            saldoawal[index] = saldoakhirgudang[index - 1]

    # print(saldoawal)
    # print(saldoakhirgudang)

    # print(asdasd)
    hargaakhirperbulan = {}
    # print(hargaakhirbulanperproduk)

    """SAVE ERROR"""
    nilaiakhir = 0
    for artikel, data in hargaakhirbulanperproduk.items():
        hargaakhirperbulan[artikel] = data["data"][index]
        nilaiakhir += data["data"][index]["hargatotal"]

    # print(hargaakhirperbulan)

    data = {
        "saldoawal": saldoawal,
        "saldoakhir": saldoakhirgudang,
        "datasaldoawal": bahanbaku,
        "hargaakhirbulanperproduk": {"data": hargaakhirperbulan, "total": nilaiakhir},
    }
    return data


"""
Perhitungan Saldo Gudang perlu di pecah antara saldo awal dan saldo akhir di laporan besar
"""


from django.db.models import Sum

def getstokbahanproduksi(last_days, stopindex, awaltahun):
    """UNTUK MENGHITUNG DATA SALDO AWAL BAHAN BAKU PADA WIP DAN FG"""
    bahanbaku = models.Produk.objects.all()
    artikels  = models.Artikel.objects.all()


    stokawalbahanbaku = {}

    totalbiayasaldoawal_wip = 0
    totalbiayasaldoawal_fg = 0

    for produk in bahanbaku:
        saldoawalobj = models.SaldoAwalBahanBaku.objects.filter(
            Tanggal__year=awaltahun.year,
            IDLokasi__NamaLokasi__in=["WIP", "FG"],
            IDBahanBaku=produk,
        )

        totalbiayaawal_wip = 0
        totalbiayaawal_fg = 0
        jumlahawal_wip = 0
        jumlahawal_fg = 0

        if saldoawalobj.exists():
            for dataproduk in saldoawalobj:
                if dataproduk.IDLokasi.NamaLokasi == "WIP":
                    totalbiayaawal_wip += dataproduk.Harga * dataproduk.Jumlah
                    jumlahawal_wip += dataproduk.Jumlah
                elif dataproduk.IDLokasi.NamaLokasi == "FG":
                    totalbiayaawal_fg += dataproduk.Harga * dataproduk.Jumlah
                    jumlahawal_fg += dataproduk.Jumlah

        produk.hargasatuanawal_wip = totalbiayaawal_wip / jumlahawal_wip if jumlahawal_wip != 0 else 0
        produk.hargasatuanawal_fg = totalbiayaawal_fg / jumlahawal_fg if jumlahawal_fg != 0 else 0

        produk.totalbiayaawal_wip = totalbiayaawal_wip
        produk.totalbiayaawal_fg = totalbiayaawal_fg

        produk.jumlahawal_wip = jumlahawal_wip
        produk.jumlahawal_fg = jumlahawal_fg

        produk.jumlahkumulatif =  produk.jumlahawal_wip +  produk.jumlahawal_fg
        produk.totalbiayakumulatif = produk.totalbiayaawal_wip + produk.totalbiayaawal_fg
        produk.hargasatuankumulatif = (produk.hargasatuanawal_wip+produk.hargasatuanawal_fg)/2

        totalbiayasaldoawal_wip += totalbiayaawal_wip
        totalbiayasaldoawal_fg += totalbiayaawal_fg


    stokawalbahanbaku[0] = {"data": bahanbaku, "total": totalbiayasaldoawal_wip}
    
    for artikel in artikels:
        saldoawalwip = models.SaldoAwalArtikel.objects.filter(
            Tanggal__year=awaltahun.year,
            IDLokasi__NamaLokasi = "WIP",
            IDArtikel=artikel,
        ).first()
        if saldoawalwip == None:
            jumlahwip = 0
        else:
            jumlahwip = saldoawalwip.Jumlah
            
        saldoawalfg = models.SaldoAwalArtikel.objects.filter(
            Tanggal__year=awaltahun.year,
            IDLokasi__NamaLokasi = "FG",
            IDArtikel=artikel,
        ).first()
        if saldoawalfg == None:
            jumlahfg = 0
        else:
            jumlahfg = saldoawalfg.Jumlah
        artikel.jumlahwip = jumlahwip
        artikel.jumlahfg = jumlahfg
        
    datasaldototalproduksi = models.SaldoAwalProduksi.objects.filter(Tanggal__year = awaltahun.year).first()
    if datasaldototalproduksi == None:
        datasaldototalproduksi= {"Saldo":0,"Tanggal":None}
     



    return datasaldototalproduksi, stokawalbahanbaku, artikels



def detaillaporanbaranstokwip(request):
    if len(request.GET) > 0:
        print(request.GET)

        bulan = request.GET["bulan"]
        waktuobj = datetime.strptime(bulan, "%Y-%m")
        print(waktuobj.month)
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = waktuobj.month
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))
        datawip = getstokbahanproduksi(last_days, index, awaltahun)
        return render(request, "ppic/detaillaporanstokwip.html")


def perhitunganpersediaan(
    last_days,
    stopindex,
    awaltahun,
    stockgudang,
    bahanbakumasuk,
    barangkeluar,
    saldofg,

):
    listbulan = [
        "Januari",
        "Februari",
        "Maret",
        "April",
        "Mei",
        "Juni",
        "Juli",
        "Agustus",
        "September",
        "Oktober",
        "Novermber",
        "Desember",
    ]
    datakirim = {}
    datasaldoawalbahanproduksi = {}
    datastokwipawal, datastokwip, datastokfg = getstokbahanproduksi(
        last_days, stopindex, awaltahun
    )
    # print(barangfg)
    # {0: {'data': {<Artikel: 1>: {'jumlah': 2640, 'hargafg': 410000.0, 'biaya': 1082400000.0}, <Artikel: 5171/AC>: {'jumlah': 13571, 'hargafg': 390.31990242, 'biaya': 5297031.39574182}, <Artikel: 5115/AC#1>: {'jumlah': 2130, 'hargafg': 803.5997990999999, 'biaya': 1711667.5720829999}, <Artikel: 9010/ACC>: {'jumlah': 19837, 'hargafg': 1071.739732065, 'biaya': 21260101.064973406}}, 'total': 1110668800.0327983}}

    # print(datastokwipawal)
    # {0: {'data': <QuerySet [<Artikel: 9010/AC>, <Artikel: 5111/EXP>, <Artikel: 1>, <Artikel: 2>, <Artikel: 5111#/AC>, <Artikel: 5171/AC>, <Artikel: 5115/AC#1>, <Artikel: 9010/ACC>, <Artikel: coba1>, <Artikel: coba2>, <Artikel: coba3>, <Artikel: coba4>, <Artikel: coba5>, <Artikel: coba6>]>, 'total': 12951899.692499999}}

    # print(stockgudang)
    # {'saldoawal': {0: 133299998.0, 1: 112960003.06, 2: 103280003.48}, 'saldoakhir': {0: 112960003.06, 1: 103280003.48, 2: 105643301.97218497}}

    # print(bahanbakumasuk)
    # {0: {'data': <QuerySet
    #  []>, 'total': 0}, 1: {'data': <QuerySet []>, 'total': 0}, 2: {'data': <QuerySet [<DetailSuratJalanPembelian: 3/SJP/I-2024 - 2024-03-21 A-101>, <DetailSuratJalanPembelian: 2 - 2024-03-24 A-101>, <DetailSuratJalanPembelian: III/SJP/I-2024 - 2024-03-25 A-001-02>, <DetailSuratJalanPembelian: III/SJP/I-2024 - 2024-03-25 A-101>, <DetailSuratJalanPembelian: IV/SJP/I-2024 - 2024-03-26 A-101>, <DetailSuratJalanPembelian: IV/SJP/I-2024 - 2024-03-26 B-001>, <DetailSuratJalanPembelian: 1 - 2024-03-14 B-001>, <DetailSuratJalanPembelian: 2/SJP/I-2024 - 2024-03-11 coba-001>]>, 'total': 38042000.0}}

    # print(barangkeluar)
    # {0: 255036.39385, 1: 0, 2: 41113622.22968718}
    # print(datastokwipawal)

    # print(sisabahanproduksifg)

    # datasaldoawalbahanproduksi[0] = datastokwipawal[0]["total"]
    datasaldoawalproduksi=models.SaldoAwalProduksi.objects.filter(Tanggal__year = awaltahun.year).first()
    if datasaldoawalproduksi != None:
        datasaldoawalbahanproduksi[0] = datasaldoawalproduksi.Saldo
    else:
        datasaldoawalbahanproduksi[0] = 0
    
    print(datasaldoawalbahanproduksi[0])
    # print(asd)
    # {0: {<Produk: A-101>: {'stok': 0, 'Hargasatuan': 80000.0, 'Hargatotal': 0}, <Produk: B-001>: {'stok': 0, 'Hargasatuan': 0, 'Hargatotal': 0}, <Produk: C-001>: {'stok': 499.6472403, 'Hargasatuan': 10000.0, 'Hargatotal': 0}, <Produk: D-001>: {'stok': 1000, 'Hargasatuan': 80000.0, 'Hargatotal': 0}, <Produk: A-001-02>: {'stok': 0, 'Hargasatuan': 0, 'Hargatotal': 0}, <Produk: A-001-03>: {'stok': 0, 'Hargasatuan': 50000.0, 'Hargatotal': 0}, <Produk: A-001-06>: {'stok': 0, 'Hargasatuan': 77727.25727272723, 'Hargatotal': 0}, <Produk: coba-001>: {'stok': 0, 'Hargasatuan': 12500.0, 'Hargatotal': 0}, <Produk: tes>: {'stok': 0, 'Hargasatuan': 0, 'Hargatotal': 0}, <Produk: tesksbb>: {'stok': 0, 'Hargasatuan': 10000.0, 'Hargatotal': 0}}, 1: {<Produk: A-101>: {'stok': 0, 'Hargasatuan': 80000.0, 'Hargatotal': 0}, <Produk: B-001>: {'stok': 0, 'Hargasatuan': 0, 'Hargatotal': 0}, <Produk: C-001>: {'stok': 499.6472403, 'Hargasatuan': 10000.0, 'Hargatotal': 0}, <Produk: D-001>: {'stok': 1000, 'Hargasatuan': 80000.0, 'Hargatotal': 0}, <Produk: A-001-02>: {'stok': 0, 'Hargasatuan': 0, 'Hargatotal': 0}, <Produk: A-001-03>: {'stok': 0, 'Hargasatuan': 50000.0, 'Hargatotal': 0}, <Produk: A-001-06>: {'stok': 0, 'Hargasatuan': 77727.25727272723, 'Hargatotal': 0}, <Produk: coba-001>: {'stok': 0, 'Hargasatuan': 12500.0, 'Hargatotal': 0}, <Produk: tes>: {'stok': 0, 'Hargasatuan': 0, 'Hargatotal': 0}, <Produk: tesksbb>: {'stok': 0, 'Hargasatuan': 10000.0, 'Hargatotal': 0}}, 2: {<Produk: A-101>: {'stok': 0, 'Hargasatuan': 79987.61015368767, 'Hargatotal': 0.0}, <Produk: B-001>: {'stok': 0, 'Hargasatuan': 125000.0, 'Hargatotal': 0.0}, <Produk: C-001>: {'stok': 553.3972403, 'Hargasatuan': 10000.0, 'Hargatotal': 5533972.403}, <Produk: D-001>: {'stok': 1000, 'Hargasatuan': 80000.0, 'Hargatotal': 80000000.0}, <Produk: A-001-02>: {'stok': 0, 'Hargasatuan': 96000.0, 'Hargatotal': 0.0}, <Produk: A-001-03>: {'stok': 0, 'Hargasatuan': 50000.0, 'Hargatotal': 0.0}, <Produk: A-001-06>: {'stok': 0, 'Hargasatuan': 77727.25727272723, 'Hargatotal': 0.0}, <Produk: coba-001>: {'stok': 0, 'Hargasatuan': 64423.07692307692, 'Hargatotal': 0.0}, <Produk: tes>: {'stok': 0, 'Hargasatuan': 0, 'Hargatotal': 0}, <Produk: tesksbb>: {'stok': 0, 'Hargasatuan': 10000.0, 'Hargatotal': 0.0}, 'total': 85533972.403}}

    # print(asd)
    for index in range(0, stopindex):
        # print(barangkeluar[index])
        # print(index)
        if index > 0:
            datasaldoawalbahanproduksi[index] = saldowip + jumlahstokfg
        try:
            jumlahbarangkeluar = barangkeluar[index]
        except KeyError:
            jumlahbarangkeluar = 0
        try:
            jumlahbarangmasuk = bahanbakumasuk[index]["total"]
        except KeyError:
            jumlahbarangmasuk = 0
        try:
            jumlahsaldoawalgudang = stockgudang["saldoawal"][index]
        except KeyError:
            jumlahsaldoawalgudang = 0
        try:
            jumlahsaldoakhirgudang = stockgudang["saldoakhir"][index]
        except KeyError:
            jumlahsaldoakhirgudang = 0
        try:
            jumlahstokfg = saldofg[index]['totalsaldo']
        except KeyError:
            jumlahstokfg = 0
        try:
            jumlahsaldoawalproduksi = datasaldoawalbahanproduksi[index]
        except KeyError:
            jumlahsaldoawalproduksi = 0
        total = (
            jumlahsaldoawalgudang
            + jumlahsaldoawalproduksi
            + jumlahbarangmasuk
            - jumlahbarangkeluar
        )

        # print("ini total : ", total)
        saldowip = total - jumlahstokfg - jumlahsaldoakhirgudang

        """BELUM MASUKIN DATA TOTAL"""
        # print(asda)
        dummy = {
            "barangkeluar": jumlahbarangkeluar,
            "barangmasuk": jumlahbarangmasuk,
            "saldoawalgudang": jumlahsaldoawalgudang,
            "saldoakhirgudang": jumlahsaldoakhirgudang,
            "stokfg": jumlahstokfg,
            "stokawalproduksi": datasaldoawalbahanproduksi[index],
            "totalsaldo": total,
            "saldowip": saldowip,
        }

        datakirim[listbulan[index]] = dummy
    # print(datakirim)
    return datakirim


def laporanpersediaan(request):
    if len(request.GET) == 0:
        return render(request, "ppic/views_laporanpersediaanperusahaan.html")
    else:
        starttime = time.time()
        bulan = request.GET["bulan"]
        waktuobj = datetime.strptime(bulan, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = int(waktuobj.month)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))

        """SECTION BARANG KELUAR"""
        awalbarangkeluar = time.time()
        (totalbiayakeluar) = (getbarangkeluar(last_days, index, awaltahun))
        akhirbarankeluar = time.time()
        """SECTION BARANG MASUK"""
        awalbarangmasuk = time.time()
        barangmasuk = getbarangmasuk(last_days, index, awaltahun)
        akhirbarangmasuk = time.time()
        """SECTION STOCK GUDANG"""
        awalsaldogudang = time.time()
        baranggudang = saldogudang(last_days, index, awaltahun)
        akhirstokgudang = time.time()
        """SECTION FG"""
        # barangfg, bahanbakusisafg = getstokartikelfg(last_days, index, awaltahun)
        # hargapurchasing = gethargapurchasingperbulan(last_days,index,awaltahun)
        awalsaldofg = time.time()
        saldofg = getstokfg(None,last_days,index,awaltahun)
        akhirsaldofg = time.time()
        # print(saldofg[0['totalsaldo']])
        datawip, datastokwiponly, datastokfgonly = getstokbahanproduksi(
    last_days, index, awaltahun
)
        """SECTION WIP (Skip dulu)"""
        awalpersediaan = time.time()
        dataperhitunganpersediaan = perhitunganpersediaan(
            last_days,
            index,
            awaltahun,
            baranggudang,
            barangmasuk,
            totalbiayakeluar[1],
            saldofg
        )
        akhirpersediaan = time.time()

        # print(asd)

        print("\n")
        print(dataperhitunganpersediaan)
        print("tes")
        endtime = time.time()
        print("Selisih waktu : ", endtime - starttime)
        print('Waktu Barang keluar : ',akhirbarankeluar - awalbarangkeluar )
        print('Waktu  Barang masuk : ',akhirbarangmasuk- awalbarangmasuk)
        print('Waktu saldo gudang : ', akhirstokgudang - awalsaldogudang)
        print('Waktu Saldo FG : ', akhirsaldofg - awalsaldofg)
        print('Waktu perhitungan persediaan : ',akhirpersediaan-awalpersediaan)
    return render(
        request,
        "ppic/views_laporanpersediaanperusahaan.html",
        {
            "modeldata": dataperhitunganpersediaan,
            "waktu": bulan,
            "tahun": waktuobj.year,
        },
    )


def detaillaporanbaranstokawalproduksi(request):
    if len(request.GET) > 0:
        bulan = request.GET["bulan"]
        waktu = request.GET["waktu"]
        waktuobj = datetime.strptime(waktu, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = listbulan.index(request.GET["bulan"]) + 1
        # print(index)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))

        datawip, datastokwip, datastokfg = getstokbahanproduksi(
            last_days, index, awaltahun
        )
        print(datawip,datastokwip,datastokfg)
        return render(
            request,
            "ppic/detailstockawalproduksi.html",
            {
                "stokawal": datawip,
                'stokawalwip':datastokwip[0],
                'stokartikel': datastokfg           },
        )


def detaillaporanbaranstokawalgudang(request):
    if len(request.GET) > 0:
        bulan = request.GET["bulan"]
        waktu = request.GET["waktu"]
        waktuobj = datetime.strptime(waktu, "%Y-%m")
        awaltahun = datetime(waktuobj.year, 1, 1)
        listbulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "Novermber",
            "Desember",
        ]
        index = listbulan.index(request.GET["bulan"]) + 1
        # print(index)
        last_days = []
        for month in range(1, 13):
            last_day = calendar.monthrange(waktuobj.year, month)[1]
            last_days.append(date(waktuobj.year, month, last_day))
        # print(last_days)
        # print(last_days[:index+1])
        datasaldoawalgudang = saldogudang(last_days, index, awaltahun)
        print(datasaldoawalgudang)
        return render(
            request,
            "ppic/detailstockawalgudang.html",
            {
                "stokawal": datasaldoawalgudang["datasaldoawal"],
                "saldoawal": datasaldoawalgudang["saldoawal"][0],
            },
        )


def read_transactionlog(request):
    waktuobj = datetime.now().date()
    
    dataobj = models.transactionlog.objects.all().order_by('-waktu')
    for i in dataobj:
        i.waktu = i.waktu.strftime("%Y-%m-%d %H:%M:%S")
    # dataobj = LogEntry.objects.filter(action_time__month = waktuobj.month)
    return render(request, "ppic/transactionlog.html", {"data": dataobj})





def exportlaporanbulananexcel(request):
    bulan = request.GET["bulan"]
    waktuobj = datetime.strptime(bulan, "%Y-%m")
    awaltahun = datetime(waktuobj.year, 1, 1)
    listbulan = [
        "Januari",
        "Februari",
        "Maret",
        "April",
        "Mei",
        "Juni",
        "Juli",
        "Agustus",
        "September",
        "Oktober",
        "Novermber",
        "Desember",
    ]
    index = int(waktuobj.month)
    last_days = []
    for month in range(1, 13):
        last_day = calendar.monthrange(waktuobj.year, month)[1]
        last_days.append(date(waktuobj.year, month, last_day))

    waktuawalproses = time.time()
    hargapurchasing = gethargapurchasingperbulan (last_days,index,awaltahun)
    waktustart = time.time()
    (
        datadetailsppb,
        totalbiayakeluar,
        datapenyusun,
        datalistbarangkelua,
        datatransaksilainlain,
        detailtransaksigold,
        detailbiaya,
        datatransaksikeluar
    ) = getbarangkeluar(last_days, index, awaltahun,hargapurchasing)
    waktugetbarangkeluar = time.time()
    datatransaksikeluar = datatransaksikeluar[index-1]
    rekapartikel,rekaptransaksidisplay,rekapdatabahankeluar,rekapgolongand,rekaptransaksilainlain,rekappemusnahanartikel,rekappemusnahanbahanbaku = create_rekaptransaksikeluar(datatransaksikeluar)
    """SECTION BARANG MASUK"""
    waktustartdatatransaksimasuk = time.time()
    barangmasuk = getbarangmasuk(last_days, index, awaltahun)
    waktubarangmasuk = time.time()
    """SECTION STOCK GUDANG"""
    waktustartbaranggudang = time.time()
    baranggudang = saldogudang(last_days, index, awaltahun,hargapurchasing)
    waktubaranggudang = time.time()
    """SECTION FG"""
    waktustartstokfg = time.time()
    stokfg = getstokfg(None,last_days,index,awaltahun,hargapurchasing)
    waktustokfg = time.time()
    """SECTION STOK AWAL PRODUKSI"""
    waktustartstokawalproduksi = time.time()
    totalbiayasaldoawalproduksi, saldoawalbahanbaku, saldoawalartikel = getstokbahanproduksi(
        last_days, index, awaltahun
    )
    waktugetstokbahanproduksi = time.time()
    """SECTION WIP (Skip dulu)"""
    waktupersediaan = time.time()
    dataperhitunganpersediaan = perhitunganpersediaan(
        last_days,
        index,
        awaltahun,
        baranggudang,
        barangmasuk,
        totalbiayakeluar,
        stokfg,

    )
    waktuperhitunganpersediaan = time.time()

    print('Waktu barang keluar : ',waktugetbarangkeluar-waktustart)
    print('Waktu barang masuk : ',waktubarangmasuk-waktustartdatatransaksimasuk)
    print('Watu Stock Gudang : ',waktubaranggudang - waktustartbaranggudang )
    print('Waktu Stock FG : ', waktustokfg-waktustartstokfg)
    print('Waktu Stok Bahan Produksi : ',   waktugetstokbahanproduksi-waktustartstokawalproduksi)
    print('Waktu Persediaan : ', waktuperhitunganpersediaan - waktuperhitunganpersediaan)
    # print(asd)


    """PEMBUATAN EXCEL"""
    """ 1. sheet untuk laporan persediaan"""
    datapersediaan = dataperhitunganpersediaan[listbulan[waktuobj.month - 1]]
    datamodelpersediaan ={
        "Barang Keluar" : [datapersediaan['barangkeluar']],
        "Barang Masuk" : [datapersediaan['barangmasuk']],
        "Saldo Awal Stock Gudang" : [datapersediaan['saldoawalgudang']],
        "Saldo Awal Bahan Produksi" : [datapersediaan['stokawalproduksi']],
        "Total Saldo":[datapersediaan['totalsaldo']],
        "Saldo Akhir Gudang" : [datapersediaan['saldoakhirgudang']],
        "Saldo Akhir FG" : [datapersediaan['stokfg']],
        "Saldo Akhir WIP" : [datapersediaan['saldowip']],
    }
    dfpersediaan = pd.DataFrame(datamodelpersediaan)

    # print(dfpersediaan)
    # print(ad)
    """2. Sheet untuk laporan barang masuk"""
    datamodelmasuk = {
        "Tanggal Masuk": [],
        "No Surat Jalan": [],
        "Supplier": [],
        "Kode Produk": [],
        "Satuan": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Harga Total": [],
    }
    for masuk in barangmasuk[index - 1]["data"]:
        # print(masuk)
        datamodelmasuk["Tanggal Masuk"].append(masuk.NoSuratJalan.Tanggal)
        datamodelmasuk["No Surat Jalan"].append(masuk.NoSuratJalan.NoSuratJalan)
        datamodelmasuk["Supplier"].append(masuk.NoSuratJalan.supplier)
        datamodelmasuk["Kode Produk"].append(masuk.KodeProduk.KodeProduk)
        datamodelmasuk["Satuan"].append(masuk.KodeProduk.unit)
        datamodelmasuk["Harga Satuan"].append(masuk.Harga)
        datamodelmasuk["Harga Total"].append(masuk.Harga * masuk.Jumlah)
        datamodelmasuk["Jumlah"].append(masuk.Jumlah)

    # print(datamodelmasuk)
    df2 = pd.DataFrame(datamodelmasuk)
    totalbiayamasuk = sum(datamodelmasuk["Harga Total"])
    # print(df2,totalbiayamasuk)
    # print(asd)
    """3. Sheet untuk Barang Keluar"""
    # print(datadetailsppb)
    datamodelkeluar = {
        "Tanggal Keluar": [],
        "No SPPB": [],
        "NO SPK": [],
        "Artikel": [],
        "Jumlah": [],
        "Harga FG": [],
        "Total Biaya": [],
    }
    print(datatransaksikeluar['SPPBArtikel'])
    for data in datatransaksikeluar['SPPBArtikel']['SPPBArtikel']:
        # print(data)
        datamodelkeluar["Jumlah"].append(data.Jumlah)
        datamodelkeluar["Artikel"].append(data.DetailSPK.KodeArtikel)
        datamodelkeluar["Harga FG"].append(data.hargafg)
        datamodelkeluar["NO SPK"].append(data.DetailSPK.NoSPK)
        datamodelkeluar["No SPPB"].append(data.NoSPPB)
        datamodelkeluar["Total Biaya"].append(data.totalharga)
        datamodelkeluar["Tanggal Keluar"].append(data.NoSPPB.Tanggal)

    # print(datamodelkeluar)
    totalbiayakeluar = sum(datamodelkeluar["Total Biaya"])
    dfdatakeluar = pd.DataFrame(datamodelkeluar)

    datamodelkeluar = {
        "Tanggal": [],
        "Artikel": [],
        'Lokasi' : [],
        "Jumlah": [],
        "Harga FG": [],
        "Total Biaya": [],
    }
    print(datatransaksikeluar['SPPBArtikel'])
    for data in datatransaksikeluar['Pemusnahanartikel']['transaksipemusnahan']:
        # print(data)

        datamodelkeluar["Tanggal"].append(data.Tanggal)
        datamodelkeluar["Jumlah"].append(data.Jumlah)
        datamodelkeluar['Lokasi'].append(data.lokasi)
        datamodelkeluar["Artikel"].append(data.KodeArtikel)
        datamodelkeluar["Harga FG"].append(data.hargafg)
        datamodelkeluar["Total Biaya"].append(data.totalharga)

    # print(datamodelkeluar)
    totalbiayapemusnahanartikel = sum(datamodelkeluar["Total Biaya"])
    dfdatapemusnahanartikel = pd.DataFrame(datamodelkeluar)
    # print(dfdatapemusnahanartikel)
    # print(asd)
    
    datamodelstransaksigold = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }
    # print(baranggudang["datasaldoawal"])
    for item in datatransaksikeluar['Transaksigolongand']['datatransaksi']:
        datamodelstransaksigold["Kode Produk"].append(item.KodeProduk.KodeProduk)
        datamodelstransaksigold["Unit"].append(item.KodeProduk.unit)
        datamodelstransaksigold["Nama Produk"].append(item.KodeProduk.NamaProduk)
        datamodelstransaksigold["Jumlah"].append(item.jumlah)
        datamodelstransaksigold["Total Biaya"].append(item.hargatotal)
        datamodelstransaksigold["Harga Satuan"].append(item.harga)

    # print(datamodelstransaksigold)
    dftransaksigold = pd.DataFrame(datamodelstransaksigold)
    totalbiayakeluargold = sum(datamodelstransaksigold["Total Biaya"])
    # print(dftransaksigold)

    # print(datatransaksilainlain)
    datamodelstransaksilainlain = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }

    for item in datatransaksikeluar['Transaksilainlain']['datatransaksi']:
        datamodelstransaksilainlain["Kode Produk"].append(item.KodeProduk.KodeProduk)
        datamodelstransaksilainlain["Unit"].append(item.KodeProduk.unit)
        datamodelstransaksilainlain["Nama Produk"].append(item.KodeProduk.NamaProduk)
        datamodelstransaksilainlain["Jumlah"].append(item.jumlah)
        datamodelstransaksilainlain["Total Biaya"].append(item.hargatotal)
        datamodelstransaksilainlain["Harga Satuan"].append(item.harga)
    
    dftransaksilainlain = pd.DataFrame(datamodelstransaksilainlain)
    # print(dftransaksilainlain)
    totalbiayakeluartransaksilainlain = sum(datamodelstransaksilainlain["Total Biaya"])
    # Transaksi Bahan Baku
    datamodelstransaksibahanbaku = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }

    for item in datatransaksikeluar['Transaksibahanbaku']['datatransaksi']:
        datamodelstransaksibahanbaku["Kode Produk"].append(item.DetailBahan.KodeProduk)
        datamodelstransaksibahanbaku["Unit"].append(item.DetailBahan.unit)
        datamodelstransaksibahanbaku["Nama Produk"].append(item.DetailBahan.NamaProduk)
        datamodelstransaksibahanbaku["Jumlah"].append(item.Jumlah)
        datamodelstransaksibahanbaku["Total Biaya"].append(item.hargatotal)
        datamodelstransaksibahanbaku["Harga Satuan"].append(item.harga)
    
    dftransaksibahanbaku = pd.DataFrame(datamodelstransaksibahanbaku)
    # print(dftransaksibahanbaku)
    totalbiayakeluartransaksibahanbaku = sum(datamodelstransaksibahanbaku["Total Biaya"])
    datamodelspemusnahanbahanbaku = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }

    for item in datatransaksikeluar['Pemusnahanbahanbaku']['datatransaksi']:
        datamodelspemusnahanbahanbaku["Kode Produk"].append(item.KodeBahanBaku.KodeProduk)
        datamodelspemusnahanbahanbaku["Unit"].append(item.KodeBahanBaku.unit)
        datamodelspemusnahanbahanbaku["Nama Produk"].append(item.KodeBahanBaku.NamaProduk)
        datamodelspemusnahanbahanbaku["Jumlah"].append(item.Jumlah)
        datamodelspemusnahanbahanbaku["Total Biaya"].append(item.hargatotal)
        datamodelspemusnahanbahanbaku["Harga Satuan"].append(item.harga)
    
    dfpemusnahanbahanbaku = pd.DataFrame(datamodelspemusnahanbahanbaku)
    print(dfpemusnahanbahanbaku)
    # print(asd)
    totalbiayapemusnahanbahanbaku = sum(datamodelspemusnahanbahanbaku["Total Biaya"])
    # Transaksi Display
    datamodeldisplay = {
        "Tanggal Keluar": [],
        "No SPPB": [],
        "NO SPK": [],
        "Artikel": [],
        "Jumlah": [],
        "Harga FG": [],
        "Total Biaya": [],
    }

    for data in datatransaksikeluar['SPPBDisplay']['SPPBDisplay']:
        # print(data)
        datamodeldisplay["Jumlah"].append(data.Jumlah)
        datamodeldisplay["Artikel"].append(data.DetailSPKDisplay.KodeDisplay)
        datamodeldisplay["Harga FG"].append(data)
        datamodeldisplay["NO SPK"].append(data.DetailSPKDisplay.NoSPK)
        datamodeldisplay["No SPPB"].append(data.NoSPPB)
        datamodeldisplay["Total Biaya"].append(data.totalbiaya)
        datamodeldisplay["Tanggal Keluar"].append(data.NoSPPB.Tanggal)

    # print(datamodeldisplay)
    dfdatakeluardisplay = pd.DataFrame(datamodeldisplay)
    totalbiayakeluardisplay = sum(datamodeldisplay["Total Biaya"])

    # print(asd)
    # Rekap Transaksi Artikel
    print(rekapartikel)
    datamodelsrekapartikelkeluar = {
        'Kode Artikel' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
    for artikel,data in rekapartikel.items():
        datamodelsrekapartikelkeluar['Kode Artikel'].append( artikel.KodeArtikel)
        datamodelsrekapartikelkeluar["Harga FG"].append( data['HargaFG'])
        datamodelsrekapartikelkeluar['Harga Total'].append( data['hargatotal'])
        datamodelsrekapartikelkeluar["Jumlah"].append( data['Jumlah'])
    
    dfrekapartikelkeluar = pd.DataFrame(datamodelsrekapartikelkeluar)
    datamodelsrekappemusnahanartikel = {
        'Kode Artikel' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
    for artikel,data in rekappemusnahanartikel.items():
        datamodelsrekappemusnahanartikel['Kode Artikel'].append( artikel.KodeArtikel)
        datamodelsrekappemusnahanartikel["Harga FG"].append( data['HargaFG'])
        datamodelsrekappemusnahanartikel['Harga Total'].append( data['hargatotal'])
        datamodelsrekappemusnahanartikel["Jumlah"].append( data['Jumlah'])
    
    dfrekappemusnahanartikel = pd.DataFrame(datamodelsrekappemusnahanartikel)
    print(dfrekappemusnahanartikel)
    # print(rekaptransaksidisplay)
    
    # Rekap transaksi display
    datamodelsdisplaykeluar = {
        'Kode Display' : [],
        'No SPK' : [],
        'Jumlah' : [],
        'Harga Satuan' : [],
        'Harga Total': []
    }
    for display, data_nospk in rekaptransaksidisplay.items():
        for spk, details in data_nospk['NoSPK'].items():
            datamodelsdisplaykeluar['Kode Display'].append(display)
            datamodelsdisplaykeluar['No SPK'].append(spk)
            datamodelsdisplaykeluar['Jumlah'].append(details['Jumlah'])
            datamodelsdisplaykeluar['Harga Satuan'].append(details['hargasatuan'])
            datamodelsdisplaykeluar['Harga Total'].append(details['Hargatotal'])
    dfrekapdisplaykeluar = pd.DataFrame(datamodelsdisplaykeluar)
    print(dfrekapdisplaykeluar)
    print(rekapdatabahankeluar)

    # Rekap Transaksi bahan baku keluar
    datamodelstransaksibahanbaku = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
    for artikel,data in rekapdatabahankeluar.items():
        datamodelstransaksibahanbaku['Kode Bahan Baku'].append( artikel.KodeProduk)
        datamodelstransaksibahanbaku['Nama Bahan Baku'].append( artikel.NamaProduk)
        datamodelstransaksibahanbaku['Satuan'].append( artikel.unit)
        datamodelstransaksibahanbaku["Harga FG"].append( data['HargaFG'])
        datamodelstransaksibahanbaku['Harga Total'].append( data['hargatotal'])
        datamodelstransaksibahanbaku["Jumlah"].append( data['Jumlah'])
    
    dfrekapbahanbakukeluar = pd.DataFrame(datamodelstransaksibahanbaku)
    print(dfrekapbahanbakukeluar)
    print(rekapgolongand)

    datamodelspemusnahanbahanbaku = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
    for artikel,data in rekappemusnahanbahanbaku.items():
        datamodelspemusnahanbahanbaku['Kode Bahan Baku'].append( artikel.KodeProduk)
        datamodelspemusnahanbahanbaku['Nama Bahan Baku'].append( artikel.NamaProduk)
        datamodelspemusnahanbahanbaku['Satuan'].append( artikel.unit)
        datamodelspemusnahanbahanbaku["Harga FG"].append( data['HargaFG'])
        datamodelspemusnahanbahanbaku['Harga Total'].append( data['hargatotal'])
        datamodelspemusnahanbahanbaku["Jumlah"].append( data['Jumlah'])
    
    dfrekappemusnahanbahanbaku = pd.DataFrame(datamodelspemusnahanbahanbaku)
    print(dfrekappemusnahanbahanbaku)
    print(rekapgolongand)

    datamodelstransaksigolongand = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
    for artikel,data in rekapgolongand.items():
        datamodelstransaksigolongand['Kode Bahan Baku'].append( artikel.KodeProduk)
        datamodelstransaksigolongand['Nama Bahan Baku'].append( artikel.NamaProduk)
        datamodelstransaksigolongand['Satuan'].append( artikel.unit)
        datamodelstransaksigolongand["Harga FG"].append( data['Harga'])
        datamodelstransaksigolongand['Harga Total'].append( data['hargatotal'])
        datamodelstransaksigolongand["Jumlah"].append( data['Jumlah'])
    
    dfrekapgolongand = pd.DataFrame(datamodelstransaksigolongand)
    print(dfrekapgolongand)
    print(rekaptransaksilainlain)

    datamodelstransaksilainlain = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
    for artikel,data in rekaptransaksilainlain.items():
        datamodelstransaksilainlain['Kode Bahan Baku'].append( artikel.KodeProduk)
        datamodelstransaksilainlain['Nama Bahan Baku'].append( artikel.NamaProduk)
        datamodelstransaksilainlain['Satuan'].append( artikel.unit)
        datamodelstransaksilainlain["Harga FG"].append( data['Harga'])
        datamodelstransaksilainlain['Harga Total'].append( data['hargatotal'])
        datamodelstransaksilainlain["Jumlah"].append( data['Jumlah'])
    
    dfrekaptransaksilainlain = pd.DataFrame(datamodelstransaksilainlain)
    print(dfrekaptransaksilainlain)

    # print(asd)
    """4. Sheet Stock Awal Gudang"""
    datamodelstockawalgudang = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }
    # print(baranggudang["datasaldoawal"])
    for item in baranggudang["datasaldoawal"]:
        datamodelstockawalgudang["Kode Produk"].append(item.KodeProduk)
        datamodelstockawalgudang["Unit"].append(item.unit)
        datamodelstockawalgudang["Nama Produk"].append(item.NamaProduk)
        datamodelstockawalgudang["Jumlah"].append(item.jumlahawal)
        datamodelstockawalgudang["Total Biaya"].append(item.totalbiayaawal)
        datamodelstockawalgudang["Harga Satuan"].append(item.hargasatuanawal)

    # print(datamodelstockawalgudang)
    dfstokgudang = pd.DataFrame(datamodelstockawalgudang)
    totalbiayaawalgudang = sum(datamodelstockawalgudang['Total Biaya'])

    """5. Sheet Stock Awal Produksi"""
    # data wip
    datamodelstockawalwip = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        'Jumlah WIP' : [],
        'Jumlah FG' : [],
        'Jumlah Total' : []
        
    }

    # print(asd)
    for stokawal in saldoawalbahanbaku[0]["data"]:
        # print(stokawal)
        datamodelstockawalwip["Harga Satuan"].append(stokawal.hargasatuanawal_wip)
        datamodelstockawalwip["Jumlah WIP"].append(stokawal.jumlahawal_wip)
        datamodelstockawalwip["Jumlah FG"].append(stokawal.jumlahawal_fg)
        datamodelstockawalwip["Jumlah Total"].append(stokawal.jumlahkumulatif)
        datamodelstockawalwip["Kode Produk"].append(stokawal.KodeProduk)
        datamodelstockawalwip["Nama Produk"].append(stokawal.NamaProduk)
        datamodelstockawalwip["Unit"].append(stokawal.unit)
    # print(datawip)
    dfstokawalwip = pd.DataFrame(datamodelstockawalwip)
    datamodelstockawalwip = {
        "Kode Artikel": [],
       "Jumlah WIP" : [],
       'Jumlah FG': []
        
    }
    # print(datastokfgonly)
    # print(datastokfgonly[0]["data"])
    for stokawal in saldoawalartikel:
        # print(stokawal)
        datamodelstockawalwip["Kode Artikel"].append(stokawal.KodeArtikel)
        datamodelstockawalwip["Jumlah WIP"].append(stokawal.jumlahwip)
        datamodelstockawalwip["Jumlah FG"].append(stokawal.jumlahfg)
    # print(datawip)
    dfstokawalfg = pd.DataFrame(datamodelstockawalwip)

    # print(asdas)
    '''STOK Gudang Sekarang'''
    # print(baranggudang)
    # print(asd)
    datamodelstockawalgudang = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }
    for bahan,item in baranggudang["hargaakhirbulanperproduk"]['data'].items():
        # print(item)
        datamodelstockawalgudang["Harga Satuan"].append(item['hargasatuan'])
        datamodelstockawalgudang["Jumlah"].append(item['jumlah'])
        datamodelstockawalgudang["Kode Produk"].append(bahan.KodeProduk)
        datamodelstockawalgudang["Nama Produk"].append(bahan.NamaProduk)
        datamodelstockawalgudang["Total Biaya"].append(item['hargatotal'])
        datamodelstockawalgudang["Unit"].append(bahan.unit)

    dfstokgudangakhir = pd.DataFrame(datamodelstockawalgudang)
    totalsaldogudang = baranggudang['hargaakhirbulanperproduk']['total']

    '''STOK FG SEKARANG'''
    # Stok Artikel
    datamodelkeluar = {
        "Kode Artikel": [],
        "Jumlah": [],
        "Harga FG": [],
        "Total Biaya": [],
    }
    # print(stokfg[-1]['Artikel'])
    # print(asd)
    for data in stokfg[-1]['Artikel']:
        # print(data)
        datamodelkeluar["Jumlah"].append(data['total'])
        datamodelkeluar["Kode Artikel"].append(data['KodeArtikel'])
        datamodelkeluar["Harga FG"].append(data['hargafg'])
        datamodelkeluar["Total Biaya"].append(data['totalsaldo'])

    # print(datamodelkeluar)
    dfstokfgartikel = pd.DataFrame(datamodelkeluar)
    totalsaldofgartikel = sum(datamodelkeluar["Total Biaya"])
    
    datamodelkeluar = {
        "Kode Display": [],
        "Jumlah": [],
        "Harga FG": [],
        "Total Biaya": [],
    }
    # print(stokfg[-1])

    # print(asd)
    for data in stokfg[-1]['Display']:
        # print(data)
        datamodelkeluar["Jumlah"].append(data['total'])
        datamodelkeluar["Kode Display"].append(data['KodeDisplay'])
        datamodelkeluar["Harga FG"].append(data['hargafg'])
        datamodelkeluar["Total Biaya"].append(data['totalsaldo'])

    # print(datamodelkeluar)
    dfstokfgdisplay = pd.DataFrame(datamodelkeluar)
    totalsaldofgdisplay = sum(datamodelkeluar["Total Biaya"])

    # print(dfstokgudang)

    datamodelkeluar = {
        "Kode Bahan Baku": [],
        "Jumlah": [],
        "Harga Satuan": [],
        "Total Biaya": [],
    }
    # print(stokfg[-1]['BahanBaku'])
    # print(asd)
    for data in stokfg[-1]['BahanBaku']:
        # print(data)
        datamodelkeluar["Jumlah"].append(data['total'])
        datamodelkeluar["Kode Bahan Baku"].append(data['KodeProduk'])
        datamodelkeluar["Harga Satuan"].append(data['hargasatuan'])
        datamodelkeluar["Total Biaya"].append(data['hargatotal'])

    # print(datamodelkeluar)
    dfstokfgbahanbaku = pd.DataFrame(datamodelkeluar)
    totalsaldofgbahanbaku = sum(datamodelkeluar["Total Biaya"])

    # print(dfstokgudang)

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Laporan Persediaan Section
        # df.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        dfpersediaan.to_excel(writer, index=False, startrow=1, sheet_name="Laporan Persediaan")
        writer.sheets["Laporan Persediaan"].cell(row=1, column = 1,value =listbulan[waktuobj.month - 1])
        maxrow = len(dfpersediaan)+1
        maxcol = len(dfpersediaan.columns)
        apply_number_format(writer.sheets['Laporan Persediaan'],2,maxrow+1,1,maxcol)
        apply_borders_thin(writer.sheets['Laporan Persediaan'],2,maxrow+1,maxcol)
        adjust_column_width(writer.sheets['Laporan Persediaan'],dfpersediaan,1,1)
        # Laporan Barang masuk Section
        '''LAPORAN BARANG MASUK'''
        if not df2.empty : 
            df2.to_excel(writer, index=False, startrow=1, sheet_name="Barang Masuk")
            maxrow = len(df2)+1
            maxcol = len(df2.columns)
            writer.sheets["Barang Masuk"].cell(row=1, column = 1,value =listbulan[waktuobj.month - 1])
            writer.sheets["Barang Masuk"].merge_cells(start_row=maxrow+2, start_column=1,end_row = maxrow+2,end_column= maxcol-1)
            writer.sheets["Barang Masuk"].cell(row=maxrow+2, column = 1).value = "Total Harga"
            writer.sheets["Barang Masuk"].cell(row=maxrow+2, column = maxcol,value = totalbiayamasuk)
            apply_number_format(writer.sheets['Barang Masuk'],2,maxrow+2,1,maxcol)
            apply_borders_thin(writer.sheets['Barang Masuk'],2,maxrow+2,maxcol)
            adjust_column_width(writer.sheets['Barang Masuk'],df2,1,1)
        '''LAPORAN REKAP BARANG KELUAR'''
        if not dfrekapartikelkeluar.empty or not dftransaksibahanbaku.empty or not dfrekapdisplaykeluar.empty or not dfrekaptransaksilainlain.empty or not dfrekapgolongand.empty or not dfrekappemusnahanartikel or not dfrekappemusnahanbahanbaku:
            # Artikel Keluar
            maxcol = 0
            maxrow = 0
            maxcolprevdf = 0

            if not dfrekapartikelkeluar.empty:
                dfrekapartikelkeluar.to_excel(writer, index=False, startrow=1, sheet_name="Rekap Barang Keluar")
                maxrow = len(dfrekapartikelkeluar) + 1
                maxcol = len(dfrekapartikelkeluar.columns)
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=1, value=listbulan[waktuobj.month - 1])
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=2, value='Artikel Keluar')
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=1, end_row=maxrow+2, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=1).value = "Total Harga"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcol, value=totalbiayakeluar)
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1, end_row=maxrow+3, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column=1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column=maxcol, value=totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain + totalbiayapemusnahanbahanbaku + totalbiayapemusnahanartikel)
                apply_number_format(writer.sheets['Rekap Barang Keluar'], 2, maxrow+3, 1, maxcol)
                apply_borders_thin(writer.sheets['Rekap Barang Keluar'], 2, maxrow+3, maxcol)
                adjust_column_width(writer.sheets['Rekap Barang Keluar'], dfrekapartikelkeluar, 1, 1)

            if not dfrekapdisplaykeluar.empty:
                dfrekapdisplaykeluar.to_excel(writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Rekap Barang Keluar")
                maxcolprevdf = maxcol
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=maxcol+2, value='Display Keluar')
                maxrow = len(dfrekapdisplaykeluar) + 1
                maxcol = len(dfrekapdisplaykeluar.columns) + maxcol + 1
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2, end_row=maxrow+2, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcol, value=totalbiayakeluardisplay)
                apply_number_format(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcolprevdf+2, maxcol)
                apply_borders_thin(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcol, maxcolprevdf+2)
                adjust_column_width(writer.sheets['Rekap Barang Keluar'], dfrekapdisplaykeluar, 1, maxcolprevdf+2)

            if not dfrekapbahanbakukeluar.empty:
                dfrekapbahanbakukeluar.to_excel(writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Rekap Barang Keluar")
                maxcolprevdf = maxcol
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=maxcol+2, value='Bahan Baku Keluar')
                maxrow = len(dfrekapbahanbakukeluar) + 1
                maxcol = len(dfrekapbahanbakukeluar.columns) + maxcol + 1
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2, end_row=maxrow+2, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcol, value=totalbiayakeluartransaksibahanbaku)
                apply_number_format(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcolprevdf+2, maxcol)
                apply_borders_thin(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcol, maxcolprevdf+2)
                adjust_column_width(writer.sheets['Rekap Barang Keluar'], dfrekapbahanbakukeluar, 1, maxcolprevdf+2)

            if not dfrekaptransaksilainlain.empty:
                dfrekaptransaksilainlain.to_excel(writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Rekap Barang Keluar")
                maxcolprevdf = maxcol
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=maxcol+2, value='Lain-lain')
                maxrow = len(dfrekaptransaksilainlain) + 1
                maxcol = len(dfrekaptransaksilainlain.columns) + maxcol + 1
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2, end_row=maxrow+2, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcol, value=totalbiayakeluartransaksilainlain)
                apply_number_format(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcolprevdf+2, maxcol)
                apply_borders_thin(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcol, maxcolprevdf+2)
                adjust_column_width(writer.sheets['Rekap Barang Keluar'], dfrekaptransaksilainlain, 1, maxcolprevdf+2)

            if not dfrekapgolongand.empty:
                dfrekapgolongand.to_excel(writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Rekap Barang Keluar")
                maxcolprevdf = maxcol
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=maxcol+2, value='Bahan Baku Golongan D Keluar')
                maxrow = len(dfrekapgolongand) + 1
                maxcol = len(dfrekapgolongand.columns) + maxcol + 1
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2, end_row=maxrow+2, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcol, value=totalbiayakeluargold)
                apply_number_format(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcolprevdf+2, maxcol)
                apply_borders_thin(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcol, maxcolprevdf+2)
                adjust_column_width(writer.sheets['Rekap Barang Keluar'], dfrekapgolongand, 1, maxcolprevdf+2)
            
            if not dfrekappemusnahanartikel.empty:
                dfrekappemusnahanartikel.to_excel(writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Rekap Barang Keluar")
                maxcolprevdf = maxcol
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=maxcol+2, value='Rekap Pemusnahan Artikel')
                maxrow = len(dfrekappemusnahanartikel) + 1
                maxcol = len(dfrekappemusnahanartikel.columns) + maxcol + 1
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2, end_row=maxrow+2, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcol, value=totalbiayapemusnahanartikel)
                apply_number_format(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcolprevdf+2, maxcol)
                apply_borders_thin(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcol, maxcolprevdf+2)
                adjust_column_width(writer.sheets['Rekap Barang Keluar'], dfrekappemusnahanartikel, 1, maxcolprevdf+2)

            if not dfrekappemusnahanbahanbaku.empty:
                dfrekappemusnahanbahanbaku.to_excel(writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Rekap Barang Keluar")
                maxcolprevdf = maxcol
                writer.sheets["Rekap Barang Keluar"].cell(row=1, column=maxcol+2, value='Rekap Pemusnahan Bahan Baku')
                maxrow = len(dfrekappemusnahanbahanbaku) + 1
                maxcol = len(dfrekappemusnahanbahanbaku.columns) + maxcol + 1
                writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2, end_row=maxrow+2, end_column=maxcol-1)
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+2, column=maxcol, value=totalbiayapemusnahanbahanbaku)
                apply_number_format(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcolprevdf+2, maxcol)
                apply_borders_thin(writer.sheets['Rekap Barang Keluar'], 2, maxrow+2, maxcol, maxcolprevdf+2)
                adjust_column_width(writer.sheets['Rekap Barang Keluar'], dfrekappemusnahanbahanbaku, 1, maxcolprevdf+2)

        '''LAPORAN BARANG KELUAR'''
        if not dfdatakeluar.empty or not dfdatakeluardisplay.empty or not dftransaksibahanbaku.empty or not dftransaksigold.empty or not dftransaksilainlain.empty or not dfpemusnahanbahanbaku or not dfdatapemusnahanartikel:
        # Artikel Keluar
            maxcol = 0
            maxrow = 0
            maxcolprevdf = 0
            if not dfdatakeluar.empty:
                dfdatakeluar.to_excel(
                    writer, index=False, startrow=1, sheet_name="Barang Keluar"
                )
                maxrow = len(dfdatakeluar)+1
                maxcol = len(dfdatakeluar.columns)
                writer.sheets["Barang Keluar"].cell(row=1, column = 1,value =listbulan[waktuobj.month - 1])
                writer.sheets["Barang Keluar"].cell(row=1, column = 2,value ='Artikel Keluar')
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=1,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = 1).value = "Total Harga"
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcol,value = totalbiayakeluar)
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain + totalbiayapemusnahanartikel + totalbiayapemusnahanbahanbaku)
                apply_number_format(writer.sheets['Barang Keluar'],2,maxrow+3,1,maxcol)
                # print(totalbiayakeluar,totalbiayakeluargold,totalbiayakeluardisplay,totalbiayakeluartransaksibahanbaku,totalbiayakeluartransaksilainlain)
                apply_borders_thin(writer.sheets['Barang Keluar'],2,maxrow+3,maxcol)
                adjust_column_width(writer.sheets['Barang Keluar'],dfdatakeluar,1,1)

            if not dfdatakeluardisplay.empty:
                # Transaksi Display
                dfdatakeluardisplay.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Barang Keluar"
                )
                maxcolprevdf = maxcol
                writer.sheets["Barang Keluar"].cell(row=1, column = maxcol+2,value ='Display Keluar')
                maxrow = len(dfdatakeluardisplay)+1
                maxcol = len(dfdatakeluardisplay.columns)+ maxcol + 1
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcol,value = totalbiayakeluardisplay)
                apply_number_format(writer.sheets['Barang Keluar'],2,maxrow+2,maxcolprevdf+2,maxcol)
                apply_borders_thin(writer.sheets['Barang Keluar'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                adjust_column_width(writer.sheets['Barang Keluar'],dfdatakeluardisplay,1,maxcolprevdf+2)

            if not dftransaksibahanbaku.empty:
            # Transaksi Bahan Baku
                dftransaksibahanbaku.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Barang Keluar"
                )
                maxcolprevdf = maxcol
                writer.sheets["Barang Keluar"].cell(row=1, column = maxcol+2,value ='Transaksi Bahan Baku')
                maxrow = len(dftransaksibahanbaku)+1
                maxcol = len(dftransaksibahanbaku.columns)+ maxcol + 1
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcol,value = totalbiayakeluartransaksibahanbaku)
                apply_number_format(writer.sheets['Barang Keluar'],2,maxrow+2,maxcolprevdf+2,maxcol)
                apply_borders_thin(writer.sheets['Barang Keluar'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                adjust_column_width(writer.sheets['Barang Keluar'],dftransaksibahanbaku,1,maxcolprevdf+2)
                
            if not dftransaksilainlain.empty:
                # Transaksi Lain-Lain
                dftransaksilainlain.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Barang Keluar"
                )
                maxcolprevdf = maxcol
                writer.sheets["Barang Keluar"].cell(row=1, column = maxcol+2,value ='Lain-lain')
                maxrow = len(dftransaksilainlain)+1
                maxcol = len(dftransaksilainlain.columns)+ maxcol + 1
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcol,value = totalbiayakeluartransaksilainlain)
                apply_number_format(writer.sheets['Barang Keluar'],2,maxrow+2,maxcolprevdf+2,maxcol)
                apply_borders_thin(writer.sheets['Barang Keluar'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                adjust_column_width(writer.sheets['Barang Keluar'],dftransaksilainlain,1,maxcolprevdf+2)

            if not dftransaksigold.empty:
            # Transaksi Golongan D
                dftransaksigold.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Barang Keluar"
                )
                maxcolprevdf = maxcol
                writer.sheets["Barang Keluar"].cell(row=1, column = maxcol+2,value ='Bahan Baku Golongan D Keluar')
                maxrow = len(dftransaksigold)+1
                maxcol = len(dftransaksigold.columns)+ maxcol + 1
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcol,value = totalbiayakeluargold)
                apply_number_format(writer.sheets['Barang Keluar'],2,maxrow+2,maxcolprevdf+2,maxcol)
                apply_borders_thin(writer.sheets['Barang Keluar'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                adjust_column_width(writer.sheets['Barang Keluar'],dftransaksigold,1,maxcolprevdf+2)

            if not dfpemusnahanbahanbaku.empty:
            # Transaksi Golongan D
                dfpemusnahanbahanbaku.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Barang Keluar"
                )
                maxcolprevdf = maxcol
                writer.sheets["Barang Keluar"].cell(row=1, column = maxcol+2,value ='Pemusnahan Bahan Baku')
                maxrow = len(dfpemusnahanbahanbaku)+1
                maxcol = len(dfpemusnahanbahanbaku.columns)+ maxcol + 1
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcol,value = totalbiayapemusnahanbahanbaku)
                apply_number_format(writer.sheets['Barang Keluar'],2,maxrow+2,maxcolprevdf+2,maxcol)
                apply_borders_thin(writer.sheets['Barang Keluar'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                adjust_column_width(writer.sheets['Barang Keluar'],dfpemusnahanbahanbaku,1,maxcolprevdf+2)

            if not dfdatapemusnahanartikel.empty:
            # Transaksi Golongan D
                dfdatapemusnahanartikel.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Barang Keluar"
                )
                maxcolprevdf = maxcol
                writer.sheets["Barang Keluar"].cell(row=1, column = maxcol+2,value ='Pemusnahan Artikel')
                maxrow = len(dfdatapemusnahanartikel)+1
                maxcol = len(dfdatapemusnahanartikel.columns)+ maxcol + 1
                writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Barang Keluar"].cell(row=maxrow+2, column = maxcol,value = totalbiayapemusnahanartikel)
                apply_number_format(writer.sheets['Barang Keluar'],2,maxrow+2,maxcolprevdf+2,maxcol)
                apply_borders_thin(writer.sheets['Barang Keluar'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                adjust_column_width(writer.sheets['Barang Keluar'],dfdatapemusnahanartikel,1,maxcolprevdf+2)

        '''LAPORAN SALDO AWAL GUDANG'''
        if not dfstokgudang.empty:
            # Stok Gudang
            dfstokgudang.to_excel(
                writer, index=False, startrow=1, sheet_name="Saldo Awal Gudang"
            )
            
            #  Transaksi lain lain belum masuk
            maxrow = len(dfstokgudang)+1
            maxcol = len(dfstokgudang.columns)
            writer.sheets["Saldo Awal Gudang"].cell(row=1, column = 1,value =listbulan[waktuobj.month - 1])
            writer.sheets["Saldo Awal Gudang"].merge_cells(start_row=maxrow+2, start_column=1,end_row = maxrow+2,end_column= maxcol-1)
            writer.sheets["Saldo Awal Gudang"].cell(row=maxrow+2, column = 1).value = "Total Harga"
            writer.sheets["Saldo Awal Gudang"].cell(row=maxrow+2, column = maxcol,value = totalbiayaawalgudang)
            apply_number_format(writer.sheets['Saldo Awal Gudang'],2,maxrow+2,1,maxcol)
            apply_borders_thin(writer.sheets['Saldo Awal Gudang'],2,maxrow+2,maxcol )
            adjust_column_width(writer.sheets['Saldo Awal Gudang'],dfstokgudang,1,1)

        # Laporan stok Produksi Section
        if not dfstokawalwip.empty or not dfstokawalfg.empty:
            maxcol = 0
            maxrow = 0
            maxcolprevdf = 0
            if not dfstokawalwip.empty:
                '''LAPORAN SALDO AWAL WIP'''
                # data stok awal WIP
                dfstokawalwip.to_excel(writer, index=False, startrow=1, sheet_name="Saldo Awal WIP")
                maxrow = len(dfstokawalwip)+1
                maxcol = len(dfstokawalwip.columns)
                writer.sheets["Saldo Awal WIP"].cell(row=1, column = 1,value =listbulan[waktuobj.month - 1])

                # writer.sheets["Saldo Awal WIP"].merge_cells(start_row=maxrow+2, start_column=1,end_row = maxrow+2,end_column= maxcol-1)
                # writer.sheets["Saldo Awal WIP"].cell(row=maxrow+2, column = 1).value = "Total Harga"
                # writer.sheets["Saldo Awal WIP"].cell(row=maxrow+2, column = maxcol,value = totalsaldoawalwip)
                apply_number_format(writer.sheets['Saldo Awal WIP'],2,maxrow+2,1,maxcol)
                apply_borders_thin(writer.sheets['Saldo Awal WIP'],2,maxrow+2,maxcol )
                adjust_column_width(writer.sheets['Saldo Awal WIP'],dfstokawalwip,1,1)
            if not dfstokawalfg.empty:
                # data stok awal FG
                dfstokawalfg.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Saldo Awal WIP"
                )
                maxcolprevdf = maxcol
                writer.sheets["Saldo Awal WIP"].cell(row=1, column = maxcol+2,value ='Saldo awal FG')
                maxrow = len(dfstokawalfg)+1
                maxcol = len(dfstokawalfg.columns)+ maxcol + 1
                # writer.sheets["Saldo Awal WIP"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                # writer.sheets["Saldo Awal WIP"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                # writer.sheets["Saldo Awal WIP"].cell(row=maxrow+2, column = maxcol,value = totalsaldoawalfg)
                apply_number_format(writer.sheets['Saldo Awal WIP'],2,maxrow+2,1,maxcol+3)
                apply_borders_thin(writer.sheets['Saldo Awal WIP'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                adjust_column_width(writer.sheets['Saldo Awal WIP'],dfstokawalfg,1,maxcolprevdf+2)

        # writer.sheets["Saldo Awal WIP"].cell(row=2, column = maxcol+2).value = "Total Harga Saldo Awal"
        # writer.sheets["Saldo Awal WIP"].cell(row=2, column = maxcol+3,value = totalsaldoawalfg+totalsaldoawalwip)
        apply_borders_thin(writer.sheets['Saldo Awal WIP'],2,2,maxcol+3 ,maxcol+2)
        
        '''LAPORAN STOK AKHIR GUDANG'''
        if not dfstokgudangakhir.empty:
            dfstokgudangakhir.to_excel(writer, index=False, startrow=1, sheet_name="Saldo Gudang")
            maxrow = len(dfstokgudangakhir)+1
            maxcol = len(dfstokgudangakhir.columns)
            writer.sheets["Saldo Gudang"].cell(row=1, column = 1,value =listbulan[waktuobj.month - 1])

            writer.sheets["Saldo Gudang"].merge_cells(start_row=maxrow+2, start_column=1,end_row = maxrow+2,end_column= maxcol-1)
            writer.sheets["Saldo Gudang"].cell(row=maxrow+2, column = 1).value = "Total Harga"
            writer.sheets["Saldo Gudang"].cell(row=maxrow+2, column = maxcol,value = totalsaldogudang)
            apply_number_format(writer.sheets['Saldo Gudang'],2,maxrow+2,1,maxcol)
            apply_borders_thin(writer.sheets['Saldo Gudang'],2,maxrow+2,maxcol )
            adjust_column_width(writer.sheets['Saldo Gudang'],dfstokgudangakhir,1,1)

        if not dfstokfgartikel.empty or not dfstokfgdisplay.empty or not dfstokfgbahanbaku.empty:
            maxcol = 0
            maxrow = 0
            maxcolprevdf = 0
            if not dfstokfgartikel.empty:
                '''LAPORAN STOK FG '''
                # Artikel
                dfstokfgartikel.to_excel(writer, index=False, startrow=1, sheet_name="Saldo FG")
                maxrow = len(dfstokfgartikel)+1
                maxcol = len(dfstokfgartikel.columns)
                writer.sheets["Saldo FG"].cell(row=1, column = 1,value =listbulan[waktuobj.month - 1])

                writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+2, start_column=1,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Saldo FG"].cell(row=maxrow+2, column = 1).value = "Total Harga"
                writer.sheets["Saldo FG"].cell(row=maxrow+2, column = maxcol,value = totalsaldofgartikel)
                writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                writer.sheets["Saldo FG"].cell(row=maxrow+3, column = 1).value = "Total Harga Saldo FG"
                writer.sheets["Saldo FG"].cell(row=maxrow+3, column = maxcol,value = totalsaldofgartikel+totalsaldofgbahanbaku+totalsaldofgdisplay)
                apply_borders_thin(writer.sheets['Saldo FG'],2,maxrow+3,maxcol )
                apply_number_format(writer.sheets['Saldo FG'],2,maxrow+3,1,maxcol)
                adjust_column_width(writer.sheets['Saldo FG'],dfstokfgartikel,1,1)
            if not dfstokfgdisplay.empty:
                # Display
                dfstokfgdisplay.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Saldo FG"
                )
                maxcolprevdf = maxcol
                writer.sheets["Saldo FG"].cell(row=1, column = maxcol+2,value ='Saldo Display FG')
                maxrow = len(dfstokfgdisplay)+1
                maxcol = len(dfstokfgdisplay.columns)+ maxcol + 1
                writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Saldo FG"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Saldo FG"].cell(row=maxrow+2, column = maxcol,value = totalsaldofgdisplay)
                apply_borders_thin(writer.sheets['Saldo FG'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                apply_number_format(writer.sheets['Saldo FG'],2,maxrow+2,maxcolprevdf+2,maxcol)
                adjust_column_width(writer.sheets['Saldo FG'],dfstokfgdisplay,1,maxcolprevdf+2)
            if not dfstokfgbahanbaku.empty:  
                # Bahan Baku
                dfstokfgbahanbaku.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Saldo FG"
                )
                maxcolprevdf = maxcol
                writer.sheets["Saldo FG"].cell(row=1, column = maxcol+2,value ='Saldo Bahan Baku FG')
                maxrow = len(dfstokfgbahanbaku)+1
                maxcol = len(dfstokfgbahanbaku.columns)+ maxcol + 1
                writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+2, start_column=maxcolprevdf + 2,end_row = maxrow+2,end_column= maxcol-1)
                writer.sheets["Saldo FG"].cell(row=maxrow+2, column = maxcolprevdf+2).value = "Total Harga"
                writer.sheets["Saldo FG"].cell(row=maxrow+2, column = maxcol,value = totalsaldofgbahanbaku)
                apply_borders_thin(writer.sheets['Saldo FG'],2,maxrow+2,maxcol ,maxcolprevdf+2)
                apply_number_format(writer.sheets['Saldo FG'],2,maxrow+2,maxcolprevdf+2,maxcol)
                adjust_column_width(writer.sheets['Saldo FG'],dfstokfgbahanbaku,1,maxcolprevdf+2)

    buffer.seek(0)
    # print('tes')
    wb = load_workbook(buffer)
   
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=laporanpersediaan{bulan}.xlsx"
    )
    
    print('Waktu generate laporan : ',time.time()-waktupersediaan)
    print('Waktu Proses : ', time.time()-waktuawalproses)
    return response

def apply_number_format(worksheet, start_row, end_row, start_col, end_col, number_format='#,##0.00'):
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Apply format only to numeric cells
                cell.number_format = number_format

def adjust_column_width(worksheet, df, start_row, start_col):
    for i, col in enumerate(df.columns):
        max_length = max(df[col].astype(str).map(len).max(), len(col))
        col_letter = get_column_letter(start_col + i)
        worksheet.column_dimensions[col_letter].width = max_length + 4

def apply_borders_thin(worksheet, start_row, end_row, max_col,min_col=1):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def create_dataframeproduksi(data):
    datamodel = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }
    for item in data["data"]:
        datamodel["Harga Satuan"].append(item.hargasatuanawal)
        datamodel["Jumlah"].append(item.jumlahawal)
        datamodel["Kode Produk"].append(item.KodeProduk)
        datamodel["Nama Produk"].append(item.NamaProduk)
        datamodel["Total Biaya"].append(item.totalbiayaawal)
        datamodel["Unit"].append(item.unit)
    return pd.DataFrame(datamodel)


# Cek Perhitungan Laporan


def exportlaporanbulananexcelkeseluruhan(request):
    bulan = request.GET["bulan"]
    waktuobj = datetime.strptime(bulan, "%Y-%m")

    awaltahun = datetime(waktuobj.year, 1, 1)
    listbulan = [
        "Januari",
        "Februari",
        "Maret",
        "April",
        "Mei",
        "Juni",
        "Juli",
        "Agustus",
        "September",
        "Oktober",
        "Novermber",
        "Desember",
    ]
    index = int(waktuobj.month)
    last_days = []
    for month in range(1, 13):
        last_day = calendar.monthrange(waktuobj.year, month)[1]
        last_days.append(date(waktuobj.year, month, last_day))
        
    datadetailsppb, totalbiayakeluar, datapenyusun, listdatadetailsppb,transaksilainlain, transaksigold, detailbiaya, datatransaksikeluar = (
        getbarangkeluar(last_days, index, awaltahun)
    )
    """SECTION HARGA AKHIR PERBULAN """
    hargaakhirbulan = gethargapurchasingperbulan(last_days, index, awaltahun)
    """SECTION BARANG MASUK"""
    barangmasuk = getbarangmasuk(last_days, index, awaltahun)
    """SECTION STOCK GUDANG"""
    baranggudang = saldogudang(last_days, index, awaltahun)
    print(baranggudang)
    # print(asdas)
    """SECTION FG"""
    # barangfg, bahanbakusisafg = getstokartikelfg(last_days, index, awaltahun)
    stokfg = getstokfg(None,last_days,index,awaltahun)
    # print(barangfg)
    # print(asdas)
    """SECTION STOK AWAL PRODUKSI"""
    nilaisaldoawalproduksi, saldoawalbahanbaku, saldoawalartikel = getstokbahanproduksi(
        last_days, index, awaltahun
    )
    """SECTION WIP (Skip dulu)"""
    dataperhitunganpersediaan = perhitunganpersediaan(
        last_days,
        index,
        awaltahun,
        baranggudang,
        barangmasuk,
        totalbiayakeluar,
        stokfg,

    )

    """PEMBUATAN EXCEL"""
    """ 1. sheet untuk laporan persediaan"""
    listpersediaan = []
    for key, value in dataperhitunganpersediaan.items():
        datamodelpersediaan ={
        "Barang Keluar" : [value['barangkeluar']],
        "Barang Masuk" : [value['barangmasuk']],
        "Saldo Awal Stock Gudang" : [value['saldoawalgudang']],
        "Saldo Awal Bahan Produksi" : [value['stokawalproduksi']],
        "Total Saldo":[value['totalsaldo']],
        "Saldo Akhir Gudang" : [value['saldoakhirgudang']],
        "Saldo Akhir FG" : [value['stokfg']],
        "Saldo Akhir WIP" : [value['saldowip']],
    }
        dfpersediaan = pd.DataFrame(datamodelpersediaan)

        # df = pd.DataFrame.from_dict(value, orient="index").T
        listpersediaan.append(dfpersediaan)
    # print(listpersediaan)
    # print(asd)
    """2. Sheet untuk laporan barang masuk"""
    # print(listpersediaan)
    # print(asdas)
    listdatamodelmasuk = []
    # print(barangmasuk)
    for bulan, value in barangmasuk.items():
        datamodelmasuk = {
            "Tanggal Masuk": [],
            "No Surat Jalan": [],
            "Supplier": [],
            "Kode Produk": [],
            "Satuan": [],
            "Harga Satuan": [],
            "Jumlah": [],
            "Harga Total": [],
        }
        key = value["data"]
        for masuk in key:
            datamodelmasuk["Tanggal Masuk"].append(masuk.NoSuratJalan.Tanggal)
            datamodelmasuk["No Surat Jalan"].append(masuk.NoSuratJalan.NoSuratJalan)
            datamodelmasuk["Supplier"].append(masuk.NoSuratJalan.supplier)
            datamodelmasuk["Kode Produk"].append(masuk.KodeProduk.KodeProduk)
            datamodelmasuk["Satuan"].append(masuk.KodeProduk.unit)
            datamodelmasuk["Harga Satuan"].append(masuk.Harga)
            datamodelmasuk["Harga Total"].append(masuk.Harga * masuk.Jumlah)
            datamodelmasuk["Jumlah"].append(masuk.Jumlah)
        # print(datamodelmasuk)
        dfmasuk = pd.DataFrame(datamodelmasuk)
        totalsaldomasuk = sum(datamodelmasuk['Harga Total'])
        listdatamodelmasuk.append([dfmasuk,totalsaldomasuk])

    
    # print(listdatamodelmasuk)
    # print(asd)

    """3. Sheet untuk Barang Keluar"""
    print("detail sppb \n\n", listdatadetailsppb)
    # print(asdasdas)

    listdatamodelkeluar = []
    listdatadisplaykeluar = []
    listdatabahanbakukeluar = []
    listtransaksilainlain = []
    listtransaksigolongand = []
    listpemusnahanbahanbaku = []
    listpemusnahanartikel = []

    listrekapdatamodelkeluar = []
    listrekapdatadisplaykeluar = []
    listrekapdatabahanbakukeluar = []
    listrekaptransaksilainlain = []
    listrekaptransaksigolongand = []
    listrekappemusnahanbahanbaku = []
    listrekappemusnahanartikel = []
    
    totalbiayatransaksiartikel = 0
    totalbiayatransaksidisplay = 0
    totalbiayatransaksibahanbaku = 0
    totalbiayatransaksigolongand = 0
    totalbiayatransaksilainlain= 0
    totalbiayapemusnahanartikel = 0
    totalbiayapemusnahanbahanbaku = 0
    for item in datatransaksikeluar:

        '''Detail Transaksi'''
        datamodelkeluar = {
            "Tanggal Keluar": [],
            "No SPPB": [],
            "NO SPK": [],
            "Artikel": [],
            "Jumlah": [],
            "Harga FG": [],
            "Total Biaya": [],
        }
        for data in item['SPPBArtikel']['SPPBArtikel']:
            print(data)
            datamodelkeluar["Jumlah"].append(data.Jumlah)
            datamodelkeluar["Artikel"].append(data.DetailSPK.KodeArtikel)
            datamodelkeluar["Harga FG"].append(data.hargafg)
            datamodelkeluar["NO SPK"].append(data.DetailSPK.NoSPK)
            datamodelkeluar["No SPPB"].append(data.NoSPPB)
            datamodelkeluar["Total Biaya"].append(data.totalharga)
            datamodelkeluar["Tanggal Keluar"].append(data.NoSPPB.Tanggal)
        df = pd.DataFrame(datamodelkeluar)
        totalbiayakeluar = sum(datamodelkeluar["Total Biaya"])
        totalbiayatransaksiartikel += totalbiayakeluar
        listdatamodelkeluar.append([df,totalbiayakeluar])

        datamodelkeluar = {
"Tanggal": [],
        "Artikel": [],
        'Lokasi' : [],
        "Jumlah": [],
        "Harga FG": [],
        "Total Biaya": [],
        }
        for data in item['Pemusnahanartikel']['transaksipemusnahan']:
            print(data)
            datamodelkeluar["Tanggal"].append(data.Tanggal)
            datamodelkeluar["Jumlah"].append(data.Jumlah)
            datamodelkeluar['Lokasi'].append(data.lokasi)
            datamodelkeluar["Artikel"].append(data.KodeArtikel)
            datamodelkeluar["Harga FG"].append(data.hargafg)
            datamodelkeluar["Total Biaya"].append(data.totalharga)
        df = pd.DataFrame(datamodelkeluar)
        biayapemusnahanartikel = sum(datamodelkeluar["Total Biaya"])
        totalbiayapemusnahanartikel += biayapemusnahanartikel
        listpemusnahanartikel.append([df,biayapemusnahanartikel])

        
        datamodelstransaksigold = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
        }
    
        for data in item['Transaksigolongand']['datatransaksi']:
            datamodelstransaksigold["Kode Produk"].append(data.KodeProduk.KodeProduk)
            datamodelstransaksigold["Unit"].append(data.KodeProduk.unit)
            datamodelstransaksigold["Nama Produk"].append(data.KodeProduk.NamaProduk)
            datamodelstransaksigold["Jumlah"].append(data.jumlah)
            datamodelstransaksigold["Total Biaya"].append(data.hargatotal)
            datamodelstransaksigold["Harga Satuan"].append(data.harga)
        dftransaksigold = pd.DataFrame(datamodelstransaksigold)
        totalbiayakeluargold = sum(datamodelstransaksigold["Total Biaya"])
        totalbiayatransaksigolongand +=totalbiayakeluargold
        listtransaksigolongand.append([dftransaksigold,totalbiayakeluargold])

        datamodelspemusnahanbahanbaku = {
                "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
        }
    
        for data in item['Pemusnahanbahanbaku']['datatransaksi']:
           
            datamodelspemusnahanbahanbaku["Kode Produk"].append(data.KodeBahanBaku.KodeProduk)
            datamodelspemusnahanbahanbaku["Unit"].append(data.KodeBahanBaku.unit)
            datamodelspemusnahanbahanbaku["Nama Produk"].append(data.KodeBahanBaku.NamaProduk)
            datamodelspemusnahanbahanbaku["Jumlah"].append(data.Jumlah)
            datamodelspemusnahanbahanbaku["Total Biaya"].append(data.hargatotal)
            datamodelspemusnahanbahanbaku["Harga Satuan"].append(data.harga)
    
        df = pd.DataFrame(datamodelspemusnahanbahanbaku)
        biayapemusnahanbahanbaku = sum(datamodelspemusnahanbahanbaku["Total Biaya"])
        totalbiayapemusnahanbahanbaku +=biayapemusnahanbahanbaku
        listpemusnahanbahanbaku.append([df,biayapemusnahanbahanbaku])


        datamodelstransaksilainlain = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }
        
        for data in item['Transaksilainlain']['datatransaksi']:
            datamodelstransaksilainlain["Kode Produk"].append(data.KodeProduk.KodeProduk)
            datamodelstransaksilainlain["Unit"].append(data.KodeProduk.unit)
            datamodelstransaksilainlain["Nama Produk"].append(data.KodeProduk.NamaProduk)
            datamodelstransaksilainlain["Jumlah"].append(data.jumlah)
            datamodelstransaksilainlain["Total Biaya"].append(data.hargatotal)
            datamodelstransaksilainlain["Harga Satuan"].append(data.harga)
    
        dftransaksilainlain = pd.DataFrame(datamodelstransaksilainlain)
        totalbiayakeluartransaksilainlain = sum(datamodelstransaksilainlain["Total Biaya"])
        totalbiayatransaksilainlain += totalbiayakeluartransaksilainlain
        listtransaksilainlain.append([dftransaksilainlain,totalbiayakeluartransaksilainlain])
    # Transaksi Bahan Baku
        datamodelstransaksibahanbaku = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }

        for data in item['Transaksibahanbaku']['datatransaksi']:
            datamodelstransaksibahanbaku["Kode Produk"].append(data.DetailBahan.KodeProduk)
            datamodelstransaksibahanbaku["Unit"].append(data.DetailBahan.unit)
            datamodelstransaksibahanbaku["Nama Produk"].append(data.DetailBahan.NamaProduk)
            datamodelstransaksibahanbaku["Jumlah"].append(data.Jumlah)
            datamodelstransaksibahanbaku["Total Biaya"].append(data.hargatotal)
            datamodelstransaksibahanbaku["Harga Satuan"].append(data.harga)
        
        dftransaksibahanbaku = pd.DataFrame(datamodelstransaksibahanbaku)
        totalbiayakeluartransaksibahanbaku = sum(datamodelstransaksibahanbaku["Total Biaya"])
        totalbiayatransaksibahanbaku +=totalbiayakeluartransaksibahanbaku
        listdatabahanbakukeluar.append([dftransaksibahanbaku,totalbiayakeluartransaksibahanbaku])

        datamodeldisplay = {
        "Tanggal Keluar": [],
        "No SPPB": [],
        "NO SPK": [],
        "Artikel": [],
        "Jumlah": [],
        "Harga FG": [],
        "Total Biaya": [],
    }

        for data in item['SPPBDisplay']['SPPBDisplay']:
            print(data)
            datamodeldisplay["Jumlah"].append(data.Jumlah)
            datamodeldisplay["Artikel"].append(data.DetailSPKDisplay.KodeDisplay)
            datamodeldisplay["Harga FG"].append(data)
            datamodeldisplay["NO SPK"].append(data.DetailSPKDisplay.NoSPK)
            datamodeldisplay["No SPPB"].append(data.NoSPPB)
            datamodeldisplay["Total Biaya"].append(data.totalbiaya)
            datamodeldisplay["Tanggal Keluar"].append(data.NoSPPB.Tanggal)


        dfdatakeluardisplay = pd.DataFrame(datamodeldisplay)
        totalbiayakeluardisplay = sum(datamodeldisplay["Total Biaya"])
        totalbiayatransaksidisplay += totalbiayakeluardisplay
        listdatadisplaykeluar.append([dfdatakeluardisplay,totalbiayakeluardisplay])

        '''Rekap Artikel Keluar'''
        rekapartikel,rekaptransaksidisplay,rekapdatabahankeluar,rekapgolongand,rekaptransaksilainlain,rekappemusnahanartikel,rekappemusnahanbahanbaku = create_rekaptransaksikeluar(item)
        datamodelsrekapartikelkeluar = {
        'Kode Artikel' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
        for artikel,data in rekapartikel.items():
            datamodelsrekapartikelkeluar['Kode Artikel'].append( artikel.KodeArtikel)
            datamodelsrekapartikelkeluar["Harga FG"].append( data['HargaFG'])
            datamodelsrekapartikelkeluar['Harga Total'].append( data['hargatotal'])
            datamodelsrekapartikelkeluar["Jumlah"].append( data['Jumlah'])
        dfrekapartikelkeluar = pd.DataFrame(datamodelsrekapartikelkeluar)
        listrekapdatamodelkeluar.append((dfrekapartikelkeluar,totalbiayatransaksiartikel))

        datamodelsrekappemusnahanartikel = {
        'Kode Artikel' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
        for artikel,data in rekappemusnahanartikel.items():
            datamodelsrekappemusnahanartikel['Kode Artikel'].append( artikel.KodeArtikel)
            datamodelsrekappemusnahanartikel["Harga FG"].append( data['HargaFG'])
            datamodelsrekappemusnahanartikel['Harga Total'].append( data['hargatotal'])
            datamodelsrekappemusnahanartikel["Jumlah"].append( data['Jumlah'])
        
        dfrekappemusnahanartikel = pd.DataFrame(datamodelsrekappemusnahanartikel)
        listrekappemusnahanartikel.append([dfrekappemusnahanartikel,totalbiayapemusnahanartikel])

        datamodelsdisplaykeluar = {
        'Kode Display' : [],
        'No SPK' : [],
        'Jumlah' : [],
        'Harga Satuan' : [],
        'Harga Total': []
    }
        for display, data_nospk in rekaptransaksidisplay.items():
            for spk, details in data_nospk['NoSPK'].items():
                datamodelsdisplaykeluar['Kode Display'].append(display)
                datamodelsdisplaykeluar['No SPK'].append(spk)
                datamodelsdisplaykeluar['Jumlah'].append(details['Jumlah'])
                datamodelsdisplaykeluar['Harga Satuan'].append(details['hargasatuan'])
                datamodelsdisplaykeluar['Harga Total'].append(details['Hargatotal'])
        dfrekapdisplaykeluar = pd.DataFrame(datamodelsdisplaykeluar)
        listrekapdatadisplaykeluar.append([dfrekapdisplaykeluar,totalbiayatransaksidisplay])

        datamodelstransaksibahanbaku = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
        for artikel,data in rekapdatabahankeluar.items():
            datamodelstransaksibahanbaku['Kode Bahan Baku'].append( artikel.KodeProduk)
            datamodelstransaksibahanbaku['Nama Bahan Baku'].append( artikel.NamaProduk)
            datamodelstransaksibahanbaku['Satuan'].append( artikel.unit)
            datamodelstransaksibahanbaku["Harga FG"].append( data['HargaFG'])
            datamodelstransaksibahanbaku['Harga Total'].append( data['hargatotal'])
            datamodelstransaksibahanbaku["Jumlah"].append( data['Jumlah'])
        
        dfrekapbahanbakukeluar = pd.DataFrame(datamodelstransaksibahanbaku)
        listrekapdatabahanbakukeluar.append([dfrekapbahanbakukeluar,totalbiayatransaksibahanbaku])

        datamodelspemusnahanbahanbaku = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
        for artikel,data in rekappemusnahanbahanbaku.items():
            datamodelspemusnahanbahanbaku['Kode Bahan Baku'].append( artikel.KodeProduk)
            datamodelspemusnahanbahanbaku['Nama Bahan Baku'].append( artikel.NamaProduk)
            datamodelspemusnahanbahanbaku['Satuan'].append( artikel.unit)
            datamodelspemusnahanbahanbaku["Harga FG"].append( data['HargaFG'])
            datamodelspemusnahanbahanbaku['Harga Total'].append( data['hargatotal'])
            datamodelspemusnahanbahanbaku["Jumlah"].append( data['Jumlah'])
        
        dfrekappemusnahanbahanbaku = pd.DataFrame(datamodelspemusnahanbahanbaku)
        listrekappemusnahanbahanbaku.append([dfrekappemusnahanbahanbaku,totalbiayapemusnahanbahanbaku])

        datamodelstransaksigolongand = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
        for artikel,data in rekapgolongand.items():
            datamodelstransaksigolongand['Kode Bahan Baku'].append( artikel.KodeProduk)
            datamodelstransaksigolongand['Nama Bahan Baku'].append( artikel.NamaProduk)
            datamodelstransaksigolongand['Satuan'].append( artikel.unit)
            datamodelstransaksigolongand["Harga FG"].append( data['Harga'])
            datamodelstransaksigolongand['Harga Total'].append( data['hargatotal'])
            datamodelstransaksigolongand["Jumlah"].append( data['Jumlah'])
        
        dfrekapgolongand = pd.DataFrame(datamodelstransaksigolongand)
        listrekaptransaksigolongand.append([dfrekapgolongand,totalbiayatransaksigolongand])

        datamodelstransaksilainlain = {
        'Kode Bahan Baku' : [],
        'Nama Bahan Baku' : [],
        'Satuan' : [],
        'Jumlah' : [],
        'Harga FG' : [],
        'Harga Total': []
    }
        for artikel,data in rekaptransaksilainlain.items():
            datamodelstransaksilainlain['Kode Bahan Baku'].append( artikel.KodeProduk)
            datamodelstransaksilainlain['Nama Bahan Baku'].append( artikel.NamaProduk)
            datamodelstransaksilainlain['Satuan'].append( artikel.unit)
            datamodelstransaksilainlain["Harga FG"].append( data['Harga'])
            datamodelstransaksilainlain['Harga Total'].append( data['hargatotal'])
            datamodelstransaksilainlain["Jumlah"].append( data['Jumlah'])
        
        dfrekaptransaksilainlain = pd.DataFrame(datamodelstransaksilainlain)
        listrekaptransaksilainlain.append([dfrekaptransaksilainlain,totalbiayatransaksilainlain])

    # print(listdatamodelkeluar)
    # print(listtransaksigolongand)
    # print(listdatabahanbakukeluar)
    print(listdatadisplaykeluar)

    # print(adsasd)
    tabeltransaksi = {
        'Total Transaksi Artikel': [totalbiayatransaksiartikel],
        'Total Transaksi Display': [totalbiayatransaksidisplay],
        'Total Transaksi Bahan Baku': [totalbiayatransaksibahanbaku],
        'Total Transaksi Golongan D': [totalbiayatransaksigolongand],
        'Total Transaksi Lain-Lain': [totalbiayatransaksilainlain],
        'Total Pemusnahan Artikel' : [totalbiayapemusnahanartikel],
        'Total Pemusnahan Bahan Baku' : [totalbiayapemusnahanbahanbaku]
    }
    dataframetransaksikeluar = pd.DataFrame(tabeltransaksi)
    """4. Sheet Stock Bahan Baku Gudang"""

    # print(hargaakhirbulan)
    data_per_bulan = {i: {} for i in range(index)}
    for produk, values in hargaakhirbulan.items():
        for bulan_index, detail in values["data"].items():
            if bulan_index not in data_per_bulan:
                data_per_bulan[bulan_index] = {}
            data_per_bulan[bulan_index][produk] = detail
    print(data_per_bulan)
    # print(asd)
    listdatagudangperbulan = []
    for bulan,item in data_per_bulan.items():
        datamodelsstokgudang = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
        }
    
        for produk,data in item.items():
            datamodelsstokgudang["Kode Produk"].append(produk.KodeProduk)
            datamodelsstokgudang["Unit"].append(produk.unit)
            datamodelsstokgudang["Nama Produk"].append(produk.NamaProduk)
            datamodelsstokgudang["Jumlah"].append(data['jumlah'])
            datamodelsstokgudang["Total Biaya"].append(data['hargatotal'])
            datamodelsstokgudang["Harga Satuan"].append(data['hargasatuan'])
        dfstokgudang = pd.DataFrame(datamodelsstokgudang)
        totalbiayastokgudang = sum(datamodelsstokgudang["Total Biaya"])
        listdatagudangperbulan.append([dfstokgudang,totalbiayastokgudang])

    print(listdatagudangperbulan)
    # print(asd)
    """5. Sheet Stock Awal Produksi"""
    # data wip
    datamodelstockawalwip = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        'Jumlah WIP' : [],
        'Jumlah FG' : [],
        'Jumlah Total' : []
        
    }

    # print(asd)
    for stokawal in saldoawalbahanbaku[0]["data"]:
        # print(stokawal)
        datamodelstockawalwip["Harga Satuan"].append(stokawal.hargasatuanawal_wip)
        datamodelstockawalwip["Jumlah WIP"].append(stokawal.jumlahawal_wip)
        datamodelstockawalwip["Jumlah FG"].append(stokawal.jumlahawal_fg)
        datamodelstockawalwip["Jumlah Total"].append(stokawal.jumlahkumulatif)
        datamodelstockawalwip["Kode Produk"].append(stokawal.KodeProduk)
        datamodelstockawalwip["Nama Produk"].append(stokawal.NamaProduk)
        datamodelstockawalwip["Unit"].append(stokawal.unit)
    # print(datawip)
    dfstokawalwip = pd.DataFrame(datamodelstockawalwip)
    datamodelstockawalwip = {
        "Kode Artikel": [],
       "Jumlah WIP" : [],
       'Jumlah FG': []
        
    }
    for stokawal in saldoawalartikel:
        # print(stokawal)
        datamodelstockawalwip["Kode Artikel"].append(stokawal.KodeArtikel)
        datamodelstockawalwip["Jumlah WIP"].append(stokawal.jumlahwip)
        datamodelstockawalwip["Jumlah FG"].append(stokawal.jumlahfg)
    # print(datawip)
    dfstokawalfg = pd.DataFrame(datamodelstockawalwip)



    """6. SHEET STOK BARANG FG"""
    # print("\n\n")
    # print(stokfg)
    # print(asd)
    listdataartikelfg = []
    listdatadisplayfg = []
    listdatabahanbakufg = []
    for item in stokfg:
        datamodelsartikelfg = {
            "Kode Artikel": [],
            "Jumlah": [],
            "Harga FG": [],
            "Total Biaya": [],
        }
        for detail in item['Artikel']:
            datamodelsartikelfg["Kode Artikel"].append(detail['KodeArtikel'])
            datamodelsartikelfg["Jumlah"].append(detail['total'])
            datamodelsartikelfg["Harga FG"].append(detail['hargafg'])
            datamodelsartikelfg["Total Biaya"].append(detail['totalsaldo'])
            # print("\n\n")
        dfstokartikelfg = pd.DataFrame(datamodelsartikelfg)
        totalsaldoartikelfg = sum(datamodelsartikelfg['Total Biaya'])
        listdataartikelfg.append([dfstokartikelfg,totalsaldoartikelfg])

        datamodelsdisplayfg = {
            "Kode Display": [],
            "Jumlah": [],
            "Harga FG": [],
            "Total Biaya": [],
        }
        for detail in item['Display']:
            datamodelsdisplayfg["Kode Display"].append(detail['KodeDisplay'])
            datamodelsdisplayfg["Jumlah"].append(detail['total'])
            datamodelsdisplayfg["Harga FG"].append(detail['hargafg'])
            datamodelsdisplayfg["Total Biaya"].append(detail['totalsaldo'])
        
        dfstokdisplayfg = pd.DataFrame(datamodelsdisplayfg)
        totalsaldodisplayfg = sum(datamodelsdisplayfg['Total Biaya'])
        listdatadisplayfg.append([dfstokdisplayfg,totalsaldodisplayfg])

        datamodelsbahanbaku = {
            "Kode Bahan Baku": [],
            "Jumlah": [],
            "Harga Satuan": [],
            "Total Biaya": [],
        }
        for detail in item['BahanBaku']:
            datamodelsbahanbaku["Kode Bahan Baku"].append(detail['KodeProduk'])
            datamodelsbahanbaku["Jumlah"].append(detail['total'])
            datamodelsbahanbaku["Harga Satuan"].append(detail['hargasatuan'])
            datamodelsbahanbaku["Total Biaya"].append(detail['hargatotal'])
        
        dfstokbahanbaku = pd.DataFrame(datamodelsbahanbaku)
        totalbiayabahanbakufg = sum(datamodelsbahanbaku['Total Biaya'])
        listdatabahanbakufg.append([dfstokbahanbaku,totalbiayabahanbakufg])
    print(listdatabahanbakufg)
    # print(asd)

    """7. Stok Awal Gudang"""
    datamodelstockawalgudang = {
        "Kode Produk": [],
        "Nama Produk": [],
        "Unit": [],
        "Harga Satuan": [],
        "Jumlah": [],
        "Total Biaya": [],
    }
    # print(baranggudang["datasaldoawal"])
    for item in baranggudang["datasaldoawal"]:
        datamodelstockawalgudang["Kode Produk"].append(item.KodeProduk)
        datamodelstockawalgudang["Unit"].append(item.unit)
        datamodelstockawalgudang["Nama Produk"].append(item.NamaProduk)
        datamodelstockawalgudang["Jumlah"].append(item.jumlahawal)
        datamodelstockawalgudang["Total Biaya"].append(item.totalbiayaawal)
        datamodelstockawalgudang["Harga Satuan"].append(item.hargasatuanawal)

    # print(datamodelstockawalgudang)
    dfstokawalgudang = pd.DataFrame(datamodelstockawalgudang)
    totalbiayasaldoawalgudang = sum(datamodelstockawalgudang["Total Biaya"])

    buffer = BytesIO()
    print(listpersediaan)
    # print(asd)

    # Use pandas to write DataFrame to the BytesIO buffer
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        '''LAPORAN PERSEDIAAN'''
        startrow = 1
        for i, datamasuk in enumerate(listpersediaan):
            datamasuk.to_excel(
                writer,
                index=False,
                startrow=startrow,
                sheet_name="Laporan Persediaan",
            )
            ws = writer.sheets["Laporan Persediaan"]
            ws.cell(
                row=startrow, column=1, value=f"Laporan Persediaan {listbulan[i]}"
            )
            # Mendapatkan jumlah baris dari DataFrame saat ini
            maxcol = len(datamasuk.columns)

            maxrow = len(datamasuk) + 1
            print(datamasuk)
            print(maxrow)
            apply_number_format(writer.sheets['Laporan Persediaan'],startrow+2,startrow+2,1,maxcol)
            apply_borders_thin(writer.sheets['Laporan Persediaan'],startrow+2,startrow+2,maxcol)
            adjust_column_width(writer.sheets['Laporan Persediaan'],dfpersediaan,startrow+2,1)
            startrow += 2+maxrow

        '''LAPORAN BARANG MASUK'''
        startrow = 0
        for i, datamasuk in enumerate(listdatamodelmasuk):
            if len(datamasuk[0])==0:
                continue
            # Set the header for the month
            
            # Write the DataFrame to the sheet
            datamasuk[0].to_excel(writer, index=False, startrow=startrow+1, sheet_name="Barang Masuk")
            writer.sheets["Barang Masuk"].cell(row=startrow+1, column=1, value=f"Laporan Barang Masuk {listbulan[i]}")
            
            # Calculate the maximum row after writing the DataFrame
            maxrow =  len(datamasuk[0]) + 1
            maxcol = len(datamasuk[0].columns)
            
            # Merge cells for the total price row
            writer.sheets["Barang Masuk"].merge_cells(start_row=startrow+2+maxrow, start_column=1, end_row=startrow+2+maxrow, end_column=maxcol-1)
            
            # Set the total price label and value
            writer.sheets["Barang Masuk"].cell(row=startrow+2+maxrow, column=1, value="Total Harga")
            writer.sheets["Barang Masuk"].cell(row=startrow+2+maxrow, column=maxcol, value=datamasuk[1])
            
            # Apply formatting and borders
            apply_number_format(writer.sheets["Barang Masuk"], startrow+2, startrow+2+maxrow, 1, maxcol)
            apply_borders_thin(writer.sheets["Barang Masuk"], startrow+2, startrow+2+maxrow, maxcol)
            adjust_column_width(writer.sheets["Barang Masuk"], datamasuk[0], 1,1)
            
            # Update startrow for the next iteration, adding an extra row for separation
            startrow += maxrow + 3
        print(len(listdatabahanbakukeluar),len(listdatamodelkeluar),len(listdatadisplaykeluar))
        print(listdatabahanbakukeluar)
        # print(asd)
        '''LAPORAN REKAP BARANG KELUAR'''

        if len(listrekapdatabahanbakukeluar) != 0 or len(listrekapdatamodelkeluar)!=0 or len(listrekapdatadisplaykeluar) !=0 or len(listrekaptransaksigolongand)!=0 or len(listrekaptransaksilainlain) !=0 or len(listrekappemusnahanbahanbaku) !=0 or len(listrekappemusnahanartikel) !=0:
            maxcolumn  = 0
            maxrow = 0
            maxcolprevdf = 0
            # ARTIKEL KELUAR (SPPB)
            if len(listrekapdatamodelkeluar) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listrekapdatamodelkeluar):
                    if len(datakeluar[0]) == 0:
                        continue 
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = 0, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+1, column = 1,value =listbulan[i])
                    writer.sheets["Rekap Barang Keluar"].cell(row=1, column = 2,value ='Artikel Keluar')
                    writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+startrow+2, start_column=1,end_row = maxrow+startrow+2,end_column= maxcol-1)
                    writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+startrow+2, column = 1).value = "Total Harga"
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcol,value = datakeluar[1])
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Rekap Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1

            # DISPLAY KELUAR (SPPB)
            if len(listrekapdatadisplaykeluar) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listrekapdatadisplaykeluar):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Rekap Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Display Keluar')
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Rekap Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # Bahan Baku Keluar (SPPB)
            if len(listrekapdatabahanbakukeluar) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listrekapdatabahanbakukeluar):
                    if len(datakeluar[0]) == 0:
                        continue 
        
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Rekap Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Bahan Baku Keluar')
                    writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = startrow+maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Rekap Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1

            # TRANSAKSI GOL D (TRANSAKSI Gudang)
            if len(listrekaptransaksigolongand) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listrekaptransaksigolongand):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Rekap Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Transaksi Golongan D')
                    writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = startrow+maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Rekap Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # TRANSAKSI LAIN LAIN (TRANSAKSI GUDANG)
            if len(listrekaptransaksilainlain) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listrekaptransaksilainlain):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Rekap Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Transaksi Lain-lain')
                    writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row =startrow+ maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Rekap Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # Pemusnahan Artikel
            if len(listrekappemusnahanartikel) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listrekappemusnahanartikel):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Rekap Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Pemusnahan Artikel')
                    writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row =startrow+ maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Rekap Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # Pemusnahan Bahan Baku
            if len(listrekappemusnahanbahanbaku) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listrekappemusnahanbahanbaku):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Rekap Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Pemusnahan Bahan Baku')
                    writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row =startrow+ maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Rekap Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Rekap Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Rekap Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Rekap Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1



            startrow = 0
            dataframetransaksikeluar.to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Rekap Barang Keluar"
                )
            apply_number_format(writer.sheets['Rekap Barang Keluar'],startrow+1,startrow+3,maxcolprevdf,maxcolprevdf+6)
            apply_borders_thin(writer.sheets['Rekap Barang Keluar'],startrow+2,startrow+3,maxcolprevdf+7,maxcolprevdf+1)
            adjust_column_width(writer.sheets['Rekap Barang Keluar'],dataframetransaksikeluar,startrow+1,maxcolprevdf+1)
        '''LAPORAN BARANG KELUAR'''

        if len(listdatabahanbakukeluar) != 0 or len(listdatamodelkeluar)!=0 or len(listdatadisplaykeluar) !=0 or len(listtransaksigolongand)!=0 or len(listtransaksilainlain) !=0 or len(listpemusnahanbahanbaku) !=0 or len(listpemusnahanartikel) !=0:
            maxcolumn  = 0
            maxrow = 0
            maxcolprevdf = 0
            # ARTIKEL KELUAR (SPPB)
            if len(listdatamodelkeluar) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listdatamodelkeluar):
                    if len(datakeluar[0]) == 0:
                        continue 
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = 0, startrow=startrow+1, sheet_name="Barang Keluar"
                )
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    writer.sheets["Barang Keluar"].cell(row=startrow+1, column = 1,value =listbulan[i])
                    writer.sheets["Barang Keluar"].cell(row=1, column = 2,value ='Artikel Keluar')
                    writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+startrow+2, start_column=1,end_row = maxrow+startrow+2,end_column= maxcol-1)
                    writer.sheets["Barang Keluar"].cell(row=maxrow+startrow+2, column = 1).value = "Total Harga"
                    writer.sheets["Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcol,value = datakeluar[1])
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1

            # DISPLAY KELUAR (SPPB)
            if len(listdatadisplaykeluar) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listdatadisplaykeluar):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Display Keluar')
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # Bahan Baku Keluar (SPPB)
            if len(listdatabahanbakukeluar) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listdatabahanbakukeluar):
                    if len(datakeluar[0]) == 0:
                        continue 
        
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Bahan Baku Keluar')
                    writer.sheets["Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = startrow+maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1

            # TRANSAKSI GOL D (TRANSAKSI Gudang)
            if len(listtransaksigolongand) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listtransaksigolongand):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Transaksi Golongan D')
                    writer.sheets["Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = startrow+maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # TRANSAKSI LAIN LAIN (TRANSAKSI GUDANG)
            if len(listtransaksilainlain) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listtransaksilainlain):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Transaksi Lain-lain')
                    writer.sheets["Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row =startrow+ maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # Pemusnahan Artikel
            if len(listpemusnahanartikel) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listpemusnahanartikel):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Pemusnahan Artikel')
                    writer.sheets["Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row =startrow+ maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
            # Pemusnahan Bahan Baku
            if len(listpemusnahanbahanbaku) != 0:
                startrow = 0
                print(maxcolprevdf)
                print(adjust_column_width)
                for i, datakeluar in enumerate(listpemusnahanbahanbaku):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Barang Keluar"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Barang Keluar"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Barang Keluar"].cell(row=1, column = maxcolprevdf+2,value ='Pemusnahan Bahan Baku')
                    writer.sheets["Barang Keluar"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row =startrow+ maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Barang Keluar"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Barang Keluar"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Barang Keluar"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Barang Keluar"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Barang Keluar'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1



            startrow = 0
            dataframetransaksikeluar.to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Barang Keluar"
                )
            apply_number_format(writer.sheets['Barang Keluar'],startrow+1,startrow+3,maxcolprevdf,maxcolprevdf+6)
            apply_borders_thin(writer.sheets['Barang Keluar'],startrow+2,startrow+3,maxcolprevdf+5,maxcolprevdf+1)
            adjust_column_width(writer.sheets['Barang Keluar'],dataframetransaksikeluar,startrow+1,maxcolprevdf+1)


                
        '''LAPORAN STOK GUDANG'''
        print(listdatagudangperbulan)
        startrow = 1
        for i, datamasuk in enumerate(listdatagudangperbulan):
            # Set the header for the month
            
            # Write the DataFrame to the sheet
            datamasuk[0].to_excel(writer, index=False, startrow=startrow, sheet_name="Saldo Gudang")
            writer.sheets["Saldo Gudang"].cell(row=startrow, column=1, value=f"Laporan Stok Gudang {listbulan[i]}")
            
            # Calculate the maximum row after writing the DataFrame
            maxrow = startrow + len(datamasuk[0]) + 1
            maxcol = len(datamasuk[0].columns)
            
            # Merge cells for the total price row
            writer.sheets["Saldo Gudang"].merge_cells(start_row=maxrow+1, start_column=1, end_row=maxrow+1, end_column=maxcol-1)
            
            # Set the total price label and value
            writer.sheets["Saldo Gudang"].cell(row=maxrow+1, column=1, value="Total Harga")
            writer.sheets["Saldo Gudang"].cell(row=maxrow+1, column=maxcol, value=datamasuk[1])
            
            # Apply formatting and borders
            apply_number_format(writer.sheets["Saldo Gudang"], startrow+1, maxrow+1, 1, maxcol)
            apply_borders_thin(writer.sheets["Saldo Gudang"], startrow+1, maxrow+1, maxcol)
            adjust_column_width(writer.sheets["Saldo Gudang"], datamasuk[0], 1,1)
            
            # Update startrow for the next iteration, adding an extra row for separation
            startrow = maxrow + 3

        '''LAPORAN SALDO FG'''

        if len(listdataartikelfg) != 0 or len(listdatadisplayfg)!=0 or len(listdatabahanbakufg) !=0 :
            maxcolumn  = 0
            maxrow = 0
            maxcolprevdf = 0
            # ARTIKEL KELUAR (SPPB)
            if len(listdataartikelfg) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listdataartikelfg):
                    if len(datakeluar[0]) == 0:
                        continue 
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = 0, startrow=startrow+1, sheet_name="Saldo FG"
                )
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    writer.sheets["Saldo FG"].cell(row=startrow+1, column = 1,value =listbulan[i])
                    writer.sheets["Saldo FG"].cell(row=1, column = 2,value ='Stok Artikel FG')
                    writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+startrow+2, start_column=1,end_row = maxrow+startrow+2,end_column= maxcol-1)
                    writer.sheets["Saldo FG"].cell(row=maxrow+startrow+2, column = 1).value = "Total Harga"
                    writer.sheets["Saldo FG"].cell(row=startrow+maxrow+2, column = maxcol,value = datakeluar[1])
                    # writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Saldo FG"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Saldo FG"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Saldo FG'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Saldo FG'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Saldo FG'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1

            # DISPLAY KELUAR (SPPB)
            if len(listdatadisplayfg) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listdatadisplayfg):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Saldo FG"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Saldo FG"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Saldo FG"].cell(row=1, column = maxcolprevdf+2,value ='Stok Display FG')
                    # writer.sheets["Saldo FG"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Saldo FG"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Saldo FG"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Saldo FG"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Saldo FG"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Saldo FG'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Saldo FG'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Saldo FG'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            # DISPLAY KELUAR (SPPB)
            if len(listdatabahanbakufg) != 0:
                startrow = 0
                for i, datakeluar in enumerate(listdatabahanbakufg):
                    if len(datakeluar[0]) == 0:
                        continue 
                    
                    datakeluar[0].to_excel(
                    writer, index=False,startcol = maxcolprevdf, startrow=startrow+1, sheet_name="Saldo FG"
                )
                    
                    maxrow = len(datakeluar[0])+1
                    maxcol = len(datakeluar[0].columns)
                    print(maxcolprevdf,startrow,maxrow,maxcol)
                    # print(asd)
                    writer.sheets["Saldo FG"].cell(row=startrow+1, column = maxcolprevdf+1,value =listbulan[i])
                    writer.sheets["Saldo FG"].cell(row=1, column = maxcolprevdf+2,value ='Stok Bahan Baku FG')
                    # writer.sheets["Saldo FG"].merge_cells(start_row=startrow+maxrow+2, start_column=maxcolprevdf+1,end_row = maxrow+2,end_column= maxcolprevdf+maxcol-1)
                    writer.sheets["Saldo FG"].cell(row=maxrow+startrow+2, column = maxcolprevdf+1).value = "Total Harga"
                    writer.sheets["Saldo FG"].cell(row=startrow+maxrow+2, column = maxcolprevdf+maxcol,value = datakeluar[1])
                    # writer.sheets["Saldo FG"].merge_cells(start_row=maxrow+3, start_column=1,end_row = maxrow+3,end_column= maxcol-1)
                    # writer.sheets["Saldo FG"].cell(row=maxrow+3, column = 1).value = f"Total Harga Keluar Bulan {listbulan[waktuobj.month - 1]}"
                    # writer.sheets["Saldo FG"].cell(row=maxrow+3, column = maxcol,value = totalbiayakeluar+totalbiayakeluargold+totalbiayakeluardisplay+totalbiayakeluartransaksibahanbaku+totalbiayakeluartransaksilainlain)
                    apply_number_format(writer.sheets['Saldo FG'],startrow+2,startrow+maxrow+2,maxcolprevdf+1,maxcol+maxcolprevdf)
                    apply_borders_thin(writer.sheets['Saldo FG'],startrow+2,startrow+maxrow+2,maxcolprevdf+maxcol,maxcolprevdf+1)
                    adjust_column_width(writer.sheets['Saldo FG'],datakeluar[0],startrow+2,maxcolprevdf+1)
                    startrow +=maxrow+3
                maxcolprevdf += maxcol+1
            
           
            

            
        if not dfstokawalwip.empty or not dfstokawalfg.empty:
            startrow = 1
            maxcolprevdf = 0
            maxrow = 0
            maxcol = 0
            if len(dfstokawalwip)!=0:
                dfstokawalwip.to_excel(writer, index=False, startrow=startrow, sheet_name="Saldo Awal Produksi")
                maxrow = len(dfstokawalwip)+1
                maxcol = len(dfstokawalwip.columns)
                writer.sheets["Saldo Awal Produksi"].cell(row=1, column = 1,value ="Saldo Awal WIP")
                apply_number_format(writer.sheets['Saldo Awal Produksi'],2,maxrow+2,1,maxcol)
                apply_borders_thin(writer.sheets['Saldo Awal Produksi'],2,maxrow+2,maxcol )
                adjust_column_width(writer.sheets['Saldo Awal Produksi'],dfstokawalwip,1,1)
                maxcolprevdf += maxcol+1

            if len(dfstokawalfg)!=0:
                # data stok awal FG
                dfstokawalfg.to_excel(
                                writer, index=False, startrow=1, startcol=maxcol+1, sheet_name="Saldo Awal Produksi"
                )
                writer.sheets["Saldo Awal Produksi"].cell(row=1, column = maxcol+2,value ='Saldo awal FG')
                maxrow = len(dfstokawalfg)+1
                maxcol = len(dfstokawalfg.columns)+ maxcol + 1
                apply_number_format(writer.sheets['Saldo Awal Produksi'],2,maxrow+2,1,maxcol+3)
                apply_borders_thin(writer.sheets['Saldo Awal Produksi'],2,maxrow+2,maxcol ,maxcolprevdf+1)
                adjust_column_width(writer.sheets['Saldo Awal Produksi'],dfstokawalfg,1,maxcolprevdf+1)

            apply_borders_thin(writer.sheets['Saldo Awal Produksi'],2,2,maxcol+3 ,maxcol+2)
        
        if not dfstokawalgudang.empty:
            dfstokawalgudang.to_excel(
                                writer, index=False, startrow=1, startcol=0, sheet_name="Saldo Awal Gudang"
                )
            writer.sheets["Saldo Awal Gudang"].cell(row=1, column = 1,value ='Saldo awal Gudang')
            maxrow = len(dfstokawalgudang)+1
            maxcol = len(dfstokawalgudang.columns)
            writer.sheets["Saldo Awal Gudang"].merge_cells(start_row=maxrow+2, start_column=1,end_row = maxrow+2,end_column= maxcol-1)
            writer.sheets["Saldo Awal Gudang"].cell(row=maxrow+2, column = 1).value = "Total Harga"
            writer.sheets["Saldo Awal Gudang"].cell(row=maxrow+2, column = maxcol,value = totalbiayasaldoawalgudang)
            apply_number_format(writer.sheets['Saldo Awal Gudang'],2,maxrow+2,1,maxcol)
            apply_borders_thin(writer.sheets['Saldo Awal Gudang'],2,maxrow+2,maxcol)
            adjust_column_width(writer.sheets['Saldo Awal Gudang'],dfstokawalgudang,1,1)

    # Load the workbook from the buffer
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Laporan Persediaan"]
    # ws2 = wb["Barang Masuk"]
    # ws3 = wb["Barang Keluar"]

    # # Insert header "Januari" above the table
    # ws["A1"] = listbulan[waktuobj.month - 1]
    # ws2["A1"] = listbulan[waktuobj.month - 1]
    # ws3["A1"] = listbulan[waktuobj.month - 1]

    # Custom WS Barang Keluar
    # baristerakhirws3 = ws3.max_row
    # kolomterakhirws3 = ws3.max_column
    # hurufkolomws3 = get_column_letter(kolomterakhirws3)
    # print("Baris Terakhir : ", baristerakhirws3)
    # print("Kolom Terakhir : ", kolomterakhirws3)
    # print("Huruf Terakhir : ", hurufkolomws3)
    # ws3.cell(row=baristerakhirws3 + 1, column=kolomterakhirws3 - 1, value="Total Biaya")
    # ws3.cell(row=baristerakhirws3 + 1, column=kolomterakhirws3, value=totalbiayakeluar)
    # # print(asdas)
    # Save the workbook back to the buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=laporanpersediaan{bulan}.xlsx"
    )

    return response


@login_required
@logindecorators.allowed_users(allowed_roles=['ppic'])
def read_saldoawalproduksi(request):
    produkobj = models.SaldoAwalProduksi.objects.all().order_by('Tanggal')
    return render(request, "ppic/read_saldoawalproduksi.html", {"produkobj": produkobj})


@login_required
@logindecorators.allowed_users(allowed_roles=["ppic"])
def create_produk(request):
    if request.method == "GET":
        return render(request, "ppic/create_saldoawalproduksi.html")
    else:
        saldo = request.POST["saldo"]
        tanggal = request.POST["tanggal"]
        tanggalobj = datetime.strptime(str(tanggal),"%Y-%m-%d")
        produkobj = models.SaldoAwalProduksi.objects.filter(Tanggal__year =tanggalobj.year)
        if len(produkobj) > 0:
            messages.error(request, f"Sudah ada data Saldo Awal Produksi pada tahun {tanggalobj.year}")
            return redirect("create_saldoawalproduksi")
        else:
            new_saldoawal = models.SaldoAwalProduksi(
                Saldo = saldo,
                Tanggal = tanggalobj
            )
            new_saldoawal.save()
            models.transactionlog(
                user="PPIC",
                waktu=datetime.now(),
                jenis="Create",
                pesan=f"Saldo awal {saldo} Tanggal {tanggalobj}",
            ).save()
            messages.success(request, "Data berhasil disimpan")
            return redirect("read_sakdoawalproduksi")


@login_required
@logindecorators.allowed_users(allowed_roles=["ppic"])
def update_produk(request, id):
    produkobj = models.SaldoAwalProduksi.objects.get(pk=id)
    produkobj.Tanggal = produkobj.Tanggal.strftime("%Y-%m-%d")
    if request.method == "GET":
        return render(
            request, "ppic/update_saldoawalproduksi.html", {"produkobj": produkobj}
        )
    else:

        saldo = request.POST["saldo"]
        tanggal = request.POST["tanggal"]
        tanggalobj = datetime.strptime(str(tanggal),"%Y-%m-%d")

        namaproduk = models.SaldoAwalProduksi.objects.filter(Tanggal__year=tanggalobj.year).exclude(pk=id).exists()
        print(namaproduk)
        print(request.POST)
        print(id)
        # print(asd)

        if namaproduk:
            messages.error(request,f'Saldo awal Produksi  {tanggalobj.year} sudah ada pada sistem')
            return redirect('update_saldoawalproduksi',id = id)
        produkbaru = models.SaldoAwalProduksi.objects.get(id=id)
        produkbaru.Saldo = saldo
        produkbaru.Tanggal = tanggalobj
       
        # print(asd)
        produkbaru.save()
        models.transactionlog(
            user="PPIC",
            waktu=datetime.now(),
            jenis="Update",
            pesan=f"Saldo Awal Produksi {tanggalobj.year} sudah di Update",
        ).save()
        messages.success(request, "Data berhasil disimpan")
        return redirect("read_sakdoawalproduksi")


@login_required
@logindecorators.allowed_users(allowed_roles=["ppic"])
def delete_produk(request, id):
    produkobj = models.SaldoAwalProduksi.objects.get(pk=id)
    # print("delete:",produkobj.KodeProduk)
    models.transactionlog(
        user="PPIC",
        waktu=datetime.now(),
        jenis="Delete",
        pesan=f"Saldo Awal Produksi {produkobj.Tanggal.year} sudah di Delete",
    ).save()
    produkobj.delete()
    messages.success(request, "Data Berhasil dihapus")
    return redirect("read_sakdoawalproduksi")